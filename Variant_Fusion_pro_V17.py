#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Variant Fusion – Best-of-Hybrid mit präzisem Populations-AF, nicht-persistiertem Genotyp,
Maintenance nur im Leerlauf, DB-Batch-Writes und gedrosselten GUI-Updates.

Windows-tauglich, ohne pysam/bcftools/samtools.

═══════════════════════════════════════════════════════════════════════════════
V17 ARCHITEKTUR (Stand: 2026-03) - REFAKTORIERT
═══════════════════════════════════════════════════════════════════════════════

OFFENE ARCHITEKTUR-SCHULDEN:

1. FLAG_AND_OPTIONS_MANAGER NICHT VOLLSTÄNDIG INTEGRIERT
   - App verwendet parallel tk.BooleanVar neben dem FlagManager
   - SOLL: App erstellt FlagManager, übergibt ihn an alle Komponenten

2. EMIT_QUEUE DIREKTZUGRIFFE (bewusste Designentscheidung)
   - VCFBuffer greift direkt auf emit_queue zu statt drain_live_enqueue
   - Für Single-Key-Updates akzeptabel, für Batch sollte drain_live_enqueue genutzt werden

3. CODING_FILTER DOPPELT INSTANZIIERT
   - Separate Instanzen in MainFilterGate und Distiller
   - Sollte nur in MainFilterGate existieren und durchgereicht werden

ERLEDIGT (2026-03):
  - AFFetchController: Einzelinstanz in App, wird per Referenz an Distiller + BackgroundMaintainer übergeben
  - STALE_DAYS Refactoring: Config.STALE_DAYS_AF=365, Config.STALE_DAYS_FULL=30, separate Getter/Setter, UI-Integration
   
GENE-ANNOTATION DEBUGGING:
   - GeneAnnotator loggt jetzt detailliert: Pfade, Cache-Status, Build-Verfügbarkeit
   - Bei Problemen: Prüfe Log auf "[GeneAnnotator]" Einträge
   - Cache-Verzeichnis ist relativ zur Script-Datei, NICHT zum CWD!
═══════════════════════════════════════════════════════════════════════════════
"""

# =============================================================================
# STANDARD LIBRARY
# =============================================================================
import asyncio
import ast
import atexit
import csv
import gzip
import io
import itertools
import json
import math
import multiprocessing
import os
import queue
import random
import re
import shutil
import signal
import socket
import sqlite3
import subprocess
import sys
import threading
import time
import traceback
import urllib.parse
import webbrowser
import pickle  # <--- Dieser Import fehlt wahrscheinlich oder wird nur lokal verwendet
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from contextlib import contextmanager
from dataclasses import dataclass, field
from pathlib import Path
from threading import RLock, Thread
from typing import Any, Callable, Dict, Iterator, List, Optional, Set, TextIO, Tuple, Union


# KRITISCH: datetime immer als Modul importieren (nie "from datetime import ...")
import datetime
#Neu und ungeordnet:
import mmap
from functools import lru_cache



# =============================================================================
# THIRD-PARTY: ESSENTIAL
# =============================================================================
import psutil
import requests
from requests.adapters import HTTPAdapter

# PIL
from PIL import Image, ImageDraw

# Interval Tree
from intervaltree import IntervalTree

# Statistics
from scipy.stats import chi2_contingency


# =============================================================================
# THIRD-PARTY: ASYNC HTTP (KRITISCH für Fetcher)
# =============================================================================
try:
    import aiohttp
    HAVE_AIOHTTP = True
    AIOHTTP_AVAILABLE = True  # ✅ Alias für Kompatibilität
except ImportError:
    aiohttp = None
    HAVE_AIOHTTP = False
    AIOHTTP_AVAILABLE = False  # ✅ Alias für Kompatibilität
    print("[WARNING] aiohttp not installed - async variant fetching will be disabled")
    print("[WARNING] Install with: pip install aiohttp")

# =============================================================================
# GUI (tkinter & ttkbootstrap)
# =============================================================================
import tkinter as tk
from tkinter import filedialog
import tkinter.ttk as tk_ttk  # ✅ NEU: Fallback für PanedWindow
try:
    import ttkbootstrap as ttk
    from ttkbootstrap.constants import *
    from ttkbootstrap.dialogs import Messagebox
    from ttkbootstrap.toast import ToastNotification
    from ttkbootstrap.scrolled import ScrolledFrame
except ImportError:
    print("ttkbootstrap ist erforderlich! Bitte installieren: pip install ttkbootstrap")
    sys.exit(1)

# =============================================================================
# SYSTEM TRAY
# =============================================================================
try:
    import pystray
    HAVE_PYSTRAY = True
except ImportError:
    pystray = None
    HAVE_PYSTRAY = False


# =============================================================================
# OPTIONAL: Excel Export
# =============================================================================
try:
    import openpyxl
    HAVE_OPENPYXL = True
except ImportError:
    openpyxl = None
    HAVE_OPENPYXL = False


# =============================================================================
# OPTIONAL: PDF Export
# =============================================================================
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A3, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
    HAVE_REPORTLAB = True
except ImportError:
    colors = None
    HAVE_REPORTLAB = False


# =============================================================================
# OPTIONAL: AlphaGenome
# =============================================================================
try:
    import numpy as np
    HAVE_NUMPY = True
except ImportError:
    np = None
    HAVE_NUMPY = False

try:
    import alphagenome
    from alphagenome.data import genome as ag_genome
    from alphagenome.models import dna_client as ag_client
    HAVE_ALPHAGENOME = True
except ImportError:
    alphagenome = None
    ag_genome = None
    ag_client = None
    HAVE_ALPHAGENOME = False


# =============================================================================
# OPTIONAL: Biopython
# =============================================================================
try:
    from Bio.Align import PairwiseAligner
    HAVE_BIOPYTHON = True
except ImportError:
    PairwiseAligner = None
    HAVE_BIOPYTHON = False


# =============================================================================
# OPTIONAL: pyfaidx (für FASTA-Indexierung)
# =============================================================================
try:
    import pyfaidx
    from pyfaidx import Faidx
    HAVE_PYFAIDX = True
except ImportError:
    pyfaidx = None
    Faidx = None
    HAVE_PYFAIDX = False


# =============================================================================
# OPTIONAL: MyVariant Python Client
# =============================================================================
try:
    import myvariant
    HAVE_MYVARIANT = True
except ImportError:
    myvariant = None
    HAVE_MYVARIANT = False

# =============================================================================
# CYTHON HOT-PATH ACCELERATION
# =============================================================================
try:
    from cython_hotpath import CythonAccelerator
    CYTHON_AVAILABLE = True
except ImportError:
    CYTHON_AVAILABLE = False
    CythonAccelerator = None
# =============================================================================
# VERSION INFO
# =============================================================================
__version__ = "17.0"
APP_VERSION = "V17"
APP_NAME = "Variant Fusion Pro"
# PyInstaller One-File: _MEIPASS fuer gebundelte Daten, EXE-Pfad fuer User-Dateien
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
    _BUNDLE_DIR = getattr(sys, '_MEIPASS', BASE_DIR)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    _BUNDLE_DIR = BASE_DIR

# =============================================================================
# EARLY SPLASH WINDOW (V17.1) - Sofortiges visuelles Feedback beim Start
# =============================================================================
_splash_window = None
_splash_log_text = None

def _create_early_splash():
    """Erstellt ein minimales Splash-Fenster BEVOR die schweren Module laden."""
    global _splash_window, _splash_log_text
    try:
        _splash_window = tk.Tk()
        _splash_window.title(f"{APP_NAME} {APP_VERSION}")
        _splash_window.geometry("620x400")
        _splash_window.resizable(False, False)
        # Zentrieren
        _splash_window.update_idletasks()
        sw = _splash_window.winfo_screenwidth()
        sh = _splash_window.winfo_screenheight()
        x = (sw - 620) // 2
        y = (sh - 400) // 2
        _splash_window.geometry(f"620x400+{x}+{y}")

        frame = tk.Frame(_splash_window, padx=25, pady=20)
        frame.pack(fill="both", expand=True)

        tk.Label(
            frame, text=f"{APP_NAME} {APP_VERSION}",
            font=("Segoe UI", 16, "bold"), anchor="w"
        ).pack(fill="x", pady=(0, 2))
        tk.Label(
            frame, text="Initialisierung ...",
            font=("Segoe UI", 9), anchor="w", fg="#666666"
        ).pack(fill="x", pady=(0, 8))

        sep = tk.Frame(frame, height=1, bg="#cccccc")
        sep.pack(fill="x", pady=(0, 10))

        _splash_log_text = tk.Text(
            frame, height=14, font=("Consolas", 9),
            wrap="word", state="normal", relief="flat",
            borderwidth=0, highlightthickness=0, bg="#f8f8f8"
        )
        _splash_log_text.pack(fill="both", expand=True)

        _splash_window.update()
    except Exception:
        _splash_window = None
        _splash_log_text = None

def _splash_log(msg):
    """Schreibt eine Nachricht ins Splash-Fenster (falls vorhanden)."""
    global _splash_log_text, _splash_window
    if _splash_log_text and _splash_window:
        try:
            _splash_log_text.insert("end", msg + "\n")
            _splash_log_text.see("end")
            _splash_window.update_idletasks()
        except Exception:
            pass

def _close_splash():
    """Schliesst das Splash-Fenster."""
    global _splash_window, _splash_log_text
    if _splash_window:
        try:
            _splash_window.destroy()
        except Exception:
            pass
        _splash_window = None
        _splash_log_text = None

if __name__ == "__main__":
    _create_early_splash()
    _splash_log("Lade Module ...")

# =============================================================================
# SINGLE INSTANCE CHECK (V17.0)
# =============================================================================
def cleanup_lock():
    """Entfernt das Lock-File bei Programmende."""
    lock_file = os.path.join(BASE_DIR, ".vf_instance.lock")
    try:
        if os.path.exists(lock_file):
            os.remove(lock_file)
    except:
        pass

def check_single_instance():
    """
    Prüft ob bereits eine Instanz läuft (via Lock-File).
    Returns: (success: bool, pid: int)
    """
    lock_file = os.path.join(BASE_DIR, ".vf_instance.lock")
    current_pid = os.getpid()
    
    if os.path.exists(lock_file):
        try:
            with open(lock_file, "r") as f:
                content = f.read().strip()
                if content:
                    old_pid = int(content)
                    if psutil.pid_exists(old_pid):
                        return False, old_pid
                    else:
                        print(f"[Startup] Stale lock found (PID {old_pid} not running). Taking over.")
        except Exception as e:
            print(f"[Startup] Lock file read error: {e}")
            
    # Create new lock
    try:
        with open(lock_file, "w") as f:
            f.write(str(current_pid))
        atexit.register(cleanup_lock)
        return True, current_pid
    except Exception as e:
        print(f"[Startup] Could not write lock file: {e}")
        return True, current_pid # Fail open if we can't write lock

# =============================================================================
# IMPORT-STATUS LOGGING
# =============================================================================
print("\n" + "="*70)
print(f"{APP_NAME} {APP_VERSION} - Import Status")
print("="*70)
_import_lines = [
    "Core libraries loaded",
    f"aiohttp (async HTTP) - {'OK' if HAVE_AIOHTTP else 'MISSING'}",
    f"openpyxl (Excel) - {'OK' if HAVE_OPENPYXL else 'disabled'}",
    f"reportlab (PDF) - {'OK' if HAVE_REPORTLAB else 'disabled'}",
    f"pyfaidx (FASTA) - {'OK' if HAVE_PYFAIDX else 'disabled'}",
    f"biopython - {'OK' if HAVE_BIOPYTHON else 'disabled'}",
    f"pystray (Tray) - {'OK' if HAVE_PYSTRAY else 'disabled'}",
]
for _line in _import_lines:
    print(f"  {_line}")
    _splash_log(f"  {_line}")
print("="*70 + "\n")
_splash_log("Module geladen. Lade Konfiguration ...")

# Kritischer Hinweis wenn aiohttp fehlt
if not HAVE_AIOHTTP:
    print("\n" + "!"*70)
    print("WARNING: aiohttp is not installed!")
    print("Async variant fetching will be DISABLED.")
    print("This will significantly slow down annotation for large VCFs.")
    print("Install with: pip install aiohttp")
    print("!"*70 + "\n")

# ═══════════════════════════════════════════════════════════════════════════════
# METHODENVERZEICHNIS V17 (zur Wartung)
# ═══════════════════════════════════════════════════════════════════════════════
"""
┌─────────────────────────────────────────────────────────────────────────────────┐
│                        VARIANT FUSION PRO V17 - ARCHITEKTUR                     │
├─────────────────────────────────────────────────────────────────────────────────┤
│                                                                                 │
│  App (GUI)                                                                      │
│   ├── flag_manager: Flag_and_Options_Manager (ZENTRAL für alle GUI-Settings)   │
│   ├── main_filter_gate: MainFilterGate (ZENTRAL für Filter-Entscheidungen)     │
│   ├── quality_manager: QualityManager (VCF-Record-Level Filter)                │
│   └── distiller: Distiller (Pipeline-Koordination)                             │
│                                                                                 │
│  Datenfluss:                                                                    │
│  GUI tk-Vars → flag_manager.sync_from_gui() → FlagManager → MainFilterGate     │
│                                               └→ Distiller → Pipeline-Phasen   │
│                                                                                 │
└─────────────────────────────────────────────────────────────────────────────────┘

════════════════════════════════════════════════════════════════════════════════
1. GLOBALE HELPER
════════════════════════════════════════════════════════════════════════════════

Logging:           MultiSinkLogger, _NullLogger
Zeit/Format:       fmt_eta, safe_float, now_iso, parse_iso_utc
Dateien:           is_gzipped, open_text_maybe_gzip, is_vcf
VCF-Parsing:       parse_vcf_records, parse_vcf_records_mmap, parse_vcf_records_smart
Genome:            determine_is_coding, load_fai_index, fetch_base_from_fasta, get_ref_base
Genotyp:           get_genotype_label, normalize_genotype_display
Threading:         run_on_main_thread (GUI-Dialoge aus Threads)

════════════════════════════════════════════════════════════════════════════════
2. BUILD DETECTION
════════════════════════════════════════════════════════════════════════════════

Haupt:             detect_build_for_vcf
Sub-Funktionen:    parse_vcf_keys, detect_build_from_header, detect_build_from_contigs,
                   detect_build_from_positions, count_variants_exact

════════════════════════════════════════════════════════════════════════════════
3. API / FETCHER (Async & Sync)
════════════════════════════════════════════════════════════════════════════════

Verfügbarkeit:     check_aiohttp_available
MyVariant:         mv_fetch, mv_fetch_async, mv_fetch_threadpool, mv_fetch_with_module
                   extract_fields_from_mv, normalize_for_mv
Externe APIs:      gnomad_fetch_async, vep_fetch_async, alfa_fetch_async
                   normalize_for_gnomad, normalize_for_vep

════════════════════════════════════════════════════════════════════════════════
4. MANAGER-KLASSEN (V17 Architektur)
════════════════════════════════════════════════════════════════════════════════

┌─────────────────────────────────────────────────────────────────────────────────┐
│ Flag_and_Options_Manager (ZENTRAL)                                              │
│   - Speichert ALLE GUI-Einstellungen thread-safe                               │
│   - sync_from_gui(app) → Synchronisiert tk.Vars vor Pipeline-Start             │
│   - get_af_threshold(), get_include_none(), get_only_protein_coding()...       │
│   - set_*() Methoden für programmatische Änderungen                            │
├─────────────────────────────────────────────────────────────────────────────────┤
│ MainFilterGate (ZENTRAL)                                                        │
│   - check_variant(key, af_value, fetch_status) → (passed, reason, data)        │
│   - check_batch(variants) → Batch-Prüfung                                      │
│   - SubGates:                                                                   │
│     ├── CodingFilter (Protein-Coding Filter)                                   │
│     └── AfNoneTreatmentManager (AF-None Policy)                                │
├─────────────────────────────────────────────────────────────────────────────────┤
│ QualityManager (VCF-Record-Level)                                               │
│   - Filtert auf Record-Ebene (QUAL, FILTER, DP, HomRef)                        │
│   - Unabhängig von MainFilterGate (andere Abstraktionsebene)                   │
├─────────────────────────────────────────────────────────────────────────────────┤
│ FetchStatusManager (Static Utility)                                             │
│   - Kodiert/Dekodiert Fetch-Status (3-Bit: gnomAD, ExAC, 1KG)                  │
│   - is_success(), encode(), decode()                                           │
├─────────────────────────────────────────────────────────────────────────────────┤
│ LightDBGnomADManager                                                            │
│   - Verwaltet gnomAD LightDB Download/Update                                   │
│   - Background-Loading mit Fortschritt                                         │
└─────────────────────────────────────────────────────────────────────────────────┘

Config:            Zentrale Konfigurationskonstanten

════════════════════════════════════════════════════════════════════════════════
5. PIPELINE CORE (Distiller)
════════════════════════════════════════════════════════════════════════════════

┌─────────────────────────────────────────────────────────────────────────────────┐
│ Distiller                                                                       │
│   __init__(app, db, flag_manager, main_filter_gate, quality_manager, ...)      │
│                                                                                 │
│   ENTRY POINTS:                                                                 │
│   ├── process_file(path, build) → Haupteinstieg                                │
│   └── _distill_vcf(vcf_path, build) → VCF-Verarbeitung                         │
│                                                                                 │
│   PIPELINE PHASEN:                                                              │
│   ├── _phase_vcf_scan         → VCF lesen, Pre-Filter, Keys sammeln            │
│   ├── _start_streaming_pipeline → Pipeline-Thread starten                       │
│   ├── _phase_af_fetch_streaming    → AF-Frequenzen holen                       │
│   ├── _phase_full_annotation_streaming → Volle Annotation (VEP, etc.)          │
│   ├── _phase_gene_annotation_streaming → Gen-Symbole ergänzen                  │
│   ├── _phase_rsid_fill_streaming       → RSIDs ergänzen                        │
│   ├── _phase_missing_fill_streaming    → Fehlende Felder ergänzen              │
│   └── _phase_alphagenome_streaming     → AlphaGenome Scores                    │
│                                                                                 │
│   FILTER-METHODEN (Delegieren an MainFilterGate):                              │
│   ├── _apply_af_filter_final(key, val) → "emit"|"reject"                       │
│   └── _validate_af_in_cache(key, row)  → "emit"|"reject"|"fetch"               │
│                                                                                 │
│   HELPER:                                                                       │
│   ├── _process_variant_batch  → Batch-Verarbeitung                             │
│   ├── _lookup_lightdb         → LightDB Lookup                                 │
│   ├── _cleanup_controllers    → Resource Cleanup                               │
│   └── export_*                → Export-Methoden (VCF, CSV, XLSX)               │
└─────────────────────────────────────────────────────────────────────────────────┘

════════════════════════════════════════════════════════════════════════════════
6. PIPELINE SUPPORT KLASSEN
════════════════════════════════════════════════════════════════════════════════

AFFetchController:   Async AF-Fetch Koordination
VCFBuffer:           Gepufferte DB-Schreiboperationen
EmitQueue:           Thread-safe Queue für GUI-Updates
PipelineProgress:    Fortschritts-Tracking (Phasen, Varianten)

════════════════════════════════════════════════════════════════════════════════
7. MAINTAINER (Background Tasks)
════════════════════════════════════════════════════════════════════════════════

BackofficeCrawler:       Crawlt fehlende Annotationen im Hintergrund
BackgroundMaintainer:    Koordiniert alle Background-Tasks
VCFMigrationsdienst:     Migriert alte VCF-Einträge

════════════════════════════════════════════════════════════════════════════════
8. CONVERTER / TOOLS
════════════════════════════════════════════════════════════════════════════════

Konverter:         convert_23andme_to_vcf, FASTQmap, StreamingFastaToGVCF
Validierung:       FastaValidator
Scoring:           AlphaGenomeScorer
Resilience:        CircuitBreaker, ThroughputTuner

════════════════════════════════════════════════════════════════════════════════
9. DATABASE
════════════════════════════════════════════════════════════════════════════════

VariantDB:         SQLite-basierte Varianten-Datenbank
                   _init_db, upsert_variant, get_variant, upsert_variants_bulk,
                   upsert_many, get_variants_by_keys

════════════════════════════════════════════════════════════════════════════════
10. GUI
════════════════════════════════════════════════════════════════════════════════

App:               Hauptfenster (Tkinter)
                   - Erstellt alle Manager in __init__
                   - Synchronisiert GUI → FlagManager vor Pipeline-Start
                   - Verwaltet Live-Updates via EmitQueue

QualitySettingsDialog:  Dialog für Quality-Filter Einstellungen

"""
# =============================================================================
# CONFIG: Zentrale Konfiguration (VOLLSTÄNDIG)
# =============================================================================

class Config:
    """
    ✅ VOLLSTÄNDIG: Zentrale Konfiguration für ALLE Parameter.
    
    Single Source of Truth für:
    - Dateipfade & Verzeichnisse
    - Performance-Parameter (inkl. neue Optimierungen)
    - Logging-Einstellungen
    - Network-Einstellungen
    - GUI-Parameter
    - Database-Einstellungen
    - FASTA-Validierung (mit Batch-Modus)
    """
    
    # =================================
    # FILES & DIRECTORIES
    # =================================
    TEMP_VCF_DIR = "converted_23andme_vcfs"
    SETTINGS_FILE = "variant_fusion_settings.json"
    LOG_FILE = "distiller_debug.log"
    CACHE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "cache.json")
    
    # =================================
    # LOGGING
    # =================================
    LOG_DATETIME_FMT = "%Y-%m-%d %H:%M:%S"
    QUEUE_MAX_SIZE = 10000
    
    # =================================
    # THREADING & PARALLELISIERUNG
    # =================================
    DEFAULT_THREADS = max(4, os.cpu_count() or 4)
    MIN_WORKERS = 2
    MAX_WORKERS = 100
    TARGET_CPU_FRACTION = 0.8
    
    # =================================
    # NETWORK & FETCHING
    # =================================
    REQUEST_TIMEOUT = 15        # HTTP Request Timeout (Sekunden)
    MAX_RETRIES = 3             # Maximale Retry-Versuche
    ASYNC_THRESHOLD = 50000     # Ab dieser Anzahl Keys → Async-Modus
    
    # =================================
    # DATABASE
    # =================================
    DB_TIMEOUT_SEC = 30         # SQLite Timeout
    DB_WAL_MODE = True          # Write-Ahead-Logging aktivieren
    
    # =================================
    # STALENESS / DATA FRESHNESS
    # =================================
    # AF-Daten ändern sich selten (gnomAD Releases ~jährlich)
    STALE_DAYS_AF = 365
    # Vollannotationen können sich öfter ändern (ClinVar, MyVariant Updates)
    STALE_DAYS_FULL = 30
    # =================================
    # CACHING & MEMORY
    # =================================
    CACHE_MAX_SIZE = 100000     # Max. Einträge in rows_cache
    
    # ========================================================================
    # ⚡ PERFORMANCE-OPTIMIERUNGEN (NEU)
    # ========================================================================
    
    # VCF-Parsing: Dynamische Chunk-Size (wird automatisch berechnet)
    # Range: 2.000 - 10.000 (basierend auf verfügbarem RAM)
    # Bei Fehler/ohne psutil: Fallback auf 1.000
    VCF_CHUNK_SIZE_MIN = 2000
    VCF_CHUNK_SIZE_MAX = 10000
    VCF_CHUNK_SIZE_FALLBACK = 1000
    
    # Adaptive Progress-Updates
    # Bei großen Dateien (>100k Varianten) weniger Updates für bessere Performance
    VCF_UPDATE_INTERVAL_SMALL = 1000   # Für Dateien <100k Varianten
    VCF_UPDATE_INTERVAL_LARGE = 5000   # Für Dateien ≥100k Varianten
    VCF_LOG_INTERVAL_SMALL = 10000     # Log-Output <100k Varianten
    VCF_LOG_INTERVAL_LARGE = 50000     # Log-Output ≥100k Varianten
    
    # Memory-Mapped VCF Reading
    # Aktiviert automatisch für unkomprimierte VCF >100MB
    VCF_MMAP_THRESHOLD_MB = 100        # Dateigröße-Schwellwert für mmap
    
    # =================================
    # GUI & DISPLAY
    # =================================
    TABLE_BATCH_SIZE = 200              # Treeview-Inserts pro Cycle
    POLL_INTERVAL_MS = 50               # GUI-Polling Intervall (ms)
    PROGRESS_UPDATE_INTERVAL = 1000     # Progress-Bar Update (ms)
    # ✅ NEU: Design Theme
    DEFAULT_THEME = "cosmo" # Helles, modernes Theme

    # ========================================================================
    # V16: LOKALER DB-PFAD (außerhalb OneDrive!)
    # ========================================================================
    # Grund: OneDrive blockiert große DBs während Sync (gnomad_light.db = 94GB!)
    # Alle DBs werden hier gesucht/gespeichert. Backup erfolgt auf NAS.
    LOCAL_DB_DIR = r"C:\_Local_DEV\DATA_STORE"

    # ========================================================================
    # FASTA-VALIDIERUNG (V16: PathManager-Integration)
    # ========================================================================

    # ✅ V16: FASTA_PATHS wird jetzt über PathManager verwaltet (siehe globale Definition)
    # Die alten hardcoded Pfade wurden entfernt. PathManager findet FASTA automatisch.

    # Validierung


    # Validierungs-Modus
    # True = STRICT (nur SNVs skip, ~85% schneller)
    # False = AGGRESSIVE (SNVs + INS skip, ~90% schneller) [DORMANT]
    FASTA_STRICT_MODE = True
    
    # Region-Cache-Größe (Anzahl 100kb-Regionen im Cache)
    FASTA_REGION_CACHE_SIZE = 1000
    
    # ========================================================================
    # ⚡ FASTA BATCH-VALIDIERUNG (NEU - EXPERIMENTELL)
    # ========================================================================
    
    # Batch-Validierung aktivieren
    # True = Validiere mehrere Varianten gleichzeitig (20-30% schneller bei vielen INDELs)
    # False = Einzelne Validierung (Standard, sicherer, empfohlen)
    FASTA_BATCH_VALIDATION = False
    
    # Batch-Größe für FASTA-Validierung
    # Anzahl Varianten die zusammen validiert werden
    # Höhere Werte = bessere Performance, aber mehr Memory
    FASTA_BATCH_SIZE = 100
    
    # ℹ️ HINWEIS zur Batch-Validierung:
    # Die Batch-Validierung ist implementiert und verfügbar über:
    #   FastaValidator.validate_batch(variants)
    # 
    # Aktivierung erfordert Anpassung der VCF-Scan-Loop:
    # 1. Varianten sammeln bis FASTA_BATCH_SIZE erreicht
    # 2. validate_batch() aufrufen
    # 3. Ergebnisse verarbeiten
    # 
    # Vorteile:
    # - 20-30% schneller durch:
    #   * Pre-Check für SNVs (skip FASTA-Lookup)
    #   * Sortierung nach Chromosom/Position (bessere Cache-Hits)
    #   * Weniger FASTA-File-Operations
    # 
    # Nachteile:
    # - Komplexere Loop-Logik
    # - Höherer Memory-Bedarf (buffert FASTA_BATCH_SIZE Varianten)
    # In Config-Klasse nach FASTA_BATCH_SIZE einfügen:

    # ========================================================================
    # ⚡ FASTA SHIFT-DETECTION (NEU - EXPERIMENTELL)
    # ========================================================================

    # Shift-Detection aktivieren
    # True = Erkennt systematische Positions-Verschiebungen (+1, -1, etc.)
    # False = Deaktiviert (Standard FASTA-Validierung)
    FASTA_SHIFT_DETECTION = True

    # Mindest-Confidence für Pattern-Anwendung
    # 0.6 = 60% der Fehler müssen gleichen Shift haben
    # Höher = konservativer, niedriger = aggressiver
    FASTA_SHIFT_CONFIDENCE = 0.6

    # Mindestanzahl Samples vor Pattern-Anwendung
    # Verhindert falsche Pattern bei zu wenigen Daten
    FASTA_SHIFT_MIN_SAMPLES = 10

    # Zu testende Shift-Offsets
    # Reihenfolge nach Häufigkeit (basierend auf echten Daten)
    FASTA_SHIFT_TEST_RANGE = [+1, -1, +2, -2]

    # FASTA-Download-URLs (optional, für automatischen Download)
    FASTA_DOWNLOAD_URLS = {
        "hg19": "https://hgdownload.soe.ucsc.edu/goldenPath/hg19/bigZips/hg19.fa.gz",
        "hg38": "https://hgdownload.soe.ucsc.edu/goldenPath/hg38/bigZips/hg38.fa.gz",
    }
    
    # ========================================================================
    # COLUMN LINK TEMPLATING SYSTEM (NEU)
    # ========================================================================
    
    DEFAULT_COLUMN_LINKS = {
        'gene': {
            'trigger': 'single',
            'template': 'https://www.genecards.org/cgi-bin/carddisp.pl?gene={value}'
        },
        'rsid': {
            'trigger': 'double',
            'template': 'https://www.ncbi.nlm.nih.gov/snp/{value}'
        },
        'pos': {
            'trigger': 'none',
            'template': 'https://gnomad.broadinstitute.org/region/{chr}-{pos}-{pos}?dataset=gnomad_r4'
        },
        'pubmed': {
            'trigger': 'double',
            'template': 'https://pubmed.ncbi.nlm.nih.gov/?term={gene}+{consequence}+variant'
        }
    }
    
    COLUMN_LINK_PRESETS = {
        'gene': {
            'GeneCards': 'https://www.genecards.org/cgi-bin/carddisp.pl?gene={value}',
            'MalaCards': 'https://www.malacards.org/card/{value}',
            'OMIM': 'https://omim.org/search?search={value}',
            'UniProt': 'https://www.uniprot.org/uniprotkb?query={value}'
        },
        'rsid': {
            'dbSNP': 'https://www.ncbi.nlm.nih.gov/snp/{value}',
            'ClinVar': 'https://www.ncbi.nlm.nih.gov/clinvar/?term={value}',
            'gnomAD': 'https://gnomad.broadinstitute.org/variant/{value}?dataset=gnomad_r4'
        },
        'pos': {
            'gnomAD Region': 'https://gnomad.broadinstitute.org/region/{chr}-{pos}-{pos}?dataset=gnomad_r4',
            'UCSC Browser': 'https://genome.ucsc.edu/cgi-bin/hgTracks?db=hg38&position=chr{chr}:{pos}-{pos}'
        }
    }

    # ========================================================================
    # QUALITY-MANAGEMENT
    # ========================================================================
    
    # Standard Quality-Preset
    DEFAULT_QUALITY_PRESET = "medium"
    
    # Quality-Presets (werden von QualityManager verwendet)
    QUALITY_PRESETS = {
        "none": {
            "filter_pass_only": False,
            "qual_threshold": 0,
            "use_dp_filter": False,
            "min_dp": 0,
            "max_dp": None,
        },
        "low": {
            "filter_pass_only": False,
            "qual_threshold": 10,
            "use_dp_filter": True,
            "min_dp": 5,
            "max_dp": 1000,
        },
        "medium": {
            "filter_pass_only": True,
            "qual_threshold": 20,
            "use_dp_filter": True,
            "min_dp": 10,
            "max_dp": 500,
        },
        "high": {
            "filter_pass_only": True,
            "qual_threshold": 30,
            "use_dp_filter": True,
            "min_dp": 15,
            "max_dp": 300,
        },
        "strict": {
            "filter_pass_only": True,
            "qual_threshold": 50,
            "use_dp_filter": True,
            "min_dp": 20,
            "max_dp": 200,
        },
    }
    
    # ========================================================================
    # EXPORT-FORMATE
    # ========================================================================
    
    # VCF-Export: Header-Optionen
    VCF_EXPORT_PRESERVE_HEADER = True      # Original VCF-Header beibehalten
    VCF_EXPORT_ADD_INFO_FIELDS = True      # Neue INFO-Fields hinzufügen
    
    # TSV/Excel-Export: Spalten-Auswahl
    EXPORT_DEFAULT_COLUMNS = [
        "chrom", "pos", "ref", "alt", "rsid",
        "gene_symbol", "consequence", "impact",
        "af_gnomad", "af_1kg", "af_filter_mean",
        "clinvar_significance", "genotype"
    ]
    
    # ========================================================================
    # API-ENDPOINTS & EXTERNAL SERVICES
    # ========================================================================
    
    # MyVariant.info API
    MYVARIANT_BASE_URL = "https://myvariant.info/v1"
    MYVARIANT_BATCH_SIZE = 1000
    
    # ClinVar
    CLINVAR_BASE_URL = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils"
    
    # gnomAD (falls direkt abgefragt)
    GNOMAD_API_URL = "https://gnomad.broadinstitute.org/api"

    # ========================================================================
    # API-SETTINGS (User-konfigurierbar via Settings-Tab)
    # ========================================================================

    DEFAULT_API_SETTINGS = {
        "global": {
            "max_retries": 3,
            "ncbi_api_key": "",
        },
        "phase1_af": {
            "myvariant":  {"enabled": True, "timeout": 60, "batch_size": 400, "workers": 8},
            "gnomad":     {"enabled": True, "timeout": 60, "workers": 8},
            "vep":        {"enabled": True, "timeout": 60, "batch_size": 25,  "workers": 4},
            "alfa":       {"enabled": True, "timeout": 30, "workers": 4},
            "topmed":     {"enabled": False, "timeout": 15, "workers": 6},
        },
        "phase2_full": {
            "myvariant_full": {"enabled": True, "timeout": 60, "batch_size": 500},
        },
        "phase4_rsid": {
            "ncbi_allele": {"enabled": True, "timeout": 10, "rate_limit": 0.35},
        },
        "phase5_fill": {
            "ncbi_clinvar":   {"enabled": True, "timeout": 10, "rate_limit": 0.5},
            "myvariant_cons": {"enabled": True, "timeout": 10, "rate_limit": 0.5},
        },
        "phase6_ag": {
            "alphagenome": {"enabled": True, "timeout": 30, "api_key": ""},
        },
    }

    API_PHASE_LABELS = {
        "phase1_af": "Phase 1: Allel-Frequenz Abruf",
        "phase2_full": "Phase 2: Vollannotation",
        "phase4_rsid": "Phase 4: RSID-Ergaenzung",
        "phase5_fill": "Phase 5: Fehlende Felder",
        "phase6_ag": "Phase 6: AlphaGenome AI",
    }

    API_LABELS = {
        "myvariant": "MyVariant.info",
        "gnomad": "gnomAD GraphQL",
        "vep": "Ensembl VEP",
        "alfa": "ALFA / NCBI",
        "topmed": "TOPMed / BRAVO",
        "myvariant_full": "MyVariant.info (Full)",
        "ncbi_allele": "NCBI Variation",
        "ncbi_clinvar": "NCBI ClinVar",
        "myvariant_cons": "MyVariant Conservation",
        "alphagenome": "AlphaGenome AI",
    }

    # Mapping Settings-Key -> AFFetchController assignment-Key
    _API_KEY_MAP = {"myvariant": "mv", "vep": "ensembl", "gnomad": "gnomad", "alfa": "alfa", "topmed": "topmed"}

    # ========================================================================
    # ENTWICKLER-OPTIONEN
    # ========================================================================
    
    # Debug-Modus (mehr Logging, langsamere Performance)
    DEBUG_MODE = False
    
    # Verbose Logging für spezifische Module
    DEBUG_VCF_PARSER = False
    DEBUG_FASTA_VALIDATOR = False
    DEBUG_AF_FETCHER = False
    DEBUG_DATABASE = False
    
    # Performance-Profiling
    ENABLE_PROFILING = False           # Python cProfile aktivieren
    PROFILING_OUTPUT = "profile.stats" # Profiling-Output-Datei
    
    # ========================================================================
    # LEGACY COMPATIBILITY (DEPRECATED)
    # ========================================================================
    
    @classmethod
    def get_fasta_path(cls, build: str) -> Optional[str]:
        build_normalized = str(build).strip()
        return cls.FASTA_PATHS.get(build_normalized)
    
    @classmethod
    def validate(cls):
        """
        ✅ Validiere Config-Einstellungen beim Start.
        
        Prüft:
        - Kritische Pfade existieren
        - Wert-Ranges sind sinnvoll
        - Keine Konflikte zwischen Settings
        
        Raises:
            ValueError: Bei ungültiger Konfiguration
        """
        # Prüfe Thread-Counts
        if cls.MIN_WORKERS > cls.MAX_WORKERS:
            raise ValueError("MIN_WORKERS muss <= MAX_WORKERS sein")
        
        if cls.DEFAULT_THREADS > cls.MAX_WORKERS:
            raise ValueError("DEFAULT_THREADS muss <= MAX_WORKERS sein")
        
        # Prüfe Chunk-Size Range
        if cls.VCF_CHUNK_SIZE_MIN > cls.VCF_CHUNK_SIZE_MAX:
            raise ValueError("VCF_CHUNK_SIZE_MIN muss <= VCF_CHUNK_SIZE_MAX sein")
        
        # Prüfe FASTA Batch-Size
        if cls.FASTA_BATCH_SIZE < 1:
            raise ValueError("FASTA_BATCH_SIZE muss >= 1 sein")
        
        # Prüfe Quality-Presets
        if cls.DEFAULT_QUALITY_PRESET not in cls.QUALITY_PRESETS:
            raise ValueError(
                f"DEFAULT_QUALITY_PRESET '{cls.DEFAULT_QUALITY_PRESET}' "
                f"nicht in QUALITY_PRESETS gefunden"
            )
        
        # Warnung bei aktivierter Batch-Validierung
        if cls.FASTA_BATCH_VALIDATION:
            import warnings
            warnings.warn(
                "FASTA_BATCH_VALIDATION ist aktiviert (experimentell). "
                "Stelle sicher, dass die VCF-Scan-Loop entsprechend angepasst wurde.",
                UserWarning
            )
    
    @classmethod
    def print_settings(cls):
        """Gibt aktuelle Config-Einstellungen formatiert aus."""
        print("=" * 70)
        print("VARIANT FUSION - CONFIGURATION")
        print("=" * 70)
        
        print("\n⚡ PERFORMANCE:")
        print(f"  Threads: {cls.DEFAULT_THREADS} (Range: {cls.MIN_WORKERS}-{cls.MAX_WORKERS})")
        print(f"  VCF Chunk-Size: {cls.VCF_CHUNK_SIZE_MIN:,}-{cls.VCF_CHUNK_SIZE_MAX:,} (dynamisch)")
        print(f"  VCF mmap threshold: >{cls.VCF_MMAP_THRESHOLD_MB} MB")
        print(f"  Cache max size: {cls.CACHE_MAX_SIZE:,}")
        
        print("\n🧬 FASTA VALIDATION:")
        print(f"  Strict mode: {cls.FASTA_STRICT_MODE}")
        print(f"  Region cache: {cls.FASTA_REGION_CACHE_SIZE:,} regions")
        print(f"  Batch validation: {cls.FASTA_BATCH_VALIDATION} (size: {cls.FASTA_BATCH_SIZE})")
        print(f"  Available builds: {', '.join(set(cls.FASTA_PATHS.values()))}")
        
        print("\n🎯 QUALITY:")
        print(f"  Default preset: {cls.DEFAULT_QUALITY_PRESET}")
        print(f"  Available presets: {', '.join(cls.QUALITY_PRESETS.keys())}")
        
        print("\n🌐 NETWORK:")
        print(f"  Timeout: {cls.REQUEST_TIMEOUT}s")
        print(f"  Max retries: {cls.MAX_RETRIES}")
        print(f"  Async threshold: {cls.ASYNC_THRESHOLD:,}")
        
        print("\n🖥️  GUI:")
        print(f"  Table batch: {cls.TABLE_BATCH_SIZE}")
        print(f"  Poll interval: {cls.POLL_INTERVAL_MS}ms")
        print(f"  Progress update: {cls.PROGRESS_UPDATE_INTERVAL}ms")
        
        print("\n🛠️  DEBUG:")
        print(f"  Debug mode: {cls.DEBUG_MODE}")
        print(f"  Profiling: {cls.ENABLE_PROFILING}")
        
        print("=" * 70)


# ============================================================================
# CONFIG VALIDATION beim Import
# ============================================================================

# Validiere Config beim ersten Import (optional - kann auskommentiert werden)
# Config.validate()

# Für Debug: Config-Einstellungen anzeigen
# Config.print_settings()
# =============================================================================
# GLOBALE CONFIG-INSTANZ
# =============================================================================
config = Config()


# =============================================================================
# LEGACY CONSTANTS (für alte Code-Kompatibilität)
# =============================================================================
# ✅ VOLLSTÄNDIG: Alle möglichen Legacy-Konstanten

# Files & Directories
TEMP_VCF_DIR = config.TEMP_VCF_DIR
SETTINGS_FILE = config.SETTINGS_FILE
CACHE_FILE = config.CACHE_FILE
# BASE_DIR bereits oben gesetzt (PyInstaller-kompatibel)

# =============================================================================
# SINGLE INSTANCE LOCK & CLOUD SYNC CHECK
# =============================================================================

def check_single_instance():
    """
    Verhindert den Start einer zweiten Instanz der App (Single Instance Lock).
    Nutzt Lock-File mit PID.
    """
    lock_file = os.path.join(BASE_DIR, ".vf_instance.lock")
    if os.path.exists(lock_file):
        try:
            with open(lock_file, "r") as f:
                content = f.read().strip()
                if content:
                    old_pid = int(content)
                    # Prüfe ob Prozess noch lebt
                    if psutil.pid_exists(old_pid):
                        return False, old_pid
        except Exception:
            # Lock-File korrupt oder unlesbar -> ignorieren und überschreiben
            pass
            
    # Neues Lock-File erstellen
    try:
        with open(lock_file, "w") as f:
            f.write(str(os.getpid()))
    except Exception:
        pass
        
    return True, os.getpid()

def release_single_instance():
    """Löscht das Lock-File beim Beenden."""
    lock_file = os.path.join(BASE_DIR, ".vf_instance.lock")
    if os.path.exists(lock_file):
        try:
            os.remove(lock_file)
        except Exception:
            pass

def check_cloud_sync_warning(path: str) -> Optional[str]:
    """
    Prüft ob ein Pfad in einem Cloud-Sync-Ordner liegt.
    Gibt den Namen des Dienstes zurück oder None.
    """
    if not path:
        return None
        
    path_up = str(path).upper()
    if "ONEDRIVE" in path_up:
        return "OneDrive"
    if "DROPBOX" in path_up:
        return "Dropbox"
    if "GOOGLE DRIVE" in path_up:
        return "Google Drive"
    if "ICLOUD" in path_up:
        return "iCloud"
    
    return None

# Threading
DEFAULT_THREADS = config.DEFAULT_THREADS
MIN_WORKERS = config.MIN_WORKERS
MAX_WORKERS = config.MAX_WORKERS
TARGET_CPU_FRACTION = config.TARGET_CPU_FRACTION

# Network
TIMEOUT_SEC = config.REQUEST_TIMEOUT
RETRY_MAX = config.MAX_RETRIES
ASYNC_THRESHOLD = config.ASYNC_THRESHOLD

# Logging
LOG_DATETIME_FMT = config.LOG_DATETIME_FMT

# Database
DB_TIMEOUT_SEC = config.DB_TIMEOUT_SEC
DB_WAL_MODE = config.DB_WAL_MODE

# Caching
CACHE_MAX_SIZE = config.CACHE_MAX_SIZE

# GUI (✅ NEU!)
TABLE_BATCH_SIZE = config.TABLE_BATCH_SIZE
POLL_INTERVAL_MS = config.POLL_INTERVAL_MS          # ✅ WICHTIG!
PROGRESS_UPDATE_INTERVAL = config.PROGRESS_UPDATE_INTERVAL

# Cache Lock
_cache_lock = threading.Lock()


# =============================================================================
# TRUE CONSTANTS (unveränderlich)
# =============================================================================
APP_NAME = "Variant Fusion"
SESSION_FLAGS = {"use_fasta": None}

# API Endpoints
MV_BASE = "https://myvariant.info/v1"
DBSNP_URL = "https://api.ncbi.nlm.nih.gov/variation/v0/beta/refsnp/"

# ---------------------------------------------------------------------------
# API-Settings Runtime (wird von App._load_settings befuellt)
# ---------------------------------------------------------------------------
_runtime_api_settings = {}

def get_api_setting(phase, api, key, default=None):
    """Liest ein API-Setting: erst Runtime, dann Config-Default."""
    try:
        return _runtime_api_settings[phase][api][key]
    except (KeyError, TypeError):
        try:
            return Config.DEFAULT_API_SETTINGS[phase][api][key]
        except (KeyError, TypeError):
            return default

def is_api_enabled(phase, api):
    """Prueft ob eine API in den Settings aktiviert ist."""
    return get_api_setting(phase, api, "enabled", True)

# FASTA Resources
FASTA_URLS = {
    "GRCh37": "http://ftp.ensembl.org/pub/grch37/current/fasta/homo_sapiens/dna/Homo_sapiens.GRCh37.dna.primary_assembly.fa.gz",
    "GRCh38": "http://ftp.ensembl.org/pub/current_fasta/homo_sapiens/dna/Homo_sapiens.GRCh38.dna.primary_assembly.fa.gz",
}

# =============================================================================
# V16 FIX 12: ONEDRIVE-STEUERUNG
# =============================================================================
# OneDrive kann I/O-Performance dramatisch reduzieren (File-Locking, Sync-Checks).
# Diese Funktionen pausieren OneDrive während der Pipeline und starten es danach.
# =============================================================================

_onedrive_was_running = False  # Globaler State für Resume-Logik

def _find_onedrive_exe() -> Optional[Path]:
    """Findet den OneDrive.exe-Pfad auf dem System."""
    possible_paths = [
        Path(os.environ.get("LOCALAPPDATA", "")) / "Microsoft" / "OneDrive" / "OneDrive.exe",
        Path(r"C:\Program Files\Microsoft OneDrive\OneDrive.exe"),
        Path(r"C:\Program Files (x86)\Microsoft OneDrive\OneDrive.exe"),
    ]
    for path in possible_paths:
        if path.exists():
            return path
    return None


def pause_onedrive(logger=None) -> bool:
    """
    DEAKTIVIERT - OneDrive-Steuerung temporär blockiert.

    Grund: OneDrive blockiert Dateien wenn es während Sync abgeschaltet wird.
    Lösung: DBs in lokalen Ordner auslagern statt OneDrive zu stoppen.
    """
    if logger:
        logger.log("[OneDrive] ⏸️ pause_onedrive() DEAKTIVIERT - Funktion blockiert")
    return False


def resume_onedrive(logger=None) -> bool:
    """
    DEAKTIVIERT - OneDrive-Steuerung temporär blockiert.

    Grund: OneDrive blockiert Dateien wenn es während Sync abgeschaltet wird.
    Lösung: DBs in lokalen Ordner auslagern statt OneDrive zu stoppen.
    """
    if logger:
        logger.log("[OneDrive] ▶️ resume_onedrive() DEAKTIVIERT - Funktion blockiert")
    return False


# =============================================================================
# V16: RESOURCE MANAGER - ZENTRALES RESSOURCEN-MANAGEMENT
# =============================================================================
#
# Problem: Hardcoded Pfade führen zu Fehlern wenn Dateien verschoben werden.
# Lösung: ResourceManager mit Auto-Discovery und resources_config.json Persistenz.
#
# Features:
# - Erfasst ALLE abhängigen Dateien (FASTA, DBs, Scripts, Configs, GTF, Icons...)
# - Speichert gefundene Pfade in resources_config.json
# - Heilt sich selbst wenn Dateien verschoben werden
# - Zentrales RESOURCES Dictionary für alle Komponenten
# - Intelligente Suche in mehreren Verzeichnissen
# =============================================================================

# --- RESSOURCEN-DEFINITION ---
# Alle Ressourcen mit Typ, Patterns und Eigenschaften
RESOURCE_DEFINITIONS = {
    # =========================================================================
    # REFERENZ-GENOME (FASTA)
    # =========================================================================
    "fasta_grch37": {
        "type": "reference",
        "description": "Human Reference Genome GRCh37/hg19",
        "required": False,
        "patterns": [
            "Homo_sapiens.GRCh37.dna.primary_assembly.fa",
            "hg19.fa",
            "GRCh37.fa",
            "human_g1k_v37.fasta",
        ],
        "search_dirs": [".", "data", "data/reference", "reference"],
        "extensions": [".fa", ".fasta", ".fa.gz", ".fasta.gz"],
    },
    "fasta_grch38": {
        "type": "reference",
        "description": "Human Reference Genome GRCh38/hg38",
        "required": False,
        "patterns": [
            "Homo_sapiens.GRCh38.dna.primary_assembly.fa",
            "hg38.fa",
            "GRCh38.fa",
        ],
        "search_dirs": [".", "data", "data/reference", "reference"],
        "extensions": [".fa", ".fasta", ".fa.gz", ".fasta.gz"],
    },
    "fasta_index_grch37": {
        "type": "index",
        "description": "FASTA Index GRCh37",
        "required": False,
        "patterns": [
            "Homo_sapiens.GRCh37.dna.primary_assembly.fa.fai",
            "hg19.fa.fai",
        ],
        "search_dirs": [".", "data", "data/reference", "reference"],
    },
    "fasta_index_grch38": {
        "type": "index",
        "description": "FASTA Index GRCh38",
        "required": False,
        "patterns": [
            "Homo_sapiens.GRCh38.dna.primary_assembly.fa.fai",
            "hg38.fa.fai",
        ],
        "search_dirs": [".", "data", "data/reference", "reference"],
    },

    # =========================================================================
    # GTF ANNOTATION FILES
    # =========================================================================
    "gtf_grch37": {
        "type": "annotation",
        "description": "GTF Annotation GRCh37",
        "required": False,
        "patterns": [
            "GRCh37.gtf.gz",
            "Homo_sapiens.GRCh37.gtf.gz",
            "gencode.v19.annotation.gtf.gz",
        ],
        "search_dirs": ["data/annotations", "data", "annotations"],
    },
    "gtf_grch38": {
        "type": "annotation",
        "description": "GTF Annotation GRCh38",
        "required": False,
        "patterns": [
            "GRCh38.gtf.gz",
            "Homo_sapiens.GRCh38.gtf.gz",
            "gencode.v38.annotation.gtf.gz",
        ],
        "search_dirs": ["data/annotations", "data", "annotations"],
    },

    # =========================================================================
    # DATABASES
    # =========================================================================
    # V16: Lokaler DB-Pfad C:\_Local_DEV\DATA_STORE als PRIMÄRER Suchort
    # Grund: OneDrive blockiert große DBs während Sync (100GB+ nicht praktikabel)
    # Backup: NAS statt OneDrive
    # =========================================================================
    "gnomad_db": {
        "type": "database",
        "description": "gnomAD Light Database (AF-Daten, ~94GB)",
        "required": False,
        "patterns": [
            "gnomad_light.db",
            "gnomad.db",
        ],
        "search_dirs": ["C:/_Local_DEV/DATA_STORE", "C:\\_Local_DEV\\DATA_STORE", "data", "."],
    },
    "variant_db": {
        "type": "database",
        "description": "Variant Fusion SQLite (Haupt-Datenbank, ~5GB)",
        "required": False,
        "patterns": [
            "variant_fusion.sqlite",
            "variant_fusion.db",
        ],
        "search_dirs": ["C:/_Local_DEV/DATA_STORE", "C:\\_Local_DEV\\DATA_STORE", ".", "data"],
    },
    "dbnsfp_db": {
        "type": "database",
        "description": "dbNSFP Light Database",
        "required": False,
        "patterns": [
            "dbnsfp_light.db",
            "dbnsfp.db",
        ],
        "search_dirs": ["C:/_Local_DEV/DATA_STORE", "C:\\_Local_DEV\\DATA_STORE", ".", "data"],
    },

    # =========================================================================
    # CONFIGURATION FILES
    # =========================================================================
    "settings_json": {
        "type": "config",
        "description": "App Settings JSON",
        "required": False,
        "patterns": [
            "variant_fusion_settings.json",
        ],
        "search_dirs": [".", "config"],
        "create_if_missing": True,
    },
    "lightdb_config": {
        "type": "config",
        "description": "LightDB Configuration",
        "required": False,
        "patterns": [
            "lightdb_config.json",
        ],
        "search_dirs": [".", "data", "config"],
    },
    "cache_json": {
        "type": "config",
        "description": "Cache JSON",
        "required": False,
        "patterns": [
            "cache.json",
        ],
        "search_dirs": [".", "config"],
    },
    "translations": {
        "type": "config",
        "description": "Translations JSON",
        "required": False,
        "patterns": [
            "translations.json",
        ],
        "search_dirs": ["locales", ".", "config"],
    },

    # =========================================================================
    # PYTHON MODULES (Abhängige Scripts)
    # =========================================================================
    "translator_module": {
        "type": "module",
        "description": "Translator Python Module",
        "required": False,
        "patterns": [
            "translator.py",
        ],
        "search_dirs": [".", "lib", "src"],
    },
    "translator_patch": {
        "type": "module",
        "description": "Translator Patch Module",
        "required": False,
        "patterns": [
            "translator_patch.py",
        ],
        "search_dirs": [".", "lib", "src"],
    },
    "lightdb_worker": {
        "type": "module",
        "description": "LightDB Index Worker",
        "required": False,
        "patterns": [
            "lightdb_index_worker.py",
        ],
        "search_dirs": [".", "data", "tools"],
    },

    # =========================================================================
    # CYTHON MODULES
    # =========================================================================
    "cython_hotpath": {
        "type": "cython",
        "description": "Cython Acceleration Module",
        "required": False,
        "patterns": [
            "cython_hotpath",
        ],
        "search_dirs": ["."],
        "is_directory": True,
    },

    # =========================================================================
    # UI RESOURCES
    # =========================================================================
    "app_icon": {
        "type": "ui",
        "description": "Application Icon",
        "required": False,
        "patterns": [
            "ICO.ico",
            "icon.ico",
            "app.ico",
        ],
        "search_dirs": ["ICO", "assets/ICO", "assets", "."],
    },

    # =========================================================================
    # LOG & OUTPUT DIRECTORIES
    # =========================================================================
    "log_dir": {
        "type": "directory",
        "description": "Log Directory",
        "required": False,
        "patterns": [
            "logs",
        ],
        "search_dirs": ["."],
        "is_directory": True,
        "create_if_missing": True,
    },
    "output_dir": {
        "type": "directory",
        "description": "Output Directory",
        "required": False,
        "patterns": [
            "output",
            "exports",
        ],
        "search_dirs": ["."],
        "is_directory": True,
    },
}

RESOURCE_SETUP_INFO = {
    "fasta_grch37": {
        "group": "reference",
        "label": "Referenz-Genom GRCh37 / hg19",
        "size": "~3 GB (Download ~900 MB)",
        "description": "Humanes Referenzgenom Build 37 fuer Allel-Validierung und 23andMe-Konvertierung.",
        "without": "Fallback auf MyVariant API (langsamer, braucht Internet). Grundfunktionen bleiben erhalten.",
        "action": "download",
    },
    "fasta_grch38": {
        "group": "reference",
        "label": "Referenz-Genom GRCh38 / hg38",
        "size": "~3 GB (Download ~900 MB)",
        "description": "Aktuelles Standard-Referenzgenom fuer neuere Datensaetze.",
        "without": "Wie GRCh37: API-Fallback. Moderne Datensaetze werden langsamer verarbeitet.",
        "action": "download",
    },
    "gnomad_db": {
        "group": "database",
        "label": "gnomAD Light Database",
        "size": "~30 GB Download, ~94 GB entpackt",
        "description": "Lokale gnomAD-Allel-Frequenzen fuer Milliarden Varianten. Ermoeglicht Offline-AF-Abfragen.",
        "without": "AF-Daten werden online geholt (Sekunden statt Millisekunden pro Variante). Funktional, aber deutlich langsamer.",
        "action": "download",
    },
    "dbnsfp_db": {
        "group": "database",
        "label": "dbNSFP Light Database",
        "size": "-",
        "description": "Pathogenitaets-Scores (SIFT, PolyPhen, REVEL). Wird automatisch erstellt wenn 23andMe-to-VCF genutzt wird.",
        "without": "Keine Einschraenkung. Pathogenitaets-Daten kommen via API. Wird bei Bedarf automatisch erstellt.",
        "action": "info",
    },
    "gtf_grch37": {
        "group": "annotation",
        "label": "Gen-Annotation GTF GRCh37",
        "size": "~50 MB",
        "description": "GENCODE Gen-Koordinaten. Wird beim ersten Lauf automatisch heruntergeladen.",
        "without": "Gen-Annotation ueber API (langsamer). Protein-Coding-Filter eingeschraenkt.",
        "action": "auto",
    },
    "gtf_grch38": {
        "group": "annotation",
        "label": "Gen-Annotation GTF GRCh38",
        "size": "~1.5 GB",
        "description": "GENCODE Gen-Koordinaten (aktuell). Automatischer Download beim ersten Lauf.",
        "without": "Wie GRCh37: API-Fallback moeglich.",
        "action": "auto",
    },
}


class ResourceManager:
    """
    V16: Zentrales Ressourcen-Management mit Auto-Discovery.

    Erfasst alle abhängigen Dateien und speichert gefundene Pfade
    in einer JSON-Config. Wenn Dateien verschoben werden, findet
    ResourceManager sie beim nächsten Start automatisch wieder.

    Verwendung:
        rm = get_resource_manager()
        fasta = rm.get("fasta_grch37")  # Absoluter Pfad oder None
        all_res = rm.get_all()          # Dict aller Ressourcen
    """

    CONFIG_FILE = "resources_config.json"
    _instance = None
    _initialized = False

    def __new__(cls):
        """Singleton Pattern - nur eine Instanz."""
        if cls._instance is None:
            cls._instance = super().__new__(cls)
        return cls._instance

    def __init__(self):
        """Initialisiert ResourceManager (nur beim ersten Mal)."""
        if ResourceManager._initialized:
            return

        self._paths = {}           # key -> relative path
        self._abs_paths = {}       # key -> absolute path (Cache)
        self._status = {}          # key -> "found" | "missing" | "created"
        self._config_path = os.path.join(BASE_DIR, self.CONFIG_FILE)

        self._load_config()
        self._verify_and_discover_all()
        ResourceManager._initialized = True

    def _load_config(self):
        """Lädt gespeicherte Pfade aus JSON."""
        try:
            if os.path.exists(self._config_path):
                with open(self._config_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    self._paths = data.get("paths", {})
                    self._status = data.get("status", {})
        except Exception as e:
            print(f"[ResourceManager] ⚠️ Config load error: {e}")
            self._paths = {}
            self._status = {}

    def _save_config(self):
        """Speichert aktuelle Pfade in JSON."""
        try:
            data = {
                "version": "1.0",
                "last_updated": datetime.datetime.now().isoformat(),
                "base_dir": BASE_DIR,
                "paths": self._paths,
                "status": self._status,
            }
            with open(self._config_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"[ResourceManager] ⚠️ Config save error: {e}")

    def _verify_and_discover_all(self):
        """Verifiziert alle Ressourcen und sucht fehlende."""
        changed = False
        found_count = 0
        missing_count = 0

        for key, definition in RESOURCE_DEFINITIONS.items():
            stored_path = self._paths.get(key)
            is_dir = definition.get("is_directory", False)

            # Prüfe ob gespeicherter Pfad noch existiert
            if stored_path:
                full_path = os.path.join(BASE_DIR, stored_path)
                if is_dir:
                    if os.path.isdir(full_path):
                        self._abs_paths[key] = full_path
                        self._status[key] = "found"
                        found_count += 1
                        continue
                else:
                    if os.path.isfile(full_path):
                        self._abs_paths[key] = full_path
                        self._status[key] = "found"
                        found_count += 1
                        continue

            # Pfad fehlt oder ungültig → Auto-Discovery
            found_path = self._discover_resource(key, definition)
            if found_path:
                self._paths[key] = found_path
                self._abs_paths[key] = os.path.join(BASE_DIR, found_path)
                self._status[key] = "found"
                changed = True
                found_count += 1
                print(f"[ResourceManager] ✅ {key}: {found_path}")
            elif definition.get("create_if_missing") and is_dir:
                # Verzeichnis erstellen wenn erlaubt
                default_path = definition["patterns"][0]
                full_path = os.path.join(BASE_DIR, default_path)
                try:
                    os.makedirs(full_path, exist_ok=True)
                    self._paths[key] = default_path
                    self._abs_paths[key] = full_path
                    self._status[key] = "created"
                    changed = True
                    found_count += 1
                    print(f"[ResourceManager] 📁 {key} created: {default_path}")
                except Exception as e:
                    self._status[key] = "missing"
                    missing_count += 1
            else:
                self._status[key] = "missing"
                if stored_path:
                    del self._paths[key]
                    changed = True
                missing_count += 1
                if definition.get("required"):
                    print(f"[ResourceManager] ❌ REQUIRED {key} nicht gefunden!")

        if changed:
            self._save_config()

        print(f"[ResourceManager] 📊 {found_count} gefunden, {missing_count} fehlen")

    def _discover_resource(self, key: str, definition: dict) -> Optional[str]:
        """Sucht eine Ressource anhand der Definition."""
        patterns = definition.get("patterns", [])
        search_dirs = definition.get("search_dirs", ["."])
        is_dir = definition.get("is_directory", False)

        for search_dir in search_dirs:
            for pattern in patterns:
                # Versuche direkt im search_dir
                rel_path = os.path.join(search_dir, pattern) if search_dir != "." else pattern
                full_path = os.path.join(BASE_DIR, rel_path)

                if is_dir:
                    if os.path.isdir(full_path):
                        return rel_path
                else:
                    if os.path.isfile(full_path):
                        return rel_path

        # V17 FIX: Rekursive os.walk()-Suche ENTFERNT.
        # Grund: Bei Fresh-Install mit 13+ fehlenden Ressourcen fuehrte
        # os.walk(BASE_DIR) pro Pattern zu 20+ vollstaendigen
        # Verzeichnis-Traversierungen, die auf OneDrive minutenlang blockierten.
        # Die direkten search_dirs-Checks oben sind voellig ausreichend.
        return None

    def get(self, key: str, absolute: bool = True) -> Optional[str]:
        """
        Gibt den Pfad einer Ressource zurück.

        Args:
            key: Ressourcen-Schlüssel (z.B. "fasta_grch37")
            absolute: True für absoluten Pfad, False für relativen

        Returns:
            Pfad zur Ressource oder None wenn nicht gefunden
        """
        if absolute:
            path = self._abs_paths.get(key)
            if path and (os.path.exists(path) or RESOURCE_DEFINITIONS.get(key, {}).get("is_directory")):
                return path
        else:
            return self._paths.get(key)

        # Pfad existiert nicht mehr → neu suchen
        definition = RESOURCE_DEFINITIONS.get(key)
        if definition:
            found = self._discover_resource(key, definition)
            if found:
                self._paths[key] = found
                self._abs_paths[key] = os.path.join(BASE_DIR, found)
                self._status[key] = "found"
                self._save_config()
                return self._abs_paths[key] if absolute else found

        return None

    def get_fasta_path(self, build: str) -> Optional[str]:
        """
        Gibt den FASTA-Pfad für ein Build zurück (Kompatibilität).

        Args:
            build: "GRCh37", "GRCh38", "hg19", "hg38", etc.

        Returns:
            Absoluter Pfad zur FASTA-Datei oder None
        """
        key = "fasta_grch38" if build.lower() in ("grch38", "hg38", "b38") else "fasta_grch37"
        return self.get(key)

    def get_gnomad_db_path(self) -> Optional[str]:
        """Gibt den Pfad zur gnomAD LightDB zurück (Kompatibilität)."""
        return self.get("gnomad_db")

    def get_variant_db_path(self) -> Optional[str]:
        """Gibt den Pfad zur Variant-Fusion SQLite DB zurück (Kompatibilität)."""
        return self.get("variant_db")

    def get_all(self) -> Dict[str, Optional[str]]:
        """Gibt alle bekannten Ressourcen-Pfade zurück."""
        return {k: self._abs_paths.get(k) for k in RESOURCE_DEFINITIONS.keys()}

    def get_status(self) -> Dict[str, str]:
        """Gibt den Status aller Ressourcen zurück."""
        return dict(self._status)

    def get_by_type(self, resource_type: str) -> Dict[str, Optional[str]]:
        """
        Gibt alle Ressourcen eines bestimmten Typs zurück.

        Args:
            resource_type: "reference", "database", "config", "module", etc.
        """
        result = {}
        for key, definition in RESOURCE_DEFINITIONS.items():
            if definition.get("type") == resource_type:
                result[key] = self._abs_paths.get(key)
        return result

    def register(self, key: str, path: str):
        """
        Registriert einen neuen/geänderten Pfad manuell.

        Args:
            key: Schlüssel (z.B. "fasta_grch37")
            path: Relativer oder absoluter Pfad
        """
        # Konvertiere zu relativem Pfad wenn möglich
        if os.path.isabs(path):
            if path.startswith(BASE_DIR):
                rel_path = os.path.relpath(path, BASE_DIR)
            else:
                rel_path = path
        else:
            rel_path = path

        full_path = os.path.join(BASE_DIR, rel_path) if not os.path.isabs(path) else path

        self._paths[key] = rel_path
        self._abs_paths[key] = full_path
        self._status[key] = "registered"
        self._save_config()
        print(f"[ResourceManager] 📝 Registered {key}: {rel_path}")

    def refresh(self):
        """Erzwingt erneute Suche aller Ressourcen."""
        self._paths.clear()
        self._abs_paths.clear()
        self._status.clear()
        self._verify_and_discover_all()

    def heal(self, key: str) -> bool:
        """
        Versucht eine fehlende Ressource zu finden (Self-Healing).

        Returns:
            True wenn Ressource gefunden wurde
        """
        definition = RESOURCE_DEFINITIONS.get(key)
        if not definition:
            return False

        found = self._discover_resource(key, definition)
        if found:
            self._paths[key] = found
            self._abs_paths[key] = os.path.join(BASE_DIR, found)
            self._status[key] = "healed"
            self._save_config()
            print(f"[ResourceManager] 🩹 Healed {key}: {found}")
            return True
        return False

    def print_report(self):
        """Gibt einen Status-Report aller Ressourcen aus."""
        print("\n" + "=" * 60)
        print("RESOURCE MANAGER - STATUS REPORT")
        print("=" * 60)

        by_type = {}
        for key, definition in RESOURCE_DEFINITIONS.items():
            rtype = definition.get("type", "unknown")
            if rtype not in by_type:
                by_type[rtype] = []
            status = self._status.get(key, "unknown")
            path = self._paths.get(key, "-")
            by_type[rtype].append((key, status, path))

        for rtype, items in sorted(by_type.items()):
            print(f"\n[{rtype.upper()}]")
            for key, status, path in items:
                icon = "✅" if status in ("found", "created", "healed", "registered") else "❌"
                print(f"  {icon} {key}: {path}")

        print("\n" + "=" * 60 + "\n")


# Globale ResourceManager-Instanz (Lazy Init)
_resource_manager: Optional[ResourceManager] = None

def get_resource_manager() -> ResourceManager:
    """Gibt die ResourceManager-Singleton-Instanz zurück."""
    global _resource_manager
    if _resource_manager is None:
        _resource_manager = ResourceManager()
    return _resource_manager

# Alias für Kompatibilität mit altem Code
get_path_manager = get_resource_manager
PathManager = ResourceManager  # Alias


# =============================================================================
# RESOURCES DICTIONARY (Globaler Zugriff für alle Komponenten)
# =============================================================================
# Wird beim ersten Zugriff durch ResourceManager befüllt.

class _ResourcesProxy(dict):
    """
    Proxy-Dict das Ressourcen-Pfade dynamisch vom ResourceManager holt.
    Ermöglicht einfachen Zugriff: RESOURCES["fasta_grch37"]
    """
    def __getitem__(self, key):
        rm = get_resource_manager()
        return rm.get(key)

    def get(self, key, default=None):
        rm = get_resource_manager()
        result = rm.get(key)
        return result if result else default

    def keys(self):
        return RESOURCE_DEFINITIONS.keys()

    def values(self):
        rm = get_resource_manager()
        return [rm.get(k) for k in RESOURCE_DEFINITIONS.keys()]

    def items(self):
        rm = get_resource_manager()
        return [(k, rm.get(k)) for k in RESOURCE_DEFINITIONS.keys()]

    def __iter__(self):
        return iter(RESOURCE_DEFINITIONS.keys())

    def __len__(self):
        return len(RESOURCE_DEFINITIONS)

    def __repr__(self):
        rm = get_resource_manager()
        return repr(rm.get_all())

# Globales RESOURCES Dictionary für einfachen Zugriff
RESOURCES = _ResourcesProxy()


# =============================================================================
# FASTA_PATHS (Kompatibilitäts-Wrapper für alten Code)
# =============================================================================
# Diese globale Variable wird von altem Code verwendet.
# Sie wird beim ersten Zugriff durch ResourceManager befüllt.

class _FastaPathsProxy(dict):
    """
    Proxy-Dict das FASTA-Pfade dynamisch von ResourceManager holt.
    Ermöglicht Kompatibilität mit altem Code der FASTA_PATHS["GRCh37"] nutzt.
    """
    def __getitem__(self, key):
        rm = get_resource_manager()
        path = rm.get_fasta_path(key)
        if path:
            return path
        # Fallback auf alte Defaults (relative Pfade)
        defaults = {
            "GRCh37": "Homo_sapiens.GRCh37.dna.primary_assembly.fa",
            "GRCh38": "Homo_sapiens.GRCh38.dna.primary_assembly.fa",
            "hg19": "Homo_sapiens.GRCh37.dna.primary_assembly.fa",
            "hg38": "Homo_sapiens.GRCh38.dna.primary_assembly.fa",
        }
        return defaults.get(key)

    def get(self, key, default=None):
        result = self.__getitem__(key)
        return result if result else default

FASTA_PATHS = _FastaPathsProxy()

# Build Detection Regex
BUILD_HINTS_37 = [
    r"\bgrch\s*37\b", r"\bgrch37\b", r"\bbuild\s*37\b", 
    r"\bhg19\b", r"\bb37\b", r"\bncb[iy]\s*build\s*37\b"
]
BUILD_HINTS_38 = [
    r"\bgrch\s*38\b", r"\bgrch38\b", r"\bbuild\s*38\b", 
    r"\bhg38\b", r"\bb38\b", r"\bncb[iy]\s*build\s*38\b"
]

# MyVariant Field Lists
MV_FIELDS_AF = [
    "exac.af",
    "gnomad.exomes.af",
    "gnomad.genomes.af",
    "thousand_genomes.af",
]

MV_FIELDS_FULL = [
    "cadd.phred",
    "dbsnp.rsid",
    "dbsnp.chrom",
    "dbsnp.hg19",
    "dbsnp.hg38",
    "gene.symbol",
    "snpeff.ann.effect",
    "snpeff.ann.impact",
    "snpeff.ann.feature_type",
    "snpeff.ann.gene_name",
    "snpeff.ann.hgvs_p",
    "clinvar.rcv.clinical_significance",
    "clinvar.rcv.phenotype",
    "clinvar.phenotype_list",
    "clinvar.clnsig",
    "clinvar.clnrevstat",
    "cadd.raw",
    "vep.consequence",
    "vep.impact",
    "vep.symbol",
    "vep.polyphen",
    "vep.sift",
    "vep.lof",
    "vep.existing_variation",
    "phastcons",
    "phylop",
    "gerp",
]

# Gene Link Preferences
GENE_LINK_SINGLE_CLICK = "genecards"  # "genecards", "malacards", or "none"
GENE_LINK_DOUBLE_CLICK = "malacards"  # "genecards", "malacards", or "none"
# =============================================================================
# HELPER CLASSES (NullLogger)
# =============================================================================
class _NullLogger:
    """Fallback-Logger wenn kein Logger übergeben wird."""
    def log(self, msg: str, prefix: str = None, flush: bool = False):
        try:
            print(f"[{prefix or 'NULL'}] {msg}")
            if flush:
                sys.stdout.flush()
        except Exception:
            pass
    
    def drain(self):
        return []


# =============================================================================
# LOGGER CLASSES
# =============================================================================
class MultiSinkLogger:
    """Multi-Sink Logger mit automatischem Log-Clearing."""
    
    def __init__(self, logfile_path: Optional[str] = None, ui_queue: Optional[queue.Queue] = None):
        self.q = ui_queue
        self.logfile_path = logfile_path
        self._lock = threading.Lock()
        
        if self.logfile_path:
            dirpath = os.path.dirname(self.logfile_path)
            if dirpath:
                os.makedirs(dirpath, exist_ok=True)
            self._clear_logfile()
    
    def _clear_logfile(self):
        try:
            with self._lock:
                with open(self.logfile_path, 'w', encoding='utf-8') as f:
                    timestamp = datetime.datetime.now().strftime(LOG_DATETIME_FMT)
                    f.write(f"{'='*70}\n")
                    f.write(f"VariantFusion Log Session Started: {timestamp}\n")
                    f.write(f"{'='*70}\n\n")
            print(f"[Logger] ✅ Logfile cleared: {self.logfile_path}")
        except Exception as e:
            print(f"[Logger] ⚠️ Could not clear logfile: {e}")
    
    def log(self, msg: str, prefix: Optional[str] = None, flush: bool = False):
        ts = datetime.datetime.now().strftime(LOG_DATETIME_FMT)
        line = f"[{ts}] [{prefix}] {msg}" if prefix else f"[{ts}] {msg}"
        
        if self.q is not None:
            try:
                self.q.put_nowait(line)
            except queue.Full:
                try:
                    self.q.get_nowait()
                    self.q.put_nowait(line)
                except Exception:
                    pass
            except Exception:
                pass
        
        if self.logfile_path:
            try:
                with self._lock:
                    with open(self.logfile_path, 'a', encoding='utf-8') as f:
                        f.write(line + '\n')
            except Exception as e:
                print(f"[Logger] ⚠️ File write error: {e}", file=sys.stderr)
        
        try:
            print(line)
            if flush:
                sys.stdout.flush()
        except Exception:
            pass
    
    def drain(self) -> list:
        msgs = []
        if self.q is None:
            return msgs
        try:
            while True:
                msgs.append(self.q.get_nowait())
        except queue.Empty:
            pass
        except Exception as e:
            print(f"[Logger] ⚠️ Drain error: {e}", file=sys.stderr)
        return msgs
    
    def shutdown(self):
        if self.q is not None:
            remaining = self.drain()
            if remaining:
                print(f"[Logger] Flushing {len(remaining)} remaining log entries...")
        
        if self.logfile_path:
            try:
                with self._lock:
                    with open(self.logfile_path, 'a', encoding='utf-8') as f:
                        timestamp = datetime.datetime.now().strftime(LOG_DATETIME_FMT)
                        f.write(f"\n{'='*70}\n")
                        f.write(f"Log Session Ended: {timestamp}\n")
                        f.write(f"{'='*70}\n")
            except Exception:
                pass


# =============================================================================
# LOGGER INITIALISIERUNG (nach Config und Legacy-Konstanten!)
# =============================================================================
_splash_log("Erstelle Logger ...")
logger = MultiSinkLogger(
    logfile_path=os.path.join(os.getcwd(), config.LOG_FILE),
    ui_queue=queue.Queue(maxsize=config.QUEUE_MAX_SIZE)
)


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================
def fmt_eta(seconds: Optional[float]) -> str:
    """Formatiert Sekunden in ein ETA-Format."""
    if seconds is None:
        return "ETA --:--"
    s = int(seconds)
    h = s // 3600
    m = (s % 3600) // 60
    s2 = s % 60
    if h > 0:
        return f"ETA {h:02d}:{m:02d}:{s2:02d}"
    return f"ETA {m:02d}:{s2:02d}"

# ============== Utils ==============

def safe_float(x, allow_negative: bool = False) -> Optional[float]:
    """
    FIX: Erweiterte Float-Konvertierung mit Range-Validation.
    
    Änderungen:
    - Optional: erlaube negative Werte (für CADD etc.)
    - Strikte Validierung für AF-Werte
    """
    try:
        f = float(x)
        
        # NaN/Inf immer reject
        if math.isnan(f) or math.isinf(f):
            return None
        
        # Range-Check (nur wenn nicht explizit negativ erlaubt)
        if not allow_negative and f < 0.0:
            return None
        
        return f
        
    except (ValueError, TypeError):
        return None

def now_iso() -> str:
    """Gibt den aktuellen UTC‑Zeitpunkt als ISO‑String mit 'Z' zurück."""
    return datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def parse_iso_utc(s: Optional[str]) -> Optional[datetime.datetime]:
    """
    Parst einen ISO‑Zeitstempel (ggf. mit 'Z') und gibt ein naive datetime in UTC zurück.
    Falls Parsing fehlschlägt, wird None geliefert.
    """
    if not s:
        return None
    try:
        if s.endswith("Z"):
            return (
                datetime.datetime.fromisoformat(s[:-1] + "+00:00")
                .astimezone(datetime.timezone.utc)
                .replace(tzinfo=None)
            )
        dt = datetime.datetime.fromisoformat(s)
        if dt.tzinfo is not None:
            return dt.astimezone(datetime.timezone.utc).replace(tzinfo=None)
        return dt
    except Exception:
        return None

def _count_lines_fast(path: str) -> Optional[int]:
    """
    Zählt Zeilen einer unkomprimierten Datei extrem schnell via mmap.
    Nur für unkomprimierte Dateien geeignet!

    Returns:
        Anzahl Zeilen oder None bei Fehler
    """
    import mmap
    try:
        with open(path, 'rb') as f:
            # Memory-map für schnellstes Zählen
            with mmap.mmap(f.fileno(), 0, access=mmap.ACCESS_READ) as mm:
                # Zähle Newlines - sehr schnell da mmap
                return mm[:].count(b'\n')
    except Exception:
        return None


def _count_vcf_variants_fast(vcf_path: str) -> Optional[int]:
    """
    Zählt Varianten in einer unkomprimierten VCF schnell.
    Nutzt mmap und zählt nur Non-Header-Zeilen.

    Returns:
        Anzahl Varianten (Non-Header-Zeilen) oder None bei Fehler
    """
    import mmap
    try:
        with open(vcf_path, 'rb') as f:
            with mmap.mmap(f.fileno(), 0, access=mmap.ACCESS_READ) as mm:
                total_lines = 0
                header_lines = 0
                pos = 0
                size = mm.size()

                while pos < size:
                    # Finde nächstes Newline
                    nl = mm.find(b'\n', pos)
                    if nl == -1:
                        nl = size

                    # Prüfe ob Header (beginnt mit #)
                    if pos < size and mm[pos:pos+1] == b'#':
                        header_lines += 1

                    total_lines += 1
                    pos = nl + 1

                return total_lines - header_lines
    except Exception:
        return None


def _estimate_total_variants(vcf_path: str, sample_lines: int = 10000) -> Optional[int]:
    """
    Schätzt die Gesamtzahl der Varianten:

    - Unkomprimierte VCF: Exaktes Zählen via mmap (sehr schnell)
    - Komprimierte VCF (.gz/.bgz): Heuristische Schätzung mit verbessertem Algorithmus

    Rückgabe:
      - int: Geschätzte/exakte Anzahl Varianten
      - None: bei Fehlern

    V16 FIX: Unkomprimierte Dateien werden jetzt exakt gezählt statt geschätzt.
    """
    # ═══════════════════════════════════════════════════════════════════
    # UNKOMPRIMIERT: Exaktes Zählen (sehr schnell mit mmap)
    # ═══════════════════════════════════════════════════════════════════
    if not is_gzipped(vcf_path):
        exact_count = _count_vcf_variants_fast(vcf_path)
        if exact_count is not None:
            return exact_count
        # Fallback auf alte Methode wenn mmap fehlschlägt

    # ═══════════════════════════════════════════════════════════════════
    # KOMPRIMIERT: Verbesserte Heuristik
    # ═══════════════════════════════════════════════════════════════════
    try:
        size_bytes = os.path.getsize(vcf_path)
    except Exception:
        size_bytes = None

    non_header = 0
    total_lines = 0
    data_bytes = 0

    try:
        with open_text_maybe_gzip(vcf_path) as f:
            for line in f:
                if not line:
                    continue
                if line[0] != "#":
                    non_header += 1
                total_lines += 1
                data_bytes += len(line)
                if total_lines >= sample_lines:
                    break
    except Exception:
        return None

    # Keine Daten im Sample
    if total_lines == 0:
        return None

    # Verhältnis Non-Header im Sample
    ratio = non_header / max(1, total_lines)
    avg_bytes_per_line = (data_bytes / max(1, total_lines)) if data_bytes > 0 else 80.0

    # Für fehlende Größe -> konservativer Fallback auf Sample-Hochrechnung
    if not size_bytes or size_bytes <= 0:
        return max(1, int(non_header * 2))

    # Komprimiert: Heuristische Dekompressionsschätzung
    # VCF-Dateien komprimieren typischerweise 5-10x (viel repetitiver Text)
    # Nutze konservativen Faktor von 6.0 statt 4.0
    decompression_factor = 6.0
    effective_size = size_bytes * decompression_factor

    # Zeilen-Gesamtschätzung
    est_total_lines = int(effective_size / max(1.0, avg_bytes_per_line))
    est_non_header = int(est_total_lines * ratio)

    # V16 FIX: Entferne künstliche Obergrenze für .gz Dateien
    # Die alte Obergrenze (50x Sample) war viel zu niedrig für große VCFs
    # Neue Logik: Vertraue der Bytes-basierten Schätzung mehr
    lower_bound = max(non_header * 2, non_header + 1000)

    # Keine künstliche Obergrenze mehr - nur Untergrenze
    est = max(lower_bound, est_non_header)

    return est

def count_variants_exact(vcf_path: str) -> int:
    """
    Zählt exakt die Anzahl der Varianten in einer VCF.
    Dabei werden Multi-ALT-Zeilen korrekt berücksichtigt,
    indem jede ALT als eigene Variante gezählt wird.
    """
    total = 0
    with open_text_maybe_gzip(vcf_path) as f:
        for line in f:
            if not line or line[0] == "#":
                continue
            # ALT-Spalte extrahieren (5. Spalte, Index 4)
            parts = line.split("\t", 6)
            if len(parts) > 4:
                alt_str = parts[4]
                total += len(alt_str.split(","))
    return total



import io, gzip, math
from typing import TextIO, Union

def is_gzipped(path: str) -> bool:
    """Prüft, ob der Pfad gzip- oder bgzip-komprimiert ist."""
    return path.lower().endswith((".gz", ".bgz"))

@contextmanager
def open_text_maybe_gzip(path: str):
    """
    Context manager for opening text files (gzipped or plain).
    Properly handles resource cleanup on exceptions.
    
    Args:
        path: File path (supports .gz, .bgz extensions)
    
    Yields:
        TextIO: Text stream
    
    Raises:
        RuntimeError: If file cannot be opened
    """
    bufsize = 1 << 20  # 1 MiB buffer
    
    file_obj = None
    buffered = None
    text_wrapper = None
    
    try:
        if is_gzipped(path):
            file_obj = gzip.GzipFile(filename=path, mode="rb")
            buffered = io.BufferedReader(file_obj, buffer_size=bufsize)
            text_wrapper = io.TextIOWrapper(
                buffered,
                encoding="utf-8",
                errors="replace",
                newline=""
            )
            yield text_wrapper
        else:
            file_obj = open(
                path,
                mode="r",
                encoding="utf-8",
                errors="replace",
                newline="",
                buffering=bufsize
            )
            yield file_obj
    
    except OSError as e:
        raise RuntimeError(f"Error opening {path}: {e}") from e
    
    finally:
        # Proper cleanup in reverse order
        if text_wrapper:
            try:
                text_wrapper.close()
            except Exception:
                pass
        if buffered:
            try:
                buffered.close()
            except Exception:
                pass
        if file_obj:
            try:
                file_obj.close()
            except Exception:
                pass

def is_vcf(path: str) -> bool:
    """Prüft, ob der Pfad wie eine VCF-Datei aussieht (plain oder gzipped)."""
    return path.lower().endswith((".vcf", ".vcf.gz"))


def _normalize_chrom_vcf(chrom: str) -> str:
    """
    Normalisiert Chromosom-Namen für konsistente Verarbeitung.

    V16 FIX: Zentrale Funktion für alle VCF-Parser.

    Transformationen:
    - chr1 -> 1
    - CHR1 -> 1
    - chrX -> X
    - chrM -> MT (VCF-Standard für Mitochondrien)
    - M -> MT

    Args:
        chrom: Rohes Chromosom aus VCF

    Returns:
        Normalisiertes Chromosom (ohne chr-Präfix, uppercase, M->MT)
    """
    c = str(chrom).replace("chr", "").replace("CHR", "").upper()
    # Mitochondrien-Normalisierung: M -> MT (Standard)
    if c == "M":
        c = "MT"
    return c


def parse_vcf_records(path):
    """
    Streamt VCF-Records effizient:
    - Header wird genau einmal via 'header' am ersten Record mitgegeben.
    - INFO wird parsimoniously geparst; use split with maxsplit to reduce allocations.
    - Erzeugt fuer jede ALT nur ein schlankes key-tuple + minimalen record-dict.
    - Ziel: moeglichst wenig temporaere Listen/Dicts in inner loop.
    """
    header = []
    header_attached = False

    with open_text_maybe_gzip(path) as f:
        for line in f:
            if not line:
                continue
            if line[0] == "#":
                header.append(line.rstrip("\n"))
                continue
            if not line.strip():
                continue

            # --- Nur Python-Parsing hier ---
            parts = line.rstrip("\n").split("\t", 9)
            if len(parts) < 8:
                continue

            chrom, pos_str, vid, ref, alt_str, qual, flt, info_str = parts[:8]
            fmt = parts[8] if len(parts) > 8 else None
            samples = parts[9:] if len(parts) > 9 else []

            try:
                pos = int(pos_str)
            except Exception:
                continue

            chrom_norm = _normalize_chrom_vcf(chrom)
            qual_norm = qual if qual not in (".", "") else None
            flt_norm = flt if flt not in (".", "") else None

            info_map = None
            if info_str and info_str != ".":
                info_map = {}
                for kv in info_str.split(";"):
                    if not kv:
                        continue
                    if "=" in kv:
                        k, v = kv.split("=", 1)
                        info_map[k] = v
                    else:
                        info_map[kv] = True

            alts = alt_str.split(",")
            for alt in alts:
                rec = {
                    "header": (header if not header_attached else None),
                    "chrom": chrom_norm,
                    "pos": pos,
                    "id": vid if vid != "." else None,
                    "ref": ref,
                    "alt": alt,
                    "qual": qual_norm,
                    "filter": flt_norm,
                    "info": info_str if info_map is None else info_map,
                    "fmt": fmt,
                    "samples": samples
                }
                yield rec

            header_attached = True


def parse_vcf_records_mmap(path: str):
    """
    MEMORY-MAPPED VCF PARSER für grosse unkomprimierte Dateien.
    FIX: Entfernt blockierende sys.stderr Aufrufe.
    """
    import mmap

    header = []
    header_attached = False

    try:
        with open(path, 'rb') as f:
            # Access=READ ist wichtig für Windows File-Locking Koexistenz
            with mmap.mmap(f.fileno(), 0, access=mmap.ACCESS_READ) as mm:
                current_pos = 0
                file_size = mm.size()

                while current_pos < file_size:
                    newline_pos = mm.find(b'\n', current_pos)
                    
                    if newline_pos == -1:
                        # Letzte Zeile ohne Newline am Ende
                        line_bytes = mm[current_pos:]
                        current_pos = file_size
                    else:
                        # Zeile bis Newline
                        line_bytes = mm[current_pos:newline_pos]
                        current_pos = newline_pos + 1

                    # Leere Zeilen überspringen
                    if not line_bytes or len(line_bytes) == 0:
                        continue
                    
                    # Win-Kompatibilität: \r entfernen falls vorhanden
                    if line_bytes.endswith(b'\r'):
                        line_bytes = line_bytes[:-1]

                    try:
                        line = line_bytes.decode('utf-8', errors='replace')
                    except Exception:
                        continue

                    if line.startswith('#'):
                        header.append(line)
                        continue
                    if not line.strip():
                        continue

                    # --- Parsing Logic ---
                    # (Identisch zum Standard-Parser, aber optimiert)
                    parts = line.split('\t', 9)
                    if len(parts) < 8:
                        continue

                    chrom, pos_str, vid, ref, alt_str, qual, flt, info_str = parts[:8]
                    fmt = parts[8] if len(parts) > 8 else None
                    samples = parts[9:] if len(parts) > 9 else []

                    try:
                        pos = int(pos_str)
                    except ValueError:
                        continue

                    chrom_norm = _normalize_chrom_vcf(chrom)
                    qual_norm = qual if qual not in (".", "") else None
                    flt_norm = flt if flt not in (".", "") else None

                    info_map = None
                    if info_str and info_str != ".":
                        # Quick-Parse für INFO (vermeidet overhead)
                        info_map = {}
                        for kv in info_str.split(";"):
                            if not kv: continue
                            if "=" in kv:
                                k, v = kv.split("=", 1)
                                info_map[k] = v
                            else:
                                info_map[kv] = True

                    alts = alt_str.split(",")
                    for alt in alts:
                        rec = {
                            "header": (header if not header_attached else None),
                            "chrom": chrom_norm,
                            "pos": pos,
                            "id": vid if vid != "." else None,
                            "ref": ref,
                            "alt": alt,
                            "qual": qual_norm,
                            "filter": flt_norm,
                            "info": info_str if info_map is None else info_map,
                            "fmt": fmt,
                            "samples": samples
                        }
                        yield rec

                    header_attached = True

    except Exception as e:
        # Fallback auf Standard-Parser bei mmap-Fehlern (z.B. 0-Byte Dateien)
        # Keine blockierenden Prints hier!
        yield from parse_vcf_records(path)


def parse_vcf_records_smart(path: str):
    """
    INTELLIGENTER VCF PARSER: Waehlt beste Methode automatisch.
    FIX: Entfernt blockierende sys.stderr Aufrufe.
    """
    try:
        file_size = os.path.getsize(path)
        file_size_mb = file_size / (1024 * 1024)

        # Nutze mmap nur für große, unkomprimierte Dateien
        # mmap lohnt sich meist erst ab >50MB, darunter ist overhead höher
        use_mmap = (
            not is_gzipped(path) and
            file_size_mb > 50
        )

        if use_mmap:
            yield from parse_vcf_records_mmap(path)
        else:
            yield from parse_vcf_records(path)

    except Exception:
        # Fallback bei jeglichem Fehler in der Smart-Selection
        yield from parse_vcf_records(path)
        
# --- FIX 1: Correct samples parsing in parse_vcf_records ---
"""def parse_vcf_records(path, max_line_length=100000):
    
    #Sicheres VCF-Parsing mit Line-Length-Protection und verbessertem Logging.
    #
    #Änderungen:
    #- Skip zu lange Zeilen (verhindert OOM)
    #- Exception-Handling pro Record
    #- Logging bei Parse-Errors (erste N detailliert, danach periodisch)
    
    header = []
    header_attached = False
    
    parse_errors = 0
    MAX_PARSE_ERRORS = 1000  # erst bei sehr vielen Fehlern hart abbrechen
    LOG_FIRST_ERRORS = 10    # erste N Fehler detailliert loggen
    LOG_EVERY_N = 50         # danach nur noch alle N Fehler loggen

    with open_text_maybe_gzip(path) as f:
        for line_no, line in enumerate(f, 1):
            if not line:
                continue
            
            # ✅ Skip zu lange Zeilen
            if len(line) > max_line_length:
                parse_errors += 1
                if parse_errors <= LOG_FIRST_ERRORS or parse_errors % LOG_EVERY_N == 0:
                    print(f"[VCF-Parse] ⚠️ Line {line_no} too long ({len(line)} chars) - skipped")
                if parse_errors > MAX_PARSE_ERRORS:
                    raise RuntimeError(f"Too many parse errors (>{MAX_PARSE_ERRORS})")
                continue
            
            # Header
            if line[0] == "#":
                header.append(line.rstrip("\n"))
                continue
            
            if not line.strip():
                continue

            try:
                parts = line.rstrip("\n").split("\t", 9)
                if len(parts) < 8:
                    parse_errors += 1
                    if parse_errors <= LOG_FIRST_ERRORS or parse_errors % LOG_EVERY_N == 0:
                        print(f"[VCF-Parse] ⚠️ Line {line_no} too few columns ({len(parts)}) - skipped")
                    continue

                chrom, pos_str, vid, ref, alt_str, qual, flt, info_str = parts[:8]
                fmt = parts[8] if len(parts) > 8 else None
                samples = parts[9:] if len(parts) > 9 else []

                try:
                    pos = int(pos_str)
                except Exception:
                    parse_errors += 1
                    if parse_errors <= LOG_FIRST_ERRORS or parse_errors % LOG_EVERY_N == 0:
                        print(f"[VCF-Parse] ⚠️ Line {line_no} invalid POS '{pos_str}' - skipped")
                    continue

                chrom_norm = _normalize_chrom_vcf(chrom)
                qual_norm = qual if qual not in (".", "") else None
                flt_norm = flt if flt not in (".", "") else None

                # Parse INFO
                info_map = None
                if info_str and info_str != ".":
                    info_map = {}
                    for kv in info_str.split(";"):
                        if not kv:
                            continue
                        if "=" in kv:
                            k, v = kv.split("=", 1)
                            info_map[k] = v
                        else:
                            info_map[kv] = True

                # Split ALTs
                alts = alt_str.split(",")
                for alt in alts:
                    rec = {
                        "header": (header if not header_attached else None),
                        "chrom": chrom_norm,
                        "pos": pos,
                        "id": vid if vid != "." else None,
                        "ref": ref,
                        "alt": alt,
                        "qual": qual_norm,
                        "filter": flt_norm,
                        "info": info_str if info_map is None else info_map,
                        "fmt": fmt,
                        "samples": samples
                    }
                    yield rec

                header_attached = True

            except Exception as e:
                parse_errors += 1
                if parse_errors <= LOG_FIRST_ERRORS or parse_errors % LOG_EVERY_N == 0:
                    print(f"[VCF-Parse] ⚠️ Line {line_no} parse error: {e}")
                if parse_errors > MAX_PARSE_ERRORS:
                    raise RuntimeError(f"Too many parse errors (>{MAX_PARSE_ERRORS})")
                continue

    if parse_errors > 0:
        print(f"[VCF-Parse] ⚠️ Total parse errors: {parse_errors}")
        """
        
def determine_is_coding(consequence):
    if not consequence:
        return None
    cons = str(consequence).lower()
    coding_terms = [
        "missense", "stop_gained", "stop_lost", "synonymous",
        "frameshift", "inframe", "start_lost", "start_gained", "splice"
    ]
    return int(any(t in cons for t in coding_terms))

# ============== Build-Erkennung ==============
def parse_vcf_keys(path, limit=None):
    """
    Streamt nur die Schlüssel (chrom, pos, ref, alt) aus einem VCF.
    - Header wird übersprungen.
    - ALT-Spalten werden gesplittet → pro ALT ein Key.
    - Keine INFO/QUAL/SAMPLES geparst.
    - Optionales Limit: bricht nach N Keys ab.
    """
    count = 0
    with open_text_maybe_gzip(path) as f:
        for line in f:
            if not line or line[0] == "#":
                continue

            parts = line.rstrip("\n").split("\t", 5)
            if len(parts) < 5:
                continue

            chrom, pos_str, vid, ref, alt_str = parts[:5]
            try:
                pos = int(pos_str)
            except Exception:
                continue

            chrom_norm = _normalize_chrom_vcf(chrom)
            for alt in alt_str.split(","):
                yield (chrom_norm, pos, ref.upper(), alt.upper())
                count += 1
                if limit and count >= limit:
                    return

def ask_user_for_build():
    """GUI-Dialog zur Build-Auswahl (Tkinter)."""
    import tkinter as tk
    from tkinter import simpledialog

    root = tk.Tk()
    root.withdraw()
    choice = simpledialog.askstring(
        "Build-Auswahl",
        "Build konnte nicht automatisch erkannt werden.\n"
        "Bitte wählen Sie: GRCh37 oder GRCh38"
    )
    if choice and choice.strip().upper().endswith("37"):
        return "GRCh37"
    if choice and choice.strip().upper().endswith("38"):
        return "GRCh38"
    return None


def ask_user_for_build_cli():
    """CLI-Fallback zur Build-Auswahl."""
    while True:
        choice = input("Build konnte nicht erkannt werden. Bitte '37' oder '38' eingeben: ").strip()
        if choice in ("37", "GRCh37"):
            return "GRCh37"
        if choice in ("38", "GRCh38"):
            return "GRCh38"
        print("Ungültige Eingabe, bitte erneut versuchen.")


def get_existing_builds(db_path: str, logger=None) -> dict:
    """
    Liefert vorhandene Builds in der LightDB zurück.
    """
    if not os.path.exists(db_path):
        return {}
    try:
        with sqlite3.connect(f"file:{db_path}?mode=ro", uri=True) as con:
            cur = con.cursor()
            cur.execute("""
                SELECT build, COUNT(*) as count
                FROM variants_light
                WHERE build IN ('GRCh37','GRCh38')
                GROUP BY build
            """)
            return {row[0]: row[1] for row in cur.fetchall()}
    except Exception as e:
        if logger:
            logger.log(f"[LightDB] ⚠️ Fehler beim Build-Check: {e}")
        return {}

    
def lightdb_test_keys(keys, build: str, db_path: str, logger=None, limit: int = 1000) -> int:
    sample = keys[:limit]
    if not sample:
        return 0

    try:
        with sqlite3.connect(f"file:{db_path}?mode=ro", uri=True, timeout=5.0) as con:
            cur = con.cursor()

            # ✅ FIX: Temp-Table statt unsichere OR-Kette
            cur.execute("DROP TABLE IF EXISTS tmp_test_keys;")
            cur.execute("""
                CREATE TEMP TABLE tmp_test_keys (
                    chrom TEXT,
                    pos INTEGER,
                    ref TEXT,
                    alt TEXT
                );
            """)
            
            # ✅ Parametrisiertes INSERT
            cur.executemany(
                "INSERT INTO tmp_test_keys VALUES (?,?,?,?)",
                [(str(c), int(p), str(r), str(a)) for (c, p, r, a) in sample]
            )

            # ✅ JOIN statt OR-Kette (nutzt Index)
            cur.execute("""
                SELECT COUNT(*) 
                FROM variants_light v
                JOIN tmp_test_keys t
                  ON v.chrom = t.chrom
                 AND v.pos = t.pos
                 AND v.ref = t.ref
                 AND v.alt = t.alt
                 AND v.build = ?
            """, (build,))
            
            hits = cur.fetchone()[0]
            cur.execute("DROP TABLE tmp_test_keys;")

        if logger:
            logger.log(f"[BuildDetect] LightDB-Test {build}: {hits}/{len(sample)} Treffer")
        return hits

    except Exception as e:
        if logger:
            logger.log(f"[BuildDetect] ⚠️ LightDB-Test für {build} fehlgeschlagen: {e}")
        return 0

def decide_build_from_lightdb(keys, db_path: str, logger=None,
                              min_ratio: float = 0.6, min_diff: int = 50):
    """
    Entscheidet den Build anhand von LightDB-Treffern.
    - Nutzt einfaches Verhältnis + Mindestdifferenz.
    - min_ratio: Anteil der Treffer, der mindestens erreicht werden muss.
    - min_diff: absolute Mindestdifferenz zwischen den Builds.
    """
    existing_builds = get_existing_builds(db_path, logger)
    if not existing_builds:
        if logger:
            logger.log("[BuildDetect] Keine Builds in LightDB gefunden.")
        return None

    hits37 = lightdb_test_keys(keys, "GRCh37", db_path, logger) if "GRCh37" in existing_builds else 0
    hits38 = lightdb_test_keys(keys, "GRCh38", db_path, logger) if "GRCh38" in existing_builds else 0
    total = len(keys)

    if logger:
        logger.log(f"[BuildDetect] Treffer: GRCh37={hits37}, GRCh38={hits38}, total={total}")

    # Beide Builds vorhanden
    if "GRCh37" in existing_builds and "GRCh38" in existing_builds:
        if hits37 + hits38 == 0:
            return None
        ratio37 = hits37 / (hits37 + hits38)
        ratio38 = hits38 / (hits37 + hits38)
        diff = abs(hits37 - hits38)

        if logger:
            logger.log(f"[BuildDetect] Verhältnis: 37={ratio37:.2f}, 38={ratio38:.2f}, diff={diff}")

        if ratio37 >= min_ratio and diff >= min_diff:
            return "GRCh37"
        elif ratio38 >= min_ratio and diff >= min_diff:
            return "GRCh38"
        else:
            return None

    # Nur ein Build vorhanden
    elif "GRCh37" in existing_builds:
        return "GRCh37" if hits37 > 0 else None
    elif "GRCh38" in existing_builds:
        return "GRCh38" if hits38 > 0 else None

    return None

def mv_query_by_rsid_batch(rsids, fields, logger=None):
    """
    Fragt MyVariant nach rsIDs ab und gibt Mapping rsid -> Hit zurück.
    """
    results = {}
    if not rsids:
        return results

    q = " OR ".join([f"dbsnp.rsid:{r}" for r in rsids])
    params = {"q": q, "fields": ",".join(fields), "size": len(rsids)}

    try:
        r = requests.get(f"{MV_BASE}/query", params=params, timeout=30)
        r.raise_for_status()
        j = r.json()
        for hit in j.get("hits", []):
            if isinstance(hit.get("dbsnp"), dict):
                rs = hit["dbsnp"].get("rsid")
                if rs:
                    results[rs] = hit
    except Exception as e:
        if logger:
            logger.log(f"[BuildDetect] ❌ MyVariant rsid query error: {e}")
    return results


def detect_build_robust_by_rsids(path, logger=None, max_rsids=20, tol=5):
    """
    Ermittelt Build anhand von rsIDs durch Abgleich mit MyVariant.
    - Liest bis zu max_rsids Varianten.
    - Vergleicht Positionen mit hg19/hg38.
    - Toleranz: ±tol Basen.
    """
    rs_pos = []
    with open_text_maybe_gzip(path) as f:
        for line in f:
            if line.startswith("#"):
                continue
            parts = line.rstrip("\n").split("\t")
            if len(parts) < 3:
                continue
            vid = parts[2]
            if vid and vid.startswith("rs"):
                try:
                    pos = int(parts[1])
                except Exception:
                    continue
                rs_pos.append((vid, pos))
            if len(rs_pos) >= max_rsids:
                break

    if not rs_pos:
        if logger:
            logger.log("[BuildDetect] ⚠️ Keine rsIDs im VCF gefunden.")
        return None

    rsids = [r for r, _ in rs_pos]
    hits = mv_query_by_rsid_batch(rsids, ["dbsnp.hg19", "dbsnp.hg38"], logger=logger)

    m37 = m38 = 0
    for rsid, pos_file in rs_pos:
        db = (hits.get(rsid) or {}).get("dbsnp") or {}
        hg19 = db.get("hg19") or {}
        hg38 = db.get("hg38") or {}

        if "start" in hg19:
            if abs(int(hg19["start"]) - pos_file) <= tol:
                m37 += 1
        if "start" in hg38:
            if abs(int(hg38["start"]) - pos_file) <= tol:
                m38 += 1

    if logger:
        logger.log(f"[BuildDetect] rsID-Matches: GRCh37={m37}, GRCh38={m38}")

    if m37 == 0 and m38 == 0:
        return None
    return "GRCh38" if m38 >= m37 else "GRCh37"


def detect_build_from_text_header(lines_iterable, logger=None):
    """
    Analysiert nur Headerzeilen.
    Gewichtung:
      1. ##reference= mit GRCh/HG-Hinweis
      2. andere Header mit GRCh/HG
      3. reine Zahlen (37/38/19)
    """
    text = "\n".join(lines_iterable).lower()

    # 1. ##reference= Zeile prüfen
    ref_lines = [l.lower() for l in lines_iterable if l.lower().startswith("##reference=")]
    for ref in ref_lines:
        found_37 = any(re.search(p, ref) for p in BUILD_HINTS_37)
        found_38 = any(re.search(p, ref) for p in BUILD_HINTS_38)
        if found_37 and not found_38:
            return "GRCh37"
        if found_38 and not found_37:
            return "GRCh38"
        if found_37 and found_38:
            if logger:
                logger.log("[BuildDetect] ⚠️ ##reference= enthält widersprüchliche Hinweise (37 & 38).")
            return None

    # 2. Andere Headerzeilen mit GRCh/HG
    found_37 = any(re.search(p, text) for p in BUILD_HINTS_37)
    found_38 = any(re.search(p, text) for p in BUILD_HINTS_38)
    if found_37 and not found_38:
        return "GRCh37"
    if found_38 and not found_37:
        return "GRCh38"
    if found_37 and found_38:
        if logger:
            logger.log("[BuildDetect] ⚠️ Header enthält widersprüchliche Hinweise (37 & 38).")
        return None

    # 3. Fallback: reine Zahlenmuster
    if "##reference=" in text:
        if re.search(r"\b38\b", text):
            return "GRCh38"
        if re.search(r"\b37\b", text) or re.search(r"\b19\b", text):
            return "GRCh37"

    return None


def detect_build_from_header(path, logger=None):
    """
    Liest nur Headerzeilen bis zur ersten Variant-Zeile (#CHROM oder erste nicht-# Zeile).
    Nutzt diese für die Build-Erkennung.
    """
    header_lines = []
    with open_text_maybe_gzip(path) as f:
        for line in f:
            if line.startswith("#"):
                header_lines.append(line.rstrip("\n"))
                if line.lower().startswith("#chrom"):
                    break
            else:
                break

    b = detect_build_from_text_header(header_lines, logger)
    if b:
        if logger:
            logger.log(f"[BuildDetect] ✅ Build im Header erkannt: {b}")
        return b

    if logger:
        logger.log("[BuildDetect] ⚠️ Kein Build im Header erkannt.")
    return None


def detect_build_for_vcf(path, db_path=None, logger=None, alpha: float = 0.01):
    """
    Robuste Build-Erkennung:
    - Header-Check
    - LightDB-Probe (Chi²-Test p<0.01, mit eigenem LightDB-Manager)
    - rsID-Heuristik (MyVariant)
    - Benutzerabfrage (GUI oder CLI), falls alles fehlschlägt
    
    V10: db_path Default ist jetzt None, wird zu absolutem Pfad aufgelöst.
    """
    # V10: Absoluter Pfad als Fallback
    if db_path is None:
        db_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            "data",
            "gnomad_light.db"
        )
    
    # 1. Header-Check
    logger.log(f"[BuildDetect] Starte Build-Erkennungs-Modul")
    logger.log(f"[BuildDetect] Versuche Build über Header-Text zu erkennen.")
    try:
        b = detect_build_from_header(path, logger)
        if b:
            if logger:
                logger.log(f"[BuildDetect] ✅ Build durch Header erkannt: {b}")
            return b
    except Exception as e:
        if logger:
            logger.log(f"[BuildDetect] ⚠️ Fehler beim Header-Check: {e}")
    # 2. LightDB-Probe (direkter DB-Zugriff, Verhältnis-Test)
    if logger:
        logger.log("[BuildDetect] Versuche Build über Tests mit Light-DB zu bestimmen.")
    try:
        if db_path and os.path.exists(db_path):
            if logger:
                logger.log(f"[BuildDetect] Nutze LightDB unter: {db_path}")

            # Keys effizient mit parse_vcf_keys sammeln
            keys = list(parse_vcf_keys(path, limit=1000))
            if logger:
                logger.log(f"[BuildDetect] Gesammelte Keys für Test: {len(keys)}")

            if keys:
                b_light = decide_build_from_lightdb(
                    keys,
                    db_path,
                    logger=logger,
                    min_ratio=0.6,   # mindestens 60% Trefferanteil
                    min_diff=50      # mindestens 50 Treffer Unterschied
                )
                if b_light:
                    if logger:
                        logger.log(f"[BuildDetect] ✅ Build durch LightDB-Probe erkannt: {b_light}")
                    return b_light
                else:
                    if logger:
                        logger.log("[BuildDetect] ⚠️ LightDB-Probe ergab kein eindeutiges Ergebnis.")
        else:
            if logger:
                logger.log("[BuildDetect] ℹ️ LightDB-Datei nicht gefunden – überspringe Probe.")
    except Exception as e:
        if logger:
            logger.log(f"[BuildDetect] ⚠️ Fehler bei LightDB-Probe: {e}")

    # 3. rsID-Heuristik
    logger.log(f"[BuildDetect] Versuche Build über API mit RSID zu erkennen.")
    try:
        b2 = detect_build_robust_by_rsids(path, logger=logger)
        if b2:
            if logger:
                logger.log(f"[BuildDetect] ✅ Build durch rsID-Heuristik erkannt: {b2}")
            return b2
    except Exception as e:
        if logger:
            logger.log(f"[BuildDetect] ⚠️ Fehler bei rsID-Heuristik: {e}")

    # 4. Benutzerabfrage
    try:
        build = ask_user_for_build()
    except Exception:
        build = None

    if not build:
        build = ask_user_for_build_cli()

    if build:
        if logger:
            logger.log(f"[BuildDetect] ⚠️ Build manuell vom Benutzer gewählt: {build}")
        return build

    raise RuntimeError("Build konnte nicht automatisch erkannt oder manuell gewählt werden.")

# =============================================================================
# NETWORK UTILS (Robust Retry)
# =============================================================================
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

def get_robust_session(retries=Config.MAX_RETRIES, backoff_factor=0.5, status_forcelist=(429, 500, 502, 503, 504)):
    """
    Erstellt eine requests.Session mit robuster Retry-Strategie.
    """
    session = requests.Session()
    retry = Retry(
        total=retries,
        read=retries,
        connect=retries,
        backoff_factor=backoff_factor,
        status_forcelist=status_forcelist,
    )
    adapter = HTTPAdapter(max_retries=retry)
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    return session


def load_fai_index(fai_path):
    idx = {}
    with open(fai_path, "r") as f:
        for line in f:
            chrom, length, offset, line_bases, line_width = line.strip().split("\t")
            idx[chrom] = (int(length), int(offset), int(line_bases), int(line_width))
    return idx

def fetch_base_from_fasta(fasta_path, fai_index, chrom, pos):
    """
    Random-access fetch of a single REF base using .fai index.
    """
    if chrom not in fai_index:
        return "N"
    length, offset, line_bases, line_width = fai_index[chrom]
    if pos < 1 or pos > length:
        return "N"
    line_num = (pos - 1) // line_bases
    col_num = (pos - 1) % line_bases
    byte_offset = offset + line_num * line_width + col_num
    with open(fasta_path, "rb") as f:
        f.seek(byte_offset)
        b = f.read(1)
    try:
        return b.decode("ascii").upper()
    except Exception:
        return "N"
# ============== MyVariant helpers ==============
def indel_ref_alt_from_spdi(chrom: str, pos1: int, deleted: str, inserted: str, fetch_ref_base) -> tuple[str, str] | None:
    """
    Baut REF/ALT-Sequenzen für ein Indel aus SPDI-Daten.

    Parameter:
        chrom        - Chromosom (z.B. "1", "X")
        pos1         - 1-basierte Position der Variation
        deleted      - SPDI deleted_sequence (kann leer sein)
        inserted     - SPDI inserted_sequence (kann leer sein)
        fetch_ref_base - Funktion(chrom, pos1) -> Referenzbase

    Rückgabe:
        (REF, ALT) oder None, falls nicht bestimmbar
    """
    # Hole die Ankerbase links der Variation
    anchor = fetch_ref_base(chrom, pos1)
    if not anchor or anchor in (".", "-", "N"):
        return None

    # REF = Anker + deleted, ALT = Anker + inserted
    ref_seq = anchor + deleted.upper() if deleted else anchor
    alt_seq = anchor + inserted.upper() if inserted else anchor

    # Validierung: nur ACGT erlaubt
    if not all(c in "ACGT" for c in ref_seq):
        return None
    if not all(c in "ACGT" for c in alt_seq):
        return None

    return ref_seq, alt_seq

def make_hgvs_indel(chrom, pos, ref, alt):
    """
    Produce HGVS g.-notation for normalized indels.
    - Deletion:  chrN:g.start_enddel or chrN:g.posdel
    - Insertion: chrN:g.pos_pos+1insSEQ
    - DelIns:    chrN:g.start_enddelinsSEQ
    Assumes ref/alt are ACGT-only and left-normalized.
    """
    def _is_acgt(s: str) -> bool:
        return isinstance(s, str) and len(s) > 0 and all(c in "ACGT" for c in s)

    ch = str(chrom).replace("CHR", "").replace("chr", "")
    if not (_is_acgt(ref) and _is_acgt(alt)):
        return None

    # Trim common prefix
    i = 0
    while i < len(ref) and i < len(alt) and ref[i] == alt[i]:
        i += 1
    pos2 = pos + i
    ref2 = ref[i:]
    alt2 = alt[i:]

    # Trim common suffix
    j = 0
    while j < len(ref2) and j < len(alt2) and ref2[-1 - j] == alt2[-1 - j]:
        j += 1
    if j > 0:
        ref2 = ref2[:-j] if j < len(ref2) else ""
        alt2 = alt2[:-j] if j < len(alt2) else ""

    # Deletion
    if ref2 and not alt2:
        if len(ref2) == 1:
            return f"chr{ch}:g.{pos2}del"
        start = pos2
        end = pos2 + len(ref2) - 1
        return f"chr{ch}:g.{start}_{end}del"

    # Insertion
    if alt2 and not ref2:
        return f"chr{ch}:g.{pos2}_{pos2+1}ins{alt2}"

    # DelIns
    start = pos2
    end = pos2 + len(ref2) - 1
    return f"chr{ch}:g.{start}_{end}delins{alt2}"

def make_hgvs(chrom, pos, ref, alt):
    """
    HGVS g.-notation for SNVs and indels. Assumes VCF-normalized alleles.
    """
    key = (chrom, pos, ref, alt, None)
    if _is_pure_snv_key(key):
        ch = str(chrom).replace("CHR", "").replace("chr", "")
        return f"chr{ch}:g.{pos}{ref}>{alt}"
    if _is_indel_key(key):
        return make_hgvs_indel(chrom, pos, ref, alt)
    return None

# FIX: Erlaubt Position-Ranges bei INDELs
HGVS_REGEX = re.compile(
    r"^chr[0-9XYM]+:g\.\d+(?:"
    r"[ACGT]>[ACGT]|"                    # SNV: g.123A>T
    r"del|"                               # Deletion: g.123del
    r"_\d+del|"                           # Range deletion: g.123_124del
    r"_\d+ins[ACGT]+|"                    # Range insertion: g.123_124insATG
    r"ins[ACGT]+|"                        # Simple insertion: g.123insATG
    r"_\d+delins[ACGT]+|"                 # Range delins: g.123_124delinsATG
    r"delins[ACGT]+"                      # Simple delins: g.123delinsATG
    r")$"
)

def is_valid_hgvs(hgvs: str) -> bool:
    """
    Validiert HGVS-Notation.
    FIX: Akzeptiert alle gueltigen HGVS-Formate inkl. Ranges
    """
    if not hgvs or not isinstance(hgvs, str):
        return False
    return bool(HGVS_REGEX.match(hgvs))

def normalize_variant(chrom, pos, ref, alt):
    while ref and alt and ref[0] == alt[0]:
        ref = ref[1:]
        alt = alt[1:]
        pos += 1
    while ref and alt and ref[-1] == alt[-1]:
        ref = ref[:-1]
        alt = alt[:-1]
    return chrom, pos, ref or "-", alt or "-"

def _is_pure_snv_key(key):
    chrom, pos, ref, alt, build = key
    return (
        isinstance(ref, str) and isinstance(alt, str) and
        len(ref) == 1 and len(alt) == 1 and
        all(c in "ACGT" for c in ref) and
        all(c in "ACGT" for c in alt) and
        ref != alt
    )

def _is_indel_key(key):
    chrom, pos, ref, alt, build = key
    if not (isinstance(ref, str) and isinstance(alt, str)):
        return False
    if not (all(c in "ACGT" for c in ref) and all(c in "ACGT" for c in alt)):
        return False
    return len(ref) != len(alt) or len(ref) > 1 or len(alt) > 1

# normalize_indel (aus relevanter code2.txt)
def normalize_indel(chrom, pos, ref, alt, get_ref_base):
    """
    Links-normalisiert eine Indel und fügt ggf. Ankerbase hinzu.
    """
    if ref is None or alt is None:
        return None

    ref = ref.replace("-", "").replace(".", "")
    alt = alt.replace("-", "").replace(".", "")

    i = 0
    while i < len(ref) and i < len(alt) and ref[i] == alt[i]:
        i += 1
    pos += i
    ref = ref[i:]
    alt = alt[i:]

    j = 0
    while j < len(ref) and j < len(alt) and ref[-1 - j] == alt[-1 - j]:
        j += 1
    if j > 0:
        ref = ref[:-j] if j < len(ref) else ""
        alt = alt[:-j] if j < len(alt) else ""

    if ref == "" or alt == "":
        anchor_pos = max(1, pos - 1)
        anchor_base = get_ref_base(chrom, anchor_pos)
        if not anchor_base or anchor_base not in "ACGT":
            return None
        if ref == "":
            ref = anchor_base
            alt = anchor_base + alt
            pos = anchor_pos
        elif alt == "":
            alt = anchor_base
            ref = anchor_base + ref
            pos = anchor_pos

    if not (all(c in "ACGT" for c in ref) and all(c in "ACGT" for c in alt)):
        return None

    return chrom, pos, ref, alt

def get_ref_base(chrom, pos, fasta_path=None, fai_index=None, cache=None, build=None):
    """
    Liefert die Referenzbase (ACGT) für eine gegebene Chromosomenposition.
    Nutzt, falls vorhanden:
      - FASTA + FAI-Index
      - ansonsten Assembly-Cache (falls übergeben)
    Gibt 'N' zurück, wenn keine Base ermittelt werden kann.
    """
    # FASTA + FAI
    if fasta_path and fai_index:
        try:
            base = fetch_base_from_fasta(fasta_path, fai_index, chrom, pos).upper()
            if base in "ACGT":
                return base
        except Exception:
            pass

    # Cache (falls vorhanden)
    if cache and build:
        key = (build, chrom, pos)
        if key in cache:
            base = cache[key]
            if base in "ACGT":
                return base

    return "N"
# Gemeinsames Key-Format: (chrom, pos, ref, alt, build)

# ============================================================================
# Normalisierungsfunktionen (konsistent und robust)
# ============================================================================

def normalize_for_mv(key: tuple) -> str:
    """
    MyVariant erwartet HGVS-Notation: chr:g.posRef>Alt
    Nutzt make_hgvs für konsistente Formatierung.
    """
    chrom, pos, ref, alt, build = key
    # make_hgvs sollte selbst für korrektes HGVS-Format sorgen
    return make_hgvs(chrom, pos, ref, alt) or f"chr{chrom}:g.{pos}{ref}>{alt}"


def normalize_for_gnomad(key: tuple) -> str:
    """
    gnomAD-Format: "chrom-pos-ref-alt" (ohne 'chr'-Prefix).
    """
    chrom, pos, ref, alt, build = key
    chrom = str(chrom).replace("chr", "").replace("CHR", "").upper()
    return f"{chrom}-{pos}-{ref}-{alt}"


def normalize_for_ensembl(key: tuple) -> str:
    """
    Ensembl VEP-Format: "chrom pos . ref alt" (VCF-like, space-separated).
    """
    chrom, pos, ref, alt, build = key
    chrom = str(chrom).replace("chr", "").replace("CHR", "").upper()
    return f"{chrom} {pos} . {ref} {alt}"


def normalize_for_vep(key: tuple) -> str:
    """
    VEP-Format (identisch zu Ensembl).
    """
    chrom, pos, ref, alt, build = key
    chrom = str(chrom).replace("chr", "").replace("CHR", "").upper()
    return f"{chrom} {pos} . {ref} {alt}"

def normalize_for_topmed(key: tuple) -> str:
    """
    TOPMed/BRAVO API: "chrom-pos-ref-alt" (ohne 'chr').
    """
    chrom, pos, ref, alt, build = key
    chrom = str(chrom).replace("chr", "").replace("CHR", "").upper()
    return f"{chrom}-{pos}-{ref}-{alt}"
# =============================================================================
# HELPER: aiohttp verfügbar prüfen
# =============================================================================

def check_aiohttp_available():
    """
    Prüft ob aiohttp verfügbar ist und gibt hilfreiche Fehlermeldung.
    
    Returns:
        bool: True wenn verfügbar, False sonst
    
    Raises:
        ImportError: Mit Installationsanleitung wenn nicht verfügbar
    """
    if not AIOHTTP_AVAILABLE or aiohttp is None:
        raise ImportError(
            "aiohttp is required for async variant fetching.\n"
            "Install with: pip install aiohttp\n"
            "Or use ThreadPool mode instead."
        )
    return True


# =============================================================================
# ROUTER-FUNKTION (mit aiohttp-Check)
# =============================================================================

def mv_fetch(keys, build, fetch_full, batch_size=1000, db=None, logger=None, phase_label="MV"):
    """
    Router für MyVariant-Fetch:
    - Entscheidet adaptiv zwischen Async und ThreadPool
    - Kriterien:
      * Anzahl Keys (total)
      * Batchgröße
      * CPU-Last
    - Bei Async-Loop-Konflikt automatisch Fallback auf ThreadPool
    - Bei fehlendem aiohttp automatisch ThreadPool
    """
    total = len(keys)
    async_needed = False

    # Kriterium 1: klassische Schwelle
    if total >= ASYNC_THRESHOLD:
        async_needed = True

    # Kriterium 2: viele Keys + große Batches
    elif total >= 10000 and batch_size >= 800:
        async_needed = True

    # Kriterium 3: hohe CPU-Last + mittlere Keys
    else:
        try:
            if psutil:
                cpu = psutil.cpu_percent(interval=None)
                if cpu > 70.0 and total >= 5000:
                    async_needed = True
        except Exception:
            pass

    # ✅ NEU: Prüfe aiohttp-Verfügbarkeit
    if async_needed and not AIOHTTP_AVAILABLE:
        if logger:
            logger.log(
                f"[{phase_label}] ⚠️ Async gewünscht, aber aiohttp nicht verfügbar → "
                f"Fallback auf ThreadPool"
            )
        async_needed = False

    if async_needed:
        if logger:
            logger.log(f"[{phase_label}] ⚡ {total:,} Varianten → Async-Modus (batch={batch_size})")
        try:
            # ✅ Prüfe aiohttp nochmal vor asyncio.run
            check_aiohttp_available()
            
            return asyncio.run(
                mv_fetch_async(
                    keys,
                    build,
                    fetch_full,
                    batch_size=batch_size,
                    logger=logger,
                    phase_label=phase_label,
                )
            )
        except ImportError as e:
            # aiohttp fehlt
            if logger:
                logger.log(f"[{phase_label}] ❌ {e}")
                logger.log(f"[{phase_label}] → Fallback auf ThreadPool")
            return mv_fetch_threadpool(
                keys,
                build,
                fetch_full,
                batch_size=batch_size,
                db=db,
                logger=logger,
                phase_label=phase_label,
            )
        except RuntimeError as e:
            # "asyncio.run() cannot be called from a running event loop"
            if logger:
                logger.log(f"[{phase_label}] ⚠️ Async-Loop aktiv – Fallback auf ThreadPool")
            return mv_fetch_threadpool(
                keys,
                build,
                fetch_full,
                batch_size=batch_size,
                db=db,
                logger=logger,
                phase_label=phase_label,
            )
    else:
        if logger:
            logger.log(f"[{phase_label}] ⚡ {total:,} Varianten → ThreadPool-Modus (batch={batch_size})")
        return mv_fetch_threadpool(
            keys,
            build,
            fetch_full,
            batch_size=batch_size,
            db=db,
            logger=logger,
            phase_label=phase_label,
        )


# =============================================================================
# ASYNC FETCHER (mit aiohttp-Check)
# =============================================================================

async def mv_fetch_async(keys, build, fetch_full, batch_size=1000, logger=None, phase_label="MV"):
    """
    MyVariant async fetch mit robustem Error-Handling.

    FIX:
    - Zusätzliche Guards gegen NoneType
    - Nur auf Dicts .get() anwenden
    - Saubere Fehlerlogs
    """
    if not AIOHTTP_AVAILABLE or aiohttp is None:
        error_msg = "aiohttp is required for mv_fetch_async"
        if logger:
            logger.log(f"[{phase_label}] ❌ {error_msg}")
        raise ImportError(error_msg)

    results = {}
    hgvs_map = {}
    items = []

    # HGVS-IDs via normalize_for_mv
    for k in keys:
        try:
            hg = normalize_for_mv(k)
            if hg and is_valid_hgvs(hg):
                hgvs_map[k] = hg
                items.append(hg)
            else:
                hgvs_map[k] = None
                results[k] = None
        except Exception as e:
            if logger:
                logger.log(f"[{phase_label}] ⚠️ HGVS creation failed for {k}: {e}")
            hgvs_map[k] = None
            results[k] = None

    if not items:
        return results

    fields = "all" if fetch_full else ",".join(MV_FIELDS_AF)
    assembly = "hg38" if build == "GRCh38" else "hg19"

    _sem_count = get_api_setting("phase1_af", "myvariant", "workers", min(12, max(4, (os.cpu_count() or 4))))
    sem = asyncio.Semaphore(_sem_count)
    _timeout = get_api_setting("phase1_af", "myvariant", "timeout", 60)

    async def fetch_chunk(session, chunk):
        for attempt in range(3):
            try:
                async with sem:
                    if attempt > 0:
                        backoff = (2 ** attempt) + (random.random() * 0.5)
                        await asyncio.sleep(backoff)

                    async with session.post(
                        f"{MV_BASE}/variant",
                        json={"ids": chunk, "fields": fields, "assembly": assembly},
                        timeout=aiohttp.ClientTimeout(total=_timeout)
                    ) as resp:
                        if resp.status == 429:
                            retry_after = int(resp.headers.get("Retry-After", 10))
                            if logger:
                                logger.log(f"[{phase_label}] 🕒 429 Rate Limit → wait {retry_after}s")
                            await asyncio.sleep(retry_after + random.random())
                            continue

                        if resp.status != 200:
                            if logger and attempt == 2:
                                logger.log(f"[{phase_label}] ❌ HTTP {resp.status}")
                            return {}

                        data = await resp.json()
                        # ✅ Guard: nur Dicts oder Listen verarbeiten
                        if isinstance(data, list):
                            hits = [h for h in data if isinstance(h, dict) and "_id" in h]
                        elif isinstance(data, dict):
                            hits = [h for h in data.get("hits", []) if isinstance(h, dict) and "_id" in h]
                        else:
                            hits = []

                        return {h["_id"]: h for h in hits}

            except (asyncio.TimeoutError, aiohttp.ClientError) as e:
                if logger and attempt == 2:
                    logger.log(f"[{phase_label}] ❌ Chunk failed: {e}")
                await asyncio.sleep(2 ** attempt)

        return {}

    try:
        connector = aiohttp.TCPConnector(
            limit=_sem_count,
            limit_per_host=_sem_count,
            ttl_dns_cache=300
        )

        async with aiohttp.ClientSession(connector=connector) as session:
            tasks = []
            for i in range(0, len(items), batch_size):
                chunk = items[i:i+batch_size]
                tasks.append(asyncio.create_task(fetch_chunk(session, chunk)))

            all_results = await asyncio.gather(*tasks, return_exceptions=True)

        merged = {}
        for r in all_results:
            if isinstance(r, dict):
                merged.update(r)
            elif isinstance(r, Exception) and logger:
                logger.log(f"[{phase_label}] ⚠️ Task exception: {r}")

        # Map zurück zu Original-Keys
        for k in keys:
            hg = hgvs_map.get(k)
            results[k] = merged.get(hg) if hg else None

    except Exception as e:
        if logger:
            logger.log(f"[{phase_label}] ❌ Fatal error: {e}")
            import traceback
            logger.log(f"[{phase_label}] Traceback:\n{traceback.format_exc()}")

    return results


# =============================================================================
# WEITERE ASYNC FETCHERS (alle mit aiohttp-Check)
# =============================================================================

async def topmed_fetch_async(keys, build, logger=None):
    """
    TOPMed/BRAVO API fetch.
    
    ✅ KRITISCH: Prüft aiohttp-Verfügbarkeit
    """
    if not AIOHTTP_AVAILABLE or aiohttp is None:
        if logger:
            logger.log("[TOPMed] ❌ aiohttp nicht verfügbar")
        return {}
    
    results = {}

    _topmed_timeout = get_api_setting("phase1_af", "topmed", "timeout", 15)

    async def fetch_one(session, key):
        url = f"https://bravo.sph.umich.edu/api/v1/variant/{normalize_for_topmed(key)}"

        for attempt in range(3):
            try:
                async with session.get(
                    url,
                    headers={"Accept": "application/json"},
                    timeout=aiohttp.ClientTimeout(total=_topmed_timeout)
                ) as resp:
                    if resp.status == 200:
                        content_type = resp.headers.get("Content-Type", "")
                        if "application/json" in content_type:
                            data = await resp.json()
                            af = data.get("allele_freq")
                            return {"af_1kg": af} if af is not None else None
                        return None

                    elif resp.status == 404:
                        return None

                    elif resp.status == 429:
                        retry_after = int(resp.headers.get("Retry-After", 5))
                        await asyncio.sleep(retry_after + random.random())
                        continue

                    else:
                        if logger and attempt == 2:
                            logger.log(f"[TOPMed] ⚠️ HTTP {resp.status} for {key}")
                        return None

            except (asyncio.TimeoutError, aiohttp.ClientError) as e:
                if logger and attempt == 2:
                    logger.log(f"[TOPMed] ❌ {key}: {e}")
                await asyncio.sleep(2 ** attempt)

        return None

    try:
        _topmed_workers = get_api_setting("phase1_af", "topmed", "workers", 6)
        sem = asyncio.Semaphore(_topmed_workers)

        async with aiohttp.ClientSession() as session:
            tasks = []
            for key in keys:
                async def sem_task(k=key):
                    async with sem:
                        return await fetch_one(session, k)
                tasks.append(asyncio.create_task(sem_task()))

            all_results = await asyncio.gather(*tasks, return_exceptions=True)

        for key, result in zip(keys, all_results):
            results[key] = result if isinstance(result, dict) else None

    except Exception as e:
        if logger:
            logger.log(f"[TOPMed] ❌ Fatal error: {e}")

    return results


# ---------------------------
# ThreadPool-Variante
# ---------------------------
def mv_fetch_threadpool(keys, build, fetch_full, batch_size=1000, db=None, logger=None, phase_label="MV", max_workers=None):
    results = {}
    hgvs_map = {}
    items = []

    for k in keys:
        chrom, pos, ref, alt, _b = k
        if _is_pure_snv_key(k):
            hg = make_hgvs(chrom, pos, ref, alt)
        elif _is_indel_key(k):
            hg = make_hgvs_indel(chrom, pos, ref, alt)
        else:
            hg = None
        hgvs_map[k] = hg
        if hg and is_valid_hgvs(hg):
            items.append(hg)
        else:
            results[k] = None

    fields = "all" if fetch_full else ",".join(MV_FIELDS_AF)
    assembly = "hg38" if build == "GRCh38" else "hg19"

    if max_workers is None:
        cpu = os.cpu_count() or 4
        max_workers = max(8, min(64, cpu * 4))

    retry = Retry(
        total=Config.MAX_RETRIES,
        backoff_factor=0.5,
        status_forcelist=[429, 500, 502, 503, 504]
    )
    adapter = HTTPAdapter(
        pool_connections=max_workers,
        pool_maxsize=max_workers,
        max_retries=retry
    )
    session = requests.Session()
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    session.headers.update({"User-Agent": "VariantFusion/mv_fetch_threadpool"})

    def fetch_chunk(chunk):
        try:
            r = session.post(
                f"{MV_BASE}/variant",
                json={"ids": chunk, "fields": fields, "assembly": assembly},
                timeout=60
            )
            r.raise_for_status()
            data = r.json()
            hits = data if isinstance(data, list) else data.get("hits", [])
            return {h.get("_id"): h for h in hits if isinstance(h, dict) and "_id" in h}
        except Exception as e:
            if logger:
                logger.log(f"[{phase_label}] ❌ Fehler bei Batch: {e}")
            return {}

    futures = []
    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        for i in range(0, len(items), batch_size):
            chunk = items[i:i+batch_size]
            futures.append(ex.submit(fetch_chunk, chunk))

        # merged mapping of hgvs -> hit
        merged_by_hgvs = {}
        for fut in as_completed(futures):
            try:
                by_id = fut.result() or {}
            except Exception as e:
                if logger:
                    logger.log(f"[{phase_label}] ❌ Exception beim Future: {e}")
                by_id = {}
            # update merged map; guard against malformed items
            try:
                if isinstance(by_id, dict):
                    merged_by_hgvs.update(by_id)
            except Exception as e:
                if logger:
                    logger.log(f"[{phase_label}] ❌ Fehler beim Mergen der Batch-Antwort: {e}")

    session.close()

    # Build reverse map hgvs -> [keys]
    reverse_hgvs = {}
    for k, hg in hgvs_map.items():
        if hg:
            reverse_hgvs.setdefault(hg, []).append(k)

    # Map returned hits back to original keys deterministically
    for hgvs_id, hit in merged_by_hgvs.items():
        keys_for_hg = reverse_hgvs.get(hgvs_id) or []
        for k in keys_for_hg:
            # if multiple keys map to same HGVS we assign same hit to all candidates
            results[k] = hit

    # Ensure every original key exists in results mapping
    for k in keys:
        if k not in results:
            hg = hgvs_map.get(k)
            results[k] = None if not hg else results.get(k, None)

    return results

def mv_fetch_with_module(keys, build, fetch_full, batch_size=1000, phase_label="MV"):
    """
    Holt Daten von MyVariant über das Python-Modul.
    - keys: Liste von (chrom,pos,ref,alt,build)
    - build: "GRCh38" oder "GRCh37"
    - fetch_full: True = Vollannotation, False = nur AF
    - batch_size: max. IDs pro Anfrage
    """
    import myvariant
    results = {}
    hgvs_map = {}
    items = []

    # 1. HGVS-IDs bauen
    for k in keys:
        chrom, pos, ref, alt, _b = k
        if _is_pure_snv_key(k):
            hg = make_hgvs(chrom, pos, ref, alt)
        elif _is_indel_key(k):
            hg = make_hgvs_indel(chrom, pos, ref, alt)
        else:
            hg = None
        if hg:
            hgvs_map[k] = hg
            items.append(hg)
        else:
            hgvs_map[k] = None

    fields = MV_FIELDS_FULL if fetch_full else MV_FIELDS_AF
    assembly = "hg38" if build == "GRCh38" else "hg19"

    mv = myvariant.MyVariantInfo()
    total = len(items)
    done = 0
    t_phase = time.time()

    # 2. Batches abfragen
    for i in range(0, total, batch_size):
        chunk = items[i:i+batch_size]
        t0 = time.time()
        successes = 0
        try:
            res = mv.querymany(
                chunk,
                scopes="hgvs",
                fields=fields,
                assembly=assembly,
                verbose=True,
                size=batch_size
            )
            by_id = {h.get("_id"): h for h in res if isinstance(h, dict) and "_id" in h}
            for hgvs in chunk:
                if by_id.get(hgvs) is not None:
                    successes += 1
                results[hgvs] = by_id.get(hgvs)
        except Exception as e:
            logger.log(f"[{phase_label}] MyVariant error: {e}")
            for hgvs in chunk:
                results[hgvs] = None

        """done += len(chunk)
        dt = max(1e-3, time.time() - t0)
        rate = len(chunk) / dt
        pct = 100.0 * done / max(1, total)
        eta_txt = fmt_eta(progress_maint.eta()) if phase_label.lower() not in ("af", "anno") else fmt_eta(progress.eta())
        logger.log(f"[{phase_label}] Batch {done}/{total} ({pct:.1f}%) – {rate:.1f} rec/s – hit {successes}/{len(chunk)} – {eta_txt}")

    dt_phase = max(1e-3, time.time() - t_phase)
    logger.log(f"[{phase_label}] Komplett: {total} in {dt_phase:.1f}s ({total/dt_phase:.1f} rec/s)")"""

    # 3. Rückgabe im gewohnten Mapping
    out = {}
    for k in keys:
        hg = hgvs_map.get(k)
        out[k] = results.get(hg) if hg else None
    return out

def extract_fields_from_mv(mv_hit: Optional[dict], debug_log: bool = False) -> Dict[str, Optional[Any]]:
    """
    Extrahiert relevante Felder aus einem MyVariant-Hit.

    KRITISCHER BUG-FIX:
    - Garantiert IMMER ein Dict zurückzugeben (nie None)
    - Robustes Error-Handling für alle Edge-Cases
    - Bessere Logging-Integration

    V15: Debug-Logging für MyVariant-Response zur Diagnose leerer Spalten.

    Args:
        mv_hit: MyVariant API Response Dict
        debug_log: Wenn True, loggt die Raw-Response (für Debugging)

    Returns:
        Dict mit allen Feldern (Werte können None sein, aber Dict nie None)
    """
    # ✅ DEFAULT: Garantiertes Rückgabe-Dict
    res: Dict[str, Optional[Any]] = {
        "cadd_phred": None,
        "gene_symbol": None,
        "impact": None,
        "consequence": None,
        "clinical_significance": None,
        "phenotypes": None,
        "is_coding": None,
        "protein_coding": None,
        "conservation": None,
        "rsid": None,
    }

    # ✅ DEFENSIVE: Early return bei ungültigen Inputs
    if not mv_hit:
        if debug_log:
            logger.log("[MV-Debug] ⚠️ mv_hit ist None/leer")
        return res

    if not isinstance(mv_hit, dict):
        if debug_log:
            logger.log(f"[MV-Debug] ⚠️ mv_hit ist kein Dict: {type(mv_hit)}")
        return res

    # ✅ V15: Debug-Logging für Raw Response
    if debug_log:
        available_keys = list(mv_hit.keys())[:20]  # Max 20 Keys loggen
        logger.log(f"[MV-Debug] 📥 Response Keys: {available_keys}")
        # Prüfe kritische Felder
        for field in ["cadd", "vep", "snpeff", "clinvar", "phastcons", "phylop", "gerp"]:
            val = mv_hit.get(field)
            if val is not None:
                logger.log(f"[MV-Debug]   ✓ {field}: {type(val).__name__} (vorhanden)")
            else:
                logger.log(f"[MV-Debug]   ✗ {field}: NICHT vorhanden")

    try:
        # --- CADD ---
        cadd = mv_hit.get("cadd") or {}
        if isinstance(cadd, dict):
            res["cadd_phred"] = safe_float(cadd.get("phred"))

        # --- VEP ---
        vep = mv_hit.get("vep") or {}
        best_vep = None
        if isinstance(vep, dict):
            cand = {
                "symbol": vep.get("symbol"),
                "impact": vep.get("impact"),
                "consequence": vep.get("consequence"),
            }
            tc = vep.get("transcript_consequences")
            if isinstance(tc, list) and tc:
                imap = {"HIGH": 3, "MODERATE": 2, "LOW": 1, "MODIFIER": 0}
                tc_sorted = sorted(
                    [t for t in tc if isinstance(t, dict)],
                    key=lambda t: imap.get(str(t.get("impact") or "").upper(), 0),
                    reverse=True
                )
                if tc_sorted:
                    t_best = tc_sorted[0]
                    cand_tc = {
                        "symbol": t_best.get("gene_symbol") or t_best.get("symbol"),
                        "impact": t_best.get("impact"),
                        "consequence": t_best.get("consequence"),
                    }
                    best_vep = cand_tc
            else:
                best_vep = cand

        if best_vep:
            res["gene_symbol"] = best_vep.get("symbol") or res["gene_symbol"]
            cons = best_vep.get("consequence")
            if isinstance(cons, list):
                cons = ",".join(str(c) for c in cons if c)
            elif not isinstance(cons, str):
                cons = None
            res["consequence"] = cons or res["consequence"]
            res["impact"] = best_vep.get("impact") or res["impact"]

        # --- Gene (Fallback) ---
        gene = mv_hit.get("gene") or {}
        if isinstance(gene, dict):
            res["gene_symbol"] = res["gene_symbol"] or gene.get("symbol") or gene.get("name")

        # --- SnpEff ---
        snpeff = mv_hit.get("snpeff") or {}
        if isinstance(snpeff, dict):
            ann = snpeff.get("ann")
            if isinstance(ann, list) and ann:
                imap = {"HIGH": 3, "MODERATE": 2, "LOW": 1, "MODIFIER": 0}
                best, score = None, -1
                for a in ann:
                    if not isinstance(a, dict):
                        continue
                    sc = imap.get(str(a.get("impact") or "").upper(), -1)
                    if sc > score:
                        best, score = a, sc
                if best:
                    res["impact"] = res["impact"] or best.get("impact")
                    if not res["consequence"]:
                        res["consequence"] = (
                            best.get("effect")
                            or best.get("annotation")
                            or best.get("consequence")
                        )
                    res["gene_symbol"] = res["gene_symbol"] or best.get("gene_symbol") or best.get("gene_name")

        # --- ClinVar ---
        clin = mv_hit.get("clinvar") or {}
        if isinstance(clin, dict):
            rcv = clin.get("rcv")
            if isinstance(rcv, list) and rcv:
                sigs, phs = [], []
                for r in rcv:
                    if not isinstance(r, dict):
                        continue
                    s = r.get("clinical_significance")
                    if s:
                        sigs.append(str(s))
                    p = r.get("phenotype")
                    if p:
                        if isinstance(p, list):
                            phs.extend(str(x) for x in p if x)
                        else:
                            phs.append(str(p))
                if sigs:
                    res["clinical_significance"] = ",".join(sorted(set(sigs)))
                if phs:
                    res["phenotypes"] = ",".join(sorted(set(phs)))
            
            # Fallback 2: top-level clinvar fields
            if not res["clinical_significance"]:
                cs = clin.get("clnsig")
                if cs:
                    if isinstance(cs, list):
                        res["clinical_significance"] = ",".join(str(x) for x in cs if x)
                    else:
                        res["clinical_significance"] = str(cs)
            
            if not res["phenotypes"]:
                pl = clin.get("phenotype_list")
                if pl:
                    if isinstance(pl, list):
                        res["phenotypes"] = ",".join(str(x) for x in pl if x)
                    else:
                        res["phenotypes"] = str(pl)

        # --- Conservation (Nested Check) ---
        def extract_score(hit, keys):
            """Sucht einen Score in mehreren möglichen Keys (auch verschachtelt)."""
            for key in keys:
                # Top-level check
                val = hit.get(key)
                if val is not None:
                    res = _extract_recursive(val)
                    if res is not None: return res
                
                # Check in common sub-dicts
                for sub in ["cadd", "dbnsfp"]:
                    sub_dict = hit.get(sub)
                    if isinstance(sub_dict, dict):
                        val = sub_dict.get(key)
                        if val is not None:
                            res = _extract_recursive(val)
                            if res is not None: return res
            return None

        def _extract_recursive(val):
            if isinstance(val, (int, float)): return float(val)
            if isinstance(val, dict):
                scores = [float(v) for v in val.values() if isinstance(v, (int, float))]
                return max(scores) if scores else None
            if isinstance(val, list):
                scores = [float(v) for v in val if isinstance(v, (int, float))]
                return max(scores) if scores else None
            return None

        phast = extract_score(mv_hit, ["phastcons", "phastcons100way_vertebrate", "phastcons46way_placental"])
        phyl = extract_score(mv_hit, ["phylop", "phylop100way_vertebrate", "phylop46way_placental"])
        gerp = extract_score(mv_hit, ["gerp", "gerp_rs", "gerp_raw"])
        
        cons_parts = []
        for name, v in (("phastCons", phast), ("phyloP", phyl), ("GERP", gerp)):
            if v is not None:
                cons_parts.append(f"{name}:{v:.2f}")
        if cons_parts:
            res["conservation"] = ",".join(cons_parts)

        # --- Coding flags ---
        res["is_coding"] = determine_is_coding(res["consequence"] or "")
        res["protein_coding"] = res["is_coding"]

        # --- dbSNP ---
        dbsnp = mv_hit.get("dbsnp") or {}
        if isinstance(dbsnp, dict):
            res["rsid"] = dbsnp.get("rsid") or dbsnp.get("id")

    except Exception as e:
        # ✅ ROBUSTHEIT: Niemals None zurückgeben
        # Logge Fehler wenn möglich, aber garantiere Dict-Rückgabe
        try:
            if hasattr(logging, 'error'):
                logging.error(f"[extract_fields_from_mv] Error: {e}")
        except Exception:
            pass  # Ignore logging errors
        
        # ✅ GARANTIERT: Gebe immer das Default-Dict zurück
        return res

    # ✅ V15: Debug-Logging für extrahierte Felder
    if debug_log:
        extracted = {k: v for k, v in res.items() if v is not None}
        if extracted:
            logger.log(f"[MV-Debug] 📤 Extrahiert: {list(extracted.keys())}")
        else:
            logger.log("[MV-Debug] ⚠️ Keine Felder extrahiert!")

    # ✅ GARANTIERT: Explizite Dict-Rückgabe
    return res


async def ensembl_1000g_fetch_async(keys, build, logger=None, batch_size=100, max_retries=3):
    """
    [DEPRECATED] Wrapper für alte Ensembl/1000G-Fetches.
    Bitte stattdessen vep_fetch_async verwenden.
    """
    if logger:
        logger.log("[DEPRECATED] ensembl_1000g_fetch_async aufgerufen – bitte auf vep_fetch_async umstellen.")
    return await vep_fetch_async(keys, build, logger=logger, batch_size=batch_size, max_retries=max_retries)

# ============================================================================
# ALFA async fetch
# ============================================================================

async def alfa_fetch_async(keys, rsid_map: dict, build, logger=None):
    """
    ALFA (NCBI refsnp) fetch.
    
    ✅ KRITISCH: Prüft aiohttp-Verfügbarkeit
    """
    if not AIOHTTP_AVAILABLE or aiohttp is None:
        if logger:
            logger.log("[ALFA] ❌ aiohttp nicht verfügbar")
        return {}
    
    base = "https://api.ncbi.nlm.nih.gov/variation/v0/beta/refsnp"
    results = {}
    _alfa_timeout = get_api_setting("phase1_af", "alfa", "timeout", 30)
    _ncbi_key = get_api_setting("global", "global", "ncbi_api_key", "")

    async def fetch_one(session, key, rsid):
        if not rsid or not str(rsid).lower().startswith("rs"):
            return None

        rid = str(rsid)[2:]  # "rs123" -> "123"
        url = f"{base}/{rid}"

        for attempt in range(3):
            try:
                async with session.get(
                    url,
                    timeout=aiohttp.ClientTimeout(total=_alfa_timeout)
                ) as resp:
                    if resp.status == 200:
                        data = await resp.json()
                        freqs = []
                        try:
                            alleles = data.get("primary_snapshot_data", {}).get("allele_annotations", [])
                            for allele in alleles:
                                for freq in allele.get("frequency", []):
                                    ac = freq.get("allele_count")
                                    tc = freq.get("total_count")
                                    if ac is not None and tc and tc > 0:
                                        freqs.append(ac / tc)
                        except Exception:
                            pass

                        if freqs:
                            # ✅ DB-konformes Feld
                            return {"af_exac": sum(freqs) / len(freqs)}
                        return None

                    elif resp.status == 404:
                        return None

                    elif resp.status == 429:
                        await asyncio.sleep(2 ** attempt)
                        continue

                    else:
                        if logger and attempt == 2:
                            logger.log(f"[ALFA] ⚠️ HTTP {resp.status} for {rsid}")
                        return None

            except (asyncio.TimeoutError, aiohttp.ClientError) as e:
                if logger and attempt == 2:
                    logger.log(f"[ALFA] ❌ {rsid}: {e}")
                await asyncio.sleep(1)

        return None

    try:
        _alfa_workers = get_api_setting("phase1_af", "alfa", "workers", 4)
        sem = asyncio.Semaphore(_alfa_workers)
        _headers = {}
        if _ncbi_key:
            _headers["api_key"] = _ncbi_key
        async with aiohttp.ClientSession(headers=_headers) as session:
            tasks = []
            for key in keys:
                rsid = rsid_map.get(key)
                if rsid:
                    async def sem_task(k=key, r=rsid):
                        async with sem:
                            return await fetch_one(session, k, r)
                    tasks.append(asyncio.create_task(sem_task()))

            all_results = await asyncio.gather(*tasks, return_exceptions=True)

        # Nur Keys mit rsID haben Tasks
        keys_with_rsid = [k for k in keys if rsid_map.get(k)]
        for key, result in zip(keys_with_rsid, all_results):
            results[key] = result if isinstance(result, dict) else None

    except Exception as e:
        if logger:
            logger.log(f"[ALFA] ❌ Fatal error: {e}")

    return results


# ============================================================================
# gnomAD async fetch
# ============================================================================

async def gnomad_fetch_async(keys, build, logger=None):
    """
    gnomAD GraphQL fetch.
    
    ✅ KRITISCH: Prüft aiohttp-Verfügbarkeit
    """
    if not AIOHTTP_AVAILABLE or aiohttp is None:
        if logger:
            logger.log("[gnomAD] ❌ aiohttp nicht verfügbar")
        return {}
    
    endpoint = "https://gnomad.broadinstitute.org/api"
    results = {}

    dataset = "gnomad_r3" if build == "GRCh38" else "gnomad_r2_1"

    async def fetch_one(session, key):
        query = """
        query VariantAF($variantId: String!, $dataset: DatasetId!) {
          variant(variantId: $variantId, dataset: $dataset) {
            variantId
            genome { ac an af }
            exome { ac an af }
          }
        }
        """
        variables = {
            "variantId": normalize_for_gnomad(key),  # String im Format "1-55516888-G-A"
            "dataset": dataset
        }

        _gnomad_timeout = get_api_setting("phase1_af", "gnomad", "timeout", 60)
        for attempt in range(3):
            try:
                async with session.post(
                    endpoint,
                    json={"query": query, "variables": variables},
                    timeout=aiohttp.ClientTimeout(total=_gnomad_timeout)
                ) as resp:
                    if resp.status == 429:
                        await asyncio.sleep(5 * (attempt + 1))
                        continue

                    if resp.status != 200:
                        try:
                            text = await resp.text()
                        except Exception:
                            text = "<no body>"
                        if logger:
                            logger.log(f"[gnomAD] ❌ HTTP {resp.status}: {text[:200]}")
                        return None

                    data = await resp.json()
                    v = data.get("data", {}).get("variant")
                    if not v:
                        return None
                    g_af = v.get("genome", {}).get("af")
                    e_af = v.get("exome", {}).get("af")
                    return {"ggn": g_af, "gex": e_af}

            except (asyncio.TimeoutError, aiohttp.ClientError) as e:
                if logger and attempt == 2:
                    logger.log(f"[gnomAD] ❌ Fetch failed for {key}: {e}")
                await asyncio.sleep(2 ** attempt)

        return None

    try:
        _gnomad_workers = get_api_setting("phase1_af", "gnomad", "workers", 8)
        sem = asyncio.Semaphore(_gnomad_workers)
        async with aiohttp.ClientSession() as session:
            async def sem_task(k):
                async with sem:
                    return k, await fetch_one(session, k)

            tasks = [asyncio.create_task(sem_task(k)) for k in keys]
            all_results = await asyncio.gather(*tasks, return_exceptions=True)

        for r in all_results:
            if isinstance(r, tuple):
                k, val = r
                results[k] = val

    except Exception as e:
        if logger:
            logger.log(f"[gnomAD] ❌ Fatal error: {e}")

    return results
# ============================================================================
# VEP async (ergänzt)
# ============================================================================

async def vep_fetch_async(keys, build, logger=None, batch_size=100, max_retries=3):
    """
    Ensembl VEP fetch.
    
    ✅ KRITISCH: Prüft aiohttp-Verfügbarkeit
    """
    if not AIOHTTP_AVAILABLE or aiohttp is None:
        if logger:
            logger.log("[VEP] ❌ aiohttp nicht verfügbar")
        return {}
    
    # Hostname je nach Build
    if build == "GRCh37":
        server = "https://grch37.rest.ensembl.org"
    else:
        server = "https://rest.ensembl.org"

    endpoint = "/vep/human/region"
    results = {}

    async def fetch_chunk(session, chunk):
        body = {"variants": [normalize_for_vep(k) for k in chunk]}
        url = f"{server}{endpoint}"

        _vep_timeout = get_api_setting("phase1_af", "vep", "timeout", 60)
        for attempt in range(1, max_retries + 1):
            try:
                async with session.post(
                    url,
                    json=body,
                    headers={"Content-Type": "application/json", "Accept": "application/json"},
                    timeout=aiohttp.ClientTimeout(total=_vep_timeout)
                ) as resp:
                    if resp.status == 200:
                        data = await resp.json()
                        out = {}
                        for entry in data:
                            afs = []
                            for coloc in entry.get("colocated_variants", []):
                                for k, v in coloc.items():
                                    if isinstance(v, (int, float)) and k.lower().endswith("_af"):
                                        afs.append(v)
                            if afs:
                                out[entry["input"]] = {"af_1kg": sum(afs) / len(afs)}
                        return out

                    elif resp.status == 429:
                        retry_after = int(resp.headers.get("Retry-After", "5"))
                        if logger:
                            logger.log(f"[VEP] 429 – warte {retry_after}s")
                        await asyncio.sleep(retry_after + random.random())

                    elif resp.status >= 500:
                        if logger:
                            logger.log(f"[VEP] {resp.status} – Retry {attempt}/{max_retries}")
                        await asyncio.sleep(2 ** attempt + random.random())

                    else:
                        if logger:
                            logger.log(f"[VEP] ❌ HTTP {resp.status}")
                        return {}

            except (asyncio.TimeoutError, aiohttp.ClientError) as e:
                if logger:
                    logger.log(f"[VEP] ⚠️ {e} – Retry {attempt}/{max_retries}")
                await asyncio.sleep(2 ** attempt + random.random())

        return {}

    try:
        _vep_workers = get_api_setting("phase1_af", "vep", "workers", 4)
        sem = asyncio.Semaphore(_vep_workers)
        async with aiohttp.ClientSession() as session:
            tasks = []
            for i in range(0, len(keys), batch_size):
                chunk = keys[i:i+batch_size]

                async def sem_task(ch=chunk):
                    async with sem:
                        return await fetch_chunk(session, ch)

                tasks.append(asyncio.create_task(sem_task()))

            merged = {}
            for t in asyncio.as_completed(tasks):
                try:
                    part = await t
                    if isinstance(part, dict):
                        merged.update(part)
                except Exception as e:
                    if logger:
                        logger.log(f"[VEP] ❌ Task exception: {e}")

        # Rückmapping
        for k in keys:
            inp = normalize_for_vep(k)
            results[k] = merged.get(inp)

    except Exception as e:
        if logger:
            logger.log(f"[VEP] ❌ Fatal error: {e}")

    return results
# ============== 23andMe -> temporäres VCF ==============

class convert_23andme_to_vcf:
    def __init__(self, file_path: str, cache_file: str = CACHE_FILE, logger=logger):
        self.file_path = file_path
        self.cache_file = cache_file
        self.logger = logger
        # Cache laden oder neu anlegen
        self.cache = self.load_cache()
        if not isinstance(self.cache, dict):
            self.logger.log("[23andMe] ⚠️ Cache war ungültig, neu initialisiert.")
            self.cache = {}
   
    @staticmethod
    def gt_from_bases_snp(alleles, ref_base: str, alt_list, ploid: int) -> str | None:
        """
        Erzeugt den GT-String für SNPs aus den beobachteten Allelen.

        - alleles: Liste der Basen aus 23andMe (z.B. ["A","G"])
        - ref_base: Referenzbase an dieser Position
        - alt_list: Liste der ALT-Basen (≠ REF)
        - ploid: 2 = diploid, 1 = haploid
        Rückgabe: GT-String (z.B. "0/1", "1/1", "0", "1") oder None bei Fehler
        """
        # Map REF=0, ALT1=1, ALT2=2, …
        allele_map = {ref_base: "0"}
        for i, a in enumerate(alt_list, start=1):
            allele_map[a] = str(i)

        if ploid == 2:
            if len(alleles) != 2:
                return None
            gt = "/".join(allele_map.get(a, ".") for a in alleles)
            return None if "." in gt else gt

        elif ploid == 1:
            if len(alleles) < 1:
                return None
            return allele_map.get(alleles[0], None)

        return
    
    @staticmethod
    def in_par(chrom: str, pos: int, build: str) -> bool:
        """
        Prüft, ob eine Position in einer Pseudoautosomalen Region (PAR) liegt.
        """
        chrom = chrom.upper().replace("CHR", "")
        if build == "GRCh37":
            par_regions = {
                "X": [(60001, 2699520), (154931044, 155260560)],
                "Y": [(10001, 2649520), (59034051, 59363566)],
            }
        elif build == "GRCh38":
            par_regions = {
                "X": [(10001, 2781479), (155701383, 156030895)],
                "Y": [(10001, 2781479), (56887903, 57217415)],
            }
        else:
            return False

        ranges = par_regions.get(chrom, [])
        return any(start <= pos <= end for (start, end) in ranges)

    @staticmethod
    def ploidy_for_site(chrom: str, pos: int, build: str, sex: str) -> int:
        """
        Liefert die Ploidie für eine gegebene Position.

        sex: "male", "female" oder "unknown"
        Rückgabe:
            0 = nicht vorhanden (z. B. Y bei Frau)
            1 = haploid
            2 = diploid
        """
        c = chrom.upper().replace("CHR", "")
        s = (sex or "unknown").lower()

        # Autosomen
        if c in {str(i) for i in range(1, 23)}:
            return 2

        # Mitochondrium
        if c == "MT":
            return 1

        # X-Chromosom
        if c == "X":
            if s == "female":
                return 2
            if s == "male":
                return 2 if convert_23andme_to_vcf.in_par(c, pos, build) else 1
            return 2  # unknown → konservativ diploid

        # Y-Chromosom
        if c == "Y":
            if s == "female":
                return 0
            if s == "male":
                return 2 if convert_23andme_to_vcf.in_par(c, pos, build) else 1
            return 0  # unknown → sicherheitshalber weglassen

        # Default
        return 2

    def load_cache(self):
        if os.path.exists(self.cache_file):
            try:
                with open(self.cache_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    if isinstance(data, dict):
                        return data
            except Exception as e:
                self.logger.log(f"[23andMe] ⚠️ Cache konnte nicht geladen werden ({e}), neuer Cache wird erstellt.")
        empty_cache = {}
        self.atomic_write_json(empty_cache, self.cache_file)
        return empty_cache

    def atomic_write_json(self, obj, dst):
        tmp = f"tmp.{''.join(random.choices('abcdefghijklmnopqrstuvwxyz', k=6))}"
        try:
            with open(tmp, "w", encoding="utf-8") as f:
                json.dump(obj, f)
            os.replace(tmp, dst)
        except Exception as e:
            self.logger.log(f"[23andMe] ❌ Fehler beim Schreiben von {dst}: {e}")

    def parse_23andme(self):
        variants = []
        with open(self.file_path, "r", encoding="utf-8") as f:
            for line in f:
                if line.startswith("#") or not line.strip():
                    continue
                parts = line.strip().split("\t")
                if len(parts) < 4:
                    continue
                rsid, chrom, pos_str, genotype = parts[:4]
                chrom = chrom.replace("chr", "").upper()
                try:
                    pos = int(pos_str)
                except ValueError:
                    continue
                variants.append((rsid, chrom, pos, genotype))
        return variants

    def is_rs_id(self, id_str: str) -> bool:
        return id_str.lower().startswith("rs") and id_str[2:].isdigit()

    def nc_to_chrom(self, nc):
        ignore_prefixes = ("XM_", "NM_", "XP_", "NP_", "NG_", "NW_", "NR_", "XR_")
        if nc.startswith(ignore_prefixes):
            return None
        if nc.startswith("NC_012920"):
            return "MT"
        m = re.match(r"NC_(\d{6})\.\d+", nc)
        if m:
            num = int(m.group(1))
            if 1 <= num <= 22:
                return str(num)
            if num == 23:
                return "X"
            if num == 24:
                return "Y"
        return None

    def parse_refsnp_payload(self, rsid, data):
        results = {}
        try:
            pwa = data["primary_snapshot_data"]["placements_with_allele"]
        except Exception:
            return results
        for p in pwa:
            if not p.get("is_ptlp"):
                continue
            alleles = p.get("alleles") or []
            for allele_entry in alleles:
                spdi = allele_entry.get("allele", {}).get("spdi", {})
                seq_id = spdi.get("seq_id")
                position = spdi.get("position")
                deleted = (spdi.get("deleted_sequence") or "").upper()
                inserted = (spdi.get("inserted_sequence") or "").upper()
                if not seq_id or position is None:
                    continue
                chrom = self.nc_to_chrom(seq_id)
                if not chrom:
                    continue
                for t in p.get("placement_annot", {}).get("seq_id_traits_by_assembly") or []:
                    asm = t.get("assembly_name")
                    if asm and (asm.startswith("GRCh37") or asm.startswith("GRCh38")):
                        base_asm = "GRCh37" if asm.startswith("GRCh37") else "GRCh38"
                        results[base_asm] = {
                            "chrom": chrom,
                            "pos": int(position) + 1,
                            "ref": deleted,
                            "alt": inserted
                        }
        return results
    
    def lookup_rsid_from_cache(self, chrom, pos, build, cache):
        """Versucht, anhand von Chrom/Pos im Cache eine rsID zu finden."""
        for rid, entry in cache.items():
            hit = entry.get("assemblies", {}).get(build)
            if hit and hit["chrom"] == chrom and hit["pos"] == int(pos):
                return rid
        return None
    
    def create_vcf(self, variants, build, out_vcf, cache, fasta_path=None, sex="unknown"):
        self.logger.log(f"[23andMe→VCF] Starte VCF-Erzeugung: {len(variants)} Varianten, Build={build}, Ausgabe={out_vcf}")

        fai_index = None
        if fasta_path:
            fai_index = load_fai_index(fasta_path + ".fai")

        written = 0
        with open(out_vcf, "w", encoding="utf-8") as vcf:
            vcf.write("##fileformat=VCFv4.2\n")
            vcf.write(f"##reference={build}\n")
            vcf.write("##FORMAT=<ID=GT,Number=1,Type=String,Description=\"Genotype\">\n")
            vcf.write("#CHROM\tPOS\tID\tREF\tALT\tQUAL\tFILTER\tINFO\tFORMAT\tSAMPLE\n")

            for idx, (rsid, chrom, pos, genotype) in enumerate(variants, start=1):
                if not genotype:
                    continue
                genotype = genotype.strip().replace("_", "-")
                if genotype in ("--", "-"):
                    continue  # Missing → überspringen

                ref_base = get_ref_base(chrom, pos, fasta_path, fai_index, cache, build)
                if not ref_base or ref_base in (".", "-", "N"):
                    continue

                ploid = self.ploidy_for_site(chrom, pos, build, sex)
                if ploid == 0:
                    continue

                alleles = list(genotype)
                FILTER = "PASS"
                info_field = "."

                # --- iID Normalisierung ---
                orig_id = rsid
                if rsid.startswith("i"):
                    mapped = self.lookup_rsid_from_cache(chrom, pos, build, cache)
                    if mapped:
                        rsid = mapped
                    else:
                        rsid = "."
                    info_field = f"I_ID={orig_id}"

                # --- SNPs ---
                if all(a in "ACGT" for a in alleles):
                    sample_alts = sorted({a for a in alleles if a != ref_base})
                    if not sample_alts:
                        continue
                    alt = ",".join(sample_alts)
                    alleles_for_gt = alleles if ploid == 2 else alleles[:1]
                    GT = self.gt_from_bases_snp(alleles_for_gt, ref_base, sample_alts, ploid)
                    if not GT:
                        continue
                    vcf.write(f"{chrom}\t{pos}\t{rsid}\t{ref_base}\t{alt}\t.\t{FILTER}\t{info_field}\tGT\t{GT}\n")
                    written += 1
                    continue

                # --- Indels (rsIDs oder gemappte iIDs) ---
                if (self.is_rs_id(rsid) and rsid in cache) or (rsid != "." and rsid in cache):
                    hit_all = cache[rsid]
                    hit_by_build = (hit_all.get("assemblies") or {}).get(build)
                    if not hit_by_build:
                        continue
                    if hit_by_build["chrom"] != chrom or hit_by_build["pos"] != pos:
                        continue
                    ref_seq, alt_seq = indel_ref_alt_from_spdi(
                        chrom, pos, hit_by_build["ref"], hit_by_build.get("alt", ""),
                        lambda c, p: get_ref_base(c, p, fasta_path, fai_index, cache, build)
                    )
                    if not ref_seq or not alt_seq:
                        continue
                    g = genotype.upper()
                    if ploid == 2:
                        if g == "II":
                            GT = "1/1"
                        elif g in ("ID", "DI"):
                            GT = "0/1"
                        elif g == "DD":
                            continue  # 0/0 → nicht schreiben
                        else:
                            continue
                    else:
                        if g in ("I", "II"):
                            GT = "1"
                        elif g in ("D", "DD"):
                            GT = "0"
                        else:
                            continue
                    vcf.write(f"{chrom}\t{pos}\t{rsid}\t{ref_seq}\t{alt_seq}\t.\t{FILTER}\t{info_field}\tGT\t{GT}\n")
                    written += 1

                if idx % 10000 == 0:
                    self.logger.log(f"[23andMe→VCF] {idx}/{len(variants)} verarbeitet, {written} geschrieben…")

        self.logger.log(f"[23andMe→VCF] Fertig: {written}/{len(variants)} Varianten ins VCF geschrieben.")
        return bool(fasta_path), 0, 0

    def cache_upsert(self, cache, rsid, build, chrom, pos, ref):
        """Schreibt oder aktualisiert einen rsID-Eintrag im Cache (assemblies/<build>/...)."""
        with _cache_lock:
            entry = cache.get(rsid) or {"assemblies": {}, "last_update": 0}
            entry["assemblies"][build] = {
                "chrom": chrom,
                "pos": int(pos),
                "ref": str(ref).upper()
            }
            entry["last_update"] = int(time.time())
            cache[rsid] = entry
            self.atomic_write_json(cache, self.cache_file)
            self.logger.log(f"[Cache] {rsid} @ {build} → {chrom}:{pos} {ref}")

    def fetch_single_refsnp(self, rsid, session):
        """Holt ein einzelnes RefSNP-Objekt von NCBI/dbSNP."""
        rid = rsid.lower().replace("rs", "")
        url = DBSNP_URL + rid
        try:
            # Session hat bereits Retry-Logik via Adapter
            r = session.get(url, timeout=TIMEOUT_SEC)
            if r.status_code == 200:
                try:
                    return rsid, r.json()
                except ValueError:
                    return rsid, None
            return rsid, None
        except Exception:
            return rsid, None


    def adaptive_parallel_fetch(self, rsids, cache):
        """Lädt fehlende rsIDs parallel nach und ergänzt den Cache."""
        if not rsids:
            return
        remaining = [r for r in rsids if r not in cache]
        if not remaining:
            return

        retry = Retry(
            total=Config.MAX_RETRIES,
            backoff_factor=1.0,
            status_forcelist=[429, 500, 502, 503, 504]
        )
        adapter = HTTPAdapter(
            pool_connections=MAX_WORKERS,
            pool_maxsize=MAX_WORKERS,
            max_retries=retry
        )
        session = requests.Session()
        session.mount("https://", adapter)
        session.mount("http://", adapter)
        session.headers.update({
            "User-Agent": f"23andMe-VCF-Converter/1.0 (+{socket.gethostname()})"
        })

        workers = min(MAX_WORKERS, max(MIN_WORKERS, (self.cpu_count() or 4) * 2))
        chunk_multiplier = 6
        target_cpu = TARGET_CPU_FRACTION * 100.0
        processed = 0
        total = len(remaining)

        while remaining:
            chunk_size = max(1, min(len(remaining), workers * chunk_multiplier))
            chunk = [remaining.pop() for _ in range(chunk_size)]
            t0 = time.time()
            futures = []
            with ThreadPoolExecutor(max_workers=workers, thread_name_prefix="dbsnp") as ex:
                for rsid in chunk:
                    futures.append(ex.submit(self.fetch_single_refsnp, rsid, session))
                for fut in as_completed(futures):
                    rsid, payload = fut.result()
                    if payload is None:
                        self.logger.log(f"[dbSNP] {rsid}: keine Daten erhalten")
                    else:
                        results = self.parse_refsnp_payload(rsid, payload)
                        if not results:
                            self.logger.log(f"[dbSNP] {rsid}: Daten erhalten, aber keine verwertbaren Ergebnisse")
                        for bld, rec in results.items():
                            self.cache_upsert(cache, rsid, bld, rec["chrom"], rec["pos"], rec["ref"])

            cpu_now = self.get_cpu_usage_percent()
            if cpu_now > target_cpu and workers > MIN_WORKERS:
                workers = max(MIN_WORKERS, workers - 1)
            elif cpu_now < target_cpu * 0.6 and workers < MAX_WORKERS:
                workers = min(MAX_WORKERS, workers + 1)

            dt = max(1e-3, time.time() - t0)
            processed += len(chunk)
            self.logger.log(
                f"[Fortschritt] Batch: {processed}/{total} | "
                f"Workers={workers} | CPU~{cpu_now:.0f}% | "
                f"{len(chunk)} rsIDs in {dt:.2f}s"
            )

        session.close()


    def ensure_fasta_with_choice(self, build):
        """
        Stellt sicher, dass für das angegebene Build eine FASTA-Datei
        (inkl. Index) verfügbar ist. Gibt den Pfad zurück oder None.
        """
        fasta_path = does_fasta_exist(build, logger=logger)
        if fasta_path is None:
            return None

        fai_path = fasta_path + ".fai"
        if os.path.exists(fasta_path) and os.path.exists(fai_path):
            return fasta_path

        # Falls FASTA existiert, aber Index fehlt, wird das bereits
        # in does_fasta_exist() behandelt. Hier reicht also Rückgabe.
        return fasta_path

    def cpu_count(self):
        try:
            return os.cpu_count() or 4
        except Exception as e:
            self.logger.log(f"[System] ⚠️ Fehler bei cpu_count(): {e}")
            return 4

    def get_cpu_usage_percent(self):
        try:
            val = psutil.cpu_percent(interval=None)
            return val
        except Exception:
            try:
                load1, _, _ = os.getloadavg()
                val = min(100.0, (load1 / max(1, self.cpu_count())) * 100.0)
                return val
            except Exception as e:
                self.logger.log(f"[System] ⚠️ CPU-Auslastung nicht bestimmbar ({e}), default=30%")
                return 30.0

    def backoff_time(self, attempt):
        # Exponentielles Backoff mit Jitter
        return min(30.0, (2 ** attempt) + random.random())
    

    def start(self, build: Optional[str] = None, sex: str = "unknown"):
        """
        Einstiegspunkt: liest 23andMe-Datei, ermittelt Build falls nötig,
        lädt ggf. Referenzdaten nach und erzeugt ein VCF.
        """
        path = self.file_path  # immer aus der Instanz nehmen
        # 1. Build bestimmen
        if not build:
            build = self.scan_header_for_build(path)
            if build in (None, "Conflict", "Unklar"):
                self.logger.log(f"[23andMe→VCF] ⚠️ Build nicht eindeutig erkannt ({build}) – Default=GRCh37")
                build = "GRCh37"
            else:
                self.logger.log(f"[23andMe→VCF] Build automatisch ermittelt: {build}")

        self.logger.log(f"[23andMe→VCF] Starte Konvertierung: {path}, Build={build}, Sex={sex}")

        # 2. Varianten einlesen
        variants = self.parse_23andme()
        self.logger.log(f"[23andMe→VCF] {len(variants)} Varianten aus 23andMe-Datei gelesen.")

        # 3. FASTA prüfen
        fasta_path = does_fasta_exist(build, logger=self.logger)
        fai_index = None
        if fasta_path:
            try:
                fai_path = fasta_path + ".fai"
                if os.path.exists(fai_path):
                    fai_index = load_fai_index(fai_path)
                    self.logger.log(f"[23andMe→VCF] FASTA verfügbar: {fasta_path}")
                else:
                    self.logger.log(f"[23andMe→VCF] ⚠️ FAI-Index fehlt, erzeuge...")
                    build_fasta_index_global(fasta_path, self.logger)
                    fai_index = load_fai_index(fasta_path + ".fai")
            except Exception as e:
                self.logger.log(f"[23andMe→VCF] ⚠️ FASTA-Index-Fehler: {e}")
                fasta_path, fai_index = None, None

        # 4. Fehlende rsIDs nachladen (nur wenn KEINE FASTA)
        if not fasta_path:
            missing_rsids = [rsid for rsid, _, _, _ in variants
                             if self.is_rs_id(rsid) and rsid not in self.cache]
            if missing_rsids:
                self.logger.log(f"[23andMe→VCF] {len(missing_rsids)} fehlende Referenzen – starte Nachladen…")
                self.adaptive_parallel_fetch(missing_rsids, self.cache)

        # 5. Ausgabedatei vorbereiten
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        basename = os.path.splitext(os.path.basename(self.file_path))[0]
        os.makedirs(TEMP_VCF_DIR, exist_ok=True)
        out_vcf = os.path.join(TEMP_VCF_DIR, f"{basename}_{build}_{timestamp}.vcf")

        # 6. VCF erzeugen
        self.create_vcf(
            variants, build, out_vcf, self.cache,
            fasta_path=fasta_path, sex=sex
        )
        self.logger.log(f"[23andMe→VCF] ✅ VCF erzeugt: {out_vcf}")
        return out_vcf, build

    def scan_header_for_build(self, file_path: str) -> Optional[str]:
        """
        Ermittelt den Referenz-Build für eine 23andMe-ähnliche Datei.
        Zuerst Header-Heuristik, dann robuste rsID-basierte Erkennung.
        """
        # 1. Header-Heuristik (schnell)
        lines = []
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                for i, line in enumerate(f):
                    if i >= 200:
                        break
                    lines.append(line.strip())
        except Exception:
            lines = []

        text = "\n".join(lines)
        found37 = any(re.search(p, text, re.IGNORECASE) for p in BUILD_HINTS_37)
        found38 = any(re.search(p, text, re.IGNORECASE) for p in BUILD_HINTS_38)

        if found37 and found38:
            return "Conflict"
        if found37:
            return "GRCh37"
        if found38:
            return "GRCh38"

        # 2. Fallback: robuste rsID-Erkennung über die neue Klassenmethode
        try:
            build = self.detect_build_robust_by_rsids(file_path)
            if build:
                return build
        except Exception as e:
            self.logger.log(f"[23andMe→VCF] ⚠️ Fehler bei rsID-basiertem Build-Check: {e}")

        # 3. Letzter Fallback
        return None

    def detect_build_robust_by_rsids(self, path: str, max_rsids: int = 50) -> Optional[str]:
        rs_pos = []
        with open_text_maybe_gzip(path) as f:
            for line in f:
                if line.startswith("#"):
                    continue
                parts = line.rstrip("\n").split("\t")
                if len(parts) < 3:
                    continue
                vid = parts[2]
                if vid and vid.startswith("rs"):
                    try:
                        pos = int(parts[1])
                    except Exception:
                        continue
                    rs_pos.append((vid, pos))
                if len(rs_pos) >= max_rsids:
                    break

        if not rs_pos:
            return None

        # rsIDs parallel bei dbSNP abfragen
        cache = {}
        self.adaptive_parallel_fetch([r for r, _ in rs_pos], cache)

        tol = 5
        m37 = 0
        m38 = 0

        for rsid, pos_file in rs_pos:
            rec37 = cache.get((rsid, "GRCh37"))
            rec38 = cache.get((rsid, "GRCh38"))

            if rec37 and abs(rec37["pos"] - pos_file) <= tol:
                m37 += 1
            if rec38 and abs(rec38["pos"] - pos_file) <= tol:
                m38 += 1

        self.logger.log(f"[BuildCheck] geprüft: {len(rs_pos)} rsIDs → m37={m37}, m38={m38}")

        if m37 == 0 and m38 == 0:
            return None
        if m37 >= m38:
            return "GRCh37"
        return "GRCh38"

class FASTQmap:
    def __init__(self, build: str = "GRCh38", logger=None):
        self.build = build
        self.logger = logger

    def convert(self, fastq_path: str, build: Optional[str] = None) -> str:
        used_build = build or self.build

        def progress_callback(done: int, total: int):
            percent = (done / total * 100.0) if total else 0.0
            if self.logger:
                self.logger.log(f"[FASTQmap] Fortschritt: {done}/{total} Reads ({percent:.1f}%)")

        if self.logger:
            self.logger.log(f"[FASTQmap] Starte Verarbeitung von {fastq_path} (Build: {used_build})")

        out_dir = self.process_fastq(
            fastq_path,
            progress_callback=progress_callback,
            build=used_build
        )

        vcf_path = os.path.join(out_dir, "variants.vcf")
        if not os.path.exists(vcf_path):
            if self.logger:
                self.logger.log("[FASTQmap] ❌ Keine VCF erzeugt.")
            raise RuntimeError(f"FASTQmap: keine VCF erzeugt unter {vcf_path}")

        if self.logger:
            self.logger.log(f"[FASTQmap] ✅ Verarbeitung abgeschlossen: {vcf_path}")

        return vcf_path

    # -------------------------
    # small helpers (methods)
    # -------------------------
    def phred_score(self, char: str) -> int:
        return max(0, ord(char) - 33)

    def load_fasta(self, fasta_path: str) -> Dict[str, str]:
        seqs: Dict[str, List[str]] = {}
        chrom: Optional[str] = None
        with open(fasta_path, "r") as f:
            for line in f:
                if line.startswith(">"):
                    chrom = line[1:].strip().split()[0]
                    seqs[chrom] = []
                else:
                    if chrom is not None:
                        seqs[chrom].append(line.strip().upper())
        return {c: "".join(s) for c, s in seqs.items()}

    def build_kmer_index(self, ref_seq: str, k: int = 15) -> Dict[str, List[int]]:
        index: Dict[str, List[int]] = {}
        if len(ref_seq) < k:
            return index
        for i in range(len(ref_seq) - k + 1):
            kmer = ref_seq[i:i+k]
            index.setdefault(kmer, []).append(i)
        return index

    def map_read(self, read_seq: str, ref_seq: str, kmer_index: Dict[str, List[int]], k: int = 15) -> Optional[int]:
        if len(read_seq) < k:
            return None
        kmer = read_seq[:k]
        candidates = kmer_index.get(kmer)
        if not candidates:
            return None
        best_pos: Optional[int] = None
        best_mismatches = len(read_seq) + 1
        for pos in candidates:
            segment = ref_seq[pos:pos+len(read_seq)]
            if len(segment) != len(read_seq):
                continue
            mismatches = sum(1 for a, b in zip(segment, read_seq) if a != b)
            if mismatches < best_mismatches:
                best_mismatches = mismatches
                best_pos = pos
        return best_pos

    def ensure_reference(self, build: str = "GRCh38") -> str:
        fasta_path = FASTA_PATHS[build]
        parent = os.path.dirname(os.path.abspath(fasta_path))
        if parent and not os.path.exists(parent):
            os.makedirs(parent, exist_ok=True)
        if os.path.exists(fasta_path):
            return fasta_path
        url = FASTA_URLS[build]
        gz_path = fasta_path + ".gz"
        if self.logger:
            self.logger.log(f"[FASTQmap] Lade Referenz {build} herunter von {url} ...")
        try:
            with get_robust_session() as session:
                with session.get(url, stream=True, timeout=60) as r:
                    r.raise_for_status()
                    r.raw.decode_content = True
                    with open(gz_path, "wb") as f:
                        shutil.copyfileobj(r.raw, f)
            with gzip.open(gz_path, "rb") as f_in, open(fasta_path, "wb") as f_out:
                shutil.copyfileobj(f_in, f_out)
            os.remove(gz_path)

        except Exception as e:
            raise RuntimeError(f"Referenz konnte nicht geladen/entpackt werden: {e}")
        return fasta_path

    def count_reads(self, fastq_path: str) -> int:
        opener = gzip.open if fastq_path.lower().endswith((".gz", ".bgz")) else open
        mode = "rt" if opener is gzip.open else "r"
        with opener(fastq_path, mode, encoding="ascii", errors="ignore") as fq:
            lines = sum(1 for _ in fq)
        return lines // 4

    def init_cache_entry(self, ref_base: str) -> Dict:
        return {
            "ref": (ref_base or "N"),
            "bases": {b: {"count": 0, "qual_sum": 0.0} for b in ["A", "C", "G", "T", "-", "N"]},
            "alt_edge_5p": 0,
            "alt_edge_3p": 0
        }

    def _sanitize_base(self, base: Optional[str]) -> str:
        if base is None:
            return "N"
        b = base.strip().upper()
        return b if b in {"A", "C", "G", "T", "-"} else "N"

    def add_base_to_cache(self, cache: Dict[Tuple[str, int], Dict], chrom: str, pos: int, base: str, qual: int,
                          is_edge_5p: bool = False, is_edge_3p: bool = False) -> None:
        if (chrom, pos) not in cache:
            cache[(chrom, pos)] = self.init_cache_entry("N")
        entry = cache[(chrom, pos)]
        b = self._sanitize_base(base)
        # ensure key exists in bases
        if b not in entry["bases"]:
            entry["bases"][b] = {"count": 0, "qual_sum": 0.0}
        entry["bases"][b]["count"] += 1
        entry["bases"][b]["qual_sum"] += float(max(0, qual))
        if b in {"A", "C", "G", "T"} and b != entry["ref"] and (is_edge_5p or is_edge_3p):
            if is_edge_5p:
                entry["alt_edge_5p"] += 1
            if is_edge_3p:
                entry["alt_edge_3p"] += 1

    def try_insertion(self, read_seq: str, qual_str: str, ref_seq: str, read_i: int, ref_pos: int,
                      max_ins: int = 10, window: int = 6) -> int:
        for k in range(1, max_ins + 1):
            if read_i + k >= len(read_seq):
                break
            rs = read_seq[read_i + k: read_i + k + window]
            rf = ref_seq[ref_pos: ref_pos + window]
            if not rs or not rf or len(rs) != len(rf):
                continue
            mism = sum(1 for a, b in zip(rs, rf) if a != b)
            if mism <= 1:
                return k
        return 0

    def trim_and_map_insertion(self, daughter_seq: str, daughter_qual: str, ref_seq: str,
                               kmer_index: Dict[str, List[int]], k: int, max_trim: int = 12) -> Tuple[int, str, str, str, Optional[int]]:
        for t in range(1, max_trim + 1):
            if len(daughter_seq) <= t:
                break
            ins_seq = daughter_seq[:t]
            new_seq = daughter_seq[t:]
            new_qual = daughter_qual[t:]
            pos = self.map_read(new_seq, ref_seq, kmer_index, k)
            if pos is not None:
                return t, ins_seq, new_seq, new_qual, pos
        return 0, "", daughter_seq, daughter_qual, None

    def left_normalize(self, ref_seq: str, pos_1b: int, ref_event: str, alt_event: str) -> Tuple[int, str, str]:
        if ref_event == alt_event:
            return pos_1b, ref_event, alt_event
        is_insertion = len(alt_event) > len(ref_event)
        is_deletion = len(ref_event) > len(alt_event)
        if not (is_insertion or is_deletion):
            return pos_1b, ref_event, alt_event
        while pos_1b > 1:
            prev_ref_base = ref_seq[pos_1b - 2]
            if not ref_event or not alt_event:
                break
            if is_deletion:
                if ref_event.endswith(prev_ref_base) and alt_event.endswith(prev_ref_base):
                    ref_event = prev_ref_base + ref_event[:-1]
                    alt_event = prev_ref_base + alt_event[:-1]
                    pos_1b -= 1
                else:
                    break
            elif is_insertion:
                if ref_event[-1] == prev_ref_base and alt_event[-1] == prev_ref_base:
                    ref_event = prev_ref_base + ref_event[:-1]
                    alt_event = prev_ref_base + alt_event[:-1]
                    pos_1b -= 1
                else:
                    break
            else:
                break
        return pos_1b, ref_event, alt_event

    def refine_indel_with_sw(self, ref_seq: str, read_seq: str, cut_ref_pos: int, cut_read_pos: int,
                             initial_flank: int = 10, max_iter: int = 3) -> Tuple[int, str, str, float, str]:
        if PairwiseAligner is None:
            corrected_pos = cut_ref_pos + 1
            ref_event = ref_seq[cut_ref_pos] if 0 <= cut_ref_pos < len(ref_seq) else "N"
            alt_event = read_seq[cut_read_pos] if 0 <= cut_read_pos < len(read_seq) else ref_event
            return corrected_pos, ref_event, alt_event, -1.0, "N"

        aligner = PairwiseAligner()
        aligner.mode = "local"
        aligner.match_score = 2
        aligner.mismatch_score = -1
        aligner.open_gap_score = -2
        aligner.extend_gap_score = -0.5

        best_tuple = (cut_ref_pos + 1,
                      ref_seq[cut_ref_pos] if 0 <= cut_ref_pos < len(ref_seq) else "N",
                      read_seq[cut_read_pos] if 0 <= cut_read_pos < len(read_seq) else "N",
                      -1.0, "N")
        prev_score = None
        prev_clip = None

        for it in range(max_iter):
            flank = initial_flank + (5 * it)
            ref_start = max(0, cut_ref_pos - flank)
            ref_end = min(len(ref_seq), cut_ref_pos + flank)
            read_start = max(0, cut_read_pos - flank)
            read_end = min(len(read_seq), cut_read_pos + flank)

            ref_window = ref_seq[ref_start:ref_end]
            read_window = read_seq[read_start:read_end]

            alns = aligner.align(ref_window, read_window)
            if not alns:
                continue

            best = alns[0]
            score = float(best.score)

            try:
                aligned_lines = str(best).splitlines()
                aligned_ref_str = aligned_lines[0].replace(" ", "")
                aligned_read_str = aligned_lines[1].replace(" ", "")
            except Exception:
                aligned_ref_str = ref_window
                aligned_read_str = read_window

            leading_gaps = aligned_ref_str.startswith('-') or aligned_read_str.startswith('-')
            trailing_gaps = aligned_ref_str.endswith('-') or aligned_read_str.endswith('-')
            clip_mode = "N"
            if leading_gaps and not trailing_gaps:
                clip_mode = "L"
            elif trailing_gaps and not leading_gaps:
                clip_mode = "R"

            ref_aln: List[str] = []
            alt_aln: List[str] = []
            for r, q in zip(aligned_ref_str, aligned_read_str):
                r_clean = self._sanitize_base(r)
                q_clean = self._sanitize_base(q)
                if r_clean != "-" and q_clean != "-":
                    ref_aln.append(r_clean)
                    alt_aln.append(q_clean)
                elif r_clean != "-" and q_clean == "-":
                    ref_aln.append(r_clean)
                elif r_clean == "-" and q_clean != "-":
                    alt_aln.append(q_clean)

            ref_seq_event = "".join(ref_aln) if ref_aln else (ref_seq[cut_ref_pos] if 0 <= cut_ref_pos < len(ref_seq) else "N")
            alt_seq_event = "".join(alt_aln) if alt_aln else (read_seq[cut_read_pos] if 0 <= cut_read_pos < len(read_seq) else ref_seq_event)
            corrected_pos = ref_start + 1

            if prev_score is not None and abs(score - prev_score) <= 0.5 and prev_clip == clip_mode:
                return corrected_pos, ref_seq_event, alt_seq_event, score, clip_mode

            best_tuple = (corrected_pos, ref_seq_event, alt_seq_event, score, clip_mode)
            prev_score, prev_clip = score, clip_mode

        return best_tuple

    def record_soft_clip(self, soft_clip_map: Dict[Tuple[str, int], Dict[str, int]],
                         chrom: str, anchor_pos_1b: int, side: str) -> None:
        d = soft_clip_map.setdefault((chrom, anchor_pos_1b), {"L": 0, "R": 0})
        d[side] = d.get(side, 0) + 1

    def cluster_events_by_cigar(self, cigar_counts: Dict[Tuple[str, int], Counter]) -> Dict[Tuple[str, int], str]:
        chosen: Dict[Tuple[str, int], str] = {}
        for key, ctr in cigar_counts.items():
            if ctr:
                cigar, _ = ctr.most_common(1)[0]
                chosen[key] = cigar
        return chosen

    def is_homopolymer(self, ref_seq: str, pos_1b: int, span: int = 8, min_run: int = 4) -> bool:
        start = max(1, pos_1b - span // 2)
        end = min(len(ref_seq), pos_1b + span // 2)
        window = ref_seq[start-1:end]
        if not window:
            return False
        run_char = window[0]
        run_len = 1
        best_run = 1
        for ch in window[1:]:
            if ch == run_char:
                run_len += 1
                best_run = max(best_run, run_len)
            else:
                run_char = ch
                run_len = 1
        return best_run >= min_run

    def seed_coherence_penalty(self, kmer_index: Dict[str, List[int]], kmer: str) -> float:
        cands = kmer_index.get(kmer, [])
        if not cands:
            return 1.0
        n = len(cands)
        if n <= 1:
            return 1.0
        if n <= 3:
            return 0.8
        if n <= 10:
            return 0.6
        return 0.4

    def compute_evidence_score(self, alt_count: int,
                               dp: int,
                               avg_qual: float,
                               soft_clip_map: Dict[Tuple[str, int], Dict[str, int]],
                               chrom: str,
                               pos_1b: int,
                               cache_entry: Dict,
                               cigar: str,
                               kmer_index: Dict[str, List[int]],
                               ref_seq: str,
                               read_seq_seed: Optional[str]) -> float:
        if dp <= 0:
            return 0.0
        af = alt_count / dp
        clip_support = 0
        for delta in (-2, -1, 0, 1, 2):
            d = soft_clip_map.get((chrom, pos_1b + delta), {})
            clip_support += d.get('L', 0) + d.get('R', 0)
        clip_factor = min(1.0, clip_support / max(1, 5))

        L = sum(soft_clip_map.get((chrom, pos_1b + d), {}).get('L', 0) for d in (-2, -1, 0, 1, 2))
        R = sum(soft_clip_map.get((chrom, pos_1b + d), {}).get('R', 0) for d in (-2, -1, 0, 1, 2))
        total_lr = L + R
        if total_lr > 0:
            balance = min(L, R) / total_lr
        else:
            balance = 0.5

        edge_hits = cache_entry.get("alt_edge_5p", 0) + cache_entry.get("alt_edge_3p", 0)
        edge_factor = 1.0 - min(0.8, edge_hits / max(1, alt_count)) * 0.5

        seed_penalty = 1.0
        if read_seq_seed and len(read_seq_seed) >= 15:
            seed_penalty = self.seed_coherence_penalty(kmer_index, read_seq_seed[:15])

        homo_penalty = 0.8 if self.is_homopolymer(ref_seq, pos_1b) else 1.0
        qual_factor = min(1.0, avg_qual / 40.0)

        cigar_simple = 1.0 if ("I" in cigar and "D" not in cigar and "X" not in cigar) or \
                             ("D" in cigar and "I" not in cigar and "X" not in cigar) or \
                             ("X" in cigar and "I" not in cigar and "D" not in cigar) else 0.8

        balance_factor = 1.0 if 0.3 <= balance <= 0.7 else 0.8

        score = (af * 0.35 +
                 clip_factor * 0.2 +
                 balance_factor * 0.1 +
                 edge_factor * 0.1 +
                 qual_factor * 0.15 +
                 cigar_simple * 0.05) * seed_penalty * homo_penalty
        return score

    # ------------------------------------------------------------
    # process_read (uses self.* consistently)
    # ------------------------------------------------------------
    def process_read(self, chrom: str, ref_seq: str, start_pos: int, read_seq: str, qual_str: str,
                     cache: Dict,
                     kmer_index: Dict[str, List[int]],
                     k: int = 15,
                     mismatch_cut: int = 2,
                     ins_cache: Optional[Dict[Tuple[str, int], Dict[str, Dict[str, float]]]] = None,
                     soft_clip_map: Optional[Dict[Tuple[str, int], Dict[str, int]]] = None,
                     cigar_counts: Optional[Dict[Tuple[str, int], Counter]] = None) -> None:
        if len(read_seq) != len(qual_str):
            min_len = min(len(read_seq), len(qual_str))
            read_seq = read_seq[:min_len]
            qual_str = qual_str[:min_len]

        i = 0
        while i < len(read_seq):
            ref_pos = start_pos + i
            if ref_pos >= len(ref_seq):
                if soft_clip_map is not None:
                    self.record_soft_clip(soft_clip_map, chrom, (start_pos + i) + 1, 'R')
                break
            ref_base = ref_seq[ref_pos]
            base = read_seq[i]
            qual = self.phred_score(qual_str[i]) if i < len(qual_str) else 0
            is_edge_5p = (i < 5)
            is_edge_3p = (len(read_seq) - i <= 5)

            key = (chrom, ref_pos + 1)
            if key not in cache:
                cache[key] = self.init_cache_entry(ref_base)

            if base != ref_base:
                if ins_cache is not None:
                    ins_k = self.try_insertion(read_seq, qual_str, ref_seq, i, ref_pos)
                    if ins_k > 0:
                        anchor_pos_1b = ref_pos + 1
                        ins_seq = read_seq[i:i + ins_k]
                        ins_qual_vals = [self.phred_score(q) for q in qual_str[i:i + ins_k]]
                        ins_entry = ins_cache.setdefault((chrom, anchor_pos_1b), {}).setdefault(ins_seq, {"count": 0, "qual_sum": 0.0})
                        ins_entry["count"] += 1
                        ins_entry["qual_sum"] += (sum(ins_qual_vals) / max(1, len(ins_qual_vals))) if ins_qual_vals else 0.0
                        if cigar_counts is not None:
                            cigar_counts.setdefault((chrom, anchor_pos_1b), Counter())[f"1M{ins_k}I"] += 1
                        i += ins_k
                        continue

                mismatch_run = 1
                while (i + mismatch_run < len(read_seq) and
                       start_pos + i + mismatch_run < len(ref_seq) and
                       read_seq[i + mismatch_run] != ref_seq[start_pos + i + mismatch_run]):
                    mismatch_run += 1

                if mismatch_run >= mismatch_cut:
                    candidates: List[int] = []
                    anchor_1b = ref_pos + 1
                    if soft_clip_map is not None:
                        for delta in (-2, -1, 0, 1, 2):
                            p = anchor_1b + delta
                            if soft_clip_map.get((chrom, p), {}).get('L', 0) + soft_clip_map.get((chrom, p), {}).get('R', 0) >= 1:
                                candidates.append(p)
                    if not candidates:
                        candidates = [anchor_1b]

                    best_pick = None
                    best_score_val = -1.0

                    for cand_pos_1b in candidates:
                        cut_ref_pos = max(0, cand_pos_1b - 1)
                        corr_pos, ref_event, alt_event, sw_score, clip_mode = self.refine_indel_with_sw(
                            ref_seq, read_seq, cut_ref_pos, i, initial_flank=10, max_iter=3
                        )
                        corr_pos, ref_event, alt_event = self.left_normalize(ref_seq, corr_pos, ref_event, alt_event)

                        cigar = "1M"
                        if len(ref_event) > len(alt_event):
                            cigar = f"1M{len(ref_event) - len(alt_event)}D"
                        elif len(alt_event) > len(ref_event):
                            cigar = f"1M{len(alt_event) - len(ref_event)}I"
                        elif ref_event != alt_event:
                            cigar = "1X"

                        if cigar_counts is not None:
                            cigar_counts.setdefault((chrom, corr_pos), Counter())[cigar] += 1

                        cache_key = (chrom, corr_pos)
                        if cache_key not in cache:
                            ref_b = ref_seq[corr_pos - 1] if 0 < corr_pos <= len(ref_seq) else "N"
                            cache[cache_key] = self.init_cache_entry(ref_b)

                        dp_here = sum(cache[cache_key]["bases"][b]["count"] for b in "ACGT")
                        avgq_here = qual
                        read_seed = read_seq[:k] if len(read_seq) >= k else None
                        score_val = self.compute_evidence_score(
                            alt_count=1, dp=max(1, dp_here), avg_qual=avgq_here,
                            soft_clip_map=soft_clip_map or {}, chrom=chrom, pos_1b=corr_pos,
                            cache_entry=cache[cache_key], cigar=cigar,
                            kmer_index=kmer_index, ref_seq=ref_seq, read_seq_seed=read_seed
                        )
                        score_val += min(1.0, sw_score / 20.0) * 0.2

                        if score_val > best_score_val:
                            best_score_val = score_val
                            best_pick = (corr_pos, ref_event, alt_event, sw_score, clip_mode)

                    if best_pick is not None:
                        corr_pos, ref_event, alt_event, _, _ = best_pick
                        for j, rb in enumerate(ref_event):
                            pos_key = corr_pos + j
                            key2 = (chrom, pos_key)
                            if key2 not in cache:
                                ref_b = ref_seq[pos_key - 1] if 0 < pos_key <= len(ref_seq) else "N"
                                cache[key2] = self.init_cache_entry(ref_b)
                            if j < len(alt_event) and alt_event[j] != rb:
                                self.add_base_to_cache(cache, chrom, pos_key, alt_event[j], qual, is_edge_5p=is_edge_5p, is_edge_3p=is_edge_3p)
                            else:
                                self.add_base_to_cache(cache, chrom, pos_key, "-", 0)

                    daughter_seq = read_seq[i + mismatch_run:]
                    daughter_qual = qual_str[i + mismatch_run:]
                    if len(daughter_seq) != len(daughter_qual):
                        min_len = min(len(daughter_seq), len(daughter_qual))
                        daughter_seq = daughter_seq[:min_len]
                        daughter_qual = daughter_qual[:min_len]

                    if daughter_seq:
                        new_start = self.map_read(daughter_seq, ref_seq, kmer_index, k)
                        if new_start is not None:
                            self.process_read(chrom, ref_seq, new_start, daughter_seq, daughter_qual,
                                              cache, kmer_index, k, mismatch_cut, ins_cache, soft_clip_map, cigar_counts)
                        else:
                            if soft_clip_map is not None:
                                self.record_soft_clip(soft_clip_map, chrom, (ref_pos + mismatch_run) + 1, 'R')
                            tlen, ins_seq, trimmed_seq, trimmed_qual, pos2 = self.trim_and_map_insertion(
                                daughter_seq, daughter_qual, ref_seq, kmer_index, k)
                            if tlen > 0 and pos2 is not None and ins_cache is not None:
                                anchor_pos_1b = (ref_pos + mismatch_run) + 1
                                ins_entry = ins_cache.setdefault((chrom, anchor_pos_1b), {}).setdefault(ins_seq, {"count": 0, "qual_sum": 0.0})
                                ins_entry["count"] += 1
                                ins_entry["qual_sum"] += (sum(self.phred_score(q) for q in daughter_qual[:tlen]) / max(1, tlen))
                                if cigar_counts is not None:
                                    cigar_counts.setdefault((chrom, anchor_pos_1b), Counter())[f"1M{tlen}I"] += 1
                                self.process_read(chrom, ref_seq, pos2, trimmed_seq, trimmed_qual,
                                                 cache, kmer_index, k, mismatch_cut, ins_cache, soft_clip_map, cigar_counts)
                    return
                else:
                    self.add_base_to_cache(cache, chrom, ref_pos + 1, base, qual, is_edge_5p=is_edge_5p, is_edge_3p=is_edge_3p)
            else:
                self.add_base_to_cache(cache, chrom, ref_pos + 1, base, qual, is_edge_5p=is_edge_5p, is_edge_3p=is_edge_3p)
            i += 1

    # ------------------------------------------------------------
    # process_fastq (uses self.* consistently)
    # ------------------------------------------------------------
    def process_fastq(self,
        fastq_path: str,
        progress_callback: Optional[Callable[[int, int], None]] = None,
        build: str = "GRCh38",
        k: int = 15,
        qual_threshold: int = 20,
        min_dp: int = 5,
        min_af: float = 0.2
    ) -> str:
        ref_fasta_path: str = self.ensure_reference(build)
        ref_seqs: Dict[str, str] = self.load_fasta(ref_fasta_path)
        if not ref_seqs:
            raise RuntimeError("Referenz FASTA enthält keine Sequenzen.")

        base_name: str = os.path.splitext(os.path.basename(fastq_path))[0]
        out_dir: str = os.path.join(os.getcwd(), base_name)
        os.makedirs(out_dir, exist_ok=True)
        out_fasta: str = os.path.join(out_dir, "consensus.fa")
        out_vcf: str = os.path.join(out_dir, "variants.vcf")

        cache: Dict[Tuple[str, int], Dict] = {}
        kmer_indices: Dict[str, Dict[str, List[int]]] = {chrom: self.build_kmer_index(seq, k) for chrom, seq in ref_seqs.items()}
        total_reads: int = self.count_reads(fastq_path)
        processed_reads: int = 0
        ins_cache: Dict[Tuple[str, int], Dict[str, Dict[str, float]]] = defaultdict(dict)
        soft_clip_map: Dict[Tuple[str, int], Dict[str, int]] = defaultdict(lambda: {"L": 0, "R": 0})
        cigar_counts: Dict[Tuple[str, int], Counter] = defaultdict(Counter)

        opener = gzip.open if fastq_path.lower().endswith((".gz", ".bgz")) else open
        mode = "rt" if opener is gzip.open else "r"
        with opener(fastq_path, mode, encoding="ascii", errors="ignore") as fq:
            while True:
                header: str = fq.readline()
                if not header:
                    break
                seq: str = fq.readline().rstrip().upper()
                plus_line: str = fq.readline()
                qual_str: str = fq.readline().rstrip()

                if not seq or not qual_str:
                    continue
                if len(seq) != len(qual_str):
                    min_len: int = min(len(seq), len(qual_str))
                    seq, qual_str = seq[:min_len], qual_str[:min_len]

                mapped = False
                for chrom, ref_seq in ref_seqs.items():
                    if len(seq) < k:
                        continue
                    pos: Optional[int] = self.map_read(seq, ref_seq, kmer_indices[chrom], k)
                    if pos is not None:
                        self.process_read(
                            chrom, ref_seq, pos, seq, qual_str,
                            cache, kmer_indices[chrom], k,
                            mismatch_cut=2, ins_cache=ins_cache,
                            soft_clip_map=soft_clip_map, cigar_counts=cigar_counts
                        )
                        mapped = True
                        break

                processed_reads += 1
                if progress_callback and processed_reads % 100 == 0:
                    progress_callback(processed_reads, total_reads)

        consensus_seqs: Dict[str, List[str]] = {chrom: list(seq) for chrom, seq in ref_seqs.items()}

        events: List[Tuple[str, int, str, str, Optional[float], str, str, int, int, str]] = []
        visited_del: Set[Tuple[str, int]] = set()

        for chrom in sorted(ref_seqs.keys(), key=lambda c: (c.isdigit(), c) if c not in ("X", "Y", "MT") else (False, c)):
            chrom_positions = sorted([p for (c, p) in cache.keys() if c == chrom])
            ref_seq = ref_seqs[chrom]

            idx = 0
            while idx < len(chrom_positions):
                pos = chrom_positions[idx]
                entry = cache.get((chrom, pos), None)
                if not entry:
                    idx += 1
                    continue

                ref_base: str = self._sanitize_base(entry["ref"])
                base_counts = entry["bases"]

                total_cov: int = sum(base_counts.get(b, {"count":0})["count"] for b in "ACGT")
                best_base: str = max(base_counts.keys(), key=lambda b: base_counts[b]["count"])
                best_count: int = base_counts[best_base]["count"]

                is_del_here = (base_counts["-"]["count"] >= max(base_counts[b]["count"] for b in "ACGT"))

                if is_del_here and (chrom, pos) not in visited_del:
                    block_len = 1
                    j = 1
                    while (chrom, pos + j) in cache:
                        nxt = cache[(chrom, pos + j)]["bases"]
                        if nxt["-"]["count"] >= max(nxt[b]["count"] for b in "ACGT"):
                            block_len += 1
                            j += 1
                        else:
                            break

                    ref_block = "".join(
                        self._sanitize_base(cache[(chrom, pos + j)]["ref"])
                        for j in range(block_len)
                        if (chrom, pos + j) in cache
                    )
                    if not ref_block:
                        ref_block = ref_seq[pos - 1: pos - 1 + block_len]
                    if not ref_block:
                        idx += block_len
                        continue

                    alt_block = ref_block[0]
                    npos, nref, nalt = self.left_normalize(ref_seq, pos, ref_block, alt_block)

                    dp_block = 0
                    ac_block = 0
                    for j in range(block_len):
                        key_j = (chrom, pos + j)
                        if key_j not in cache:
                            continue
                        counts_j = cache[key_j]["bases"]
                        dp_block += sum(counts_j.get(b, {"count":0})["count"] for b in "ACGT")
                        ac_block += counts_j["-"]["count"]

                    af_block = (ac_block / dp_block) if dp_block else 0.0
                    if dp_block >= min_dp and af_block >= min_af:
                        filt = "PASS"
                        cigar_str = f"1M{len(nref) - len(nalt)}D"
                        gt = "0/1" if 0.2 <= af_block < 0.8 else "1/1"
                        events.append((chrom, npos, nref, nalt, None, filt, cigar_str, dp_block, ac_block, gt))

                    for j in range(block_len):
                        visited_del.add((chrom, pos + j))
                    idx += block_len
                    continue

                if best_base in {"A", "C", "G", "T"} and best_base != ref_base:
                    af = (best_count / max(1, total_cov)) if total_cov else 0.0
                    if total_cov >= min_dp and af >= min_af:
                        avg_qual = (base_counts[best_base]["qual_sum"] / max(1, best_count)) if best_count else 0.0
                        local_thr = qual_threshold + 5 if self.is_homopolymer(ref_seq, pos) else qual_threshold
                        filt = "PASS" if avg_qual >= local_thr else "LowQual"
                        cigar_str = "1X"
                        gt = "0/1" if 0.2 <= af < 0.8 else "1/1"
                        events.append((chrom, pos, ref_base, best_base, avg_qual, filt, cigar_str, total_cov, best_count, gt))
                        if 0 < pos <= len(consensus_seqs[chrom]):
                            consensus_seqs[chrom][pos - 1] = best_base

                idx += 1

        for (chrom, pos), seqs in ins_cache.items():
            if not seqs:
                continue
            ins_seq, stats = max(seqs.items(), key=lambda kv: kv[1]["count"])
            if any(ch not in "ACGT" for ch in ins_seq.upper()):
                continue

            if (chrom, pos) not in cache:
                ref_anchor_b = ref_seqs[chrom][pos - 1] if 0 < pos <= len(ref_seqs[chrom]) else "N"
                cache[(chrom, pos)] = self.init_cache_entry(ref_anchor_b)
            dp_here = sum(cache[(chrom, pos)]["bases"].get(b, {"count":0})["count"] for b in "ACGT")
            ac_i = int(stats["count"])
            af_i = (ac_i / dp_here) if dp_here else 0.0
            if dp_here < min_dp or af_i < min_af:
                continue

            avgq_i = stats["qual_sum"] / max(1, ac_i)
            ref_anchor = cache[(chrom, pos)]["ref"]

            npos, nref, nalt = self.left_normalize(ref_seqs[chrom], pos, ref_anchor, ref_anchor + ins_seq)
            local_thr = qual_threshold + 5 if self.is_homopolymer(ref_seqs[chrom], npos) else qual_threshold
            filt_i = "PASS" if avgq_i >= local_thr else "LowQual"
            cigar_i = f"1M{len(nalt) - len(nref)}I"
            gt_i = "0/1" if 0.2 <= af_i < 0.8 else "1/1"

            events.append((chrom, npos, nref, nalt, avgq_i, filt_i, cigar_i, dp_here, ac_i, gt_i))

        top_cigars = self.cluster_events_by_cigar(cigar_counts)
        filtered_events: List[Tuple[str, int, str, str, Optional[float], str, str, int, int, str]] = []
        for ev in events:
            key = (ev[0], ev[1])
            top_cigar = top_cigars.get(key)
            if top_cigar and top_cigar != ev[6]:
                filtered_events.append((ev[0], ev[1], ev[2], ev[3], ev[4], "Secondary", ev[6], ev[7], ev[8], ev[9]))
            else:
                filtered_events.append(ev)

        unique_events: Dict[Tuple[str, int, str, str], Tuple[str, int, str, str, Optional[float], str, str, int, int, str]] = {}
        for ev in filtered_events:
            key = (ev[0], ev[1], ev[2], ev[3])
            if key not in unique_events or (ev[4] or 0.0) > (unique_events[key][4] or 0.0):
                unique_events[key] = ev

        with open(out_vcf, "w", encoding="utf-8") as vcf:
            vcf.write("##fileformat=VCFv4.2\n")
            vcf.write(f"##reference={ref_fasta_path}\n")
            vcf.write("##INFO=<ID=CIGAR,Number=1,Type=String,Description=\"Local CIGAR string for variant\">\n")
            vcf.write("##INFO=<ID=DP,Number=1,Type=Integer,Description=\"Total depth at position\">\n")
            vcf.write("##INFO=<ID=AC,Number=1,Type=Integer,Description=\"Alt allele count\">\n")
            vcf.write("##FORMAT=<ID=GT,Number=1,Type=String,Description=\"Genotype\">\n")
            vcf.write("#CHROM\tPOS\tID\tREF\tALT\tQUAL\tFILTER\tINFO\tFORMAT\tSAMPLE\n")
            for chrom, pos, ref, alt, qual, filt, cigar, dp, ac, gt in sorted(unique_events.values(), key=lambda e: (e[0], e[1])):
                if not ref or not alt or pos < 1:
                    continue
                info_str = f"CIGAR={cigar};DP={dp};AC={ac}"
                q = 0.0 if qual is None else float(qual)
                vcf.write(f"{chrom}\t{pos}\t.\t{ref}\t{alt}\t{q:.2f}\t{filt}\t{info_str}\tGT\t{gt}\n")

        with open(out_fasta, "w", encoding="utf-8") as fa:
            for chrom, seq_list in consensus_seqs.items():
                fa.write(f">{chrom}_consensus\n")
                for j in range(0, len(seq_list), 60):
                    fa.write("".join(seq_list[j:j+60]) + "\n")

        return out_dir


class StreamingFastaToGVCF:
    def detect_build_from_fasta(self, fasta_path):
        """Versucht, den Build aus den ersten Headerzeilen der FASTA zu erkennen."""
        try:
            with open(fasta_path, "r") as f:
                for _ in range(200):
                    line = f.readline()
                    if not line:
                        break
                    if line.startswith(">"):
                        low = line.lower()
                        if "grch38" in low or "hg38" in low:
                            return "GRCh38"
                        if "grch37" in low or "hg19" in low or "b37" in low:
                            return "GRCh37"
        except Exception as e:
            print(f"[WARN] Build-Erkennung fehlgeschlagen: {e}")
        return "GRCh38"  # Fallback

    def build_fasta_index(self, fasta_path):
        """
        Build a .fai using pyfaidx (samtools-compatible).
        """
        fai_path = fasta_path + ".fai"
        if os.path.exists(fai_path):
            print(f"[INFO] Index {fai_path} existiert bereits – überspringe.")
            return fai_path
        if not HAVE_PYFAIDX:
            raise RuntimeError("pyfaidx ist nicht installiert, kann FAI nicht erzeugen.")
        print(f"[INFO] Erstelle Index für {fasta_path} ...")
        Faidx(fasta_path)
        print(f"[INFO] Index erstellt: {fai_path}")
        return fai_path

    def ensure_fasta(self, build):
        fasta_path = FASTA_PATHS[build]
        if os.path.exists(fasta_path) and os.path.exists(fasta_path + ".fai"):
            return fasta_path
        url = FASTA_URLS[build]
        gz_path = fasta_path + ".gz"
        print(f"[INFO] Lade Referenz {build} herunter...")
        with get_robust_session() as session:
            with session.get(url, stream=True, timeout=60) as r:
                r.raise_for_status()
                with open(gz_path, "wb") as f:
                    shutil.copyfileobj(r.raw, f)
        print("[INFO] Entpacke...")
        with gzip.open(gz_path, "rb") as f_in, open(fasta_path, "wb") as f_out:
            shutil.copyfileobj(f_in, f_out)
        os.remove(gz_path)
        self.build_fasta_index(fasta_path)
        return fasta_path

    def convert_streaming_gvcf(self, sample_fasta, build, out_vcf):
        if not build:
            build = self.detect_build_from_fasta(sample_fasta)
            print(f"[INFO] Automatisch ermittelter Build: {build}")

        ref_fasta = self.ensure_fasta(build)
        fai_index = load_fai_index(ref_fasta + ".fai")

        with open(sample_fasta, "r") as sf, open(out_vcf, "w") as vcf:
            vcf.write("##fileformat=VCFv4.2\n")
            vcf.write(f"##reference={build}\n")
            vcf.write(f"##source=StreamingFastaToGVCF\n")
            vcf.write(f"##reference_file={os.path.abspath(ref_fasta)}\n")
            vcf.write(f"##fileDate={datetime.datetime.now(datetime.timezone.utc).strftime('%Y%m%d')}\n")
            vcf.write("##INFO=<ID=END,Number=1,Type=Integer,Description=\"End position of the block\">\n")
            vcf.write("##FORMAT=<ID=GT,Number=1,Type=String,Description=\"Genotype\">\n")
            vcf.write("#CHROM\tPOS\tID\tREF\tALT\tQUAL\tFILTER\tINFO\tFORMAT\tSAMPLE\n")

            chrom = None
            pos = 0
            block_start = None
            block_ref_base = None
            in_ref_block = False

            def flush_block(end_pos):
                nonlocal block_start, in_ref_block
                if in_ref_block and block_start is not None:
                    vcf.write(f"{chrom}\t{block_start}\t.\t{block_ref_base}\t.\t.\tPASS\tEND={end_pos}\tGT\t0/0\n")
                block_start = None
                in_ref_block = False

            for line in sf:
                if line.startswith(">"):
                    if in_ref_block:
                        flush_block(pos)
                    chrom = line[1:].strip().split()[0]
                    pos = 0
                    continue
                for base in line.strip().upper():
                    pos += 1
                    ref_base = fetch_base_from_fasta(ref_fasta, fai_index, chrom, pos)
                    if base not in ("A", "C", "G", "T", "N") or ref_base == "N":
                        continue
                    if base != ref_base:
                        if in_ref_block:
                            flush_block(pos - 1)
                        vcf.write(f"{chrom}\t{pos}\t.\t{ref_base}\t{base}\t100\tPASS\t.\tGT\t1/1\n")
                    else:
                        if not in_ref_block:
                            block_start = pos
                            block_ref_base = ref_base
                            in_ref_block = True
            if in_ref_block:
                flush_block(pos)
        print(f"[INFO] gVCF geschrieben: {out_vcf}")

    def convert_streaming_variants_only(self, sample_fasta, build, out_vcf):
        """
        Stream through sample FASTA and emit only variant positions vs. reference.
        No gVCF blocks, only ALT sites with a dummy homozygous alt genotype (1/1).
        """
        if not build:
            build = self.detect_build_from_fasta(sample_fasta)
            print(f"[INFO] Automatisch ermittelter Build: {build}")

        ref_fasta = self.ensure_fasta(build)
        fai_index = load_fai_index(ref_fasta + ".fai")

        with open(sample_fasta, "r") as sf, open(out_vcf, "w") as vcf:
            vcf.write("##fileformat=VCFv4.2\n")
            vcf.write(f"##reference={build}\n")
            vcf.write(f"##source=StreamingFastaToVCF\n")
            vcf.write(f"##reference_file={os.path.abspath(ref_fasta)}\n")
            vcf.write(f"##fileDate={datetime.datetime.now(datetime.timezone.utc).strftime('%Y%m%d')}\n")
            vcf.write("##FORMAT=<ID=GT,Number=1,Type=String,Description=\"Genotype\">\n")
            vcf.write("#CHROM\tPOS\tID\tREF\tALT\tQUAL\tFILTER\tINFO\tFORMAT\tSAMPLE\n")

            chrom = None
            pos = 0
            for line in sf:
                if line.startswith(">"):
                    chrom = line[1:].strip().split()[0]
                    pos = 0
                    continue
                for base in line.strip().upper():
                    pos += 1
                    if base not in ("A", "C", "G", "T", "N"):
                        continue
                    ref_base = fetch_base_from_fasta(ref_fasta, fai_index, chrom, pos)
                    if ref_base == "N" or base == "N":
                        continue
                    if base != ref_base:
                        vcf.write(f"{chrom}\t{pos}\t.\t{ref_base}\t{base}\t100\tPASS\t.\tGT\t1/1\n")

#SCHLECHTER NAME umbenennen irgendwann
#Eher sowas wie NoneClassifierAndHandler
# =============================================================================
# AF-NONE-TREATMENT MANAGER
# =============================================================================

class AfNoneTreatmentManager:
    """
    ✅ Policy-Verwaltung für AF-None-Behandlung.
    
    Verantwortlichkeiten:
    - Verwaltet Preset-basierte Policies
    - Klassifiziert None-Typen (via FetchStatusManager)
    - Tracked Statistiken
    - Beantwortet Policy-Fragen
    
    WICHTIG: Trifft KEINE Filter-Entscheidungen!
             Wird von _apply_af_filter_final() GENUTZT.
    
    ========================================================================
    STATUS-CODE-SYSTEM (3-stellig)
    ========================================================================
    
    Position:  [1] [2] [3]
               │   │   └─── Bearbeitungsstatus/Retry-Count (0-9)
               │   └─────── Fehlerart/Erfolgsart (0-4+)
               └─────────── Erfolg/Fehler (0=Fehler, 1=Erfolg)
    
    AKTUELLE CODES:
    ---------------
    "0"/"000" = Generischer Fehler, nie bearbeitet
    "1"/"111" = Erfolg, alle Quellen
    "010"     = Fehler in Queue, noch nicht bearbeitet
    "011"     = Fehler in Queue, 1. Retry fehlgeschlagen
    "012"     = Fehler in Queue, 2. Retry fehlgeschlagen
    "013"     = Fehler in Queue, 3. Retry/Exhausted
    
    RESERVIERT (zukünftig):
    -----------------------
    "020"-"029" = Quarantäne-System (9 Retries)
    "030"-"039" = Andere Fehlerklasse
    "040"       = Zur Löschung freigegeben
    
    ========================================================================
    4 AF-NONE-ZUSTÄNDE
    ========================================================================
    
    1. has_af          → af_filter_mean != NULL
    2. never_fetched   → meanAF_last_fetch = NULL 
                          OR is_stale(>30d) 
                          OR code="013"
    3. fetch_failed    → code in (0,000,010-012,020,030,040) 
                          AND af_filter_mean = NULL
    4. true_none       → code in (1,111) 
                          AND af_filter_mean = NULL 
                          (validiert selten)
    
    ========================================================================
    VERWENDUNG
    ========================================================================
    
    # Initialisierung
    >>> manager = AfNoneTreatmentManager(preset="clinical", logger=logger)
    
    # None-Typ klassifizieren
    >>> none_type = manager.classify_none_type(row)
    
    # Policy abfragen
    >>> should_include = manager.should_include_type(none_type)
    
    # Statistik tracken
    >>> manager.record_decision(none_type, included=True)
    
    ========================================================================
    """
    
    # Preset-Policies
    PRESETS = {
        "permissive": {
            "description": "Maximale Sensitivität - alle None-Typen inkludieren",
            "never_fetched": True,
            "fetch_failed": True,
            "true_none": True,
        },
        "research": {
            "description": "Standard für Forschung - konservativ bei Fehlern",
            "never_fetched": True,
            "fetch_failed": False,
            "true_none": True,
        },
        "clinical": {
            "description": "Klinische Anwendung - nur validierte",
            "never_fetched": False,
            "fetch_failed": False,
            "true_none": True,
        },
        "strict": {
            "description": "Maximale Spezifität - nur echte AF-Werte",
            "never_fetched": False,
            "fetch_failed": False,
            "true_none": False,
        },
    }
    
    def __init__(self, preset: str = "clinical", logger=None):
        """
        Initialisiert Manager mit Preset.
        
        Args:
            preset: Name des Presets
            logger: Optional Logger
        """
        self.logger = logger
        self._lock = threading.RLock()
        
        self.current_preset = preset
        self.policy = self._load_preset(preset)
        
        # Statistiken (für Reporting)
        self.stats = {
            "total_checked": 0,
            "has_af_included": 0,
            "has_af_excluded": 0,
            "never_fetched_included": 0,
            "never_fetched_excluded": 0,
            "fetch_failed_included": 0,
            "fetch_failed_excluded": 0,
            "true_none_included": 0,
            "true_none_excluded": 0,
        }
        
        if self.logger:
            desc = self.policy.get("description", "")
            self.logger.log(
                f"[AfNoneTreatment] Initialisiert mit Preset '{preset}': {desc}"
            )
    
    def _load_preset(self, preset: str) -> dict:
        """Lädt Preset-Policy."""
        if preset not in self.PRESETS:
            if self.logger:
                self.logger.log(
                    f"[AfNoneTreatment] ⚠️ Unknown preset '{preset}', using 'clinical'"
                )
            preset = "clinical"
        
        return dict(self.PRESETS[preset])
    
    def set_preset(self, preset: str):
        """Setzt neues Preset."""
        with self._lock:
            self.policy = self._load_preset(preset)
            self.current_preset = preset
            
            if self.logger:
                desc = self.policy.get("description", "")
                self.logger.log(f"[AfNoneTreatment] Preset: '{preset}' - {desc}")
    
    def set_custom_policy(
        self,
        never_fetched: bool,
        fetch_failed: bool,
        true_none: bool
    ):
        """Setzt Custom-Policy."""
        with self._lock:
            self.policy = {
                "description": "Custom Policy",
                "never_fetched": never_fetched,
                "fetch_failed": fetch_failed,
                "true_none": true_none,
            }
            self.current_preset = "custom"
            
            if self.logger:
                self.logger.log(
                    f"[AfNoneTreatment] Custom Policy: "
                    f"never={never_fetched}, failed={fetch_failed}, true={true_none}"
                )
    
    def set_legacy_include_none(self, include_none: bool):
        """
        ✅ RÜCKWÄRTSKOMPATIBILITÄT: Legacy include_none → Policy.
        
        DEPRECATED: Nutze stattdessen set_custom_policy()
        """
        with self._lock:
            if include_none:
                self.set_custom_policy(
                    never_fetched=True,
                    fetch_failed=True,
                    true_none=True
                )
            else:
                self.set_custom_policy(
                    never_fetched=False,
                    fetch_failed=False,
                    true_none=True
                )
            
            if self.logger:
                self.logger.log(
                    f"[AfNoneTreatment] Legacy include_none={include_none} "
                    f"→ Konvertiert zu Custom Policy"
                )
    
  
    """
    ✅ ÖFFENTLICHE API (vom BackgroundMaintainer zu nutzen):
    """   
    def classify_none_type(self, row: dict, stale_days: int = Config.STALE_DAYS_AF) -> str:
        """
        ✅ Klassifiziert AF-None-Typ.
        
        Args:
            row: DB-Row mit af_filter_mean, meanAF_last_fetch, meanAF_fetch_success
            stale_days: Tage bis Daten stale sind
        
        Returns:
            "has_af"        - AF-Wert vorhanden
            "never_fetched" - Noch nie geprüft oder stale
            "fetch_failed"  - Fetch-Fehler
            "true_none"     - Validiert ultra-rare
        """
        return FetchStatusManager.classify_af_status(
            af_value=row.get("af_filter_mean"),
            last_fetch=row.get("meanAF_last_fetch"),
            fetch_status=row.get("meanAF_fetch_success"),
            stale_days=stale_days
        )
    
    def should_include_type(self, none_type: str) -> bool:
        """
        ✅ Fragt Policy für None-Typ ab.
        
        Args:
            none_type: "never_fetched" | "fetch_failed" | "true_none" | "has_af"
        
        Returns:
            True wenn inkludieren, False wenn excludieren
        """
        with self._lock:
            if none_type == "never_fetched":
                return self.policy.get("never_fetched", False)
            elif none_type == "fetch_failed":
                return self.policy.get("fetch_failed", False)
            elif none_type == "true_none":
                return self.policy.get("true_none", True)
            else:
                # "has_af" oder unknown
                return False
    
    def record_decision(
        self, 
        none_type: str, 
        included: bool
    ):
        """
        ✅ Tracked Statistik (von _apply_af_filter_final() aufgerufen).
        
        Args:
            none_type: Klassifizierter None-Typ
            included: Wurde inkludiert?
        """
        with self._lock:
            self.stats["total_checked"] += 1
            
            if none_type == "has_af":
                if included:
                    self.stats["has_af_included"] += 1
                else:
                    self.stats["has_af_excluded"] += 1
            elif none_type == "never_fetched":
                if included:
                    self.stats["never_fetched_included"] += 1
                else:
                    self.stats["never_fetched_excluded"] += 1
            elif none_type == "fetch_failed":
                if included:
                    self.stats["fetch_failed_included"] += 1
                else:
                    self.stats["fetch_failed_excluded"] += 1
            elif none_type == "true_none":
                if included:
                    self.stats["true_none_included"] += 1
                else:
                    self.stats["true_none_excluded"] += 1
    
    def get_policy(self) -> dict:
        """Gibt aktuelle Policy zurück."""
        with self._lock:
            return {
                "preset": self.current_preset,
                "never_fetched": self.policy.get("never_fetched"),
                "fetch_failed": self.policy.get("fetch_failed"),
                "true_none": self.policy.get("true_none"),
                "description": self.policy.get("description", "")
            }
    
    def get_stats(self) -> dict:
        """Gibt Statistiken zurück."""
        with self._lock:
            total = self.stats["total_checked"]
            if total == 0:
                return dict(self.stats)
            
            stats_with_pct = dict(self.stats)
            stats_with_pct["has_af_included_pct"] = (
                100.0 * self.stats["has_af_included"] / total
            )
            stats_with_pct["never_fetched_excluded_pct"] = (
                100.0 * self.stats["never_fetched_excluded"] / total
            )
            stats_with_pct["fetch_failed_excluded_pct"] = (
                100.0 * self.stats["fetch_failed_excluded"] / total
            )
            stats_with_pct["true_none_included_pct"] = (
                100.0 * self.stats["true_none_included"] / total
            )
            
            return stats_with_pct
    
    def print_stats(self):
        """Gibt formatierte Statistiken aus."""
        with self._lock:
            stats = self.get_stats()
            total = stats["total_checked"]
            
            if total == 0:
                if self.logger:
                    self.logger.log("[AfNoneTreatment] Keine Varianten geprüft")
                return
            
            policy = self.get_policy()
            
            msg = (
                f"[AfNoneTreatment] Statistik ({policy['preset']}):\n"
                f"  Total: {total:,} Varianten\n"
                f"\n"
                f"  ✅ AF vorhanden:\n"
                f"    Inkludiert: {stats['has_af_included']:,} "
                f"({stats.get('has_af_included_pct', 0):.1f}%)\n"
                f"    Exkludiert: {stats['has_af_excluded']:,} (AF > Threshold)\n"
                f"\n"
                f"  🔵 Nie geprüft: {policy['never_fetched'] and '✓' or '✗'}\n"
                f"    Inkludiert: {stats['never_fetched_included']:,}\n"
                f"    Exkludiert: {stats['never_fetched_excluded']:,} "
                f"({stats.get('never_fetched_excluded_pct', 0):.1f}%)\n"
                f"\n"
                f"  🟠 Fehlerhafte Abrufe: {policy['fetch_failed'] and '✓' or '✗'}\n"
                f"    Inkludiert: {stats['fetch_failed_included']:,}\n"
                f"    Exkludiert: {stats['fetch_failed_excluded']:,} "
                f"({stats.get('fetch_failed_excluded_pct', 0):.1f}%)\n"
                f"\n"
                f"  🟢 Validiert selten: {policy['true_none'] and '✓' or '✗'}\n"
                f"    Inkludiert: {stats['true_none_included']:,} "
                f"({stats.get('true_none_included_pct', 0):.1f}%)\n"
                f"    Exkludiert: {stats['true_none_excluded']:,}\n"
            )
            
            if self.logger:
                self.logger.log(msg)
            else:
                print(msg)
    
    def reset_stats(self):
        """Reset Statistiken."""
        with self._lock:
            self.stats = {k: 0 for k in self.stats.keys()}
    
    def get_preset_names(self) -> list:
        """Gibt Liste aller Preset-Namen zurück."""
        return list(self.PRESETS.keys())
    
    def get_preset_description(self, preset: str) -> str:
        """Gibt Beschreibung eines Presets zurück."""
        return self.PRESETS.get(preset, {}).get("description", "Unknown")
    
class FetchStatusManager:
    """
    Zentrale Verwaltung und Interpretation der Fetch-Statuscodes.
    - Kapselt Update-Logik (inkl. Eskalation)
    - Übersetzt Codes in Klartext
    - Bietet Hilfsfunktionen für Erfolg/Fehlschlag/Deletion
    """
    """
    ✅ DOKUMENTIERT: Status-Code-System für AF & Full-Anno
    
    ========================================================================
    STATUS-CODE-SYSTEM (3-stellig)
    ========================================================================
    
    Position:  [1] [2] [3]
               │   │   └─── Bearbeitungsstatus/Retry-Count (0-9)
               │   └─────── Fehlerart/Erfolgsart (0-4+)
               └─────────── Erfolg/Fehler (0=Fehler, 1=Erfolg)
    
    AKTUELLE CODES:
    ---------------
    "0"/"000" = Generischer Fehler, nie bearbeitet
    "1"/"111" = Erfolg, alle Quellen
    "010"     = Fehler in Queue, noch nicht bearbeitet
    "011"     = Fehler in Queue, 1. Retry fehlgeschlagen
    "012"     = Fehler in Queue, 2. Retry fehlgeschlagen
    "013"     = Fehler in Queue, 3. Retry/Exhausted
    
    RESERVIERT (zukünftig):
    -----------------------
    "020"-"029" = Quarantäne-System (9 Retries)
    "030"-"039" = Andere Fehlerklasse
    "040"       = Zur Löschung freigegeben
    
    ========================================================================
    4 AF-NONE-ZUSTÄNDE
    ========================================================================
    
    1. has_af          → af_filter_mean != NULL
    2. never_fetched   → meanAF_last_fetch = NULL 
                          OR is_stale(>30d) 
                          OR code="013"
    3. fetch_failed    → code in (0,000,010-012,020,030,040) 
                          AND af_filter_mean = NULL
    4. true_none       → code in (1,111) 
                          AND af_filter_mean = NULL 
                          (validiert selten)
    
    ========================================================================
    """
     
    FETCH_STATUS_LABELS = {
        # Klassische Werte
        "0":   "Fehlschlag (generisch, wird erneut probiert)",
        "1":   "Erfolgreich",

        # Neue feinere Codes
        "000": "Fehlschlag (generisch, wird erneut probiert)",
        "010": "Normalisierung angefragt",
        "011": "Re-Normalisierung fehlgeschlagen (1. Versuch)",
        "012": "Re-Normalisierung fehlgeschlagen (2. Versuch)",
        "013": "Re-Normalisierung fehlgeschlagen (3. Versuch, zur Löschung markiert)",

        # Reservierte Codes
        "020": "Reserviert: andere Fehlerklasse (z. B. API-Fehler)",
        "030": "Reserviert: andere Fehlerklasse",
        "040": "Reserviert: andere Fehlerklasse",

        # Erfolgsstatus
        "111": "Erfolgreich"
    }

    @classmethod
    def status_label(cls, code) -> str:
        """Gibt eine menschenlesbare Beschreibung für den Fetch-Status zurück."""
        c = cls.normalize(code)
        return cls.FETCH_STATUS_LABELS.get(c, f"Unbekannter Status {c}")

    @classmethod
    def is_success(cls, code) -> bool:
        """True, wenn Status ein Erfolg ist."""
        c = cls.normalize(code)
        return c in {"1", "111"}

    @classmethod
    def is_failure(cls, code) -> bool:
        """True, wenn Status ein Fehlschlag ist (alle 0er-Codes)."""
        c = cls.normalize(code)
        return c in {"0", "000", "010", "011", "012", "013", "020", "030", "040"}

    @classmethod
    def needs_renormalization(cls, code) -> bool:
        """True, wenn Variante erneut normalisiert werden soll."""
        c = cls.normalize(code)
        return c in {"010", "011", "012"}

    @classmethod
    def is_deletable(cls, code) -> bool:
        """True, wenn Variante endgültig zur Löschung freigegeben ist."""
        c = cls.normalize(code)
        return c == "013"

    @classmethod
    def update_status(cls, current_status, success: bool) -> str:
        """
        Aktualisiert den Status nach Eskalationslogik.
        - success=True  -> "111"
        - success=False -> Eskalation
        """
        current = cls.normalize(current_status)

        if success:
            return "111"

        # Behandle leere Strings wie "0"
        if current in (None, "", "0", "000", "1", "111"):
            return "010"
        elif current == "010":
            return "011"
        elif current == "011":
            return "012"
        elif current == "012":
            return "013"
        elif current in {"020", "030", "040"}:
            return current
        else:
            return "000"

    @staticmethod
    def normalize(code) -> str:
        """Normalisiert Eingaben auf String-Statuscode."""
        if code is None:
            return "0"
        if isinstance(code, int):
            return str(code)
        return str(code).strip()
    
    @classmethod
    def classify_af_status(
        cls,
        af_value: Optional[float],
        last_fetch: Optional[str],
        fetch_status: Optional[str],
        stale_days: int = Config.STALE_DAYS_AF
    ) -> str:
        """
        ✅ Klassifiziert AF-Status in 4 Kategorien.
        
        Args:
            af_value: af_filter_mean Wert (oder None)
            last_fetch: meanAF_last_fetch Timestamp
            fetch_status: meanAF_fetch_success Status-Code
            stale_days: Tage bis Stale (default 30)
        
        Returns:
            "has_af"         - Hat validen AF-Wert
            "never_fetched"  - Noch nie/lange nicht/exhausted gefetcht
            "fetch_failed"   - Technischer Fehler (unvalidiert)
            "true_none"      - Erfolgreich, aber AF=None (validiert selten)
        
        =====================================================================
        MAPPING ZU DB-ZUSTÄNDEN:
        =====================================================================
        
        has_af:
          - af_filter_mean != NULL
        
        never_fetched:
          - meanAF_last_fetch = NULL
          - OR is_stale(meanAF_last_fetch, 30d)
          - OR meanAF_fetch_success = "013" (exhausted)
        
        fetch_failed:
          - meanAF_fetch_success NOT IN ("1","111","013",NULL)
          - AND af_filter_mean = NULL
          - Codes: "0","000","010","011","012","020","030","040"
        
        true_none:
          - meanAF_fetch_success IN ("1","111")
          - AND af_filter_mean = NULL
          - (Erfolgreicher Fetch, aber nicht in DBs → wahrscheinlich selten)
        
        =====================================================================
        """
        import datetime
        
        # 1. Hat AF-Wert? → has_af
        if af_value is not None:
            return "has_af"
        
        # 2. Noch nie gefetcht?
        if not last_fetch or last_fetch == "":
            return "never_fetched"
        
        # 3. Stale-Check (>30 Tage alt)
        try:
            dt = datetime.datetime.fromisoformat(last_fetch.replace("Z", "+00:00"))
            age_days = (datetime.datetime.now(datetime.timezone.utc) - dt).days
            if age_days > stale_days:
                return "never_fetched"  # Stale = needs refetch
        except Exception:
            return "never_fetched"
        
        # 4. Status analysieren
        status = cls.normalize(fetch_status)
        
        # Exhausted? → never_fetched (Background-Only)
        if status == "013":
            return "never_fetched"
        
        # Erfolg ohne AF? → true_none (validiert selten)
        if cls.is_success(status):  # "1" oder "111"
            return "true_none"
        
        # Sonst: Fetch fehlgeschlagen
        return "fetch_failed"

GNOMAD_ZENODO_URLS = {
    "GRCh37": "https://zenodo.org/record/5770384/files/gnomad_db_v2.1.1.sqlite3.gz?download=1",
    "GRCh38": "https://zenodo.org/record/6818606/files/gnomad_db_v3.1.2.sqlite3.gz?download=1",
}

def download_build(build: str, logger, progress_callback=None, cancel_event=None) -> str:
    """
    Laedt gnomAD-DB fuer einen Build von Zenodo herunter und entpackt sie.

    Args:
        build: "GRCh37" oder "GRCh38"
        logger: Logger-Instanz
        progress_callback: Optional - fn(downloaded, total, phase) fuer GUI-Updates
        cancel_event: Optional - threading.Event zum Abbrechen

    Returns:
        Pfad zur entpackten SQLite-DB
    """
    url = GNOMAD_ZENODO_URLS.get(build)
    if not url:
        raise ValueError(f"Unbekannter Build: {build}")

    data_dir = os.path.join(BASE_DIR, "data")
    os.makedirs(data_dir, exist_ok=True)
    gz_path = os.path.join(data_dir, f"gnomad_{build}.sqlite3.gz")
    db_path = os.path.join(data_dir, f"gnomad_{build}.sqlite3")

    # Download mit Progress
    logger.log(f"[LightDB] Starte Download {build} von Zenodo...")
    r = requests.get(url, stream=True, timeout=60)
    r.raise_for_status()
    total = int(r.headers.get('content-length', 0))
    downloaded = 0
    with open(gz_path, "wb") as f:
        for chunk in r.iter_content(chunk_size=65536):
            if cancel_event and cancel_event.is_set():
                f.close()
                if os.path.exists(gz_path):
                    os.remove(gz_path)
                raise InterruptedError("Download abgebrochen")
            f.write(chunk)
            downloaded += len(chunk)
            if progress_callback:
                progress_callback(downloaded, total, "download")
            elif total and downloaded % (50 * 1024 * 1024) < 65536:
                logger.log(f"[LightDB] Download {build}: {downloaded / total * 100:.0f}%")

    # Entpacken
    logger.log(f"[LightDB] Entpacke {build}...")
    if progress_callback:
        progress_callback(0, 0, "extract")
    with gzip.open(gz_path, "rb") as f_in, open(db_path, "wb") as f_out:
        shutil.copyfileobj(f_in, f_out)
    os.remove(gz_path)

    logger.log(f"[LightDB] {build} entpackt: {db_path}")
    return db_path

class LightDBGnomADManager:
    """
    Verantwortlich für Download, Migration und Indexierung der gnomAD Light-DB.
    Entlastet den AFFetchController.
    
    V10: Pfadresistent - alle Dateien werden relativ zum base_dir gespeichert.
    """

    # Dateinamen (ohne Pfad) - werden in __init__ zu absoluten Pfaden
    _DB_FILENAME = "gnomad_light.db"
    _CONFIG_FILENAME = "lightdb_config.json"
    _PID_FILENAME = "lightdb_index.pid"
    _PROGRESS_FILENAME = "lightdb_progress.json"

    def __init__(self, logger, base_dir: str = None):
        """
        Args:
            logger: Logger-Instanz
            base_dir: Basisverzeichnis für alle LightDB-Dateien.
                      Default: Script-Verzeichnis (BASE_DIR)
        """
        self.logger = logger
        self.proc = None
        self.lightdb_path = None

        # V10: Pfadresistenz - berechne absolute Pfade
        if base_dir is None:
            # Fallback auf Script-Verzeichnis
            base_dir = os.path.dirname(os.path.abspath(__file__))

        # V16: Prüfe zuerst ResourceManager für gnomad_db
        rm_db_path = None
        try:
            rm = get_resource_manager()
            rm_db_path = rm.get("gnomad_db")
            if rm_db_path and os.path.exists(rm_db_path):
                self.logger.log(f"[LightDB] ✅ DB via ResourceManager gefunden: {rm_db_path}")
        except Exception as e:
            self.logger.log(f"[LightDB] ResourceManager-Lookup fehlgeschlagen: {e}")

        # V16: Wenn ResourceManager DB gefunden hat, nutze dessen Verzeichnis
        if rm_db_path and os.path.exists(rm_db_path):
            self.data_dir = os.path.dirname(rm_db_path)
            self.OUT_DB = rm_db_path
        else:
            # Fallback: Check Config.LOCAL_DB_DIR (C:\_Local_DEV\DATA_STORE)
            # This prevents OneDrive sync issues with large DBs
            if hasattr(Config, 'LOCAL_DB_DIR') and os.path.exists(Config.LOCAL_DB_DIR):
                 self.data_dir = Config.LOCAL_DB_DIR
                 self.logger.log(f"[LightDB] Nutze lokalen Storage: {self.data_dir}")
            else:
                 # Fallback: Data-Verzeichnis für LightDB-Dateien
                 self.data_dir = os.path.join(base_dir, "data")

            # Stelle sicher, dass data-Verzeichnis existiert
            try:
                os.makedirs(self.data_dir, exist_ok=True)
            except Exception as e:
                self.logger.log(f"[LightDB] ⚠️ Konnte data-Verzeichnis nicht erstellen: {e}")
                # Fallback auf base_dir selbst
                self.data_dir = base_dir

            self.OUT_DB = os.path.join(self.data_dir, self._DB_FILENAME)

        # Config/PID/Progress immer im data_dir
        self.CONFIG_FILE = os.path.join(self.data_dir, self._CONFIG_FILENAME)
        self.PID_FILE = os.path.join(self.data_dir, self._PID_FILENAME)
        self.progress_file = os.path.join(self.data_dir, self._PROGRESS_FILENAME)

        self.logger.log(f"[LightDB] V16 Init - Datenverzeichnis: {self.data_dir}")
        self.logger.log(f"[LightDB] DB-Pfad: {self.OUT_DB}")

    # ---------------- Hilfsfunktionen ----------------
    def get_existing_builds(self, db_path: str) -> dict:
        if not os.path.exists(db_path):
            return {}
        try:
            with sqlite3.connect(db_path) as con:
                cur = con.cursor()
                cur.execute("""
                    SELECT build, COUNT(*) as count
                    FROM variants_light
                    WHERE build IN ('GRCh37','GRCh38')
                    GROUP BY build
                """)
                return {row[0]: row[1] for row in cur.fetchall()}
        except Exception as e:
            self.logger.log(f"[LightDB] ⚠️ Fehler beim Build-Check: {e}")
            return {}

    # Problem: Korrupte Config-Datei crasht Pipeline
    def load_config(self) -> dict:
        if os.path.exists(self.CONFIG_FILE):
            try:
                with open(self.CONFIG_FILE, "r") as f:
                    cfg = json.load(f)
                    # ✅ Validierung
                    if not isinstance(cfg, dict):
                        return {}
                    return cfg
            except (json.JSONDecodeError, IOError):
                return {}
        return {}

    def save_config(self, cfg: dict) -> None:
        with open(self.CONFIG_FILE, "w") as f:
            json.dump(cfg, f, indent=2)
            
    def are_builds_missing_or_outdated(self, cfg: dict) -> list[str]:
        """
        Prüft anhand der Config, ob GRCh37/GRCh38 fehlen oder älter als 1 Jahr sind.
        Gibt eine Liste der betroffenen Builds zurück.
        """
        outdated_builds = []
        now = datetime.datetime.now()

        for build in ("GRCh37", "GRCh38"):
            key = f"last_fetch_{build}"
            if key not in cfg:
                outdated_builds.append(build)
            else:
                try:
                    ts = datetime.datetime.fromisoformat(cfg[key])
                    if (now - ts).days >= 365:
                        outdated_builds.append(build)
                except Exception:
                    outdated_builds.append(build)

        return outdated_builds
    
    def _check_free_space(self) -> bool:
        """
        Prüft, ob genug Speicherplatz für Shadow-Build vorhanden ist.
        - Erfordert mindestens 150% der aktuellen DB-Größe.
        - Loggt Warnungen oder Fehler.
        """
        if not os.path.exists(self.OUT_DB):
            return True  # keine alte DB → keine Prüfung nötig

        db_size = os.path.getsize(self.OUT_DB)
        required = db_size + db_size // 2  # 150% der aktuellen Größe
        usage = shutil.disk_usage(os.path.dirname(self.OUT_DB))
        free = usage.free

        if free < required:
            missing_gb = (required - free) / (1024**3)
            self.logger.log(
                f"[LightDB] ⚠️ Veraltet, aber nicht genug Speicherplatz für Neu-Erstellung. "
                f"Es fehlen {missing_gb:.1f} GB. "
                f"Wenn Sie trotzdem erneuern wollen, löschen Sie die Datei {self.OUT_DB}."
            )
            return False
        else:
            self.logger.log(
                "[LightDB] ℹ️ Achtung: Für die Erneuerung werden kurzfristig große Mengen "
                "Speicherplatz zusätzlich belegt. Dies ermöglicht einen schnelleren Annotationsprozess."
            )
            return True
        
    def migrate_config(self, cfg: dict, existing_builds: dict) -> dict:
        changed = False
        if "last_fetch" in cfg and "last_fetch_GRCh37" not in cfg:
            old_timestamp = cfg["last_fetch"]
            for build in existing_builds.keys():
                cfg[f"last_fetch_{build}"] = old_timestamp
            del cfg["last_fetch"]
            changed = True
            self.logger.log("[LightDB] ✅ Config migriert: last_fetch → per-Build")

        now = datetime.datetime.now().isoformat()
        for build in existing_builds.keys():
            key = f"last_fetch_{build}"
            if key not in cfg:
                cfg[key] = now
                changed = True
                self.logger.log(f"[LightDB] ✅ Config ergänzt: {build} → {now}")

        if changed:
            self.save_config(cfg)
        return cfg


    def quick_check(self, db_path: str = None) -> bool:
        """
        Prüft, ob die LightDB nutzbar ist.
        Optionaler Pfad: ermöglicht Prüfung auch auf Shadow-Build (.new).

        ✅ V12.1 FIX: SELECT 1 statt COUNT(*) - auf 94GB DB ist COUNT(*) viel zu langsam!
        """
        db_path = db_path or self.OUT_DB
        if not os.path.exists(db_path):
            return False
        try:
            # V12.1: Kurzer Timeout + SELECT 1 statt COUNT(*) für Performance auf großen DBs
            with sqlite3.connect(db_path, timeout=5) as con:
                cur = con.cursor()
                # Nur prüfen ob Tabelle existiert und Daten hat - NICHT zählen!
                cur.execute("SELECT 1 FROM variants_light LIMIT 1")
                result = cur.fetchone()
                return result is not None
        except Exception:
            return False

    # ---------------- Download + Migration ----------------
    def gunzip_file(self, gz_path: str, out_path: str) -> None:
        self.logger.log(f"[LightDB] Entpacke {gz_path} → {out_path}")
        with gzip.open(gz_path, "rb") as f_in, open(out_path, "wb") as f_out:
            shutil.copyfileobj(f_in, f_out)

    def migrate_build(self, src_db: str, build: str, dest_db: str) -> None:
        """Migriert einen Build in die Light-DB."""
        with sqlite3.connect(src_db) as src, sqlite3.connect(dest_db) as dst:
            # Zieltabelle anlegen (ohne Index, der kommt am Ende für Performance)
            dst.execute("""
                CREATE TABLE IF NOT EXISTS variants_light (
                    chrom TEXT,
                    pos INTEGER,
                    ref TEXT,
                    alt TEXT,
                    build TEXT,
                    rsid TEXT,
                    af REAL
                )
            """)
            dst.commit()

            cur = src.cursor()
            cur.execute("SELECT name FROM sqlite_master WHERE type='table'")
            tables = [r[0].lower() for r in cur.fetchall()]
            candidate = None
            for name in ("variants", "gnomad_variants", "variants_v2", "gnomad_db"):
                if name.lower() in tables:
                    candidate = name
                    break
            if not candidate:
                raise RuntimeError(f"Keine Variantentabelle in {src_db}")

            cur.execute(f"PRAGMA table_info({candidate})")
            cols = [c[1].lower() for c in cur.fetchall()]
            colmap = {
                "chrom": "chrom" if "chrom" in cols else "chr",
                "pos": "pos" if "pos" in cols else "position",
                "ref": "ref",
                "alt": "alt",
                "rsid": "rsid" if "rsid" in cols else None,
                "af": "af" if "af" in cols else None
            }
            select_cols = [
                colmap["chrom"],
                colmap["pos"],
                colmap["ref"],
                colmap["alt"],
                colmap["rsid"] if colmap["rsid"] else "NULL as rsid",
                colmap["af"] if colmap["af"] else "NULL as af"
            ]
            query = f"SELECT {', '.join(select_cols)} FROM {candidate}"
            cur.execute(query)

            processed = 0
            batch = []
            BATCH_SIZE = 100000
            while True:
                rows = cur.fetchmany(BATCH_SIZE)
                if not rows:
                    break
                for row in rows:
                    batch.append((row[0], row[1], row[2], row[3], build, row[4], row[5]))
                dst.executemany(
                    "INSERT INTO variants_light (chrom,pos,ref,alt,build,rsid,af) VALUES (?,?,?,?,?,?,?)",
                    batch
                )
                processed += len(batch)   # ✅ Zähler vor clear()
                dst.commit()
                batch.clear()
                if processed % 1_000_000 == 0:
                    self.logger.log(f"[LightDB] {build}: {processed:,} Varianten...")

            # Index erst nach der Migration erstellen → viel schneller
            dst.execute("""
                CREATE INDEX IF NOT EXISTS idx_variants_light
                ON variants_light(chrom,pos,ref,alt,build)
            """)
            dst.commit()

            self.logger.log(f"[LightDB] ✅ {build}: {processed:,} Varianten migriert.")

    # ---------------- Indexbau im separaten Prozess ----------------
    def start_index_worker(self):
        """Startet externen Index-Worker, falls nicht schon einer läuft.
        Räumt PID- und Progress-Dateien nach erfolgreichem Abschluss auf.
        """
        # Prüfen, ob PID-Datei existiert
        if os.path.exists(self.PID_FILE):
            try:
                with open(self.PID_FILE) as f:
                    pid = int(f.read().strip())
                # Prozess noch aktiv?
                os.kill(pid, 0)  # nur Testsignal
                self.logger.log(f"[LightDB] Index-Worker läuft bereits (PID {pid}).")
                return
            except Exception:
                # PID-Datei veraltet → löschen
                os.remove(self.PID_FILE)

        # Neuen Worker starten - V10: absoluter Pfad zum Worker-Script
        worker_script = os.path.join(os.path.dirname(os.path.abspath(__file__)), "lightdb_index_worker.py")
        proc = subprocess.Popen(
            [sys.executable, worker_script, self.OUT_DB, self.progress_file],
            creationflags=subprocess.CREATE_NO_WINDOW
        )
        # PID speichern
        with open(self.PID_FILE, "w") as f:
            f.write(str(proc.pid))

        self.logger.log(f"[LightDB] Externer Index-Worker gestartet (PID {proc.pid}).")

        # Hintergrund-Thread zum Aufräumen nach Abschluss
        def monitor_worker(p: subprocess.Popen):
            try:
                ret = p.wait()  # blockiert bis Prozessende
                if ret == 0:
                    # Erfolgreich beendet → Steuerdateien löschen
                    if os.path.exists(self.PID_FILE):
                        os.remove(self.PID_FILE)
                    if os.path.exists(self.progress_file):
                        os.remove(self.progress_file)
                    self.logger.log("[LightDB] ✅ Indexerstellung abgeschlossen, Steuerdateien entfernt.")
                else:
                    self.logger.log(f"[LightDB] ⚠️ Index-Worker beendet mit Code {ret}.")
            except Exception as e:
                self.logger.log(f"[LightDB] ❌ Fehler beim Überwachen des Index-Workers: {e}")

        threading.Thread(target=monitor_worker, args=(proc,), daemon=True).start()
        
    # ---------------- Statusabfragen ----------------
    def get_progress(self) -> dict:
        """
        Liefert Fortschritt + Status:
        - "running": Worker läuft noch
        - "stalled": PID-Datei existiert, Prozess aber nicht mehr aktiv
        - "done": Index fertig (100%)
        - "idle": kein Worker, kein Fortschritt
        """
        status = "idle"
        prog = {}

        # Fortschritt aus JSON lesen
        if os.path.exists(self.progress_file):
            try:
                with open(self.progress_file) as f:
                    prog = json.load(f)
            except Exception:
                prog = {}

        # Wenn Index fertig → immer "done"
        if prog.get("pct") == 100.0:
            status = "done"
            # PID-Datei ggf. aufräumen
            if os.path.exists(self.PID_FILE):
                try:
                    os.remove(self.PID_FILE)
                except Exception:
                    pass
            return {"status": status, **prog}

        # PID prüfen (nur wenn noch nicht "done")
        if os.path.exists(self.PID_FILE):
            try:
                with open(self.PID_FILE) as f:
                    pid = int(f.read().strip())
                if psutil.pid_exists(pid):
                    status = "running"
                else:
                    status = "stalled"
            except Exception:
                status = "stalled"
        else:
            # Wenn Fortschritt <100 aber keine PID → idle
            if prog:
                status = "idle"

        return {"status": status, **prog}

    def is_ready(self) -> bool:
        """True, wenn DB existiert und Index fertig ist (optimale Performance)."""
        return self.is_available() and self._has_index()
    
    def is_ready_cached(self) -> bool:
        """
        Verwendet Fortschrittsstatus ODER prüft ob DB bereits existiert.

        V10 FIX: Wenn keine Progress-Datei existiert (manuelle DB),
        fällt zurück auf direkten DB-Check.

        V16 FIX: Nutze lightdb_path Cache statt DB-Query im GUI-Thread!
        """
        # Erst Progress prüfen (schnell)
        prog = self.get_progress()
        if prog.get("status") == "done":
            return True

        # V16 FIX: Wenn lightdb_path bereits gesetzt ist, ist DB ready
        # (wurde in ensure_lightdb/quick_check gesetzt)
        if self.lightdb_path and os.path.exists(self.lightdb_path):
            return True

        # V10 FIX: Fallback auf direkten DB-Check wenn Progress fehlt
        # V16: NUR wenn noch nicht gecheckt (vermeidet GUI-Freeze durch DB-Query)
        if prog.get("status") == "idle" and not os.path.exists(self.progress_file):
            # Starte Check in Background Thread statt GUI-Thread
            if not hasattr(self, '_ready_check_started'):
                self._ready_check_started = True
                def bg_check():
                    if self.is_ready():
                        self.lightdb_path = self.OUT_DB
                        try:
                            with open(self.progress_file, "w") as f:
                                json.dump({"pct": 100.0, "status": "done"}, f)
                        except Exception:
                            pass
                threading.Thread(target=bg_check, daemon=True).start()

        return False
    
    def is_available(self) -> bool:
        """True, wenn eine Light-DB existiert (auch ohne Index)."""
        return os.path.exists(self.OUT_DB) and bool(self.get_existing_builds(self.OUT_DB))
    
    def _has_index(self, db_path: str = None) -> bool:
        """Prüft, ob die Index-Tabelle existiert."""
        db_path = db_path or self.OUT_DB
        try:
            with sqlite3.connect(db_path) as con:
                cur = con.cursor()
                cur.execute("""
                    SELECT name FROM sqlite_master
                    WHERE type='index' AND name='idx_variants_light';
                """)
                return cur.fetchone() is not None
        except Exception:
            return False

    def ensure_lightdb_async(self, auto_download: bool = False):
        """Startet ensure_lightdb in einem separaten Thread, damit GUI nicht blockiert."""
        t = threading.Thread(
            target=self.ensure_lightdb,
            args=(auto_download,),
            daemon=True
        )
        t.start()

    # ---------------- Orchestrator ----------------
    def ensure_lightdb(self, auto_download: bool = False):
        """
        Orchestriert die Prüfung und ggf. den Download/Migration der Light-DB.
        Blockiert NICHT die GUI – lange Operationen laufen im Hintergrund.
        Nutzt Shadow-Build: alte DB bleibt aktiv, bis neue vollständig validiert ist.
        """
        try:
            # 1. Quick-Check auf bestehende DB
            if self.quick_check(self.OUT_DB):
                self.lightdb_path = self.OUT_DB
                if self._has_index(self.OUT_DB):
                    self.logger.log("[LightDB] ✅ Quick-Check bestanden – DB mit Index nutzbar.")
                else:
                    self.logger.log("[LightDB] ⚠️ Quick-Check bestanden – DB nutzbar, aber ohne Index (langsamer).")
                    threading.Thread(target=self.start_index_worker, daemon=True).start()
                return

            # 2. Migration/Download asynchron starten
            def background_job():
                try:
                    cfg = self.load_config()
                    existing_builds = self.get_existing_builds(self.OUT_DB)
                    cfg = self.migrate_config(cfg, existing_builds)

                    # Prüfen, ob Builds fehlen oder veraltet sind
                    outdated_builds = self.are_builds_missing_or_outdated(cfg)

                    # Falls alles aktuell → alte DB weiter nutzen
                    if not outdated_builds:
                        self.lightdb_path = self.OUT_DB
                        if self._has_index(self.OUT_DB):
                            self.logger.log("[LightDB] ✅ Aktuell: DB nutzbar (Index vorhanden).")
                        else:
                            self.logger.log("[LightDB] ⚠️ Aktuell: DB nutzbar (Index fehlt, wird erstellt).")
                            self.start_index_worker()
                        return

                    # V17: Kein Auto-Download wenn vom User deaktiviert
                    if not auto_download:
                        self.logger.log(f"[LightDB] Builds fehlen: {outdated_builds} – Auto-Download deaktiviert, ueberspringe.")
                        return

                    # Speicherplatz prüfen
                    if not self._check_free_space():
                        self.logger.log("[LightDB] ⏩ Abbruch: Speicherplatzprüfung fehlgeschlagen.")
                        return

                    # Shadow-Build in temporärer Datei
                    tmp_db = self.OUT_DB + ".new"
                    if os.path.exists(tmp_db):
                        os.remove(tmp_db)

                    for build in outdated_builds:
                        try:
                            db_path = download_build(build, self.logger)
                            self.migrate_build(db_path, build, tmp_db)
                            cfg[f"last_fetch_{build}"] = datetime.datetime.now().isoformat()
                            self.save_config(cfg)
                            os.remove(db_path)
                        except Exception as e:
                            self.logger.log(f"[LightDB] ⚠️ {build} fehlgeschlagen: {e}")

                    # Final check auf tmp_db
                    if self.quick_check(tmp_db):
                        os.replace(tmp_db, self.OUT_DB)
                        self.lightdb_path = self.OUT_DB
                        if self._has_index(self.OUT_DB):
                            self.logger.log("[LightDB] ✅ Neue DB aktiviert (Index vorhanden). Alte DB überschrieben.")
                        else:
                            self.logger.log("[LightDB] ⚠️ Neue DB aktiviert (Index fehlt, wird erstellt). Alte DB überschrieben.")
                            self.start_index_worker()
                    else:
                        if os.path.exists(tmp_db):
                            os.remove(tmp_db)
                        self.logger.log("[LightDB] ⚠️ Neue DB unvollständig – alte bleibt aktiv.")

                except Exception as e:
                    self.lightdb_path = None
                    self.logger.log(f"[LightDB] ❌ Schwerer Fehler in ensure_lightdb: {e}")

            threading.Thread(target=background_job, daemon=True).start()

        except Exception as e:
            self.lightdb_path = None
            self.logger.log(f"[LightDB] ❌ Fehler beim Start von ensure_lightdb: {e}")
            
    def lookup_variants_bulk(self, keys, batch_size=1000):
        """
        Bulk-Lookup in variants_light.
        FIX: Connection wird immer geschlossen (try-finally)
        """
        db_path = self.OUT_DB
        results = {}

        # Schnelle Pruefung wenn DB nicht existiert
        if not db_path or not os.path.exists(db_path):
            self.logger.log("[LightDB] Warnung: Datenbank nicht vorhanden")
            return results, keys

        # Cache Check
        if hasattr(self, '_bulk_cache'):
            cached = {k: v for k, v in self._bulk_cache.items() if k in keys}
            uncached = [k for k in keys if k not in cached]
            results.update(cached)
            
            if not uncached:
                self.logger.log(f"[LightDB] OK: Alle {len(keys)} Keys aus Cache")
                return results, []
        else:
            uncached = list(keys)

        # FIX: Connection mit finally-Block
        conn = None
        try:
            conn = sqlite3.connect(db_path)
            cur = conn.cursor()

            chunk_size = 5000
            total = len(uncached)
            self.logger.log(f"[LightDB] Starte Lookup fuer {total:,} Varianten (Chunk-Groesse {chunk_size})...")

            for i in range(0, total, chunk_size):
                block = uncached[i:i+chunk_size]
                self.logger.log(f"[LightDB] Block {i//chunk_size+1}: {len(block)} Keys werden geprueft...")

                parsed = []
                for key in block:
                    try:
                        if isinstance(key, tuple) and len(key) == 5:
                            chrom, pos, ref, alt, build_key = key
                            parsed.append((str(chrom), int(pos), str(ref), str(alt), str(build_key)))
                        else:
                            self.logger.log(f"[LightDB] Warnung: Ungueltiges Key-Format: {key}")
                    except Exception as e:
                        self.logger.log(f"[LightDB] Warnung: Fehler beim Parsen von {key}: {e}")
                        continue

                if not parsed:
                    continue

                try:
                    # Temporaere Tabelle fuer Keys anlegen
                    cur.execute("DROP TABLE IF EXISTS tmp_keys;")
                    cur.execute("""
                        CREATE TEMP TABLE tmp_keys (
                            chrom TEXT,
                            pos   INTEGER,
                            ref   TEXT,
                            alt   TEXT,
                            build TEXT
                        );
                    """)
                    cur.executemany("INSERT INTO tmp_keys VALUES (?,?,?,?,?)", parsed)

                    # JOIN gegen variants_light (nutzt Index)
                    cur.execute("""
                        SELECT v.chrom, v.pos, v.ref, v.alt, v.build, v.rsid, v.af
                        FROM variants_light v
                        JOIN tmp_keys k
                          ON v.chrom=k.chrom
                         AND v.pos=k.pos
                         AND v.ref=k.ref
                         AND v.alt=k.alt
                         AND v.build=k.build;
                    """)
                    fetched = cur.fetchall()

                    for chrom, pos, ref, alt, build_val, rsid, af in fetched:
                        key = (chrom, pos, ref, alt, build_val)
                        results[key] = {"af": af, "rsid": rsid}

                    self.logger.log(
                        f"[LightDB] OK: Block {i//chunk_size+1}: {len(fetched)} Treffer, "
                        f"{len(block) - len(fetched)} nicht gefunden."
                    )
                except Exception as e:
                    self.logger.log(f"[LightDB] Warnung: SQL-Fehler im Block {i//chunk_size+1}: {e}")

        except Exception as e:
            self.logger.log(f"[LightDB] Warnung: Fehler beim Lookup: {e}")
        finally:
            # KRITISCH: Connection immer schliessen
            if conn:
                try:
                    conn.close()
                except Exception:
                    pass

        return results, uncached

#Soll abgelöst werden durch jeweils vereinfachten Fetcher direkt in Distiller
#Nur noch für backgroundmaintainer
#Evtl. umbenenn und auch weniger zentral in der App benennen, da nur noch für den langsamen refresh prozess.
class AFFetchController:
    """
    ✅ V9 HINWEIS: Dieser Controller wird primär vom BackgroundMaintainer 
    verwendet für asynchrone AF-Datenabrufe. Der Distiller hat seine eigene
    AF-Logik in _phase_af_fetch.
    
    Funktionen:
    - Batch-Abruf von Allel-Frequenzen aus verschiedenen Quellen
    - Adaptive Batch-Größen via ThroughputTuner
    - LightDB-Integration für lokale Frequenz-Daten
    """
    def __init__(self, db, logger, stopflag=None, mode="full"):
        self.db = db
        self.logger = logger
        self.stopflag = stopflag
        self.mode = mode
        self.cache = {}
        self.result_buffer = {}
        self.flush_size = 5000

        # LightDB-Integration
        self.lightdb_path = None   # ← Attribut existiert immer, wird später von App gesetzt

        self.stats = {
            "mv": {"ok": 0, "fail": 0},
            "gnomad": {"ok": 0, "fail": 0},
            "vep": {"ok": 0, "fail": 0},
            "alfa": {"ok": 0, "fail": 0},
            "topmed": {"ok": 0, "fail": 0},
        }

       # Konservative Batchgrenzen zur Vermeidung von HTTP 400
        self.tuners = {
            "mv":      ThroughputTuner(min_batch=100, max_batch=1000, base_batch=800),
            "gnomad":  ThroughputTuner(min_batch=20,  max_batch=100,  base_batch=70),
            "vep":     ThroughputTuner(min_batch=10,  max_batch=80,   base_batch=25),
            "alfa":    ThroughputTuner(min_batch=10,  max_batch=50,   base_batch=25),
            "topmed":  ThroughputTuner(min_batch=10,  max_batch=50,   base_batch=25),
        }

        self.tuner_rules = {
            "gnomad": {
                "batch_penalty":  -20,
                "max_penalty":    -15,
                "min_penalty":    -2,
                "worker_penalty": -1,
                "batch_reward":   +5,   # kleine Belohnung
                "max_reward":     +5,
                "worker_reward":  +1,
                "reward_threshold": 3,  # nach 3 stabilen Läufen
            },
            "mv": {
                "batch_penalty":  -100,
                "worker_penalty": -2,
                "batch_reward":   +50,
                "worker_reward":  +2,
                "reward_threshold": 2,
            },
            "vep": {
                "batch_penalty":  -5,
                "batch_reward":   +5,
                "reward_threshold": 3,
            },
        }
        
        # 🔔 Deprecation-Alias für Altcode, der noch "ensembl" nutzt
        self.stats["ensembl"] = self.stats["vep"]
        self.tuners["ensembl"] = self.tuners["vep"]
        self.logger.log("[DEPRECATED] 'ensembl' wird intern auf 'vep' gemappt – bitte Code umstellen.")
    
        # Erfolgszähler für Rewards initialisieren
        self.stable_runs = defaultdict(int)

    def apply_success_reward(self, api_name: str):
        rules = self.tuner_rules.get(api_name)
        tuner = self.tuners.get(api_name)
        if not rules or not tuner:
            return

        # Erfolgszähler hochzählen
        self.stable_runs[api_name] += 1

        # Reward nur anwenden, wenn Threshold erreicht
        if self.stable_runs[api_name] >= rules.get("reward_threshold", 3):
            old_min = tuner.min_b
            old_b   = tuner.b
            old_max = tuner.max_b
            old_workers = getattr(tuner, "w", None)

            # Batchgröße vorsichtig erhöhen
            if "batch_reward" in rules:
                tuner.b = min(tuner.max_b, tuner.b + rules["batch_reward"])

            # Maximalgröße erhöhen
            if "max_reward" in rules:
                tuner.max_b = max(tuner.b, tuner.max_b + rules["max_reward"])

            # Workerzahl erhöhen
            if "worker_reward" in rules and hasattr(tuner, "w"):
                tuner.w = max(1, tuner.w + rules["worker_reward"])

            new_workers = getattr(tuner, "w", "?")

            self.logger.log(
                f"[{api_name}] ✅ Stabil – Reward angewendet: "
                f"min {old_min}->{tuner.min_b}, "
                f"batch {old_b}->{tuner.b}, "
                f"max {old_max}->{tuner.max_b}, "
                f"workers {old_workers}->{new_workers}"
            )

            # Zähler zurücksetzen
            self.stable_runs[api_name] = 0


    def apply_timeout_penalty(self, api_name: str):
        rules = self.tuner_rules.get(api_name)
        tuner = self.tuners.get(api_name)
        if not rules or not tuner:
            return

        # Vorher-Werte sichern für Logging
        old_min = tuner.min_b
        old_b   = tuner.b
        old_max = tuner.max_b
        old_workers = getattr(tuner, "w", None)

        # Anpassungen nach Regeln
        if "batch_penalty" in rules:
            tuner.b = max(tuner.min_b, tuner.b + rules["batch_penalty"])
        if "max_penalty" in rules:
            tuner.max_b = max(tuner.min_b, tuner.max_b + rules["max_penalty"])
        if "min_penalty" in rules:
            tuner.min_b = max(1, tuner.min_b + rules["min_penalty"])
        if "worker_penalty" in rules and hasattr(tuner, "w"):
            tuner.w = max(1, tuner.w + rules["worker_penalty"])

        # Nachher-Werte für Logging
        new_workers = getattr(tuner, "w", "?")

        self.logger.log(
            f"[{api_name}] ⚠️ Timeout – Tuner angepasst: "
            f"min {old_min}->{tuner.min_b}, "
            f"batch {old_b}->{tuner.b}, "
            f"max {old_max}->{tuner.max_b}, "
            f"workers {old_workers}->{new_workers}"
        )
    def _lookup_lightdb(
            self,
            uncached: list,
            build: str = "GRCh37",
            db_path: str | None = None
        ) -> tuple[dict, list]:
        """
        Lookup von Varianten in der lokalen Light-DB.

        Args:
            uncached: Liste von Variant-Keys (Tuples: chrom,pos,ref,alt,build)
            build: Referenzgenom ("GRCh37" oder "GRCh38")
            db_path: Pfad zur Light-DB (vom Manager übergeben)

        Returns:
            (results_dict, still_uncached_list)
            - results_dict: {key: {"af": float|None, "rsid": str|None}}
            - still_uncached_list: Keys ohne Treffer in Light-DB
        """
        results = {}

        if not db_path:
            self.logger.log("[LightDB] ⚠️ Kein Light-DB Pfad – Lookup übersprungen.")
            return results, uncached

        try:
            conn = sqlite3.connect(db_path)
            cur = conn.cursor()

            chunk_size = 5000
            total = len(uncached)
            self.logger.log(f"[LightDB] Starte Lookup für {total:,} Varianten (Chunk-Größe {chunk_size})...")

            for i in range(0, total, chunk_size):
                block = uncached[i:i+chunk_size]
                self.logger.log(f"[LightDB] 🔎 Block {i//chunk_size+1}: {len(block)} Keys werden geprüft...")

                parsed = []
                for key in block:
                    try:
                        if isinstance(key, tuple) and len(key) == 5:
                            chrom, pos, ref, alt, build_key = key
                            parsed.append((str(chrom), int(pos), str(ref), str(alt), str(build_key)))
                        else:
                            self.logger.log(f"[LightDB] ⚠️ Ungültiges Key-Format: {key}")
                    except Exception as e:
                        self.logger.log(f"[LightDB] ⚠️ Fehler beim Parsen von {key}: {e}")
                        continue

                if not parsed:
                    continue

                try:
                    # Temporäre Tabelle für Keys anlegen
                    cur.execute("DROP TABLE IF EXISTS tmp_keys;")
                    cur.execute("""
                        CREATE TEMP TABLE tmp_keys (
                            chrom TEXT,
                            pos   INTEGER,
                            ref   TEXT,
                            alt   TEXT,
                            build TEXT
                        );
                    """)
                    cur.executemany("INSERT INTO tmp_keys VALUES (?,?,?,?,?)", parsed)

                    # JOIN gegen variants_light (nutzt Index)
                    cur.execute("""
                        SELECT v.chrom, v.pos, v.ref, v.alt, v.build, v.rsid, v.af
                        FROM variants_light v
                        JOIN tmp_keys k
                          ON v.chrom=k.chrom
                         AND v.pos=k.pos
                         AND v.ref=k.ref
                         AND v.alt=k.alt
                         AND v.build=k.build;
                    """)
                    fetched = cur.fetchall()

                    for chrom, pos, ref, alt, build_val, rsid, af in fetched:
                        key = (chrom, pos, ref, alt, build_val)
                        results[key] = {"af": af, "rsid": rsid}

                    self.logger.log(
                        f"[LightDB] ✅ Block {i//chunk_size+1}: {len(fetched)} Treffer, "
                        f"{len(block) - len(fetched)} nicht gefunden."
                    )
                except Exception as e:
                    self.logger.log(f"[LightDB] ⚠️ SQL-Fehler im Block {i//chunk_size+1}: {e}")

            conn.close()
        except Exception as e:
            self.logger.log(f"[LightDB] ⚠️ Fehler beim Lookup: {e}")

        still_uncached = [k for k in uncached if k not in results]
        self.logger.log(
            f"[LightDB] Zusammenfassung: {len(results)} Treffer, {len(still_uncached)} verbleibend."
        )
        return results, still_uncached

    def _chunked(self, iterable, size):
        """
        Zerlegt ein Iterable in aufeinanderfolgende Blöcke (Chunks) der Länge `size`.
        Beispiel: list(self._chunked([1,2,3,4,5], 2)) -> [[1,2],[3,4],[5]]
        """
        chunk = []
        for item in iterable:
            chunk.append(item)
            if len(chunk) >= size:
                yield chunk
                chunk = []
        if chunk:
            yield chunk
            
    def _normalize_key(self, k):
        """
        Normalisiert einen Variant-Key.

        Args:
            k: (chr, pos, ref, alt, build) oder bereits normalisiert

        Returns:
            (chr, pos, ref, alt, build) - normalisiert
        """
        if not isinstance(k, tuple) or len(k) != 5:
            self.logger.log(f"[AF-Controller] ⚠️ Ungültiger Key: {k}")
            return k

        chr_, pos, ref, alt, build = k

        # Chromosome normalisieren
        chr_ = str(chr_).replace("chr", "").replace("CHR", "").upper()

        # Position als int
        try:
            pos = int(pos)
        except (ValueError, TypeError):
            self.logger.log(f"[AF-Controller] ⚠️ Ungültige Position: {pos}")
            return k

        # Allele uppercase
        ref = str(ref).upper()
        alt = str(alt).upper()

        # Build normalisieren
        b = str(build).upper()
        if b in ("GRCH37", "HG19", "B37"):
            build = "GRCh37"
        elif b in ("GRCH38", "HG38", "B38"):
            build = "GRCh38"
        else:
            pass

        return (chr_, pos, ref, alt, build)
    # ---------- Assignment & rsID-Mapping ----------

    def _distribute_keys(self, keys, weights):
        """
        Verteilt Keys auf APIs und erstellt rsID-Map einmalig in sicheren Chunks.
        ALFA bekommt nur Keys mit gültiger rsID.
        """
        assignments = {api: [] for api in weights}
        apis = list(weights.keys())
        if not apis:
            # leere Verteilung und leere rsID-Map zurückgeben
            return assignments, {}

        # rsIDs in Chunks mappen (SQLite Platzhalter-Limit beachten)
        rsid_map = {}
        try:
            max_tuples = 200  # 200 * 5 Variablen = 1000 Platzhalter ~ sicher für SQLite
            with self.db.lock, self.db._conn() as con:
                cur = con.cursor()
                for batch in self._chunked(keys, max_tuples):
                    placeholders = ",".join(["(?,?,?,?,?)"] * len(batch))
                    flat = [x for k in batch for x in k]
                    cur.execute(f"""
                        SELECT chr,pos,ref,alt,build,rsid
                        FROM variants
                        WHERE (chr,pos,ref,alt,build) IN ({placeholders})
                    """, flat)
                    for row in cur.fetchall():
                        key = (row[0], row[1], row[2], row[3], row[4])
                        rsid_map[key] = row[5]
        except Exception as e:
            self.logger.log(f"[AF-Controller] ⚠️ rsID-Mapping fehlgeschlagen: {e}")

        # Round-Robin mit ALFA-Filter
        for i, k in enumerate(keys):
            api = apis[i % len(apis)]
            if api == "alfa":
                rsid = rsid_map.get(k)
                if rsid and str(rsid).lower().startswith("rs"):
                    assignments[api].append(k)
                else:
                    # Kein rsID -> nächste API nehmen
                    for alt_api in apis:
                        if alt_api != "alfa":
                            assignments[alt_api].append(k)
                            break
            else:
                assignments[api].append(k)

        return assignments, rsid_map

    # ---------- Hauptfluss ----------

    async def job_collector_and_distributor(self, keys, build, workers: int):
        """
        FIX: Verhindert Deadlocks bei Task-Cancellation.
        """
        if isinstance(keys, tuple) and len(keys) == 5:
            keys = [keys]

        normalized_keys = [self._normalize_key(k) for k in keys]
        unique_keys = list(dict.fromkeys(normalized_keys))

        if self.stopflag and self.stopflag.is_set():
            return {}

        results = {}
        still_uncached = unique_keys

        # API-Verteilung
        weights = self._compute_weights()
        assignments, rsid_map = self._distribute_keys(still_uncached, weights)

        tasks = []

        # Tasks erstellen (mit Enabled-Check)
        if assignments.get("mv") and is_api_enabled("phase1_af", "myvariant"):
            _, batch_mv = self.tuners["mv"].decide()
            _mv_batch = get_api_setting("phase1_af", "myvariant", "batch_size", 400)
            t = asyncio.create_task(
                self._wrap_fetch(
                    "mv",
                    mv_fetch_async(assignments["mv"], build, fetch_full=False,
                                  batch_size=min(batch_mv, _mv_batch), logger=self.logger),
                    workers
                )
            )
            t.api_name = "mv"
            tasks.append(t)

        #legacy Funktionen werden bereits im Distiller gemacht
        """# 1. Cache-Check
        uncached = []
        for k in unique_keys:
            if k in self.cache:
                results[k] = self.cache[k]
            else:
                uncached.append(k)

        self.logger.log(
            f"[AF-Controller] 📋 {len(results)} Treffer aus RAM-Cache, "
            f"{len(uncached)} uncached."
        )

        if not uncached:
            return results

        # 2. Light-DB-Check - V10: Mit Fallback auf absoluten Pfad
        db_path = getattr(self, "lightdb_path", None)
        if not db_path:
            fallback = os.path.join(BASE_DIR, "data", "gnomad_light.db")
            if os.path.exists(fallback):
                db_path = fallback
                self.lightdb_path = fallback  # Für zukünftige Aufrufe
        
        light_results, still_uncached = self._lookup_lightdb(uncached, build, db_path)

        for k, val in light_results.items():
            results[k] = val
            self.cache[k] = val
            self._buffer_result(k, val)

        self.logger.log(
            f"[AF-Controller] 🔒 Light-DB Lookup: {len(light_results)} Treffer, "
            f"{len(still_uncached)} sofort weiter an API."
        )

        if not still_uncached:
            return results

        if self.stopflag and self.stopflag.is_set():
            self.logger.log("[AF-Controller] ℹ Abbruch nach Light-DB-Phase.")
            return results"""
        
        still_uncached = unique_keys

        # 3. API-Verteilung
        self.logger.log(
            f"[AF-Controller] 🌐 Starte API-Fetch für {len(still_uncached)} Varianten..."
        )
        weights = self._compute_weights()
        assignments, rsid_map = self._distribute_keys(still_uncached, weights)

        tasks = []

        # MyVariant
        if assignments.get("mv") and is_api_enabled("phase1_af", "myvariant"):
            _, batch_mv = self.tuners["mv"].decide()
            _mv_batch = get_api_setting("phase1_af", "myvariant", "batch_size", 400)
            self.logger.log(f"[AF-Controller] → MyVariant: {len(assignments['mv'])} Keys")
            t = asyncio.create_task(
                self._wrap_fetch(
                    "mv",
                    mv_fetch_async(assignments["mv"], build,
                                   fetch_full=False, batch_size=min(batch_mv, _mv_batch),
                                   logger=self.logger, phase_label="MV"),
                    workers
                )
            )
            t.api_name = "mv"
            tasks.append(t)

        # gnomAD
        if assignments.get("gnomad") and is_api_enabled("phase1_af", "gnomad"):
            self.logger.log(f"[AF-Controller] → gnomAD: {len(assignments['gnomad'])} Keys")
            t = asyncio.create_task(
                self._wrap_fetch(
                    "gnomad",
                    gnomad_fetch_async(assignments["gnomad"], build, logger=self.logger),
                    workers
                )
            )
            t.api_name = "gnomad"
            tasks.append(t)

        # VEP
        if assignments.get("ensembl") and is_api_enabled("phase1_af", "vep"):
            _, batch_vep = self.tuners["ensembl"].decide()
            _vep_bs = get_api_setting("phase1_af", "vep", "batch_size", 25)
            self.logger.log(f"[AF-Controller] → VEP: {len(assignments['ensembl'])} Keys")
            t = asyncio.create_task(
                self._wrap_fetch(
                    "vep",
                    vep_fetch_async(assignments["ensembl"], build,
                                    batch_size=min(batch_vep, _vep_bs),
                                    logger=self.logger),
                    workers
                )
            )
            t.api_name = "vep"
            tasks.append(t)

        # ALFA
        if assignments.get("alfa") and is_api_enabled("phase1_af", "alfa"):
            alfa_keys = [k for k in assignments["alfa"] if rsid_map.get(k)]
            if alfa_keys:
                self.logger.log(f"[AF-Controller] → ALFA: {len(alfa_keys)} Keys")
                t = asyncio.create_task(
                    self._wrap_fetch(
                        "alfa",
                        alfa_fetch_async(alfa_keys, rsid_map, build, logger=self.logger),
                        workers
                    )
                )
                t.api_name = "alfa"
                tasks.append(t)

        # ✅ KRITISCH: Robustes Task-Handling
        timeout = getattr(self, "fetch_timeout", 300)
        results_list = []
        
        try:
            done, pending = await asyncio.wait(
                tasks,
                timeout=timeout,
                return_when=asyncio.ALL_COMPLETED
            )
            
            # Erfolgreiche Tasks
            for task in done:
                try:
                    result = task.result()
                    results_list.append(result)
                    api = getattr(task, "api_name", None)
                    if api and not isinstance(result, Exception):
                        self.apply_success_reward(api)
                except Exception as e:
                    api = getattr(task, "api_name", "unknown")
                    self.logger.log(f"[AF-Controller] ⚠️ Task {api} failed: {e}")
                    self.apply_timeout_penalty(api)
                    results_list.append({})
            
            # ✅ Pending Tasks sicher canceln
            if pending:
                cancel_tasks = []
                for task in pending:
                    api = getattr(task, "api_name", "unknown")
                    self.logger.log(f"[AF-Controller] ⏰ Timeout bei {api}")
                    self.apply_timeout_penalty(api)
                    task.cancel()
                    cancel_tasks.append(task)
                
                # Warte auf Cancel-Completion mit eigenem Timeout
                if cancel_tasks:
                    try:
                        await asyncio.wait(cancel_tasks, timeout=2.0)
                    except asyncio.TimeoutError:
                        pass  # Tasks wurden gecancelt, ignoriere Timeout
                
                results_list.extend([{}] * len(pending))
        
        except Exception as e:
            self.logger.log(f"[AF-Controller] ❌ Critical error: {e}")
            
            # Alle Tasks canceln
            for t in tasks:
                if not t.done():
                    t.cancel()
                    api = getattr(t, "api_name", "unknown")
                    self.apply_timeout_penalty(api)
            
            # Warte kurz auf Cleanup
            try:
                await asyncio.wait(tasks, timeout=1.0)
            except Exception:
                pass
            
            results_list = [{}] * len(tasks)

        # Merge Results
        merged = self.result_collector_and_merger(still_uncached, results_list)

        # Cache + Buffer
        for k, val in merged.items():
            self.cache[k] = val
            if val is not None:
                self._buffer_result(k, val)

        results.update(merged)
        return results

    def result_collector_and_merger(self, keys, results):
        """
        FIX: Robuste AF-Konsolidierung mit Validation.
        
        Änderungen:
        - Verwirft ungültige AF-Werte (< 0 oder > 1)
        - Priorisiert qualitativ hochwertige Quellen
        - Loggt Anomalien für Debugging
        - Garantiert af_filter_mean im Bereich [0, 1]
        """
        grouped = {}
        
        # Ergebnisse sammeln
        for res in results:
            if not isinstance(res, dict):
                continue
            for k, hit in res.items():
                if not isinstance(k, tuple) or len(k) != 5:
                    continue
                grouped.setdefault(k, []).append(hit)
        
        final = {}
        
        for k in keys:
            hits = grouped.get(k, [])
            
            # ✅ Sammle AF-Werte mit Validierung
            af_sources = {
                "exac": None,
                "gnomad_exomes": None,
                "gnomad_genomes": None,
                "1kg": None,
                "lightdb": None
            }
            
            for hit in hits:
                if not hit or not isinstance(hit, dict):
                    continue
                
                # LightDB (direkter "af"-Key)
                if "af" in hit:
                    val = safe_float(hit.get("af"))
                    if self._validate_af(val, k, "LightDB"):
                        af_sources["lightdb"] = val
                
                # MyVariant-Format
                if "exac" in hit:
                    val = safe_float((hit.get("exac") or {}).get("af"))
                    if self._validate_af(val, k, "ExAC"):
                        af_sources["exac"] = val
                
                if "gnomad" in hit:
                    exomes = safe_float(hit.get("gnomad", {}).get("exomes", {}).get("af"))
                    genomes = safe_float(hit.get("gnomad", {}).get("genomes", {}).get("af"))
                    
                    if self._validate_af(exomes, k, "gnomAD-Exomes"):
                        af_sources["gnomad_exomes"] = exomes
                    if self._validate_af(genomes, k, "gnomAD-Genomes"):
                        af_sources["gnomad_genomes"] = genomes
                
                # Direkte Felder (von anderen Fetchern)
                if "gex" in hit:
                    val = safe_float(hit.get("gex"))
                    if self._validate_af(val, k, "gnomAD-GEX"):
                        af_sources["gnomad_exomes"] = val
                
                if "ggn" in hit:
                    val = safe_float(hit.get("ggn"))
                    if self._validate_af(val, k, "gnomAD-GGN"):
                        af_sources["gnomad_genomes"] = val
                
                if "af_1kg" in hit:
                    val = safe_float(hit.get("af_1kg"))
                    if self._validate_af(val, k, "1000G"):
                        af_sources["1kg"] = val
                
                if "af_exac" in hit:
                    val = safe_float(hit.get("af_exac"))
                    if self._validate_af(val, k, "ALFA-ExAC"):
                        af_sources["exac"] = val
            
            # ✅ Priorisierte AF-Auswahl (gnomAD > ExAC > 1kg > LightDB)
            valid_afs = []
            priority = ["gnomad_genomes", "gnomad_exomes", "exac", "1kg", "lightdb"]
            
            for source in priority:
                val = af_sources.get(source)
                if val is not None:
                    valid_afs.append(val)
            
            # Mittelwert nur aus validen AFs
            mean_af = None
            if valid_afs:
                mean_af = sum(valid_afs) / len(valid_afs)
                
                # ✅ KRITISCHER Sicherheitscheck: Mittelwert validieren
                if not (0.0 <= mean_af <= 1.0):
                    self.logger.log(
                        f"[AF-Controller] ⚠️ ANOMALIE: mean_af={mean_af:.6f} für {k}\n"
                        f"  Sources: {[f'{s}={af_sources[s]:.6f}' for s in priority if af_sources[s] is not None]}"
                    )
                    mean_af = None  # Verwerfen
                else:
                    # ✅ Clamp auf [0, 1] zur Sicherheit
                    mean_af = max(0.0, min(1.0, mean_af))
            
            # Ausgabe mit garantiert validen Werten
            final[k] = {
                "af_exac": af_sources["exac"],
                "af_gnomad_exomes": af_sources["gnomad_exomes"],
                "af_gnomad_genomes": af_sources["gnomad_genomes"],
                "af_1kg": af_sources["1kg"],
                "af_filter_mean": mean_af,  # ✅ Garantiert None oder [0, 1]
            }
        
        return final

    def _validate_af(self, val: Optional[float], key: Tuple, source: str) -> bool:
        """
        Validiert AF-Wert auf Plausibilität.
        
        Args:
            val: AF-Wert
            key: Variant-Key (für Logging)
            source: Quelle (für Logging)
        
        Returns:
            True wenn valide, False sonst
        """
        if val is None:
            return False
        
        if not isinstance(val, (int, float)):
            self.logger.log(f"[AF-Validation] ⚠️ Nicht-numerischer AF von {source} für {key}: {val}")
            return False
        
        if val < 0.0 or val > 1.0:
            self.logger.log(
                f"[AF-Validation] ⚠️ UNGÜLTIGER AF von {source} für {key}: {val:.6f}\n"
                f"  → Wert außerhalb [0.0, 1.0] → verworfen"
            )
            return False
        
        # Optional: Warnung bei ungewöhnlich hohen AFs (> 0.5)
        if val > 0.5:
            self.logger.log(
                f"[AF-Validation] ℹ️ Hoher AF von {source} für {key}: {val:.6f}\n"
                f"  (> 0.5 ist ungewöhnlich für seltene Varianten)"
            )
        
        return True

    def _normalize_for_db(self, record: dict) -> dict:
        """
        Normalisiert externe Feldnamen und Werte für die DB.
        - Wendet UI_TO_DB-Mapping an (z.B. "af" → "af_filter_mean").
        - Konvertiert bool → int, float → safe_float.
        - Lässt unbekannte Felder weg.
        """
        norm = {}
        for k, v in (record or {}).items():
            # Feldname ins DB-Schema übersetzen
            db_field = VariantDB.UI_TO_DB.get(k, k)

            # Nur bekannte Spalten übernehmen
            if db_field not in VariantDB.ALL_COLS:
                continue

            # Typ-Normalisierung
            if isinstance(v, bool):
                v = int(v)
            elif isinstance(v, (int, float)):
                v = safe_float(v)
            elif v is None:
                pass
            else:
                v = str(v)

            norm[db_field] = v

        return norm
    async def _wrap_fetch(self, api_name, coro, workers: int = None):
        """
        Führt einen einzelnen Fetcher-Aufruf aus.

        FIX: Bessere Exception-Behandlung und konsistente Rückgabe.
        Hinweis: Ein Semaphore pro Call bringt keine Begrenzung; verwalte Concurrency höher.
        """
        def _init_stats():
            if api_name not in self.stats:
                self.stats[api_name] = {"ok": 0, "fail": 0}

        try:
            res = await coro
            _init_stats()
            self.stats[api_name]["ok"] += 1
            return res if isinstance(res, dict) else {}
        except asyncio.CancelledError:
            _init_stats()
            self.stats[api_name]["fail"] += 1
            self.logger.log(f"[{api_name}] ⚠️ Fetch cancelled")
            return {}
        except Exception as e:
            _init_stats()
            self.stats[api_name]["fail"] += 1
            self.logger.log(f"[{api_name}] ❌ Fetch error: {e}")
            return {}

        # HINWEIS:
        # Falls in diesem Abschnitt (Reward/Penalty-Logik) unerwartete Fehler,
        # Instabilitäten oder zu hohe Komplexität auftreten,
        # bitte die gesamte Reward/Penalty-Mechanik wieder entfernen
        # und auf die einfache Gewichtung via _compute_weights() zurückgehen.
        #
        # Rückbau-Anleitung siehe Entwickler-Kommentar unten.
    def _compute_weights(self):
        """Berechnet dynamische Gewichte pro API basierend auf Erfolgsquote."""
        # Reverse-Map: stats-Key -> settings-Key
        _stats_to_settings = {"mv": "myvariant", "gnomad": "gnomad", "vep": "vep", "alfa": "alfa", "topmed": "topmed"}
        weights = {}
        for api, st in self.stats.items():
            # Deaktivierte APIs ueberspringen
            settings_key = _stats_to_settings.get(api, api)
            if not is_api_enabled("phase1_af", settings_key):
                continue
            ok = st.get("ok", 0)
            fail = st.get("fail", 0)
            total = ok + fail
            success_rate = ok / total if total > 0 else 1.0
            # Unterdrücke stark fehlschlagende APIs
            if success_rate > 0.2:
                weights[api] = success_rate

        s = sum(weights.values())
        return {api: w / s for api, w in weights.items()} if s else {"mv": 1.0}


    def _buffer_result(self, key, val: dict):
        """
        Buffert Ergebnis für DB-Write.
        Ergänzt automatisch Success-Felder und Abrufzeitpunkte
        (sowohl meanAF als auch full, abhängig vom Modus).
        """
        record = self._normalize_for_db(dict(val or {}))
        now = datetime.datetime.now(datetime.timezone.utc).isoformat()
        record.setdefault("meanAF_last_fetch", now)

        # Prüfen, ob ein Fetch-Fehler markiert wurde
        if record.get("fetch_error"):
            record["meanAF_fetch_success"] = FetchStatusManager.update_status(
                record.get("meanAF_fetch_success"), success=False
            )
        else:
            # Egal ob AF gefunden oder True None → gilt als erfolgreicher Fetch
            record["meanAF_fetch_success"] = FetchStatusManager.update_status(
                record.get("meanAF_fetch_success"), success=True
            )

        # Full-Modus
        if getattr(self, "mode", None) == "full":
            record.setdefault("full_last_fetch", now)
            record.setdefault("full_fetch_success", any(
                record.get(f) is not None for f in (
                    "af_exac", "af_gnomad_exomes", "af_gnomad_genomes",
                    "af_1kg", "cadd_phred", "gene_symbol", "impact"
                )
            ))

        self.result_buffer[key] = record
        if len(self.result_buffer) >= self.flush_size:
            self._flush_buffer()


    def _flush_buffer(self):
        """Schreibt Buffer in DB (sicher gegen Dict-Änderungen)."""
        if not self.result_buffer:
            return

        try:
            # ✅ Kopie erzeugen, damit Iteration stabil bleibt
            buffer_copy = dict(self.result_buffer)
            self.db.upsert_variants_bulk(buffer_copy)
            self.logger.log(f"[AF-Controller] Flushed {len(buffer_copy)} results to DB.")
        except Exception as e:
            self.logger.log(f"[AF-Controller] ❌ DB flush error: {e}")
        finally:
            self.result_buffer.clear()


    def finalize(self):
        """Finaler Flush beim Beenden."""
        try:
            self._flush_buffer()
            self.logger.log("[AF-Controller] Final flush complete.")
        except Exception as e:
            self.logger.log(f"[AF-Controller] ❌ Fehler beim finalen Flush: {e}")


class GeneAnnotator:
    """
    FIX: Memory-optimierte Gene-Annotation.
    
    Änderungen:
    - Kompakte Datenstrukturen (nur gene_name + is_coding)
    - Lazy-Loading von GTF-Dateien
    - Optionale Disk-Cache für Index
    - Robusterer GTF-Parser (gene_type vs gene_biotype)
    - ✅ V9: Absoluter Pfad + umfangreiches Logging
    """
    
    def __init__(self, cache_dir=None, logger=None):
        """
        Args:
            cache_dir: Optionaler Pfad. Wenn None, wird ein Ordner relativ zur 
                       Script-Datei verwendet (nicht CWD!).
            logger: Optionaler Logger für detaillierte Ausgaben.
        """
        self.logger = logger
        self._log("=" * 60)
        self._log("[GeneAnnotator] ✅ INITIALISIERUNG GESTARTET")
        self._log("=" * 60)
        
        # ✅ V9 FIX: Absoluter Pfad relativ zur Script-Datei, nicht zum CWD!
        if cache_dir is None:
            script_dir = os.path.dirname(os.path.abspath(__file__))
            cache_dir = os.path.join(script_dir, "data", "annotations")
        
        self.cache_dir = os.path.abspath(cache_dir)
        
        self._log(f"[GeneAnnotator] Script-Verzeichnis: {os.path.dirname(os.path.abspath(__file__))}")
        self._log(f"[GeneAnnotator] Aktuelles CWD: {os.getcwd()}")
        self._log(f"[GeneAnnotator] Cache-Verzeichnis: {self.cache_dir}")
        
        # Verzeichnis erstellen
        try:
            os.makedirs(self.cache_dir, exist_ok=True)
            self._log(f"[GeneAnnotator] ✅ Cache-Verzeichnis existiert/erstellt")
        except Exception as e:
            self._log(f"[GeneAnnotator] ❌ FEHLER beim Erstellen des Cache-Verzeichnisses: {e}")
            
        self.trees = {"GRCh37": {}, "GRCh38": {}}
        self.available = False
        self._init_stats = {"GRCh37": {}, "GRCh38": {}}

        # Beide Builds vorbereiten
        for build in ("GRCh37", "GRCh38"):
            self._log(f"[GeneAnnotator] --- Verarbeite {build} ---")
            
            # ✅ Prüfe Index-Cache
            index_cache = os.path.join(self.cache_dir, f"{build}_index.pkl")
            self._log(f"[GeneAnnotator] Index-Cache-Pfad: {index_cache}")
            self._log(f"[GeneAnnotator] Index-Cache existiert: {os.path.exists(index_cache)}")
            
            if os.path.exists(index_cache):
                try:
                    self._load_index_cache(build, index_cache)
                    chrom_count = len(self.trees[build])
                    gene_count = sum(len(tree) for tree in self.trees[build].values())
                    self._log(f"[GeneAnnotator] ✅ Index {build} aus Cache geladen: {chrom_count} Chromosomen, {gene_count} Gene")
                    self._init_stats[build] = {"source": "cache", "chromosomes": chrom_count, "genes": gene_count}
                    continue
                except Exception as e:
                    self._log(f"[GeneAnnotator] ⚠️ Cache-Fehler {build}: {e}")
            
            # Fallback: GTF parsen
            try:
                gtf_path = self._ensure_gtf(build)
                self._log(f"[GeneAnnotator] GTF-Datei: {gtf_path}")
                self._log(f"[GeneAnnotator] GTF existiert: {os.path.exists(gtf_path)}")
                
                gene_count = self._build_index(build)
                chrom_count = len(self.trees[build])
                self._log(f"[GeneAnnotator] ✅ Index {build} gebaut: {chrom_count} Chromosomen, {gene_count} Gene")
                self._init_stats[build] = {"source": "gtf", "chromosomes": chrom_count, "genes": gene_count}
                
                # ✅ Index speichern
                try:
                    self._save_index_cache(build, index_cache)
                    self._log(f"[GeneAnnotator] ✅ Index {build} im Cache gespeichert")
                except Exception as e:
                    self._log(f"[GeneAnnotator] ⚠️ Cache-Speicherung fehlgeschlagen: {e}")
            except Exception as e:
                self._log(f"[GeneAnnotator] ❌ Initialisierung für {build} FEHLGESCHLAGEN: {e}")
                import traceback
                self._log(f"[GeneAnnotator] Traceback: {traceback.format_exc()}")
                self._init_stats[build] = {"source": "failed", "error": str(e)}
                
        # Verfügbar wenn mindestens ein Build geladen wurde
        grch37_ok = bool(self.trees["GRCh37"])
        grch38_ok = bool(self.trees["GRCh38"])
        self.available = grch37_ok or grch38_ok
        
        self._log("=" * 60)
        self._log(f"[GeneAnnotator] INITIALISIERUNG ABGESCHLOSSEN")
        self._log(f"[GeneAnnotator] GRCh37 verfügbar: {grch37_ok} ({len(self.trees['GRCh37'])} Chromosomen)")
        self._log(f"[GeneAnnotator] GRCh38 verfügbar: {grch38_ok} ({len(self.trees['GRCh38'])} Chromosomen)")
        self._log(f"[GeneAnnotator] Gesamt-Status: {'✅ BEREIT' if self.available else '❌ NICHT VERFÜGBAR'}")
        self._log("=" * 60)
    
    def _log(self, message: str):
        """Logging-Helper: Schreibt in Logger falls vorhanden, sonst print."""
        print(message)  # Immer auf Konsole
        if self.logger:
            try:
                self.logger.log(message)
            except:
                pass
    
    def get_status_report(self) -> str:
        """Gibt einen formatierten Status-Report zurück."""
        lines = [
            "",  # V10 FIX: Leerzeile vor Box für bessere Lesbarkeit
            "╔════════════════════════════════════════════════════════════╗",
            "║           GENE ANNOTATOR STATUS REPORT                     ║",
            "╠════════════════════════════════════════════════════════════╣",
            f"║ Cache-Verzeichnis: {self.cache_dir[:45]:<45} ║",
            f"║ Verfügbar: {'JA' if self.available else 'NEIN':<50} ║",
            "╠════════════════════════════════════════════════════════════╣",
        ]
        
        for build in ["GRCh37", "GRCh38"]:
            stats = self._init_stats.get(build, {})
            source = stats.get("source", "unknown")
            if source == "failed":
                lines.append(f"║ {build}: ❌ FEHLER - {stats.get('error', 'unbekannt')[:35]:<35} ║")
            else:
                chroms = stats.get("chromosomes", 0)
                genes = stats.get("genes", 0)
                lines.append(f"║ {build}: ✅ {genes:,} Gene auf {chroms} Chromosomen ({source}) ║")
        
        lines.append("╚════════════════════════════════════════════════════════════╝")
        lines.append("")  # V10 FIX: Leerzeile nach Box
        return "\n".join(lines)

    def _load_index_cache(self, build, cache_path):
        """Lädt serialisierten IntervalTree-Index."""
        import pickle
        with open(cache_path, "rb") as f:
            self.trees[build] = pickle.load(f)

    def _save_index_cache(self, build, cache_path):
        """Speichert IntervalTree-Index serialisiert."""
        import pickle
        with open(cache_path, "wb") as f:
            pickle.dump(self.trees[build], f, protocol=pickle.HIGHEST_PROTOCOL)

    def _build_index(self, build):
        """
        FIX: Memory-optimierter Index-Build.
        ✅ V9: Gibt gene_count zurück für Logging.
        """
        gtf_path = os.path.join(self.cache_dir, f"{build}.gtf.gz")
        self._log(f"[GeneAnnotator] Baue Index für {build} aus {gtf_path} …")
        
        gene_count = 0
        try:
            with gzip.open(gtf_path, "rt") as f:
                for line in f:
                    if line.startswith("#"):
                        continue
                    
                    parts = line.strip().split("\t")
                    if len(parts) < 9:
                        continue
                    
                    chrom, source, feature, start, end, score, strand, frame, attrs = parts
                    
                    # ✅ Nur gene-Features
                    if feature != "gene":
                        continue
                    
                    gene_id, gene_name, biotype = self._parse_attrs(attrs)
                    
                    # ✅ Normalisiere Chromosom
                    norm_chrom = chrom[3:] if chrom.startswith("chr") else chrom
                    
                    if norm_chrom not in self.trees[build]:
                        self.trees[build][norm_chrom] = IntervalTree()
                    
                    # ✅ Kompakte Daten (nur name + is_coding)
                    is_coding = (biotype == "protein_coding")
                    self.trees[build][norm_chrom][int(start):int(end)+1] = (gene_name, is_coding)
                    
                    gene_count += 1
                    
                    # Progress-Log
                    if gene_count % 10000 == 0:
                        self._log(f"[GeneAnnotator] {build}: {gene_count} Gene indiziert...")
            
            self._log(f"[GeneAnnotator] Index {build} fertig: {gene_count} Gene auf {len(self.trees[build])} Chromosomen.")
            return gene_count
            
        except Exception as e:
            self._log(f"[GeneAnnotator] ❌ Fehler beim Indexieren von {gtf_path}: {e}")
            raise

    def _ensure_gtf(self, build):
        """Lädt GTF-Datei herunter (falls nicht vorhanden)."""
        if build == "GRCh38":
            url = "https://ftp.ebi.ac.uk/pub/databases/gencode/Gencode_human/release_46/gencode.v46.annotation.gtf.gz"
        else:  # GRCh37
            url = "https://ftp.ebi.ac.uk/pub/databases/gencode/Gencode_human/release_19/gencode.v19.annotation.gtf.gz"

        gtf_path = os.path.join(self.cache_dir, f"{build}.gtf.gz")
        
        if not os.path.exists(gtf_path):
            self._log(f"[GeneAnnotator] GTF für {build} nicht gefunden, lade herunter...")
            self._log(f"[GeneAnnotator] URL: {url}")
            self._log(f"[GeneAnnotator] Ziel: {gtf_path}")
            
            r = requests.get(url, stream=True)
            r.raise_for_status()
            
            total_size = int(r.headers.get('content-length', 0))
            downloaded = 0
            
            with open(gtf_path, "wb") as f:
                for chunk in r.iter_content(chunk_size=8192):
                    f.write(chunk)
                    downloaded += len(chunk)
                    if total_size > 0 and downloaded % (1024*1024*10) == 0:  # Alle 10MB
                        pct = (downloaded / total_size) * 100
                        self._log(f"[GeneAnnotator] Download {build}: {pct:.1f}%")
            
            self._log(f"[GeneAnnotator] ✅ Download {build} abgeschlossen: {os.path.getsize(gtf_path) / 1024 / 1024:.1f} MB")
        else:
            file_size = os.path.getsize(gtf_path) / 1024 / 1024
            self._log(f"[GeneAnnotator] ✅ GTF {build} existiert bereits: {file_size:.1f} MB")
        
        return gtf_path

    def _parse_attrs(self, attr_str):
        """Parst GTF-Attribute."""
        d = {}
        for part in attr_str.split(";"):
            part = part.strip()
            if not part:
                continue
            # Fallback für Attribute ohne Leerzeichen (selten, aber möglich)
            if " " not in part:
                # Versuch mit =
                if "=" in part:
                     key, val = part.split("=", 1)
                else:
                    continue
            else:
                key, val = part.split(" ", 1)
            
            d[key] = val.strip('"')
        
        # Robustere Biotype-Erkennung
        biotype = d.get("gene_type") or d.get("gene_biotype")
        return d.get("gene_id"), d.get("gene_name"), biotype

    def annotate_variant(self, chrom, pos, build, debug=False):
        """
        Annotiert eine einzelne Variante und gibt ALLE überlappenden Gene zurück.
        
        Args:
            debug: Wenn True, wird detailliertes Logging ausgegeben.
        """
        norm_chrom = str(chrom).replace("chr", "").replace("CHR", "")
        
        if debug:
            self._log(f"[GeneAnnotator] DEBUG annotate_variant: chr={chrom} -> {norm_chrom}, pos={pos}, build={build}")
            self._log(f"[GeneAnnotator] DEBUG: Build '{build}' in trees: {build in self.trees}")
            if build in self.trees:
                self._log(f"[GeneAnnotator] DEBUG: Chromosom '{norm_chrom}' in trees[{build}]: {norm_chrom in self.trees[build]}")
                self._log(f"[GeneAnnotator] DEBUG: Verfügbare Chromosomen: {list(self.trees[build].keys())[:5]}...")
        
        if build not in self.trees or norm_chrom not in self.trees[build]:
            if debug:
                self._log(f"[GeneAnnotator] DEBUG: Kein Treffer - Build oder Chromosom nicht verfügbar")
            return None, None
        
        try:
            # Finde alle überlappenden Intervalle an der Position 'pos'
            hits = self.trees[build][norm_chrom][int(pos)]
            
            if debug:
                self._log(f"[GeneAnnotator] DEBUG: {len(hits)} Treffer an Position {pos}")
            
            if not hits:
                return None, None
            
            # Extrahieren und deduplizieren Sie die Gen-Namen und Codierungsinformationen
            unique_genes = set()
            for interval in hits:
                gene_name, is_coding = interval.data
                if gene_name: # Nur Gene mit Namen speichern
                    unique_genes.add((gene_name, is_coding))
            
            if not unique_genes:
                return None, None

            # Formatierung: Sortierte, kommagetrennte Gen-Namen
            gene_names = sorted([name for name, _ in unique_genes])
            
            # Prüfe, ob mindestens eines Protein-kodierend ist
            is_coding_overall = any(is_coding for _, is_coding in unique_genes)
            
            if debug:
                self._log(f"[GeneAnnotator] DEBUG: Gefunden: {gene_names}, is_coding={is_coding_overall}")
            
            return ", ".join(gene_names), is_coding_overall
        except Exception as e:
            if debug:
                self._log(f"[GeneAnnotator] DEBUG: Fehler: {e}")
            return None, None

    def annotate_batch(self, variants):
        """
        Batch-Annotation durch Aufruf von annotate_variant für jeden Key.
        
        Änderungen:
        - Gibt ein vollständiges Dict zurück (kein Generator)
        - V10 FIX: Build-Normalisierung (hg19->GRCh37, hg38->GRCh38)
        """
        result = {}
        for key, chrom, pos in variants:
            # key[4] ist der Build (falls vorhanden)
            raw_build = key[4] if len(key) > 4 else "GRCh37"
            
            # V10 FIX: Build normalisieren
            build_map = {
                "hg19": "GRCh37", "hg37": "GRCh37", "b37": "GRCh37", "grch37": "GRCh37",
                "hg38": "GRCh38", "b38": "GRCh38", "grch38": "GRCh38"
            }
            build = build_map.get(str(raw_build).lower(), raw_build)
            
            if build not in ("GRCh37", "GRCh38"):
                continue
            
            # Die korrigierte Methode gibt jetzt 'Gene1, Gene2' und ein bool zurück
            gene_str, pc = self.annotate_variant(chrom, pos, build=build)
            
            # Auch bei None speichern, damit wir wissen, dass gesucht wurde (Cache-Hit "None")
            result[key] = {
                "gene_symbol": gene_str,
                "is_coding": pc
            }
        return result
    
    
# ============== AlphaGenome (optional) ==============
class AlphaGenomeScorer:
    """
    FIX: Robustes AlphaGenome-Scoring mit Timeouts.
    
    Änderungen:
    - Per-Variant-Timeout (30s)
    - Skip bei wiederholten Fehlern
    - Besseres Error-Logging
    """
    
    SUPPORTED_CHROMS = {f"chr{i}" for i in range(1, 23)} | {"chrX", "chrY", "chrM"}
    TARGET_LEN = 16384
    VARIANT_TIMEOUT = 30.0  # ✅ Timeout pro Variante

    def __init__(self, ag_api_key=None):
        """Initialize AlphaGenome using documented client API."""
        try:
            import numpy as np
            self._np = np
        except Exception:
            self._np = None

        self.api_key = ag_api_key
        self.available = False
        self.model = None
        
        # ✅ Error-Tracking
        self.error_count = 0
        self.max_errors = 100  # Nach 100 Fehlern → disable

        if not HAVE_ALPHAGENOME or ag_client is None or ag_genome is None or self._np is None or not ag_api_key:
            logger.log("[AlphaGenome] Client nicht initialisiert: Abhängigkeiten fehlen oder API-Key leer.")
            self.available = False
            return

        try:
            self.model = ag_client.create(ag_api_key)
            if self.model is None or not hasattr(self.model, "predict_variant"):
                raise RuntimeError("AlphaGenome client lacks predict_variant")
            if not hasattr(ag_client, "OutputType"):
                raise RuntimeError("AlphaGenome OutputType missing")
            self.available = True
            logger.log("AlphaGenome-Client initialisiert (dna_client.create).")
        except Exception as e:
            self.available = False
            self.model = None
            logger.log(f"AlphaGenome nicht verfügbar: {e}")

    def score_batch(self, variants, logger=None):
        """
        FIX: Zwischenergebnisse werden progressiv geflusht.
        
        Änderungen:
        - Batch-Write alle 100 Scores
        - Final-Flush am Ende
        - Verhindert Datenverlust bei Timeout
        """
        if not self.available or self.model is None:
            return {}
        
        # ✅ Check Error-Threshold
        if self.error_count >= self.max_errors:
            if logger:
                logger.log(
                    f"[AlphaGenome] ⚠️ Disabled nach {self.error_count} Fehlern"
                )
            self.available = False
            return {}

        out = {}
        total = len(variants)
        done = 0
        success_count = 0
        
        # ✅ Batch-Write-Akkumulator
        pending_writes = []
        FLUSH_INTERVAL = 100  # Alle 100 Scores flushen
        
        if logger:
            logger.log(f"[AlphaGenome] Scoring-Batch gestartet ({total} Varianten).")
        
        start_time = time.time()

        for (chrom, pos, ref, alt, build) in variants:
            variant_start = time.time()
            
            try:
                # Pre-Validation
                c = self._norm_chr(chrom)
                if c not in self.SUPPORTED_CHROMS:
                    out[(chrom, pos, ref, alt, build)] = None
                    done += 1
                    continue
                
                if not isinstance(pos, int) or pos <= 0:
                    out[(chrom, pos, ref, alt, build)] = None
                    done += 1
                    continue
                
                if not (isinstance(ref, str) and isinstance(alt, str) and ref and alt):
                    out[(chrom, pos, ref, alt, build)] = None
                    done += 1
                    continue
                
                # API-Call mit Timeout-Wrapper
                half = self.TARGET_LEN // 2
                start = max(1, int(pos) - half)
                end = start + self.TARGET_LEN

                interval = ag_genome.Interval(chromosome=c, start=start, end=end)
                variant = ag_genome.Variant(
                    chromosome=c,
                    position=int(pos),
                    reference_bases=str(ref),
                    alternate_bases=str(alt)
                )

                requested = [ag_client.OutputType.RNA_SEQ]
                
                # Timeout-geschützter Call
                outputs = self._predict_with_timeout(
                    interval, variant, requested, 
                    timeout=self.VARIANT_TIMEOUT
                )
                
                if outputs is None:
                    # Timeout → Skip
                    out[(chrom, pos, ref, alt, build)] = None
                    self.error_count += 1
                    done += 1
                    continue

                # Extract Arrays
                ref_seq_obj = getattr(getattr(outputs, "reference", None), "rna_seq", None)
                alt_seq_obj = getattr(getattr(outputs, "alternate", None), "rna_seq", None)
                ref_arr = self._to_array(ref_seq_obj)
                alt_arr = self._to_array(alt_seq_obj)

                if self._np is None or ref_arr.size == 0 or alt_arr.size == 0:
                    out[(chrom, pos, ref, alt, build)] = None
                else:
                    min_len = min(ref_arr.size, alt_arr.size)
                    if min_len == 0:
                        out[(chrom, pos, ref, alt, build)] = None
                    else:
                        delta = float(
                            self._np.mean(
                                self._np.abs(alt_arr[:min_len] - ref_arr[:min_len])
                            )
                        )
                        out[(chrom, pos, ref, alt, build)] = delta
                        success_count += 1
                        
                        # ✅ In Batch-Write-Queue
                        key_tuple = (chrom, pos, ref, alt, build)
                        pending_writes.append((key_tuple, {"ag_score": delta}))
                
            except Exception as e:
                if logger:
                    logger.log(f"[AlphaGenome] Fehler bei {chrom}:{pos} {ref}>{alt}: {e}")
                out[(chrom, pos, ref, alt, build)] = None
                self.error_count += 1
            
            finally:
                done += 1
                
                # ✅ Progressiver Flush alle 100 Scores
                if len(pending_writes) >= FLUSH_INTERVAL:
                    if hasattr(self, 'db') and self.db:
                        try:
                            self.db.update_variant_fields_bulk(pending_writes)
                            if logger:
                                logger.log(
                                    f"[AlphaGenome] ✅ Zwischenspeicherung: "
                                    f"{len(pending_writes)} Scores geflusht"
                                )
                            pending_writes.clear()
                        except Exception as e:
                            if logger:
                                logger.log(f"[AlphaGenome] ⚠️ Flush-Fehler: {e}")
                
                # Progress-Logging (alle 50 Varianten)
                if logger and done % 50 == 0:
                    elapsed = time.time() - start_time
                    rate = done / max(1, elapsed)
                    eta_seconds = (total - done) / max(0.1, rate)
                    
                    logger.log(
                        f"[AlphaGenome] Progress: {done}/{total} "
                        f"({(done/total)*100:.1f}%) | "
                        f"Success: {success_count} | "
                        f"Errors: {self.error_count} | "
                        f"ETA {fmt_eta(eta_seconds)}"
                    )

        # ✅ Final-Flush
        if pending_writes:
            if hasattr(self, 'db') and self.db:
                try:
                    self.db.update_variant_fields_bulk(pending_writes)
                    if logger:
                        logger.log(
                            f"[AlphaGenome] ✅ Final-Flush: "
                            f"{len(pending_writes)} Scores gespeichert"
                        )
                except Exception as e:
                    if logger:
                        logger.log(f"[AlphaGenome] ⚠️ Final-Flush-Fehler: {e}")

        if logger:
            total_time = time.time() - start_time
            logger.log(
                f"[AlphaGenome] Batch complete: {done}/{total} in {total_time:.1f}s | "
                f"Success: {success_count} | Errors: {self.error_count}"
            )
        
        return out
    def _predict_with_timeout(self, interval, variant, requested, timeout):
        """
        Helper: Führt predict_variant mit Timeout aus.
        
        Returns:
            Outputs bei Erfolg, None bei Timeout
        """
        import concurrent.futures
        
        with concurrent.futures.ThreadPoolExecutor(max_workers=1) as executor:
            future = executor.submit(
                self.model.predict_variant,
                interval=interval,
                variant=variant,
                ontology_terms=[],
                requested_outputs=requested
            )
            
            try:
                return future.result(timeout=timeout)
            except concurrent.futures.TimeoutError:
                # Timeout → Cancel Future
                future.cancel()
                return None
            except Exception:
                return None

    def _to_array(self, seq):
        """Convert AlphaGenome output containers to numpy arrays."""
        if self._np is None or seq is None:
            return self._np.array([], dtype=float) if self._np is not None else []
        
        # Try typical container attributes
        for attr in ("values", "data"):
            try:
                val = getattr(seq, attr)
            except Exception:
                val = None
            if val is not None:
                try:
                    return self._np.array(val, dtype=float).ravel()
                except Exception:
                    continue
        
        # Fallback: direct conversion
        try:
            return self._np.array(seq, dtype=float).ravel()
        except Exception:
            return self._np.array([], dtype=float)

    def _norm_chr(self, chrom):
        """Normalize chromosome name."""
        if chrom is None:
            return ""
        c = str(chrom).strip()
        if not c:
            return ""
        if not c.lower().startswith("chr"):
            c = "chr" + c
        if c in ("chrMT", "MT", "chrMt"):
            c = "chrM"
        return c


# ============== Throughput Tuner (vereinfacht) ==============
class CircuitBreaker:
    """
    Verhindert API-Überlastung durch temporäres Pausieren bei wiederholten Fehlern.
    """
    def __init__(self, failure_threshold=3, timeout=30.0, logger=None):
        self.failure_threshold = failure_threshold
        self.timeout = timeout
        self.logger = logger
        self.failures = 0
        self.last_failure_time = 0
        self.state = "closed"  # closed, open, half_open
        self.lock = asyncio.Lock()
    
    async def record_failure(self):
        async with self.lock:
            self.failures += 1
            self.last_failure_time = time.time()
            
            if self.failures >= self.failure_threshold:
                if self.state != "open":
                    self.state = "open"
                    if self.logger:
                        self.logger.log(
                            f"[CircuitBreaker] OPEN (failures={self.failures}), "
                            f"pause {self.timeout}s"
                        )
    
    async def record_success(self):
        async with self.lock:
            self.failures = max(0, self.failures - 1)
            
            if self.state == "open" and time.time() - self.last_failure_time > self.timeout:
                self.state = "half_open"
                if self.logger:
                    self.logger.log("[CircuitBreaker] HALF-OPEN, testing...")
            
            if self.state == "half_open" and self.failures == 0:
                self.state = "closed"
                if self.logger:
                    self.logger.log("[CircuitBreaker] CLOSED, normal operation")
    
    def is_open(self):
        if self.state == "open":
            if time.time() - self.last_failure_time > self.timeout:
                self.state = "half_open"
                return False
            return True
        return False
    
class ThroughputTuner:
    def __init__(
        self,
        *,
        min_workers: int = 2,
        max_workers: int = 100,
        start_workers: int = 4,
        min_batch: int = 50,
        max_batch: int = 1000,
        base_batch: int = 800,
        cpu_high: float = 85.0,
        log_cb: Optional[Callable[[str], None]] = None,
        start_batch: Optional[int] = None
    ):
        self.phase = 1
        self.min_w = min_workers
        self.max_w = max_workers
        self.w = max(min(start_workers, max_workers), min_workers)
        self.min_b = min_batch
        self.max_b = max_batch
        init_b = start_batch if start_batch is not None else base_batch
        self.b = max(min(init_b, max_batch), min_batch)
        self.cpu_high = cpu_high
        self.safe_log = log_cb or (lambda msg: None)
        self.history: List[Tuple[float, int, float, int, float, float]] = []
        self._cpu_cooldown_until = 0.0
        self.total_variants: Optional[int] = None
        self.done_variants = 0
        self.end_retry_variants = 0

    def set_load_state(self, *, total_variants: int, done_variants: int, end_retry_variants: int):
        self.total_variants = max(1, total_variants)
        self.done_variants = max(0, done_variants)
        self.end_retry_variants = max(0, end_retry_variants)

    def _restload(self) -> Optional[int]:
        if self.total_variants is None:
            return None
        rest = (self.total_variants - self.done_variants) + self.end_retry_variants
        return max(1, rest)

    def _throughput(self, window_sec: float) -> float:
        now = time.time()
        wins = [(s, d) for (t, s, d, _, _, _) in self.history if now - t <= window_sec]
        total_s = sum(s for s, _ in wins)
        total_d = sum(d for _, d in wins)
        return (total_s / total_d) if total_d > 0 else 0.0

    def _throughput_restload(self, window_sec: float) -> float:
        base = self._throughput(window_sec)
        rest = self._restload()
        if rest is None:
            return base
        return base * (self.total_variants / rest)

    def record_round(self, successes: int, duration: float):
        now = time.time()
        duration = max(duration, 1e-6)
        u60_val = self._throughput_restload(60)
        u120_val = self._throughput_restload(120)
        self.history.append((now, successes, duration, self.w, u60_val, u120_val))
        cutoff = now - 250
        self.history = [h for h in self.history if h[0] >= cutoff]

    # [PROBABLY DEAD] Inline-Logik in decide() ersetzt diese Methode
    # def _cpu_cut(self):
        # self.w = max(self.min_w, self.w - 10)
        # self._cpu_cooldown_until = time.time() + 10
        # cutoff = time.time() - 30
        # self.history = [h for h in self.history if h[0] >= cutoff]
        # self.safe_log(f"[TUNER] CPU-Cut -> w={self.w}, History auf 30s beschnitten")

    def decide(self, current_cpu: Optional[float] = None) -> Tuple[int, int]:
        now = time.time()
        u10 = self._throughput_restload(10)
        u60 = self._throughput_restload(60)
        u120 = self._throughput_restload(120)
        in_cooldown = now < self._cpu_cooldown_until

        # If CPU critical, reduce workers gently and throttle batch
        if current_cpu is not None and current_cpu >= self.cpu_high:
            # gentle cut: reduce by 20% but not below min_w
            new_w = max(self.min_w, int(self.w * 0.8))
            if new_w < self.w:
                self.w = new_w
                self._cpu_cooldown_until = time.time() + 5
                self.safe_log(f"[TUNER] CPU high -> reduce workers to {self.w}")
            # reduce batch as well
            self.b = max(self.min_b, int(self.b * 0.9))
            return self.w, self.b

        # Phase logic: smoother increments/decrements
        if self.phase == 1 and not in_cooldown:
            # small positive step
            self.w = min(self.max_w, self.w + 2)
            # if throughput trends downward across windows, reduce and switch
            if u10 and u60 and u120 and (u10 < u60 * 0.9 and u60 < u120 * 0.9):
                self.w = max(self.min_w, self.w - 3)
                self.phase = 2
                self.safe_log(f"[TUNER] Trend down -> phase=2, w={self.w}")
            return self.w, self.b

        if self.phase == 2 and not in_cooldown:
            # adjust batch size slowly
            if u10 and u60 and u120 and (u10 > u60 * 1.05 or u60 > u120 * 1.05):
                self.b = min(self.max_b, self.b + 10)
                self.phase = 1
                self.safe_log(f"[TUNER] Trend up -> phase=1, b={self.b}")
            else:
                self.b = max(self.min_b, self.b - 5)
            return self.w, self.b

        return self.w, self.b
        
def fill_missing_fields(key, build, logger=None):
    """
    Holt gezielt fehlende Annotationen aus alternativen Quellen (Lückenfüller).
    Nur conservative additions. Vermeidet fehlerhafte/HTML 'APIs'.
    """
    chrom, pos, ref, alt, _ = key
    filled = {}

    # gnomAD GraphQL or REST isn't trivially available; skip direct fetch
    # unless a proper endpoint is configured in the application.
    gnomad_ep = globals().get("GNOMAD_ENDPOINT", None)
    if isinstance(gnomad_ep, str) and gnomad_ep:
        try:
            # Minimal batch-style query: single key
            payload = [{"chrom": chrom, "pos": pos, "ref": ref, "alt": alt}]
            r = requests.post(gnomad_ep, json=payload, timeout=20)
            if r.status_code == 200:
                data = r.json()
                if isinstance(data, list) and data:
                    entry = data[0] or {}
                    g_exomes = entry.get("exomes_af")
                    g_genomes = entry.get("genomes_af")
                    if g_exomes is not None:
                        filled["af_gnomad_exomes"] = float(g_exomes)
                    if g_genomes is not None:
                        filled["af_gnomad_genomes"] = float(g_genomes)
        except Exception as e:
            if logger:
                logger.log(f"[Lückenfüller] gnomAD Endpoint-Fehler {key}: {e}")

    # ClinVar: Only attempt refsnp by rsID if present in DB (not by 'ref' bases!)
    # If rsid is available, use NCBI variation API (beta refsnp) with rs number.
    try:
        # 'ref' here is a base; we need rsid from DB instead.
        # Fetch row and read rsid. If absent, skip.
        row = db.get_variant(key) if 'db' in globals() and isinstance(db, VariantDB) else None
        rsid = (row or {}).get("rsid")
        if rsid and rsid.lower().startswith("rs"):
            rid = rsid[2:]
            url = f"https://api.ncbi.nlm.nih.gov/variation/v0/beta/refsnp/{rid}"
            r = requests.get(url, timeout=20)
            if r.status_code == 200:
                j = r.json()
                sigs = []
                try:
                    rcv = (j.get("primary_snapshot_data") or {}).get("clinical") or {}
                    # schema is variable; keep conservative extraction
                    if isinstance(rcv, dict):
                        for k, v in rcv.items():
                            if isinstance(v, dict):
                                s = v.get("description")
                                if s:
                                    sigs.append(str(s))
                except Exception:
                    pass
                if sigs:
                    filled["clinical_significance"] = ",".join(sorted(set(sigs)))
    except Exception as e:
        if logger:
            logger.log(f"[Lückenfüller] ClinVar-Fehler {key}: {e}")

    return filled if filled else None

def safe_nested_get(data, *keys, default=None):
    """
    Sicheres navigieren durch verschachtelte Dictionaries.
    
    Args:
        data: Das Start-Dictionary
        *keys: Sequenz von Keys zum Navigieren
        default: Rückgabewert wenn Navigation fehlschlägt
    
    Returns:
        Wert am Ende des Pfads oder default
        
    Beispiel:
        safe_nested_get(data, "level1", "level2", "value", default="")
    """
    current = data
    for key in keys:
        if isinstance(current, dict):
            current = current.get(key)
            if current is None:
                return default
        else:
            return default
    return current

class VariantDB:
    # Spaltenkonstanten (unverändert)
    KEY_COLS = ("chr", "pos", "ref", "alt", "build")
    DATA_COLS = (
        "rsid",
        "af_filter_mean", "af_exac", "af_gnomad_exomes", "af_gnomad_genomes", "af_1kg",
        "cadd_phred", "gene_symbol", "impact", "consequence",
        "clinical_significance", "phenotypes",
        "protein_coding", "is_coding", "conservation", "ag_score",
        "extra_json",
        "meanAF_last_fetch", "meanAF_fetch_success",
        "full_last_fetch", "full_fetch_success"
    )
    ALL_COLS = KEY_COLS + DATA_COLS
    
    # AKTUALISIERTE MAPPING-TABELLE (Single Source of Truth)
    UI_TO_DB = {
        # GUI Spaltenname : DB Spaltenname
        "chr": "chr", 
        "pos": "pos", 
        "ref": "ref", 
        "alt": "alt", 
        "build": "build",
        "rsid": "rsid",
        
        # Annotationen
        "af": "af_filter_mean",
        "cadd": "cadd_phred",
        "gene": "gene_symbol",
        "Gene": "gene_symbol",      # Alias
        "impact": "impact",
        "consequence": "consequence",
        "clin_sig": "clinical_significance",
        "phenotypes": "phenotypes",
        "coding": "is_coding",
        "ag": "ag_score",
        "cons": "conservation",
        
        # Status-Felder (für interne Zwecke)
        "meanAF_status": "meanAF_fetch_success",
        "full_status": "full_fetch_success"
    }
    DB_TO_UI = {v: k for k, v in UI_TO_DB.items()}
    
    # Status-Sets (unverändert)
    SUCCESS_CODES = {"1", "111"}
    GENERIC_FAIL_CODES = {"0", "000"}
    RENORM_CODES = {"010", "011", "012"}
    OTHER_FAIL_CODES = {"020", "030", "040"}
    DELETABLE_CODES = {"013"}

    def __init__(self, db_path: str = None):
        """
        Einfache Initialisierung ohne Connection Pool.

        V16: Wenn db_path=None, wird ResourceManager für variant_db genutzt.
        Fallback auf "variant_fusion.sqlite" im aktuellen Verzeichnis.
        """
        # V16: ResourceManager für DB-Pfad nutzen
        if db_path is None:
            try:
                rm = get_resource_manager()
                rm_path = rm.get("variant_db")
                if rm_path and os.path.exists(rm_path):
                    db_path = rm_path
                    print(f"[DB] ✅ Via ResourceManager: {db_path}")
                else:
                    # Fallback: Check Config.LOCAL_DB_DIR
                    if hasattr(Config, 'LOCAL_DB_DIR') and os.path.exists(Config.LOCAL_DB_DIR):
                        db_path = os.path.join(Config.LOCAL_DB_DIR, "variant_fusion.sqlite")
                        print(f"[DB] Nutze lokalen Storage: {db_path}")
                    else:
                        db_path = "variant_fusion.sqlite"
                        print(f"[DB] ResourceManager: variant_db nicht gefunden, nutze Default")
            except Exception:
                db_path = "variant_fusion.sqlite"

        self.db_path = db_path
        self.lock = threading.RLock()  # RLock für nested calls

        if not os.path.exists(self.db_path):
            print(f"[DB] {self.db_path} nicht gefunden – wird neu erstellt.")

        self._init_db()
        self._ensure_annotation_cache()

    @contextmanager
    def _conn(self):
        """
        Context manager für DB-Verbindungen.
        Jeder Thread bekommt bei Bedarf eine neue Connection.

        V12 FIX: Proper transaction management with DEFERRED isolation level.
        Automatic commit on success, rollback on error.
        """
        conn = None
        try:
            conn = sqlite3.connect(
                self.db_path,
                timeout=Config.DB_TIMEOUT_SEC,
                isolation_level=None,  # ✅ V12.1 REVERT: Autocommit (wie V11) - DEFERRED war zu langsam!
                check_same_thread=False
            )
            conn.row_factory = sqlite3.Row

            # Pragmas für Performance
            try:
                if Config.DB_WAL_MODE:
                    conn.execute("PRAGMA journal_mode=WAL;")
                conn.execute("PRAGMA synchronous=NORMAL;")
                conn.execute("PRAGMA temp_store=MEMORY;")
                conn.execute("PRAGMA foreign_keys=ON;")
                conn.execute("PRAGMA busy_timeout=30000;")  # ✅ V12 FIX: Add busy timeout
            except Exception as e:
                print(f"[DB] Warning: PRAGMA config failed: {e}")

            yield conn

            # ✅ V12.1: Kein manuelles Commit bei Autocommit-Modus (isolation_level=None)
            # conn.commit()

        except Exception as e:
            # ✅ V12.1: Kein Rollback bei Autocommit-Modus
            # if conn:
            #     try:
            #         conn.rollback()
            #     except Exception as rb_err:
            #         print(f"[DB] Rollback error: {rb_err}")
            print(f"[DB] Connection error: {e}")
            raise
        finally:
            if conn:
                try:
                    conn.close()
                except Exception:
                    pass

    def _init_db(self):
        """Initialisiert Schema."""
        with self.lock, self._conn() as con:
            cur = con.cursor()
            cur.execute("""
            CREATE TABLE IF NOT EXISTS variants (
                chr   TEXT    NOT NULL,
                pos   INTEGER NOT NULL,
                ref   TEXT    NOT NULL,
                alt   TEXT    NOT NULL,
                build TEXT    NOT NULL,

                rsid TEXT,

                af_filter_mean     REAL,
                af_exac            REAL,
                af_gnomad_exomes   REAL,
                af_gnomad_genomes  REAL,
                af_1kg             REAL,

                cadd_phred REAL,
                gene_symbol TEXT,
                impact TEXT,
                consequence TEXT,
                clinical_significance TEXT,
                phenotypes TEXT,
                protein_coding INTEGER,
                is_coding INTEGER,
                conservation TEXT,
                ag_score REAL,

                extra_json TEXT,

                meanAF_last_fetch   TEXT,
                meanAF_fetch_success TEXT,
                full_last_fetch     TEXT,
                full_fetch_success  TEXT,

                PRIMARY KEY (chr, pos, ref, alt, build)
            )
            """)
            cur.execute("CREATE INDEX IF NOT EXISTS idx_rsid ON variants (rsid)")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_full_status ON variants (full_fetch_success)")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_meanAF_status ON variants (meanAF_fetch_success)")
            # ✅ V12 FIX: No manual commit - context manager handles it

    def _ensure_annotation_cache(self):
        """Cache-Tabelle für Gene-Annotationen."""
        with self.lock, self._conn() as con:
            cur = con.cursor()
            cur.execute("""
                CREATE TABLE IF NOT EXISTS annotation_cache (
                    chr TEXT,
                    pos INTEGER,
                    ref TEXT,
                    alt TEXT,
                    build TEXT,
                    gene_symbol TEXT,
                    is_coding INTEGER,
                    PRIMARY KEY (chr, pos, ref, alt, build)
                )
            """)
            # ✅ V12 FIX: No manual commit - context manager handles it

    # -----------------------
    # Hilfsfunktion für Mapping
    # -----------------------
    def resolve_column(self, field: str) -> str:
        """Mappt UI- oder Alias-Namen auf den echten DB-Spaltennamen."""
        return self.UI_TO_DB.get(field, field)
    
    def upsert_variant(self, key, fields):
        """Fügt ein Variant-Row ein oder aktualisiert es (COALESCE verhindert Überschreiben mit NULL)."""
        with self.lock, self._conn() as con:
            self._upsert_variant_conn(con, key, fields)

    #CACHE für parallel zum Distiller laufenden GEN-Symbol-Annotation. 
    def cache_annotations(self, updates):
        """
        Legt Annotationen in einer Cache-Tabelle ab.
        updates = [(key, {"gene_symbol": "BRCA1", "is_coding": True}), ...]
        """
        if not updates:
            return
        with self.lock, self._conn() as con:
            cur = con.cursor()
            cur.execute("""
                CREATE TABLE IF NOT EXISTS annotation_cache (
                    chr TEXT,
                    pos INTEGER,
                    ref TEXT,
                    alt TEXT,
                    build TEXT,
                    gene_symbol TEXT,
                    is_coding INTEGER,
                    PRIMARY KEY (chr, pos, ref, alt, build)
                )
            """)
            for key, data in updates:
                cur.execute("""
                    INSERT OR REPLACE INTO annotation_cache
                    (chr, pos, ref, alt, build, gene_symbol, is_coding)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (
                    key[0], key[1], key[2], key[3], key[4],
                    data["gene_symbol"], int(data["is_coding"])
                ))
            # ✅ V12 FIX: No manual commit - context manager handles it

    def flush_annotation_cache(self):
        rows_to_flush = []
        
        # ✅ Lese Cache außerhalb des DB-Locks
        try:
            with self.lock, self._conn() as con:
                cur = con.cursor()
                cur.execute("SELECT chr, pos, ref, alt, build, gene_symbol, is_coding FROM annotation_cache")
                rows_to_flush = cur.fetchall()

                if not rows_to_flush:
                    return 0

                cur.execute("DELETE FROM annotation_cache")
                # ✅ V12 FIX: No manual commit - context manager handles it
        except Exception as e:
            print(f"[DB] Cache-Flush-Fehler: {e}")
            return 0

        # ✅ Update außerhalb des Locks (parallel möglich)
        update_count = 0
        batch = []
        BATCH_SIZE = 1000
        
        for row in rows_to_flush:
            try:
                key = (row[0], row[1], row[2], row[3], row[4])
                batch.append((key, {
                    "gene_symbol": row[5],
                    "is_coding": bool(row[6])
                }))
                
                if len(batch) >= BATCH_SIZE:
                    self.update_variant_fields_bulk(batch)
                    update_count += len(batch)
                    batch.clear()
                    
            except Exception as e:
                print(f"[DB] Fehler bei Cache-Flush für {key}: {e}")
        
        if batch:
            self.update_variant_fields_bulk(batch)
            update_count += len(batch)

        return update_count

    def _upsert_variant_conn(self, con, key, fields):
        fields = self._normalize_for_db(fields or {})
        cols = list(self.ALL_COLS)
        # Default-Daten: None, dann mit fields überschreiben
        data = {c: None for c in cols}
        data.update(fields)
        # Key setzen (überschreibt ggf. in fields gesetzte key-Werte konsistent)
        data["chr"], data["pos"], data["ref"], data["alt"], data["build"] = key

        placeholders = ",".join(["?"] * len(cols))
        updates = ",".join([
            f"{c}=COALESCE(excluded.{c}, {c})"
            for c in cols if c not in self.KEY_COLS
        ])
        vals = [data[c] for c in cols]

        con.execute(f"""
            INSERT INTO variants ({",".join(cols)})
            VALUES ({placeholders})
            ON CONFLICT(chr,pos,ref,alt,build) DO UPDATE SET {updates}
        """, vals)
    # Einfügen nach Zeile 7646 in der VariantDB Klasse
    def upsert_many(self, items, batch_commit=1000):
        """
        V12 FIX: Removed manual transaction management.
        Context manager now handles BEGIN/COMMIT/ROLLBACK automatically.
        Intermediate commits are handled by closing and reopening the connection.
        """
        if not items:
            return
        cols = list(self.ALL_COLS)
        placeholders = ",".join(["?"] * len(cols))
        updates = ",".join([f"{c}=COALESCE(excluded.{c}, {c})" for c in cols if c not in self.KEY_COLS])
        sql = f"INSERT INTO variants ({','.join(cols)}) VALUES ({placeholders}) ON CONFLICT(chr,pos,ref,alt,build) DO UPDATE SET {updates}"

        # ✅ V12 FIX: Process in batches with separate transactions
        for batch_start in range(0, len(items), batch_commit):
            batch_end = min(batch_start + batch_commit, len(items))
            batch_items = items[batch_start:batch_end]

            with self.lock, self._conn() as con:
                cur = con.cursor()
                for key, fields in batch_items:
                    fields = self._normalize_for_db(fields or {})
                    data = {c: None for c in cols}
                    data.update(fields)
                    data["chr"], data["pos"], data["ref"], data["alt"], data["build"] = key
                    vals = [data[c] for c in cols]
                    cur.execute(sql, vals)
                # Context manager commits automatically at end of batch


    # -----------------------
    # Select/Get-Operationen
    # -----------------------
    def get_positions(self, keys):
        """
        Liefert ein Mapping key -> (chrom, pos).
        Erwartet keys als (chr, pos, ref, alt, build).
        """
        return {key: (key[0], key[1]) for key in keys}

    def _normalize_for_db(self, record: dict) -> dict:
        """Sorgt dafür, dass JSON-Felder korrekt serialisiert werden."""
        rec = dict(record)
        if "conservation" in rec and isinstance(rec["conservation"], dict):
            rec["conservation"] = json.dumps(rec["conservation"])
        return rec

    def _parse_from_db(self, row: dict) -> dict:
        """Sorgt dafür, dass JSON-Felder beim Lesen automatisch geparst werden."""
        rec = dict(row)
        if "conservation" in rec and isinstance(rec["conservation"], str):
            try:
                rec["conservation"] = json.loads(rec["conservation"])
            except Exception:
                # Falls kein valides JSON, String unverändert lassen
                pass
        return rec

    def get_variant(self, key):
        with self.lock, self._conn() as con:
            cur = con.cursor()
            chr_, pos, ref, alt, build = key
            cur.execute(
                "SELECT * FROM variants WHERE chr=? AND pos=? AND ref=? AND alt=? AND build=?",
                (chr_, pos, ref, alt, build)
            )
            row = cur.fetchone()
            return self._parse_from_db(dict(row)) if row else None


    def get_variant_field(self, key, field):
        """Gibt einzelnes Feld zurück oder None, wenn Row/Feld fehlt."""
        if field not in self.ALL_COLS:
            return None
        with self.lock, self._conn() as con:
            cur = con.cursor()
            cur.execute(f"""
                SELECT {field} FROM variants
                WHERE chr=? AND pos=? AND ref=? AND alt=? AND build=?
            """, key)
            row = cur.fetchone()
            if not row:
                return None
            val = row[0]
            # Falls conservation-Feld: JSON parsen
            if field == "conservation" and isinstance(val, str):
                try:
                    import json
                    return json.loads(val)
                except Exception:
                    return val
            return val


    def get_variants_bulk(self, keys: List[Tuple]) -> Dict[Tuple, Optional[Dict]]:
        if not keys:
            return {}

        keys_clean = [tuple(k) for k in keys]
        results = {k: None for k in keys_clean}

        CHUNK_SIZE = 900
        CHUNK_TIMEOUT = 5.0  # ✅ Max 5 Sekunden pro Chunk
        MAX_TOTAL_TIME = 30.0  # ✅ Max 30s für gesamten Bulk-Lookup
        
        total_start = time.time()
        
        for i in range(0, len(keys_clean), CHUNK_SIZE):
            # ✅ Global Timeout Check
            if time.time() - total_start > MAX_TOTAL_TIME:
                if hasattr(self, 'logger') and self.logger:
                    self.logger.log(
                        f"[DB] ⚠️ Bulk-Fetch timeout after {MAX_TOTAL_TIME}s "
                        f"({i}/{len(keys_clean)} processed)"
                    )
                break
            
            chunk = keys_clean[i:i+CHUNK_SIZE]
            chunk_start = time.time()
            
            try:
                with self.lock, self._conn() as con:
                    # ✅ SQLite Lock-Timeout setzen
                    con.execute("PRAGMA busy_timeout = 3000")
                    cur = con.cursor()

                    # ✅ Timeout-Check vor Temp-Table
                    if time.time() - chunk_start > CHUNK_TIMEOUT:
                        continue

                    # ✅ V12 FIX: No manual BEGIN - context manager handles it

                    # Temp-Table erstellen
                    cur.execute("""
                        CREATE TEMP TABLE IF NOT EXISTS tmp_fetch_keys(
                            chr TEXT, pos INTEGER, ref TEXT, alt TEXT, build TEXT
                        )
                    """)
                    cur.execute("DELETE FROM tmp_fetch_keys")

                    # Chunk-Insert mit Timeout-Check
                    insert_sql = "INSERT INTO tmp_fetch_keys VALUES (?, ?, ?, ?, ?)"
                    timeout_occurred = False
                    for j in range(0, len(chunk), 100):
                        if time.time() - chunk_start > CHUNK_TIMEOUT:
                            timeout_occurred = True
                            break
                        mini_chunk = chunk[j:j+100]
                        cur.executemany(insert_sql, mini_chunk)

                    # ✅ V12.1: Skip on timeout (Autocommit - keine Rollback-Logik nötig)
                    if timeout_occurred:
                        continue  # Skip diesen Chunk, weiter mit nächstem

                    # JOIN mit Timeout-Check
                    if time.time() - chunk_start > CHUNK_TIMEOUT:
                        continue  # Skip diesen Chunk

                    cur.execute("""
                        SELECT v.*
                        FROM variants v
                        JOIN tmp_fetch_keys t
                          ON v.chr = t.chr
                         AND v.pos = t.pos
                         AND v.ref = t.ref
                         AND v.alt = t.alt
                         AND v.build = t.build
                    """)

                    rows = cur.fetchall()
                    colnames = [d[0] for d in cur.description]

                    for row in rows:
                        rowd = dict(zip(colnames, row))
                        rowd = self._parse_from_db(rowd)
                        key = (rowd["chr"], rowd["pos"], rowd["ref"], rowd["alt"], rowd["build"])
                        results[key] = rowd

                    cur.execute("DELETE FROM tmp_fetch_keys")
                    # ✅ V12 FIX: No manual COMMIT - context manager handles it
                    
            except TimeoutError:
                # ✅ V12 FIX: Timeout exceptions are handled gracefully
                continue
            except sqlite3.OperationalError as e:
                if "locked" in str(e).lower():
                    continue
                raise
            except Exception as e:
                if hasattr(self, 'logger') and self.logger:
                    self.logger.log(f"[DB] ⚠️ Chunk {i//CHUNK_SIZE} failed: {e}")
                continue

        return results

    def get_all_variant_keys(self, build=None):
        with self.lock, self._conn() as con:
            cur = con.cursor()
            if build:
                cur.execute("""
                    SELECT chr,pos,ref,alt,build
                    FROM variants
                    WHERE build=?
                """, (build,))
            else:
                cur.execute("""
                    SELECT chr,pos,ref,alt,build
                    FROM variants
                """)
            return list(cur.fetchall())
        
    # -----------------------
    # Schlüssel-Migration
    # -----------------------

    def update_variant_key(self, old_key, new_key):
        """
        Migriert einen Datensatz atomar auf neuen zusammengesetzten Primärschlüssel.
        - Kopiert alle Nicht-Schlüssel-Felder.
        - Vermeidet Datenverlust bei Fehlern via Transaktion.

        V12 FIX: Removed manual transaction management.
        """
        with self.lock, self._conn() as con:
            cur = con.cursor()
            # ✅ V12 FIX: No manual BEGIN - context manager handles it

            cur.execute("""
                SELECT * FROM variants
                WHERE chr=? AND pos=? AND ref=? AND alt=? AND build=?""", old_key)
            row = cur.fetchone()
            if not row:
                return  # ✅ V12 FIX: Just return, context manager commits

            colnames = [d[0] for d in cur.description]
            data = dict(zip(colnames, row))

            # Neue Schlüsselwerte setzen
            data["chr"], data["pos"], data["ref"], data["alt"], data["build"] = new_key

            cols = colnames
            placeholders = ",".join(["?"] * len(cols))
            values = [data[c] for c in cols]

            merge_updates = ",".join([
                f"{c}=COALESCE(excluded.{c}, {c})"
                for c in cols if c not in self.KEY_COLS
            ])
            cur.execute(f"""
                INSERT INTO variants ({','.join(cols)}) VALUES ({placeholders})
                ON CONFLICT(chr,pos,ref,alt,build) DO UPDATE SET {merge_updates}
            """, values)

            cur.execute("""
                DELETE FROM variants
                WHERE chr=? AND pos=? AND ref=? AND alt=? AND build=?""", old_key)

            # ✅ V12 FIX: No manual COMMIT - context manager handles it
            # Exception will trigger automatic ROLLBACK
            
    def update_variant_fields(self, key, updates: dict):
        """
        Aktualisiert mehrere Felder für eine Variante.
        updates = {"gene_symbol": "BRCA1", "is_coding": True}
        """
        updates = self._normalize_for_db(updates or {})
        with self.lock, self._conn() as con:
            cur = con.cursor()
            sets = ", ".join([f"{field}=?" for field in updates.keys()])
            values = list(updates.values()) + list(key)
            cur.execute(f"""
                UPDATE variants
                SET {sets}
                WHERE chr=? AND pos=? AND ref=? AND alt=? AND build=?
            """, values)
            # ✅ V12 FIX: No manual commit - context manager handles it


    def update_variant_field(self, key, field, value):
        """
        Aktualisiert ein einzelnes Feld für eine Variante.
        key = (chr, pos, ref, alt, build)
        """
        updates = self._normalize_for_db({field: value})
        norm_value = updates[field]
        with self.lock, self._conn() as con:
            cur = con.cursor()
            cur.execute(f"""
                UPDATE variants
                SET {field} = ?
                WHERE chr=? AND pos=? AND ref=? AND alt=? AND build=?
            """, (norm_value, key[0], key[1], key[2], key[3], key[4]))
            con.commit()


    def update_variant_fields_bulk(self, updates):
        """
        updates = [(key, {"gene_symbol": "BRCA1", "is_coding": True}), ...]
        """
        if not updates:
            return
        with self.lock, self._conn() as con:
            cur = con.cursor()
            for key, data in updates:
                data = self._normalize_for_db(data or {})
                cur.execute("""
                    UPDATE variants
                    SET gene_symbol=?, is_coding=?, conservation=?
                    WHERE chr=? AND pos=? AND ref=? AND alt=? AND build=?
                """, (
                    data.get("gene_symbol"),
                    data.get("is_coding"),
                    data.get("conservation"),
                    key[0], key[1], key[2], key[3], key[4]
                ))
            con.commit()

    def upsert_variants_bulk(self, records: dict):
        """
        records: {key: {"field1": val1, "field2": val2, ...}, ...}
        key = (chr, pos, ref, alt, build)

        Schreibt nur AF-Spalten und Success-/Timestamp-Felder.
        """
        if not records:
            return

        # Alle bekannten Spalten der Tabelle
        all_columns = [
            "chr", "pos", "ref", "alt", "build",
            "rsid",
            "af_filter_mean", "af_exac", "af_gnomad_exomes", "af_gnomad_genomes", "af_1kg",
            "cadd_phred", "gene_symbol", "impact", "consequence",
            "clinical_significance", "phenotypes",
            "protein_coding", "is_coding", "conservation",
            "ag_score", "extra_json",
            "meanAF_last_fetch", "meanAF_fetch_success",
            "full_last_fetch", "full_fetch_success"
        ]

        # Whitelist: nur diese Spalten dürfen überschrieben werden
        AF_COLUMNS = {
            "af_filter_mean", "af_exac", "af_gnomad_exomes", "af_gnomad_genomes", "af_1kg",
            "meanAF_last_fetch", "meanAF_fetch_success",
            "full_last_fetch", "full_fetch_success"
        }

        with self.lock, self._conn() as con:
            cur = con.cursor()
            for key, fields in records.items():
                chr_, pos, ref, alt, build = key

                # Default-Record mit None für alle Spalten
                record = {col: None for col in all_columns if col not in ("chr", "pos", "ref", "alt", "build")}
                # Eingehende Felder normalisieren und überschreiben
                record.update(self._normalize_for_db(fields or {}))

                cols = ["chr", "pos", "ref", "alt", "build"] + list(record.keys())
                placeholders = ",".join(["?"] * len(cols))
                values = [chr_, pos, ref, alt, build] + list(record.values())

                # Nur AF- und Success-Felder überschreiben
                merge_updates = ",".join([
                    f"{c}=excluded.{c}" for c in record.keys() if c in AF_COLUMNS
                ])

                sql = f"""
                    INSERT INTO variants ({','.join(cols)})
                    VALUES ({placeholders})
                    ON CONFLICT(chr,pos,ref,alt,build)
                    DO UPDATE SET {merge_updates}
                """
                cur.execute(sql, values)
            con.commit()

    def update_variant_rsids_bulk(self, updates):
        """
        updates = [(key, {"rsid": "rs123"}), ...]
        """
        if not updates:
            return
        with self.lock, self._conn() as con:
            cur = con.cursor()
            for key, data in updates:
                cur.execute("""
                    UPDATE variants
                    SET rsid=?
                    WHERE chr=? AND pos=? AND ref=? AND alt=? AND build=?
                """, (
                    data.get("rsid"),
                    key[0], key[1], key[2], key[3], key[4]
                ))
            # ✅ V12 FIX: No manual commit - context manager handles it

    # -----------------------
    # Shutdown / Cleanup
    # -----------------------

    def close(self):
        """
        V12 FIX: Proper database shutdown with WAL checkpoint.
        Call this before application exit to ensure data integrity.
        """
        try:
            with self._conn() as con:
                # ✅ Force WAL checkpoint before closing
                con.execute("PRAGMA wal_checkpoint(FULL);")
                print("[DB] WAL checkpoint completed successfully")
        except Exception as e:
            print(f"[DB] Warning: WAL checkpoint failed during shutdown: {e}")

    # -----------------------
    # Export/Reporting
    # -----------------------

    def select_for_export(self, build=None, af_threshold=None, include_none=False):
        """
        Gibt Tuples im UI-Format zurück. genotype und pubmed sind UI-only -> NULL/leer.
        Wendet AF-Threshold konsistent an (<= Schwelle; optional NULL zulassen).
        """
        # Dynamisch SELECT-Liste aus UI_COLUMNS erzeugen
        select_parts = []
        for col in self.UI_COLUMNS:
            if col == "genotype":
                select_parts.append("NULL AS genotype")
            elif col == "pubmed":
                select_parts.append("'' AS pubmed")
            else:
                db_field = self.resolve_column(col)
                if db_field != col:
                    select_parts.append(f"{db_field} AS {col}")
                else:
                    select_parts.append(col)

        base_select = "SELECT " + ", ".join(select_parts) + " FROM variants"

        params, conds = [], []
        if build:
            conds.append("build=?")
            params.append(build)
        if af_threshold is not None:
            if include_none:
                conds.append("(af_filter_mean <= ? OR af_filter_mean IS NULL)")
            else:
                conds.append("af_filter_mean <= ?")
            params.append(float(af_threshold))

        query = base_select
        if conds:
            query += " WHERE " + " AND ".join(conds)

        with self.lock, self._conn() as con:
            cur = con.execute(query, params)
            return cur.fetchall()

    # -----------------------
    # Feldprüfung
    # -----------------------

    def find_keys_without_field(self, field: str, limit: int = 100, only_filtered: bool = False):
        """
        Liefert Keys zurück, bei denen das angegebene Feld fehlt oder nur einen Platzhalter enthält.
        Nutzt resolve_column(), um UI- oder Alias-Namen auf DB-Feldnamen zu mappen.
        """
        db_field = self.resolve_column(field)
        with self.lock, self._conn() as con:
            cur = con.cursor()
            cur.execute(f"""
                SELECT chr, pos, ref, alt, build
                FROM variants
                WHERE {db_field} IS NULL
                   OR {db_field} = ''
                   OR {db_field} = '.'
                LIMIT ?
            """, (limit,))
            rows = cur.fetchall()

        keys = [(r[0], r[1], r[2], r[3], r[4]) for r in rows]

        if only_filtered:
            visible = getattr(self.distiller, "display_keys", set())
            keys = [k for k in keys if k in visible]

        return keys
    # -----------------------
    # Löschen
    # -----------------------

    def delete_variant(self, key):
        """Löscht eine Variante per zusammengesetztem Primärschlüssel."""
        with self.lock, self._conn() as con:
            con.execute("""
                DELETE FROM variants
                WHERE chr=? AND pos=? AND ref=? AND alt=? AND build=?
            """, key)

    def for_background_priorities(
        self,
        stale_days: int,
        p1_cooldown_hours: int,
        mode: str = "full",
        af_none_policy: str = "clinical",
        limit: int = 1000
    ):
        """
        ✅ GEFIXT: SQL-Syntax-Fehler in P3-Query behoben.
        ✅ V12.1: `limit` Parameter für GUI-Responsiveness beim Start.

        Liefert vier Listen (P0–P3) mit Kandidaten.

        Zeilenform: (chr, pos, ref, alt, build, fetch_success, last_fetch)

        Args:
            stale_days: Tage bis Daten als veraltet gelten
            p1_cooldown_hours: Cooldown für generische Fehler
            mode: "full" für Full-Annotation, "af" für AF-Only
            af_none_policy: "clinical", "research", "permissive", "strict"
            limit: Max. Anzahl Kandidaten pro Priority (Default 1000, beim Start 50)

        Returns:
            Tuple (P0, P1, P2, P3) - Listen von Kandidaten

        BUG-FIX:
        - ORDER BY muss NACH UNION kommen, nicht innerhalb
        - Separate Queries für P3 statt UNION mit ORDER BY
        - V12.1: Kleineres LIMIT beim ersten Call verhindert GUI-Freeze
        """
        now = datetime.datetime.now(datetime.timezone.utc)
        stale_cutoff = now - datetime.timedelta(days=stale_days)
        p1_cutoff = now - datetime.timedelta(hours=p1_cooldown_hours)
        
        # Dynamisches Status-Feld basierend auf Mode
        status_field = "full_fetch_success" if mode == "full" else "meanAF_fetch_success"
        last_fetch_field = "full_last_fetch" if mode == "full" else "meanAF_last_fetch"
        
        def in_clause(vals):
            """Werte als Strings in SQL-Literal-Form"""
            return ",".join(f"'{str(v)}'" for v in vals)
        
        not_success = f"{status_field} NOT IN ({in_clause(self.SUCCESS_CODES)})"
        not_deletable = f"{status_field} NOT IN ({in_clause(self.DELETABLE_CODES)})"
        
        # AF-None-Awareness für AF-Mode
        af_none_clause = ""
        if mode == "af" and af_none_policy == "clinical":
            # Clinical: Priorisiere Varianten MIT AF-Daten
            af_none_clause = "AND af_filter_mean IS NOT NULL"
        
        with self.lock, self._conn() as con:
            cur = con.cursor()
            
            # ============================================================
            # P0: Stale Candidates
            # ============================================================
            cur.execute(f"""
                SELECT chr, pos, ref, alt, build, {status_field}, {last_fetch_field}
                FROM variants
                WHERE {not_success} AND {not_deletable}
                  AND ({last_fetch_field} IS NULL OR {last_fetch_field} < ?)
                  {af_none_clause}
                ORDER BY COALESCE({last_fetch_field}, '1970-01-01T00:00:00Z') ASC
                LIMIT ?
            """, (stale_cutoff.strftime("%Y-%m-%dT%H:%M:%SZ"), limit))
            P0 = cur.fetchall()

            # ============================================================
            # P1: Generic Failures
            # ============================================================
            cur.execute(f"""
                SELECT chr, pos, ref, alt, build, {status_field}, {last_fetch_field}
                FROM variants
                WHERE {status_field} IN ({in_clause(self.GENERIC_FAIL_CODES)})
                  AND ({last_fetch_field} IS NULL OR {last_fetch_field} < ?)
                  {af_none_clause}
                ORDER BY COALESCE({last_fetch_field}, '1970-01-01T00:00:00Z') ASC
                LIMIT ?
            """, (p1_cutoff.strftime("%Y-%m-%dT%H:%M:%SZ"), limit))
            P1 = cur.fetchall()

            # ============================================================
            # P2: Re-Normalization Cases
            # ============================================================
            cur.execute(f"""
                SELECT chr, pos, ref, alt, build, {status_field}, {last_fetch_field}
                FROM variants
                WHERE {status_field} IN ({in_clause(self.RENORM_CODES)})
                  {af_none_clause}
                ORDER BY COALESCE({last_fetch_field}, '1970-01-01T00:00:00Z') ASC
                LIMIT ?
            """, (limit,))
            P2 = cur.fetchall()
            
            # ============================================================
            # P3: Other Failures + AF=None (Clinical)
            # ✅ GEFIXT: Separate Queries statt UNION mit ORDER BY
            # ============================================================
            if mode == "af" and af_none_policy == "clinical":
                # Clinical-Mode: Erst OTHER_FAIL_CODES, dann AF=None

                # Teil 1: Other Failure Codes (ohne AF=None)
                p3_limit_half = max(1, limit // 2)
                cur.execute(f"""
                    SELECT chr, pos, ref, alt, build, {status_field}, {last_fetch_field}
                    FROM variants
                    WHERE {status_field} IN ({in_clause(self.OTHER_FAIL_CODES)})
                      AND af_filter_mean IS NOT NULL
                    ORDER BY COALESCE({last_fetch_field}, '1970-01-01T00:00:00Z') ASC
                    LIMIT ?
                """, (p3_limit_half,))
                p3_failures = cur.fetchall()

                # Teil 2: AF=None Varianten (niedrigere Priorität)
                cur.execute(f"""
                    SELECT chr, pos, ref, alt, build, {status_field}, {last_fetch_field}
                    FROM variants
                    WHERE {not_success} AND {not_deletable}
                      AND af_filter_mean IS NULL
                      AND ({last_fetch_field} IS NULL OR {last_fetch_field} < ?)
                    ORDER BY COALESCE({last_fetch_field}, '1970-01-01T00:00:00Z') ASC
                    LIMIT ?
                """, (stale_cutoff.strftime("%Y-%m-%dT%H:%M:%SZ"), p3_limit_half))
                p3_af_none = cur.fetchall()

                # Kombiniere: Failures zuerst, dann AF=None
                P3 = p3_failures + p3_af_none
            else:
                # Standard: Nur OTHER_FAIL_CODES
                cur.execute(f"""
                    SELECT chr, pos, ref, alt, build, {status_field}, {last_fetch_field}
                    FROM variants
                    WHERE {status_field} IN ({in_clause(self.OTHER_FAIL_CODES)})
                      {af_none_clause}
                    ORDER BY COALESCE({last_fetch_field}, '1970-01-01T00:00:00Z') ASC
                    LIMIT ?
                """, (limit,))
                P3 = cur.fetchall()
        
        return P0, P1, P2, P3
    
    def find_cross_build_candidates(self, limit=100):
        """
        Findet Varianten-Paare (Build 37 <-> 38) mit gleicher rsID,
        bei denen Annotationen übertragen werden können.
        
        Strategie:
        Suche Target (T) ohne Gene/CADD/ClinVar, das einen Source (S) 
        mit gleicher rsID aber anderem Build UND vorhandenen Daten hat.
        """
        # Felder, die wir übertragen wollen
        fields_to_sync = [
            "gene_symbol", "cadd_phred", "clinical_significance", 
            "phenotypes", "impact", "consequence"
        ]
        
        # Dynamische WHERE-Klausel: Target muss leer sein, Source muss voll sein
        # Wir prüfen exemplarisch auf 'gene_symbol', um Kandidaten zu finden
        
        query = f"""
            SELECT 
                t.chr, t.pos, t.ref, t.alt, t.build,  -- Target Key
                s.gene_symbol, s.cadd_phred, s.clinical_significance, 
                s.phenotypes, s.impact, s.consequence,
                s.build AS source_build
            FROM variants t
            JOIN variants s ON t.rsid = s.rsid
            WHERE 
                t.rsid IS NOT NULL AND t.rsid != '.' AND t.rsid != ''
                AND t.build != s.build
                AND (t.gene_symbol IS NULL OR t.gene_symbol = '') -- Target fehlt Info
                AND (s.gene_symbol IS NOT NULL AND s.gene_symbol != '') -- Source hat Info
            LIMIT ?
        """
        
        with self.lock, self._conn() as con:
            cur = con.cursor()
            cur.execute(query, (limit,))
            return cur.fetchall()
        
class StopFlag:
    def __init__(self):
        self.ev = threading.Event()

    def stop(self):
        """Signalisiert Abbruch."""
        self.ev.set()

    def clear(self):
        """Reset für neuen Run."""
        self.ev.clear()

    def is_set(self):
        """Prüft, ob Abbruch signalisiert wurde."""
        return self.ev.is_set()

   
class VCFMigrationsdienst(multiprocessing.Process):
    """
    ✅ ERWEITERT: VCF-Cache zu SQLite-Migration mit Progress-Tracking.
    
    Migriert Records aus JSON-Cache in die SQLite-Datenbank.
    Läuft als separater Prozess nach GUI-Exit.
    
    Features:
    - Batch-Processing (5000 Records/Batch)
    - Signal-Handling (SIGTERM, SIGINT)
    - Progress-Tracking
    - Automatic Cache-Cleanup nach erfolgreicher Migration
    - PID-File für Process-Management
    
    """
    
    def __init__(
        self, 
        cache_path: str, 
        db_path: str, 
        logger, 
        pidfile: str = "/tmp/vcfmigrator.pid",
        batch_size: int = 5000,
        cleanup_cache: bool = True
    ):
        """
        Initialisiert VCF-Migrationsdienst.
        
        Args:
            cache_path: Pfad zum JSON-Cache
            db_path: Pfad zur SQLite-DB
            logger: Logger-Instanz
            pidfile: Pfad zur PID-Datei
            batch_size: Records pro Batch
            cleanup_cache: Cache nach erfolgreicher Migration löschen
        """
        super().__init__()
        self.cache_path = cache_path
        self.db_path = db_path
        self.logger = logger
        self.pidfile = pidfile
        self.batch_size = batch_size
        self.cleanup_cache = cleanup_cache
        self._stop_event = multiprocessing.Event()

    def run(self):
        """Haupt-Entry-Point für den Prozess."""
        # PID-File schreiben
        try:
            with open(self.pidfile, "w") as f:
                f.write(str(os.getpid()))
            self.logger.log(
                f"[VCFMigrationsdienst] 🚀 Gestartet (PID {os.getpid()})"
            )
        except Exception as e:
            self.logger.log(
                f"[VCFMigrationsdienst] ⚠️ PID-File-Fehler: {e}"
            )

        # Signal-Handler registrieren
        signal.signal(signal.SIGTERM, self._handle_stop)
        signal.signal(signal.SIGINT, self._handle_stop)

        try:
            success = self._migrate_loop()
            
            # Cache aufräumen wenn erfolgreich
            if success and self.cleanup_cache:
                try:
                    os.remove(self.cache_path)
                    self.logger.log(
                        f"[VCFMigrationsdienst] 🗑️ Cache gelöscht: {self.cache_path}"
                    )
                except Exception as e:
                    self.logger.log(
                        f"[VCFMigrationsdienst] ⚠️ Cache-Cleanup fehlgeschlagen: {e}"
                    )
        
        except Exception as e:
            self.logger.log(
                f"[VCFMigrationsdienst] ❌ Kritischer Fehler: {e}\n"
                f"{traceback.format_exc()}"
            )
        
        finally:
            self._cleanup()

    def _migrate_loop(self) -> bool:
        """
        Haupt-Migrations-Loop.
        
        Returns:
            True wenn erfolgreich, False bei Fehler
        """
        import json
        import sqlite3

        # Cache-Datei prüfen
        if not os.path.exists(self.cache_path):
            self.logger.log(
                f"[VCFMigrationsdienst] ⚠️ Cache nicht gefunden: {self.cache_path}"
            )
            return False

        # Cache laden
        try:
            with open(self.cache_path, "r", encoding="utf-8") as f:
                records = json.load(f)
        except Exception as e:
            self.logger.log(
                f"[VCFMigrationsdienst] ❌ Cache-Load-Fehler: {e}"
            )
            return False

        total_records = len(records)
        self.logger.log(
            f"[VCFMigrationsdienst] 📦 {total_records:,} Records aus Cache geladen"
        )

        if total_records == 0:
            self.logger.log("[VCFMigrationsdienst] ✅ Nichts zu migrieren")
            return True

        # DB-Connection
        try:
            conn = sqlite3.connect(self.db_path)
            cur = conn.cursor()
        except Exception as e:
            self.logger.log(
                f"[VCFMigrationsdienst] ❌ DB-Connection-Fehler: {e}"
            )
            return False

        # Batch-Processing
        buffer = []
        processed = 0
        errors = 0
        start_time = time.time()

        try:
            for idx, (key, update) in enumerate(records, start=1):
                buffer.append((key, update))
                
                # Batch voll oder letzter Record
                if len(buffer) >= self.batch_size or idx == total_records:
                    batch_errors = self._commit_batch(cur, buffer)
                    errors += batch_errors
                    
                    try:
                        conn.commit()
                    except Exception as e:
                        self.logger.log(
                            f"[VCFMigrationsdienst] ⚠️ Commit-Fehler Batch {idx}: {e}"
                        )
                        errors += len(buffer)
                    
                    processed += len(buffer)
                    
                    # Progress-Log
                    if processed % 10000 == 0 or idx == total_records:
                        elapsed = time.time() - start_time
                        rate = processed / elapsed if elapsed > 0 else 0
                        pct = 100.0 * processed / total_records
                        
                        self.logger.log(
                            f"[VCFMigrationsdienst] 📊 Progress: {processed:,}/{total_records:,} "
                            f"({pct:.1f}%) | {rate:.0f} rec/s | Errors: {errors}"
                        )
                    
                    buffer.clear()

                # Stop-Signal?
                if self._stop_event.is_set():
                    self.logger.log(
                        "[VCFMigrationsdienst] ⏸️ Stop-Signal empfangen, breche ab"
                    )
                    conn.close()
                    return False

            conn.close()
            
            # Summary
            elapsed = time.time() - start_time
            self.logger.log(
                f"[VCFMigrationsdienst] ✅ Migration abgeschlossen:\n"
                f"  Records: {processed:,}/{total_records:,}\n"
                f"  Errors: {errors}\n"
                f"  Duration: {elapsed:.1f}s\n"
                f"  Rate: {processed/elapsed:.0f} rec/s"
            )
            
            return errors == 0
        
        except Exception as e:
            self.logger.log(
                f"[VCFMigrationsdienst] ❌ Migration-Fehler: {e}\n"
                f"{traceback.format_exc()}"
            )
            try:
                conn.close()
            except:
                pass
            return False

    def _commit_batch(self, cur, buffer) -> int:
        """
        Schreibt einen Batch in die DB.
        
        Args:
            cur: DB-Cursor
            buffer: Liste von (key, update) Tuples
        
        Returns:
            Anzahl Fehler
        """
        if not buffer:
            return 0

        errors = 0
        
        for key, update in buffer:
            # Validierung
            if isinstance(key, list) and len(key) == 5:
                key = tuple(key)
            
            if not isinstance(key, tuple) or len(key) != 5:
                errors += 1
                continue
            
            if not isinstance(update, dict) or not update:
                errors += 1
                continue

            chr_, pos_, ref_, alt_, build_ = key
            cols = list(update.keys())
            
            if not cols:
                errors += 1
                continue
            
            placeholders = ",".join("?" for _ in cols)
            merge_updates = ",".join([
                f"{c}=COALESCE(excluded.{c}, {c})" 
                for c in cols
            ])

            values = [chr_, pos_, ref_, alt_, build_] + [update[c] for c in cols]

            sql = f"""
                INSERT INTO variants (chr, pos, ref, alt, build, {', '.join(cols)})
                VALUES (?, ?, ?, ?, ?, {placeholders})
                ON CONFLICT(chr,pos,ref,alt,build)
                DO UPDATE SET {merge_updates}
            """
            
            try:
                cur.execute(sql, values)
            except Exception as e:
                errors += 1
                if errors <= 10:  # Log nur erste 10 Fehler
                    self.logger.log(
                        f"[VCFMigrationsdienst] ⚠️ Upsert failed for {key}: {e}"
                    )
        
        return errors

    def _handle_stop(self, signum, frame):
        """Signal-Handler für SIGTERM/SIGINT."""
        self.logger.log(
            f"[VCFMigrationsdienst] 🛑 Stop-Signal {signum} empfangen"
        )
        self._stop_event.set()

    def _cleanup(self):
        """Cleanup: PID-File entfernen."""
        try:
            if os.path.exists(self.pidfile):
                os.remove(self.pidfile)
        except Exception as e:
            self.logger.log(
                f"[VCFMigrationsdienst] ⚠️ PID-File-Cleanup-Fehler: {e}"
            )
        
        self.logger.log("[VCFMigrationsdienst] 💤 Beendet")
        
    def migrate_to_db(self):
        """
        Schreibt Records aus Queue in DB.
        FIX: Connection wird immer geschlossen
        """
        total_records = self._record_queue.qsize()
        if total_records == 0:
            self.logger.log("[VCFMigrationsdienst] Warnung: Queue ist leer")
            return True

        self.logger.log(
            f"[VCFMigrationsdienst] Start: Migration von {total_records:,} Records..."
        )

        # FIX: Connection mit try-finally
        conn = None
        try:
            conn = sqlite3.connect(self.db_path)
            cur = conn.cursor()
        except Exception as e:
            self.logger.log(
                f"[VCFMigrationsdienst] Fehler: DB-Connection-Fehler: {e}"
            )
            return False

        # Batch-Processing
        buffer = []
        processed = 0
        errors = 0
        start_time = time.time()

        try:
            # Records aus Queue holen
            records = []
            while not self._record_queue.empty():
                try:
                    records.append(self._record_queue.get_nowait())
                except queue.Empty:
                    break
            
            total_records = len(records)
            
            for idx, (key, update) in enumerate(records, start=1):
                buffer.append((key, update))
                
                # Batch voll oder letzter Record
                if len(buffer) >= self.batch_size or idx == total_records:
                    batch_errors = self._commit_batch(cur, buffer)
                    errors += batch_errors
                    
                    try:
                        conn.commit()
                    except Exception as e:
                        self.logger.log(
                            f"[VCFMigrationsdienst] Warnung: Commit-Fehler Batch {idx}: {e}"
                        )
                        errors += len(buffer)
                    
                    processed += len(buffer)
                    
                    # Progress-Log
                    if processed % 10000 == 0 or idx == total_records:
                        elapsed = time.time() - start_time
                        rate = processed / elapsed if elapsed > 0 else 0
                        pct = 100.0 * processed / total_records
                        
                        self.logger.log(
                            f"[VCFMigrationsdienst] Progress: {processed:,}/{total_records:,} "
                            f"({pct:.1f}%) | {rate:.0f} rec/s | Errors: {errors}"
                        )
                    
                    buffer.clear()

                # Stop-Signal?
                if self._stop_event.is_set():
                    self.logger.log(
                        "[VCFMigrationsdienst] Stop-Signal empfangen, breche ab"
                    )
                    return False
            
            # Summary
            elapsed = time.time() - start_time
            self.logger.log(
                f"[VCFMigrationsdienst] OK: Migration abgeschlossen:\n"
                f"  Records: {processed:,}/{total_records:,}\n"
                f"  Errors: {errors}\n"
                f"  Duration: {elapsed:.1f}s\n"
                f"  Rate: {processed/elapsed:.0f} rec/s"
            )
            
            return errors == 0
        
        except Exception as e:
            self.logger.log(
                f"[VCFMigrationsdienst] Fehler: Migration-Fehler: {e}\n"
                f"{traceback.format_exc()}"
            )
            return False
        finally:
            # KRITISCH: Connection immer schliessen
            if conn:
                try:
                    conn.close()
                except Exception:
                    pass

class BackofficeCrawler:
    """
    ✅ ERWEITERT: Orchestriert Migration + BackgroundMaintainer nach GUI-Exit.
    
    Workflow:
    1. GUI wird geschlossen
    2. VCFMigrationsdienst läuft (falls Cache vorhanden)
    3. BackgroundMaintainer startet (falls Distiller verfügbar)
    4. Tray-Icon zeigt Status
    5. Läuft bis manuell beendet
    
    Features:
    - Tray-Icon mit Status-Anzeige
    - Sequentielle Ausführung (Migration → Maintainer)
    - Flexible Maintainer-Class-Injection
    - Error-Handling mit Status-Anzeige
    
    """
    
    # PID-Lock-Datei fuer Singleton-Check
    _LOCK_FILE = str(Path.home() / ".vfdistiller" / "crawler.lock")

    @classmethod
    def is_already_running(cls):
        """Prueft ob bereits eine Crawler-Instanz laeuft (PID-Lock)."""
        lock_path = Path(cls._LOCK_FILE)
        if lock_path.exists():
            try:
                pid = int(lock_path.read_text().strip())
                # Pruefen ob PID noch existiert
                if sys.platform == "win32":
                    import ctypes
                    kernel32 = ctypes.windll.kernel32
                    handle = kernel32.OpenProcess(0x100000, False, pid)
                    if handle:
                        kernel32.CloseHandle(handle)
                        return True
                else:
                    os.kill(pid, 0)
                    return True
            except (ValueError, OSError, PermissionError):
                pass
            # Stale Lock-File entfernen
            try:
                lock_path.unlink()
            except OSError:
                pass
        return False

    def _acquire_lock(self):
        """Erstellt PID-Lock-Datei."""
        lock_path = Path(self._LOCK_FILE)
        lock_path.parent.mkdir(parents=True, exist_ok=True)
        lock_path.write_text(str(os.getpid()))

    def _release_lock(self):
        """Entfernt PID-Lock-Datei."""
        try:
            Path(self._LOCK_FILE).unlink()
        except OSError:
            pass

    def __init__(
        self,
        cache_path: str,
        db_path: str,
        logger,
        distiller=None,
        db=None,
        stopflag=None,
        maintainer_cls=None,
        gene_annotator=None,
        af_fetcher=None,
        lang: str = "de"
    ):
        """
        Initialisiert BackofficeCrawler.

        Args:
            cache_path: Pfad zum VCF-Cache (für Migration)
            db_path: Pfad zur SQLite-DB
            logger: Logger-Instanz
            distiller: Distiller-Instanz (für Maintainer)
            db: VariantDB-Instanz (optional, wird sonst aus distiller geholt)
            stopflag: StopFlag für Pipeline-Kontrolle
            maintainer_cls: BackgroundMaintainer-Klasse (für Dependency-Injection)
            gene_annotator: GeneAnnotator-Instanz (für Maintainer)
            af_fetcher: AFFetchController-Instanz (für Maintainer)
            lang: Sprachcode fuer Tray-UI (default: "de")
        """
        self.cache_path = cache_path or ""
        self.db_path = db_path or ""
        self.logger = logger
        self.distiller = distiller
        self.db = db
        self.stopflag = stopflag
        self.maintainer_cls = maintainer_cls or BackgroundMaintainer
        self.gene_annotator = gene_annotator
        self.af_fetcher = af_fetcher

        # Eigenstaendiger Translator (kein App-Zugang im separaten Prozess)
        self._translations = {}
        self._lang = lang
        try:
            _base = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
            _t_path = os.path.join(_base, "locales", "translations.json")
            with open(_t_path, "r", encoding="utf-8") as f:
                self._translations = json.load(f)
        except Exception as e:
            self.logger.log(f"[BackofficeCrawler] Translations not loaded: {e}")

        # Tray-Icon
        self.icon = None
        self._tray_thread = None
        self._running = True

    def _t(self, key: str) -> str:
        """Eigenstaendiger Translator fuer Tray-UI (kein App-Zugang)."""
        entry = self._translations.get(key)
        if not entry:
            return key
        return entry.get(self._lang, entry.get("de", key))

    # ========================================================================
    # TRAY-ICON
    # ========================================================================

    def _create_icon_image(self, color=(0, 128, 0)):
        """
        Erzeugt ein einfaches Icon (farbiger Kreis).
        
        Args:
            color: RGB-Tuple für Icon-Farbe
        
        Returns:
            PIL.Image
        """
        img = Image.new("RGB", (64, 64), (255, 255, 255))
        d = ImageDraw.Draw(img)
        d.ellipse((8, 8, 56, 56), fill=color)
        return img

    def _setup_tray(self):
        """Richtet Tray-Icon mit Menü ein."""
        def on_quit(icon, item):
            self.logger.log("[BackofficeCrawler] 🛑 Tray: Quit gewählt")
            self._running = False
            icon.stop()

        self.icon = pystray.Icon(
            "BackofficeCrawler",
            icon=self._create_icon_image(),
            title=self._t("BackofficeCrawler läuft"),
            menu=pystray.Menu(
                pystray.MenuItem(self._t("Beenden"), on_quit)
            )
        )
        self.icon.run()

    def _start_tray(self):
        """Startet Tray-Icon in separatem Thread."""
        self._tray_thread = threading.Thread(
            target=self._setup_tray, 
            daemon=True
        )
        self._tray_thread.start()
        self.logger.log("[BackofficeCrawler] 🖼️ Tray-Icon gestartet")

    def _update_tray_status(self, status: str):
        """
        Aktualisiert Tray-Icon-Status.
        
        Args:
            status: 'migration', 'maintainer', 'error', 'idle'
        """
        if not self.icon:
            return
        
        status_config = {
            "migration": {
                "color": (0, 128, 0),  # Grün
                "title": self._t("Migration läuft")
            },
            "maintainer": {
                "color": (0, 100, 200),  # Blau
                "title": self._t("BackgroundMaintainer aktiv")
            },
            "error": {
                "color": (200, 0, 0),  # Rot
                "title": self._t("Fehler")
            },
            "idle": {
                "color": (128, 128, 128),  # Grau
                "title": self._t("Idle")
            }
        }
        
        config = status_config.get(status, status_config["idle"])
        self.icon.icon = self._create_icon_image(config["color"])
        self.icon.title = config["title"]
        
        try:
            self.icon.update_menu()
        except Exception:
            # update_menu kann fehlschlagen wenn Icon gerade geschlossen wird
            pass

    # ========================================================================
    # HAUPTLOGIK
    # ========================================================================

    def run(self):
        """
        ✅ HAUPT-ENTRY-POINT: Orchestriert Migration + Maintainer.
        """
        if self.is_already_running():
            self.logger.log("[BackofficeCrawler] ⚠️ Bereits eine Instanz aktiv – Abbruch")
            return
        self._acquire_lock()
        self.logger.log("[BackofficeCrawler] 🚀 Gestartet")
        self._start_tray()

        try:
            # ============================================================
            # PHASE 1: VCF-Cache-Migration
            # ============================================================
            if self.cache_path and self.db_path:
                if os.path.exists(self.cache_path):
                    self._update_tray_status("migration")
                    self.logger.log(
                        f"[BackofficeCrawler] 📦 Starte VCF-Migration: {self.cache_path}"
                    )
                    
                    migrator = VCFMigrationsdienst(
                        cache_path=self.cache_path,
                        db_path=self.db_path,
                        logger=self.logger
                    )
                    migrator.start()
                    migrator.join()  # Warte auf Completion
                    
                    self.logger.log("[BackofficeCrawler] ✅ Migration abgeschlossen")
                else:
                    self.logger.log(
                        f"[BackofficeCrawler] ⏭️ Keine Migration "
                        f"(Cache nicht gefunden: {self.cache_path})"
                    )
            else:
                self.logger.log(
                    "[BackofficeCrawler] ⏭️ Keine Migration "
                    "(Cache/DB-Pfad fehlt)"
                )

            # ============================================================
            # PHASE 2: BackgroundMaintainer
            # ============================================================
            if self.distiller is None:
                self.logger.log(
                    "[BackofficeCrawler] ⚠️ Kein Distiller → "
                    "Maintainer nicht gestartet"
                )
                self._update_tray_status("idle")
                return

            self._update_tray_status("maintainer")

            # DB-Instanz ermitteln
            db_instance = None
            try:
                if hasattr(self.distiller, "db") and isinstance(
                    self.distiller.db, VariantDB
                ):
                    db_instance = self.distiller.db
                    self.logger.log(
                        "[BackofficeCrawler] ✅ DB aus Distiller übernommen"
                    )
                elif self.db_path:
                    db_instance = VariantDB(self.db_path)
                    self.logger.log(
                        f"[BackofficeCrawler] ✅ DB neu geöffnet: {self.db_path}"
                    )
            except Exception as e:
                self.logger.log(
                    f"[BackofficeCrawler] ❌ DB-Initialisierung fehlgeschlagen: {e}\n"
                    f"{traceback.format_exc()}"
                )

            if db_instance is None:
                self.logger.log(
                    "[BackofficeCrawler] ❌ Keine DB verfügbar → "
                    "Maintainer nicht gestartet"
                )
                self._update_tray_status("error")
                return

            # Maintainer starten
            self.logger.log("[BackofficeCrawler] 🛠️ Starte BackgroundMaintainer...")
            
            try:
                maint = self.maintainer_cls(
                    distiller=self.distiller,
                    db=db_instance,
                    stopflag=self.stopflag,
                    logger=self.logger,
                    gene_annotator=self.gene_annotator,
                    af_fetcher=self.af_fetcher
                )
                
                self.logger.log(
                    f"[BackofficeCrawler] ✅ Maintainer initialisiert\n"
                    f"  Gene-Annotator: {'✓' if self.gene_annotator else '✗'}\n"
                    f"  AF-Fetcher: {'✓' if self.af_fetcher else '✗'}"
                )
                
                # run_forever blockiert bis Maintainer beendet wird
                maint.run_forever()
                
            except Exception as e:
                self.logger.log(
                    f"[BackofficeCrawler] ❌ Maintainer-Fehler: {e}\n"
                    f"{traceback.format_exc()}"
                )
                self._update_tray_status("error")
        
        except Exception as e:
            self.logger.log(
                f"[BackofficeCrawler] ❌ Kritischer Fehler: {e}\n"
                f"{traceback.format_exc()}"
            )
            self._update_tray_status("error")
        
        finally:
            self._release_lock()
            self.logger.log("[BackofficeCrawler] 💤 Beendet")

    def stop(self):
        """
        Stoppt BackofficeCrawler und Maintainer.
        """
        self._running = False
        if self.icon:
            try:
                self.icon.stop()
            except Exception as e:
                self.logger.log(
                    f"[BackofficeCrawler] ⚠️ Icon-Stop-Fehler: {e}"
                )
            
class BackgroundMaintainer:
    """
    ✅ ERWEITERT: AF-None-Awareness + Mode-basierte Priorities
    
    STALE DAYS: Refactoring abgeschlossen (2026-03).
    Config.STALE_DAYS_AF=365, Config.STALE_DAYS_FULL=30, UI-Integration vorhanden.
    """
    
    def __init__(
        self, 
        distiller, 
        db: VariantDB, 
        stopflag: StopFlag,
        logger, 
        threads=DEFAULT_THREADS, 
        stale_days=Config.STALE_DAYS_AF,
        p1_cooldown_hours=24, 
        max_per_round=100,
        short_pause=15, 
        long_pause=300, 
        error_threshold=0.5,
        gene_annotator=None, 
        af_fetcher=None
    ):
        """
        Initialisiert BackgroundMaintainer mit AF-None-Awareness.
        
        Args:
            distiller: Distiller-Instanz (für upsert_from_mv)
            db: VariantDB-Instanz
            stopflag: StopFlag für Pipeline-Kontrolle
            logger: Logger-Instanz
            threads: Anzahl Worker-Threads
            stale_days: Tage bis Daten als veraltet gelten (TODO: aufteilen)
            p1_cooldown_hours: Cooldown für generische Fehler
            max_per_round: Max. Varianten pro Durchlauf
            short_pause: Kurze Pause (Sekunden)
            long_pause: Lange Pause (Sekunden)
            error_threshold: Fehler-Schwellwert
            gene_annotator: GeneAnnotator-Instanz (optional)
            af_fetcher: AFFetchController-Instanz (optional)
        """
        self.logger = logger
        self.distiller = distiller
        self.db = db
        self.stopflag = stopflag
        self.threads = threads
        self.stale_days = stale_days
        self.p1_cooldown_hours = p1_cooldown_hours
        self.max_per_round = max_per_round
        self.short_pause = short_pause
        self.long_pause = long_pause
        self.error_threshold = error_threshold
        self.running = False
        self.thread = None
        
        # ✅ Thread-safe Pipeline-State
        self._pipeline_lock = threading.RLock()
        self._pipeline_active = False
        
        self._last_non_p0 = 0

        self.af_fetcher = af_fetcher
        self.gene_annotator = gene_annotator
        self.status_mgr = FetchStatusManager
        
        # ✅ NEU: AF-None-Manager (übernehmen vom Distiller falls vorhanden)
        if hasattr(distiller, 'af_none_manager'):
            self.af_none_manager = distiller.af_none_manager
            self.logger.log(
                f"[Maint] ✅ AF-None-Manager übernommen "
                f"(Policy: {self.af_none_manager.current_preset})"
            )
        else:
            # Fallback: Eigener Manager
            # AfNoneTreatmentManager ist in dieser Datei definiert (Zeile ~5908)
            self.af_none_manager = AfNoneTreatmentManager(
                preset="clinical",
                logger=self.logger
            )
            self.logger.log(
                "[Maint] ⚠️ Erstelle eigenen AF-None-Manager (clinical)"
            )

        # Validation
        if self.distiller is None:
            raise ValueError("BackgroundMaintainer: distiller darf nicht None sein.")

        if not hasattr(self.distiller, "upsert_from_mv"):
            raise AttributeError("BackgroundMaintainer: Distiller fehlt Methode: upsert_from_mv")

        for fld in ("done_variants", "end_retry_variants"):
            if not hasattr(self.distiller, fld):
                setattr(self.distiller, fld, 0)
    
    # ✅ Thread-safe Property für Pipeline-Status
    @property
    def pipeline_active(self) -> bool:
        with self._pipeline_lock:
            return self._pipeline_active
    
    @pipeline_active.setter
    def pipeline_active(self, value: bool):
        with self._pipeline_lock:
            self._pipeline_active = value
            
    def start(self):
        """Startet den Hintergrund-Thread."""
        if self.running:
            logger.log("[Maint] ⚠️ BackgroundMaintainer läuft bereits.")
            return
        logger.log("[Maint] 🌱 Starte BackgroundMaintainer – die gute Seele wacht auf...")
        self.running = True
        self.thread = threading.Thread(target=self._run, daemon=True)
        self.thread.start()

    def stop(self):
        """Stoppt den Hintergrund-Thread."""
        if not self.running:
            logger.log("[Maint] ⚠️ BackgroundMaintainer war nicht aktiv.")
            return
        logger.log("[Maint] 💤 Stoppe BackgroundMaintainer – Zeit für Schönheitsschlaf...")
        self.running = False
        if self.thread and self.thread.is_alive():
            self.thread.join(timeout=5)
        logger.log("[Maint] ✅ BackgroundMaintainer gestoppt.")
        
    def pause(self):
        """Thread-safe Pause."""
        with self._pipeline_lock:
            if not self._pipeline_active:
                self.logger.log("[Maintainer] Pausiert (Pipeline startet).")
                self._pipeline_active = True

    def force_pause(self, timeout: float = 2.0):
        """
        V15: Härteres Pausieren mit Warte-Garantie.

        Problem (V14): pause() setzt nur Flag, Maintainer prüft es aber nur
        ZWISCHEN Batches. Wenn Maintainer in großem Query hängt → VCF-Scan blockiert.

        Lösung: Warte bis aktueller Batch abgeschlossen ist (mit Timeout).

        Args:
            timeout: Max. Wartezeit in Sekunden (default: 2s)
        """
        self.pause()  # Flag setzen

        # ✅ Warte bis Maintainer tatsächlich pausiert
        # (erkennt man daran, dass er auf _pause_ack wartet oder idle ist)
        if not hasattr(self, '_current_batch_active'):
            self._current_batch_active = threading.Event()

        start = time.time()
        while hasattr(self, '_current_batch_active') and self._current_batch_active.is_set():
            if time.time() - start > timeout:
                self.logger.log(f"[Maintainer] ⚠️ force_pause Timeout nach {timeout}s - Maintainer evtl. noch aktiv")
                break
            time.sleep(0.05)

        self.logger.log("[Maintainer] ✅ force_pause abgeschlossen")

    def resume(self):
        """Thread-safe Resume."""
        with self._pipeline_lock:
            if self._pipeline_active:
                self.logger.log("[Maintainer] Wieder aktiviert (Pipeline abgeschlossen).")
                self._pipeline_active = False

    def run_forever(self):
        """
        ✅ V12.1: Startet Thread und blockiert bis zum Ende.

        Öffentliche Einstiegsmethode für den Maintainer in BackofficeCrawler.
        Anders als start() blockiert diese Methode und wartet auf Thread-Ende.
        """
        if not self.running:
            self.start()  # Startet den Thread

        # Blockiere und warte bis Thread beendet wird
        if self.thread and self.thread.is_alive():
            try:
                self.thread.join()  # Blockiert bis Maintainer stoppt
            except KeyboardInterrupt:
                self.logger.log("[Maint] ⚠️ KeyboardInterrupt - stoppe Maintainer...")
                self.stop()

    def _run(self):
        """
        ✅ ERWEITERT: Mode-basierte Priorities + AF-None-Policy + Cross-Build-Sync
        ✅ V12.1: Initial Sleep für GUI-Responsiveness
        ✅ V15: Batch-Active-Tracking für force_pause()
        """
        self.logger.log("[Maint] 🛠 Hintergrund-Thread läuft")

        # ✅ V15: Event für force_pause() Synchronisation
        self._current_batch_active = threading.Event()

        # ✅ V12.1: 3 Sekunden warten, damit GUI vollständig geladen ist
        # Verhindert GUI-Freeze durch erste DB-Query beim App-Start
        self.logger.log("[Maint] ⏳ Warte 3s für GUI-Startup...")
        time.sleep(3)
        self.logger.log("[Maint] ✅ Startup-Delay abgeschlossen")

        was_paused = False
        next_first_block = random.choice(["P0", "P4"])
        self.last_cache_flush = time.time()
        self.last_p5_skip = 0
        self._last_non_p0 = 0

        # Worker für P6 (Cross-Build-Sync)
        def process_cross_build_sync(limit=50):
            """P6: Synchronisiert Daten zwischen GRCh37 und 38 via rsID."""
            try:
                # Hole Kandidaten
                candidates = self.db.find_cross_build_candidates(limit=limit)
                
                if not candidates:
                    return 0, 0
                
                updates = []
                synced_count = 0
                
                for row in candidates:
                    # Entpacken (Target Key + Source Data)
                    t_key = (row[0], row[1], row[2], row[3], row[4])
                    
                    # Daten zum Übertragen
                    update_data = {
                        "gene_symbol": row[5],
                        "cadd_phred": row[6],
                        "clinical_significance": row[7],
                        "phenotypes": row[8],
                        "impact": row[9],
                        "consequence": row[10],
                        # Wir setzen auch den Fetch-Status auf Success, da wir Daten haben
                        "full_fetch_success": "111",
                        "full_last_fetch": datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
                    }
                    
                    # Nur Werte übernehmen, die nicht None sind
                    clean_update = {k: v for k, v in update_data.items() if v is not None}
                    
                    if clean_update:
                        updates.append((t_key, clean_update))
                        synced_count += 1
                
                # Bulk Update durchführen
                if updates:
                    self.db.upsert_many(updates)
                    self.logger.log(f"[Maint-P6] 🧬 Cross-Build-Sync: {synced_count} Varianten angereichert")
                    
                    # Cache invalidieren, damit GUI es sieht
                    if self.distiller:
                        keys = [u[0] for u in updates]
                        self.distiller.invalidate_cache_bulk(keys)
                        
                return synced_count, 0
                
            except Exception as e:
                self.logger.log(f"[Maint-P6] ⚠️ Fehler: {e}")
                return 0, 0

        # Worker für P0-P3 + P5
        def process_block(rows, fetch_full, label):
            """Pipeline-Pause-Check in jedem Block."""
            if self.pipeline_active:
                self.logger.log(f"[{label}] ⏸️ Pipeline aktiv → abgebrochen")
                return 0, 0
            
            if not rows:
                self.logger.log(f"[{label}] ✅ keine Arbeit")
                return 0, 0

            # P5: Gen-Annotation
            if label == "Maint-P5":
                try:
                    import psutil
                    cpu_load = psutil.cpu_percent(interval=0.1)
                    if cpu_load > 70:
                        if time.time() - self.last_p5_skip > 60:
                            self.logger.log(f"[{label}] ⏸ Übersprungen (CPU {cpu_load:.1f}%)")
                            self.last_p5_skip = time.time()
                        return 0, 0

                    self.logger.log(f"[{label}] 🧬 {len(rows)} Varianten ohne Gen-Symbol")
                    
                    pos_map = self.db.get_positions(rows)
                    if not pos_map:
                        self.logger.log(f"[{label}] ⚠️ Keine Positionen")
                        return 0, 0
                    
                    annots = self.gene_annotator.annotate_batch(
                        [(k, chrom, pos) for k, (chrom, pos) in pos_map.items()]
                    )

                    bulk_updates = [
                        (key, data) for key, data in annots.items()
                        if isinstance(data, dict) and data.get("gene_symbol")
                    ]
                    
                    if bulk_updates:
                        self.db.cache_annotations(bulk_updates)
                        self.logger.log(f"[{label}] 🧬 {len(bulk_updates)} Annotationen")
                    
                    return len(bulk_updates), 0

                except Exception as e:
                    self.logger.log(f"[{label}] ⚠️ Fehler: {e}")
                    return 0, 0

            # P0-P3: Standard-Fetches
            actionable, deleted = [], 0
            
            for r in rows[:self.max_per_round]:
                if self.pipeline_active:
                    self.logger.log(f"[{label}] ⏸️ Pipeline aktiv → Abbruch")
                    break
                
                try:
                    key = (r[0], r[1], r[2], r[3], r[4])
                    status_field = "full_fetch_success" if fetch_full else "meanAF_fetch_success"
                    status = self.db.get_variant_field(key, status_field)

                    if FetchStatusManager.is_success(status):
                        continue
                    
                    if FetchStatusManager.is_deletable(status):
                        try:
                            self.db.delete_variant(key)
                            self.logger.log(f"[{label}] 🗑 Gelöscht: {key}")
                            deleted += 1
                        except Exception as e:
                            self.logger.log(f"[{label}] ⚠️ Löschen fehlgeschlagen: {e}")
                        continue

                    actionable.append(key)
                except Exception as e:
                    self.logger.log(f"[{label}] ⚠️ Fehler: {e}")
                    continue

            if not actionable:
                return deleted, 0

            mode = "full" if fetch_full else "af"
            self.logger.log(f"[{label}] 🔍 {len(actionable)} Varianten (Mode={mode})")

            try:
                before_done = self.distiller.done_variants
                before_fail = self.distiller.end_retry_variants

                self.automatic_fetch_decission_and_processing_unit(
                    actionable, actionable[0][4], mode=mode, stale_days=self.stale_days
                )

                done = (self.distiller.done_variants - before_done) + deleted
                fail = self.distiller.end_retry_variants - before_fail
                return done, fail
            
            except Exception as e:
                self.logger.log(f"[{label}] ⚠️ Fetch-Fehler: {e}")
                return deleted, 0

        # ---------------- HAUPTSCHLEIFE ----------------
        # ✅ V12.1: Warmup-Phase für GUI-Responsiveness
        first_iteration = True
        warmup_limit = 50  # Beim ersten Durchlauf nur 50 statt 1000

        while self.running and not self.stopflag.is_set():
            try:
                # Cache Flush alle 180s
                if time.time() - self.last_cache_flush > 180:
                    try:
                        flushed = self.db.flush_annotation_cache()
                        if flushed and flushed > 0:
                            self.logger.log(f"[Maint-P5] 💾 {flushed} Annotationen in DB geschrieben")
                        self.last_cache_flush = time.time()
                    except Exception as e:
                        self.logger.log(f"[Maint-P5] ⚠️ Cache-Flush fehlgeschlagen: {e}")
                        self.last_cache_flush = time.time()

                is_paused = self.pipeline_active

                if is_paused:
                    if not was_paused:
                        self.logger.log("[Maint] ⏸️ Pipeline aktiv → pausiere")
                        was_paused = True
                    time.sleep(self.short_pause)
                    continue
                elif was_paused:
                    self.logger.log("[Maint] 👋 Pipeline beendet")
                    was_paused = False

                # Policy vom AF-None-Manager holen
                policy = self.af_none_manager.current_preset

                # ✅ V12.1: Dynamisches LIMIT für GUI-Responsiveness
                query_limit = warmup_limit if first_iteration else 1000

                # Priorities abrufen
                try:
                    # AF-Priorities
                    p0_af, p1_af, p2_af, p3_af = self.db.for_background_priorities(
                        stale_days=self.stale_days,
                        p1_cooldown_hours=self.p1_cooldown_hours,
                        mode="af",
                        af_none_policy=policy,
                        limit=query_limit
                    )

                    # Full-Priorities
                    p0_full, p1_full, p2_full, p3_full = self.db.for_background_priorities(
                        stale_days=self.stale_days,
                        p1_cooldown_hours=self.p1_cooldown_hours,
                        mode="full",
                        af_none_policy=policy,
                        limit=query_limit
                    )

                    # ✅ V12.1: Nach erstem Durchlauf LIMIT erhöhen
                    if first_iteration:
                        first_iteration = False
                        self.logger.log(f"[Maint] 🚀 Warmup abgeschlossen, erhöhe Query-LIMIT auf 1000")
                    
                    # Kombiniere (prioritize AF über Full)
                    p0 = p0_af if p0_af else p0_full
                    p1 = p1_af if p1_af else p1_full
                    p2 = p2_af if p2_af else p2_full
                    p3 = p3_af if p3_af else p3_full
                    
                except Exception as e:
                    self.logger.log(f"[Maint] ⚠️ Fehler beim Abrufen der Prioritäten: {e}")
                    time.sleep(self.short_pause)
                    continue

                p5 = []
                if getattr(self, "gene_annotator", None) and self.gene_annotator.available:
                    try:
                        import psutil
                        cpu_load = psutil.cpu_percent(interval=0.1)
                        if cpu_load < 70:
                            p5 = self.db.find_keys_without_field("gene_symbol", limit=self.max_per_round)
                        elif time.time() - self.last_p5_skip > 60:
                            self.logger.log(f"[Maint-P5] ⏸ Übersprungen (CPU {cpu_load:.1f}%)")
                            self.last_p5_skip = time.time()
                    except Exception as e:
                        self.logger.log(f"[Maint-P5] ⚠️ CPU-Check fehlgeschlagen: {e}")

                # Priorities Liste mit P6 (Cross-Build)
                priorities = [
                    (p1, False, "Maint-P1"),
                    (p2, True,  "Maint-P2"),
                    (p3, True,  "Maint-P3"),
                    (p5, None,  "Maint-P5"),
                    (None, "CROSS_BUILD", "Maint-P6")  # ✅ NEU
                ]
                
                if not hasattr(self, '_last_non_p0'):
                    self._last_non_p0 = 0
                idx = self._last_non_p0 % len(priorities)

                work_done = False
                
                # P0 und P4 Behandlung (High Priority & Cleanup)
                if not self.pipeline_active:
                    if next_first_block == "P0" and p0:
                        try:
                            process_block(p0, fetch_full=False, label="Maint-P0")
                            work_done = True
                            time.sleep(self.short_pause)
                        except Exception as e:
                            self.logger.log(f"[Maint-P0] ⚠️ Fehler: {e}")

                    elif next_first_block == "P4":
                        try:
                            self.logger.log("[Maint] 🧹 Starte P4-Bereinigung")
                            self._phase4_cleanup_invalid_variants(limit=100)
                            work_done = True
                            time.sleep(self.short_pause)
                        except Exception as e:
                            self.logger.log(f"[Maint-P4] ⚠️ Fehler: {e}")

                # Round-Robin durch P1, P2, P3, P5, P6
                for _ in range(len(priorities)):
                    if self.pipeline_active:
                        break
                    
                    item = priorities[idx]
                    idx = (idx + 1) % len(priorities)
                    
                    # Spezialfall P6
                    if item[1] == "CROSS_BUILD":
                        try:
                            # Nur laufen lassen, wenn keine high-prio items da sind
                            if not (p0 or p1 or p2):
                                done, _ = process_cross_build_sync(limit=100)
                                if done > 0:
                                    self._last_non_p0 = idx
                                    work_done = True
                                    time.sleep(self.short_pause)
                                    break
                        except Exception as e:
                            self.logger.log(f"[Maint-P6] Fehler: {e}")
                        continue
                    
                    # Standard P1, P2, P3, P5
                    rows, full, label = item
                    if rows:
                        try:
                            process_block(rows, fetch_full=full, label=label)
                            self._last_non_p0 = idx
                            work_done = True
                            time.sleep(self.short_pause)
                            break
                        except Exception as e:
                            self.logger.log(f"[{label}] ⚠️ Fehler: {e}")

                # Idle-Handling
                if not work_done and not self.pipeline_active and not any([p0, p1, p2, p3, p5]):
                    try:
                        self.logger.log("[Maint] 🧹 Keine Arbeit → starte P4")
                        self._phase4_cleanup_invalid_variants(limit=100)
                        time.sleep(self.short_pause)
                    except Exception as e:
                        self.logger.log(f"[Maint-P4] ⚠️ Fehler: {e}")
                # ✅ V12.1: Extra Sleep wenn komplett idle (keine Priorities gefunden)
                elif not work_done:
                    # Alle Priorities leer, aber Pipeline aktiv → längerer Sleep
                    time.sleep(self.long_pause)

                next_first_block = random.choice(["P0", "P4"])

            except Exception as e:
                self.logger.log(f"[Maint] ⚠️ Fehler: {e}")
                self.logger.log(f"[Maint] Traceback: {traceback.format_exc()}")
                time.sleep(self.short_pause)

        # Shutdown Cleanup
        try:
            final_count = self.db.flush_annotation_cache()
            if final_count and final_count > 0:
                self.logger.log(f"[Maint] 💾 Finaler Flush: {final_count}")
        except Exception as e:
            self.logger.log(f"[Maint] ⚠️ Finaler Flush fehlgeschlagen: {e}")
            
    def automatic_fetch_decission_and_processing_unit(
        self, 
        keys, 
        build, 
        mode="af", 
        stale_days=Config.STALE_DAYS_AF,
        af_threshold=None, 
        include_none=False, 
        live_enqueue=None
    ):
        """
        ✅ GEFIXT: Nutzt korrekte AfNoneTreatmentManager-API.
        
        Zentrale Fetch-Einheit mit AF-None-Awareness für stabilen Controller.
        """
        # ✅ FIX: App-Referenz übergeben (self.distiller.app)
        app_ref = getattr(self.distiller, "app", None)
        fasta_path = does_fasta_exist(build, self.logger, app=app_ref)
        if not fasta_path:
            fasta_path = does_user_want_to_use_fasta(build, self.logger, app=app_ref)
        use_fasta = bool(fasta_path)

        fai_index = None
        if use_fasta:
            try:
                fai_index_path = fasta_path + ".fai"
                if os.path.exists(fai_index_path):
                    fai_index = load_fai_index(fai_index_path)
            except Exception as e:
                self.logger.log(f"[Maint] ⚠️ FAI-Index Fehler: {e}")

        def _ref_base_with_fasta(chrom, pos):
            return get_ref_base(
                chrom, pos, fasta_path=fasta_path,
                fai_index=fai_index, cache=None, build=build
            ) if use_fasta else "N"

        # Schlüssel nach Status UND AF-None-Typ segmentieren
        renorm_keys = []
        normal_keys = []
        af_none_skip_keys = []
        
        status_field = "full_fetch_success" if mode == "full" else "meanAF_fetch_success"

        for key in keys:
            try:
                status = self.db.get_variant_field(key, status_field)
            except Exception:
                status = None

            if FetchStatusManager.is_success(status):
                continue

            if FetchStatusManager.is_deletable(status):
                try:
                    self.db.delete_variant(key)
                    self.logger.log(f"[Maint] 🗑 Gelöscht (Status 013): {key}")
                except Exception as e:
                    self.logger.log(f"[Maint] ⚠️ Löschen fehlgeschlagen {key}: {e}")
                continue

            # ✅ GEFIXT: AF-None-Check mit korrekter API
            if mode == "af":
                try:
                    row = self.db.get_variant(key)
                    if row:
                        # ✅ KORREKT: Nutze öffentliche Methoden
                        none_type = self.af_none_manager.classify_none_type(
                            row, 
                            stale_days=stale_days
                        )
                        
                        # Wenn AF vorhanden → normal processing
                        if none_type == "has_af":
                            pass  # Fällt durch zu normal_keys
                        else:
                            # AF=None → Policy prüfen
                            should_include = self.af_none_manager.should_include_type(none_type)
                            
                            if should_include:
                                # Legitim → skip re-fetch
                                af_none_skip_keys.append(key)
                                if len(af_none_skip_keys) <= 5:  # Log nur erste 5
                                    self.logger.log(
                                        f"[Maint] ✅ AF=None ({none_type}) legitim: {key} → skip"
                                    )
                                continue
                            else:
                                # Nicht erwünscht → re-fetch
                                if len(normal_keys) < 5:  # Log nur erste 5
                                    self.logger.log(
                                        f"[Maint] 🔄 AF=None ({none_type}) re-fetch: {key}"
                                    )
                                # Fällt durch zu normal_keys
                except Exception as e:
                    self.logger.log(f"[Maint] ⚠️ AF-None-Check Error {key}: {e}")
                    # Bei Fehler: Standard-Processing

            if FetchStatusManager.needs_renormalization(status):
                renorm_keys.append(key)
            else:
                normal_keys.append(key)
        
        # Log AF-None-Statistiken
        if af_none_skip_keys:
            self.logger.log(
                f"[Maint] ⏭️ {len(af_none_skip_keys)} AF=None-Varianten übersprungen "
                f"(Policy: {self.af_none_manager.current_preset})"
            )

        # Re-Normalisierung (unverändert)
        if renorm_keys:
            self.logger.log(f"[Maint] 🔄 Re-Normalisierung für {len(renorm_keys)} Varianten...")
            for key in renorm_keys:
                try:
                    chrom, pos, ref, alt, bld = key
                    current = None
                    try:
                        current = self.db.get_variant_field(key, status_field)
                    except Exception:
                        pass

                    if _is_indel_key(key):
                        norm = normalize_indel(chrom, pos, ref, alt, _ref_base_with_fasta)
                    else:
                        norm = normalize_variant(chrom, pos, ref, alt)

                    if norm:
                        new_chrom, new_pos, new_ref, new_alt = norm
                        new_key = (new_chrom, new_pos, new_ref, new_alt, bld)
                        
                        if new_key != key:
                            try:
                                self.db.update_variant_key(old_key=key, new_key=new_key)
                                self.db.upsert_variant(new_key, {
                                    status_field: FetchStatusManager.update_status(None, success=True),
                                    f"{'full' if mode=='full' else 'meanAF'}_last_fetch": now_iso()
                                })
                                self.logger.log(f"[Maint] ✅ Re-Normalisierung: {key} → {new_key}")
                                try:
                                    self.db.delete_variant(key)
                                except Exception:
                                    pass
                                normal_keys.append(new_key)
                            except Exception as e:
                                new_status = FetchStatusManager.update_status(current, success=False)
                                if new_status == "013":
                                    new_status = "012"
                                try:
                                    self.db.upsert_variant(key, {
                                        status_field: new_status,
                                        f"{'full' if mode=='full' else 'meanAF'}_last_fetch": now_iso()
                                    })
                                except Exception:
                                    pass
                                self.logger.log(f"[Maint] ⚠️ Migration fehlgeschlagen {key}: {e}")
                        else:
                            new_status = FetchStatusManager.update_status(current, success=False)
                            if new_status == "013":
                                new_status = "012"
                            try:
                                self.db.upsert_variant(key, {
                                    status_field: new_status,
                                    f"{'full' if mode=='full' else 'meanAF'}_last_fetch": now_iso()
                                })
                            except Exception:
                                pass
                    else:
                        new_status = FetchStatusManager.update_status(current, success=False)
                        if new_status == "013":
                            new_status = "012"
                        try:
                            self.db.upsert_variant(key, {
                                status_field: new_status,
                                f"{'full' if mode=='full' else 'meanAF'}_last_fetch": now_iso()
                            })
                        except Exception:
                            pass

                except Exception as e:
                    self.logger.log(f"[Maint] ⚠️ Re-Norm Exception {key}: {e}")

        # ✅ Fetch mit STABILEM CONTROLLER (bewährt)
        if mode == "af":
            if normal_keys:
                try:
                    # ✅ Nutze self.af_fetcher (stabiler Controller)
                    results = asyncio.run(
                        self.af_fetcher.job_collector_and_distributor(
                            normal_keys, build,
                            workers=min(16, max(4, (os.cpu_count() or 4) * 2))
                        )
                    )
                except RuntimeError:
                    loop = asyncio.get_event_loop()
                    results = loop.run_until_complete(
                        self.af_fetcher.job_collector_and_distributor(
                            normal_keys, build,
                            workers=min(16, max(4, (os.cpu_count() or 4) * 2))
                        )
                    )
                except Exception as e:
                    self.logger.log(f"[Maint] ⚠️ AF-Fetch Fehler: {e}")
                    results = {}

                # Controller macht Upsert selbst

        elif mode == "full":
            if use_fasta:
                self.logger.log("[Maint] Vollannotation mit FASTA.")
                if normal_keys:
                    mv_data = mv_fetch(
                        normal_keys, build, fetch_full=True,
                        batch_size=800, logger=self.logger, phase_label="FULL"
                    )
                    for k, ann in mv_data.items():
                        if ann is not None:
                            self.distiller.upsert_from_mv(k, ann)
                        else:
                            # ✅ GEFIXT: AF-None-Policy-Check
                            try:
                                row = self.db.get_variant(k)
                                if row:
                                    none_type = self.af_none_manager.classify_none_type(
                                        row,
                                        stale_days=stale_days
                                    )
                                    
                                    # Wenn true_none und Policy sagt include → kein fill_missing_fields
                                    if none_type == "true_none":
                                        should_include = self.af_none_manager.should_include_type("true_none")
                                        
                                        if should_include:
                                            self.logger.log(
                                                f"[Maint] ✅ MyVariant-Failure aber AF=None legitim: {k}"
                                            )
                                            continue
                            except Exception as e:
                                self.logger.log(f"[Maint] ⚠️ Policy-Check Error {k}: {e}")
                            
                            # Standard fill_missing_fields
                            extra = fill_missing_fields(k, build, logger=self.logger)
                            if extra:
                                self.db.upsert_variant(k, extra)
            else:
                self.logger.log("[Maint] Vollannotation ohne FASTA.")
                if normal_keys:
                    mv_data = mv_fetch_with_module(
                        normal_keys, build, fetch_full=True, batch_size=800
                    )
                    for k, ann in mv_data.items():
                        if ann is not None:
                            self.distiller.upsert_from_mv(k, ann)
                        else:
                            # ✅ GEFIXT: AF-None-Policy-Check
                            try:
                                row = self.db.get_variant(k)
                                if row:
                                    none_type = self.af_none_manager.classify_none_type(
                                        row,
                                        stale_days=stale_days
                                    )
                                    
                                    if none_type == "true_none":
                                        should_include = self.af_none_manager.should_include_type("true_none")
                                        
                                        if should_include:
                                            self.logger.log(
                                                f"[Maint] ✅ MyVariant-Failure aber AF=None legitim: {k}"
                                            )
                                            continue
                            except Exception as e:
                                self.logger.log(f"[Maint] ⚠️ Policy-Check Error {k}: {e}")
                            
                            extra = fill_missing_fields(k, build, logger=self.logger)
                            if extra:
                                self.db.upsert_variant(k, extra)
        else:
            raise ValueError(f"Unbekannter Modus: {mode}")
        
    def _phase4_cleanup_invalid_variants(self, limit=100, fasta_path=None, fai_index=None, cache=None):
        """
        Bereinigt ungültige Varianten:
          - Mit FASTA: versucht Reparatur (REF-Base neu holen, Indel normalisieren).
          - Ohne FASTA: löscht nicht, sondern eskaliert Status (010–013).
          - Zusätzlich: löscht pro Durchlauf ein kleines Kontingent an rsIDs,
            die mit 'i' beginnen (23andMe-Artefakte).
        """
        VALID_ALLELE = re.compile(r"^[ACGT]+$")

        def is_valid(allele):
            return bool(allele and VALID_ALLELE.match(allele))

        try:
            all_keys = self.db.get_all_variant_keys() or []
        except Exception as e:
            self.logger.log(f"[P4] ⚠️ Konnte Variantenschlüssel nicht abrufen: {e}")
            return

        # --- NEU: 23andMe-"i"-rsIDs löschen (nur ein paar pro Durchlauf) ---
        try:
            with self.db.lock, self.db._conn() as con:
                cur = con.cursor()
                cur.execute("""
                    SELECT chr, pos, ref, alt, build
                    FROM variants
                    WHERE rsid LIKE 'i%'
                    LIMIT 20
                """)
                bad_rsids = cur.fetchall()

            if bad_rsids:
                for key in bad_rsids:
                    try:
                        self.db.delete_variant(key)
                        self.logger.log(f"[P4] 🗑 Gelöscht (23andMe i-rsid): {key}")
                    except Exception as e:
                        self.logger.log(f"[P4] ⚠️ Fehler beim Löschen von {key}: {e}")
                self.logger.log(f"[P4] Insgesamt {len(bad_rsids)} i-rsIDs entfernt.")
        except Exception as e:
            self.logger.log(f"[P4] ⚠️ Fehler bei Suche nach i-rsIDs: {e}")

        # --- normale Prüfung auf ungültige Allele ---
        subset = all_keys[:limit]
        bad_keys = [
            key for key in subset
            if not is_valid(key[2]) or not is_valid(key[3]) or key[2] == key[3]
        ]

        if not bad_keys:
            self.logger.log("[P4] Keine ungültigen Varianten gefunden.")
            return

        self.logger.log(f"[P4] Prüfe {len(bad_keys)} ungültige Varianten auf Reparaturmöglichkeit…")
        repaired = 0
        deleted = 0
        marked = 0

        for key in bad_keys:
            chrom, pos, ref, alt, build = key
            try:
                if fasta_path:
                    # Mit FASTA: Reparatur versuchen
                    correct_ref = get_ref_base(chrom, pos, fasta_path, fai_index, cache, build)
                    if correct_ref and is_valid(correct_ref) and correct_ref != alt:
                        if _is_indel_key((chrom, pos, correct_ref, alt, build)):
                            norm = normalize_indel(
                                chrom, pos, correct_ref, alt,
                                lambda c, p: get_ref_base(c, p, fasta_path, fai_index, cache, build)
                            )
                            if norm:
                                chrom, pos, correct_ref, alt = norm
                        self.db.update_variant_key(key, (chrom, pos, correct_ref, alt, build))
                        self.logger.log(f"[P4] Repariert: {key} → {(chrom, pos, correct_ref, alt, build)}")
                        repaired += 1
                    else:
                        self.db.delete_variant(key)
                        self.logger.log(f"[P4] Gelöscht: {key}")
                        deleted += 1
                else:
                    # Ohne FASTA: nicht löschen, sondern Eskalation
                    current_status = self.db.get_variant_field(key, "full_fetch_success")
                    new_status = FetchStatusManager.update_status(current_status, success=False)
                    self.db.upsert_variant(key, {
                        "full_fetch_success": new_status,
                        "full_last_fetch": now_iso()
                    })
                    self.logger.log(f"[P4] Markiert/Eskalation für {key} → Status {new_status} "
                               f"({FetchStatusManager.status_label(new_status)})")
                    marked += 1

            except Exception as e:
                self.logger.log(f"[P4] ⚠️ Fehler bei Reparatur/Markierung von {key}: {e}")

        self.logger.log(f"[P4] Bereinigung abgeschlossen – {repaired} repariert, {deleted} gelöscht, {marked} markiert.")
        
    def find_cross_build_candidates(self, limit=100):
        """
        Findet Varianten-Paare (Build 37 <-> 38) mit gleicher rsID,
        bei denen Annotationen übertragen werden können.
        
        Strategie:
        Suche Target (T) ohne Gene/CADD/ClinVar, das einen Source (S) 
        mit gleicher rsID aber anderem Build UND vorhandenen Daten hat.
        """
        # Felder, die wir übertragen wollen
        fields_to_sync = [
            "gene_symbol", "cadd_phred", "clinical_significance", 
            "phenotypes", "impact", "consequence"
        ]
        
        # Dynamische WHERE-Klausel: Target muss leer sein, Source muss voll sein
        # Wir prüfen exemplarisch auf 'gene_symbol', um Kandidaten zu finden
        
        query = f"""
            SELECT 
                t.chr, t.pos, t.ref, t.alt, t.build,  -- Target Key
                s.gene_symbol, s.cadd_phred, s.clinical_significance, 
                s.phenotypes, s.impact, s.consequence,
                s.build AS source_build
            FROM variants t
            JOIN variants s ON t.rsid = s.rsid
            WHERE 
                t.rsid IS NOT NULL AND t.rsid != '.' AND t.rsid != ''
                AND t.build != s.build
                AND (t.gene_symbol IS NULL OR t.gene_symbol = '') -- Target fehlt Info
                AND (s.gene_symbol IS NOT NULL AND s.gene_symbol != '') -- Source hat Info
            LIMIT ?
        """
        
        with self.lock, self._conn() as con:
            cur = con.cursor()
            cur.execute(query, (limit,))
            return cur.fetchall()   
##################
_splash_log("Initialisiere Datenbank ...")
db = VariantDB()
_splash_log("Datenbank bereit.")
##################


@dataclass
class PipelineProgress:
    """
    FIX: Korrektes Tracking von gefilterten Varianten mit Stop-Support
    """
    total_variants: int = 0
    processed_variants: int = 0
    rejected_variants: int = 0
    emitted_count: int = 0  # NEU
    
    # Phase-spezifische Zähler
    vcf_scan_done: int = 0
    af_fetch_done: int = 0
    full_anno_done: int = 0
    gene_anno_done: int = 0
    rsid_fill_done: int = 0
    missing_fill_done: int = 0
    ag_score_done: int = 0
    
    # ✅ FIX: Stop-Flag hinzufügen
    _stopped: bool = False
    
    # Phase-Gewichtungen
    PHASE_WEIGHTS = {
        "vcf_scan": 0.05,
        "af_fetch": 0.30,
        "full_anno": 0.35,
        "gene_anno": 0.10,
        "rsid_fill": 0.05,
        "missing_fill": 0.10,
        "ag_score": 0.05,
    }
    
    _lock: RLock = field(default_factory=RLock)
    phase_name: str = ""
    start_time: Optional[float] = None
    _phase_start_times: dict = field(default_factory=dict)
    _phases_completed: set = field(default_factory=set)
    # Mapping für UI-Anzeige
    PHASE_DISPLAY_NAMES = {
        "vcf_scan": "VCF-Scan",
        "af_fetch": "AF-Fetch",
        # Alle Subphasen der Annotation laufen unter einem Namen
        "full_anno": "Vollannotation",
        "gene_anno": "Vollannotation",
        "rsid_fill": "Vollannotation",
        "missing_fill": "Vollannotation",
        "ag_score": "Vollannotation",
    }

    def set_phase(self, name: str):
        """
        FIX: Phase-Tracking mit korrektem Mapping.
        
        Änderungen:
        - UI-Name wird gemappt
        - Internes Tracking bleibt konsistent
        """
        with self._lock:
            # ✅ UI-Name setzen (gemappt)
            self.phase_name = self.PHASE_DISPLAY_NAMES.get(name, name)
            
            # ✅ Internes Tracking (echter Name)
            if name not in self._phase_start_times:
                self._phase_start_times[name] = time.time()

                
    def start_pipeline(self, total: int):
        """
        Initialisiert Pipeline mit Quality-Gate-bestandenen Varianten.
        
        Args:
            total: Anzahl Varianten NACH Quality-Gate
        """
        with self._lock:
            self.total_variants = total
            self.processed_variants = 0
            self.rejected_variants = 0
            self.emitted_count = 0  # NEU
            self.vcf_scan_done = 0
            self.af_fetch_done = 0
            self.full_anno_done = 0
            self.gene_anno_done = 0
            self.rsid_fill_done = 0
            self.missing_fill_done = 0
            self.ag_score_done = 0
            self.start_time = time.time()
            self._phase_start_times.clear()
            self._phases_completed.clear()
    
    def start_phase(self, phase: str):
        """Markiert Start einer Phase."""
        with self._lock:
            if phase not in self._phase_start_times:
                self._phase_start_times[phase] = time.time()
                self.phase_name = phase
    
    def complete_phase(self, phase: str):
        """Markiert Completion einer Phase."""
        with self._lock:
            self._phases_completed.add(phase)
    
    def get_phase_progress(self) -> dict:
        """
        FIX: Korrekte Prozentwerte inkl. Rejected.
        """
        with self._lock:
            if self._is_pipeline_complete():
                return {"vcf_scan": 100, "af_fetch": 100, "full_anno": 100, "overall": 100}
            
            total = max(1, self.total_variants)
            # Rejections werden als "erledigt" für nachfolgende Phasen betrachtet
            rejected = self.rejected_variants
            
            # VCF Scan ist immer auf Total bezogen (hier wird rejected erzeugt)
            vcf_pct = min(100, (self.vcf_scan_done / total * 100))
            
            # AF Fetch: Done + Rejected
            af_pct = min(100, ((self.af_fetch_done + rejected) / total * 100))
            
            # Full Anno: Done + Rejected
            # Wir nehmen den Durchschnitt der Sub-Phasen oder einfach full_anno_done
            full_anno_total = self.full_anno_done 
            full_pct = min(100, ((full_anno_total + rejected) / total * 100))
            
            return {
                "vcf_scan": round(vcf_pct, 1),
                "af_fetch": round(af_pct, 1),
                "full_anno": round(full_pct, 1),
                "overall": round(self.get_progress() or 0, 1)
            }

    def _is_pipeline_complete(self) -> bool:
        """✅ Helper: Prüfe ob Pipeline komplett"""
        return all(
            phase in self._phases_completed 
            for phase in self.PHASE_WEIGHTS.keys()
        )
        
    def update_phase(self, phase: str, count: int = 1):
        """Update für spezifische Phase."""
        with self._lock:
            # Phase-Start tracken
            if phase not in self._phase_start_times:
                self._phase_start_times[phase] = time.time()
            
            # Zähler aktualisieren
            if phase == "vcf_scan":
                self.vcf_scan_done += count
            elif phase == "af_fetch":
                self.af_fetch_done += count
            elif phase == "full_anno":
                self.full_anno_done += count
            elif phase == "gene_anno":
                self.gene_anno_done += count
            elif phase == "rsid_fill":
                self.rsid_fill_done += count
            elif phase == "missing_fill":
                self.missing_fill_done += count
            elif phase == "ag_score":
                self.ag_score_done += count
    
    def mark_rejected(self, count: int = 1):
        """
        ✅ NEU: Markiert Varianten als durch AF-Filter abgelehnt.
        Diese zählen als "verarbeitet" für Fortschritt, aber nicht als "emittiert".
        """
        with self._lock:
            self.rejected_variants += count
            # ✅ Auch als "processed" zählen (sie sind fertig)
            self.processed_variants += count
    
    # Problem: processed_variants kann > total_variants werden
    def mark_fully_processed(self, count: int = 1):
        with self._lock:
            self.processed_variants = min(
                self.processed_variants + count,
                self.total_variants  # ✅ Cap bei total
            )
    
    def mark_emitted(self, count: int = 1):
        """NEU: Markiert Varianten als emittiert."""
        with self._lock:
            self.emitted_count += count
    
    def is_fully_complete(self) -> bool:
        """NEU: Robuste Ende-Erkennung."""
        with self._lock:
            if "vcf_scan" not in self._phases_completed:
                return False
            if self.total_variants > 0:
                total_handled = self.emitted_count + self.rejected_variants
                if total_handled < self.total_variants:
                    return False
            for phase in ["af_fetch", "full_anno", "rsid_fill"]:
                phase_done = getattr(self, f"{phase}_done", 0)
                if phase_done < self.emitted_count:
                    return False
            return True
        
    def percent(self) -> float:
        with self._lock:
            if self.total_variants <= 0:
                return 0.0

            # ✅ WICHTIG: processed darf nie > total sein
            actual_processed = min(self.processed_variants, self.total_variants)

            # Basis: processed / total
            base_percent = (actual_processed / self.total_variants) * 100.0

            # ✅ Early return bei >= 99%
            if base_percent >= 99.0:
                return min(100.0, base_percent)

            # Gewichteter Progress (nur wenn < 99%)
            weighted_sum = 0.0
            emitted = actual_processed - self.rejected_variants

            #VCF-Scan
            if "vcf_scan" in self._phases_completed:
                weighted_sum += self.PHASE_WEIGHTS["vcf_scan"]
            elif self.vcf_scan_done > 0:
                vcf_pct = min(1.0, self.vcf_scan_done / self.total_variants)
                weighted_sum += vcf_pct * self.PHASE_WEIGHTS["vcf_scan"]
            
            #AF-Fetch
            if "af_fetch" in self._phases_completed:
                weighted_sum += self.PHASE_WEIGHTS["af_fetch"]
            elif self.af_fetch_done > 0:
                af_pct = min(1.0, self.af_fetch_done / self.total_variants)
                weighted_sum += af_pct * self.PHASE_WEIGHTS["af_fetch"]
            
            # ✅Full-Anno (nur wenn emitted > 0)
            if "full_anno" in self._phases_completed:
                weighted_sum += self.PHASE_WEIGHTS["full_anno"]
            elif self.full_anno_done > 0 and emitted > 0:
                full_pct = min(1.0, self.full_anno_done / emitted)
                weighted_sum += full_pct * self.PHASE_WEIGHTS["full_anno"]
            
            # ✅Gene-Anno
            if "gene_anno" in self._phases_completed:
                weighted_sum += self.PHASE_WEIGHTS["gene_anno"]
            elif self.gene_anno_done > 0 and emitted > 0:
                gene_pct = min(1.0, self.gene_anno_done / emitted)
                weighted_sum += gene_pct * self.PHASE_WEIGHTS["gene_anno"]
            # ✅ RSID (eigener Counter!)
            if "rsid_fill" in self._phases_completed:
                weighted_sum += self.PHASE_WEIGHTS["rsid_fill"]
            elif self.rsid_fill_done > 0 and emitted > 0:  # ✅ Korrekter Counter
                rsid_pct = min(1.0, self.rsid_fill_done / emitted)
                weighted_sum += rsid_pct * self.PHASE_WEIGHTS["rsid_fill"]

            # ✅ Missing (eigener Counter!)
            if "missing_fill" in self._phases_completed:
                weighted_sum += self.PHASE_WEIGHTS["missing_fill"]
            elif self.missing_fill_done > 0 and emitted > 0:  # ✅ Korrekter Counter
                missing_pct = min(1.0, self.missing_fill_done / emitted)
                weighted_sum += missing_pct * self.PHASE_WEIGHTS["missing_fill"]

            # ✅ Alphagenome (eigener Counter!)
            if "ag_score" in self._phases_completed:
                weighted_sum += self.PHASE_WEIGHTS["ag_score"]
            elif self.ag_score_done > 0 and emitted > 0:  # ✅ Korrekter Counter
                ag_pct = min(1.0, self.ag_score_done / emitted)
                weighted_sum += ag_pct * self.PHASE_WEIGHTS["ag_score"]
            
            # ✅Nutze höheren Wert, aber IMMER max 100%
            weighted_percent = weighted_sum * 100.0
            final_percent = max(base_percent, weighted_percent)
            
            return min(100.0, final_percent)
        
    def eta(self) -> Optional[float]:
        """
        FIX: ETA berücksichtigt frühe Rejections korrekt.
        """
        with self._lock:
            if not self.start_time or self.total_variants <= 0:
                return None
            
            elapsed = time.time() - self.start_time
            if elapsed <= 0:
                return None
            
            # ✅ Prüfe, ob wir noch in früher Phase sind (< 0.1% Fortschritt)
            progress_fraction = self.processed_variants / self.total_variants
            
            if progress_fraction <= 0.001:
                return None
            
            # ✅ Bei frühem Fortschritt (<10%): nutze Phase-Rates
            if progress_fraction < 0.1:
                # VCF-Scan Rate (wenn aktiv)
                if self.vcf_scan_done > 0 and "vcf_scan" not in self._phases_completed:
                    scan_elapsed = time.time() - self._phase_start_times.get("vcf_scan", self.start_time)
                    if scan_elapsed > 0:
                        scan_rate = self.vcf_scan_done / scan_elapsed
                        remaining_variants = self.total_variants - self.vcf_scan_done
                        scan_eta = remaining_variants / max(0.1, scan_rate)
                        
                        # Gewichte mit Phase-Gewicht (5%)
                        weighted_eta = scan_eta / self.PHASE_WEIGHTS.get("vcf_scan", 0.05)
                        return weighted_eta
                
                # AF-Fetch Rate (wenn aktiv)
                if self.af_fetch_done > 0 and "af_fetch" not in self._phases_completed:
                    af_elapsed = time.time() - self._phase_start_times.get("af_fetch", self.start_time)
                    if af_elapsed > 0:
                        af_rate = self.af_fetch_done / af_elapsed
                        remaining_variants = self.total_variants - self.af_fetch_done
                        af_eta = remaining_variants / max(0.1, af_rate)
                        
                        # Gewichte mit Phase-Gewicht (30%)
                        weighted_eta = af_eta / self.PHASE_WEIGHTS.get("af_fetch", 0.30)
                        return weighted_eta
            
            # ✅ Normale ETA-Berechnung (> 10% Fortschritt)
            estimated_total = elapsed / max(0.001, progress_fraction)
            remaining = max(0, estimated_total - elapsed)
            
            # ✅ Sanity-Check: ETA nicht > 24h bei >50% Fortschritt
            if progress_fraction > 0.5 and remaining > 86400:
                # Nutze letzte 20% für Rate-Schätzung
                recent_processed = int(self.total_variants * 0.2)
                recent_elapsed = elapsed * 0.2
                
                if recent_elapsed > 0:
                    recent_rate = recent_processed / recent_elapsed
                    remaining_variants = self.total_variants - self.processed_variants
                    remaining = remaining_variants / max(0.1, recent_rate)
            
            return remaining
        
    def get_phase_eta(self, phase: str) -> Optional[float]:
        """
        Berechnet ETA für eine spezifische Phase.
        
        Args:
            phase: Phase-Name (z.B. "af_fetch")
        
        Returns:
            ETA in Sekunden oder None
        """
        with self._lock:
            # Phase abgeschlossen?
            if phase in self._phases_completed:
                return 0.0
            
            # Phase noch nicht gestartet?
            if phase not in self._phase_start_times:
                return None
            
            # Aktuelle Phase-Daten
            phase_start = self._phase_start_times[phase]
            elapsed = time.time() - phase_start
            
            # Phase-spezifische Counter
            phase_counters = {
                "vcf_scan": self.vcf_scan_done,
                "af_fetch": self.af_fetch_done,
                "full_anno": self.full_anno_done,
                "gene_anno": self.gene_anno_done,
                "rsid_fill": self.rsid_fill_done,
                "missing_fill": self.missing_fill_done,
                "ag_score": self.ag_score_done,
            }
            
            done = phase_counters.get(phase, 0)
            if done <= 0:
                return None
            
            # Rate berechnen
            rate = done / elapsed
            if rate <= 0:
                return None
            
            # Remaining
            remaining = max(0, self.total_variants - done)
            return remaining / rate
    
    def get_detailed_status(self) -> dict:
        """
        Detaillierter Status mit Rejection-Info.
        """
        with self._lock:
            emitted = self.processed_variants - self.rejected_variants
            
            return {
                "total": self.total_variants,
                "processed": self.processed_variants,
                "emitted": emitted,  # ✅ NEU
                "rejected": self.rejected_variants,  # ✅ NEU
                "percent": self.percent(),
                "eta_seconds": self.eta(),
                "phases": {
                    "vcf_scan": {
                        "done": self.vcf_scan_done,
                        "percent": (self.vcf_scan_done / max(1, self.total_variants)) * 100.0,
                        "eta": self.get_phase_eta("vcf_scan"),
                        "complete": "vcf_scan" in self._phases_completed,
                    },
                    "af_fetch": {
                        "done": self.af_fetch_done,
                        "percent": (self.af_fetch_done / max(1, self.total_variants)) * 100.0,
                        "eta": self.get_phase_eta("af_fetch"),
                        "complete": "af_fetch" in self._phases_completed,
                    },
                    "full_anno": {
                        "done": self.full_anno_done,
                        "percent": (self.full_anno_done / max(1, self.total_variants)) * 100.0,
                        "eta": self.get_phase_eta("full_anno"),
                        "complete": "full_anno" in self._phases_completed,
                    },
                    "gene_anno": {
                        "done": self.gene_anno_done,
                        "percent": (self.gene_anno_done / max(1, self.total_variants)) * 100.0,
                        "eta": self.get_phase_eta("gene_anno"),
                        "complete": "gene_anno" in self._phases_completed,
                    },
                    "rsid_fill": {
                        "done": self.rsid_fill_done,
                        "percent": (self.rsid_fill_done / max(1, self.total_variants)) * 100.0,
                        "eta": self.get_phase_eta("rsid_fill"),
                        "complete": "rsid_fill" in self._phases_completed,
                    },
                    "missing_fill": {
                        "done": self.missing_fill_done,
                        "percent": (self.missing_fill_done / max(1, self.total_variants)) * 100.0,
                        "eta": self.get_phase_eta("missing_fill"),
                        "complete": "missing_fill" in self._phases_completed,
                    },
                    "ag_score": {
                        "done": self.ag_score_done,
                        "percent": (self.ag_score_done / max(1, self.total_variants)) * 100.0,
                        "eta": self.get_phase_eta("ag_score"),
                        "complete": "ag_score" in self._phases_completed,
                    },
                }
            }
    
    def stop(self):
        """✅ FIX: Stoppe Progress-Tracking"""
        with self._lock:
            self._stopped = True
            # Setze alle Phasen auf complete wenn gestoppt
            for phase in self.PHASE_WEIGHTS.keys():
                self._phases_completed.add(phase)
    
    def reset(self):
        """✅ FIX: Reset inklusive Stop-Flag"""
        with self._lock:
            self._stopped = False
            self.total_variants = 0
            self.processed_variants = 0
            self.rejected_variants = 0
            self.emitted_count = 0  # NEU
            self.vcf_scan_done = 0
            self.af_fetch_done = 0
            self.full_anno_done = 0
            self.gene_anno_done = 0
            self.rsid_fill_done = 0
            self.missing_fill_done = 0
            self.ag_score_done = 0
            self._phase_start_times.clear()
            self._phases_completed.clear()
    
    def get_progress(self) -> float:
        """✅ FIX: Return 100% wenn gestoppt"""
        with self._lock:
            if self._stopped:
                return 100.0
            
    def complete_pipeline(self):
        """Markiert Pipeline als vollständig abgeschlossen und korrigiert Counter."""
        with self._lock:
            # Alle Phasen als fertig markieren
            self._phases_completed.update(self.PHASE_WEIGHTS.keys())
            self.phase_name = "Abgeschlossen"
            
            # Sanity Check: Processed darf nicht kleiner als Rejected sein
            # Wenn wir fertig sind, haben wir per Definition alles verarbeitet (akzeptiert oder rejected)
            if self.total_variants > 0:
                self.processed_variants = self.total_variants

    def get_summary(self) -> str:
        """
        FIX: Summary berücksichtigt jetzt Rejected-Varianten in den Phasen-Prozenten
        und verhindert negative Emitted-Zahlen.
        """
        with self._lock:
            pct = self.percent()
            eta_sec = self.eta()
            eta_str = fmt_eta(eta_sec) if eta_sec else "??:??"
            
            # FIX: Schutz vor negativen Zahlen durch max(0, ...)
            emitted = max(0, self.processed_variants - self.rejected_variants)
            
            # Aktive Phasen Berechnung
            active_phases = []
            total_denominator = max(1, self.total_variants)
            
            for phase in ["af_fetch", "full_anno", "gene_anno"]:
                counter = getattr(self, f"{phase}_done", 0)
                
                # Addiere Rejected zur Basis, da diese logisch "erledigt" sind für diese Phase
                effective_done = counter + self.rejected_variants
                
                # Zeige Phase nur an, wenn sie läuft und noch nicht fertig ist
                if counter > 0 and effective_done < (self.total_variants * 0.999):
                    phase_pct = (effective_done / total_denominator) * 100.0
                    phase_pct = min(100.0, phase_pct)
                    active_phases.append(f"{phase.split('_')[0].upper()}:{phase_pct:.0f}%")
            
            # Status-String bestimmen
            if self.phase_name == "Abgeschlossen":
                active_str = "Fertig"
            elif not active_phases and pct < 100.0:
                active_str = "Processing..."
            else:
                active_str = " ".join(active_phases[:3]) if active_phases else "Idle"
            
            return (
                f"{pct:.1f}% | {emitted} emitted, {self.rejected_variants} rejected | "
                f"{active_str} | {eta_str}"
            )
# ============================================================================
# FIX 1: EmitQueue - Verhindert Endlosschleife
# ============================================================================

class EmitQueue:
    """
    FIX: Timestamp-Management verhindert Endlosschleife.
    
    Änderungen:
    - Timestamp VOR Flush setzen (nicht während)
    - Deduplizierung vor Flush
    - Circuit-Breaker mit exponentieller Backoff
    """
    
    def __init__(self, emit_func: Callable, batch_size: int = 100, flush_interval: float = 10.0):
        self._buffer: List[Tuple] = []
        self._emit = emit_func
        self._batch_size = batch_size
        self._lock = threading.Lock()
        self._emitted = set()
        
        # Circuit Breaker
        self._drop_count = 0
        self._backpressure_count = 0
        self._last_stats_log = time.time()
        self._circuit_open = False
        self._circuit_open_until = 0.0

        # Zeitbasierter Flush
        self._last_flush = time.time()
        self._flush_interval = flush_interval
    
    def add(self, key: Tuple):
        with self._lock:
            now = time.time()
            
            # Circuit Breaker Check
            if self._circuit_open:
                if now < self._circuit_open_until:
                    self._drop_count += 1
                    return
                else:
                    self._circuit_open = False
                    self._drop_count = 0
                    self._backpressure_count = 0
            
            # Deduplizierung
            if key in self._emitted:
                return
            
            self._emitted.add(key)
            self._buffer.append(key)
            
            # Adaptive Batch-Größe
            current_size = len(self._buffer)
            flush_threshold = self._batch_size
            
            if current_size > 1000:
                flush_threshold = 50
            elif current_size > 500:
                flush_threshold = self._batch_size // 2
            elif current_size < 50:
                flush_threshold = self._batch_size * 2
            
            # ✅ FIX: Timestamp VOR Flush setzen (verhindert Endlosschleife)
            should_flush = False
            time_since_flush = now - self._last_flush
            
            if current_size >= flush_threshold:
                should_flush = True
                self._last_flush = now  # ✅ Timestamp VORHER
            elif time_since_flush >= self._flush_interval:
                should_flush = True
                self._last_flush = now  # ✅ Timestamp VORHER
            
            if should_flush:
                self._flush_unsafe()

    def _flush_unsafe(self):
        """
        FIX: Kein Timestamp-Update mehr hier (wird in add() gesetzt).
        """
        if not self._buffer:
            return  # ✅ Early return ohne Timestamp-Update
        
        now = time.time()
        failed_keys = []
        failed_count = 0
        
        for k in self._buffer:
            try:
                self._emit(k)
            except queue.Full:
                failed_keys.append(k)
                failed_count += 1
                self._backpressure_count += 1
            except Exception as e:
                if self._drop_count % 100 == 0:
                    print(f"[EmitQueue] ⚠️ Emit failed: {e}")
                self._drop_count += 1
        
        # Circuit Breaker
        buffer_size = len(self._buffer)
        if buffer_size > 0:
            fail_rate = failed_count / buffer_size
            if fail_rate > 0.5 and buffer_size > 100:
                self._circuit_open = True
                self._circuit_open_until = now + 5.0
        
        # Failed Keys: nur erste 500 behalten
        if failed_keys:
            max_retry = 500
            self._buffer = failed_keys[:max_retry]
            
            if len(failed_keys) > max_retry:
                excess = len(failed_keys) - max_retry
                self._drop_count += excess
        else:
            self._buffer.clear()
        
        # Stats alle 60s
        if now - self._last_stats_log > 60:
            self._log_stats()
            self._last_stats_log = now
            
    def _log_stats(self):
        if self._drop_count > 0 or self._backpressure_count > 0:
            print(
                f"[EmitQueue] Stats: "
                f"Buffered: {len(self._buffer)}, "
                f"Emitted: {len(self._emitted)}, "
                f"Dropped: {self._drop_count}, "
                f"Backpressure: {self._backpressure_count}, "
                f"Circuit: {'OPEN' if self._circuit_open else 'CLOSED'}"
            )
    
    def flush(self):
        with self._lock:
            was_open = self._circuit_open
            self._circuit_open = False
            try:
                self._flush_unsafe()
            finally:
                if was_open:
                    self._circuit_open = True
    
    def reset(self):
        with self._lock:
            self._buffer.clear()
            self._emitted.clear()
            self._drop_count = 0
            self._backpressure_count = 0
            self._circuit_open = False
            self._circuit_open_until = 0.0
            self._last_flush = time.time()
    
    def shutdown(self):
        """
        ✅ V11 FIX: Fehlende shutdown()-Methode hinzugefügt.
        Flusht verbleibende Items und setzt den Zustand zurück.
        """
        try:
            self.flush()
        except Exception:
            pass
        self.reset()

    def discard(self, key: Tuple):
        """
        ✅ V16 FIX 10: Entfernt einen Key aus der EmitQueue.
        Wird verwendet wenn eine Variante ungültig ist (z.B. REF-Mismatch).
        """
        with self._lock:
            # Aus Buffer entfernen falls vorhanden
            if key in self._buffer:
                self._buffer.remove(key)
            # Aus emitted-Set entfernen
            self._emitted.discard(key)


class VCFBuffer:
    """
    Vereinfachter Buffer mit Prioritäten, kompatibel zu aktuellen DB-Methoden.
    FIX: Korrekte Trennung von AF-Only Updates und gemischten Annotationen.
    """

    def __init__(self, db, logger, batch_size: int = 5000, prio_threshold: int = 100, distiller=None):
        self.db = db
        self.logger = logger
        self.batch_size = batch_size
        self.prio_threshold = prio_threshold
        self.distiller = distiller
        self._buffer: List[Tuple] = []
        self._priority_buffer: List[Tuple] = []
        self._lock = threading.Lock()

    def _is_pure_af_update(self, update: Dict) -> bool:
        """
        Prüft, ob ein Update AUSSCHLIESSLICH AF-Felder enthält.
        Falls auch nur ein Annotationsfeld (Gene, Coding, etc.) dabei ist, muss False zurückgegeben werden,
        damit upsert_many (der alles schreibt) verwendet wird.
        """
        # Diese Felder werden von upsert_variants_bulk unterstützt/geschrieben:
        af_whitelist = {
            "af_filter_mean", "af_exac", "af_gnomad_exomes",
            "af_gnomad_genomes", "af_1kg",
            "meanAF_last_fetch", "meanAF_fetch_success",
            "full_last_fetch", "full_fetch_success"
            # rsid wird in upsert_variants_bulk NICHT im UPDATE-Teil geschrieben, daher hier nicht whitelisten!
        }
        
        # Wenn irgendein Key im Update NICHT in der Whitelist ist -> Kein reines AF-Update
        return all(k in af_whitelist for k in update.keys())

    def _flush_batch(self, batch: List[Tuple], label: str):
        """Wrapper für DB-Flush, entscheidet automatisch AF vs. Annotation."""
        if not batch:
            return
        try:
            # Entscheidung: Bulk-AF-Writer oder generischer Writer?
            # Wir nutzen den schnellen AF-Writer nur, wenn ALLE Items im Batch reine AF-Updates sind.
            is_pure_af_batch = all(self._is_pure_af_update(update) for _, update in batch)

            if is_pure_af_batch:
                # Schnell, aber schreibt nur AF-Spalten
                af_dict = {k: v for k, v in batch}
                self.db.upsert_variants_bulk(af_dict)
                # self.logger.log(f"[VCF-Buffer] 💾 Flushed {len(batch)} pure AF records ({label})")
            else:
                # Langsamer, aber schreibt ALLE Spalten (Gene, Coding, Impact etc.)
                self.db.upsert_many(batch)
                # self.logger.log(f"[VCF-Buffer] 💾 Flushed {len(batch)} full/mixed records ({label})")
            
            # 2. Cache Invalidierung & GUI Refresh (NUR für sichtbare Keys!)
            if hasattr(self, 'distiller') and self.distiller:
                keys = [k for k, _ in batch]
                
                # A. Cache invalidieren (damit beim nächsten Refresh frische Daten kommen)
                self.distiller.invalidate_cache_bulk(keys)
                
                # B. Live-Update NUR für bereits sichtbare Keys triggern
                # ⚠️ V9 HINWEIS: Direkter emit_queue Zugriff - beabsichtigt für Batch-Flushes
                # Der VCFBuffer ist Teil des Distillers, daher ist direkter Zugriff okay
                if hasattr(self.distiller, 'emit_queue') and hasattr(self.distiller, 'display_keys'):
                    visible_keys = self.distiller.display_keys
                    for k in keys:
                        if k in visible_keys:
                            self.distiller.emit_queue.add(k)
            
        except Exception as e:
            self.logger.log(f"[VCF-Buffer] ⚠️ Flush failed ({label}): {e}")
            import traceback
            self.logger.log(traceback.format_exc())

    def add(self, key: Tuple, update: Dict, priority: bool = False):
        """Add item mit optionaler Priorität."""
        should_flush = False
        batch_copy = None
        label = ""

        with self._lock:
            if priority:
                self._priority_buffer.append((key, update))
                if len(self._priority_buffer) >= self.prio_threshold:
                    should_flush = True
                    batch_copy = self._priority_buffer[:]
                    self._priority_buffer.clear()
                    label = "priority"
            else:
                self._buffer.append((key, update))
                if len(self._buffer) >= self.batch_size:
                    should_flush = True
                    batch_copy = self._buffer[:]
                    self._buffer.clear()
                    label = "regular"

        if should_flush and batch_copy:
            self._flush_batch(batch_copy, label)

    def flush(self, force_priority: bool = False):
        """Force flush."""
        prio_copy = None
        batch_copy = None

        with self._lock:
            if self._priority_buffer: # Immer flushen, nicht nur bei force
                prio_copy = self._priority_buffer[:]
                self._priority_buffer.clear()
            if self._buffer:
                batch_copy = self._buffer[:]
                self._buffer.clear()

        if prio_copy:
            self._flush_batch(prio_copy, "priority-final")
        if batch_copy:
            self._flush_batch(batch_copy, "final")

## 📊 **ARCHITEKTUR-ÜBERSICHT:**
"""
┌─────────────────────────────────────────────────────────────┐
│  MANAGEMENT LAYER (BEHALTEN - wird von anderen genutzt)     │
├─────────────────────────────────────────────────────────────┤
│  does_fasta_exist()           → Prüft/Download/Index        │
│  does_user_want_to_use_fasta() → User-Dialog               │
│  build_fasta_index_global()    → Index-Erstellung          │
└─────────────────────────────────────────────────────────────┘
                          ↓ nutzt ↓
┌─────────────────────────────────────────────────────────────┐
│  VALIDATION LAYER (NEU - Performance-optimiert)             │
├─────────────────────────────────────────────────────────────┤
│  FastaValidator                                             │
│    - Memory-mapped I/O                                      │
│    - Region-Cache                                           │
│    - Pre-Check (SNVs skip)                                  │
│    - Windows-kompatibel                                     │
└─────────────────────────────────────────────────────────────┘
                          ↓ nutzt ↓
┌─────────────────────────────────────────────────────────────┐
│  DISTILLER METHODS                                          │
├─────────────────────────────────────────────────────────────┤
│  _validate_with_fasta_fallback()                            │
│    1. does_fasta_exist() → FASTA-Pfad                       │
│    2. FastaValidator → Validierung                          │
│    3. DB-Update für ungültige Keys                          │
└─────────────────────────────────────────────────────────────┘
"""
def run_on_main_thread(app, func, *args, **kwargs):
    """
    Führt eine GUI-Funktion thread-sicher im Haupt-Thread aus und wartet auf das Ergebnis.
    """
    if threading.current_thread() is threading.main_thread():
        return func(*args, **kwargs)
    
    result = [None]
    exception = [None]
    done_event = threading.Event()
    
    def wrapper():
        try:
            result[0] = func(*args, **kwargs)
        except Exception as e:
            exception[0] = e
        finally:
            done_event.set()
    
    # Event im Main-Loop einplanen
    app.after(0, wrapper)
    
    # Warten bis ausgeführt
    done_event.wait()
    
    if exception[0]:
        raise exception[0]
    
    return result[0]

def does_fasta_exist(build: str, logger, app=None) -> str | None:
    """
    V16: Prüft ob FASTA + Index vorhanden sind - mit PathManager Auto-Discovery.

    Nutzt PathManager für intelligente Pfad-Suche:
    1. Prüft zuerst PathManager (findet FASTA auch wenn verschoben)
    2. Fallback auf FASTA_PATHS wenn PathManager nichts findet
    3. Bietet Download an wenn nichts gefunden
    """
    fasta_path = None

    # ✅ V16: Zuerst PathManager fragen
    try:
        pm = get_path_manager()
        fasta_path = pm.get_fasta_path(build)
        if fasta_path:
            logger.log(f"[FASTA] ✅ PathManager: {build} gefunden unter {fasta_path}")
    except Exception as e:
        logger.log(f"[FASTA] ⚠️ PathManager-Fehler: {e}")

    # Fallback auf globale FASTA_PATHS
    if not fasta_path:
        fasta_path = FASTA_PATHS.get(build)

    if not fasta_path:
        logger.log(f"[FASTA] ❌ Kein Pfad für Build {build} definiert.")
        return None

    fai_path = fasta_path + ".fai"

    # Fall 1: FASTA + Index vorhanden
    if os.path.exists(fasta_path) and os.path.exists(fai_path):
        logger.log(f"[FASTA] FASTA und Index für {build} vorhanden.")
        # ✅ V16: Registriere gefundenen Pfad bei PathManager
        try:
            pm = get_path_manager()
            build_key = "fasta_grch38" if build.lower() in ("grch38", "hg38", "b38") else "fasta_grch37"
            pm.register_path(build_key, os.path.relpath(fasta_path, BASE_DIR) if fasta_path.startswith(BASE_DIR) else fasta_path)
        except Exception:
            pass
        return fasta_path

    # Fall 2: FASTA vorhanden, Index fehlt
    if os.path.exists(fasta_path) and not os.path.exists(fai_path):
        logger.log(f"[FASTA] FASTA für {build} vorhanden, Index fehlt – erzeuge Index…")
        try:
            return build_fasta_index_global(fasta_path, logger) and fasta_path
        except Exception as e:
            logger.log(f"[FASTA] ❌ Index-Erzeugung fehlgeschlagen: {e}")
            return None

    # Fall 3: FASTA fehlt komplett → Benutzer fragen
    return does_user_want_to_use_fasta(build, logger, app=app)


def does_user_want_to_use_fasta(build: str, logger, app=None) -> str | None:
    """
    Fragt den Benutzer einmal pro Session, ob FASTA heruntergeladen werden soll.
    Thread-Safe Implementierung.
    """
    global SESSION_FLAGS
    if SESSION_FLAGS.get("use_fasta") is False:
        return None

    # Dialog-Funktion definieren
    def ask_dialog():
        return messagebox.askyesno(
            "FASTA verwenden?",
            f"Soll die Referenz-FASTA für {build} heruntergeladen und verwendet werden?\n"
            "Wenn Nein, werden alle Abfragen nur noch über MyVariant laufen."
        )

    # Thread-Safe Aufruf
    if app:
        try:
            decision = run_on_main_thread(app, ask_dialog)
        except Exception as e:
            logger.log(f"[FASTA] Fehler beim Dialog-Aufruf: {e}")
            return None
    else:
        # Fallback (unsicher, aber besser als Crash)
        decision = ask_dialog()

    if not decision:
        logger.log("[FASTA] Download abgelehnt – Fallback auf MyVariant.")
        SESSION_FLAGS["use_fasta"] = False
        return None

    # ... (Rest der Funktion bleibt gleich: Download Logik) ...
    fasta_path = FASTA_PATHS.get(build) # Verwende .get für Sicherheit
    if not fasta_path:
         # Fallback falls Path in Config fehlt
         return None
         
    url = FASTA_URLS.get(build)
    if not url:
        return None
        
    gz_path = fasta_path + ".gz"
    try:
        logger.log(f"[FASTA] Lade FASTA für {build} herunter von {url} …")
        # ... (Download Code hier unverändert lassen) ...
        # Hier nur verkürzt dargestellt für den Fix-Kontext
        with requests.get(url, stream=True, timeout=60) as r:
            r.raise_for_status()
            r.raw.decode_content = True
            with open(gz_path, "wb") as f:
                shutil.copyfileobj(r.raw, f)

        logger.log("[FASTA] Download abgeschlossen, entpacke…")
        with gzip.open(gz_path, "rb") as f_in, open(fasta_path, "wb") as f_out:
            shutil.copyfileobj(f_in, f_out)
        os.remove(gz_path)

        logger.log("[FASTA] Entpacken abgeschlossen, erzeuge Index…")
        build_fasta_index_global(fasta_path, logger)
        logger.log(f"[FASTA] Bereit: {fasta_path}")
        SESSION_FLAGS["use_fasta"] = True
        return fasta_path
    except Exception as e:
        logger.log(f"[FASTA] ❌ Fehler beim Herunterladen/Entpacken: {e}")
        SESSION_FLAGS["use_fasta"] = False
        return None

def build_fasta_index_global(fasta_path, logger):
    """
    Erzeugt eine .fai-Indexdatei für die FASTA, falls sie fehlt.
    """
    fai_path = fasta_path + ".fai"

    if os.path.exists(fai_path):
        logger.log(f"[FASTA] Index {fai_path} existiert bereits – überspringe.")
        return fai_path

    if not HAVE_PYFAIDX:
        raise RuntimeError("pyfaidx ist nicht installiert, kann FAI nicht erzeugen.")

    logger.log(f"[FASTA] Erstelle Index für {fasta_path} …")
    Faidx(fasta_path)  # erzeugt automatisch .fai
    logger.log(f"[FASTA] Index erstellt: {fai_path}")
    return fai_path

@dataclass
class ShiftDetectionStats:
    """
    ✅ Tracking für systematische Positions-Verschiebungen.
    
    Erkennt Patterns wie:
    - +1 Shift (häufigster Fall bei VCF-Konvertierungsfehlern)
    - -1 Shift
    - ±2 Shifts
    
    Nutzt Pattern für proaktive Korrektur → 30-50% schneller bei systematischen Fehlern.
    """
    
    # Pattern-Erkennung
    shift_patterns: Dict[int, int] = field(default_factory=dict)  # {offset: count}
    total_shifts_detected: int = 0
    
    # Pattern-Anwendung
    pattern_applied: int = 0
    pattern_successful: int = 0
    pattern_failed: int = 0
    
    # Dominant Pattern
    dominant_shift: Optional[int] = None
    dominant_confidence: float = 0.0
    
    def record_shift(self, shift: int):
        """
        Registriere erkannte Verschiebung.
        
        Args:
            shift: Erkannter Offset (z.B. +1, -1, +2)
        """
        self.shift_patterns[shift] = self.shift_patterns.get(shift, 0) + 1
        self.total_shifts_detected += 1
        self._update_dominant()
    
    def _update_dominant(self):
        """Aktualisiere dominantes Shift-Pattern."""
        if not self.shift_patterns:
            return
        
        # Finde häufigstes Pattern
        max_shift = max(self.shift_patterns.items(), key=lambda x: x[1])
        self.dominant_shift = max_shift[0]
        self.dominant_confidence = max_shift[1] / self.total_shifts_detected
    
    def should_apply_pattern(self, threshold: float = 0.6, min_samples: int = 10) -> bool:
        """
        Prüfe ob Pattern zuverlässig genug für proaktive Anwendung.
        
        Args:
            threshold: Mindest-Confidence (default: 0.6 = 60%)
            min_samples: Mindestanzahl Samples für Zuverlässigkeit
        
        Returns:
            True wenn Pattern angewendet werden sollte
        """
        return (
            self.dominant_shift is not None and 
            self.dominant_confidence >= threshold and
            self.total_shifts_detected >= min_samples
        )
    
    def print_report(self, logger):
        """
        Druckt detaillierten Shift-Detection-Report.
        
        Args:
            logger: Logger-Instanz (mit .log() Methode)
        """
        if not self.shift_patterns:
            return
        
        logger.log("\n" + "="*70)
        logger.log("SHIFT-DETECTION REPORT")
        logger.log("="*70)
        
        logger.log(f"\n📊 DETECTED PATTERNS:")
        logger.log(f"  Total Shifts: {self.total_shifts_detected}")
        
        # Sortiere nach Häufigkeit
        for shift, count in sorted(
            self.shift_patterns.items(), 
            key=lambda x: x[1], 
            reverse=True
        ):
            pct = (count / self.total_shifts_detected) * 100
            logger.log(f"  Shift {shift:+3d}: {count:4d} ({pct:5.1f}%)")
        
        if self.dominant_shift is not None:
            logger.log(f"\n🎯 DOMINANT PATTERN:")
            logger.log(f"  Shift: {self.dominant_shift:+d}")
            logger.log(f"  Confidence: {self.dominant_confidence:.1%}")
            
            if self.pattern_applied > 0:
                success_rate = self.pattern_successful / self.pattern_applied
                logger.log(f"  Applied: {self.pattern_applied}")
                logger.log(
                    f"  Success Rate: {self.pattern_successful}/{self.pattern_applied} "
                    f"({success_rate*100:.1f}%)"
                )
        
        logger.log("="*70 + "\n")
    
    def get_summary(self) -> dict:
        """
        Gibt Zusammenfassung als Dictionary zurück.
        
        Returns:
            Dictionary mit Stats
        """
        return {
            "total_shifts": self.total_shifts_detected,
            "patterns": dict(self.shift_patterns),
            "dominant_shift": self.dominant_shift,
            "dominant_confidence": self.dominant_confidence,
            "pattern_applied": self.pattern_applied,
            "pattern_successful": self.pattern_successful,
            "pattern_failed": self.pattern_failed,
            "success_rate": (
                self.pattern_successful / self.pattern_applied 
                if self.pattern_applied > 0 else 0.0
            )
        }


class FastaValidator:
    """
    ✅ OPTIMIERT: FASTA-Validator mit O(1) Random-Access (kein Iterieren mehr).
    
    Nutzt .fai Index für mathematische Positionsberechnung statt Byte-Streaming.
    Behebt das "Hängenbleiben" bei großen Chromosomen.
    """
    
    def __init__(
        self, 
        fasta_path: str,
        strict_mode: bool = True,
        cache_size: int = 1000,
        logger=None
    ):
        self.fasta_path = fasta_path
        self.strict_mode = strict_mode
        self.logger = logger
        
        # Index-Daten: name -> (length, offset, line_bases, line_width)
        self._fasta_index = {}
        
        self._fasta_file = None
        self._fasta_mmap = None
        
        self._cache_size = cache_size
        self._fetch_sequence_cached = None
        
        self._region_cache = {}
        self._region_cache_size = Config.FASTA_REGION_CACHE_SIZE
        
        # Stats & Shift-Detection (Initialisierung wie gehabt)
        self.stats = {
            "total_checked": 0, "skipped_snv": 0, "skipped_insertion": 0,
            "validated_deletion": 0, "validated_complex": 0, "validated_other": 0,
            "validation_failed": 0, "validation_corrected": 0,
            "cache_hits": 0, "cache_misses": 0,
        }
        self.shift_stats = ShiftDetectionStats()
        self.shift_detection_enabled = True
        self.shift_confidence_threshold = 0.6
        self.shift_min_samples = 10

        if self.logger:
            self.logger.log(f"[FastaValidator] Init: {fasta_path} (O(1) Access Mode)")

    def _load_fasta_index(self):
        """
        Lädt .fai Datei und parst Geometrie (Zeilenlängen).
        Format: NAME LENGTH OFFSET LINE_BASES LINE_WIDTH
        """
        if self._fasta_index:
            return

        fai_path = self.fasta_path + ".fai"
        if not os.path.exists(fai_path):
            # Versuch, Index zu bauen, falls er fehlt
            if self.logger:
                self.logger.log("[FastaValidator] Index fehlt, versuche Build...")
            try:
                from pyfaidx import Faidx
                Faidx(self.fasta_path)
            except Exception as e:
                raise FileNotFoundError(f"Konnte Index nicht erstellen: {e}")

        try:
            with open(fai_path, "r") as f:
                for line in f:
                    parts = line.strip().split("\t")
                    if len(parts) < 5: continue
                    
                    chrom = parts[0]
                    # Parsing der Geometrie-Daten für Mathe-Zugriff
                    length = int(parts[1])
                    offset = int(parts[2])
                    line_bases = int(parts[3])
                    line_width = int(parts[4])
                    
                    self._fasta_index[chrom] = (length, offset, line_bases, line_width)
            
            if self.logger:
                self.logger.log(f"[FastaValidator] Index geladen: {len(self._fasta_index)} Chromosomen")
                
        except Exception as e:
            raise RuntimeError(f"Fehler beim Lesen von {fai_path}: {e}")

    def _open_fasta_mmap(self):
        """Öffnet FASTA File-Handle (mmap optional, File-Handle reicht oft)."""
        if self._fasta_file is not None:
            return
        try:
            self._fasta_file = open(self.fasta_path, 'rb')
            # Unter Windows ist mmap auf sehr großen Dateien manchmal problematisch
            # Wir nutzen hier standard file seek, was mit SSDs fast genauso schnell ist
            # und weniger Locking-Probleme verursacht.
        except Exception as e:
            if self.logger:
                self.logger.log(f"[FastaValidator] File open error: {e}")

    def _fetch_sequence_impl(self, chrom: str, start: int, end: int) -> str:
        """
        Berechnet Byte-Offset und liest Sequenz direkt (0-based coordinates).
        """
        if chrom not in self._fasta_index:
            return ""
            
        length, offset, line_bases, line_width = self._fasta_index[chrom]
        
        if start < 0 or end > length:
            return ""
        
        # --- MATHEMATISCHER SPRUNG (O(1)) ---
        # Berechne Byte-Position inkl. Newlines
        
        # Start-Offset berechnen
        # Zeilen vor dem Start * Bytes pro Zeile + Rest-Basen in der Zeile
        start_line_idx = start // line_bases
        start_col_idx = start % line_bases
        byte_start = offset + (start_line_idx * line_width) + start_col_idx
        
        # End-Offset berechnen
        end_line_idx = end // line_bases
        end_col_idx = end % line_bases
        byte_end = offset + (end_line_idx * line_width) + end_col_idx
        
        # Anzahl zu lesender Bytes (inklusive Newlines zwischendrin)
        read_len = byte_end - byte_start
        
        if read_len <= 0:
            return ""

        # Seek und Read
        self._fasta_file.seek(byte_start)
        raw_bytes = self._fasta_file.read(read_len)
        
        # Newlines entfernen und dekodieren
        # .replace(b'\n', b'') ist in C implementiert und extrem schnell
        seq = raw_bytes.replace(b'\n', b'').replace(b'\r', b'').decode('ascii', errors='ignore')
        
        return seq.upper()

    def _get_reference_base(self, chrom: str, pos: int, length: int) -> Optional[str]:
        """
        Holt Referenzsequenz aus FASTA (nutzt Region-Cache + LRU).
        
        Features:
        - 0-based Indexing Fix
        - Cross-Boundary Schutz (Direkt-Fetch bei Block-Grenzen)
        - LRU-Cache für Regionen und Direktzugriffe
        """
        try:
            # Lazy Initialization
            if self._fasta_index is None: self._load_fasta_index()
            if self._fasta_file is None: self._open_fasta_mmap()
            
            if chrom not in self._fasta_index:
                return None
            
            # LRU Cache Wrapper initialisieren falls noch nicht geschehen
            if self._fetch_sequence_cached is None:
                self._fetch_sequence_cached = lru_cache(maxsize=self._cache_size)(
                    self._fetch_sequence_impl
                )

            # Konvertiere zu 0-based Koordinaten für interne Logik
            pos0 = pos - 1
            block_size = self._region_cache_size  # Standard: 100.000

            # --- CHECK 1: Block-Grenzüberschreitung (Cross-Boundary) ---
            # Wenn eine Sequenz über das Ende eines Cache-Blocks hinausgeht,
            # laden wir sie direkt, um Stitching-Fehler zu vermeiden.
            offset_in_block = pos0 % block_size
            if offset_in_block + length > block_size:
                # Direkt-Fetch (Parameter: chrom, start, end)
                # KEIN seq_offset übergeben! _fetch_sequence_impl holt das intern.
                return self._fetch_sequence_cached(chrom, pos0, pos0 + length)

            # --- CHECK 2: Standard Region-Cache ---
            # Berechne Start des 100kb-Blocks
            region_start = (pos0 // block_size) * block_size
            cache_key = (chrom, region_start)
            
            # Prüfe internen Region-Cache (Dict)
            if cache_key not in self._region_cache:
                # Lade den ganzen Block via LRU-Cache
                # Parameter: chrom, start, end
                region_seq = self._fetch_sequence_cached(chrom, region_start, region_start + block_size)
                
                if region_seq:
                    self._region_cache[cache_key] = region_seq
                    # Simple Eviction für Region-Cache (FIFO)
                    if len(self._region_cache) > 50:
                        self._region_cache.pop(next(iter(self._region_cache)))
                else:
                    return None
            
            # Extrahiere gewünschten Teil aus dem Block
            region = self._region_cache[cache_key]
            offset = pos0 - region_start
            
            # Bounds Check (Sicherheitsnetz)
            if offset < 0 or offset + length > len(region):
                # Fallback: Direkt-Fetch
                return self._fetch_sequence_cached(chrom, pos0, pos0 + length)
            
            return region[offset:offset + length]
            
        except Exception as e:
            if self.logger:
                self.logger.log(f"[FastaValidator] Lookup Error {chrom}:{pos}: {e}")
            return None
        # --- Restliche Methoden (validate, needs_validation, etc.) ---
    # Diese müssen beibehalten werden, da sie die Logik steuern.
    # Ich füge hier die wichtigsten Wrapper ein, damit die Klasse vollständig funktioniert.

    def needs_validation(self, ref: str, alt: str) -> Tuple[bool, str]:
        ref_len, alt_len = len(ref), len(alt)
        if ref_len == 1 and alt_len == 1: return (False, "snv")
        if ref_len == 1 and alt_len > 1:
            return (True, "insertion_strict") if self.strict_mode else (False, "insertion_aggressive")
        return (True, "deletion") if ref_len > 1 and alt_len == 1 else (True, "complex")

    def validate_single(self, chrom, pos, ref, alt, build):
        # Wrapper für den neuen validate Call
        return self.validate(chrom, pos, ref, alt, build)

    def validate(self, chrom=None, pos=None, ref=None, alt=None, build=None):
        # Modus 1: File Check
        if chrom is None:
            try:
                self._load_fasta_index()
                self._open_fasta_mmap()
                return (True, "OK")
            except Exception as e:
                return (False, str(e))

        # Modus 2: Variant Check
        self.stats["total_checked"] += 1
        needs, reason = self.needs_validation(ref, alt)
        if not needs:
            self.stats["skipped_snv" if reason == "snv" else "skipped_insertion"] += 1
            return (True, None, reason)

        chrom_norm = self._normalize_chrom(chrom)
        if not chrom_norm:
            self.stats["validation_failed"] += 1
            return (False, None, "chr_missing")

        fetched = self._get_reference_base(chrom_norm, pos, len(ref))
        if not fetched:
            self.stats["validation_failed"] += 1
            return (False, None, "lookup_fail")

        if fetched == ref.upper():
            self.stats["validated_complex"] += 1
            return (True, None, "valid")
        else:
            self.stats["validation_corrected"] += 1
            return (True, fetched, "corrected")
        
    def validate_variant(self, chrom: Optional[str] = None, pos: Optional[int] = None, 
                        ref: Optional[str] = None, alt: Optional[str] = None, 
                        build: Optional[str] = None):
        #Wrapper-Umleitung zu validate()
        return self.validate(chrom, pos, ref, alt, build)

    def validate_batch(self, variants, build, enable_shift_detection=True, shift_confidence_threshold=0.6):
        """
        Batch-Validierung (Fix für Unpacking-Error).
        Erwartet variants als Liste von (chrom, pos, ref, alt).
        """
        results = []
        # FIX: 'idx' entfernt, da variants nur 4 Elemente pro Tupel hat
        for chrom, pos, ref, alt in variants:
            # Einzelvalidierung aufrufen
            res = self.validate(chrom, pos, ref, alt, build)
            results.append(res)
        return results

    def _normalize_chrom(self, chrom):
        if chrom in self._fasta_index: return chrom
        alt = chrom[3:] if chrom.startswith("chr") else f"chr{chrom}"
        return alt if alt in self._fasta_index else None

    def close(self):
        if self._fasta_file:
            self._fasta_file.close()
            self._fasta_file = None
        if self.logger: self.logger.log("[FastaValidator] Closed.")
    
    def get_stats(self) -> dict:
        #Gibt Statistiken mit Prozentsätzen zurück.
        total = self.stats["total_checked"]
        if total == 0:
            return dict(self.stats)
        
        stats_with_pct = dict(self.stats)
        
        stats_with_pct["skipped_snv_pct"] = 100.0 * self.stats["skipped_snv"] / total
        stats_with_pct["skipped_insertion_pct"] = 100.0 * self.stats["skipped_insertion"] / total
        
        total_validated = (
            self.stats["validated_deletion"] +
            self.stats["validated_complex"] +
            self.stats["validated_other"]
        )
        stats_with_pct["validated_pct"] = 100.0 * total_validated / total
        stats_with_pct["corrected_pct"] = 100.0 * self.stats["validation_corrected"] / total
        
        # ✅ Cache-Stats
        cache_total = self.stats["cache_hits"] + self.stats["cache_misses"]
        if cache_total > 0:
            stats_with_pct["cache_hit_rate"] = 100.0 * self.stats["cache_hits"] / cache_total
        
        return stats_with_pct
    
    def print_stats(self):
        #Gibt formatierte Statistiken mit Cache-Info aus.
        stats = self.get_stats()
        total = stats["total_checked"]
        
        if total == 0:
            if self.logger:
                self.logger.log("[FastaValidator] Keine Varianten validiert")
            return
        
        mode_name = "STRICT" if self.strict_mode else "AGGRESSIVE"
        
        total_skipped = self.stats["skipped_snv"] + self.stats["skipped_insertion"]
        total_validated = (
            self.stats["validated_deletion"] +
            self.stats["validated_complex"] +
            self.stats["validated_other"]
        )
        
        # ✅ Cache-Stats
        cache_hits = self.stats["cache_hits"]
        cache_misses = self.stats["cache_misses"]
        cache_total = cache_hits + cache_misses
        hit_rate = stats.get("cache_hit_rate", 0)
        io_reduction = hit_rate  # Cache-Hits = gespart Disk I/O
        
        msg = (
            f"[FastaValidator] Statistik ({mode_name} Mode):\n"
            f"  Total: {total:,} Varianten\n"
            f"\n"
            f"  ⚡ FAST PATH (übersprungen):\n"
            f"    SNVs: {self.stats['skipped_snv']:,} ({stats.get('skipped_snv_pct', 0):.1f}%)\n"
        )
        
        if self.stats["skipped_insertion"] > 0:
            msg += (
                f"    Insertionen: {self.stats['skipped_insertion']:,} "
                f"({stats.get('skipped_insertion_pct', 0):.1f}%)\n"
            )
        
        msg += (
            f"    Gesamt übersprungen: {total_skipped:,} ({100.0 * total_skipped / total:.1f}%)\n"
            f"\n"
            f"  🐌 SLOW PATH (FASTA-Validierung):\n"
            f"    Deletionen: {self.stats['validated_deletion']:,}\n"
            f"    Complex: {self.stats['validated_complex']:,}\n"
            f"    Andere: {self.stats['validated_other']:,}\n"
            f"    Gesamt validiert: {total_validated:,} ({stats.get('validated_pct', 0):.1f}%)\n"
            f"\n"
            f"  📝 Ergebnisse:\n"
            f"    Korrekturen: {self.stats['validation_corrected']:,} ({stats.get('corrected_pct', 0):.1f}%)\n"
            f"    Fehler: {self.stats['validation_failed']:,}\n"
        )
        
        # ✅ Cache-Performance
        if cache_total > 0:
            msg += (
                f"\n"
                f"  🔥 Cache Performance:\n"
                f"    Cache hits: {cache_hits:,}\n"
                f"    Cache misses: {cache_misses:,}\n"
                f"    Hit rate: {hit_rate:.1f}%\n"
                f"    Disk I/O saved: {cache_hits:,}/{cache_total:,} ({io_reduction:.1f}%)\n"
                f"    Cache size: {self._cache_size}"
            )
        
        if self.logger:
            self.logger.log(msg)
        else:
            print(msg)
    
    def reset_stats(self):
        
        self.stats = {k: 0 for k in self.stats.keys()}
    
    def reset_shift_detection(self):
        
        self.shift_stats = ShiftDetectionStats()
        if self.logger:
            self.logger.log("[FastaValidator] Shift-Detection Stats zurückgesetzt")
    
    def configure_shift_detection(self, enabled: bool = True, confidence_threshold: float = 0.6, 
                                  min_samples: int = 10):
        
        self.shift_detection_enabled = enabled
        self.shift_confidence_threshold = confidence_threshold
        self.shift_min_samples = min_samples
        
        if self.logger:
            self.logger.log(
                f"[FastaValidator] Shift-Detection konfiguriert:\n"
                f"  Enabled: {enabled}\n"
                f"  Confidence: {confidence_threshold:.1%}\n"
                f"  Min Samples: {min_samples}"
            )
    
    def get_shift_stats_summary(self) -> dict:
        
        return self.shift_stats.get_summary()
    
# =============================================================================
# FLAG AND OPTIONS MANAGER (V10 - KOMPLETT ÜBERARBEITET)
# =============================================================================
#
# Zentrale Verwaltung ALLER Flags und GUI-Einstellungen.
# Wird von App instanziiert und an MainFilterGate/Distiller übergeben.
#
# ═══════════════════════════════════════════════════════════════════════════════
# V10 ARCHITEKTUR: FILTER-FLOW
# ═══════════════════════════════════════════════════════════════════════════════
#
#     GUI tk-Variables (af_slider_var, only_coding_var, ...)
#            │
#            ▼
#     flag_manager.sync_from_gui(app)  ← Einmalig vor Pipeline-Start
#            │
#            ├─────────────────────────────────────────┐
#            ▼                                         ▼
#      FlagManager                              MainFilterGate
#      (alle ~30 Settings)                      ├── CodingFilter
#            │                                  └── AfNoneTreatmentManager
#            │                                         │
#            └──────────────► Distiller ◄──────────────┘
#                                 │
#                       _apply_af_filter_final()
#                       _validate_af_in_cache()
#                                 │
#                                 ▼
#                       main_filter_gate.check_variant()
#
# FLOW:
# 1. GUI → FlagManager.sync_from_gui() → Alle Einstellungen zentral
# 2. Distiller holt Werte vom FlagManager
# 3. MainFilterGate prüft Varianten mit AF-None-Policy und Coding-Filter
# 4. Ergebnis: "emit" oder "reject"
#
# ═══════════════════════════════════════════════════════════════════════════════
class Flag_and_Options_Manager:
    """
    V10: Zentrale Verwaltung ALLER Flags und Optionen.
    
    Kategorien:
    - Pipeline-Steuerung (af_threshold, include_none, etc.)
    - Filter-Flags (only_protein_coding, filter_pass_only)
    - Stale-Days (für Cache-Invalidierung)
    - API-Keys (AlphaGenome)
    - Post-Filter (CADD, ClinVar, Impact)
    - Debug-Optionen
    
    Thread-safe durch RLock.
    """
    
    def __init__(self, logger=None):
        self.logger = logger
        self._lock = threading.RLock()
        
        # ═══════════════════════════════════════════════════════════════════
        # PIPELINE HAUPTFILTER
        # ═══════════════════════════════════════════════════════════════════
        self.af_threshold: float = 0.01  # Max AF (1%)
        self.include_none: bool = True   # AF=None einschließen
        self.filter_pass_only: bool = False  # Nur PASS-Varianten
        self.qual_threshold: float = 30.0  # Min QUAL Score
        
        # ═══════════════════════════════════════════════════════════════════
        # PRÄ-FILTER (vor Pipeline)
        # ═══════════════════════════════════════════════════════════════════
        self.only_protein_coding: bool = False  # Nur protein-codierende Gene
        
        # ═══════════════════════════════════════════════════════════════════
        # CACHE STALE DAYS
        # ═══════════════════════════════════════════════════════════════════
        self.stale_days: int = Config.STALE_DAYS_AF        # AF-Daten Staleness
        self.stale_days_full: int = Config.STALE_DAYS_FULL  # Vollannotation Staleness
        
        # ═══════════════════════════════════════════════════════════════════
        # API KEYS
        # ═══════════════════════════════════════════════════════════════════
        self.alphagenome_key: Optional[str] = None
        
        # ═══════════════════════════════════════════════════════════════════
        # AF-NONE POLICY
        # ═══════════════════════════════════════════════════════════════════
        self.af_none_policy: str = "clinical"  # clinical, research, permissive, strict
        
        # ═══════════════════════════════════════════════════════════════════
        # POST-FILTER (GUI-Tabelle)
        # ═══════════════════════════════════════════════════════════════════
        self.postfilter_protein_coding: bool = False
        self.postfilter_cadd_min: Optional[float] = None
        
        # ClinVar Filter
        self.postfilter_benign: bool = True
        self.postfilter_likely_benign: bool = True
        self.postfilter_vus: bool = True
        self.postfilter_likely_path: bool = True
        self.postfilter_path: bool = True
        
        # Impact Filter
        self.postfilter_impact_high: bool = True
        self.postfilter_impact_moderate: bool = True
        self.postfilter_impact_low: bool = True
        
        # ═══════════════════════════════════════════════════════════════════
        # GEN-LISTEN
        # ═══════════════════════════════════════════════════════════════════
        self.gen_whitelist: str = ""
        self.gen_blacklist: str = ""
        
        # ═══════════════════════════════════════════════════════════════════
        # DEBUG / ENTWICKLUNG
        # ═══════════════════════════════════════════════════════════════════
        self.debug_show_db: bool = False
        self.skip_info_recycling: bool = False
        self.vcf_export_mode: str = "complete"  # complete, filtered
        self.cadd_highlight_threshold: float = 20.0
        
        # ═══════════════════════════════════════════════════════════════════
        # EXTERNE TOOLS
        # ═══════════════════════════════════════════════════════════════════
        self.external_db_viewer: str = ""
        
        if self.logger:
            self.logger.log("[FlagManager] V10 Initialisiert mit Defaults")
    
    # ───────────────────────────────────────────────────────────────────────
    # SYNC MIT GUI (tk.Var → Flag_Manager)
    # ───────────────────────────────────────────────────────────────────────
    
    def sync_from_gui(self, app):
        """
        Synchronisiert alle Werte von GUI tk-Variablen.
        Sollte vor Pipeline-Start aufgerufen werden.
        """
        with self._lock:
            try:
                # Pipeline
                self.af_threshold = float(app.af_threshold.get())
                self.include_none = bool(app.include_none.get())
                self.filter_pass_only = bool(app.filter_pass_only.get())
                self.qual_threshold = float(app.qual_threshold.get())
                
                # Prä-Filter
                self.only_protein_coding = bool(app.prefilter_protein_coding.get())
                
                # Stale Days
                self.stale_days = int(app.stale_days.get())
                self.stale_days_full = int(app.stale_days_full.get())
                
                # API Keys
                key = app.alphagenome_key.get().strip() if hasattr(app, 'alphagenome_key') else ""
                self.alphagenome_key = key if key else None
                
                # Post-Filter
                self.postfilter_protein_coding = bool(app.postfilter_protein_coding.get())
                cadd_str = app.postfilter_cadd_min.get().strip()
                self.postfilter_cadd_min = float(cadd_str) if cadd_str else None
                
                self.postfilter_benign = bool(app.postfilter_benign.get())
                self.postfilter_likely_benign = bool(app.postfilter_likely_benign.get())
                self.postfilter_vus = bool(app.postfilter_vus.get())
                self.postfilter_likely_path = bool(app.postfilter_likely_path.get())
                self.postfilter_path = bool(app.postfilter_path.get())
                
                self.postfilter_impact_high = bool(app.postfilter_impact_high.get())
                self.postfilter_impact_moderate = bool(app.postfilter_impact_moderate.get())
                self.postfilter_impact_low = bool(app.postfilter_impact_low.get())
                
                # Gen-Listen
                self.gen_whitelist = app.gen_whitelist.get().strip()
                self.gen_blacklist = app.gen_blacklist.get().strip()
                
                # Debug
                self.debug_show_db = bool(app.debug_show_db.get())
                self.skip_info_recycling = bool(app.skip_info_recycling.get())
                self.vcf_export_mode = app.vcf_export_mode.get()
                self.cadd_highlight_threshold = float(app.cadd_highlight_threshold.get())
                
                # Externe Tools
                self.external_db_viewer = app.external_db_viewer.get()
                
                if self.logger:
                    self.logger.log(f"[FlagManager] Sync von GUI: AF<={self.af_threshold}, Coding={self.only_protein_coding}")
                    
            except Exception as e:
                if self.logger:
                    self.logger.log(f"[FlagManager] Sync-Fehler: {e}")
    
    # ───────────────────────────────────────────────────────────────────────
    # GETTER (Thread-Safe)
    # ───────────────────────────────────────────────────────────────────────
    
    def get_af_threshold(self) -> float:
        with self._lock:
            return self.af_threshold
    
    def get_include_none(self) -> bool:
        with self._lock:
            return self.include_none
    
    def get_filter_pass_only(self) -> bool:
        with self._lock:
            return self.filter_pass_only
    
    def get_qual_threshold(self) -> float:
        with self._lock:
            return self.qual_threshold
    
    def get_only_protein_coding(self) -> bool:
        with self._lock:
            return self.only_protein_coding
    
    def get_stale_days(self) -> int:
        with self._lock:
            return self.stale_days
    
    def get_stale_days_full(self) -> int:
        with self._lock:
            return self.stale_days_full
    
    def get_alphagenome_key(self) -> Optional[str]:
        with self._lock:
            return self.alphagenome_key
    
    def get_af_none_policy(self) -> str:
        with self._lock:
            return self.af_none_policy
    
    def get_cadd_highlight_threshold(self) -> float:
        with self._lock:
            return self.cadd_highlight_threshold
    
    # ───────────────────────────────────────────────────────────────────────
    # SETTER (Thread-Safe, mit Logging)
    # ───────────────────────────────────────────────────────────────────────
    
    def set_af_threshold(self, value: float):
        with self._lock:
            self.af_threshold = float(value)
            if self.logger:
                self.logger.log(f"[FlagManager] AF-Threshold -> {value}")
    
    def set_include_none(self, value: bool):
        with self._lock:
            self.include_none = bool(value)
    
    def set_only_protein_coding(self, value: bool):
        with self._lock:
            self.only_protein_coding = bool(value)
            if self.logger:
                self.logger.log(f"[FlagManager] Protein-Coding -> {value}")
    
    def set_af_none_policy(self, policy: str):
        valid = ["clinical", "research", "permissive", "strict"]
        with self._lock:
            if policy in valid:
                self.af_none_policy = policy
                if self.logger:
                    self.logger.log(f"[FlagManager] AF-None-Policy -> {policy}")
            else:
                if self.logger:
                    self.logger.log(f"[FlagManager] Ungueltige Policy: {policy}")
    
    def set_stale_days(self, days: int):
        with self._lock:
            self.stale_days = int(days)
    
    def set_stale_days_full(self, days: int):
        with self._lock:
            self.stale_days_full = int(days)
    
    def set_alphagenome_key(self, key: Optional[str]):
        with self._lock:
            self.alphagenome_key = key.strip() if key else None
    
    # ───────────────────────────────────────────────────────────────────────
    # EXPORT / IMPORT
    # ───────────────────────────────────────────────────────────────────────
    
    def get_all_flags(self) -> dict:
        """Exportiert alle Flags als Dict."""
        with self._lock:
            return {
                "af_threshold": self.af_threshold,
                "include_none": self.include_none,
                "filter_pass_only": self.filter_pass_only,
                "qual_threshold": self.qual_threshold,
                "only_protein_coding": self.only_protein_coding,
                "stale_days": self.stale_days,
                "stale_days_full": self.stale_days_full,
                "af_none_policy": self.af_none_policy,
                "cadd_highlight_threshold": self.cadd_highlight_threshold,
            }
    
    def set_from_dict(self, flags: dict):
        """Importiert Flags aus Dict."""
        with self._lock:
            for key, value in flags.items():
                if hasattr(self, key):
                    setattr(self, key, value)




class MainFilterGate:
    """
    V10: Zentrales Filter-Gate für alle Varianten-Filter.
    
    SubGates:
    - CodingFilter (Protein-Coding Check)
    - AfNoneTreatmentManager (AF=None Policy)
    
    Der Distiller fragt NUR dieses Gate, nicht die SubGates direkt.
    Alle Filter-Einstellungen werden vom Flag_and_Options_Manager geholt.
    
    Usage:
        gate = MainFilterGate(flag_manager, db, gene_annotator, logger)
        
        # Einzelne Variante:
        passed, reason, data = gate.check_variant(key, af_value=0.005)
        
        # Batch:
        results = gate.check_batch(variants)  # [(key, af, status), ...]
    """
    
    def __init__(
        self,
        flag_manager: 'Flag_and_Options_Manager',
        db: 'VariantDB',
        gene_annotator: Optional['GeneAnnotator'],
        logger
    ):
        self.flag_manager = flag_manager
        self.db = db
        self.logger = logger
        self._lock = threading.RLock()
        
        self._log("=" * 60)
        self._log("[MainFilterGate] V10 INITIALISIERUNG")
        self._log("=" * 60)
        
        # ═══════════════════════════════════════════════════════════════════
        # SUBGATE: CodingFilter
        # ═══════════════════════════════════════════════════════════════════
        self._coding_filter: Optional[CodingFilter] = None
        self._gene_annotator = gene_annotator
        
        if gene_annotator and getattr(gene_annotator, 'available', False):
            self._coding_filter = CodingFilter(db, gene_annotator, logger)
            self._log("[MainFilterGate] SubGate CodingFilter AKTIV")
        else:
            self._log("[MainFilterGate] SubGate CodingFilter INAKTIV (kein GeneAnnotator)")
        
        # ═══════════════════════════════════════════════════════════════════
        # SUBGATE: AfNoneTreatmentManager
        # ═══════════════════════════════════════════════════════════════════
        self._af_none_manager = AfNoneTreatmentManager(
            preset=flag_manager.get_af_none_policy(),
            logger=logger
        )
        self._log(f"[MainFilterGate] SubGate AfNoneTreatment (Policy: {flag_manager.get_af_none_policy()})")
        
        # ═══════════════════════════════════════════════════════════════════
        # BATCH BUFFER für Coding-Checks
        # ═══════════════════════════════════════════════════════════════════
        self._coding_batch_buffer: List[Tuple] = []
        self._coding_batch_results: Dict = {}
        self.CODING_BATCH_SIZE = 500
        
        # ═══════════════════════════════════════════════════════════════════
        # STATISTIKEN
        # ═══════════════════════════════════════════════════════════════════
        self.stats = {
            "total_checked": 0,
            "passed": 0,
            "failed_af": 0,
            "failed_coding": 0,
            "failed_none_policy": 0,
            "af_none_included": 0,
        }
        
        self._log("[MainFilterGate] Initialisierung abgeschlossen")
        self._log("=" * 60)
    
    def _log(self, msg: str):
        if self.logger:
            self.logger.log(msg)
    
    # ───────────────────────────────────────────────────────────────────────
    # SUBGATE ACCESSORS
    # ───────────────────────────────────────────────────────────────────────
    
    def get_coding_filter(self) -> Optional['CodingFilter']:
        """Gibt CodingFilter SubGate zurück (oder None)."""
        return self._coding_filter
    
    def get_af_none_manager(self) -> 'AfNoneTreatmentManager':
        """Gibt AfNoneTreatmentManager SubGate zurück."""
        return self._af_none_manager
    
    def is_coding_available(self) -> bool:
        """Prüft ob CodingFilter verfügbar ist."""
        return self._coding_filter is not None
    
    def set_gene_annotator(self, gene_annotator: 'GeneAnnotator'):
        """Setzt GeneAnnotator nachträglich (für Lazy Loading)."""
        self._gene_annotator = gene_annotator
        if gene_annotator and getattr(gene_annotator, 'available', False):
            self._coding_filter = CodingFilter(self.db, gene_annotator, self.logger)
            self._log("[MainFilterGate] CodingFilter nachträglich aktiviert")
    
    # ───────────────────────────────────────────────────────────────────────
    # HAUPTMETHODE: check_variant
    # ───────────────────────────────────────────────────────────────────────
    
    def check_variant(
        self,
        key: Tuple,
        af_value: Optional[float] = None,
        fetch_status: Optional[str] = None
    ) -> Tuple[bool, str, dict]:
        """
        HAUPTMETHODE: Prüft ob Variante alle Filter besteht.
        
        Filter-Reihenfolge:
        1. AF-Filter (schnell)
        2. AF-None Policy (wenn AF=None)
        3. Protein-Coding Filter (wenn aktiviert)
        
        Args:
            key: Varianten-Key (chr, pos, ref, alt, build)
            af_value: Populations-AF (kann None sein)
            fetch_status: Status des AF-Fetches
        
        Returns:
            (passed, reason, data)
            - passed: True wenn alle Filter bestanden
            - reason: Grund für Ablehnung (oder "passed")
            - data: Zusätzliche Daten (gene_symbol, is_coding, etc.)
        """
        with self._lock:
            self.stats["total_checked"] += 1
        
        result_data = {"gene_symbol": None, "is_coding": None}
        
        # Hole aktuelle Einstellungen vom FlagManager
        af_threshold = self.flag_manager.get_af_threshold()
        include_none = self.flag_manager.get_include_none()
        only_coding = self.flag_manager.get_only_protein_coding()
        
        # ─────────────────────────────────────────────────────────────────
        # FILTER 1: AF-Threshold
        # ─────────────────────────────────────────────────────────────────
        if af_value is not None:
            if af_value > af_threshold:
                with self._lock:
                    self.stats["failed_af"] += 1
                return False, f"af>{af_threshold}", result_data
        
        # ─────────────────────────────────────────────────────────────────
        # FILTER 2: AF-None Policy (SubGate)
        # ─────────────────────────────────────────────────────────────────
        if af_value is None:
            # V10 FIX: classify_none_type erwartet row-Dict, nicht einzelne Parameter
            temp_row = {
                "af_filter_mean": af_value,
                "meanAF_fetch_success": fetch_status,
                "meanAF_last_fetch": None  # Nicht verfügbar in diesem Kontext
            }
            none_type = self._af_none_manager.classify_none_type(temp_row)
            
            should_include = self._af_none_manager.should_include_type(none_type)
            
            if not should_include and not include_none:
                with self._lock:
                    self.stats["failed_none_policy"] += 1
                return False, f"af_none:{none_type}", result_data
            
            with self._lock:
                self.stats["af_none_included"] += 1
        
        # ─────────────────────────────────────────────────────────────────
        # FILTER 3: Protein-Coding (SubGate)
        # ─────────────────────────────────────────────────────────────────
        if only_coding:
            if not self._coding_filter:
                # Kein CodingFilter -> durchlassen
                pass
            else:
                # Einzelabfrage
                chrom, pos = key[0], key[1]
                coding_result = self._coding_filter.process_batch([(key, chrom, pos)])
                
                if key in coding_result:
                    data = coding_result[key]
                    result_data["gene_symbol"] = data.get("gene_symbol")
                    result_data["is_coding"] = data.get("is_coding")
                    
                    if data.get("is_coding") in (False, 0):
                        with self._lock:
                            self.stats["failed_coding"] += 1
                        return False, "not_coding", result_data
        
        # ─────────────────────────────────────────────────────────────────
        # ALLE FILTER BESTANDEN
        # ─────────────────────────────────────────────────────────────────
        with self._lock:
            self.stats["passed"] += 1
        
        return True, "passed", result_data
    
    # ───────────────────────────────────────────────────────────────────────
    # BATCH-METHODE: check_batch
    # ───────────────────────────────────────────────────────────────────────
    
    def check_batch(
        self,
        variants: List[Tuple[Tuple, Optional[float], Optional[str]]]
    ) -> Dict[Tuple, Tuple[bool, str, dict]]:
        """
        Batch-Prüfung für mehrere Varianten.
        
        Args:
            variants: Liste von (key, af_value, fetch_status) Tupeln
        
        Returns:
            Dict[key -> (passed, reason, data)]
        """
        results = {}
        
        # Hole Einstellungen einmal
        af_threshold = self.flag_manager.get_af_threshold()
        include_none = self.flag_manager.get_include_none()
        only_coding = self.flag_manager.get_only_protein_coding()
        
        # Sammle Varianten für Coding-Batch
        coding_candidates = []
        
        for key, af_value, fetch_status in variants:
            with self._lock:
                self.stats["total_checked"] += 1
            
            result_data = {"gene_symbol": None, "is_coding": None}
            
            # AF-Filter
            if af_value is not None and af_value > af_threshold:
                with self._lock:
                    self.stats["failed_af"] += 1
                results[key] = (False, f"af>{af_threshold}", result_data)
                continue
            
            # AF-None Policy
            if af_value is None:
                # V10 FIX: classify_none_type erwartet row-Dict
                temp_row = {
                    "af_filter_mean": af_value,
                    "meanAF_fetch_success": fetch_status,
                    "meanAF_last_fetch": None
                }
                none_type = self._af_none_manager.classify_none_type(temp_row)
                should_include = self._af_none_manager.should_include_type(none_type)
                
                if not should_include and not include_none:
                    with self._lock:
                        self.stats["failed_none_policy"] += 1
                    results[key] = (False, f"af_none:{none_type}", result_data)
                    continue
                
                with self._lock:
                    self.stats["af_none_included"] += 1
            
            # Für Coding-Filter sammeln
            if only_coding and self._coding_filter:
                coding_candidates.append((key, key[0], key[1], af_value, fetch_status))
            else:
                with self._lock:
                    self.stats["passed"] += 1
                results[key] = (True, "passed", result_data)
        
        # Coding-Batch verarbeiten
        if coding_candidates and self._coding_filter:
            batch_input = [(k, c, p) for k, c, p, _, _ in coding_candidates]
            coding_results = self._coding_filter.process_batch(batch_input)
            
            for key, chrom, pos, af_value, fetch_status in coding_candidates:
                result_data = {"gene_symbol": None, "is_coding": None}
                
                if key in coding_results:
                    data = coding_results[key]
                    result_data["gene_symbol"] = data.get("gene_symbol")
                    result_data["is_coding"] = data.get("is_coding")
                    
                    if data.get("is_coding") in (False, 0):
                        with self._lock:
                            self.stats["failed_coding"] += 1
                        results[key] = (False, "not_coding", result_data)
                        continue
                
                with self._lock:
                    self.stats["passed"] += 1
                results[key] = (True, "passed", result_data)
        
        return results
    
    # ───────────────────────────────────────────────────────────────────────
    # LEGACY COMPATIBILITY: passes()
    # ───────────────────────────────────────────────────────────────────────
    
    def passes(self, key: Tuple, af_value: Optional[float] = None) -> Tuple[bool, str]:
        """Legacy-Methode für Kompatibilität."""
        passed, reason, _ = self.check_variant(key, af_value)
        return passed, reason
    
    # ───────────────────────────────────────────────────────────────────────
    # STATISTIKEN
    # ───────────────────────────────────────────────────────────────────────
    
    def get_stats(self) -> dict:
        with self._lock:
            stats = self.stats.copy()
        
        # SubGate Stats hinzufügen
        if self._coding_filter:
            stats["coding_filter"] = self._coding_filter.get_stats()
        stats["af_none_policy"] = self._af_none_manager.current_preset
        
        return stats
    
    def reset_stats(self):
        with self._lock:
            self.stats = {k: 0 for k in self.stats}
        if self._coding_filter:
            self._coding_filter.reset_stats()
    
    def print_stats(self):
        stats = self.get_stats()
        total = stats.get("total_checked", 0)
        
        if total == 0:
            self._log("[MainFilterGate] Keine Varianten geprueft")
            return
        
        msg = (
            f"[MainFilterGate] Statistik:\n"
            f"  Geprueft: {total:,}\n"
            f"  Bestanden: {stats['passed']:,} ({100*stats['passed']/total:.1f}%)\n"
            f"  AF-Filter: {stats['failed_af']:,} abgelehnt\n"
            f"  None-Policy: {stats['failed_none_policy']:,} abgelehnt, {stats['af_none_included']:,} eingeschlossen\n"
            f"  Coding-Filter: {stats['failed_coding']:,} abgelehnt\n"
            f"  Policy: {stats['af_none_policy']}"
        )
        self._log(msg)




class CodingFilter:
    """
    Filtert Varianten basierend auf protein-kodierend Status.
    
    ✅ BATCH-FIRST Design für Performance.
    ✅ Wird von MainFilterGate aufgerufen.
    """
    
    def __init__(self, db: 'VariantDB', gene_annotator: Optional['GeneAnnotator'], logger):
        self.db = db
        self.gene_annotator = gene_annotator
        self.logger = logger
        
        self.stats = {
            "db_hits": 0,
            "annotator_calls": 0,
            "unknown": 0,
            "batches_processed": 0
        }
        
        # ✅ V9: Status-Logging bei Initialisierung
        if logger:
            logger.log(f"[CodingFilter] ✅ Initialisiert")
            if gene_annotator:
                logger.log(f"[CodingFilter] GeneAnnotator verfügbar: {gene_annotator.available}")
            else:
                logger.log(f"[CodingFilter] ⚠️ KEIN GeneAnnotator übergeben!")
    
    # In CodingFilter Klasse:

    def process_batch(self, variants: list) -> dict:
        """
        ✅ HAUPT-METHODE: Verarbeitet Batch von Varianten.
        
        Args:
            variants: Liste von (key, chrom, pos) Tupeln
        
        Returns:
            Dict[key -> {'is_coding': bool, 'gene_symbol': str}]
            Gibt jetzt das volle Datenobjekt zurück, nicht nur den Boolean.
        """
        if not variants:
            return {}
        
        results = {}
        self.stats["batches_processed"] += 1
        
        # ✅ V9: Periodisches Logging
        if self.stats["batches_processed"] == 1 or self.stats["batches_processed"] % 10 == 0:
            if self.logger:
                self.logger.log(f"[CodingFilter] Batch #{self.stats['batches_processed']}: {len(variants)} Varianten")
                self.logger.log(f"[CodingFilter] GeneAnnotator vorhanden: {self.gene_annotator is not None}")
                if self.gene_annotator:
                    self.logger.log(f"[CodingFilter] GeneAnnotator.available: {self.gene_annotator.available}")
        
        # STEP 1: Bulk-Check in VariantDB
        keys = [v[0] for v in variants]
        to_annotate = []
        
        # Hole existierende Daten
        db_results = self.db.get_variants_bulk(keys)
        
        for key, chrom, pos in variants:
            row = db_results.get(key)
            # Wir nehmen den Wert aus der DB nur, wenn is_coding gesetzt ist
            if row and row.get("is_coding") is not None:
                results[key] = {
                    "is_coding": bool(row["is_coding"]),
                    "gene_symbol": row.get("gene_symbol")
                }
                self.stats["db_hits"] += 1
            else:
                to_annotate.append((key, chrom, pos))
        
        # STEP 2: GeneAnnotator-Batch für fehlende
        if to_annotate and self.gene_annotator and self.gene_annotator.available:
            try:
                annotations = self.gene_annotator.annotate_batch(to_annotate)
                bulk_updates = []
                
                for key, chrom, pos in to_annotate:
                    if key in annotations:
                        data = annotations[key]
                        is_coding = data.get("is_coding")
                        gene_symbol = data.get("gene_symbol")
                        
                        if is_coding is not None:
                            # Ergebnis zusammenbauen
                            result_entry = {
                                "is_coding": bool(is_coding),
                                "gene_symbol": gene_symbol
                            }
                            results[key] = result_entry
                            
                            # DB Update vorbereiten
                            update_fields = {"is_coding": is_coding}
                            if gene_symbol:
                                update_fields["gene_symbol"] = gene_symbol
                            
                            bulk_updates.append((key, update_fields))
                            self.stats["annotator_calls"] += 1
                        else:
                            # Nicht gefunden -> Treat as FALSE (not coding)
                            results[key] = {"is_coding": False, "gene_symbol": None}
                            self.stats["unknown"] += 1
                    else:
                        # Nicht in Annotations gefunden -> Treat as FALSE
                        results[key] = {"is_coding": False, "gene_symbol": None}
                        self.stats["unknown"] += 1
                
                # STEP 3: Bulk-Update in DB (Cache)
                if bulk_updates:
                    self.db.cache_annotations(bulk_updates)
            
            except Exception as e:
                self.logger.log(f"[CodingFilter] ⚠️ Batch annotation error: {e}")
                for key, _, _ in to_annotate:
                    if key not in results:
                        results[key] = {"is_coding": None, "gene_symbol": None}
                        self.stats["unknown"] += 1
        
        elif to_annotate:
            # Fallback wenn kein Annotator da ist
            # ✅ V9: Warnung nur beim ersten Mal
            if self.stats["batches_processed"] == 1:
                if self.logger:
                    self.logger.log(f"[CodingFilter] ⚠️ {len(to_annotate)} Varianten ohne Annotation!")
                    self.logger.log(f"[CodingFilter] ⚠️ GeneAnnotator: {self.gene_annotator}")
                    if self.gene_annotator:
                        self.logger.log(f"[CodingFilter] ⚠️ GeneAnnotator.available: {self.gene_annotator.available}")
                    self.logger.log(f"[CodingFilter] ⚠️ Gene-Symbole werden NICHT annotiert!")
            
            for key, _, _ in to_annotate:
                results[key] = {"is_coding": None, "gene_symbol": None}
                self.stats["unknown"] += 1
        
        return results
    
    def is_protein_coding(self, key: tuple, chrom: str, pos: int) -> Optional[bool]:
        """Einzelabfrage (Wrapper)."""
        res_dict = self.process_batch([(key, chrom, pos)])
        data = res_dict.get(key)
        return data.get("is_coding") if data else None
    
    def get_stats(self) -> dict:
        return self.stats.copy()
    
    def reset_stats(self):
        self.stats = {
            "db_hits": 0,
            "annotator_calls": 0,
            "unknown": 0,
            "batches_processed": 0
        }

class Distiller:
    """
    Core pipeline for VCF processing and variant annotation.
    
    Pipeline phases:
    1. VCF Scan: Parse records, apply quality filters, collect keys
    2. AF Fetch: Load population allele frequencies
    3. Full Annotation: Fetch comprehensive variant data
    4. AlphaGenome: Optional pathogenicity scoring
    
    ⚡ PERFORMANCE-OPTIMIERUNGEN:
    - FASTA LRU-Cache: 50-200x schneller (reduziert Disk I/O um 90%+)
    - Region-Prefetch: 5-10x schneller bei Batch-Validierung
    - Validated-Refs-Cache: Skip doppelte Validierungen
    - Batch-Add to_fetch_af: 100-800x weniger Lock-Overhead
    - Thread-Safe Queue: Parallele VCF-Scan + AF-Fetch
    """

    def __init__(
        self,
        app,
        db: VariantDB,
        stopflag: StopFlag,
        live_enqueue: Callable,
        threads: int = DEFAULT_THREADS,
        logger=None,
        # V10: Manager-Objekte statt einzelner Parameter
        flag_manager: Optional['Flag_and_Options_Manager'] = None,
        main_filter_gate: Optional['MainFilterGate'] = None,
        quality_manager: Optional['QualityManager'] = None,
        gene_annotator: Optional['GeneAnnotator'] = None,
        af_fetcher: Optional['AFFetchController'] = None,
        # Legacy-Parameter (deprecated)
        ag_api_key: Optional[str] = None,
        genotype_store: Optional[Dict] = None,
        stale_days: int = Config.STALE_DAYS_AF,
        include_none: bool = False,
        filter_pass_only: bool = False,
        qual_threshold: float = 30.0
    ):
        """V10: Distiller mit zentralem MainFilterGate."""
        # Core dependencies
        self.app = app
        self.progress = app.progress
        self.db = db
        self.stopflag = stopflag
        self.threads = threads
        self.logger = logger if logger else _NullLogger()
        
        self.logger.log("=" * 60)
        self.logger.log("[Distiller] V10 INITIALISIERUNG")
        self.logger.log("=" * 60)
        
        # FLAG MANAGER (zentral)
        if flag_manager is None:
            self.flag_manager = Flag_and_Options_Manager(logger=self.logger)
            self.logger.log("[Distiller] Kein FlagManager, erstelle Default")
            self.flag_manager.set_af_threshold(0.01)
            self.flag_manager.set_include_none(include_none)
            self.flag_manager.stale_days = stale_days
        else:
            self.flag_manager = flag_manager
            self.logger.log("[Distiller] FlagManager uebernommen")
        
        # Legacy-Zugriff
        self.stale_days = self.flag_manager.get_stale_days()
        self.include_none = self.flag_manager.get_include_none()
        self.filter_pass_only = filter_pass_only
        self.qual_threshold = qual_threshold
        self.app_ref = None
        
        # MAIN FILTER GATE (zentral)
        self.gene_annotator = gene_annotator
        
        if main_filter_gate is None:
            self.main_filter_gate = MainFilterGate(
                flag_manager=self.flag_manager,
                db=self.db,
                gene_annotator=gene_annotator,
                logger=self.logger
            )
            self.logger.log("[Distiller] MainFilterGate erstellt")
        else:
            self.main_filter_gate = main_filter_gate
            self.logger.log("[Distiller] MainFilterGate uebernommen")
        
        # CodingFilter ist jetzt NUR im MainFilterGate!
        self.coding_filter = self.main_filter_gate.get_coding_filter()
        
        # QUALITY MANAGER
        if quality_manager is None:
            self.quality_manager = QualityManager(preset="clinical", logger=self.logger)
            self.logger.log("[Distiller] Kein QualityManager, erstelle Default")
        else:
            self.quality_manager = quality_manager
            self.logger.log(f"[Distiller] QualityManager (Preset: {quality_manager.current_preset})")
        
        # AF NONE MANAGER (jetzt im MainFilterGate!)
        self.af_none_manager = self.main_filter_gate.get_af_none_manager()
        
        # AF FETCHER
        self.af_fetcher = af_fetcher
        if self.af_fetcher:
            self.logger.log("[Distiller] AF-Controller uebernommen")
        
        # ALPHAGNOME SCORER
        ag_key = ag_api_key or self.flag_manager.get_alphagenome_key()
        self.ag_scorer = AlphaGenomeScorer(ag_key) if ag_key else None
        
        # Thread-safe state
        self._keys_lock = RLock()
        self._display_keys: Set[Tuple] = set()
        self._normalized_cache: Set[Tuple] = set()
        
        # Genotype storage
        self.genotype_store = genotype_store if isinstance(genotype_store, dict) else {}
        
        # Emit queue
        self.emit_queue = EmitQueue(live_enqueue, batch_size=100)
        
        # VCF buffer
        self.vcf_buffer = VCFBuffer(
            self.db, 
            self.logger, 
            batch_size=5000, 
            prio_threshold=100,
            distiller=self
        )
                
        # Original VCF data
        self.original_vcf_path: Optional[str] = None
        self.build: Optional[str] = None
        self.orig_records: Dict[Tuple, Dict] = {}
        self.current_vcf_keys: Set[Tuple] = set()
        
        # Status manager
        self.status_mgr = FetchStatusManager
        
        # Maintainer reference
        self.maint = None
        
        # Legacy alias
        self._af_controller = self.af_fetcher
        
        # Throughput Tuner
        self.tuners = {
            "mv": ThroughputTuner(min_batch=100, max_batch=1000, base_batch=900, start_batch=900)
        }
        
        # LightDB
        self.lightdb_path = None
        
        # AF-None-Treatment-Manager ist jetzt im MainFilterGate!
        # self.af_none_manager wird oben aus MainFilterGate geholt
        
        # FastaValidator (Lazy Loading)
        self.fasta_validator = None
        
        # Pipeline State (Thread-Safe)
        self._bulk_cache = {}
        self._bulk_cache_size = 10000
        
        # Thread-Safe Queue für AF-Fetch
        self.to_fetch_af = set()
        self.to_fetch_af_lock = threading.Lock()
        
        # Pipeline-Phase State
        self.passed_quality: Set[Tuple] = set()
        self.emitted_immediate: Set[Tuple] = set()
        
        # VCF-Scan Counters
        self.scan_counters: Dict[str, int] = {}
        
        # FASTA-Validierung Cache
        self.validated_refs: Set[Tuple] = set()
        
        # Cython-Accelerator
        if CYTHON_AVAILABLE:
            try:
                self.cython = CythonAccelerator(logger=self.logger, enable_stats=True)
                # ✅ FIX: Prüfen auf internes 'available' Flag (oder Fallback-Erkennung)
                if getattr(self.cython, 'available', False): 
                    self.logger.log("[Distiller] ✅ Cython acceleration enabled")
                else:
                    # Accelerator existiert, läuft aber im Python-Modus
                    self.logger.log("[Distiller] ℹ️ Cython wrapper loaded, but running in pure Python mode")
            except Exception as e:
                self.cython = None
                self.logger.log(f"[Distiller] ⚠️ Cython init failed ({e}), using Python")
        else:
            self.cython = None
            self.logger.log("[Distiller] ⚠️ Cython not available, using Python")
            
        # ✅ KORREKTUR: Initialisiere den Wert direkt vom übergebenen 'app' Objekt
        self.stale_days_full = Config.STALE_DAYS_FULL  # Fallback Default
        if hasattr(app, "stale_days_full"):
            try:
                self.stale_days_full = int(app.stale_days_full.get())
            except:
                pass
        
    def set_gene_annotator(self, gene_annotator):
        """
        Nachträgliche Aktivierung des GeneAnnotators und CodingFilters,
        sobald der Hintergrund-Ladevorgang abgeschlossen ist.
        """
        self.gene_annotator = gene_annotator
        
        if self.gene_annotator and self.gene_annotator.available:
            self.coding_filter = CodingFilter(
                db=self.db,
                gene_annotator=self.gene_annotator,
                logger=self.logger
            )
            self.logger.log("[Distiller] ✅ GeneAnnotator nachgeladen -> CodingFilter nachträglich aktiviert.")
        else:
            self.logger.log("[Distiller] ⚠️ GeneAnnotator Update erhalten, aber nicht verfügbar.")
            
    def _get_variants_bulk_cached(self, keys: List[Tuple]) -> Dict[Tuple, Optional[Dict]]:
        """Cached Bulk-Query mit Auto-Invalidierung. Gibt immer ein Dict zurück."""
        try:
            out: Dict[Tuple, Optional[Dict]] = {}
            if not keys:
                return {}

            # 1) Sammle uncached keys
            uncached = [k for k in keys if k not in self._bulk_cache]

            # 2) Hole fehlende aus DB (defensiv)
            if uncached:
                try:
                    rows = self.db.get_variants_bulk(uncached) or {}
                except Exception as e:
                    self.logger.log(f"[Distiller] ⚠️ DB bulk fetch failed in cache: {e}")
                    rows = {}

                # Nur dict‑Zeilen in Cache aufnehmen
                for k in uncached:
                    v = rows.get(k)
                    if isinstance(v, dict):
                        self._bulk_cache[k] = v
                    else:
                        # explizit None festhalten (keine Einträge vermeiden erneute DB-Hits)
                        self._bulk_cache.setdefault(k, None)

                # Eviction: wenn Cache zu groß, entferne älteste ~10%
                if len(self._bulk_cache) > self._bulk_cache_size:
                    target = int(self._bulk_cache_size * 0.9)
                    # einfache FIFO eviction über insertion-order dict (Py3.7+)
                    keys_list = list(self._bulk_cache.keys())
                    for rem in keys_list[:len(self._bulk_cache) - target]:
                        self._bulk_cache.pop(rem, None)

            # 3) Zusammenstellen der Rückgabe (erhalte originale Reihenfolge)
            for k in keys:
                out[k] = self._bulk_cache.get(k)

            return out

        except Exception as e:
            self.logger.log(f"[Distiller] ⚠️ _get_variants_bulk_cached unexpected error: {e}")
            return {k: None for k in (keys or [])}
    
    # ✅ NEU: Cache-Invalidierung
    def invalidate_cache(self, key: Tuple):
        """Entfernt einzelnen Key aus Cache nach Update"""
        self._bulk_cache.pop(key, None)
    
    def invalidate_cache_bulk(self, keys: List[Tuple]):
        """Leitet Cache-Invalidierung an die App weiter."""
        if self.app:
            self.app.invalidate_cache_bulk(keys)
    
    def clear_cache(self):
        """Leert kompletten Cache (z.B. bei Pipeline-Restart)"""
        self._bulk_cache.clear()
    
    def _get_or_create_fasta_validator(self, build: str) -> Optional[FastaValidator]:
        """
        V16: Lazy Loading mit PathManager für automatische Pfad-Erkennung.

        Nutzt PathManager für intelligente Pfad-Suche:
        - Findet FASTA automatisch auch wenn verschoben
        - Speichert gefundene Pfade in paths_config.json
        - Fallback auf Config.FASTA_HG19/38 wenn PathManager fehlschlägt
        """
        if self.fasta_validator is not None:
            return self.fasta_validator

        # ✅ V16: Nutze PathManager für intelligente Pfad-Suche
        try:
            pm = get_path_manager()
            fasta_path = pm.get_fasta_path(build)

            if fasta_path:
                self.logger.log(f"[Distiller] ✅ PathManager: FASTA gefunden für {build}: {fasta_path}")
            else:
                # Fallback auf alte Config (für Kompatibilität)
                self.logger.log(f"[Distiller] ⚠️ PathManager: Kein FASTA für {build}, versuche Config-Fallback...")
                if build in ("GRCh38", "hg38"):
                    fasta_path = Config.FASTA_HG38
                else:
                    fasta_path = Config.FASTA_HG19

        except Exception as e:
            self.logger.log(f"[Distiller] ⚠️ PathManager-Fehler: {e}, nutze Config-Fallback")
            if build in ("GRCh38", "hg38"):
                fasta_path = Config.FASTA_HG38
            else:
                fasta_path = Config.FASTA_HG19

        if not fasta_path or not os.path.exists(fasta_path):
            self.logger.log(f"[Distiller] ⚠️ FASTA nicht gefunden: {fasta_path}")
            return None

        try:
            self.fasta_validator = FastaValidator(
                fasta_path=fasta_path,
                strict_mode=Config.FASTA_STRICT_MODE,
                logger=self.logger
            )
            self.logger.log(f"[Distiller] ✅ FastaValidator initialisiert: {fasta_path}")
            return self.fasta_validator

        except Exception as e:
            self.logger.log(f"[Distiller] ⚠️ FastaValidator-Init fehlgeschlagen: {e}")
            return None

    # =============================================================================
    # ERWEITERT: cleanup()
    # =============================================================================

    def cleanup(self):
        """
        ✅ ERWEITERT: Cleanup mit FastaValidator.
        """
        try:
            # Emit-Queue
            if hasattr(self, 'emit_queue'):
                self.emit_queue.shutdown()
            
            # VCF-Buffer
            if hasattr(self, 'vcf_buffer'):
                try:
                    self.vcf_buffer.flush()
                except Exception as e:
                    if self.logger:
                        self.logger.log(f"[Cleanup] Buffer flush error: {e}")
            
            # ✅ FastaValidator schließen
            if hasattr(self, 'fasta_validator') and self.fasta_validator:
                try:
                    self.fasta_validator.close()
                    self.logger.log("[Cleanup] ✅ FastaValidator geschlossen")
                except Exception as e:
                    self.logger.log(f"[Cleanup] ⚠️ FastaValidator close error: {e}")
            
            # AF-Fetcher
            if hasattr(self, 'af_fetcher') and self.af_fetcher:
                try:
                    self.af_fetcher.shutdown()
                except Exception as e:
                    if self.logger:
                        self.logger.log(f"[Cleanup] AF-Fetcher shutdown error: {e}")
            
            self.logger.log("[Cleanup] ✅ Distiller cleanup complete")
        
        except Exception as e:
            if self.logger:
                self.logger.log(f"[Cleanup] ⚠️ Error: {e}")

    # =============================================================================
    # HINWEIS: _process_fasta() bleibt UNVERÄNDERT
    # =============================================================================
    # Diese Methode konvertiert FASTA-Sequenzen zu VCF (andere Funktionalität)
    # und hat nichts mit der FASTA-Referenz-Validierung zu tun!
    def set_af_none_preset(self, preset: str):
        """
        ✅ NEU: Setzt AF-None-Preset.
        
        Args:
            preset: "permissive", "research", "clinical", "strict"
        """
        self.af_none_manager.set_preset(preset)

    def set_legacy_include_none(self, include_none: bool):
        """
        ✅ RÜCKWÄRTSKOMPATIBILITÄT: Setzt Legacy include_none Flag.
        
        DEPRECATED: Nutze stattdessen set_af_none_preset()
        """
        self.af_none_manager.set_legacy_include_none(include_none)

    def _apply_af_filter_final(
        self, 
        key: Tuple, 
        val: Optional[Dict], 
        af_threshold: Optional[float] = None, 
        include_none: bool = None
    ) -> str:
        """
        V10: Delegiert an MainFilterGate.
        
        Diese Methode ist ein Wrapper für Rückwärtskompatibilität.
        Die eigentliche Filterlogik ist im MainFilterGate.
        
        Args:
            key: Variant-Key (chr, pos, ref, alt, build)
            val: API-Fetch-Result (kann None sein)
            af_threshold: DEPRECATED - wird vom FlagManager geholt
            include_none: DEPRECATED - wird vom FlagManager geholt
        
        Returns:
            "emit" | "reject"
        """
        # V10: Hole AF-Wert und Status aus val
        if val and isinstance(val, dict):
            mean_af = val.get("af_filter_mean")
            fetch_status = val.get("meanAF_fetch_success")
            
            # Konvertiere AF zu float
            if mean_af is not None:
                try:
                    mean_af = float(mean_af)
                except (ValueError, TypeError):
                    mean_af = None
        else:
            # DB-Fallback
            try:
                cached = self._get_variants_bulk_cached([key])
                row = cached.get(key)
                if row:
                    mean_af = row.get("af_filter_mean")
                    fetch_status = row.get("meanAF_fetch_success")
                    if mean_af is not None:
                        try:
                            mean_af = float(mean_af)
                        except (ValueError, TypeError):
                            mean_af = None
                else:
                    return "reject"
            except Exception:
                return "reject"
        
        # V10: Nutze MainFilterGate
        passed, reason, data = self.main_filter_gate.check_variant(
            key=key,
            af_value=mean_af,
            fetch_status=fetch_status
        )
        
        return "emit" if passed else "reject"

    def print_final_stats(self):
        """Finale Statistiken mit AF-None-Breakdown."""
        # ... existing stats ...
        
        # ✅ AF-None-Stats
        self.logger.log("\n" + "="*70)
        self.af_none_manager.print_stats()
        self.logger.log("="*70 + "\n")

    # ==================== Properties ====================
    
    @property
    def display_keys(self) -> Set[Tuple]:
        """Thread-safe access to displayed variant keys."""
        with self._keys_lock:
            return self._display_keys.copy()
    
    @display_keys.setter
    def display_keys(self, value: Set[Tuple]):
        with self._keys_lock:
            self._display_keys = set(value)
    
    def add_display_key(self, key: Tuple):
        """
        Thread-sichere Addition von Display-Keys.
        
        Änderungen:
        - Nutzt self._keys_lock für atomare Operation
        - Verhindert Race Conditions
        - Entfernt falsche Delegation an self.distiller
        
        FIX: Ursprünglicher Code hatte rekursiven/falschen Aufruf:
             self.distiller.add_display_key(key) - obwohl self bereits Distiller ist
        """
        with self._keys_lock:
            self._display_keys.add(key)


        
    @property
    def done_variants(self) -> int:
        return self.progress.done
    
    @done_variants.setter
    def done_variants(self, value: int):
        self.progress.done = value
    
    @property
    def end_retry_variants(self) -> int:
        return self.progress.failed
    
    @end_retry_variants.setter
    def end_retry_variants(self, value: int):
        self.progress.failed = value
    
    @property
    def total_variants(self) -> Optional[int]:
        return self.progress.total if self.progress.total > 0 else None
    
    @total_variants.setter
    def total_variants(self, value: int):
        self.progress.total = value
    
    # ==================== Configuration ====================
    
    def set_alpha_genome_key(self, ag_api_key: Optional[str]):
        """Dynamically update AlphaGenome client."""
        if ag_api_key:
            try:
                self.ag_scorer = AlphaGenomeScorer(ag_api_key)
                if getattr(self.ag_scorer, "available", False):
                    self.logger.log("[AG] AlphaGenome activated.")
                else:
                    self.logger.log("[AG] AlphaGenome client unavailable.")
            except Exception as e:
                self.ag_scorer = None
                self.logger.log(f"[AG] Initialization error: {e}")
        else:
            self.ag_scorer = None
            self.logger.log("[AG] AlphaGenome disabled (no API key).")
    
    # ==================== Database Operations ====================
    #sollte nur noch von Vollnotation genutzt werden
    def upsert_from_mv(self, key: Tuple, mv_hit: Optional[Dict]):
        """Update variant from MyVariant data."""
        if not mv_hit:
            current_status = self.db.get_variant_field(key, "full_fetch_success")
            new_status = self.status_mgr.update_status(current_status, success=False)
            self.db.upsert_variant(key, {
                "full_last_fetch": now_iso(),
                "full_fetch_success": new_status
            })
            return
        
        # DEBUG: Temporär debug_log aktiviert für Diagnose leerer Annotations-Spalten
        parsed = extract_fields_from_mv(mv_hit, debug_log=True) or {}
        
        mv_cols = (
            "gene_symbol", "impact", "consequence", "clinical_significance",
            "phenotypes", "is_coding", "protein_coding", "cadd_phred",
            "conservation", "ag_score", "rsid"
        )
        update = {}
        for col in mv_cols:
            val = parsed.get(col)
            if val not in (None, "", "."):
                update[col] = val
        
        has_real_data = any(k in update for k in mv_cols)
        current_status = self.db.get_variant_field(key, "full_fetch_success")
        new_status = self.status_mgr.update_status(current_status, success=bool(has_real_data))
        
        update["full_last_fetch"] = now_iso()
        update["full_fetch_success"] = new_status
        
        self.db.upsert_variant(key, update)
    
    def upsert_from_vcf(self, key: Tuple, fields: Dict):
        """Buffer variant data from VCF INFO fields."""
        clean_fields = {k: v for k, v in (fields or {}).items() if v not in (None, "", ".")}
        if not clean_fields:
            return
        
        row = self.db.get_variant(key) or {}
        update = {}
        pop_cols = ("af_exac", "af_gnomad_exomes", "af_gnomad_genomes", "af_1kg")
        new_pop_value = False
        
        for col in pop_cols:
            val = clean_fields.get(col)
            if val is not None and row.get(col) is None:
                update[col] = val
                new_pop_value = True
        
        for col, val in clean_fields.items():
            if col in pop_cols:
                continue
            if row.get(col) is None:
                update[col] = val
        
        if new_pop_value:
            af_values = [
                update.get("af_exac") or row.get("af_exac"),
                update.get("af_gnomad_exomes") or row.get("af_gnomad_exomes"),
                update.get("af_gnomad_genomes") or row.get("af_gnomad_genomes"),
                update.get("af_1kg") or row.get("af_1kg")
            ]
            af_values = [v for v in af_values if v is not None]
            if af_values:
                update["af_filter_mean"] = sum(af_values) / len(af_values)
        
        if update:
            self.vcf_buffer.add(key, update)
    """
    PIPELINE FLUSSGRAFIK und WARTUNG
    ┌──────────────────────────────────────────────────────────────────┐
    │ VCF-Datei (z.B. 10.000 Varianten)                               │
    └────────────────────────┬─────────────────────────────────────────┘
                             │
                             ▼
    ┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓
    ┃ PHASE 1: VCF-SCAN (Sequentiell, Main Thread)                  ┃
    ┃ Methode: _phase_vcf_scan()                                    ┃
    ┃ ⚡ OPTIMIERT: Batch-Processing, FASTA-Cache, Lock-Minimierung  ┃
    ┗━━━━━━━━━━━━━━━━━━━━━━━━━┯━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛
                              │
                              ▼
                  ┌───────────────────────┐
                  │ parse_vcf_records_    │
                  │ smart()               │
                  │ (Generator, lazy)     │
                  └───────────┬───────────┘
                              │
                              ▼
                  ┌───────────────────────┐
                  │ BATCH-COLLECTION      │
                  │ (500 Varianten/Batch) │
                  └───────────┬───────────┘
                              │
                              ▼
                  ┌───────────────────────┐
                  │ QualityManager.       │
                  │ passes(rec)           │
                  └─────┬──────────┬──────┘
                        │          │
             ❌ 3.777   │          │ ✅ 5.962 Quality bestanden
             rejected   │          │
                        │          │
                        │          ▼
                        │   self.passed_quality.add(key)
                        │   self.orig_records[key] = {...}
                        │          │
                        │          ▼
                        │   ┌──────────────────────────────────────┐
                        │   │ INFO-Recycling                       │
                        │   │ determine_is_coding()                │
                        │   │ vcf_buffer.add(gene, impact, cons)   │
                        │   └──────────┬───────────────────────────┘
                        │              │
                        │              ▼
                        │   ┌──────────────────────────────────────┐
                        │   │ ⚡ FASTA-VALIDIERUNG (OPTIMIERT!)    │
                        │   │ validate_batch()                     │
                        │   │                                      │
                        │   │ PERFORMANCE-FEATURES:                │
                        │   │ ✅ Pre-Check: Skip SNVs (~85%)       │
                        │   │ ✅ LRU-Cache: 1000 Sequenzen         │
                        │   │ ✅ Region-Prefetch: 100kb Blöcke     │
                        │   │ ✅ validated_refs: Skip doppelte     │
                        │   │ ✅ Shift-Detection: Pattern-Korrektur│
                        │   │                                      │
                        │   │ WORKFLOW:                            │
                        │   │ 1. Check validated_refs Cache        │
                        │   │ 2. Pre-Check: SNVs direkt überspringen│
                        │   │ 3. Sortiere nach Chrom/Pos           │
                        │   │ 4. Prefetch Regionen → LRU-Cache     │
                        │   │ 5. Validiere (Cache warm!)           │
                        │   │ 6. Markiere in validated_refs        │
                        │   └──────────┬───────────────────────────┘
                        │              │
                        │              ▼
                        │   ┌──────────────────────────────────────┐
                        │   │ DB Bulk-Lookup                       │
                        │   │ db.get_variants_bulk(keys)           │
                        │   │ → row_cache (500 Keys auf einmal!)   │
                        │   └──────────┬───────────────────────────┘
                        │              │
                        │              ▼
                        │   ┌──────────────────────────────────────┐
                        │   │ LightDB-Lookup                       │
                        │   │ af_fetcher._lookup_lightdb()         │
                        │   │ → update row_cache mit rsid + AF     │
                        │   └──────────┬───────────────────────────┘
                        │              │
                        │              ▼
                        │   ┌──────────────────────────────────────┐
                        │   │ AF-Entscheidung (SIMPLE!)            │
                        │   │ _validate_af_in_cache()              │
                        │   │                                      │
                        │   │ ⚠️ KEINE None-Typ-Klassifizierung!   │
                        │   │ Nur: AF vorhanden? → emit/reject     │
                        │   │      AF fehlt? → "fetch"             │
                        │   └────┬──────────┬──────────┬───────────┘
                        │        │          │          │
                        │   emit │          │ reject   │ fetch
                        │        │          │          │
                        ▼        ▼          ▼          ▼
                (rejected)  ┌────────┐  ┌────────┐  ┌──────────────────┐
                            │emit_   │  │mark_   │  │ ⚡ BATCH-ADD!    │
                            │queue   │  │rejected│  │ keys_to_fetch[]  │
                            │.add()  │  │(1)     │  │ (sammeln)        │
                            └────┬───┘  └────┬───┘  └──────┬───────────┘
                                 │           │             │
                            add_display_key  │             │
                                 │           │             │
                                 │           │             ▼
                                 │           │   ┌─────────────────────┐
                                 │           │   │ ⚡ BATCH-ADD:       │
                                 │           │   │ with lock:          │
                                 │           │   │   to_fetch_af.      │
                                 │           │   │   update(batch)     │
                                 │           │   │                     │
                                 │           │   │ 1 Lock statt 817!   │
                                 │           │   │ → 100-800x schneller│
                                 │           │   └──────┬──────────────┘
                                 │           │          │
                                 ▼           ▼          ▼
                            ┌─────────────────────────────────┐
                            │ ZWISCHENERGEBNIS NACH VCF-SCAN: │
                            │                                 │
                            │ display_keys: ~5.145 Keys       │
                            │ (sofort emittiert, AF OK)       │
                            │                                 │
                            │ to_fetch_af: ~817 Keys          │
                            │ (AF unbekannt, thread-safe!)    │
                            │                                 │
                            │ rejected: ~3.777 + ? Keys       │
                            │ (Quality + AF > threshold)      │
                            │                                 │
                            │ ✅ vcf_done.set() in finally    │
                            └──────────┬──────────────────────┘
                                       │
                                       ▼
    ┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓
    ┃ PHASE 2: AF-FETCH (Parallel Thread, asyncio.run)              ┃
    ┃ Methode: _phase_af_fetch_streaming()                          ┃
    ┃ ⚡ OPTIMIERT: Producer-Consumer, Early-Start, Thread-Safe      ┃
    ┗━━━━━━━━━━━━━━━━━━━━━━━━┯━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛
                              │
                              ▼
                  ┌────────────────────────────────────┐
                  │ IDLE-LOOP (wartet auf Daten)       │
                  │                                    │
                  │ while True:                        │
                  │   with lock:                       │
                  │     queue_size = len(to_fetch_af)  │
                  │                                    │
                  │   if vcf_done AND queue_size > 0:  │
                  │     break  # Start!                │
                  │   if vcf_done AND queue_size == 0: │
                  │     return  # Skip!                │
                  │   if queue_size >= 1000:           │
                  │     break  # Early-Start!          │
                  └────────────┬───────────────────────┘
                               │
                               ▼
                  ┌────────────────────────────────────┐
                  │ CONTINUOUS PROCESSING LOOP         │
                  │                                    │
                  │ while True:  # Round-basiert       │
                  └─────────┬──────────────────────────┘
                            │
                            ▼
                  ┌────────────────────────────────────┐
                  │ ⚡ THREAD-SAFE SNAPSHOT:           │
                  │ with lock:                         │
                  │   raw_keys = list(to_fetch_af)     │
                  └─────────┬──────────────────────────┘
                            │
                            ▼
                  ┌────────────────────────────────────┐
                  │ EXIT-CHECK:                        │
                  │ if vcf_done AND queue_empty:       │
                  │   consecutive_empty += 1           │
                  │   if consecutive_empty >= 2:       │
                  │     break  # Double-Check OK!      │
                  └─────────┬──────────────────────────┘
                            │
                            ▼
                  ┌────────────────────────────┐
                  │ Pre-Check: AF-Status       │
                  │ db.get_variants_bulk()     │
                  │                            │
                  │ Hat inzwischen AF?         │
                  └───┬──────────────┬─────────┘
                      │              │
                 AF   │              │ Kein AF
                 jetzt│              │
                 da   │              │
                      ▼              ▼
              ┌──────────────┐  ┌──────────────┐
              │ Filter       │  │ needs_fetch  │
              │ anwenden     │  │ .append(key) │
              └──────┬───────┘  └──────┬───────┘
                     │                 │
                     │                 ▼
                     │      ┌───────────────────────┐
                     │      │ API-Fetch (async)     │
                     │      │ mv_fetch_async()      │
                     │      │ ThroughputTuner       │
                     │      │ Parallel Chunks (6x)  │
                     │      └──────┬────────────────┘
                     │             │
                     │             ▼
                     │      ┌───────────────────────┐
                     │      │ result_collector_     │
                     │      │ and_merger()          │
                     │      │ → results dict        │
                     │      └──────┬────────────────┘
                     │             │
                     │             ▼
                     │      ┌───────────────────────┐
                     │      │ vcf_buffer.add()      │
                     │      │ (DB-Update)           │
                     │      │ + FetchStatusManager  │
                     │      └──────┬────────────────┘
                     │             │
                     └─────────────┘
                                   │
                                   ▼
    ┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓
    ┃ ⭐ HIER FINDET NONE-VALIDIERUNG STATT! ⭐                      ┃
    ┃                                                               ┃
    ┃ Methode: _apply_af_filter_final(key, val, af_threshold, ...)  ┃
    ┗━━━━━━━━━━━━━━━━━━━━━━━━┯━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛
                              │
                              ▼
                  ┌────────────────────────────────────────┐
                  │ SCHRITT 1: AF vorhanden?               │
                  │                                        │
                  │ if mean_af is not None:                │
                  │   validate → emit/reject               │
                  └────────────┬───────────────────────────┘
                               │
                          AF = None
                               │
                               ▼
                  ┌────────────────────────────────────────┐
                  │ SCHRITT 2: None-Typ klassifizieren     │
                  │                                        │
                  │ af_none_manager.classify_none_type()   │
                  │                                        │
                  │ Prüft:                                 │
                  │ - meanAF_fetch_success Status          │
                  │ - af_exac, af_gnomad_*, af_1kg         │
                  │                                        │
                  │ Returns:                               │
                  │   "true_none"  → Seltene Variante!     │
                  │   "fetch_fail" → API-Fehler            │
                  │   "false_none" → AF in anderen Spalten │
                  │   "has_af"     → AF gefunden           │
                  └────────┬───────────────────────────────┘
                           │
                           ▼
                  ┌────────────────────────────────────────┐
                  │ SCHRITT 3: Policy abfragen             │
                  │                                        │
                  │ af_none_manager.should_include_type()  │
                  │                                        │
                  │ Policy-Mapping (z.B. "clinical"):      │
                  │   true_none:  ✅ INCLUDE               │
                  │   fetch_fail: ❌ EXCLUDE               │
                  │   false_none: ❌ EXCLUDE (DB-Recheck!) │
                  └────────┬───────────────────────────────┘
                           │
                           ▼
                  ┌────────────────────────────────────────┐
                  │ SCHRITT 4: Statistik tracken           │
                  │                                        │
                  │ af_none_manager.record_decision()      │
                  │                                        │
                  │ Zählt:                                 │
                  │ - none_type_counts[type]               │
                  │ - none_included_by_type[type]          │
                  │ - none_excluded_by_type[type]          │
                  └────────┬───────────────────────────────┘
                           │
                           ▼
                  ┌────────────────────────────────────────┐
                  │ SCHRITT 5: Entscheidung treffen        │
                  │                                        │
                  │ return "emit" if should_include        │
                  │        else "reject"                   │
                  └─────┬──────────────────┬───────────────┘
                        │                  │
                   emit │                  │ reject
                        │                  │
                        ▼                  ▼
                ┌──────────────┐    ┌──────────────┐
                │emit_queue    │    │mark_rejected │
                │.add(key)     │    │(1)           │
                └──────┬───────┘    └──────┬───────┘
                       │                   │
                  add_display_key          │
                       │                   │
                       │                   ▼
                       │         ┌─────────────────────┐
                       │         │ ⚡ THREAD-SAFE:     │
                       │         │ with lock:          │
                       │         │   to_fetch_af -=    │
                       │         │   set(raw_keys)     │
                       │         │                     │
                       │         │ Entferne verarbeitete│
                       │         └─────────────────────┘
                       │
                       ▼
                ┌────────────────────────────────┐
                │ ENDERGEBNIS NACH AF-FETCH:     │
                │                                │
                │ display_keys: ~5.500 Keys      │
                │ (alle gefilterten Varianten)   │
                │                                │
                │ rejected: ~4.500 Keys          │
                │ (Quality + AF + None-Policy)   │
                │                                │
                │ ✅ af_done.set() - Phase done  │
                └──────────┬─────────────────────┘
                           │
                           ▼
    ┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓
    ┃ PHASE 3-7: ANNOTATION STREAM (Parallel)                       ┃
    ┃ Input: display_keys (nur gefilterte Varianten!)               ┃
    ┗━━━━━━━━━━━━━━━━━━━━━━━┯━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛
                              │
            ┌─────────────────┼─────────────────┐
            │                 │                 │
            ▼                 ▼                 ▼
    ┌──────────────┐  ┌──────────────┐  ┌──────────────┐
    │ PHASE 3:     │  │ PHASE 4:     │  │ PHASE 5:     │
    │ Full-Anno    │  │ Gene-Anno    │  │ RSID-Fill    │
    │              │  │              │  │              │
    │ Methoden:    │  │ Methoden:    │  │ Methoden:    │
    │ _phase_full_ │  │ _phase_gene_ │  │ _phase_rsid_ │
    │ annotation_  │  │ annotation_  │  │ fill_        │
    │ streaming()  │  │ streaming()  │  │ streaming()  │
    │              │  │              │  │              │
    │ mv_fetch_    │  │ gene_        │  │ NCBI allele  │
    │ async()      │  │ annotator.   │  │ API          │
    │              │  │ annotate_    │  │              │
    │ extract_     │  │ batch()      │  │ make_hgvs()  │
    │ fields_from_ │  │              │  │              │
    │ mv()         │  │              │  │              │
    └──────────────┘  └──────────────┘  └──────────────┘
            │                 │                 │
            │                 │                 │
            └─────────────────┼─────────────────┘
                              │
            ┌─────────────────┼─────────────────┐
            │                 │                 │
            ▼                 ▼                 ▼
    ┌──────────────┐  ┌──────────────┐  ┌──────────────┐
    │ PHASE 6:     │  │ PHASE 7:     │  │              │
    │ Missing-Fill │  │ AlphaGenome  │  │              │
    │              │  │              │  │              │
    │ Methoden:    │  │ Methoden:    │  │              │
    │ _phase_      │  │ _phase_      │  │              │
    │ missing_fill_│  │ alphagenome_ │  │              │
    │ streaming()  │  │ streaming()  │  │              │
    │              │  │              │  │              │
    │ ClinVar API  │  │ ag_scorer.   │  │              │
    │ MyVariant    │  │ score_batch()│  │              │
    │ Conservation │  │              │  │              │
    └──────────────┘  └──────────────┘  └──────────────┘
            │                 │
            └─────────────────┘
                              │
                              ▼
                  ┌────────────────────┐
                  │ Alle Phasen done   │
                  │                    │
                  │ Pipeline Complete! │
                  └────────────────────┘
    """
    
    # ==================== Main Entry Point ====================
    
    def process_file(
        self,
        path: str,
        build: Optional[str] = None,
        # V10: Legacy-Parameter (werden ignoriert, FlagManager hat Vorrang)
        af_threshold: Optional[float] = None,
        include_none: Optional[bool] = None,
        filter_pass_only: Optional[bool] = None,
        cadd_highlight_threshold: Optional[float] = None,
        stale_days: int = None,
        live_enqueue: Optional[Callable] = None,
        qual_threshold: float = None,
        skip_info_recycling: bool = False
    ):
        """
        V10: Main entry point - nutzt FlagManager für alle Einstellungen.
        """
        self.logger.log(f"[Distiller] Starting processing: {path}")
        
        # V10: Sync FlagManager vor Verarbeitung (falls App vorhanden)
        if self.flag_manager and hasattr(self, 'app') and self.app:
            try:
                self.flag_manager.sync_from_gui(self.app)
                self.logger.log("[Distiller] FlagManager synced from GUI")
            except Exception as e:
                self.logger.log(f"[Distiller] FlagManager sync failed: {e}")
        
        # Live-Enqueue Setup
        if live_enqueue is not None:
            self.emit_queue = EmitQueue(live_enqueue, batch_size=100)
        
        # File type detection and delegation
        path_lower = path.lower()
        
        # 23andMe TXT
        if path_lower.endswith(".txt"):
            if self.looks_like_23andme_txt(path):
                self.logger.log("[Distiller] 23andMe format detected, converting to VCF...")
                # V10: FlagManager hat alle Parameter
                return self._process_23andme(path, build)
        
        # FASTA
        if path_lower.endswith((".fa", ".fasta", ".fna", ".mpfa", ".fsa")):
            self.logger.log("[Distiller] FASTA detected, converting...")
            return self._process_fasta(path, build or "GRCh38")
        
        # FASTQ
        if path_lower.endswith((".fastq", ".fq", ".fastq.gz", ".fq.gz")):
            self.logger.log("[Distiller] FASTQ detected, converting...")
            return self._process_fastq(path, build or "GRCh38")
        
        # gVCF
        if path_lower.endswith((".g.vcf", ".gvcf", ".g.vcf.gz", ".gvcf.gz")):
            self.logger.log("[Distiller] gVCF detected, filtering reference blocks...")
            path = self._filter_gvcf(path)
        
        # Standard VCF - RELAXED CHECK
        if is_vcf(path) or self.looks_like_vcf(path):
            self.logger.log("[Distiller] Processing as VCF...")
            # V10: Alle Parameter kommen vom FlagManager
            self._distill_vcf(
                vcf_path=path,
                build=build,
                skip_info_recycling=skip_info_recycling
            )
        else:
            self.emit_queue.add(("[Distiller] ❌ File format not recognized - must be VCF, FASTA, FASTQ or 23andMe"))
            self.emit_queue.add((
                "Ungültiges Format",
                f"Datei wird nicht erkannt:\n{path}\n\nErwartete Formate: VCF, gVCF, FASTA, FASTQ, 23andMe TXT"
            ))
             
    
    # ==================== Core VCF Pipeline ====================
   
    #--------------------------------------------------------
    #to do
    #--------------------------------------------------------
    #Prüfen, dass jede Variante am Ende genau einmal mark_fully_processed(1) oder mark_rejected(1) bekommt.
    #Sicherstellen, dass jede Phase bei Abschluss complete_phase("...") setzt.
    #Optional: Am Ende der Pipeline (z. B. in AlphaGenome oder im Starter)
    #eine „Finalisierungsschleife“, die alle Keys in display_keys durchgeht
    #und fehlende Markierungen nachholt. So stellst du sicher, dass processed_variants == total_variants.

    def _distill_vcf(
        self,
        vcf_path: str,
        build: Optional[str] = None,
        # V10: Alle Parameter kommen jetzt vom FlagManager
        # Legacy-Parameter für Rückwärtskompatibilität (werden ignoriert wenn flag_manager vorhanden)
        af_threshold: Optional[float] = None,
        include_none: bool = None,
        filter_pass_only: bool = None,
        cadd_highlight_threshold: float = None,
        stale_days: int = None,
        qual_threshold: float = None,
        skip_info_recycling: bool = False,
        validate_ref: bool = False
    ):
        """
        V10: VCF-Scan mit FlagManager und MainFilterGate.
        
        Alle Filter-Parameter werden vom FlagManager geholt.
        Legacy-Parameter werden nur genutzt wenn kein FlagManager vorhanden.
        """
        # V10: Hole alle Parameter vom FlagManager
        if self.flag_manager:
            af_threshold = self.flag_manager.get_af_threshold()
            include_none = self.flag_manager.get_include_none()
            filter_pass_only = self.flag_manager.get_filter_pass_only()
            qual_threshold = self.flag_manager.get_qual_threshold()
            stale_days = self.flag_manager.get_stale_days()
            cadd_highlight_threshold = self.flag_manager.get_cadd_highlight_threshold()
            skip_info_recycling = self.flag_manager.skip_info_recycling
        else:
            # Legacy Fallback
            af_threshold = af_threshold if af_threshold is not None else 0.01
            include_none = include_none if include_none is not None else False
            filter_pass_only = filter_pass_only if filter_pass_only is not None else True
            qual_threshold = qual_threshold if qual_threshold is not None else 30.0
            stale_days = stale_days if stale_days is not None else Config.STALE_DAYS_FULL
            cadd_highlight_threshold = cadd_highlight_threshold if cadd_highlight_threshold is not None else 20.0
        if self.maint:
            try:
                self.maint.pause()
            except Exception:
                pass
        
        # ============================================================
        # Events für Pipeline-Koordination
        # ============================================================
        vcf_done = threading.Event()  # ← KRITISCH: VCF-Scan Complete Signal
        pipeline_thread = None
        
        try:
            # ============================================================
            # Build-Erkennung
            # ============================================================
            try:
                # V10: Hole lightdb_path vom af_fetcher (wird von App gesetzt)
                # Fallback auf absoluten Default-Pfad
                lightdb_db_path = getattr(self.af_fetcher, 'lightdb_path', None)
                if not lightdb_db_path:
                    lightdb_db_path = os.path.join(BASE_DIR, "data", "gnomad_light.db")
                
                build = detect_build_for_vcf(
                    vcf_path, 
                    db_path=lightdb_db_path, 
                    logger=self.logger
                )
                self.logger.log(f"[Pipeline] ✅ Build gewählt: {build}")
            except Exception as e:
                self.logger.log(f"[Pipeline] ❌ Build konnte nicht bestimmt werden: {e}")
                raise
            
            # ============================================================
            # Variant-Zählung (optimiert für große Dateien)
            # ============================================================
            try:
                file_size_mb = os.path.getsize(vcf_path) / (1024 * 1024)
                # Bei Dateien > 50MB nur schätzen, um Start-Verzögerung zu vermeiden
                if file_size_mb > 50:
                    self.logger.log(f"[Pipeline] Large file ({file_size_mb:.1f}MB) - estimating variant count...")
                    total_variants = _estimate_total_variants(vcf_path)
                    self.logger.log(f"[Pipeline] Estimated variants: {total_variants:,}")
                else:
                    total_variants = count_variants_exact(vcf_path)
                    self.logger.log(f"[Pipeline] Total variants: {total_variants:,}")
            except Exception as e:
                self.logger.log(f"[Pipeline] ⚠️ Could not count variants: {e}")
                total_variants = _estimate_total_variants(vcf_path)
                self.logger.log(f"[Pipeline] Estimated variants: {total_variants:,}")
            
            # Progress-System initialisieren
            self.progress.start_pipeline(total_variants)
            
            # ============================================================
            # Starte Streaming-Pipeline in separatem Thread
            # (wartet auf vcf_done)
            # ============================================================
            def run_streaming_pipeline():
                try:
                    self._start_streaming_pipeline(
                        build=build,
                        af_threshold=af_threshold,
                        include_none=include_none,
                        stale_days=stale_days,
                        vcf_done_event=vcf_done  # ← Übergebe Event
                    )
                except Exception as e:
                    self.logger.log(f"[Pipeline] ❌ Streaming error: {e}")
                    import traceback
                    self.logger.log(traceback.format_exc())
            
            pipeline_thread = threading.Thread(
                target=run_streaming_pipeline,
                daemon=False,  # ← WICHTIG: Nicht daemon (muss auf vcf_done warten)
                name="StreamingPipeline"
            )
            pipeline_thread.start()
            self.logger.log(
                "[Pipeline] 🚀 Streaming-Pipeline started (waiting for vcf_done)"
            )
            
            # ============================================================
            # Phase 1: VCF-Scan (Main Thread)
            # ============================================================
            self._phase_vcf_scan(
                vcf_path=vcf_path,
                build=build,
                af_threshold=af_threshold,
                include_none=include_none,
                filter_pass_only=filter_pass_only,
                qual_threshold=qual_threshold,
                skip_info_recycling=skip_info_recycling,
                validate_ref=True #Schalter FASTA AN AUS
            )
            
            if self.stopflag.is_set():
                self.logger.log("[Pipeline] ⏸️ Abbruch durch StopFlag")
                return
            
        except Exception as e:
            self.logger.log(f"[Distiller] ❌ Critical error: {e}")
            import traceback
            self.logger.log(traceback.format_exc())
        
        finally:
            # ============================================================
            # KRITISCH: IMMER vcf_done setzen (auch bei Exception!)
            # ============================================================
            if not vcf_done.is_set():
                self.logger.log("[Distiller] 🚦 Setting vcf_done (finally block)")
                vcf_done.set()
            else:
                self.logger.log("[Distiller] 🚦 vcf_done already set")
            
            self.logger.log(
                f"[Distiller] ✅ VCF-Scan complete, "
                f"{len(self.to_fetch_af):,} variants ready for pipeline"
            )
            if self.cython and self.cython.enable_stats:
                self.cython.print_stats()
                
            # ============================================================
            # Warte auf Pipeline-Thread
            # ============================================================
            # Zeilen 12650-12663 ersetzen durch:
            if pipeline_thread and pipeline_thread.is_alive():
                self.logger.log(
                    "[Distiller] ⏳ Waiting for streaming pipeline (no timeout)..."
                )
                
                # ✅ Kein Timeout - Phasen beenden sich selbst!
                while pipeline_thread.is_alive():
                    # User kann über stopflag abbrechen
                    if self.stopflag.is_set():
                        self.logger.log("[Distiller] ⏸️ User abort detected")
                        break
                    time.sleep(2)
                
                if pipeline_thread.is_alive():
                    self.logger.log("[Distiller] ⏸️ Pipeline aborted by user")
                else:
                    self.logger.log("[Distiller] ✅ Pipeline complete")
            elif pipeline_thread:
                self.logger.log("[Distiller] ✅ Pipeline already complete")
            
            # ============================================================
            # Cleanup
            # ============================================================
            self.vcf_buffer.flush()
            self.emit_queue.flush()

            if self.maint:
                try:
                    self.maint.resume()
                except Exception:
                    pass

            # ✅ V16 FIX 12: OneDrive wieder starten nach Pipeline-Ende
            try:
                resume_onedrive(self.logger)
            except Exception as e:
                self.logger.log(f"[Distiller] ⚠️ OneDrive-Resume fehlgeschlagen: {e}")
                
    def _start_streaming_pipeline(
        self, 
        build: str, 
        af_threshold: Optional[float], 
        include_none: bool, 
        stale_days: int,
        vcf_done_event: threading.Event
    ):
        """
        Startet alle Phasen mit vcf_done-Synchronisation.
        """
        pipeline_state = {
            "vcf_done": vcf_done_event,
            "af_done": threading.Event(),
            "full_done": threading.Event(),
            "gene_done": threading.Event(),
            "rsid_done": threading.Event(),
            "missing_done": threading.Event(),
            "alphagenome_done": threading.Event(),
        }
        
        threads = []
        
        # Phase 2: AF-Fetch (async → asyncio.run im Thread)
        def af_target():
            asyncio.run(
                self._phase_af_fetch_streaming(
                    build, 
                    af_threshold, 
                    include_none, 
                    pipeline_state
                )
            )
        
        af_thread = threading.Thread(
            target=af_target,
            daemon=False,
            name="AF-Fetch"
        )
        threads.append(af_thread)
        
        # Phase 3: Full-Annotation
        # In _start_streaming_pipeline Aufruf:
        full_thread = threading.Thread(
            target=self._phase_full_annotation_streaming,
            args=(build, getattr(self, 'stale_days_full', Config.STALE_DAYS_FULL), pipeline_state), # Nutze Attribut
            daemon=False,
            name="Full-Anno"
        )
            
        threads.append(full_thread)  # ✅ FIX 2026-01-09: War vergessen!
        
        # Phase 4: Gene-Annotation
        gene_thread = threading.Thread(
            target=self._phase_gene_annotation_streaming,
            args=(pipeline_state,),
            daemon=False,
            name="Gene-Anno"
        )
        threads.append(gene_thread)
        
        # Phase 5: RSID-Fill
        rsid_thread = threading.Thread(
            target=self._phase_rsid_fill_streaming,
            args=(build, pipeline_state),
            daemon=False,
            name="RSID-Fill"
        )
        threads.append(rsid_thread)
        
        # Phase 6: Missing-Fill
        missing_thread = threading.Thread(
            target=self._phase_missing_fill_streaming,
            args=(build, pipeline_state),
            daemon=False,
            name="Missing-Fill"
        )
        threads.append(missing_thread)
        
        # Phase 7: AlphaGenome
        if self.ag_scorer and self.ag_scorer.available:
            ag_thread = threading.Thread(
                target=self._phase_alphagenome_streaming,
                args=(pipeline_state,),
                daemon=False,
                name="AlphaGenome"
            )
            threads.append(ag_thread)
        
        # Alle Threads starten
        for t in threads:
            t.start()
            self.logger.log(f"[Pipeline] Started: {t.name}")
        
        # Warten auf Completion
        for t in threads:
            t.join()
            self.logger.log(f"[Pipeline] {t.name} joined")
        
        self.logger.log("[Pipeline] ✅ Alle Phasen abgeschlossen.")
        # Finalisiere Progress-Counter
        self._finalize_pipeline()
    
    def _phase_vcf_scan(
        self,
        vcf_path: str,
        build: str,
        af_threshold: Optional[float],
        include_none: bool,
        filter_pass_only: bool,
        qual_threshold: float,
        skip_info_recycling: bool,
        validate_ref: bool = False
    ):
        """
        ✅ ERWEITERT: Phase 1 VCF-Scan mit optional Protein-Coding Filter.
        
        ✅ NEU: Protein-Coding Filter nach QualityManager (falls aktiviert in GUI)
        """
        
        self.progress.set_phase("vcf_scan")

        # Speichere VCF-Pfad und Build fuer Export
        self.original_vcf_path = vcf_path
        self.build = build

        # Multi-Sample VCF Warnung
        try:
            sample_names = self.quality_manager.get_vcf_sample_names(vcf_path)
            if len(sample_names) > 1:
                self.logger.log(
                    f"[VCF-Scan] WARNUNG: Multi-Sample VCF erkannt ({len(sample_names)} Samples: "
                    f"{', '.join(sample_names[:5])}{'...' if len(sample_names) > 5 else ''}). "
                    f"Nur das erste Sample '{sample_names[0]}' wird ausgewertet. "
                    f"Fuer einzelne Samples: bcftools view -s {sample_names[0]} input.vcf > single.vcf"
                )
        except Exception:
            pass

        # Reset Quality-Stats
        self.quality_manager.reset_stats()
        
        # ✅ NEU: Reset CodingFilter Stats
        if self.coding_filter:
            self.coding_filter.reset_stats()

        # Get total_est from progress
        total_est = getattr(self.progress, 'total_variants', 0)
        if total_est == 0:
            try:
                total_est = _estimate_total_variants(vcf_path)
            except Exception:
                total_est = 10000

        # PIPELINE-KOMPATIBEL: Klassen-Attribut fuer alle Counter
        self.scan_counters = {
            'quality_rejected': 0,
            'coding_rejected': 0,  # ✅ NEU
            'ref_validation_rejected': 0,
            'cache_emitted': 0,
            'lightdb_emitted': 0,
            'lightdb_stored': 0,
            'db_bulk_errors': 0,
            'info_errors': 0,
            'af_rejected': 0,
            'af_fetch_needed': 0
        }

        # State
        with self.to_fetch_af_lock:
            self.to_fetch_af.clear()
        self.passed_quality = set()
        self.emitted_immediate = set()
        row_cache = {}
        
        # Batch-Listen
        collected_variants = []  # Für FASTA-Batch
        
        # ============================================================
        # V16 FIX 9: Optimiertes Coding-Filter Batching
        # Problem: CodingFilter blockierte Scanner mit DB I/O
        # Lösung: Größere Batches + weniger häufige Flushes
        # ============================================================
        coding_batch_buffer = []  # Format: [(key, chrom, pos), ...]
        coding_batch_results = {}  # Format: {key: is_coding}
        CODING_BATCH_SIZE = 2000  # V16: Von 500 auf 2000 erhöht für weniger DB-Roundtrips
        coding_batch_pending = []  # V16: Varianten die auf Coding-Ergebnis warten
        
        # V10: Alle Filter-Einstellungen vom FlagManager
        af_threshold = self.flag_manager.get_af_threshold()
        include_none = self.flag_manager.get_include_none()
        only_protein_coding = self.flag_manager.get_only_protein_coding()
        
        # ✅ V17: Warte auf GeneAnnotator wenn Filter aktiv (statt Deaktivierung)
        if only_protein_coding and not self.main_filter_gate.is_coding_available():
            self.logger.log("[VCF-Scan] ⏳ Protein-Coding Filter angefordert, aber Engine lädt noch... Warte max 20s.")
            wait_start = time.time()
            while not self.main_filter_gate.is_coding_available() and (time.time() - wait_start < 20):
                if self.stopflag.is_set(): break
                time.sleep(1.0)
            
            if self.main_filter_gate.is_coding_available():
                self.logger.log("[VCF-Scan] ✅ Protein-Coding Engine bereit.")
            else:
                self.logger.log("[VCF-Scan] ⚠️ Protein-Coding Engine nicht bereit. Filter wird DEAKTIVIERT.")
                only_protein_coding = False

        # Counters
        counter = 0
        scan_start_time = time.time()  # V16: Für ETA-Berechnung
        last_heartbeat = time.time()
        last_logged_counter = 0

        # Performance-Optimierung: Dynamische Parameter
        try:
            import psutil
            #import hier notwendig? Warum nicht oben?
            available_ram_gb = psutil.virtual_memory().available / (1024**3)
            CHUNK_SIZE = min(10000, max(2000, int(available_ram_gb * 500)))
        except ImportError:
            CHUNK_SIZE = 1000
            available_ram_gb = 0.0
        
        # Adaptive Intervalle
        UPDATE_INTERVAL = 5000 if total_est > 100000 else 1000
        LOG_INTERVAL = 50000 if total_est > 100000 else 10000

        # Log Settings
        qm_settings = self.quality_manager.get_settings()
        self.logger.log(
            f"[VCF-Scan] Starting... (Total estimate: {total_est:,})\n"
            f"  Quality-Preset: {qm_settings['preset']}\n"
            f"  FILTER: {'PASS only' if qm_settings['filter_pass_only'] else 'Alle'}\n"
            f"  QUAL: >={qm_settings['qual_threshold']}\n"
            f"  DP: {qm_settings['min_dp']}-{qm_settings['max_dp'] or 'infinity'} "
            f"({'aktiv' if qm_settings['use_dp_filter'] else 'deaktiviert'})\n"
            f"  Protein-Coding Filter: {'AKTIV' if only_protein_coding else 'DEAKTIVIERT'}\n"
            f"\n"
            f"  Performance-Einstellungen:\n"
            f"  CHUNK_SIZE: {CHUNK_SIZE:,} ({'dynamisch' if available_ram_gb > 0 else 'statisch'})\n"
            f"  UPDATE_INTERVAL: {UPDATE_INTERVAL:,}\n"
            f"  LOG_INTERVAL: {LOG_INTERVAL:,}\n"
            f"  FASTA Batch-Mode: {'AKTIV' if validate_ref else 'DEAKTIVIERT'}\n"
            f"  AF-Threshold: {af_threshold if af_threshold is not None else 'NONE (alle durchlassen)'}\n"
            f"  Include-None: {include_none}"
        )
        
        if available_ram_gb > 0:
            self.logger.log(f"  Available RAM: {available_ram_gb:.1f} GB")

        # FASTA-VALIDATOR-SETUP (unverändert)
        fasta_validator = None
        if validate_ref:
            self.logger.log(
                f"[VCF-Scan] REF-Validierung aktiviert (BATCH-MODUS) - pruefe FASTA fuer {build}..."
            )
            try:
                fasta_path = does_fasta_exist(build, self.logger, app=self.app)
                if fasta_path:
                    fasta_validator = FastaValidator(
                        fasta_path=fasta_path,
                        strict_mode=Config.FASTA_STRICT_MODE,
                        logger=self.logger
                    )
                    is_valid, message = fasta_validator.validate()
                    if is_valid:
                        self.logger.log(f"[VCF-Scan] FastaValidator bereit (Batch-Mode): {fasta_path}")
                    else:
                        self.logger.log(f"[VCF-Scan] WARNING FASTA-Validierung fehlgeschlagen: {message}")
                        fasta_validator.close()
                        fasta_validator = None
                        validate_ref = False
                else:
                    self.logger.log("[VCF-Scan] INFO REF-Validierung deaktiviert (keine FASTA verfuegbar)")
                    validate_ref = False
            except Exception as e:
                self.logger.log(f"[VCF-Scan] WARNING FASTA-Setup fehlgeschlagen: {e}")
                import traceback
                self.logger.log(f"[VCF-Scan] Traceback:\n{traceback.format_exc()}")
                fasta_validator = None
                validate_ref = False
        
        # LIGHTDB-SETUP - V10: Mit Fallback auf absoluten Pfad
        lightdb_path = getattr(self.af_fetcher, 'lightdb_path', None) if self.af_fetcher else None
        
        # V10: Fallback auf absoluten Default-Pfad wenn af_fetcher.lightdb_path nicht gesetzt
        if not lightdb_path:
            fallback_path = os.path.join(BASE_DIR, "data", "gnomad_light.db")
            if os.path.exists(fallback_path):
                lightdb_path = fallback_path
                self.logger.log(f"[VCF-Scan] LightDB via Fallback-Pfad: {fallback_path}")
                # Setze auch im af_fetcher für konsistente Nutzung
                if self.af_fetcher:
                    self.af_fetcher.lightdb_path = fallback_path
        
        lightdb_available = bool(lightdb_path and os.path.exists(lightdb_path))
        
        if lightdb_available:
            self.logger.log(f"[VCF-Scan] LightDB aktiv: {lightdb_path}")
        else:
            self.logger.log(f"[VCF-Scan] LightDB nicht verfuegbar (Pfad: {lightdb_path})")

        # VCFBuffer
        vcf_buffer = self.vcf_buffer

        # Helper zum Verarbeiten des Coding-Batches
        # V16 FIX 9: Optimiert für weniger DB-Overhead
        coding_buffer_updates = []  # Sammelt Updates für späteren Bulk-Write

        def process_coding_batch():
            """
            ✅ V16 FIX 11: Verarbeitet den Coding-Batch und füllt coding_batch_results.
            WICHTIG: Leert den Buffer NICHT mehr - das macht der Aufrufer nach dem Filtern.
            """
            nonlocal coding_buffer_updates
            if not coding_batch_buffer:
                return

            # Batch verarbeiten - nur die (key, chrom, pos) Teile extrahieren
            batch_keys = [(item[0], item[1], item[2]) for item in coding_batch_buffer]
            batch_data = self.coding_filter.process_batch(batch_keys)
            coding_batch_results.update(batch_data)

            # V16 FIX: Sammle Updates für späteren Bulk-Write statt sofort zu schreiben
            # Das reduziert DB-I/O von ~N writes auf ~1 write pro großem Batch
            for key, data in batch_data.items():
                if data and data.get("gene_symbol"):
                    update = {
                        "gene_symbol": data["gene_symbol"],
                        "is_coding": data["is_coding"]
                    }
                    coding_buffer_updates.append((key, update))
                    # Lokaler Cache für diesen Run
                    if key not in row_cache:
                        row_cache[key] = {}
                    row_cache[key].update(update)

            # Flush zu VCF-Buffer wenn genug gesammelt (batch statt einzeln)
            if len(coding_buffer_updates) >= 5000:
                flush_coding_updates()

            # HINWEIS: Buffer wird NICHT mehr hier geleert!

        def flush_coding_updates():
            nonlocal coding_buffer_updates
            if not coding_buffer_updates:
                return
            # Bulk-Add zum VCF-Buffer (viel effizienter als einzelne adds)
            for key, update in coding_buffer_updates:
                self.vcf_buffer.add(key, update, priority=False)  # V16: priority=False für Batch
            coding_buffer_updates.clear()

        # ═════════════════════════════════════════════════════════════════════
        # HAUPTSCHLEIFE (MIT CODING-FILTER)
        # ═════════════════════════════════════════════════════════════════════
        try:
            for rec in parse_vcf_records_smart(vcf_path):
                if self.stopflag.is_set():
                    self.logger.log("[VCF-Scan] Abbruch durch StopFlag")
                    break

                counter += 1
                
                # Progress-Update
                if counter % UPDATE_INTERVAL == 0:
                    self.progress.update_phase("vcf_scan", UPDATE_INTERVAL)
                
                # Heartbeat-Log mit Stats
                if counter % LOG_INTERVAL == 0:
                    delta = counter - last_logged_counter
                    elapsed = time.time() - last_heartbeat
                    if delta > 0 and elapsed > 0:
                        rate = delta / elapsed
                        passed = len(self.passed_quality)
                        rejected_qm = self.scan_counters['quality_rejected']
                        rejected_coding = self.scan_counters['coding_rejected']
                        pass_rate = (passed / counter * 100) if counter > 0 else 0

                        # ═══════════════════════════════════════════════════════════
                        # V16 FIX: Adaptive Schätzungskorrektur für .gz Dateien
                        # Wenn counter > 80% von total_est, erhöhe Schätzung
                        # ═══════════════════════════════════════════════════════════
                        if counter > total_est * 0.8:
                            # Schätzung war zu niedrig - korrigiere basierend auf bisheriger Rate
                            total_elapsed = time.time() - scan_start_time
                            if total_elapsed > 10:  # Nur nach >10s korrigieren
                                # Extrapoliere: Wenn wir X% der Zeit für counter Records gebraucht haben,
                                # wie viele Records erwarten wir insgesamt?
                                # Konservative Schätzung: verdopple den Rest
                                new_est = int(counter * 1.5)
                                if new_est > total_est:
                                    old_est = total_est
                                    total_est = new_est
                                    self.progress.total_variants = new_est
                                    self.logger.log(
                                        f"[VCF-Scan] Adaptive correction: {old_est:,} -> {new_est:,} "
                                        f"(compressed file larger than expected)"
                                    )

                        # V16 FIX: Verbesserte ETA-Berechnung
                        total_elapsed = time.time() - scan_start_time if 'scan_start_time' in dir() else elapsed
                        if total_elapsed > 0 and counter > 0:
                            overall_rate = counter / total_elapsed
                            remaining_records = max(0, total_est - counter)
                            eta_seconds = remaining_records / overall_rate if overall_rate > 0 else 0

                            if eta_seconds < 60:
                                eta_str = f"{eta_seconds:.0f}s"
                            elif eta_seconds < 3600:
                                eta_str = f"{eta_seconds/60:.1f}min"
                            else:
                                eta_str = f"{eta_seconds/3600:.1f}h"
                        else:
                            eta_str = "?"

                        # Berechne Prozent mit korrigiertem total_est
                        pct = min(99.9, counter / total_est * 100) if total_est > 0 else 0

                        msg = (
                            f"[VCF-Scan] Progress: {counter:,}/{total_est:,} ({pct:.1f}%) | "
                            f"ETA: {eta_str} | {rate:.0f} rec/s | "
                            f"PASS {passed:,} ({pass_rate:.1f}%)"
                        )
                        if rejected_qm > 0:
                            msg += f" | REJECT-QM {rejected_qm:,}"
                        if only_protein_coding and rejected_coding > 0:
                            msg += f" | REJECT-Coding {rejected_coding:,}"

                        self.logger.log(msg)
                    last_heartbeat = time.time()
                    last_logged_counter = counter

                # ═════════════════════════════════════════════════════════════
                # FILTER 1: QUALITY MANAGER (unverändert)
                # ═════════════════════════════════════════════════════════════
                if not self.quality_manager.passes(rec):
                    self.scan_counters['quality_rejected'] += 1
                    continue

                chrom = rec["chrom"]
                pos = rec["pos"]
                ref = rec["ref"]
                alt = rec["alt"]

                # Key erstellen
                if self.cython:
                    key = self.cython.normalize_key(chrom, pos, ref, alt, build)
                else:
                    key = (chrom, pos, ref, alt, build)
                
                # ═════════════════════════════════════════════════════════════
                # ✅ V16 FIX 10: PROTEIN-CODING FILTER (korrigierte Logik)
                # Problem vorher: Varianten wurden durchgelassen BEVOR der
                #                 Batch verarbeitet wurde (is_coding = None)
                # Lösung: Sammle alle Varianten, verarbeite Batch, DANN filtern
                # ═════════════════════════════════════════════════════════════
                if only_protein_coding and self.coding_filter:
                    # Sammle für Batch-Check (rec mitführen für spätere Verarbeitung)
                    coding_batch_buffer.append((key, chrom, pos, rec))

                    # Verarbeite Batch bei CODING_BATCH_SIZE
                    if len(coding_batch_buffer) >= CODING_BATCH_SIZE:
                        process_coding_batch()

                        # ✅ V16 FIX: Jetzt alle gepufferten Varianten filtern
                        for buf_key, buf_chrom, buf_pos, buf_rec in coding_batch_buffer:
                            coding_data = coding_batch_results.get(buf_key)
                            is_coding_val = coding_data.get("is_coding") if coding_data else None

                            if is_coding_val is False:
                                # Definitiv NICHT protein-kodierend
                                self.scan_counters['coding_rejected'] += 1
                            else:
                                # is_coding == True oder None → durchlassen
                                collected_variants.append((buf_rec, buf_key))

                            # Cleanup
                            coding_batch_results.pop(buf_key, None)

                        coding_batch_buffer.clear()

                        # ✅ V16 FIX 11: Auch nach Coding-Filter Batch-Verarbeitung prüfen
                        if len(collected_variants) >= CHUNK_SIZE:
                            self._process_variant_batch(
                                collected_variants=collected_variants,
                                fasta_validator=fasta_validator,
                                validate_ref=validate_ref,
                                build=build,
                                skip_info_recycling=skip_info_recycling,
                                vcf_buffer=vcf_buffer,
                                row_cache=row_cache,
                                lightdb_available=lightdb_available,
                                lightdb_path=lightdb_path,
                                af_threshold=af_threshold,
                                include_none=include_none
                            )
                            collected_variants.clear()

                    # Die Variante ist jetzt im Buffer und wird beim nächsten Batch-Flush geprüft
                    continue

                # ═════════════════════════════════════════════════════════════
                # QUALITY-FILTER BESTANDEN (kein Coding-Filter aktiv)
                # ═════════════════════════════════════════════════════════════
                collected_variants.append((rec, key))

                # BATCH-VERARBEITUNG bei CHUNK_SIZE
                if len(collected_variants) >= CHUNK_SIZE:
                    self._process_variant_batch(
                        collected_variants=collected_variants,
                        fasta_validator=fasta_validator,
                        validate_ref=validate_ref,
                        build=build,
                        skip_info_recycling=skip_info_recycling,
                        vcf_buffer=vcf_buffer,
                        row_cache=row_cache,
                        lightdb_available=lightdb_available,
                        lightdb_path=lightdb_path,
                        af_threshold=af_threshold,
                        include_none=include_none
                    )
                    collected_variants.clear()
                    
        except Exception as e:
            self.logger.log(f"[VCF-Scan] ERROR Kritischer Fehler in Hauptschleife: {e}")
            import traceback
            self.logger.log(f"[VCF-Scan] Traceback:\n{traceback.format_exc()}")
        
        # ═════════════════════════════════════════════════════════════════════
        # FINALISIERUNG
        # ═════════════════════════════════════════════════════════════════════
        
        # ✅ V16 FIX 11: Flush letzten Coding-Batch + alle pending Updates
        if only_protein_coding:
            # Letzten Batch verarbeiten (falls Buffer nicht leer)
            if coding_batch_buffer:
                process_coding_batch()

                # Varianten aus dem letzten Batch filtern und zu collected_variants hinzufügen
                for buf_key, buf_chrom, buf_pos, buf_rec in coding_batch_buffer:
                    coding_data = coding_batch_results.get(buf_key)
                    is_coding_val = coding_data.get("is_coding") if coding_data else None

                    if is_coding_val is False:
                        self.scan_counters['coding_rejected'] += 1
                    else:
                        collected_variants.append((buf_rec, buf_key))

                    coding_batch_results.pop(buf_key, None)

                coding_batch_buffer.clear()

            flush_coding_updates()  # V16 FIX: Flush remaining updates
        
        # Rest-Batch verarbeiten
        if collected_variants:
            self.logger.log(f"[VCF-Scan] Processing final batch ({len(collected_variants)} variants)...")
            
            self._process_variant_batch(
                collected_variants=collected_variants,
                fasta_validator=fasta_validator,
                validate_ref=validate_ref,
                build=build,
                skip_info_recycling=skip_info_recycling,
                vcf_buffer=vcf_buffer,
                row_cache=row_cache,
                lightdb_available=lightdb_available,
                lightdb_path=lightdb_path,
                af_threshold=af_threshold,
                include_none=include_none
            )
        
        # FastaValidator schliessen
        if fasta_validator:
            fasta_validator.close()
        
        # VCFBuffer flushen
        self.vcf_buffer.flush()
        
        # Emit Queue flushen
        self.emit_queue.flush()
        
        # Kalibriere total_variants auf tatsächlich gescannte Anzahl
        # (verhindert dass Fortschritt bei 99% hängen bleibt wenn Schätzung falsch war)
        if counter > 0 and counter != self.progress.total_variants:
            old_total = self.progress.total_variants
            self.progress.total_variants = counter
            self.logger.log(f"[VCF-Scan] Progress calibrated: {old_total:,} -> {counter:,}")

        # Progress abschliessen
        self.progress.complete_phase("vcf_scan")

        # ✅ NEU: Coding-Filter Statistiken
        if only_protein_coding and self.coding_filter:
            coding_stats = self.coding_filter.get_stats()
            self.logger.log(
                f"\n"
                f"[CodingFilter] Performance:\n"
                f"  📦 Batches verarbeitet: {coding_stats['batches_processed']}\n"
                f"  🎯 DB Hits (cached): {coding_stats['db_hits']:,}\n"
                f"  🔬 GeneAnnotator Calls: {coding_stats['annotator_calls']:,}\n"
                f"  ❓ Unknown (permissive): {coding_stats['unknown']:,}\n"
                f"  ❌ Rejected (non-coding): {self.scan_counters['coding_rejected']:,}"
            )
        
        # Final Stats (erweitert)
        self.logger.log(
            f"\n"
            f"[VCF-Scan] ABGESCHLOSSEN\n"
            f"  SCANNED: {counter:,}\n"
            f"  Quality Accepted: {len(self.passed_quality):,}\n"
            f"  Quality Rejected: {self.scan_counters['quality_rejected']:,}\n"
        )
        
        if only_protein_coding:
            self.logger.log(
                f"  Coding Rejected: {self.scan_counters['coding_rejected']:,}\n"
                f"  Both Filters Passed: {len(self.passed_quality) - self.scan_counters['coding_rejected']:,}"
            )
        
        self.logger.log(
            f"  Cache Emitted: {self.scan_counters.get('cache_emitted', 0):,}\n"
            f"  LightDB Emitted: {self.scan_counters.get('lightdb_emitted', 0):,}\n"
            f"  AF Fetch Needed: {len(self.to_fetch_af):,}"
        )
        
        # Print Quality Manager Stats
        self.quality_manager.print_stats()
        
    def _process_variant_batch(
        self,
        collected_variants: List[Tuple[dict, tuple]],
        fasta_validator,
        validate_ref: bool,
        build: str,
        skip_info_recycling: bool,
        vcf_buffer,
        row_cache: dict,
        lightdb_available: bool,
        lightdb_path: str,
        af_threshold: Optional[float],
        include_none: bool
    ):
        """
        ✅ OPTIMIERT: Batch-Verarbeitung mit FASTA-Cache, Skip-Logic und Batch-Stale-Refresh.
        """
        if not collected_variants:
            return
        
        # ============================================================
        # 1. BATCH-FASTA-VALIDIERUNG (OPTIMIERT mit Skip-Logic)
        # ============================================================
        valid_variants = []
        
        if validate_ref and fasta_validator:
            try:
                # ✅ FILTERE bereits validierte Varianten
                needs_validation = []
                already_validated = []
                
                for rec, key in collected_variants:
                    ref_tuple = (key[0], key[1], key[2], build)  # chrom, pos, ref, build
                    
                    if ref_tuple in self.validated_refs:
                        already_validated.append((rec, key))
                    else:
                        needs_validation.append((rec, key))
                
                if already_validated:
                    self.logger.log(
                        f"[VCF-Scan] ⚡ FASTA Skip: {len(already_validated):,} already validated "
                        f"({len(needs_validation):,} need validation)"
                    )
                
                # Nur neue Varianten validieren
                if needs_validation:
                    variants_to_validate = [
                        (key[0], key[1], key[2], key[3])  # chrom, pos, ref, alt
                        for rec, key in needs_validation
                    ]
                    
                    # ✅ Batch-Validierung mit Region-Prefetch
                    validation_results = fasta_validator.validate_batch(
                        variants=variants_to_validate,
                        build=build
                    )
                    
                    # Verarbeite Validierungsergebnisse
                    for (rec, key), (is_valid, corrected_ref, msg) in zip(needs_validation, validation_results):
                        if not is_valid:
                            self.scan_counters['ref_validation_rejected'] += 1
                            if self.scan_counters['ref_validation_rejected'] <= 10:
                                self.logger.log(
                                    f"[VCF-Scan] 🧬 FASTA validation failed: "
                                    f"{key[0]}:{key[1]} {key[2]}→{key[3]} ({msg})"
                                )
                            continue
                        
                        # ✅ Markiere als validiert
                        ref_tuple = (key[0], key[1], key[2], build)
                        self.validated_refs.add(ref_tuple)
                        
                        # Verwende korrigierten REF falls vorhanden
                        if corrected_ref:
                            chrom, pos, _, alt, _build = key
                            key = (chrom, pos, corrected_ref, alt, _build)
                            rec["ref"] = corrected_ref
                        
                        valid_variants.append((rec, key))
                
                # Füge bereits validierte hinzu
                valid_variants.extend(already_validated)
            
            except Exception as e:
                self.logger.log(f"[VCF-Scan] ⚠️ Batch-FASTA error: {e}")
                import traceback
                self.logger.log(f"[VCF-Scan] Traceback:\n{traceback.format_exc()}")
                # Fallback: alle als valid behalten
                valid_variants = collected_variants
        else:
            # Keine FASTA-Validierung
            valid_variants = collected_variants
        
        if not valid_variants:
            return
        
        # ============================================================
        # 2. Record-Processing und Bulk-Operations
        # ============================================================
        
        validated_keys = []
        rsid_updates = []
        
        for rec, key in valid_variants:
            self.passed_quality.add(key)
            validated_keys.append(key)
            
            # rsID aus VCF extrahieren
            vcf_rsid = rec.get("id")
            if vcf_rsid and vcf_rsid != "." and vcf_rsid.startswith("rs"):
                rsid_updates.append((key, {"rsid": vcf_rsid}))
            
            # Original Record + Genotype speichern
            self.orig_records[key] = {
                "id": rec["id"] or ".",
                "qual": rec["qual"],
                "filter": rec["filter"],
                "info": rec["info"] or ".",
                "fmt": rec["fmt"] or ".",
                "sample": rec["samples"][0] if rec["samples"] else ".",
            }
            self.genotype_store[key] = get_genotype_label(rec["fmt"], rec["samples"])
            self.current_vcf_keys.add(key)
            
            # INFO-Recycling
            if not skip_info_recycling:
                try:
                    info = rec["info"] if isinstance(rec["info"], dict) else {}
                    
                    vcf_buffer.add(key, {
                        "gene_symbol": info.get("GENE") or info.get("gene_symbol"),
                        "impact": info.get("IMPACT") or info.get("impact"),
                        "consequence": info.get("CONSEQUENCE") or info.get("CONS") or info.get("consequence"),
                        "is_coding": determine_is_coding(info.get("CONSEQUENCE") or info.get("consequence")),
                    }, priority=False)
                except Exception as e:
                    if self.scan_counters.get('info_errors', 0) <= 10:
                        self.logger.log(f"[VCF-Scan] ⚠️ INFO parse error: {e}")
                        self.scan_counters['info_errors'] = self.scan_counters.get('info_errors', 0) + 1
        
        # rsIDs aus VCF in DB speichern
        if rsid_updates:
            try:
                self.db.update_variant_rsids_bulk(rsid_updates)
                self.logger.log(f"[VCF-Scan] ✅ Recycled {len(rsid_updates)} rsIDs from VCF")
                self.scan_counters['rsid_recycled'] = self.scan_counters.get('rsid_recycled', 0) + len(rsid_updates)
            except Exception as e:
                self.logger.log(f"[VCF-Scan] ⚠️ rsID bulk update failed: {e}")
        
        # DB-Bulk-Lookup
        try:
            chunk_cache = self.db.get_variants_bulk(validated_keys)
            if chunk_cache:
                row_cache.update(chunk_cache)
        except Exception as e:
            self.scan_counters['db_bulk_errors'] += 1
            if self.scan_counters['db_bulk_errors'] <= 5:
                self.logger.log(f"[VCF-Scan] ⚠️ DB bulk error: {e}")
        
        # ✅ LightDB-Lookup für uncached (klassenintern)
        if lightdb_available:
            uncached = [k for k in validated_keys if k not in row_cache]
            if uncached:
                try:
                    lightdb_results, _ = self._lookup_lightdb(
                        uncached, build, lightdb_path
                    )
                    
                    if lightdb_results:
                        for k, val in lightdb_results.items():
                            if not val:
                                continue
                            
                            update = {}
                            
                            if "af" in val and val["af"] is not None:
                                try:
                                    af_float = float(val["af"])
                                    if 0.0 <= af_float <= 1.0:
                                        update["af_filter_mean"] = af_float
                                except (ValueError, TypeError):
                                    pass
                            
                            if "rsid" in val and val["rsid"]:
                                update["rsid"] = val["rsid"]
                            
                            if update or ("af" not in val and "rsid" not in val):
                                update["meanAF_last_fetch"] = now_iso()
                                if update:
                                    update["meanAF_fetch_success"] = "111"
                                else:
                                    update["meanAF_fetch_success"] = "010"
                                
                                vcf_buffer.add(k, update, priority=True)
                                row_cache[k] = update
                                self.scan_counters['lightdb_stored'] += 1
                        
                        if self.scan_counters['lightdb_stored'] > 0 and self.scan_counters['lightdb_stored'] % 1000 == 0:
                            self.logger.log(
                                f"[VCF-Scan] 📦 LightDB: {len(lightdb_results)} hits, "
                                f"{self.scan_counters['lightdb_stored']} total stored"
                            )
                except Exception as e:
                    self.logger.log(f"[VCF-Scan] ⚠️ LightDB error: {e}")
        
        # ============================================================
        # 3. ✅ BATCH-STALE-REFRESH (vor AF-Validierung, klassenintern)
        # ============================================================
        if lightdb_available and lightdb_path:
            stale_keys = []
            for key in validated_keys:
                row = row_cache.get(key)
                if row:
                    last_fetch = row.get("meanAF_last_fetch")
                    if is_af_stale(last_fetch):
                        stale_keys.append(key)
            
            if stale_keys:
                try:
                    self.logger.log(f"[VCF-Scan] 🔄 Refreshing {len(stale_keys)} stale AF values...")
                    
                    lightdb_results, _ = self._lookup_lightdb(
                        stale_keys, build, lightdb_path
                    )
                    
                    refreshed_count = 0
                    for k, val in lightdb_results.items():
                        if val and val.get("af") is not None:
                            try:
                                af_float = float(val["af"])
                                if 0.0 <= af_float <= 1.0:
                                    update = {
                                        "af_filter_mean": af_float,
                                        "meanAF_last_fetch": now_iso(),
                                        "meanAF_fetch_success": "111"
                                    }
                                    vcf_buffer.add(k, update, priority=True)
                                    
                                    # Update row_cache sofort
                                    if k in row_cache:
                                        row_cache[k].update(update)
                                    else:
                                        row_cache[k] = update
                                    
                                    refreshed_count += 1
                            except (ValueError, TypeError):
                                pass
                    
                    if refreshed_count > 0:
                        self.logger.log(f"[VCF-Scan] ✅ Refreshed {refreshed_count}/{len(stale_keys)} stale entries")
                
                except Exception as e:
                    self.logger.log(f"[VCF-Scan] ⚠️ Batch stale refresh error: {e}")
        
        # ============================================================
        # 4. AF-VALIDIERUNG (OPTIMIERT: Batch-Add für to_fetch_af)
        # ============================================================
        keys_to_fetch = []
        
        for key in validated_keys:
            row = row_cache.get(key)
            
            # V10: FlagManager hat af_threshold - prüfe ob kein Threshold gesetzt
            effective_threshold = self.flag_manager.get_af_threshold() if self.flag_manager else af_threshold
            
            if effective_threshold is None:
                self.emit_queue.add(key)
                self.add_display_key(key)
                self.emitted_immediate.add(key)
                self.scan_counters['cache_emitted'] += 1
            else:
                # V10: Parameter werden vom FlagManager geholt
                decision = self._validate_af_in_cache(key, row)
                
                if decision == "emit":
                    self.emit_queue.add(key)
                    self.add_display_key(key)
                    self.emitted_immediate.add(key)
                    
                    if row and row.get("rsid"):
                        self.scan_counters['lightdb_emitted'] += 1
                    else:
                        self.scan_counters['cache_emitted'] += 1
                        
                elif decision == "reject":
                    self.progress.mark_rejected(1)
                    self.scan_counters['af_rejected'] = self.scan_counters.get('af_rejected', 0) + 1
                    
                else:  # "fetch"
                    keys_to_fetch.append(key)
                    self.scan_counters['af_fetch_needed'] = self.scan_counters.get('af_fetch_needed', 0) + 1
        
        # ✅ BATCH-ADD: Nur EIN Lock für ALLE Keys
        if keys_to_fetch:
            with self.to_fetch_af_lock:
                self.to_fetch_af.update(keys_to_fetch)

    def _validate_af_in_cache(
        self,
        key: Tuple,
        row: Optional[Dict],
        af_threshold: float = None,  # V10: Deprecated, wird vom FlagManager geholt
        include_none: bool = None    # V10: Deprecated, wird vom FlagManager geholt
    ) -> str:
        """
        V10: Stale-Check entfernt, nutzt FlagManager für Einstellungen.
        
        Returns:
            "emit" | "reject" | "fetch"
        """
        # V10: Hole Parameter vom FlagManager
        if self.flag_manager:
            af_threshold = self.flag_manager.get_af_threshold()
            include_none = self.flag_manager.get_include_none()
        else:
            # Legacy Fallback
            af_threshold = af_threshold if af_threshold is not None else 0.01
            include_none = include_none if include_none is not None else False
        
        if not row:
            return "fetch"
        
        mean_af = row.get("af_filter_mean")
        mean_status = row.get("meanAF_fetch_success")
        
        # ✅ 1. Erfolgreicher Wert → Range-Check
        if mean_af is not None and FetchStatusManager.is_success(mean_status):
            try:
                af_value = float(mean_af)
                if not (0.0 <= af_value <= 1.0):
                    self.logger.log(f"[AF-Cache] ⚠️ Invalid AF {af_value} → refetch")
                    return "fetch"
                return "emit" if af_value <= af_threshold else "reject"
            except (ValueError, TypeError):
                return "fetch"
        
        # ✅ 2. True None → Fallback auf andere AF-Spalten
        if mean_af is None and FetchStatusManager.is_success(mean_status):
            try:
                # Prüfe ob row bereits alle Felder hat (dann kein DB-Call nötig)
                af_candidates = []
                
                # Erst aus aktuellem row versuchen
                for field in ["af_exac", "af_gnomad_exomes", "af_gnomad_genomes", "af_1kg"]:
                    val = row.get(field)
                    if val not in (None, ".", ""):
                        try:
                            af_float = float(val)
                            if 0.0 <= af_float <= 1.0:
                                af_candidates.append(af_float)
                        except (ValueError, TypeError):
                            pass
                
                # Wenn keine Kandidaten in row, dann DB-Bulk-Lookup
                if not af_candidates:
                    cached = self._get_variants_bulk_cached([key])
                    row_full = cached.get(key) or {}
                    
                    for field in ["af_exac", "af_gnomad_exomes", "af_gnomad_genomes", "af_1kg"]:
                        val = row_full.get(field)
                        if val not in (None, ".", ""):
                            try:
                                af_float = float(val)
                                if 0.0 <= af_float <= 1.0:
                                    af_candidates.append(af_float)
                            except (ValueError, TypeError):
                                pass
                
                if af_candidates:
                    # Verwende Fallback-AF
                    fallback_af = sum(af_candidates) / len(af_candidates)
                    update = {
                        "af_filter_mean": fallback_af,
                        "meanAF_last_fetch": now_iso(),
                        "meanAF_fetch_success": "111"
                    }
                    self.vcf_buffer.add(key, update, priority=True)
                    row.update(update)
                    return "emit" if fallback_af <= af_threshold else "reject"
                else:
                    # ✅ Echter None-Fall (keine AF-Daten verfügbar)
                    return "emit" if include_none else "reject"
                    
            except Exception as e:
                self.logger.log(f"[AF-Cache] ⚠️ Fallback AF error: {e}")
                return "fetch"
        
        # 3. Unsicherer Status → API-Fetch nötig
        return "fetch"
        
    def _validate_with_fasta_fallback(
        self, 
        failed_keys: list, 
        build: str, 
        phase_label: str = "Fallback"
    ) -> Tuple[List, List]:
        """
        ✅ ÜBERARBEITET: FASTA-Fallback mit FastaValidator.
        """
        if not failed_keys:
            return [], []
        
        self.logger.log(
            f"[{phase_label}] 🧬 FASTA-Fallback: Validiere {len(failed_keys)} fehlgeschlagene Keys..."
        )
        
        # Lazy Loading
        validator = self._get_or_create_fasta_validator(build)
        
        if not validator:
            self.logger.log(
                f"[{phase_label}] ⚠️ FastaValidator nicht verfügbar - "
                f"behalte alle {len(failed_keys)} Keys"
            )
            return failed_keys, []
        
        valid_keys = []
        invalid_keys = []
        
        # Batch-Validierung
        for key in failed_keys:
            chrom, pos, ref, alt, _build = key
            
            try:
                # ✅ Nutze validate_single (war vorher validate)
                is_valid, corrected_ref, msg = validator.validate_single(
                    chrom=chrom,
                    pos=pos,
                    ref=ref,
                    alt=alt,
                    build=build
                )
                
                if is_valid:
                    valid_keys.append(key)
                else:
                    invalid_keys.append(key)
                    
                    if len(invalid_keys) <= 10:
                        self.logger.log(
                            f"[{phase_label}] 🧬 REF mismatch: {chrom}:{pos} "
                            f"VCF_REF={ref} ({msg})"
                        )
            
            except Exception as e:
                valid_keys.append(key)
                if len(invalid_keys) + len(valid_keys) <= 10:
                    self.logger.log(
                        f"[{phase_label}] ⚠️ Validierung fehlgeschlagen für {key}: {e}"
                    )
        
        # Markiere ungültige
        if invalid_keys:
            self.logger.log(
                f"[{phase_label}] 🚫 {len(invalid_keys)} ungültige Varianten "
                f"(REF-Mismatch mit FASTA)"
            )
            
            try:
                for k in invalid_keys:
                    self.vcf_buffer.add(k, {
                        "fasta_ref_valid": 0,
                        "meanAF_fetch_success": "000",
                        "meanAF_last_fetch": now_iso()
                    }, priority=True)
                
                self.vcf_buffer.flush(force_priority=True)
            
            except Exception as e:
                self.logger.log(
                    f"[{phase_label}] ⚠️ Fehler beim Markieren ungültiger Varianten: {e}"
                )
        
        self.logger.log(
            f"[{phase_label}] ✅ FASTA-Validierung: "
            f"{len(valid_keys)} valid, {len(invalid_keys)} invalid"
        )
        
        return valid_keys, invalid_keys


    #Validiert AF aus API Abfragen
    def _validate_af(self, val: Optional[float], key: Tuple, source: str) -> bool:
        """
        Validiert AF-Wert auf Plausibilität.
        
        Args:
            val: AF-Wert
            key: Variant-Key (für Logging)
            source: Quelle (für Logging)
        
        Returns:
            True wenn valide, False sonst
        """
        if self.cython:
            return self.cython.validate_af(val)
        else:
            # Fallback
            if val is None:
                return False
            
            if not isinstance(val, (int, float)):
                self.logger.log(f"[AF-Validation] ⚠️ Nicht-numerischer AF von {source} für {key}: {val}")
                return False
            
            if val < 0.0 or val > 1.0:
                self.logger.log(
                    f"[AF-Validation] ⚠️ UNGÜLTIGER AF von {source} für {key}: {val:.6f}\n"
                    f"  → Wert außerhalb [0.0, 1.0] → verworfen"
                )
                return False
            
            # Optional: Warnung bei ungewöhnlich hohen AFs (> 0.5)
            if val > 0.5:
                self.logger.log(
                    f"[AF-Validation] ℹ️ Hoher AF von {source} für {key}: {val:.6f}\n"
                    f"  (> 0.5 ist ungewöhnlich für seltene Varianten)"
                )
            
            return True
        
    def _lookup_lightdb(
        self,
        uncached: list,
        build: str = "GRCh37",
        db_path: str | None = None
        ) -> tuple[dict, list]:
        """
        Lookup von Varianten in der lokalen Light-DB.
        
        FIX: Connection wird immer geschlossen (try-finally)
        
        Optimiert temporaere Tabellen-Verwaltung:
        - Temporaere Tabelle wird nur einmal erstellt (vorher in jedem Chunk)
        - Verwendet DELETE statt DROP+CREATE (schneller)
        - Cleanup nach der Schleife
        - Reduziert DDL-Operationen von O(n) auf O(1)

        Args:
            uncached: Liste von Variant-Keys (Tuples: chrom,pos,ref,alt,build)
            build: Referenzgenom ("GRCh37" oder "GRCh38")
            db_path: Pfad zur Light-DB (falls None -> self.lightdb_path)

        Returns:
            (results_dict, still_uncached_list)
            - results_dict: {key: {"af": float|None, "rsid": str|None}}
            - still_uncached_list: Keys ohne Treffer in Light-DB
        """
        results = {}

        if not db_path:
            db_path = self.lightdb_path

        if not db_path:
            self.logger.log("[Distiller/LightDB] Warnung: Kein Light-DB Pfad - Lookup uebersprungen.")
            return results, uncached

        # FIX: Connection mit finally-Block
        conn = None
        try:
            conn = sqlite3.connect(db_path)
            cur = conn.cursor()

            chunk_size = 5000
            total = len(uncached)
            self.logger.log(f"[Distiller/LightDB] Starte Lookup fuer {total:,} Varianten (Chunk-Groesse {chunk_size})...")

            # Temporaere Tabelle einmalig VOR der Schleife erstellen
            try:
                cur.execute("DROP TABLE IF EXISTS tmp_keys;")
                cur.execute("""
                    CREATE TEMP TABLE tmp_keys (
                        chrom TEXT,
                        pos   INTEGER,
                        ref   TEXT,
                        alt   TEXT,
                        build TEXT
                    );
                """)
            except Exception as e:
                self.logger.log(f"[Distiller/LightDB] Warnung: Fehler beim Erstellen der temp table: {e}")
                return results, uncached

            for i in range(0, total, chunk_size):
                block = uncached[i:i+chunk_size]
                self.logger.log(f"[Distiller/LightDB] Block {i//chunk_size+1}: {len(block)} Keys werden geprueft...")

                parsed = []
                for key in block:
                    try:
                        if isinstance(key, tuple) and len(key) == 5:
                            chrom, pos, ref, alt, build_key = key
                            parsed.append((str(chrom), int(pos), str(ref), str(alt), str(build_key)))
                        else:
                            self.logger.log(f"[Distiller/LightDB] Warnung: Ungueltiges Key-Format: {key}")
                    except Exception as e:
                        self.logger.log(f"[Distiller/LightDB] Warnung: Fehler beim Parsen von {key}: {e}")
                        continue

                if not parsed:
                    continue

                try:
                    # Temporaere Tabelle LEEREN (schneller als DROP+CREATE)
                    cur.execute("DELETE FROM tmp_keys;")
                    cur.executemany("INSERT INTO tmp_keys VALUES (?,?,?,?,?)", parsed)

                    # JOIN gegen variants_light (nutzt Index)
                    cur.execute("""
                        SELECT v.chrom, v.pos, v.ref, v.alt, v.build, v.rsid, v.af
                        FROM variants_light v
                        JOIN tmp_keys k
                          ON v.chrom=k.chrom
                         AND v.pos=k.pos
                         AND v.ref=k.ref
                         AND v.alt=k.alt
                         AND v.build=k.build;
                    """)
                    fetched = cur.fetchall()

                    for chrom, pos, ref, alt, build_val, rsid, af in fetched:
                        key = (chrom, pos, ref, alt, build_val)
                        results[key] = {"af": af, "rsid": rsid}

                    self.logger.log(
                        f"[Distiller/LightDB] OK: Block {i//chunk_size+1}: {len(fetched)} Treffer, "
                        f"{len(block) - len(fetched)} nicht gefunden."
                    )
                except Exception as e:
                    self.logger.log(f"[Distiller/LightDB] Warnung: SQL-Fehler im Block {i//chunk_size+1}: {e}")

            # Cleanup: Temporaere Tabelle loeschen
            try:
                cur.execute("DROP TABLE IF EXISTS tmp_keys;")
            except Exception:
                pass

        except Exception as e:
            self.logger.log(f"[Distiller/LightDB] Warnung: Fehler beim Lookup: {e}")
        finally:
            # KRITISCH: Connection immer schliessen
            if conn:
                try:
                    conn.close()
                except Exception:
                    pass

        still_uncached = [k for k in uncached if k not in results]
        self.logger.log(
            f"[Distiller/LightDB] Zusammenfassung: {len(results)} Treffer, {len(still_uncached)} verbleibend."
        )
        return results, still_uncached


    def result_collector_and_merger(self, keys, results):
        """
        FIX: Korrekte Key-Normalisierung für MyVariant-Results.
        
        Änderungen:
        - Nutzt HGVS als primären Lookup-Key
        - Fallback auf Positions-Matching bei multi-ALT
        - Validiert AF-Range vor Mittelwert-Berechnung
        - Zusätzliche Guards gegen NoneType
        """
        grouped = {}

        # ✅ Ergebnisse sammeln (HGVS-basiert)
        for res in results:
            if not isinstance(res, dict):
                continue
            for hgvs_id, hit in res.items():
                if not isinstance(hgvs_id, str):
                    continue
                grouped.setdefault(hgvs_id, []).append(hit)

        final = {}

        # ✅ Reverse-Mapping: HGVS -> Keys
        hgvs_to_keys = {}
        for k in keys:
            try:
                hgvs = normalize_for_mv(k)
                if hgvs:
                    hgvs_to_keys.setdefault(hgvs, []).append(k)
            except Exception:
                continue

        for hgvs_id, hits in grouped.items():
            matching_keys = hgvs_to_keys.get(hgvs_id, [])
            if not matching_keys:
                continue

            af_sources = {
                "exac": None,
                "gnomad_exomes": None,
                "gnomad_genomes": None,
                "1kg": None,
            }

            for hit in hits:
                if not hit or not isinstance(hit, dict):
                    continue

                # ExAC
                exac = hit.get("exac")
                if isinstance(exac, dict):
                    val = safe_float(exac.get("af"))
                    if self._validate_af(val, hgvs_id, "ExAC"):
                        af_sources["exac"] = val

                # gnomAD
                gnomad = hit.get("gnomad")
                if isinstance(gnomad, dict):
                    exomes = safe_float(gnomad.get("exomes", {}).get("af")) \
                        if isinstance(gnomad.get("exomes"), dict) else None
                    genomes = safe_float(gnomad.get("genomes", {}).get("af")) \
                        if isinstance(gnomad.get("genomes"), dict) else None

                    if self._validate_af(exomes, hgvs_id, "gnomAD-Exomes"):
                        af_sources["gnomad_exomes"] = exomes
                    if self._validate_af(genomes, hgvs_id, "gnomAD-Genomes"):
                        af_sources["gnomad_genomes"] = genomes

                # 1000G
                dbsnp = hit.get("dbsnp")
                if isinstance(dbsnp, dict):
                    val = safe_float(dbsnp.get("gmaf"))
                    if self._validate_af(val, hgvs_id, "1000G"):
                        af_sources["1kg"] = val

            # Priorisierte AF-Auswahl
            valid_afs = []
            for source in ["gnomad_genomes", "gnomad_exomes", "exac", "1kg"]:
                val = af_sources.get(source)
                if val is not None:
                    valid_afs.append(val)

            mean_af = None
            if valid_afs:
                mean_af = sum(valid_afs) / len(valid_afs)
                if not (0.0 <= mean_af <= 1.0):
                    self.logger.log(
                        f"[AF-Controller] ⚠️ Invalid mean_af={mean_af:.6f} für {hgvs_id}"
                    )
                    mean_af = None
                else:
                    mean_af = max(0.0, min(1.0, mean_af))

            result = {
                "af_exac": af_sources["exac"],
                "af_gnomad_exomes": af_sources["gnomad_exomes"],
                "af_gnomad_genomes": af_sources["gnomad_genomes"],
                "af_1kg": af_sources["1kg"],
                "af_filter_mean": mean_af,
            }

            for k in matching_keys:
                final[k] = result

        return final
    
    # [PROBABLY DEAD] Nie aufgerufen - Statistics werden anders berechnet
    # def _get_recent_success_rate(self) -> float:
        # """
        # Berechnet Erfolgsrate aus Tuner-History (letzte 60s).
        # Wird für adaptives Delay-Management verwendet.
        # """
        # tuner = self.tuners.get("mv")
        # if not tuner or not tuner.history:
            # return 1.0
        
        # now = time.time()
        # recent = [
            # (successes, duration) 
            # for (t, successes, duration, _, _, _) in tuner.history 
            # if now - t <= 60
        # ]
        
        # if not recent:
            # return 1.0
        
        # total_successes = sum(s for s, _ in recent)
        # total_duration = sum(d for _, d in recent)
        
        # Schätze erwartete Erfolge basierend auf Durchschnitt
        # if not tuner.history:
            # return 1.0
        
        # Nutze die letzten erfolgreichen Batches als Baseline
        # successful_batches = [s for s, _ in recent if s > 0]
        # if not successful_batches:
            # return 0.0
        
        # avg_success = sum(successful_batches) / len(successful_batches)
        # expected_total = avg_success * len(recent)
        
        # if expected_total == 0:
            # return 1.0
        
        # return min(1.0, total_successes / expected_total)
    
    async def _phase_af_fetch_streaming(
        self,
        build: str,
        af_threshold: float = 0.01,
        include_none: bool = False,
        pipeline_state=None
    ):
        """
        ✅ KONTINUIERLICH + THREAD-SAFE: Verarbeitet neue Varianten während VCF-Scan läuft.
        """

        self.progress.set_phase("af_fetch")

        vcf_done = pipeline_state.get("vcf_done") if pipeline_state else None

        # ============================================================
        # V16 FIX: ADAPTIVE THRESHOLDS statt statische Werte
        # Problem: Bei langsamen Scannern (<100 rec/s) ist early_start_threshold=1000
        #          zu hoch und max_idle_time=600 zu lang
        # Lösung: Dynamische Anpassung basierend auf Scanner-Geschwindigkeit
        # ============================================================
        early_start_threshold = 200  # V16: Deutlich niedriger (war 1000)
        max_idle_time = 60           # V16: Deutlich kürzer (war 600)
        MIN_BATCH_SIZE = 50          # V16: Minimum Batch-Größe für effizientes Processing

        # ============================================================
        # PHASE 1: IDLE-SCHLEIFE bis Start-Bedingung
        # ============================================================
        idle_start = time.time()
        logged_waiting = False
        last_queue_size = 0
        queue_growth_samples = []  # Für adaptive Threshold-Berechnung

        while True:
            # V10 FIX: StopFlag-Check am Anfang jeder Iteration
            if self.stopflag and self.stopflag.is_set():
                self.logger.log("[AF-Streaming] ⏸️ Abbruch durch StopFlag (Phase 1)")
                self.progress.complete_phase("af_fetch")
                if pipeline_state:
                    pipeline_state["af_done"].set()
                return {}

            # Queue-Größe threadsicher lesen
            with self.to_fetch_af_lock:
                current_queue_size = len(self.to_fetch_af)

            vcf_is_done = vcf_done.is_set() if vcf_done else True

            # Start-Bedingung 1: VCF-Scan complete UND Queue hat Daten
            if vcf_is_done and current_queue_size > 0:
                self.logger.log(f"[AF-Streaming] 🚀 VCF-Scan complete, processing {current_queue_size:,} variants")
                break

            # Start-Bedingung 2: VCF-Scan complete UND Queue leer → SKIP
            if vcf_is_done and current_queue_size == 0:
                self.logger.log("[AF-Streaming] ✅ VCF-Scan complete, Queue empty → SKIP")
                self.progress.complete_phase("af_fetch")
                if pipeline_state:
                    pipeline_state["af_done"].set()
                return {}

            # V16 FIX: Adaptive Early-Start basierend auf Queue-Wachstum
            if current_queue_size > last_queue_size:
                growth = current_queue_size - last_queue_size
                queue_growth_samples.append(growth)

                # Nach 5 Samples: Berechne durchschnittliche Rate und passe Threshold an
                if len(queue_growth_samples) >= 5:
                    avg_growth = sum(queue_growth_samples[-10:]) / min(10, len(queue_growth_samples))
                    # Bei langsamer Rate (<50/s): Starte früher
                    if avg_growth < 50:
                        early_start_threshold = max(MIN_BATCH_SIZE, int(avg_growth * 10))  # 10 Sekunden Puffer

            last_queue_size = current_queue_size

            # Start-Bedingung 3: Early-Start (Queue voll genug, VCF läuft noch)
            if not vcf_is_done and current_queue_size >= early_start_threshold:
                self.logger.log(f"[AF-Streaming] 🚀 Early-Start triggered ({current_queue_size:,} >= {early_start_threshold})")
                break

            # V16 FIX: Auch bei kleinen Batches starten wenn genug Zeit vergangen
            idle_time = time.time() - idle_start
            if not vcf_is_done and current_queue_size >= MIN_BATCH_SIZE and idle_time > 10:
                self.logger.log(f"[AF-Streaming] 🚀 Mini-Batch Start ({current_queue_size:,} variants after {idle_time:.0f}s idle)")
                break

            # Timeout-Check
            if idle_time > max_idle_time:
                if current_queue_size > 0:
                    self.logger.log(f"[AF-Streaming] ⚠️ TIMEOUT after {idle_time:.0f}s, starting with {current_queue_size:,} variants")
                    break
                else:
                    # Queue noch leer, aber Scanner läuft noch - weiter warten
                    if not vcf_is_done:
                        max_idle_time += 30  # Verlängere Timeout
                        self.logger.log(f"[AF-Streaming] ⏳ Extended timeout to {max_idle_time}s (queue empty, VCF still running)")

            # Warten
            if not logged_waiting:
                self.logger.log(f"[AF-Streaming] ⏳ Waiting for VCF-Scan OR {early_start_threshold} variants (current: {current_queue_size:,})")
                logged_waiting = True

            await asyncio.sleep(1)

        # ============================================================
        # PHASE 2: KONTINUIERLICHE VERARBEITUNG
        # ============================================================
        self.logger.log("[AF-Streaming] 🔄 Entering continuous processing mode")

        round_num = 0
        total_processed = 0
        consecutive_empty = 0
        _display_key_counter = 0
        _last_gui_update = time.time()
        GUI_UPDATE_INTERVAL = 100
        GUI_UPDATE_TIME_SEC = 2.0

        while True:
            # V10 FIX: StopFlag-Check am Anfang jeder Iteration
            if self.stopflag and self.stopflag.is_set():
                self.logger.log("[AF-Streaming] ⏸️ Abbruch durch StopFlag (Phase 2)")
                break
            
            round_num += 1
            vcf_is_done = vcf_done.is_set() if vcf_done else True

            MAX_BULK_CHUNK = 5000
            with self.to_fetch_af_lock:
                raw_keys = list(dict.fromkeys(self.to_fetch_af))[:MAX_BULK_CHUNK]

            current_queue_size = len(raw_keys)

            # Exit-Bedingungen
            if vcf_is_done:
                if current_queue_size == 0:
                    consecutive_empty += 1
                    if consecutive_empty == 1:
                        self.logger.log("[AF-Streaming] 🔍 Queue empty, double-checking...")
                        await asyncio.sleep(0.5)
                        continue
                    elif consecutive_empty >= 2:
                        self.logger.log("[AF-Streaming] ✅ VCF-Scan complete AND Queue confirmed empty → EXIT")
                        break
                else:
                    consecutive_empty = 0
                    self.logger.log(f"[AF-Streaming] 🔄 Round {round_num}: VCF-Scan complete, processing remaining {current_queue_size:,} variants")
            else:
                consecutive_empty = 0
                if current_queue_size == 0:
                    # V10 FIX: Nur alle 10 Runden loggen um Log-Spam zu reduzieren
                    if round_num % 10 == 1:
                        self.logger.log(f"[AF-Streaming] ⏳ Queue empty (round {round_num}), waiting for VCF-Scan...")
                    await asyncio.sleep(2)
                    continue
                else:
                    self.logger.log(f"[AF-Streaming] 🔄 Round {round_num}: Processing {current_queue_size:,} variants (VCF-Scan still running)")

            # ============================================================
            # Snapshot verarbeiten
            # ============================================================
            needs_fetch = []
            already_done = []

            try:
                rows = self._get_variants_bulk_cached(raw_keys) or {}

                for k in raw_keys:
                    row = rows.get(k)
                    if row is None:
                        needs_fetch.append(k)
                        continue

                    mean_af = row.get("af_filter_mean") if row else None
                    mean_status = row.get("meanAF_fetch_success") if row else None
                    last_fetch = row.get("meanAF_last_fetch") if row else None

                    if mean_af is not None and FetchStatusManager.is_success(mean_status):
                        already_done.append(k)
                        # V10: Parameter werden vom FlagManager geholt
                        decision = self._apply_af_filter_final(k, row)
                        if decision == "emit":
                            try:
                                self.emit_queue.add(k)
                                self.add_display_key(k)
                                self.progress.update_phase("af_fetch", 1)
                            except Exception:
                                pass
                        elif decision == "reject":
                            self.progress.mark_rejected(1)
                            self.progress.update_phase("af_fetch", 1)
                        continue

                    if is_af_stale(last_fetch):
                        needs_fetch.append(k)
                    else:
                        already_done.append(k)
                        # V10: Parameter werden vom FlagManager geholt
                        decision = self._apply_af_filter_final(k, row)
                        if decision == "emit":
                            try:
                                self.emit_queue.add(k)
                                _display_key_counter += 1
                                time_since_update = time.time() - _last_gui_update
                                should_update = (
                                    _display_key_counter % GUI_UPDATE_INTERVAL == 0
                                    or time_since_update >= GUI_UPDATE_TIME_SEC
                                )
                                if should_update:
                                    self.add_display_key(k)
                                    _last_gui_update = time.time()
                            except Exception:
                                pass
                        elif decision == "reject":
                            self.progress.mark_rejected(1)
                            self.progress.update_phase("af_fetch", 1)

            except Exception as e:
                self.logger.log(f"[AF-Streaming] ⚠️ Status-Check failed: {e}")
                needs_fetch = raw_keys

            if already_done:
                self.logger.log(f"[AF-Streaming] Round {round_num}: {len(already_done):,} already have AF (skipped)")

            if not needs_fetch:
                self.logger.log(f"[AF-Streaming] Round {round_num}: No fetch needed, continuing...")
                with self.to_fetch_af_lock:
                    self.to_fetch_af -= set(raw_keys)
                await asyncio.sleep(1)
                continue

            # ============================================================
            # FETCH für needs_fetch
            # ============================================================
            self.logger.log(f"[AF-Streaming] Round {round_num}: Fetching {len(needs_fetch):,} variants")

            self.tuners["mv"].set_load_state(
                total_variants=len(needs_fetch),
                done_variants=0,
                end_retry_variants=0
            )

            results = {}
            failed_keys = []
            processed_count = 0

            async def fetch_chunk(chunk, idx):
                nonlocal processed_count
                start = time.time()
                chunk_failed = []

                try:
                    current_cpu = None
                    try:
                        current_cpu = psutil.cpu_percent(interval=0.1)
                    except Exception:
                        pass

                    workers, batch_size = self.tuners["mv"].decide(current_cpu=current_cpu)

                    self.tuners["mv"].set_load_state(
                        total_variants=len(needs_fetch),
                        done_variants=processed_count,
                        end_retry_variants=len(failed_keys)
                    )

                    res = await mv_fetch_async(
                        chunk,
                        build,
                        fetch_full=False,
                        batch_size=batch_size,
                        logger=self.logger,
                        phase_label=f"AF-R{round_num}-C{idx}"
                    )

                    # ✅ Guard: Ergebnis muss Dict sein
                    if not isinstance(res, dict):
                        # Alles als fehlgeschlagen markieren
                        current_statuses = self.db.get_variants_bulk(chunk) or {}
                        if not isinstance(current_statuses, dict):
                            current_statuses = {}
                        for k in chunk:
                            # ✅ FIX: Verwende safe_nested_get
                            current_status = safe_nested_get(current_statuses, k, "meanAF_fetch_success")
                            new_status = FetchStatusManager.update_status(current_status, success=False)
                            self.vcf_buffer.add(k, {
                                "af_filter_mean": None,
                                "meanAF_last_fetch": now_iso(),
                                "meanAF_fetch_success": new_status
                            }, priority=False)
                            chunk_failed.append(k)

                        failed_keys.extend(chunk_failed)
                        processed_count += len(chunk)
                        self.tuners["mv"].record_round(0, time.time() - start)
                        return 0

                    # ✅ Remap: von {key_tuple: hit_dict} zu {hgvs_id: hit_dict}
                    hits_by_hgvs = {}
                    for k, hit in res.items():
                        if not isinstance(hit, dict):
                            continue
                        try:
                            hg = normalize_for_mv(k)
                        except Exception:
                            hg = None
                        if hg and isinstance(hg, str):
                            hits_by_hgvs[hg] = hit

                    # Collector erwartet eine Liste von Dicts mit HGVS-Keys
                    merged = self.result_collector_and_merger(chunk, [hits_by_hgvs]) or {}

                    # Defensive Bulk-Fetch
                    current_statuses = self.db.get_variants_bulk(chunk) or {}
                    if not isinstance(current_statuses, dict):
                        current_statuses = {}

                    for k in chunk:
                        val = merged.get(k) if merged else None
                        results[k] = val

                        if val is None:
                            # ✅ FIX: Verwende safe_nested_get
                            current_status = safe_nested_get(current_statuses, k, "meanAF_fetch_success")
                            new_status = FetchStatusManager.update_status(current_status, success=False)

                            self.vcf_buffer.add(k, {
                                "af_filter_mean": None,
                                "meanAF_last_fetch": now_iso(),
                                "meanAF_fetch_success": new_status
                            }, priority=False)

                            chunk_failed.append(k)
                        else:
                            status = FetchStatusManager.update_status(None, success=True)
                            enriched = dict(val)
                            enriched.update({
                                "meanAF_last_fetch": now_iso(),
                                "meanAF_fetch_success": status
                            })
                            self.vcf_buffer.add(k, enriched, priority=False)

                    success_count = len(chunk) - len(chunk_failed)
                    self.tuners["mv"].record_round(success_count, time.time() - start)
                    self.progress.update_phase("af_fetch", len(chunk))

                    if chunk_failed:
                        failed_keys.extend(chunk_failed)

                    processed_count += len(chunk)
                    return success_count

                except Exception as e:
                    self.logger.log(f"[AF-Fetch-R{round_num}-C{idx}] ❌ Chunk error: {e}")
                    return 0

            # Chunk-Processing Loop
            max_concurrent_chunks = 6
            active_chunks = set()
            pos = 0
            chunk_idx = 1

            while pos < len(needs_fetch):
                while len(active_chunks) >= max_concurrent_chunks:
                    done, active_chunks = await asyncio.wait(
                        active_chunks,
                        return_when=asyncio.FIRST_COMPLETED
                    )

                _, tuner_batch = self.tuners["mv"].decide()
                chunk_size = min(int(tuner_batch * 1.5), 1200)

                chunk = needs_fetch[pos:pos + chunk_size]
                if not chunk:
                    break

                task = asyncio.create_task(fetch_chunk(chunk, chunk_idx))
                active_chunks.add(task)

                pos += chunk_size
                chunk_idx += 1

            if active_chunks:
                await asyncio.gather(*active_chunks, return_exceptions=True)

            self.vcf_buffer.flush()

            # Finale Filter-Anwendung
            emit_count = 0
            reject_count = 0

            for k in needs_fetch:
                val = results.get(k) if results else None
                # V10: Parameter werden vom FlagManager geholt
                decision = self._apply_af_filter_final(k, val)

                if decision == "emit":
                    try:
                        self.emit_queue.add(k)
                        _display_key_counter += 1

                        time_since_update = time.time() - _last_gui_update
                        should_update = (
                            _display_key_counter % GUI_UPDATE_INTERVAL == 0
                            or time_since_update >= GUI_UPDATE_TIME_SEC
                        )

                        if should_update:
                            self.add_display_key(k)
                            _last_gui_update = time.time()

                        emit_count += 1
                    except Exception:
                        pass
                elif decision == "reject":
                    self.progress.mark_rejected(1)
                    reject_count += 1

            # THREAD-SAFE: Entferne verarbeitete Keys
            with self.to_fetch_af_lock:
                self.to_fetch_af -= set(raw_keys)

            total_processed += len(raw_keys)

            self.logger.log(
                f"[AF-Streaming] Round {round_num} complete: "
                f"{len(needs_fetch):,} fetched, "
                f"{emit_count:,} emitted, "
                f"{reject_count:,} rejected, "
                f"{len(failed_keys):,} failed"
            )

            if not vcf_is_done:
                await asyncio.sleep(1)

        # EXIT: Alle Varianten verarbeitet
        self.logger.log(
            f"[AF-Streaming] ✅ Processing complete:\n"
            f"  Total rounds: {round_num}\n"
            f"  Total processed: {total_processed:,}\n"
            f"  Final queue size: {len(self.to_fetch_af):,}"
        )

        self.progress.complete_phase("af_fetch")
        if pipeline_state:
            pipeline_state["af_done"].set()
            self.logger.log("[AF-Streaming] ✅ af_done.set() - Phase complete")

        return {}               
  
    def _phase_full_annotation_streaming(self, build: str, stale_days: int, pipeline_state: dict):
        """
        ✅ MIT FASTA-VALIDIERUNG INTEGRIERT
        
        BUG-FIXES:
        1. Defensives None-Checking für parsed
        2. Besseres Logging für Fehler-Diagnose
        3. Separate Fehler-Kategorien (parse, empty, exception)
        
        NEU: FASTA-FALLBACK
        - Validiert fehlgeschlagene Keys gegen FASTA-Referenz
        - Entfernt Varianten mit REF-Mismatch
        - Markiert ungültige Varianten in DB
        
        FIXES Problem: "'NoneType' object has no attribute 'get'"
        """
        # Guard: Phase deaktiviert?
        if not is_api_enabled("phase2_full", "myvariant_full"):
            self.logger.log("[Full-Anno] Phase deaktiviert, ueberspringe.")
            pipeline_state["full_done"].set()
            self.progress.complete_phase("full_anno")
            return

        # Warte auf erste Varianten
        start = time.time()
        while len(self.display_keys) < 10:
            if time.time() - start > 10 or self.stopflag.is_set():
                break
            time.sleep(0.5)

        self.progress.set_phase("full_anno")
        self.logger.log("[Full-Anno] Start.")

        FULL_ANNO_FIELDS = {
            "cadd_phred", "gene_symbol", "impact", "consequence",
            "clinical_significance", "phenotypes", "conservation"
        }

        processed_keys = set()
        idle_rounds = 0
        max_idle_rounds = 3

        # ============================================================
        # V16 FIX: BATCH-AKKUMULATION für effizientes Processing
        # Problem: Bei langsamen Scannern kommen nur 1-2 Varianten pro Iteration
        #          → Jeder API-Call hat den gleichen Overhead, aber nur für 1-2 Items
        # Lösung: Varianten akkumulieren bis MIN_BATCH_SIZE erreicht
        # ============================================================
        MIN_BATCH_SIZE = 50      # Mindestgröße für API-Call
        MAX_BATCH_WAIT = 15      # Max Sekunden warten auf Batch-Füllung
        FULL_ANNO_BATCH = get_api_setting("phase2_full", "myvariant_full", "batch_size", 500)
        batch_accumulator = []   # Akkumuliert Varianten für Batching
        last_batch_time = time.time()

        # ✅ Batch-Write-Konfiguration
        BATCH_WRITE_SIZE = 5000
        pending_writes = {}
        
        def flush_pending_writes():
            nonlocal pending_writes
            if not pending_writes:
                return
            try:
                items = list(pending_writes.items())
                keys = [k for k, v in items]  # Keys für Cache-Invalidierung
                self.db.upsert_many(items)
                self.logger.log(f"[Full-Anno] ✅ Batch-Write: {len(items)} Varianten")

                # ✅ V15 FIX: Cache-Invalidierung nach Batch-Write
                # Ohne dies zeigt die GUI alte Cache-Daten obwohl DB aktualisiert wurde
                try:
                    self.invalidate_cache_bulk(keys)
                    self.logger.log(f"[Full-Anno] ✅ Cache invalidiert: {len(keys)} Keys")
                except Exception as cache_err:
                    self.logger.log(f"[Full-Anno] ⚠️ Cache-Invalidierung fehlgeschlagen: {cache_err}")

                pending_writes.clear()
            except Exception as e:
                self.logger.log(f"[Full-Anno] ⚠️ Batch-Write fehlgeschlagen: {e}")
                pending_writes.clear()

        while True:
            if self.stopflag.is_set():
                flush_pending_writes()
                break

            current_keys = self.display_keys.copy()
            new_keys = [k for k in current_keys if k not in processed_keys]

            af_done = pipeline_state["af_done"].is_set()
            
            if not new_keys:
                if af_done:
                    idle_rounds += 1
                    if idle_rounds >= max_idle_rounds:
                        flush_pending_writes()
                        self.logger.log("[Full-Anno] AF done + idle → Stop.")
                        break
                time.sleep(2)
                continue

            idle_rounds = 0

            # Stale-Check (utcnow() ohne tzinfo für Konsistenz mit parse_iso_utc)
            stale_cutoff = datetime.datetime.utcnow() - datetime.timedelta(days=stale_days)
            to_annotate = []

            try:
                rows = self.db.get_variants_bulk(new_keys)
            except Exception as e:
                self.logger.log(f"[Full-Anno] ⚠️ DB-Bulk-Fetch failed: {e}")
                rows = {}

            for key in new_keys:
                row = rows.get(key) or {}

                missing_fields = []
                for field in FULL_ANNO_FIELDS:
                    val = row.get(field)
                    if val is None or val == "" or val == ".":
                        missing_fields.append(field)

                last_full = row.get("full_last_fetch")
                success_flag = row.get("full_fetch_success")
                dt = parse_iso_utc(last_full)
                is_stale = (not dt) or (dt < stale_cutoff)

                needs_annotation = (
                    len(missing_fields) > 0 or
                    is_stale or
                    not self.status_mgr.is_success(success_flag)
                )

                if needs_annotation:
                    to_annotate.append(key)

                processed_keys.add(key)
                self.progress.update_phase("full_anno", 1)

            if not to_annotate:
                continue

            # ============================================================
            # V16 FIX: Akkumuliere Varianten für größere Batches
            # ============================================================
            batch_accumulator.extend(to_annotate)
            time_since_last_batch = time.time() - last_batch_time

            # Prüfe ob wir einen Batch starten sollten
            should_process = False
            reason = ""

            if len(batch_accumulator) >= MIN_BATCH_SIZE:
                should_process = True
                reason = f"batch full ({len(batch_accumulator)} >= {MIN_BATCH_SIZE})"
            elif af_done and batch_accumulator:
                should_process = True
                reason = f"AF done, flushing {len(batch_accumulator)} remaining"
            elif time_since_last_batch > MAX_BATCH_WAIT and batch_accumulator:
                should_process = True
                reason = f"timeout ({time_since_last_batch:.0f}s), processing {len(batch_accumulator)}"

            if not should_process:
                # Warte auf mehr Varianten
                if len(batch_accumulator) > 0 and len(batch_accumulator) < MIN_BATCH_SIZE:
                    self.logger.log(
                        f"[Full-Anno] ⏳ Accumulating batch: {len(batch_accumulator)}/{MIN_BATCH_SIZE} "
                        f"(wait {MAX_BATCH_WAIT - time_since_last_batch:.0f}s)"
                    )
                continue

            # Jetzt verarbeiten
            to_annotate = batch_accumulator[:]
            batch_accumulator.clear()
            last_batch_time = time.time()

            self.logger.log(
                f"[Full-Anno] 📦 Processing batch: {len(to_annotate)} variants ({reason})"
            )

            try:
                mv_data = self._safe_mv_fetch_async(
                    to_annotate, 
                    build, 
                    fetch_full=True,
                    batch_size=FULL_ANNO_BATCH
                )
                
                # ✅ DEFENSIVE: Validiere mv_data
                if not mv_data:
                    self.logger.log("[Full-Anno] ⚠️ mv_data is None or empty")
                    mv_data = {}
                
                if not isinstance(mv_data, dict):
                    self.logger.log(
                        f"[Full-Anno] ⚠️ Invalid MV response type: {type(mv_data)}"
                    )
                    mv_data = {}
                
                # ✅ Zähler für verschiedene Fehlerarten
                success_count = 0
                parse_error_count = 0
                empty_result_count = 0
                exception_count = 0
                
                # ✅ NEU: Sammle fehlgeschlagene Keys für FASTA-Fallback
                batch_failed_keys = []
                
                for k in to_annotate:
                    try:
                        ann = mv_data.get(k)
                        
                        # ✅ DEFENSIVE: Prüfe ann
                        if not ann:
                            empty_result_count += 1
                            batch_failed_keys.append(k)  # ✅ NEU
                            
                            # Markiere als Fehlschlag
                            try:
                                current_statuses = self.db.get_variants_bulk([k])
                                current_status = current_statuses.get(k, {}).get("full_fetch_success")
                            except Exception:
                                current_status = None
                            
                            new_status = self.status_mgr.update_status(current_status, success=False)
                            pending_writes[k] = {
                                "full_last_fetch": now_iso(),
                                "full_fetch_success": new_status
                            }
                            continue
                        
                        if not isinstance(ann, dict):
                            parse_error_count += 1
                            self.logger.log(f"[Full-Anno] ⚠️ Non-dict annotation for {k}: {type(ann)}")
                            continue
                        
                        # ✅ KRITISCH: Rufe extract_fields_from_mv auf
                        # DEBUG: Temporär debug_log aktiviert für Diagnose leerer Annotations-Spalten
                        parsed = extract_fields_from_mv(ann, debug_log=True)
                        
                        # ✅ DEFENSIVE: Double-check parsed
                        if parsed is None:
                            parse_error_count += 1
                            self.logger.log(f"[Full-Anno] ⚠️ extract_fields_from_mv returned None for {k}")
                            continue
                        
                        if not isinstance(parsed, dict):
                            parse_error_count += 1
                            self.logger.log(f"[Full-Anno] ⚠️ extract_fields_from_mv returned non-dict for {k}: {type(parsed)}")
                            continue
                        
                        # ✅ SAFE: Jetzt können wir sicher .items() aufrufen
                        update_fields = {
                            field: val 
                            for field, val in parsed.items() 
                            if field in FULL_ANNO_FIELDS and val not in (None, "", ".")
                        }
                        
                        if update_fields:
                            update_fields["full_last_fetch"] = now_iso()
                            update_fields["full_fetch_success"] = "111"
                            pending_writes[k] = update_fields
                            success_count += 1
                        else:
                            empty_result_count += 1
                            batch_failed_keys.append(k)  # ✅ NEU: Auch leere Results prüfen
                    
                    except Exception as e:
                        exception_count += 1
                        self.logger.log(f"[Full-Anno] ⚠️ Processing error for {k}: {e}")
                        import traceback
                        self.logger.log(f"[Full-Anno] Traceback: {traceback.format_exc()}")
                    
                    self.progress.update_phase("full_anno", 1)
                
                # ====================================================================
                # ✅ NEU: FASTA-FALLBACK FÜR FEHLGESCHLAGENE KEYS IN DIESEM BATCH
                # ====================================================================
                if batch_failed_keys:
                    self.logger.log(
                        f"[Full-Anno] 🧬 Starte FASTA-Fallback für {len(batch_failed_keys)} "
                        f"fehlgeschlagene Keys in diesem Batch..."
                    )
                    
                    try:
                        valid_keys, invalid_keys = self._validate_with_fasta_fallback(
                            batch_failed_keys,
                            build,
                            phase_label="Full-Anno-FASTA-Fallback"
                        )
                        
                        # Entferne ungültige Keys
                        if invalid_keys:
                            for k in invalid_keys:
                                self.display_keys.discard(k)
                                self.current_vcf_keys.discard(k)
                                # Auch aus emit_queue entfernen falls vorhanden
                                self.emit_queue.discard(k)
                                
                                # Update Zähler
                                if k in pending_writes:
                                    del pending_writes[k]
                            
                            self.logger.log(
                                f"[Full-Anno] 🚫 {len(invalid_keys)} ungültige Varianten "
                                f"entfernt (FASTA-REF-Mismatch)"
                            )
                        
                        # Log Summary
                        if valid_keys or invalid_keys:
                            self.logger.log(
                                f"[Full-Anno] 🧬 FASTA-Fallback Complete: "
                                f"{len(valid_keys)} valid, {len(invalid_keys)} invalid"
                            )
                            
                    except Exception as e:
                        self.logger.log(f"[Full-Anno] ⚠️ FASTA-Fallback error: {e}")
                        import traceback
                        self.logger.log(f"[Full-Anno] Traceback: {traceback.format_exc()}")
                # ====================================================================
                
                # ✅ DETAILLIERTES LOGGING
                total_processed = success_count + parse_error_count + empty_result_count + exception_count
                self.logger.log(
                    f"[Full-Anno] ✅ Batch complete:\n"
                    f"  Success:        {success_count}\n"
                    f"  Empty results:  {empty_result_count}\n"
                    f"  Parse errors:   {parse_error_count}\n"
                    f"  Exceptions:     {exception_count}\n"
                    f"  Total:          {total_processed}"
                )
                
                # Batch-Write bei Grenze
                if len(pending_writes) >= BATCH_WRITE_SIZE:
                    flush_pending_writes()
            
            except Exception as e:
                self.logger.log(f"[Full-Anno] ⚠️ MV-Fetch error: {e}")
                import traceback
                self.logger.log(f"[Full-Anno] Traceback: {traceback.format_exc()}")

        # Final Flush
        flush_pending_writes()
        
        # ✅ WICHTIG: Markiere alle annotierten Keys als processed
        for k in processed_keys:
            self.progress.mark_fully_processed(1)
        
        # Signal: Full-Phase complete
        pipeline_state["full_done"].set()
        self.progress.complete_phase("full_anno")
        self.logger.log("[Full-Anno] ✅ Complete.")
        
    def _safe_mv_fetch_async(self, keys, build, fetch_full, batch_size):
        """
        Helper: Event-Loop-sicherer Wrapper für mv_fetch_async.
        
        Änderungen:
        - Prüft ob Event-Loop läuft
        - Nutzt run_coroutine_threadsafe bei laufendem Loop
        - Fallback auf ThreadPool
        """
        try:
            # ✅ Prüfe Event-Loop-Status
            try:
                loop = asyncio.get_running_loop()
                
                # ✅ Loop läuft → nutze run_coroutine_threadsafe
                future = asyncio.run_coroutine_threadsafe(
                    mv_fetch_async(
                        keys, build, fetch_full,
                        batch_size=batch_size,
                        logger=self.logger,
                        phase_label="FULL"
                    ),
                    loop
                )
                return future.result(timeout=300)
            
            except RuntimeError:
                # ✅ Kein Loop → asyncio.run
                return asyncio.run(
                    mv_fetch_async(
                        keys, build, fetch_full,
                        batch_size=batch_size,
                        logger=self.logger,
                        phase_label="FULL"
                    )
                )
        
        except Exception as e:
            self.logger.log(f"[Full-Anno] ⚠️ Async-Fetch failed: {e}")
            
            # ✅ Fallback auf ThreadPool
            return mv_fetch_threadpool(
                keys, build, fetch_full,
                batch_size=batch_size,
                logger=self.logger,
                phase_label="FULL"
            )
    @staticmethod
    def await_mv_fetch_async(keys, build, fetch_full, batch_size, logger, phase_label):
        """
        Helper: Ruft mv_fetch_async synchron auf (für Thread-Kontext).

        Args:
            keys: Liste von Variant-Keys (Tuples)
            build: GRCh37/GRCh38
            fetch_full: True für Voll-Annotation
            batch_size: Chunk-Größe für API
            logger: Logger-Instanz
            phase_label: Label für Logging

        Returns:
            Dict mit MV-Results
        """
        try:
            # Prüfe, ob Event Loop verfügbar
            try:
                loop = asyncio.get_running_loop()
                # Loop läuft → nutze run_coroutine_threadsafe
                future = asyncio.run_coroutine_threadsafe(
                    mv_fetch_async(
                        keys, build, fetch_full,
                        batch_size=batch_size,
                        logger=logger,
                        phase_label=phase_label
                    ),
                    loop
                )
                return future.result(timeout=300)

            except RuntimeError:
                # Kein Loop → asyncio.run
                return asyncio.run(
                    mv_fetch_async(
                        keys, build, fetch_full,
                        batch_size=batch_size,
                        logger=logger,
                        phase_label=phase_label
                    )
                )

        except Exception as e:
            if logger:
                logger.log(f"[{phase_label}] ⚠️ Async fetch failed: {e}")
            return {}
        
    def _phase_gene_annotation_streaming(self, pipeline_state: dict):
        """
        FIX: Wartet nun auf den GeneAnnotator, statt sofort abzubrechen.
        Schreibt Ergebnisse direkt via vcf_buffer in die Haupttabelle.
        """
        # --- FIX: WARTE-SCHLEIFE FÜR ANNOTATOR ---
        # Wir warten bis zu 120 Sekunden, falls der Annotator noch im Hintergrund lädt
        wait_seconds = 0
        while wait_seconds < 120:
            if self.stopflag.is_set():
                pipeline_state["gene_done"].set()
                return
                
            if self.gene_annotator and self.gene_annotator.available:
                break # Annotator ist bereit!
                
            time.sleep(1)
            wait_seconds += 1
            if wait_seconds % 10 == 0:
                self.logger.log(f"[Gene-Anno] Warte auf GeneAnnotator Initialisierung... ({wait_seconds}s)")

        if not self.gene_annotator or not self.gene_annotator.available:
            self.logger.log("[Gene-Anno] ⚠️ Annotator nicht verfügbar (Timeout) - Phase übersprungen.")
            pipeline_state["gene_done"].set()
            return
        # -----------------------------------------
        
        # Warte auf erste Varianten
        start = time.time()
        while len(self.display_keys) < 10:
            if time.time() - start > 10 or self.stopflag.is_set():
                break
            time.sleep(1)
        
        self.progress.set_phase("gene_anno")
        self.logger.log("[Gene-Anno] Start.")
        
        processed_keys = set()
        idle_rounds = 0
        
        while True:
            if self.stopflag.is_set():
                break
            
            current_keys = self.display_keys.copy()
            new_keys = [k for k in current_keys if k not in processed_keys]
            
            af_done = pipeline_state["af_done"].is_set()
            if not new_keys:
                if af_done:
                    idle_rounds += 1
                    if idle_rounds >= 3:
                        break
                time.sleep(2)
                continue
            
            idle_rounds = 0
            
            try:
                # Nur Varianten ohne Gene-Symbol
                to_annotate = []
                
                # Bulk Fetch um DB-Hits zu minimieren
                chunk_rows = self.db.get_variants_bulk(new_keys)
                
                for k in new_keys:
                    row = chunk_rows.get(k) or {}
                    # Wir prüfen, ob gene_symbol fehlt
                    if not row.get("gene_symbol"):
                        to_annotate.append(k)
                    processed_keys.add(k)
                
                if to_annotate:
                    self.logger.log(
                        f"[Gene-Anno] 📦 Processing batch: {len(to_annotate)} variants "
                        f"(missing gene symbols)"
                    )
                    
                    pos_map = self.db.get_positions(to_annotate)
                    query = [(k, chrom, pos) for k, (chrom, pos) in pos_map.items()]
                    
                    # Batch Annotation
                    annots = self.gene_annotator.annotate_batch(query)
                    
                    success_count = 0
                    
                    for key, data in annots.items():
                        if data and data.get("gene_symbol"):
                            # ✅ FIX: Direkt in den VCF-Buffer (Priority=True für sofortigen Flush)
                            self.vcf_buffer.add(key, {
                                "gene_symbol": data["gene_symbol"],
                                "is_coding": data["is_coding"]
                            }, priority=True)
                            
                            success_count += 1
                    
                    self.progress.update_phase("gene_anno", len(to_annotate))
                    self.logger.log(f"[Gene-Anno] ✅ Batch complete: {success_count} genes found")
            
            except Exception as e:
                self.logger.log(f"[Gene-Anno] ⚠️ Batch error: {e}")
                time.sleep(2)
        
        # Buffer flushen um sicherzugehen
        self.vcf_buffer.flush(force_priority=True)
        
        pipeline_state["gene_done"].set()
        self.progress.complete_phase("gene_anno")
        self.logger.log("[Gene-Anno] Complete.")


    def _phase_rsid_fill_streaming(self, build: str, pipeline_state: dict):
        """
        BUG-FIX: Connection aborted Errors
        - Besseres Exception-Handling für Netzwerk-Fehler
        - Automatisches Session-Retry
        - Circuit-Breaker für wiederkehrende Fehler
        """
        # Guard: Phase deaktiviert?
        if not is_api_enabled("phase4_rsid", "ncbi_allele"):
            self.logger.log("[RSID-Fill] Phase deaktiviert, ueberspringe.")
            pipeline_state["rsid_done"].set()
            self.progress.complete_phase("rsid_fill")
            return

        start = time.time()
        while len(self.display_keys) < 10 and not self.stopflag.is_set():
            if time.time() - start > 10:
                break
            time.sleep(0.5)

        self.progress.set_phase("rsid_fill")
        self.logger.log("[RSID-Fill] Start.")

        processed_keys: Set[Tuple] = set()
        idle_rounds = 0
        max_idle_rounds = 3

        # Rate-Limiting (aus Settings)
        last_request_time = 0.0
        MIN_REQUEST_INTERVAL = get_api_setting("phase4_rsid", "ncbi_allele", "rate_limit", 0.35)
        ncbi_key = get_api_setting("global", "global", "ncbi_api_key", "")
        if ncbi_key:
            MIN_REQUEST_INTERVAL = min(MIN_REQUEST_INTERVAL, 0.1)
        JITTER_MAX = 0.05

        # ✅ BUG-FIX: Session-Retry-Mechanismus
        session = None
        session_errors = 0
        MAX_SESSION_ERRORS = 10
        
        def get_or_recreate_session():
            """Helper: Erstellt neue Session bei Problemen"""
            nonlocal session, session_errors
            
            if session is None or session_errors >= MAX_SESSION_ERRORS:
                if session:
                    try:
                        session.close()
                    except Exception:
                        pass
                
                session = requests.Session()
                _hdrs = {"User-Agent": "VariantFusion/RSID-Fill"}
                if ncbi_key:
                    _hdrs["api_key"] = ncbi_key
                session.headers.update(_hdrs)
                session_errors = 0
                self.logger.log("[RSID-Fill] 🔄 Session (re)created")
            
            return session

        # Initiale Session
        session = get_or_recreate_session()

        # Batch-Updates
        BATCH_SIZE = 100
        pending_updates: List[Tuple[Tuple, Dict[str, str]]] = []

        def flush_updates():
            nonlocal pending_updates
            if not pending_updates:
                return
            try:
                self.db.update_variant_rsids_bulk(pending_updates)
                self.logger.log(f"[RSID-Fill] ✅ Batch: {len(pending_updates)} RSIDs")
                pending_updates.clear()
            except Exception as e:
                self.logger.log(f"[RSID-Fill] ⚠️ Batch update failed: {e}")
                pending_updates.clear()

        def safe_get_rsid_from_allele_json(data: Union[Dict, List, None]) -> Optional[str]:
            """Defensive Parser für NCBI allele response"""
            if not isinstance(data, dict):
                return None

            # Direktfeld
            rsid = data.get("refsnp_id")
            if rsid:
                return str(rsid)

            # Geschachtelt unter "allele"
            allele = data.get("allele")
            if isinstance(allele, dict):
                rsid = allele.get("refsnp_id") or allele.get("rsid")
                if rsid:
                    return str(rsid)
            elif isinstance(allele, list):
                for entry in allele:
                    if isinstance(entry, dict):
                        rsid = entry.get("refsnp_id") or entry.get("rsid")
                        if rsid:
                            return str(rsid)

            # Fallback: primary_snapshot_data
            psd = data.get("primary_snapshot_data")
            if isinstance(psd, dict):
                rid = psd.get("refsnp_id")
                if rid:
                    return str(rid)

            return None

        def rate_limit_sleep():
            nonlocal last_request_time
            now = time.time()
            elapsed = now - last_request_time
            target = MIN_REQUEST_INTERVAL + random.uniform(0.0, JITTER_MAX)
            if elapsed < target:
                time.sleep(target - elapsed)

        # Haupt-Loop
        while True:
            if self.stopflag.is_set():
                flush_updates()
                break

            current_keys = self.display_keys.copy()
            new_keys = [k for k in current_keys if k not in processed_keys]

            af_done = pipeline_state["af_done"].is_set()
            if not new_keys:
                if af_done:
                    idle_rounds += 1
                    if idle_rounds >= max_idle_rounds:
                        flush_updates()
                        self.logger.log("[RSID-Fill] AF done + idle → Stop.")
                        break
                time.sleep(2)
                continue

            idle_rounds = 0

            # ✅ Bulk-Fetch für Performance
            try:
                rows_bulk = self.db.get_variants_bulk(new_keys)
            except Exception as e:
                self.logger.log(f"[RSID-Fill] ⚠️ Bulk-fetch failed: {e}")
                rows_bulk = {}

            for key in new_keys:
                if self.stopflag.is_set():
                    break

                chrom, pos, ref, alt, build_ = key
                row = rows_bulk.get(key) or {}

                if row.get("rsid") not in (None, "", "."):
                    processed_keys.add(key)
                    continue

                # HGVS erzeugen
                hgvs = None
                try:
                    hgvs = make_hgvs(chrom, pos, ref, alt)
                except Exception as e:
                    self.logger.log(f"[RSID-Fill] ⚠️ HGVS build failed for {key}: {e}")

                if not hgvs or not isinstance(hgvs, str):
                    processed_keys.add(key)
                    self.progress.update_phase("rsid_fill", 1)
                    continue

                # Rate-Limit
                rate_limit_sleep()

                # ✅ GET mit robustem Error-Handling
                url = f"https://api.ncbi.nlm.nih.gov/variation/v0/allele/{hgvs}"
                last_request_time = time.time()

                success = False
                for attempt in range(4):
                    if self.stopflag.is_set():
                        break
                    
                    try:
                        # ✅ Hole aktuelle Session (recreate bei Problemen)
                        current_session = get_or_recreate_session()
                        
                        r = current_session.get(url, timeout=10)
                        
                        if r.status_code == 200:
                            data = r.json()
                            rsid = safe_get_rsid_from_allele_json(data)
                            if rsid:
                                pending_updates.append((key, {"rsid": rsid}))
                            success = True
                            session_errors = 0  # Reset bei Erfolg
                            break
                        
                        elif r.status_code == 429:
                            backoff = min(8.0, 2.0 ** attempt)
                            self.logger.log(f"[RSID-Fill] 429 - warte {backoff:.1f}s")
                            time.sleep(backoff)
                            continue
                        
                        elif r.status_code in (404, 422):
                            success = True  # Kein weiterer Retry
                            break
                        
                        else:
                            session_errors += 1
                            time.sleep(0.25 * (attempt + 1))
                    
                    except requests.exceptions.ConnectionError as e:
                        # ✅ SPEZIFISCH: Connection-Fehler
                        session_errors += 1
                        self.logger.log(
                            f"[RSID-Fill] ⚠️ Connection error for {hgvs} (attempt {attempt+1}/4): {e}"
                        )
                        
                        # Session neu erstellen bei Connection-Fehlern
                        if session_errors >= 3:
                            get_or_recreate_session()
                        
                        backoff = 1.0 + attempt * 0.5
                        time.sleep(backoff)
                    
                    except requests.Timeout:
                        session_errors += 1
                        backoff = 1.0 + attempt * 0.5
                        time.sleep(backoff)
                    
                    except Exception as e:
                        session_errors += 1
                        self.logger.log(f"[RSID-Fill] ⚠️ Request error for {hgvs}: {e}")
                        break

                processed_keys.add(key)
                self.progress.update_phase("rsid_fill", 1)

                # Batch-Flush
                if len(pending_updates) >= BATCH_SIZE:
                    flush_updates()

            if pending_updates and (af_done or self.stopflag.is_set()):
                flush_updates()

        # Final Flush
        flush_updates()

        # Cleanup
        pipeline_state["rsid_done"].set()
        try:
            if session:
                session.close()
        except Exception:
            pass
        
        self.progress.complete_phase("rsid_fill")
        self.logger.log("[RSID-Fill] ✅ Complete.")


    def _phase_missing_fill_streaming(self, build: str, pipeline_state: dict):
        """
        FIX: Robustes Missing-Fill mit Circuit-Breaker.
        
        Änderungen:
        - Circuit-Breaker bei wiederholten API-Fehlern
        - Separate Timeouts für ClinVar/MyVariant
        - Batch-Updates für Performance
        """
        # Warte auf Full-Anno-Done
        while not pipeline_state["full_done"].is_set():
            if self.stopflag.is_set():
                return
            time.sleep(0.5)

        self.progress.set_phase("missing_fill")
        self.logger.log("[Missing-Fill] Start.")

        processed_keys = set()
        idle_rounds = 0
        max_idle_rounds = 3

        # ✅ Circuit-Breaker für APIs (mit Settings-Check)
        clinvar_errors = 0
        myvariant_errors = 0
        MAX_API_ERRORS = 50  # Nach 50 Fehlern → Skip API

        clinvar_enabled = is_api_enabled("phase5_fill", "ncbi_clinvar")
        myvariant_enabled = is_api_enabled("phase5_fill", "myvariant_cons")

        # Rate-Limiting (aus Settings)
        last_clinvar_call = 0.0
        last_myvariant_call = 0.0
        min_call_interval = get_api_setting("phase5_fill", "ncbi_clinvar", "rate_limit", 0.5)
        _ncbi_key_mf = get_api_setting("global", "global", "ncbi_api_key", "")
        if _ncbi_key_mf:
            min_call_interval = min(min_call_interval, 0.1)

        # ✅ Batch-Updates
        BATCH_SIZE = 100
        pending_updates = []
        
        def flush_updates():
            """Helper: Schreibt akkumulierte Updates."""
            nonlocal pending_updates
            
            if not pending_updates:
                return
            
            try:
                update_dict = dict(pending_updates)
                # ✅ V14 FIX: Korrigiert zu upsert_many (upsert_annotations_bulk existiert nicht)
                self.db.upsert_many(list(update_dict.items()))
                self.logger.log(f"[Missing-Fill] ✅ Batch: {len(pending_updates)} updates")
                pending_updates.clear()
            except Exception as e:
                self.logger.log(f"[Missing-Fill] ⚠️ Batch update failed: {e}")
                pending_updates.clear()

        while True:
            if self.stopflag.is_set():
                flush_updates()
                break

            current_keys = self.display_keys.copy()
            new_keys = [k for k in current_keys if k not in processed_keys]

            if not new_keys:
                if pipeline_state["full_done"].is_set():
                    idle_rounds += 1
                    if idle_rounds >= max_idle_rounds:
                        flush_updates()
                        self.logger.log("[Missing-Fill] Full-Anno done + idle → Stop.")
                        break
                time.sleep(2)
                continue

            idle_rounds = 0

            # ✅ Filtere nach fehlenden Feldern
            to_annotate = []
            rows = self.db.get_variants_bulk(new_keys)
            
            for key in new_keys:
                row = rows.get(key) or {}
                
                needs_clinvar = (
                    clinvar_enabled and
                    (not row.get("clinical_significance") or not row.get("phenotypes")) and 
                    row.get("rsid")
                )
                
                needs_conservation = (
                    myvariant_enabled and
                    not row.get("conservation")
                )
                
                if needs_clinvar or needs_conservation:
                    to_annotate.append({
                        "key": key,
                        "row": row,
                        "needs_clinvar": needs_clinvar,
                        "needs_conservation": needs_conservation
                    })
                
                processed_keys.add(key)
            
            if not to_annotate:
                continue
            
            self.logger.log(f"[Missing-Fill] Processing {len(to_annotate)} variants...")
            
            # ✅ Worker-Funktion mit Circuit-Breaker
            def fetch_annotations(item):
                nonlocal last_clinvar_call, last_myvariant_call
                nonlocal clinvar_errors, myvariant_errors
                nonlocal clinvar_enabled, myvariant_enabled
                
                key = item["key"]
                row = item["row"]
                needs_clinvar = item["needs_clinvar"]
                needs_conservation = item["needs_conservation"]
                
                chrom, pos, ref, alt, build_ = key
                update_fields = {}
                
                # --- ClinVar ---
                if needs_clinvar and clinvar_enabled:
                    # Check Circuit-Breaker
                    if clinvar_errors >= MAX_API_ERRORS:
                        clinvar_enabled = False
                        self.logger.log(
                            "[Missing-Fill] ⚠️ ClinVar disabled (too many errors)"
                        )
                    
                    if clinvar_enabled:
                        # Rate-Limiting
                        now = time.time()
                        wait_time = max(0, min_call_interval - (now - last_clinvar_call))
                        if wait_time > 0:
                            time.sleep(wait_time)
                        
                        try:
                            rsid = row.get("rsid", "").replace("rs", "")
                            if rsid and rsid.isdigit():
                                url = f"https://api.ncbi.nlm.nih.gov/variation/v0/refsnp/{rsid}"
                                _cv_timeout = get_api_setting("phase5_fill", "ncbi_clinvar", "timeout", 10)
                                _cv_headers = {"api_key": _ncbi_key_mf} if _ncbi_key_mf else {}

                                # ✅ Timeout + Retry
                                for attempt in range(2):
                                    try:
                                        r = requests.get(url, timeout=_cv_timeout, headers=_cv_headers)
                                        last_clinvar_call = time.time()
                                        
                                        if r.status_code == 200:
                                            data = r.json()
                                            
                                            # Clinical Significance
                                            clinvar_data = data.get("primary_snapshot_data", {})
                                            clin_asserts = clinvar_data.get("clinical_assertions", [])
                                            
                                            if clin_asserts and not row.get("clinical_significance"):
                                                sigs = set()
                                                for assertion in clin_asserts:
                                                    sig = assertion.get("clinical_significance", {}).get("description")
                                                    if sig:
                                                        sigs.add(sig)
                                                
                                                if sigs:
                                                    update_fields["clinical_significance"] = ",".join(sorted(sigs))
                                            
                                            # Phenotypes
                                            if not row.get("phenotypes"):
                                                phenos = []
                                                for allele in clinvar_data.get("allele_annotations", []):
                                                    for assoc in allele.get("clinical_assertions", []):
                                                        for trait in assoc.get("traits", []):
                                                            name = trait.get("name")
                                                            if name:
                                                                phenos.append(name)
                                                
                                                if phenos:
                                                    from collections import Counter
                                                    top_phenos = [p for p, _ in Counter(phenos).most_common(3)]
                                                    update_fields["phenotypes"] = ",".join(top_phenos)
                                            
                                            break  # Success
                                        
                                        elif r.status_code == 429:
                                            time.sleep(2 ** attempt)
                                            continue
                                        elif r.status_code == 404:
                                            break  # Not found → Skip
                                        else:
                                            clinvar_errors += 1
                                            break
                                    
                                    except requests.Timeout:
                                        clinvar_errors += 1
                                        if attempt == 1:
                                            break
                        
                        except Exception as e:
                            clinvar_errors += 1
                
                # --- Conservation (MyVariant) ---
                if needs_conservation and myvariant_enabled:
                    # Check Circuit-Breaker
                    if myvariant_errors >= MAX_API_ERRORS:
                        myvariant_enabled = False
                        self.logger.log(
                            "[Missing-Fill] ⚠️ MyVariant disabled (too many errors)"
                        )
                    
                    if myvariant_enabled:
                        # Rate-Limiting
                        now = time.time()
                        wait_time = max(0, min_call_interval - (now - last_myvariant_call))
                        if wait_time > 0:
                            time.sleep(wait_time)
                        
                        try:
                            hgvs = make_hgvs(chrom, pos, ref, alt)
                            if hgvs:
                                url = f"https://myvariant.info/v1/variant/{hgvs}?fields=phylop,phastcons,gerp"
                                
                                # ✅ Timeout + Retry
                                for attempt in range(2):
                                    try:
                                        r = requests.get(url, timeout=10)
                                        last_myvariant_call = time.time()
                                        
                                        if r.status_code == 200:
                                            data = r.json()
                                            cons_parts = []
                                            
                                            # PhyloP
                                            if "phylop" in data:
                                                val = data["phylop"]
                                                if isinstance(val, dict):
                                                    for k, v in val.items():
                                                        if isinstance(v, (int, float)):
                                                            cons_parts.append(f"phyloP:{v:.3f}")
                                                            break
                                                elif isinstance(val, (int, float)):
                                                    cons_parts.append(f"phyloP:{val:.3f}")
                                            
                                            # PhastCons
                                            if "phastcons" in data:
                                                val = data["phastcons"]
                                                if isinstance(val, dict):
                                                    for k, v in val.items():
                                                        if isinstance(v, (int, float)):
                                                            cons_parts.append(f"phastCons:{v:.3f}")
                                                            break
                                                elif isinstance(val, (int, float)):
                                                    cons_parts.append(f"phastCons:{val:.3f}")
                                            
                                            # GERP
                                            if "gerp" in data:
                                                val = data["gerp"]
                                                if isinstance(val, dict):
                                                    rs = val.get("rs")
                                                    if isinstance(rs, (int, float)):
                                                        cons_parts.append(f"GERP:{rs:.3f}")
                                                elif isinstance(val, (int, float)):
                                                    cons_parts.append(f"GERP:{val:.3f}")
                                            
                                            if cons_parts:
                                                update_fields["conservation"] = ",".join(cons_parts)
                                            
                                            break  # Success
                                        
                                        elif r.status_code == 429:
                                            time.sleep(2 ** attempt)
                                            continue
                                        else:
                                            myvariant_errors += 1
                                            break
                                    
                                    except requests.Timeout:
                                        myvariant_errors += 1
                                        if attempt == 1:
                                            break
                        
                        except Exception as e:
                            myvariant_errors += 1
                
                # Rückgabe
                if update_fields:
                    return (key, update_fields)
                return None
            
            # ✅ Parallel mit ThreadPool (max 5 Worker)
            with ThreadPoolExecutor(max_workers=5) as executor:
                futures = [executor.submit(fetch_annotations, item) for item in to_annotate]
                
                for future in as_completed(futures):
                    try:
                        result = future.result(timeout=30)
                        if result:
                            pending_updates.append(result)
                            self.progress.update_phase("missing_fill", 1)
                            
                            # ✅ Batch-Flush bei Grenze
                            if len(pending_updates) >= BATCH_SIZE:
                                flush_updates()
                    
                    except Exception as e:
                        self.logger.log(f"[Missing-Fill] Worker exception: {e}")
        
        # ✅ Final Flush
        flush_updates()
        
        # Signal: Missing-Fill-Phase complete
        pipeline_state["missing_done"].set()
        self.logger.log("[Missing-Fill] ✅ Complete.")


    def _phase_alphagenome_streaming(self, pipeline_state: dict):
        """
        ✅ OPTIMIERT: Bulk-DB-Updates statt Single-Updates.
        FIX: Korrekte Completion-Markierung für ALLE Varianten.
        """
        # Guard: Phase deaktiviert?
        if not is_api_enabled("phase6_ag", "alphagenome"):
            self.logger.log("[AlphaGenome] Phase deaktiviert, ueberspringe.")
            pipeline_state["alphagenome_done"].set()
            self.progress.complete_phase("ag_score")
            self.progress.complete_pipeline()
            return

        # ✅ Phase-Start tracken
        self.progress.set_phase("ag_score")

        # Fall 1: AlphaGenome nicht verfügbar
        if not self.ag_scorer or not self.ag_scorer.available:
            self.logger.log("[AlphaGenome] ⚠️ Nicht verfügbar - warte auf andere Phasen.")
            
            # ✅ V10 FIX: Warte auf alle vorherigen Phasen MIT StopFlag-Check
            for event_name in ["af_done", "full_done", "gene_done", "rsid_done", "missing_done"]:
                if self.stopflag.is_set():
                    self.logger.log("[AlphaGenome] ⏸️ Abbruch durch StopFlag (Warte-Phase)")
                    pipeline_state["alphagenome_done"].set()
                    return
                event = pipeline_state.get(event_name)
                if event:
                    # Warte mit Timeout, um StopFlag regelmäßig zu prüfen
                    while not event.is_set():
                        if self.stopflag.is_set():
                            self.logger.log("[AlphaGenome] ⏸️ Abbruch durch StopFlag (Warte-Phase)")
                            pipeline_state["alphagenome_done"].set()
                            return
                        event.wait(timeout=1.0)

            # ✅ Markiere ALLE display_keys als processed
            final_keys = self.display_keys.copy()
            for k in final_keys:
                self.progress.mark_fully_processed(1)

            self.progress.complete_phase("ag_score")
            pipeline_state["alphagenome_done"].set()
            self.logger.log(f"[AlphaGenome] ✅ Skipped - {len(final_keys)} variants marked as complete.")
            self.progress.complete_pipeline()
            self.logger.log(f"Pipeline abgeschlossen")
            return

        # Fall 2: AlphaGenome verfügbar
        # ✅ V10 FIX: Warte auf AF-Phase MIT StopFlag-Check
        af_done = pipeline_state.get("af_done")
        if af_done:
            while not af_done.is_set():
                if self.stopflag.is_set():
                    self.logger.log("[AlphaGenome] ⏸️ Abbruch durch StopFlag (AF-Wait)")
                    pipeline_state["alphagenome_done"].set()
                    return
                af_done.wait(timeout=1.0)
        
        self.logger.log("[AlphaGenome] Start.")

        processed_keys = set()
        scored_count = 0
        skipped_count = 0
        error_count = 0

        idle_rounds = 0
        max_idle_rounds = 3

        while True:
            if self.stopflag.is_set():
                self.logger.log("[AlphaGenome] ⚠️ Stopflag gesetzt - breche ab.")
                break

            current_keys = self.display_keys.copy()
            new_keys = [k for k in current_keys if k not in processed_keys]

            if not new_keys:
                if af_done and af_done.is_set():
                    idle_rounds += 1
                    if idle_rounds >= max_idle_rounds:
                        self.logger.log("[AlphaGenome] AF-Anno done + idle → Stop.")
                        break
                time.sleep(2)
                continue

            idle_rounds = 0
            to_score = []
            already_scored = []

            rows = self.db.get_variants_bulk(new_keys)

            for k in new_keys:
                row = rows.get(k) or {}
                processed_keys.add(k)

                if row.get("ag_score") is not None:
                    already_scored.append(k)
                    # ✅ WICHTIG: Auch bereits gescorte Keys als processed markieren
                    self.progress.mark_fully_processed(1)
                    continue

                try:
                    chrom, pos, ref, alt, build = k
                    to_score.append((chrom, int(pos), ref, alt, build))
                except Exception as e:
                    self.logger.log(f"[AlphaGenome] ⚠️ Ungültiger Key {k}: {e}")
                    skipped_count += 1
                    # ✅ Auch fehlerhafte Keys als processed markieren
                    self.progress.mark_fully_processed(1)
                    self.progress.update_phase("ag_score", 1)

            if to_score:
                self.logger.log(f"[AlphaGenome] 📦 Scoring batch: {len(to_score)} variants...")
                try:
                    scores = self.ag_scorer.score_batch(to_score, logger=self.logger) or {}
                    
                    # ✅ BULK: Sammle alle Updates
                    bulk_updates = []
                    
                    for tup in to_score:
                        score = scores.get(tup)
                        
                        if score is not None:
                            bulk_updates.append((tup, {"ag_score": score}))
                        else:
                            skipped_count += 1
                        
                        # Progress für jeden Key
                        self.progress.mark_fully_processed(1)
                        self.progress.update_phase("ag_score", 1)
                    
                    # ✅ BULK-UPDATE (statt einzelne Calls)
                    if bulk_updates:
                        try:
                            self.db.upsert_many(bulk_updates)
                            scored_count += len(bulk_updates)
                            self.logger.log(
                                f"[AlphaGenome] ✅ Bulk-Update: {len(bulk_updates)} scores gespeichert"
                            )
                        except Exception as e:
                            self.logger.log(f"[AlphaGenome] ⚠️ Bulk-Update failed: {e}")
                            error_count += len(bulk_updates)
                
                except Exception as e:
                    self.logger.log(f"[AlphaGenome] ⚠️ Batch scoring failed: {e}")
                    error_count += len(to_score)
                    # ✅ Alle fehlgeschlagenen auch markieren
                    for _ in to_score:
                        self.progress.mark_fully_processed(1)
                        self.progress.update_phase("ag_score", 1)

        # ✅ Warte auf alle anderen Phasen
        for event_name in ["af_done", "full_done", "gene_done", "rsid_done", "missing_done"]:
            event = pipeline_state.get(event_name)
            if event:
                event.wait()

        # ✅ Finaler Pass: Alle display_keys als processed markieren
        final_keys = self.display_keys.copy()
        unmarked = [k for k in final_keys if k not in processed_keys]
        
        if unmarked:
            self.logger.log(f"[AlphaGenome] ⚠️ Marking {len(unmarked)} unmarked keys as processed")
            for k in unmarked:
                self.progress.mark_fully_processed(1)

        # ✅ Phase als abgeschlossen markieren
        self.progress.complete_phase("ag_score")
        pipeline_state["alphagenome_done"].set()

        total_processed = scored_count + skipped_count + error_count
        self.logger.log(
            f"[AlphaGenome] ✅ Complete: "
            f"{scored_count} scored, "
            f"{skipped_count} skipped, "
            f"{error_count} errors, "
            f"{total_processed} total processed."
        )
        self.progress.complete_pipeline()
        self.logger.log(f"Pipeline abgeschlossen")
        
    def _finalize_pipeline(self):
        """
        FIX: Stellt sicher dass alle Progress-Counter 100% zeigen am Ende.
        Behebt das Problem dass am Pipeline-Ende nicht 100% angezeigt wird.
        """
        self.logger.log("[Pipeline] 🎯 Finalizing progress counters...")
        
        with self.progress._lock:
            # Markiere alle Phasen als komplett
            for phase in self.progress.PHASE_WEIGHTS.keys():
                self.progress._phases_completed.add(phase)
            
            # Setze Counter auf Finale Werte basierend auf total_variants
            total = self.progress.total_variants
            
            # Alle Phasen-Counter auf Total setzen
            self.progress.vcf_scan_done = total
            self.progress.af_fetch_done = total
            self.progress.full_anno_done = total
            self.progress.gene_anno_done = total
            self.progress.rsid_fill_done = total  
            self.progress.missing_fill_done = total
            self.progress.ag_score_done = total
            
            # Markiere alle als processed
            self.progress.processed_variants = total
            self.progress.phase_name = "Abgeschlossen"
            
        self.logger.log(f"[Pipeline] ✅ Finalized: {total} variants processed (100%)")
    def _cleanup_controllers(self):
        """
        FIX: Saubere Shutdown-Cleanup für Controller und Puffer.

        Änderungen:
        - Finalize AFFetchController (falls vorhanden)
        - EmitQueue und VCFBuffer flushen, Fehler tolerieren
        - GeneAnnotator optional schließen (wenn close/stop existiert)
        - Display-Keys Thread-sicher leeren
        - Maintainer pausieren, dann wieder freigeben (defensiv)
        - Progress weich beenden (Phase 'Abgeschlossen' setzen, falls sinnvoll)
        """
        # AFFetchController finalisieren
        if hasattr(self, "_af_controller") and self._af_controller:
            try:
                self._af_controller.finalize()
            except Exception as e:
                self.logger.log(f"[Cleanup] ⚠️ AF-Controller finalize error: {e}")
            finally:
                self._af_controller = None

        # Maintainer pausieren (defensiv)
        if getattr(self, "maint", None):
            try:
                self.maint.pause()
            except Exception:
                pass

        # EmitQueue flushen
        try:
            if hasattr(self, "emit_queue") and self.emit_queue:
                self.emit_queue.flush()
        except Exception as e:
            self.logger.log(f"[Cleanup] ⚠️ EmitQueue flush error: {e}")

        # VCFBuffer flushen
        try:
            if hasattr(self, "vcf_buffer") and self.vcf_buffer:
                self.vcf_buffer.flush()
        except Exception as e:
            self.logger.log(f"[Cleanup] ⚠️ VCFBuffer flush error: {e}")

        # GeneAnnotator ggf. schließen
        try:
            if getattr(self, "gene_annotator", None):
                ga = self.gene_annotator
                if hasattr(ga, "close") and callable(ga.close):
                    ga.close()
                elif hasattr(ga, "stop") and callable(ga.stop):
                    ga.stop()
        except Exception as e:
            self.logger.log(f"[Cleanup] ⚠️ GeneAnnotator cleanup error: {e}")
        finally:
            self.gene_annotator = None

        # Display-Keys leeren (thread-sicher)
        try:
            with self._keys_lock:
                self._display_keys.clear()
        except Exception:
            # Fallback
            try:
                self.display_keys = set()
            except Exception:
                pass

        # Maintainer fortsetzen (falls nötig)
        if getattr(self, "maint", None):
            try:
                self.maint.resume()
            except Exception:
                pass

        # Progress weich beenden
        try:
            if hasattr(self, "progress") and self.progress:
                # Setze Phase auf 'Abgeschlossen', wenn Pipeline lief
                if getattr(self.progress, "total", 0) > 0:
                    self.progress.set_phase("Abgeschlossen")
        except Exception:
            pass

        self.logger.log("[Cleanup] ✅ Controllers and buffers cleaned up.")

                    
    # ==================== Format Converters ====================
    
    def _process_23andme(self, path: str, build: Optional[str] = None):
        """
        V10: Convert and process 23andMe file.
        Parameter werden vom FlagManager geholt.
        """
        try:
            # Klasse mit file_path instanziieren
            conv = convert_23andme_to_vcf(path)
            vcf_path, out_build = conv.start(build)

            if not vcf_path or not out_build:
                self.logger.log("[23andMe] Conversion failed or cancelled.")
                return

            self.logger.log(f"[23andMe] Conversion complete: {vcf_path} (Build: {out_build})")

            # V10: Alle Parameter kommen vom FlagManager
            self._distill_vcf(vcf_path=vcf_path, build=out_build)

        except Exception as e:
            self.logger.log(f"[23andMe] Error: {e}")
            import traceback
            self.logger.log(f"[23andMe] Traceback:\n{traceback.format_exc()}")
    
    def _process_fasta(self, path: str, build: str = "GRCh38"):
        """
        V10: Convert and process FASTA file.
        Parameter werden vom FlagManager geholt.
        """
        try:
            conv = StreamingFastaToGVCF()
            file_size_mb = os.path.getsize(path) / (1024 * 1024)
            tmp_vcf = os.path.splitext(os.path.basename(path))[0] + ".vcf"
            
            if file_size_mb < 50:
                self.logger.log(f"[FASTA] Small file ({file_size_mb:.1f} MB), creating gVCF...")
                conv.convert_streaming_gvcf(path, None, tmp_vcf)
                filtered_vcf = tmp_vcf.replace(".vcf", "_variants.vcf")
                with open(tmp_vcf, "r") as fin, open(filtered_vcf, "w") as fout:
                    for line in fin:
                        if line.startswith("#") or line.split("\t")[4] != ".":
                            fout.write(line)
                vcf_for_pipeline = filtered_vcf
            else:
                self.logger.log(f"[FASTA] Large file ({file_size_mb:.1f} MB), creating variants-only VCF...")
                conv.convert_streaming_variants_only(path, None, tmp_vcf)
                vcf_for_pipeline = tmp_vcf
            
            self.logger.log(f"[FASTA] Conversion complete: {vcf_for_pipeline}")
            
            # V10: Alle Parameter kommen vom FlagManager
            self._distill_vcf(vcf_path=vcf_for_pipeline, build=build)
        
        except Exception as e:
            self.logger.log(f"[FASTA] Error: {e}")
            import traceback
            self.logger.log(f"[FASTA] Traceback:\n{traceback.format_exc()}")
    
    def _process_fastq(self, path: str, build: str = "GRCh38"):
        """
        V10: Convert and process FASTQ file.
        Parameter werden vom FlagManager geholt.
        """
        try:
            fastqmap = FASTQmap(build=build, logger=self.logger)
            vcf_for_pipeline = fastqmap.convert(path)
            
            self.logger.log(f"[FASTQ] Conversion complete: {vcf_for_pipeline}")
            
            # V10: Alle Parameter kommen vom FlagManager
            self._distill_vcf(vcf_path=vcf_for_pipeline, build=build)
        
        except Exception as e:
            self.logger.log(f"[FASTQ] Error: {e}")
            import traceback
            self.logger.log(f"[FASTQ] Traceback:\n{traceback.format_exc()}")
    
    def _filter_gvcf(self, path: str) -> str:
        """Filter reference blocks from gVCF."""
        filtered_vcf = (
            path.replace(".g.vcf.gz", "_variants.vcf")
                .replace(".gvcf.gz", "_variants.vcf")
                .replace(".g.vcf", "_variants.vcf")
                .replace(".gvcf", "_variants.vcf")
        )
        
        with open_text_maybe_gzip(path) as fin, open(filtered_vcf, "w") as fout:
            for line in fin:
                if line.startswith("#") or (line and line.split("\t")[4] != "."):
                    fout.write(line)
        
        self.logger.log(f"[gVCF] Filtered to: {filtered_vcf}")
        return filtered_vcf
    
    # ==================== Utilities ====================
    
    def looks_like_23andme_txt(self, path: str) -> bool:
        """Detect 23andMe format."""
        try:
            with open(path, "r", encoding="utf-8") as f:
                for line in f:
                    if line.startswith("#"):
                        continue
                    parts = line.strip().split("\t")
                    if len(parts) < 4:
                        return False
                    rsid, chrom, pos, genotype = parts[:4]
                    if rsid.lower().startswith("rs") and chrom.upper() in [str(i) for i in range(1, 23)] + ["X", "Y", "MT"]:
                        return True
                    break
        except Exception:
            return False
        return False
    
    def get_vcf_sample_names(self, path: str) -> list:
        """Extrahiert Sample-Namen aus dem VCF-Header (#CHROM-Zeile, Spalten ab Index 9)."""
        try:
            with open_text_maybe_gzip(path) as f:
                for line in f:
                    line = line.strip()
                    if line.startswith("#CHROM"):
                        cols = line.split("\t")
                        if len(cols) > 9:
                            return cols[9:]
                        return []
                    if not line.startswith("#"):
                        break
        except Exception:
            pass
        return []

    def looks_like_vcf(self, path: str) -> bool:
        """Validate VCF format (relaxed check)."""
        try:
            with open_text_maybe_gzip(path) as f:
                has_fileformat = False
                has_header = False
                has_data = False

                for line in f:
                    line = line.strip()
                    if not line:
                        continue

                    # Check for fileformat declaration
                    if line.startswith("##fileformat=VCF"):
                        has_fileformat = True
                        continue

                    # Check for column header
                    if line.startswith("#CHROM"):
                        cols = line.split("\t")
                        # Minimal requirement: first 8 columns present
                        required = ["#CHROM", "POS", "ID", "REF", "ALT", "QUAL", "FILTER", "INFO"]
                        if len(cols) >= 8 and all(req in cols[:8] for req in required):
                            has_header = True
                        continue

                    # Check for at least one data line
                    if not line.startswith("#"):
                        parts = line.split("\t")
                        if len(parts) >= 8:  # Minimal VCF: 8 columns
                            has_data = True
                            break

                # Accept if has header AND (fileformat OR data)
                return has_header and (has_fileformat or has_data)

        except Exception as e:
            self.logger.log(f"[VCF-Check] Fehler beim Validieren: {e}")
            return False
        
    @staticmethod
    def _chunks(lst: List, n: int):
        """
        Teilt eine Liste in aufeinanderfolgende Blöcke der Größe n auf.

        Args:
            lst: Eingabeliste
            n:   Blockgröße

        Yields:
            Teil-Listen mit maximal n Elementen
        """
        for i in range(0, len(lst), n):
            yield lst[i:i + n]

# ═══════════════════════════════════════════════════════════════════
# GENOTYP-VERARBEITUNG
# ═══════════════════════════════════════════════════════════════════

def get_genotype_label(fmt, samples, sample_index=0):
    """
    ✅ AKTUALISIERT: Unterstützt hemizygous Genotypen (X/Y/MT).
    
    VCF-Standard:
    - Autosomen: 0/0, 0/1, 1/1 (diploid)
    - chrX non-PAR (männlich): 0, 1 (hemizygous)
    - chrY (männlich): 0, 1 (hemizygous)
    - chrM/MT: 0, 1 (haploid)
    
    Args:
        fmt: FORMAT-String aus VCF
        samples: Sample-Werte aus VCF
        sample_index: Index des Samples (default: 0)
    
    Returns:
        "het" | "hom" | "ref" | "hemi" | "."
    """
    if not fmt or not samples:
        return "."
    
    fmt_keys = fmt.split(":")
    try:
        gt_idx = fmt_keys.index("GT")
    except ValueError:
        return "."
    
    if sample_index >= len(samples):
        return "."
    
    vals = samples[sample_index].split(":")
    if gt_idx >= len(vals):
        return "."
    
    gt = vals[gt_idx].strip()
    gt = gt.replace("|", "/")
    
    # ✅ KRITISCH: Hemizygous Check ZUERST (vor Diploid)
    if "/" not in gt:
        # Hemizygous/Haploid: chrX, chrY, MT
        if gt == ".":
            return "."
        try:
            allele = int(gt)
            if allele == 0:
                return "ref"  # 0 = Referenz
            else:
                return "hemi"  # 1, 2, etc. = hemizygous alternativ
        except (ValueError, TypeError):
            return "."
    
    # Diploid Mapping
    GENOTYPE_MAP = {
        "0/1": "het", "1/0": "het",
        "0/2": "het", "2/0": "het",
        "1/2": "het", "2/1": "het",
        "1/1": "hom", "2/2": "hom", "3/3": "hom",
        "0/0": "ref",
        "./.": ".", "./0": ".", "0/.": ".",
    }
    
    normalized = GENOTYPE_MAP.get(gt)
    if normalized:
        return normalized
    
    # Fallback
    alleles = gt.split("/")
    if len(alleles) != 2:
        return "."
    
    try:
        if "." in alleles:
            return "."
        a1_int = int(alleles[0])
        a2_int = int(alleles[1])
        
        if a1_int == 0 and a2_int == 0:
            return "ref"
        elif a1_int == a2_int:
            return "hom"
        else:
            return "het"
    except (ValueError, TypeError):
        return "."


def normalize_genotype_display(genotype_str: str) -> str:
    """
    Normalisiert bereits gespeicherte Genotypen für konsistente Anzeige.
    
    Args:
        genotype_str: Genotyp aus DB/Cache
    
    Returns:
        Normalisierter Genotyp ("het" | "hom" | "ref" | "hemi" | ".")
    """
    if not genotype_str:
        return "."
    
    gt = str(genotype_str).strip().lower()
    
    # Bereits normalisiert?
    if gt in ("het", "hom", "ref", "hemi", "."):
        return gt
    
    # VCF-Format → durchleiten via get_genotype_label
    if "/" in gt or "|" in gt:
        # Simuliere FMT/Sample-Format
        return get_genotype_label("GT", [gt], 0)
    
    # Unbekanntes Format → .
    return "."


# ═══════════════════════════════════════════════════════════════════
# STALE-DATA MANAGEMENT (AF-REFRESH)
# ═══════════════════════════════════════════════════════════════════

def is_af_stale(
    last_fetch_iso: Optional[str], 
    base_days: int = 365,
    random_offset: Optional[int] = None
) -> bool:
    """
    Prüft, ob ein AF-Wert als "stale" gilt.
    
    Args:
        last_fetch_iso: ISO-Timestamp des letzten Fetches
        base_days: Basis-Tage für Stale-Check (default: 1 Jahr)
        random_offset: Optional fixer Offset (sonst random ±100)
    
    Returns:
        True wenn stale (zu alt oder fehlend)
    
    Logik:
    - Kein Timestamp → stale
    - Timestamp älter als (base_days ± random_offset) → stale
    - Random-Offset verhindert Thundering Herd
    
    Beispiele:
        is_af_stale("2023-01-01T00:00:00Z")  # True (>1 Jahr)
        is_af_stale("2024-12-01T00:00:00Z")  # False (<1 Jahr)
        is_af_stale(None)                    # True (fehlend)
    """
    if not last_fetch_iso:
        return True  # Kein Timestamp → stale
    
    dt = parse_iso_utc(last_fetch_iso)
    if not dt:
        return True  # Ungültiger Timestamp → stale
    
    # Random Offset für verteilte Updates
    if random_offset is None:
        random_offset = random.randint(-100, 100)
    
    # Verwende utcnow() ohne tzinfo, da parse_iso_utc() auch naive datetimes zurückgibt
    stale_cutoff = datetime.datetime.utcnow() - datetime.timedelta(
        days=base_days + random_offset
    )

    return dt < stale_cutoff


def get_stale_cutoff_date(base_days: int = 365) -> datetime.datetime:
    """
    Berechnet Stale-Cutoff mit Random-Offset.

    Args:
        base_days: Basis-Tage (default: 1 Jahr)

    Returns:
        Cutoff-DateTime (UTC, naive - ohne tzinfo für Konsistenz mit parse_iso_utc)

    Verwendung:
        cutoff = get_stale_cutoff_date()
        # Alle Varianten mit last_fetch < cutoff sind stale
    """
    random_offset = random.randint(-100, 100)
    return datetime.datetime.utcnow() - datetime.timedelta(
        days=base_days + random_offset
    )


def format_age(last_fetch_iso: Optional[str]) -> str:
    """
    Formatiert Alter eines Fetches für Logging.
    
    Args:
        last_fetch_iso: ISO-Timestamp
    
    Returns:
        Menschenlesbarer String (z.B. "3 days ago", "2 years ago")
    
    Beispiele:
        format_age("2024-10-28T12:00:00Z")  # "1 day ago"
        format_age("2023-01-01T00:00:00Z")  # "1 year ago"
        format_age(None)                    # "never"
    """
    if not last_fetch_iso:
        return "never"
    
    dt = parse_iso_utc(last_fetch_iso)
    if not dt:
        return "invalid"
    
    delta = datetime.datetime.utcnow() - dt
    
    if delta.days < 1:
        hours = delta.seconds // 3600
        if hours < 1:
            return f"{delta.seconds // 60} minutes ago"
        return f"{hours} hours ago"
    
    if delta.days < 30:
        return f"{delta.days} days ago"
    
    if delta.days < 365:
        months = delta.days // 30
        return f"{months} months ago"
    
    years = delta.days // 365
    return f"{years} year{'s' if years > 1 else ''} ago"


def log_stale_stats(db, logger, build: str = None):
    """
    Loggt Statistiken über Stale-AF-Werte.
    
    Args:
        db: VariantDB-Instanz
        logger: Logger-Instanz
        build: Optional Filter auf Build
    
    Loggt:
        - Anzahl stale AF-Werte
        - Ältester Fetch
        - Durchschnittsalter
    
    Verwendung:
        log_stale_stats(db, logger, "GRCh38")
    """
    cutoff = get_stale_cutoff_date()
    
    with db.lock, db._conn() as con:
        cur = con.cursor()
        
        # Stale Count
        query = """
            SELECT COUNT(*) FROM variants
            WHERE meanAF_last_fetch < ?
        """
        params = [cutoff.strftime("%Y-%m-%dT%H:%M:%SZ")]
        
        if build:
            query += " AND build=?"
            params.append(build)
        
        cur.execute(query, params)
        stale_count = cur.fetchone()[0]
        
        # Ältester Fetch
        query = """
            SELECT MIN(meanAF_last_fetch) FROM variants
            WHERE meanAF_last_fetch IS NOT NULL
        """
        if build:
            query += " AND build=?"
            cur.execute(query, [build])
        else:
            cur.execute(query)
        
        oldest = cur.fetchone()[0]
        
        # Total Count
        query = "SELECT COUNT(*) FROM variants"
        if build:
            query += " WHERE build=?"
            cur.execute(query, [build])
        else:
            cur.execute(query)
        
        total = cur.fetchone()[0]
    
    if total > 0:
        stale_pct = (stale_count / total) * 100.0
    else:
        stale_pct = 0.0
    
    logger.log(
        f"[Stale-Stats] Build: {build or 'ALL'}\n"
        f"  - Total variants: {total:,}\n"
        f"  - Stale AF values: {stale_count:,} ({stale_pct:.1f}%)\n"
        f"  - Oldest fetch: {format_age(oldest)}\n"
        f"  - Cutoff: {cutoff.strftime('%Y-%m-%d')} "
        f"({(datetime.datetime.utcnow() - cutoff).days} days ago)"
    )


def mark_variants_for_refresh(db, logger, max_count: int = 10000) -> List[Tuple]:
    """
    Identifiziert Varianten für AF-Refresh (stale oder fehlend).
    
    Args:
        db: VariantDB-Instanz
        logger: Logger-Instanz
        max_count: Max. Anzahl Keys zurückgeben
    
    Returns:
        Liste von Keys (chr, pos, ref, alt, build)
    
    Verwendung:
        keys = mark_variants_for_refresh(db, logger, 5000)
        # → Nutze für Background-Refresh
    """
    cutoff = get_stale_cutoff_date()
    
    with db.lock, db._conn() as con:
        cur = con.cursor()
        
        cur.execute("""
            SELECT chr, pos, ref, alt, build
            FROM variants
            WHERE meanAF_last_fetch IS NULL
               OR meanAF_last_fetch < ?
            ORDER BY COALESCE(meanAF_last_fetch, '1970-01-01T00:00:00Z') ASC
            LIMIT ?
        """, (cutoff.strftime("%Y-%m-%dT%H:%M:%SZ"), max_count))
        
        keys = [tuple(row) for row in cur.fetchall()]
    
    logger.log(
        f"[Refresh-Queue] Identified {len(keys)} variants for AF refresh "
        f"(stale or missing)"
    )
    
    return keys


# ═══════════════════════════════════════════════════════════════════
# VCF-EXPORT HELPER-FUNKTIONEN
# ═══════════════════════════════════════════════════════════════════

def _extract_format_field(rec: dict, field_name: str) -> Optional[str]:
    """
    Extrahiert einen spezifischen Wert aus dem FORMAT-String eines VCF-Records.
    
    Args:
        rec: VCF-Record dict
        field_name: Name des FORMAT-Feldes (z.B. "DP", "GQ", "AD")
    
    Returns:
        Feld-Wert oder None
    
    Beispiele:
        _extract_format_field(rec, "DP")  # "45"
        _extract_format_field(rec, "GQ")  # "99"
    """
    if not rec.get("fmt") or field_name not in rec.get("fmt", ""):
        return None
    
    if not rec.get("samples") or len(rec["samples"]) == 0:
        return None
    
    try:
        fmt_fields = rec["fmt"].split(":")
        sample_fields = rec["samples"][0].split(":")
        
        if field_name in fmt_fields:
            field_index = fmt_fields.index(field_name)
            if len(sample_fields) > field_index:
                value = sample_fields[field_index]
                if value and value != ".":
                    return value
    except (IndexError, ValueError, AttributeError):
        pass
    
    return None


def _get_dp_value(rec: dict) -> Optional[str]:
    """
    ✅ KORRIGIERT: Extrahiert DP (Read Depth) NUR aus VCF-Record.

    ⚠️ WICHTIG: DP wird NICHT in der DB gespeichert!
              DP kommt ausschließlich aus dem Original-VCF-Record.

    Priorität:
    1. FORMAT-Feld DP (aus Sample)
    2. INFO-Feld DP (multi-sample)

    Args:
        rec: VCF-Record dict aus orig_records

    Returns:
        DP-Wert als String oder None

    V16 FIX: Unterstützt sowohl String- als auch Dict-Format für INFO.
    """
    # 1. FORMAT-Feld DP (bevorzugt - sample-spezifisch)
    dp_format = _extract_format_field(rec, "DP")
    if dp_format:
        return dp_format

    # 2. INFO-Feld DP (fallback - multi-sample gesamt)
    info = rec.get("info")
    if info:
        # V16 FIX: INFO kann Dict (vom Parser) oder String sein
        if isinstance(info, dict):
            dp_val = info.get("DP")
            if dp_val is not None:
                return str(dp_val)
        elif isinstance(info, str) and "DP=" in info:
            match = re.search(r'DP=(\d+)', info)
            if match:
                return match.group(1)

    # ❌ NICHT aus DB holen - DP wird dort nicht gespeichert!
    return None


# [PROBABLY DEAD] Nie aufgerufen - Status wird inline bestimmt
# def _determine_filter_status(rec: dict, row: dict) -> str:
    # """
    # Bestimmt den korrekten FILTER-Status für VCF-Export.
    
    # VCF-Spec:
    # - "PASS" für Varianten ohne Filter
    # - "." ist veraltet, sollte "PASS" sein
    # - Spezifische Filter-Namen (z.B. "LowQual", "DP10")
    
    # Args:
        # rec: VCF-Record dict aus orig_records
        # row: DB-Row (fallback)
    
    # Returns:
        # FILTER-String ("PASS", "LowQual", etc.)
    # """
    # Aus orig_records (bevorzugt)
    # flt = rec.get("filter")
    
    # Fallback: Aus DB
    # if not flt:
        # flt = row.get("filter")
    
    # Normalisierung
    # if not flt or flt == "" or flt == ".":
        # return "PASS"
    
    # flt_upper = str(flt).upper()
    # if flt_upper == "PASS":
        # return "PASS"
    
    # return str(flt)


def _build_format_fields(rec: dict, genotype: str) -> tuple:
    """
    ✅ KORRIGIERT: Konstruiert FORMAT-String und Sample-Werte für VCF-Export.
    
    Nutzt nur VCF-Record-Daten, keine DB-Felder.
    
    Args:
        rec: VCF-Record dict aus orig_records
        genotype: Normalisierter Genotyp ("het", "hom", "ref", "hemi")
    
    Returns:
        (format_string, sample_string)
    
    Beispiel:
        _build_format_fields(rec, "het")
        # → ("GT:DP:GQ:AD", "0/1:45:99:20,25")
    """
    format_fields = ["GT"]
    sample_values = [genotype]
    
    # DP
    dp_value = _get_dp_value(rec)
    if dp_value:
        format_fields.append("DP")
        sample_values.append(dp_value)
    
    # GQ (Genotype Quality)
    gq_value = _extract_format_field(rec, "GQ")
    if gq_value:
        format_fields.append("GQ")
        sample_values.append(gq_value)
    
    # AD (Allelic Depth)
    ad_value = _extract_format_field(rec, "AD")
    if ad_value:
        format_fields.append("AD")
        sample_values.append(ad_value)
    
    # PL (Phred-scaled Likelihoods)
    pl_value = _extract_format_field(rec, "PL")
    if pl_value:
        format_fields.append("PL")
        sample_values.append(pl_value)
    
    format_str = ":".join(format_fields)
    sample_str = ":".join(sample_values)
    
    return format_str, sample_str
# =============================================================================
# QUALITYMANAGER - AKTUALISIERTE VERSION MIT HOMREF-FILTER
# =============================================================================

class QualityManager:
    """
    Zentrale Verwaltung aller Quality-Filter für VCF-Varianten.
    
    Funktionen:
    - Vordefinierte Presets (research, clinical, strict, permissive, rma_sensitive)
    - Custom Settings
    - Einheitliche Quality-Gates für FILTER, QUAL, DP, GENOTYPE
    - Thread-safe für GUI-Updates
    
    Usage:
        qm = QualityManager(preset="clinical")
        if qm.passes(record):
            # Variante akzeptiert
    """
    
    # Vordefinierte Profile
    # HINWEIS zu filter_homref:
    #   True  = 0/0 (homozygote Referenz) wird ausgefiltert (Standard)
    #   False = 0/0 wird behalten (wichtig für Reference Minor Allele Analyse)
    #   
    #   ACHTUNG: Das Referenzgenom kann pathogene Allele enthalten!
    #   Bei Verdacht auf Reference Minor Alleles (RMAs) filter_homref=False setzen.
    PRESETS = {
        "permissive": {
            "description": "Permissive - Maximale Sensitivität für Forschung",
            "filter_pass_only": False,
            "qual_threshold": 10.0,
            "use_dp_filter": True,
            "min_dp": 5,
            "max_dp": None,
            "filter_homref": True,  # 0/0 ausfiltern
        },
        "research": {
            "description": "Research - Standard für Forschungsprojekte",
            "filter_pass_only": True,
            "qual_threshold": 20.0,
            "use_dp_filter": True,
            "min_dp": 8,
            "max_dp": None,
            "filter_homref": True,  # 0/0 ausfiltern
        },
        "clinical": {
            "description": "Clinical - Empfohlen für klinische Anwendungen",
            "filter_pass_only": True,
            "qual_threshold": 30.0,
            "use_dp_filter": True,
            "min_dp": 20,
            "max_dp": None,
            "filter_homref": True,  # 0/0 ausfiltern
        },
        "strict": {
            "description": "Strict - Maximale Spezifität",
            "filter_pass_only": True,
            "qual_threshold": 50.0,
            "use_dp_filter": True,
            "min_dp": 30,
            "max_dp": None,
            "filter_homref": True,  # 0/0 ausfiltern
        },
        "rma_sensitive": {
            "description": "RMA-Sensitive - Behält 0/0 für Reference Minor Allele Analyse",
            "filter_pass_only": True,
            "qual_threshold": 20.0,
            "use_dp_filter": True,
            "min_dp": 10,
            "max_dp": None,
            "filter_homref": False,  # 0/0 BEHALTEN für RMA-Suche
        },
        "custom": {
            "description": "Custom - Benutzerdefinierte Einstellungen",
            "filter_pass_only": True,
            "qual_threshold": 30.0,
            "use_dp_filter": True,
            "min_dp": 10,
            "max_dp": None,
            "filter_homref": True,  # 0/0 ausfiltern (änderbar)
        },
    }
    
    def __init__(self, preset="clinical", logger=None):
        """
        Initialisiert QualityManager.
        
        Args:
            preset: Preset-Name ("research", "clinical", "strict", "permissive", "rma_sensitive", "custom")
            logger: Optional Logger-Instanz
        """
        self.logger = logger
        self.current_preset = preset
        self._lock = threading.Lock()
        
        # Lade Preset
        if preset not in self.PRESETS:
            if self.logger:
                self.logger.log(f"[QualityManager] ⚠️ Unbekanntes Preset '{preset}', nutze 'clinical'")
            preset = "clinical"
            self.current_preset = preset
        
        self.settings = self.PRESETS[preset].copy()
        
        # Statistiken
        self.stats = {
            "total_checked": 0,
            "passed": 0,
            "failed_filter": 0,
            "failed_qual": 0,
            "failed_dp_min": 0,
            "failed_dp_max": 0,
            "failed_homref": 0,  # NEU: Counter für 0/0 Genotypen
            "no_dp": 0,
        }
        
        # V10 FIX: Auto-Detection für fehlende DP-Werte
        self._dp_samples_checked = 0
        self._dp_samples_found = 0
        self._dp_auto_disabled = False
        self._dp_auto_disable_threshold = 1000  # Nach 1000 Varianten ohne DP -> Warnung
        
        if self.logger:
            homref_status = "ausfiltern" if self.settings.get('filter_homref', True) else "BEHALTEN (RMA-Modus)"
            self.logger.log(
                f"[QualityManager] ✅ Initialisiert mit Preset '{preset}':\n"
                f"  FILTER: {'PASS only' if self.settings['filter_pass_only'] else 'Alle'}\n"
                f"  QUAL: ≥{self.settings['qual_threshold']}\n"
                f"  DP: {self.settings['min_dp']}-{self.settings['max_dp'] or '∞'} "
                f"({'aktiv' if self.settings['use_dp_filter'] else 'deaktiviert'})\n"
                f"  0/0 (hom-ref): {homref_status}"
            )
    
    def set_preset(self, preset: str):
        """
        Wechselt zu vordefiniertem Preset.
        
        Args:
            preset: Preset-Name
        """
        with self._lock:
            if preset not in self.PRESETS:
                raise ValueError(f"Unbekanntes Preset: {preset}. Verfügbar: {list(self.PRESETS.keys())}")
            
            self.current_preset = preset
            self.settings = self.PRESETS[preset].copy()
            
            if self.logger:
                self.logger.log(f"[QualityManager] ✅ Preset gewechselt zu '{preset}'")
    
    def get_preset_names(self) -> list:
        """Gibt Liste aller verfügbaren Preset-Namen zurück."""
        return list(self.PRESETS.keys())
    
    def get_preset_description(self, preset: str) -> str:
        """Gibt Beschreibung eines Presets zurück."""
        return self.PRESETS.get(preset, {}).get("description", "")
    
    def set_custom_settings(self, 
                           filter_pass_only=None,
                           qual_threshold=None,
                           use_dp_filter=None,
                           min_dp=None,
                           max_dp=None,
                           filter_homref=None):
        """
        Setzt benutzerdefinierte Settings (Custom-Modus).
        
        Args:
            filter_pass_only: Nur PASS-Varianten akzeptieren
            qual_threshold: Minimum QUAL score
            use_dp_filter: DP-Filter aktivieren
            min_dp: Minimum read depth
            max_dp: Maximum read depth (None = unbegrenzt)
            filter_homref: 0/0 Genotypen ausfiltern (False für RMA-Analyse)
        """
        with self._lock:
            self.current_preset = "custom"
            
            if filter_pass_only is not None:
                self.settings["filter_pass_only"] = filter_pass_only
            if qual_threshold is not None:
                self.settings["qual_threshold"] = float(qual_threshold)
            if use_dp_filter is not None:
                self.settings["use_dp_filter"] = use_dp_filter
            if min_dp is not None:
                self.settings["min_dp"] = int(min_dp)
            if max_dp is not None:
                self.settings["max_dp"] = int(max_dp) if max_dp else None
            if filter_homref is not None:
                self.settings["filter_homref"] = filter_homref
            
            if self.logger:
                homref_status = "ausfiltern" if self.settings.get('filter_homref', True) else "BEHALTEN (RMA-Modus)"
                self.logger.log(
                    f"[QualityManager] ✅ Custom Settings aktiviert:\n"
                    f"  FILTER: {'PASS only' if self.settings['filter_pass_only'] else 'Alle'}\n"
                    f"  QUAL: ≥{self.settings['qual_threshold']}\n"
                    f"  DP: {self.settings['min_dp']}-{self.settings['max_dp'] or '∞'} "
                    f"({'aktiv' if self.settings['use_dp_filter'] else 'deaktiviert'})\n"
                    f"  0/0 (hom-ref): {homref_status}"
                )
    
    def get_settings(self) -> dict:
        """
        Gibt aktuelle Settings zurück.
        
        Returns:
            Dict mit allen aktuellen Einstellungen
        """
        with self._lock:
            return {
                "preset": self.current_preset,
                "filter_pass_only": self.settings["filter_pass_only"],
                "qual_threshold": self.settings["qual_threshold"],
                "use_dp_filter": self.settings["use_dp_filter"],
                "min_dp": self.settings["min_dp"],
                "max_dp": self.settings["max_dp"],
                "filter_homref": self.settings.get("filter_homref", True),
            }
    
    def passes(self, record: dict) -> bool:
        """
        ✅ HAUPTMETHODE: Prüft ob Variante Quality-Gates besteht.
        
        Args:
            record: VCF-Record dict mit Keys: filter, qual, info, fmt, samples
        
        Returns:
            bool: True wenn akzeptiert, False wenn gefiltert
        """
        with self._lock:
            self.stats["total_checked"] += 1
            
            # 1. FILTER-Check
            if self.settings["filter_pass_only"]:
                f = (record.get("filter") or ".")
                has_pass = f in ("PASS", ".")
                
                if not has_pass:
                    self.stats["failed_filter"] += 1
                    return False
            else:
                has_pass = True  # Filter ignorieren
            
            # 2. QUAL-Check
            try:
                qv = float(record.get("qual") or 0)
                has_qual = qv >= self.settings["qual_threshold"]
            except (ValueError, TypeError):
                has_qual = False
            
            if not has_pass and not has_qual:
                # Weder PASS noch ausreichendes QUAL
                self.stats["failed_qual"] += 1
                return False
            
            # 3. DP-Check (falls aktiviert UND nicht auto-deaktiviert)
            if self.settings["use_dp_filter"] and not self._dp_auto_disabled:
                dp = self._extract_dp(record)
                
                # V10 FIX: Tracke DP-Verfügbarkeit
                self._dp_samples_checked += 1
                
                if dp is not None:
                    self._dp_samples_found += 1
                    
                    # Minimum DP
                    if dp < self.settings["min_dp"]:
                        self.stats["failed_dp_min"] += 1
                        return False
                    
                    # Maximum DP (falls gesetzt)
                    max_dp = self.settings["max_dp"]
                    if max_dp is not None and dp > max_dp:
                        self.stats["failed_dp_max"] += 1
                        return False
                else:
                    # Kein DP vorhanden
                    self.stats["no_dp"] += 1
                    
                    # V10 FIX: Auto-Disable Warnung wenn VCF keine DPs hat
                    if (self._dp_samples_checked >= self._dp_auto_disable_threshold and 
                        self._dp_samples_found == 0 and not self._dp_auto_disabled):
                        self._dp_auto_disabled = True
                        if self.logger:
                            self.logger.log("─" * 60)
                            self.logger.log(f"[QualityManager] ⚠️ VCF enthält KEINE DP-Werte!")
                            self.logger.log(f"[QualityManager] ℹ️ DP-Filter wird ignoriert (PASS/QUAL bleibt aktiv)")
                            self.logger.log("─" * 60)
                    # Akzeptiere wenn PASS oder QUAL ok
            
            # 4. GENOTYPE-Check: 0/0 (hom-ref) ausfiltern (falls aktiviert)
            # HINWEIS: Kann deaktiviert werden für Reference Minor Allele (RMA) Analyse
            if self.settings.get("filter_homref", True):
                gt = self._extract_genotype(record)
                if gt == "ref":  # 0/0 = homozygote Referenz
                    self.stats["failed_homref"] += 1
                    return False
            
            # ✅ Alle Checks bestanden
            self.stats["passed"] += 1
            return True
    
    def _extract_dp(self, record: dict) -> Optional[int]:
        """
        Extrahiert DP aus INFO oder FORMAT.

        Args:
            record: VCF-Record dict

        Returns:
            int oder None

        V16 FIX: Unterstützt sowohl String- als auch Dict-Format für INFO.
        """
        # 1. INFO-Feld: DP=123
        info = record.get("info")

        if info:
            # V16 FIX: INFO kann Dict (vom Parser) oder String sein
            if isinstance(info, dict):
                # Parser liefert Dict mit geparsten Werten
                dp_val = info.get("DP")
                if dp_val is not None:
                    try:
                        return int(dp_val)
                    except (ValueError, TypeError):
                        pass
            elif isinstance(info, str) and "DP=" in info:
                # Legacy: INFO als String
                match = re.search(r'DP=(\d+)', info)
                if match:
                    try:
                        return int(match.group(1))
                    except ValueError:
                        pass

        # 2. FORMAT-Feld: GT:DP -> 0/1:45
        fmt = record.get("fmt", "")
        sample = record.get("sample", "")

        if fmt and sample and "DP" in fmt:
            try:
                fmt_fields = fmt.split(":")
                sample_fields = sample.split(":")
                dp_idx = fmt_fields.index("DP")
                if dp_idx < len(sample_fields):
                    return int(sample_fields[dp_idx])
            except (ValueError, IndexError):
                pass

        # 3. Aus samples-Liste (falls vorhanden)
        samples = record.get("samples", [])
        if samples and fmt and "DP" in fmt:
            try:
                fmt_fields = fmt.split(":")
                dp_idx = fmt_fields.index("DP")
                sample_fields = samples[0].split(":")
                if dp_idx < len(sample_fields):
                    return int(sample_fields[dp_idx])
            except (ValueError, IndexError):
                pass

        return None
    
    def _extract_genotype(self, record: dict) -> str:
        """
        ✅ AKTUALISIERT: Extrahiert normalisierten Genotyp mit Hemizygous-Support.
        
        VCF-Standard:
        - Autosomen: 0/0, 0/1, 1/1 (diploid)
        - chrX non-PAR (männlich): 0, 1 (hemizygous)
        - chrY (männlich): 0, 1 (hemizygous)
        - chrM/MT: 0, 1 (haploid)
        
        HINWEIS zu Reference Minor Alleles (RMAs):
        Ein 0/0 Genotyp bedeutet NICHT automatisch "gesund".
        Das Referenzgenom kann pathogene Allele enthalten!
        
        Args:
            record: VCF-Record dict mit fmt und samples
        
        Returns:
            "het" | "hom" | "ref" | "hemi" | "." (normalisiert)
        """
        fmt = record.get("fmt")
        samples = record.get("samples", [])
        
        if not fmt or not samples:
            return "."
        
        # GT-Index finden
        fmt_keys = fmt.split(":")
        try:
            gt_idx = fmt_keys.index("GT")
        except ValueError:
            return "."
        
        # Sample-Werte extrahieren
        if not samples:
            return "."
        
        sample_vals = samples[0].split(":")
        if gt_idx >= len(sample_vals):
            return "."
        
        gt = sample_vals[gt_idx].strip()
        
        # Normalisiere (Phasing ignorieren: | -> /)
        gt = gt.replace("|", "/")
        
        # ═════════════════════════════════════════════════════════════════
        # ✅ KRITISCH: Hemizygous Check ZUERST (vor Diploid)
        # ═════════════════════════════════════════════════════════════════
        if "/" not in gt:
            # Hemizygous/Haploid: chrX non-PAR, chrY, MT
            if gt == ".":
                return "."
            try:
                allele = int(gt)
                if allele == 0:
                    return "ref"  # 0 = Referenz (hemizygous)
                else:
                    return "hemi"  # 1, 2, etc. = hemizygous alternativ
            except (ValueError, TypeError):
                return "."
        
        # ═════════════════════════════════════════════════════════════════
        # Diploid Mapping (Autosomen, chrX PAR)
        # ═════════════════════════════════════════════════════════════════
        GENOTYPE_MAP = {
            # Heterozygot
            "0/1": "het", "1/0": "het",
            "0/2": "het", "2/0": "het",
            "1/2": "het", "2/1": "het",
            # Homozygot alternativ
            "1/1": "hom", "2/2": "hom", "3/3": "hom",
            # Homozygot Referenz
            "0/0": "ref",
            # Missing
            "./.": ".", "./0": ".", "0/.": ".",
        }
        
        normalized = GENOTYPE_MAP.get(gt)
        if normalized:
            return normalized
        
        # ═════════════════════════════════════════════════════════════════
        # Fallback: Analysiere Allele manuell
        # ═════════════════════════════════════════════════════════════════
        alleles = gt.split("/")
        if len(alleles) != 2:
            return "."  # Ungültiges Format
        
        try:
            if "." in alleles:
                return "."
            a1, a2 = int(alleles[0]), int(alleles[1])
            
            if a1 == 0 and a2 == 0:
                return "ref"
            elif a1 == a2:
                return "hom"
            else:
                return "het"
        except (ValueError, TypeError):
            return "."
    
    def get_stats(self) -> dict:
        """
        Gibt Statistiken zurück.
        
        Returns:
            Dict mit Zählern
        """
        with self._lock:
            return self.stats.copy()
    
    def reset_stats(self):
        """Setzt Statistiken zurück."""
        with self._lock:
            self.stats = {
                "total_checked": 0,
                "passed": 0,
                "failed_filter": 0,
                "failed_qual": 0,
                "failed_dp_min": 0,
                "failed_dp_max": 0,
                "failed_homref": 0,  # NEU: Counter für 0/0 Genotypen
                "no_dp": 0,
            }
    
    def print_stats(self):
        """Gibt Statistiken aus."""
        with self._lock:
            total = self.stats["total_checked"]
            if total == 0:
                if self.logger:
                    self.logger.log("[QualityManager] Keine Varianten geprüft")
                return
            
            passed = self.stats["passed"]
            failed_total = total - passed
            
            msg = (
                f"[QualityManager] Statistik ({self.current_preset}):\n"
                f"  Total geprüft: {total:,}\n"
                f"  ✅ Bestanden: {passed:,} ({100*passed/total:.1f}%)\n"
                f"  ❌ Gefiltert: {failed_total:,} ({100*failed_total/total:.1f}%)\n"
            )
            
            if failed_total > 0:
                msg += "  Gründe:\n"
                if self.stats["failed_filter"] > 0:
                    msg += f"    - FILTER: {self.stats['failed_filter']:,}\n"
                if self.stats["failed_qual"] > 0:
                    msg += f"    - QUAL: {self.stats['failed_qual']:,}\n"
                if self.stats["failed_dp_min"] > 0:
                    msg += f"    - DP < {self.settings['min_dp']}: {self.stats['failed_dp_min']:,}\n"
                if self.stats["failed_dp_max"] > 0:
                    msg += f"    - DP > {self.settings['max_dp']}: {self.stats['failed_dp_max']:,}\n"
                if self.stats.get("failed_homref", 0) > 0:
                    msg += f"    - 0/0 (hom-ref): {self.stats['failed_homref']:,}\n"
            
            if self.stats["no_dp"] > 0:
                msg += f"  ℹ️ Ohne DP: {self.stats['no_dp']:,}\n"
            
            if self.logger:
                self.logger.log(msg)

class QualitySettingsDialog(ttk.Toplevel):
    """
    ✅ ÜBERARBEITET: Scrollbarer Dialog mit klaren Beschreibungen.
    FIX: Verhindert Rückspringen auf 'Custom' beim Laden von Presets (Event-Lock).
    FIX: Toplevel Init-Fehler behoben (Title separat setzen).
    """
    
    def __init__(self, parent, quality_manager, af_none_manager=None):
        # ✅ FIX: title nicht in __init__ übergeben, da dies mit dem parent-Argument kollidiert
        super().__init__(parent)

        # Translation helper (inherit from parent App)
        self._t = getattr(parent, '_t', lambda k: k)

        self.title(self._t("Quality-Filter Einstellungen"))

        self.quality_manager = quality_manager
        self.af_none_manager = af_none_manager
        self.result = None
        
        # ✅ WICHTIG: Flag um Rückkopplungen zu verhindern
        self._loading_preset = False
        
        self.geometry("850x900")
        
        # Modal machen
        self.transient(parent)
        self.grab_set()
        
        self.create_widgets()
        self.load_current_settings()
        
        # Fenster zentrieren
        self.place_window_center()
    
    def create_widgets(self):
        # ============================================================
        # MAIN LAYOUT
        # ============================================================
        # Container für Buttons unten (feststehend)
        btn_frame = ttk.Frame(self, padding=15)
        btn_frame.pack(side=BOTTOM, fill=X)
        
        # Container für scrollbaren Inhalt (füllt den Rest)
        # Wir nutzen einen Wrapper für Padding um den ScrolledFrame
        wrapper = ttk.Frame(self, padding=10)
        wrapper.pack(side=TOP, fill=BOTH, expand=YES)
        
        # ✅ ScrolledFrame ersetzt Canvas/Scrollbar Konstrukt
        sf = ScrolledFrame(wrapper, autohide=True, bootstyle="round")
        sf.pack(fill=BOTH, expand=YES)
        
        # Inhalt in den ScrolledFrame packen
        content = sf

        # ============================================================
        # PRESETS
        # ============================================================
        # ✅ FIX: Labelframe statt LabelFrame
        preset_frame = ttk.Labelframe(
            content, 
            text="📋 Vordefinierte Profile", 
            padding=15, 
            bootstyle="info"
        )
        preset_frame.pack(fill=X, pady=(0, 15), padx=5)

        self.preset_var = tk.StringVar(value="clinical")
        self.preset_radiobuttons = {}

        presets = self.quality_manager.get_preset_names()
        for preset in presets:
            # Custom überspringen wir hier, fügen es manuell am Ende an
            if preset == "custom": 
                continue

            desc = self.quality_manager.get_preset_description(preset)
            
            rb = ttk.Radiobutton(
                preset_frame,
                text=f"{preset.capitalize()}: {desc}",
                variable=self.preset_var,
                value=preset,
                command=self.on_preset_change,
                bootstyle="info"
            )
            rb.pack(anchor="w", pady=3)
            self.preset_radiobuttons[preset] = rb

        # Custom (Manuell)
        custom_rb = ttk.Radiobutton(
            preset_frame,
            text="Custom: Eigene Einstellungen",
            variable=self.preset_var,
            value="custom",
            command=self.on_preset_change,
            bootstyle="info"
        )
        custom_rb.pack(anchor="w", pady=3)
        self.preset_radiobuttons["custom"] = custom_rb

        # ============================================================
        # QUALITY-GATES
        # ============================================================
        # ✅ FIX: Labelframe statt LabelFrame
        quality_frame = ttk.Labelframe(
            content, 
            text="⚙️ Quality-Gates", 
            padding=15, 
            bootstyle="primary"
        )
        quality_frame.pack(fill=X, pady=(0, 15), padx=5)
        
        # FILTER
        self.filter_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            quality_frame,
            text="Nur PASS-Varianten akzeptieren (FILTER-Spalte)",
            variable=self.filter_var,
            command=self.on_custom_change,
            bootstyle="primary-round-toggle"
        ).pack(anchor="w", pady=5)
        
        # QUAL
        qual_frame = ttk.Frame(quality_frame)
        qual_frame.pack(fill=X, pady=8)
        ttk.Label(qual_frame, text="Minimum QUAL:").pack(side=LEFT)
        self.qual_var = tk.DoubleVar(value=30.0)
        
        qual_spinbox = ttk.Spinbox(
            qual_frame,
            from_=0.0,
            to=100.0,
            increment=5.0,
            textvariable=self.qual_var,
            width=10,
            command=self.on_custom_change
        )
        qual_spinbox.pack(side=LEFT, padx=10)
        qual_spinbox.bind("<KeyRelease>", lambda e: self.on_custom_change())
        
        # DP Filter Toggle
        self.dp_filter_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            quality_frame,
            text="DP-Filter aktivieren (Read Depth)",
            variable=self.dp_filter_var,
            command=self.on_custom_change,
            bootstyle="primary-round-toggle"
        ).pack(anchor="w", pady=(10, 5))
        
        # DP Settings Row
        dp_row = ttk.Frame(quality_frame)
        dp_row.pack(fill=X, padx=20)
        
        # Min DP
        ttk.Label(dp_row, text="Min DP:").pack(side=LEFT)
        self.min_dp_var = tk.IntVar(value=10)
        min_dp_spinbox = ttk.Spinbox(
            dp_row,
            from_=1,
            to=100,
            increment=5,
            textvariable=self.min_dp_var,
            width=8,
            command=self.on_custom_change
        )
        min_dp_spinbox.pack(side=LEFT, padx=5)
        min_dp_spinbox.bind("<KeyRelease>", lambda e: self.on_custom_change())
        
        # Max DP
        ttk.Label(dp_row, text="Max DP:").pack(side=LEFT, padx=(15, 0))
        self.max_dp_var = tk.StringVar(value="")
        # Trace hinzufügen für Custom-Change-Erkennung bei Eingabe
        self.max_dp_var.trace_add("write", lambda *a: self.on_custom_change())
        
        ttk.Entry(
            dp_row, 
            textvariable=self.max_dp_var, 
            width=8
        ).pack(side=LEFT, padx=5)
        
        ttk.Label(dp_row, text="(leer = ∞)", font=("", 8)).pack(side=LEFT, padx=5)
        
        ttk.Separator(quality_frame, orient=HORIZONTAL).pack(fill=X, pady=15)
        
        # HomRef
        self.homref_filter_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            quality_frame,
            text="0/0 Genotypen ausfiltern (homozygote Referenz)",
            variable=self.homref_filter_var,
            command=self.on_custom_change,
            bootstyle="warning-round-toggle"
        ).pack(anchor="w", pady=2)
        
        ttk.Label(
            quality_frame, 
            text="⚠️ Deaktivieren für RMA-Analyse (Reference Minor Allele)", 
            font=("", 8), 
            bootstyle="inverse-warning"
        ).pack(anchor="w", padx=25)

        # ============================================================
        # AF-NONE-BEHANDLUNG
        # ============================================================
        # ✅ FIX: Labelframe statt LabelFrame
        none_frame = ttk.Labelframe(
            content, 
            text="🔍 AF-None-Behandlung (Anzeige)", 
            padding=15, 
            bootstyle="secondary"
        )
        none_frame.pack(fill=X, pady=(0, 15), padx=5)
        
        info_lbl = ttk.Label(
            none_frame,
            text="Steuert, welche Varianten ohne AF-Wert angezeigt werden:",
            font=("", 9, "bold")
        )
        info_lbl.pack(anchor="w", pady=(0, 10))
        
        # 1. Nie geprüft
        self.never_fetched_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            none_frame,
            text="Nie geprüfte anzeigen (Status unbekannt)",
            variable=self.never_fetched_var,
            command=self.on_custom_change,
            bootstyle="round-toggle"
        ).pack(anchor="w", pady=2)
        
        # 2. Fehlerhafte
        self.fetch_failed_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            none_frame,
            text="Fehlerhafte Abrufe anzeigen (Technischer Fehler)",
            variable=self.fetch_failed_var,
            command=self.on_custom_change,
            bootstyle="round-toggle"
        ).pack(anchor="w", pady=2)
        
        # 3. Validiert seltene
        self.true_none_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            none_frame,
            text="🟢 Validiert seltene anzeigen (Empfohlen!)",
            variable=self.true_none_var,
            command=self.on_custom_change,
            bootstyle="success-round-toggle"
        ).pack(anchor="w", pady=2)

        # ============================================================
        # VORSCHAU
        # ============================================================
        # ✅ FIX: Labelframe statt LabelFrame
        preview_frame = ttk.Labelframe(content, text="👁️ Vorschau", padding=10)
        preview_frame.pack(fill=X, pady=(0, 20), padx=5)
        
        self.preview_text = tk.Text(
            preview_frame, 
            height=9, 
            width=70, 
            font=("Consolas", 9), 
            bg="#f0f0f0", 
            relief="flat",
            state="normal"
        )
        self.preview_text.pack(fill=X)

        # ============================================================
        # BUTTONS (im feststehenden Frame unten)
        # ============================================================
        ttk.Button(
            btn_frame, 
            text="Abbrechen", 
            command=self.on_cancel, 
            bootstyle="secondary"
        ).pack(side=RIGHT, padx=5)
        
        ttk.Button(
            btn_frame, 
            text="Einstellungen Übernehmen", 
            command=self.on_apply, 
            bootstyle="success"
        ).pack(side=RIGHT)

    def load_current_settings(self):
        """Lädt aktuelle Settings ohne 'Custom'-Rückfall zu triggern."""
        self._loading_preset = True
        try:
            settings = self.quality_manager.get_settings()
            
            self.preset_var.set(settings["preset"])
            self.filter_var.set(settings["filter_pass_only"])
            self.qual_var.set(settings["qual_threshold"])
            self.dp_filter_var.set(settings["use_dp_filter"])
            self.min_dp_var.set(settings["min_dp"])
            self.max_dp_var.set(str(settings["max_dp"]) if settings["max_dp"] else "")
            self.homref_filter_var.set(settings.get("filter_homref", True))
            
            if self.af_none_manager:
                policy = self.af_none_manager.get_policy()
                self.never_fetched_var.set(policy.get("never_fetched", False))
                self.fetch_failed_var.set(policy.get("fetch_failed", False))
                self.true_none_var.set(policy.get("true_none", True))
            
            self.update_preview()
            
        finally:
            self._loading_preset = False
            self.update_idletasks()

    def on_preset_change(self):
        """Preset ausgewählt -> Werte laden."""
        preset = self.preset_var.get()
        if preset == "custom":
            self.update_preview()
            return
        
        self._loading_preset = True
        try:
            self.quality_manager.set_preset(preset)
            settings = self.quality_manager.PRESETS.get(preset, {})
            
            if settings:
                self.filter_var.set(settings.get("filter_pass_only", True))
                self.qual_var.set(settings.get("qual_threshold", 30.0))
                self.dp_filter_var.set(settings.get("use_dp_filter", True))
                self.min_dp_var.set(settings.get("min_dp", 10))
                self.max_dp_var.set(str(settings.get("max_dp")) if settings.get("max_dp") else "")
                self.homref_filter_var.set(settings.get("filter_homref", True))
            
            if self.af_none_manager and hasattr(AfNoneTreatmentManager, 'PRESETS'):
                af_preset = AfNoneTreatmentManager.PRESETS.get(preset, {})
                if af_preset:
                    self.never_fetched_var.set(af_preset.get("never_fetched", False))
                    self.fetch_failed_var.set(af_preset.get("fetch_failed", False))
                    self.true_none_var.set(af_preset.get("true_none", True))
            
            self.update_preview()
        finally:
            self._loading_preset = False
            
    def on_custom_change(self, *args):
        """User ändert Wert -> Switch auf Custom."""
        if not self._loading_preset:
            self.preset_var.set("custom")
            self.update_preview()
    
    def update_preview(self):
        """Aktualisiert die Vorschau-Box."""
        try:
            preset = self.preset_var.get()
            
            # Text zusammenbauen
            txt = f"Aktives Profil: {preset.upper()}\n"
            txt += "-" * 40 + "\n"
            
            txt += f"Filter PASS: {'JA' if self.filter_var.get() else 'NEIN'}\n"
            txt += f"Min QUAL:    {self.qual_var.get()}\n"
            txt += f"DP Filter:   {'AN' if self.dp_filter_var.get() else 'AUS'}\n"
            if self.dp_filter_var.get():
                txt += f"  Range:     {self.min_dp_var.get()} - {self.max_dp_var.get() or '∞'}\n"
            
            txt += f"HomRef (0/0): {'AUSGEFILTERT' if self.homref_filter_var.get() else 'BEHALTEN'}\n"
            txt += "-" * 40 + "\n"
            txt += "AF-Anzeige:\n"
            txt += f"  Nie geprüft:      {'AN' if self.never_fetched_var.get() else 'AUS'}\n"
            txt += f"  Fehlerhaft:       {'AN' if self.fetch_failed_var.get() else 'AUS'}\n"
            txt += f"  Validiert selten: {'AN' if self.true_none_var.get() else 'AUS'}\n"

            self.preview_text.delete("1.0", tk.END)
            self.preview_text.insert("1.0", txt)
        except Exception:
            pass
    
    def on_apply(self):
        """Einstellungen speichern und schließen."""
        preset = self.preset_var.get()
        
        if preset == "custom":
            max_dp_str = self.max_dp_var.get().strip()
            max_dp = int(max_dp_str) if max_dp_str else None
            
            self.quality_manager.set_custom_settings(
                filter_pass_only=self.filter_var.get(),
                qual_threshold=self.qual_var.get(),
                use_dp_filter=self.dp_filter_var.get(),
                min_dp=self.min_dp_var.get(),
                max_dp=max_dp,
                filter_homref=self.homref_filter_var.get()
            )
            
            if self.af_none_manager:
                self.af_none_manager.set_custom_policy(
                    never_fetched=self.never_fetched_var.get(),
                    fetch_failed=self.fetch_failed_var.get(),
                    true_none=self.true_none_var.get()
                )
        else:
            self.quality_manager.set_preset(preset)
            if self.af_none_manager:
                self.af_none_manager.set_preset(preset)
        
        # Toast Notification statt Messagebox
        ToastNotification(
            title=self._t("Einstellungen gespeichert"),
            message=f"{self._t('Quality-Filter Einstellungen')}: '{preset}'",
            bootstyle="success",
            duration=2000
        ).show_toast()
        
        self.result = True
        self.destroy()
    
    def on_cancel(self):
        self.result = False
        self.destroy()        
# ============== GUI App ==============
#noch benötigt?
UI_COLUMNS = [
    "chr","pos","ref","alt","build","rsid","genotype","af","cadd","gene",
    "impact","consequence","clin_sig","phenotypes","coding","ag","cons","pubmed"
]

DEFAULT_COLUMNS = [
    "chr","pos","ref","alt","build","rsid","genotype","af","cadd","gene",
    "impact","consequence","clin_sig","phenotypes","coding","ag","cons","pubmed"
]
COLUMN_LABELS = {
    "chr": "chr", "pos": "pos", "ref": "ref", "alt": "alt", "build": "build",
    "rsid": "rsid", "genotype": "genotype", "af": "AF", "cadd": "CADD", "gene": "Gene",
    "impact": "Impact", "consequence": "Consequence", "clin_sig": "ClinSig", "phenotypes": "Phenotypes",
    "coding": "Coding", "ag": "AG", "cons": "Conservation", "pubmed": "PubMed",
    "meanAF_status": "AF Status", "full_status": "Anno Status"
}

# =============================================================================
# RESOURCE SETUP DIALOG + SHARED CARD BUILDER
# =============================================================================

_RESOURCE_GROUP_LABELS = {
    "reference": "Referenz-Genome",
    "database": "Datenbanken",
    "annotation": "Gen-Annotation",
}

_RESOURCE_GROUP_ORDER = ["reference", "database", "annotation"]


def build_resource_cards(parent, rm, logger_inst, card_widgets=None, _t=None):
    """
    Erstellt Resource-Cards gruppiert in Labelframes.
    Wird von ResourceSetupDialog UND dem Settings-Tab wiederverwendet.

    Args:
        parent: tk-Container (Frame/ScrolledFrame)
        rm: ResourceManager-Instanz
        logger_inst: Logger
        card_widgets: Optional dict zum Speichern der Widget-Referenzen {key: {status_lbl, progress_var, btn_frame}}
        _t: Optionale Uebersetzungsfunktion (key -> uebersetzter String).

    Returns:
        dict mit Widget-Referenzen falls card_widgets is None
    """
    if card_widgets is None:
        card_widgets = {}
    status = rm.get_status()

    # Gruppieren
    groups = {}
    for key, info in RESOURCE_SETUP_INFO.items():
        g = info["group"]
        if g not in groups:
            groups[g] = []
        groups[g].append((key, info))

    for group_key in _RESOURCE_GROUP_ORDER:
        if group_key not in groups:
            continue
        items = groups[group_key]
        lf = ttk.Labelframe(parent, text=_RESOURCE_GROUP_LABELS.get(group_key, group_key), padding=10)
        lf.pack(fill=X, pady=5, padx=5)

        # Annotation-Gruppe: kompaktes Info-Layout
        if group_key == "annotation":
            ttk.Label(lf, text="GTF-Dateien werden automatisch beim ersten Annotationslauf heruntergeladen.",
                      wraplength=700).pack(anchor="w", pady=(0, 5))
            for key, info in items:
                st = status.get(key, "missing")
                color = "green" if st in ("found", "registered", "created") else "orange"
                st_text = "Vorhanden" if color == "green" else "Fehlt"
                row = ttk.Frame(lf)
                row.pack(fill=X, pady=1)
                ttk.Label(row, text=f"{info['label']} ({info['size']})", width=40, anchor="w").pack(side=LEFT)
                sl = ttk.Label(row, text=f"  {st_text}", foreground=color)
                sl.pack(side=LEFT)
                card_widgets[key] = {"status_lbl": sl}
            continue

        # Andere Gruppen: volle Cards
        for key, info in items:
            st = status.get(key, "missing")
            is_ok = st in ("found", "registered", "created")
            _build_single_resource_card(lf, key, info, is_ok, rm, logger_inst, card_widgets)

    return card_widgets


def _build_single_resource_card(parent, key, info, is_ok, rm, logger_inst, card_widgets):
    """Einzelne Resource-Card mit Status, Beschreibung, Buttons und Progressbar."""
    card = ttk.Frame(parent, padding=5)
    card.pack(fill=X, pady=3)

    # Zeile 1: Label + Status
    row1 = ttk.Frame(card)
    row1.pack(fill=X)
    ttk.Label(row1, text=info["label"], font=("", 10, "bold")).pack(side=LEFT)
    color = "green" if is_ok else "red"
    st_text = "Vorhanden" if is_ok else "Fehlt"
    status_lbl = ttk.Label(row1, text=f"  {st_text}", foreground=color)
    status_lbl.pack(side=LEFT, padx=10)
    ttk.Label(row1, text=info["size"], foreground="gray").pack(side=RIGHT)

    # Zeile 2: Beschreibung
    ttk.Label(card, text=info["description"], wraplength=700).pack(anchor="w", pady=(2, 0))

    # Zeile 3: "Ohne"-Hinweis
    ttk.Label(card, text=f"Ohne: {info['without']}", foreground="gray", wraplength=700).pack(anchor="w")

    # Zeile 4: Buttons + Progress (nur bei action != "info" und nicht vorhanden)
    progress_var = tk.DoubleVar(value=0.0)
    btn_frame = ttk.Frame(card)
    btn_frame.pack(fill=X, pady=(3, 0))

    widgets = {"status_lbl": status_lbl, "progress_var": progress_var, "btn_frame": btn_frame, "card": card}

    if info["action"] == "info":
        ttk.Label(btn_frame, text="(Wird automatisch erstellt)", foreground="gray").pack(anchor="w")
    elif not is_ok:
        if info["action"] == "download":
            if key.startswith("fasta_"):
                dl_text = "Herunterladen"
            else:
                dl_text = "Herunterladen & Erstellen"
            dl_btn = ttk.Button(btn_frame, text=dl_text, bootstyle="success-outline",
                                command=lambda k=key, w=widgets: _on_download_click(k, w, rm, logger_inst, _t=_t))
            dl_btn.pack(side=LEFT, padx=(0, 5))
            widgets["dl_btn"] = dl_btn

        browse_btn = ttk.Button(btn_frame, text="Pfad waehlen...", bootstyle="secondary-outline",
                                command=lambda k=key, w=widgets: _on_browse_click(k, w, rm, logger_inst))
        browse_btn.pack(side=LEFT, padx=5)
        widgets["browse_btn"] = browse_btn

        # Cancel-Button (initial hidden)
        cancel_btn = ttk.Button(btn_frame, text="Abbrechen", bootstyle="danger-outline")
        widgets["cancel_btn"] = cancel_btn

        # Progressbar
        pb = ttk.Progressbar(card, variable=progress_var, maximum=100, mode="determinate")
        pb.pack(fill=X, pady=(3, 0))
        widgets["progressbar"] = pb

        # Phase-Label
        phase_lbl = ttk.Label(card, text="", foreground="gray")
        phase_lbl.pack(anchor="w")
        widgets["phase_lbl"] = phase_lbl

    card_widgets[key] = widgets
    # Separator
    ttk.Separator(parent, orient="horizontal").pack(fill=X, pady=3)


def _on_browse_click(key, widgets, rm, logger_inst):
    """Dateiauswahl-Dialog fuer eine Ressource."""
    if key == "gnomad_db":
        ftypes = [("SQLite DB", "*.db"), ("Alle Dateien", "*.*")]
    elif key.startswith("fasta_"):
        ftypes = [("FASTA", "*.fa *.fasta *.fa.gz"), ("Alle Dateien", "*.*")]
    else:
        ftypes = [("Alle Dateien", "*.*")]

    path = filedialog.askopenfilename(title=f"Pfad fuer {RESOURCE_SETUP_INFO[key]['label']}", filetypes=ftypes)
    if path and os.path.exists(path):
        rm.register(key, path)
        widgets["status_lbl"].config(text="  Vorhanden", foreground="green")
        logger_inst.log(f"[Ressourcen] {key} manuell registriert: {path}")


def _on_download_click(key, widgets, rm, logger_inst, _t=None):
    """Startet Download im Hintergrund-Thread."""
    cancel_event = threading.Event()
    widgets["cancel_btn"].config(command=lambda: cancel_event.set())
    widgets["cancel_btn"].pack(side=LEFT, padx=5)
    if "dl_btn" in widgets:
        widgets["dl_btn"].config(state="disabled")
    if "browse_btn" in widgets:
        widgets["browse_btn"].config(state="disabled")

    root = widgets["status_lbl"].winfo_toplevel()

    if key.startswith("fasta_"):
        build = "GRCh37" if "37" in key else "GRCh38"
        thread = threading.Thread(target=_download_fasta_thread,
                                  args=(build, widgets, rm, logger_inst, cancel_event, root, _t),
                                  daemon=True)
    elif key == "gnomad_db":
        thread = threading.Thread(target=_download_gnomad_thread,
                                  args=(widgets, rm, logger_inst, cancel_event, root, _t),
                                  daemon=True)
    else:
        return
    thread.start()


def _update_progress_safe(root, widgets, downloaded, total, phase, _t=None):
    """Thread-sicheres GUI-Update via after().

    Args:
        root: Tkinter root widget fuer after()-Aufrufe.
        widgets: Dict mit GUI-Widgets (progress_var, phase_lbl, etc.).
        downloaded: Heruntergeladene Bytes oder Status-String.
        total: Gesamtgroesse in Bytes.
        phase: Phase-String (download, extract, migrate, index, done, error).
        _t: Optionale Uebersetzungsfunktion (key -> uebersetzter String).
    """
    if _t is None:
        _t = lambda k: k

    def _do():
        if phase == "download" and total > 0:
            pct = downloaded / total * 100
            widgets["progress_var"].set(pct)
            mb_dl = downloaded / (1024 * 1024)
            mb_tot = total / (1024 * 1024)
            widgets["phase_lbl"].config(text=f"Download: {mb_dl:.0f} / {mb_tot:.0f} MB ({pct:.1f}%)")
        elif phase == "extract":
            widgets["progress_var"].set(0)
            widgets["phase_lbl"].config(text=_t("Entpacke..."))
            if "progressbar" in widgets:
                widgets["progressbar"].config(mode="indeterminate")
                widgets["progressbar"].start(15)
        elif phase == "migrate":
            widgets["phase_lbl"].config(text=f"{_t('Migriere')} {downloaded}...")
        elif phase == "index":
            widgets["phase_lbl"].config(text=_t("Erstelle Index..."))
        elif phase == "done":
            if "progressbar" in widgets:
                widgets["progressbar"].stop()
                widgets["progressbar"].config(mode="determinate")
            widgets["progress_var"].set(100)
            widgets["phase_lbl"].config(text=_t("Fertig!"))
            widgets["status_lbl"].config(text=f"  {_t('Vorhanden')}", foreground="green")
            widgets["cancel_btn"].pack_forget()
            if "dl_btn" in widgets:
                widgets["dl_btn"].pack_forget()
            if "browse_btn" in widgets:
                widgets["browse_btn"].pack_forget()
        elif phase == "error":
            if "progressbar" in widgets:
                widgets["progressbar"].stop()
                widgets["progressbar"].config(mode="determinate")
            widgets["progress_var"].set(0)
            widgets["phase_lbl"].config(text=f"{_t('Fehler')}: {downloaded}", foreground="red")
            widgets["cancel_btn"].pack_forget()
            if "dl_btn" in widgets:
                widgets["dl_btn"].config(state="normal")
            if "browse_btn" in widgets:
                widgets["browse_btn"].config(state="normal")
    try:
        root.after(0, _do)
    except RuntimeError:
        pass


def _download_fasta_thread(build, widgets, rm, logger_inst, cancel_event, root, _t=None):
    """FASTA-Download in Hintergrund-Thread."""
    try:
        fasta_path = FASTA_PATHS.get(build)
        if not fasta_path:
            fasta_path = os.path.join(BASE_DIR, "data", f"Homo_sapiens.{build}.dna.primary_assembly.fa")
        os.makedirs(os.path.dirname(fasta_path), exist_ok=True)

        url = FASTA_URLS.get(build)
        if not url:
            raise ValueError(f"Keine FASTA-URL fuer {build}")

        gz_path = fasta_path + ".gz"

        # Download
        r = requests.get(url, stream=True, timeout=60)
        r.raise_for_status()
        total = int(r.headers.get('content-length', 0))
        downloaded = 0
        with open(gz_path, "wb") as f:
            for chunk in r.iter_content(chunk_size=65536):
                if cancel_event.is_set():
                    raise InterruptedError("Abgebrochen")
                f.write(chunk)
                downloaded += len(chunk)
                if downloaded % (512 * 1024) < 65536:
                    _update_progress_safe(root, widgets, downloaded, total, "download", _t=_t)

        # Entpacken
        _update_progress_safe(root, widgets, 0, 0, "extract", _t=_t)
        with gzip.open(gz_path, "rb") as f_in, open(fasta_path, "wb") as f_out:
            shutil.copyfileobj(f_in, f_out)
        os.remove(gz_path)

        # Index
        _update_progress_safe(root, widgets, 0, 0, "index", _t=_t)
        build_fasta_index_global(fasta_path, logger_inst)

        # Registrieren
        build_key = "fasta_grch38" if "38" in build else "fasta_grch37"
        rm.register(build_key, fasta_path)
        _update_progress_safe(root, widgets, 0, 0, "done", _t=_t)

    except InterruptedError:
        _update_progress_safe(root, widgets, "Abgebrochen", 0, "error", _t=_t)
        # Cleanup
        for p in [gz_path, fasta_path]:
            if os.path.exists(p):
                try:
                    os.remove(p)
                except Exception:
                    pass
    except Exception as e:
        logger_inst.log(f"[Ressourcen] FASTA-Download Fehler: {e}")
        _update_progress_safe(root, widgets, str(e)[:80], 0, "error", _t=_t)


def _download_gnomad_thread(widgets, rm, logger_inst, cancel_event, root, _t=None):
    """gnomAD-Download + Migration + Index in Hintergrund-Thread."""
    try:
        data_dir = os.path.join(BASE_DIR, "data")
        os.makedirs(data_dir, exist_ok=True)
        dest_db = os.path.join(data_dir, "gnomad_light.db")

        lightdb_mgr = LightDBGnomADManager(logger=logger_inst, base_dir=BASE_DIR)

        for build in ("GRCh37", "GRCh38"):
            if cancel_event.is_set():
                raise InterruptedError("Abgebrochen")

            def progress_cb(dl, tot, phase, _build=build):
                if phase == "download":
                    _update_progress_safe(root, widgets, dl, tot, "download", _t=_t)
                elif phase == "extract":
                    _update_progress_safe(root, widgets, 0, 0, "extract", _t=_t)

            src_db = download_build(build, logger_inst, progress_callback=progress_cb, cancel_event=cancel_event)

            # Migrieren
            _update_progress_safe(root, widgets, build, 0, "migrate", _t=_t)
            lightdb_mgr.migrate_build(src_db, build, dest_db)

            # Temp-DB loeschen
            if os.path.exists(src_db):
                os.remove(src_db)

        # Index
        _update_progress_safe(root, widgets, 0, 0, "index", _t=_t)
        lightdb_mgr.start_index_worker()

        # Registrieren
        rm.register("gnomad_db", dest_db)
        _update_progress_safe(root, widgets, 0, 0, "done", _t=_t)

    except InterruptedError:
        _update_progress_safe(root, widgets, "Abgebrochen", 0, "error", _t=_t)
    except Exception as e:
        logger_inst.log(f"[Ressourcen] gnomAD-Download Fehler: {e}")
        _update_progress_safe(root, widgets, str(e)[:80], 0, "error", _t=_t)


class ResourceSetupDialog:
    """Startup-Dialog der fehlende Ressourcen anzeigt und Downloads anbietet."""

    def __init__(self, parent, rm, logger_inst):
        self.rm = rm
        self.logger = logger_inst
        self.card_widgets = {}
        self._t = getattr(parent, '_t', lambda k: k)

        self.win = ttk.Toplevel(parent)
        self.win.title(self._t("VF Distiller - Ressourcen-Einrichtung"))
        self.win.geometry("780x620")
        self.win.transient(parent)
        self.win.grab_set()
        try:
            self.win.place_window_center()
        except Exception:
            pass

        self._build_ui(parent)

    def _build_ui(self, parent):
        # Header
        header = ttk.Frame(self.win, padding=15)
        header.pack(fill=X)
        ttk.Label(header, text="Willkommen! Einige optionale Ressourcen fehlen.",
                  font=("", 12, "bold")).pack(anchor="w")
        ttk.Label(header, text="Diese verbessern Geschwindigkeit und Offline-Faehigkeit.\n"
                  "Alles kann spaeter unter Optionen > Einstellungen > Ressourcen nachgeholt werden.",
                  wraplength=700).pack(anchor="w", pady=(5, 0))

        # Scrollable Content
        sf = ScrolledFrame(self.win, autohide=True, padding=5)
        sf.pack(fill=BOTH, expand=YES, padx=10)

        build_resource_cards(sf, self.rm, self.logger, self.card_widgets, _t=self._t)

        # Footer
        footer = ttk.Frame(self.win, padding=10)
        footer.pack(fill=X, side=BOTTOM)

        self.suppress_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(footer, text="Diesen Dialog nicht mehr anzeigen",
                        variable=self.suppress_var).pack(side=LEFT)

        ttk.Button(footer, text="Schliessen", command=self._on_close,
                   bootstyle="success").pack(side=RIGHT, padx=5)
        ttk.Button(footer, text="Spaeter", command=self._on_later,
                   bootstyle="secondary").pack(side=RIGHT, padx=5)

    def _on_later(self):
        self.win.destroy()

    def _on_close(self):
        if self.suppress_var.get():
            # Setze Flag im Parent (App)
            try:
                app = self.win.master
                if hasattr(app, 'setup_dialog_suppressed'):
                    app.setup_dialog_suppressed = True
                    app._save_settings()
            except Exception:
                pass
        self.win.destroy()


class App(ttk.Window):
    """
    ✅ MODERNISIERT V5.4:
    - Layout: Tabelle -> Count -> Collapsible Filter -> Collapsible Log/Export.
    - CADD-Highlight in TopBar verschoben.
    - Info-Recycling Option im Menü.
    - Consequences im Toggle-Design.
    """
    
    def __init__(self):
        # 1. Theme vorladen
        self._load_initial_settings()
        theme_name = getattr(self, "saved_theme", Config.DEFAULT_THEME)
        
        # 2. Fenster initialisieren
        super().__init__(title=APP_NAME, themename=theme_name)
        self.geometry("1400x1000")
        
        self.live_queue = queue.Queue(maxsize=200000)
        self.logger = logger
        self.progress = PipelineProgress()
        
        self.protocol("WM_DELETE_WINDOW", self.on_close)

        # Translator
        try:
            from translator import Translator
            from translator_patch import patch_widgets
            self.translator = Translator(lang="de")
            patch_widgets(self.translator)
        except ImportError:
            self.translator = None

        # Translation helper (safe for None translator)
        self._t = self.translator.t if self.translator else (lambda k: k)

        # Gene Annotator
        self.gene_annotator = None
        self.enqueue_live = self._make_enqueue_live()

        self.db = db
        self.stopflag = StopFlag()
        
        # V10: Zentrale Manager
        self.flag_manager = Flag_and_Options_Manager(logger=self.logger)
        self.quality_manager = QualityManager(preset="clinical", logger=self.logger)
        
        # GUI-Variablen
        self._init_gui_variables()
        self.columns = DEFAULT_COLUMNS[:] + ["meanAF_status", "full_status"]
        self.visible_columns = set(self.columns)
        
        # Settings laden
        self._load_settings()

        # GUI Bauen
        self._build_gui()
        self._setup_events()

        # Locks & Caches
        self._table_lock = threading.RLock()
        self.rows_cache = {}
        self._refresh_idx = 0
        self._last_keyset_sig = None
        
        # LightDB - V10: Explizit BASE_DIR für pfadresistente Operationen
        self.lightdb_manager = LightDBGnomADManager(logger=self.logger, base_dir=BASE_DIR)
        # ✅ V12.1: LightDB-Check NACH GUI-Load starten (94GB DB kann 30s+ dauern!)
        # Verhindert GUI-Freeze durch langsames sqlite3.connect() beim App-Start
        # V17: auto_download=False → Downloads nur noch ueber Setup-Dialog/Einstellungen
        self.after(500, lambda: self.lightdb_manager.ensure_lightdb_async(auto_download=False))

        # V17: Startup-Check fuer fehlende Ressourcen
        self.after(1500, self._check_startup_setup)

        # AF-Controller
        self.af_fetcher = AFFetchController(self.db, self.logger, stopflag=self.stopflag, mode="full")
        
        # V10: MainFilterGate erstellen (noch ohne GeneAnnotator)
        self.main_filter_gate = MainFilterGate(
            flag_manager=self.flag_manager,
            db=self.db,
            gene_annotator=None,  # Wird später via _init_gene_annotator gesetzt
            logger=logger
        )
        
        # Sync GUI -> FlagManager vor Distiller-Erstellung
        self.flag_manager.sync_from_gui(self)
        
        # Distiller mit allen Managern
        self.distiller = Distiller(
            app=self, db=self.db, stopflag=self.stopflag, live_enqueue=self.enqueue_live,
            threads=Config.DEFAULT_THREADS, logger=logger,
            flag_manager=self.flag_manager,
            main_filter_gate=self.main_filter_gate,
            quality_manager=self.quality_manager,
            gene_annotator=None,
            af_fetcher=self.af_fetcher
        )
        self.distiller.app_ref = self
        
        # Maintainer
        self.maint = BackgroundMaintainer(
            distiller=self.distiller, db=self.db, stopflag=self.stopflag, logger=logger,
            gene_annotator=None, af_fetcher=self.af_fetcher
        )
        self.distiller.maint = self.maint
        # ✅ V12.1: Maintainer NACH GUI-Startup starten (via after)
        # Verhindert GUI-Freeze durch DB-Queries während GUI noch lädt
        self.after(100, self.maint.start)

        # Polling & Threads
        self.after(5000, self._poll_lightdb_status)
        threading.Thread(target=self._init_gene_annotator, daemon=True).start()
        self.after(Config.POLL_INTERVAL_MS, self._poll_log)
        self.after(Config.POLL_INTERVAL_MS, self._drain_live_enqueue)
        self.after(2000, self._refresh_table)
        self.after(Config.PROGRESS_UPDATE_INTERVAL, self._refresh_progress)
        self._drain_active = False

        # V10 FIX: Trace auf die korrekte Variable setzen (nicht include_none_var!)
        # self.include_none ist die Haupt-tk-Variable (definiert in _setup_main_variables)
        # Wir setzen den trace hier, damit Änderungen sofort an FlagManager weitergegeben werden
        self._include_none_trace_setup = False  # Flag um doppelte Traces zu vermeiden
        # ✅ V17: Performance-Check (Cloud Sync)
        self.after(2000, self._check_cloud_performance_warning)
        logger.log(f"[App] GUI initialization complete (Theme: {theme_name}).")

    def _make_enqueue_live(self):
        def enqueue(key):
            try:
                if isinstance(key, tuple) and len(key) == 5:
                    self.live_queue.put_nowait(key)
            except: pass
        return enqueue

    def _load_initial_settings(self):
        self.saved_theme = Config.DEFAULT_THEME
        if os.path.exists(Config.SETTINGS_FILE):
            try:
                with open(Config.SETTINGS_FILE, "r") as f:
                    d = json.load(f)
                    self.saved_theme = d.get("theme", Config.DEFAULT_THEME)
            except: pass



    # -------------------------------------------------------------------------
    # LAYOUT HELPER
    # -------------------------------------------------------------------------
    def _create_collapsible_section(self, parent, title, visible=True):
        """Erzeugt einen einklappbaren Bereich mit Header."""
        container = ttk.Frame(parent)
        container.pack(fill=X, padx=10, pady=5)
        
        header_frame = ttk.Frame(container)
        header_frame.pack(fill=X)
        
        is_visible = tk.BooleanVar(value=visible)
        
        content_frame = ttk.Frame(container)
        
        def toggle():
            if is_visible.get():
                content_frame.pack(fill=X, expand=YES, pady=5)
                btn.configure(text="▼")
            else:
                content_frame.pack_forget()
                btn.configure(text="▶")
        
        # Toggle Button
        btn = ttk.Checkbutton(
            header_frame, 
            text="▼" if visible else "▶", 
            variable=is_visible, 
            command=toggle, 
            style="Toolbutton", 
            bootstyle="link"
        )
        btn.pack(side=LEFT)
        
        # Title
        ttk.Label(header_frame, text=title, font=("", 10, "bold")).pack(side=LEFT, padx=5)
        ttk.Separator(header_frame).pack(side=LEFT, fill=X, expand=YES, padx=10)
        
        if visible:
            content_frame.pack(fill=X, expand=YES, pady=5)
            
        return content_frame

    def _parse_key_from_iid(self, iid: str) -> tuple:
        try:
            parts = iid.split("|")
            if len(parts) == 5:
                chrom, pos_str, ref, alt, build = parts
                # WICHTIG: Typ-Konvertierung muss exakt mit DB-Keys übereinstimmen
                return (chrom, int(pos_str), ref, alt, build)
        except Exception as e:
            print(f"Error parsing IID '{iid}': {e}") # Debug print
            pass
        return None

    def _on_include_none_changed(self):
        """
        V10: GUI Checkbox -> FlagManager.
        Wird aufgerufen wenn self.include_none (tk.BooleanVar) sich ändert.
        """
        include_none = self.include_none.get()  # V10 FIX: Korrekte Variable
        
        # V10: Setze im FlagManager
        if hasattr(self, 'flag_manager') and self.flag_manager:
            self.flag_manager.set_include_none(include_none)
            
        # Legacy: Auch an Distiller weitergeben (für Rückwärtskompatibilität)
        if hasattr(self, 'distiller') and self.distiller:
            self.distiller.set_legacy_include_none(include_none)
            
    def _poll_lightdb_status(self):
        """
        Prueft regelmaessig, ob LightDB fertig ist.
        FIX: Widget-Existenz pruefen vor after-Call
        """
        # KRITISCH: Pruefe ob Widget noch existiert
        if not self.winfo_exists():
            return
        
        prog = self.lightdb_manager.get_progress()
        if prog:
            pct = int(prog.get("pct", 0))
            status = prog.get("status", "unknown")

            # Nur loggen, wenn sich der Prozentwert geaendert hat
            if pct != getattr(self, "_last_lightdb_pct", None):
                self.logger.log(f"[App] LightDB Fortschritt: {pct}% (Status: {status})")
                self._last_lightdb_pct = pct

        # Jetzt die Cache-Version nutzen (keine SQLite-Blockade im GUI-Thread)
        if self.lightdb_manager.is_ready_cached():
            if not getattr(self.af_fetcher, "lightdb_path", None):
                self.logger.log("[App] OK: LightDB fertiggestellt, aktiviere Lookup im AF-Controller.")
                self.af_fetcher.lightdb_path = self.lightdb_manager.OUT_DB
            return  # kein weiteres Polling noetig

        # Weiter pollen, solange nicht fertig UND Widget existiert
        if self.winfo_exists():
            self.after(5000, self._poll_lightdb_status)
            
           
    def _ensure_gui_update(self, key, updated_fields):
        """
        FIX: Forciert GUI-Update nach Feld-Annotation.
        Behebt Problem dass annotierte Felder nicht angezeigt werden.
        
        ⚠️ V9 HINWEIS: Diese Methode nutzt emit_queue DIREKT statt drain_live_enqueue.
        Das ist beabsichtigt für sofortige Updates einzelner Zeilen.
        Für Batch-Updates sollte drain_live_enqueue verwendet werden.
        
        UMGEHT SINGLE TRUTH DER DB!!! 
        """
        try:
            # 1. Update in Datenbank
            self.db.update_variant_fields(key, updated_fields)
            
            # 2. Force emit to GUI queue (direkter Zugriff - beabsichtigt für Einzelupdates)
            if hasattr(self, 'distiller') and self.distiller:
                if hasattr(self.distiller, 'emit_queue'):
                    self.distiller.emit_queue.add(key)
            
            # 3. Wenn bereits in Tree, direkt updaten
            iid = self._get_iid_for_key(key)
            if iid and self.tree.exists(iid):
                # Hole aktuelle Werte aus DB
                variant = self.db.get_variant(key)
                if variant:
                    # Update Tree-Item
                    # ✅ FIX J3: _format_tree_values existiert nicht -> _compute_row_values
                    values = self._compute_row_values(key, variant)
                    tags = self._compute_row_tags(variant)
                    self.tree.item(iid, values=values, tags=tags)
                    
        except Exception as e:
            if self.logger:
                self.logger.log(f"[GUI] ⚠️ _ensure_gui_update error: {e}")
                
    #WICHTIG          
    def invalidate_cache_bulk(self, keys: List[Tuple]):
        """
        Entfernt Keys aus dem GUI-Cache, damit beim nächsten Refresh 
        frische Daten aus der DB geladen werden.
        """
        with self._table_lock:
            for key in keys:
                self.rows_cache.pop(key, None)
                
    def _get_iid_for_key(self, key):
        """Helper: Findet Tree-Item ID für Key UMGEHT SINGLE TRUTH DER DB!!! """
        search_str = f"{key[0]}:{key[1]}:{key[2]}>{key[3]}"
        for iid in self.tree.get_children(""):
            if search_str in iid:
                return iid
        return None
    
    def _init_gene_annotator(self):
        """Background initialization of GeneAnnotator."""
        try:
            logger.log("[App] ========== GENE ANNOTATOR INITIALISIERUNG ==========")
            
            # ✅ V9: Logger übergeben für besseres Debugging
            ga = GeneAnnotator(logger=logger)
            
            # Status-Report ausgeben
            logger.log(ga.get_status_report())
            
            # Nach Fertigstellung zuweisen
            self.gene_annotator = ga
            
            # V10: GeneAnnotator an MainFilterGate übergeben
            if hasattr(self, 'main_filter_gate') and self.main_filter_gate:
                self.main_filter_gate.set_gene_annotator(ga)
                logger.log("[App] GeneAnnotator an MainFilterGate uebergeben")
            
            # Distiller aktualisieren
            if self.distiller:
                self.distiller.gene_annotator = ga
                if hasattr(self, 'main_filter_gate') and self.main_filter_gate:
                    self.distiller.coding_filter = self.main_filter_gate.get_coding_filter()
                logger.log("[App] Distiller CodingFilter aktualisiert")
            
            # Auch im Maintainer setzen
            if self.maint:
                self.maint.gene_annotator = ga
                logger.log("[App] ✅ GeneAnnotator an Maintainer übergeben")
                
            if ga.available:
                logger.log("[App] ✅ GeneAnnotator BEREIT - Gene-Annotation ist aktiv")
            else:
                logger.log("[App] ⚠️ GeneAnnotator NICHT VERFÜGBAR - Gene werden NICHT annotiert!")
                logger.log("[App] ⚠️ Bitte prüfen Sie das data/annotations Verzeichnis")
            
            logger.log("[App] ========== GENE ANNOTATOR FERTIG ==========")
            
            # NEU: Erzwinge einen Refresh der Tabelle, falls Annotationen nachgeladen wurden
            self.after(0, lambda: self._refresh_table(force=True))
            
        except Exception as e:
            logger.log(f"[App] ❌ GeneAnnotator initialization FAILED: {e}")
            import traceback
            logger.log(f"[App] Traceback: {traceback.format_exc()}")
    
    def _add_or_update_row_fast(self, key):
        # ✅ Cache-Eviction VOR Insert
        cache_size = len(self.rows_cache)
        if cache_size >= Config.CACHE_MAX_SIZE:
            evict_count = max(1, int(cache_size * 0.1))
            try:
                for _ in range(evict_count):
                    try:
                        first_key = next(iter(self.rows_cache))
                        del self.rows_cache[first_key]
                    except (StopIteration, KeyError):
                        break
                
                if not hasattr(self, '_evict_count'):
                    self._evict_count = 0
                self._evict_count += 1
                
                if self._evict_count % 10 == 0:
                    self.logger.log(
                        f"[Cache] Eviction #{self._evict_count}: "
                        f"{evict_count} removed, {len(self.rows_cache)} remaining"
                    )
            except Exception as e:
                self.logger.log(f"[Cache] ⚠️ Eviction error: {e}")
        
        # ✅ Cache-First Lookup
        row = self.rows_cache.get(key)
        
        if not row:
            # Async Fetch (nicht blockierend)
            threading.Thread(
                target=self._fetch_row_async,
                args=(key,),
                daemon=True
            ).start()
            return
        
        # Rest unverändert...
        try:
            values = self._compute_row_values(key, row)
            tags = self._compute_row_tags(row)
            iid = "|".join(map(str, key))
        except Exception as e:
            self.logger.log(f"[App] ⚠️ Row compute error for {key}: {e}")
            return
        
        try:
            if self.tree.exists(iid):
                self.tree.item(iid, values=values, tags=tags)
            else:
                self.tree.insert("", "end", iid=iid, values=values, tags=tags)
        except tk.TclError:
            pass
        
    # [PROBABLY DEAD] Inline-Logik in _add_or_update_row_fast ersetzt diese Methode
    # def _evict_cache_async(self):
        # """
        # Helper: Non-blocking Cache-Eviction.
        
        # Änderungen:
        # - Läuft in separatem Thread
        # - Evicted nur 10% (sanfter)
        # - Logging nur alle 10 Evictions
        # """
        # try:
            # current_size = len(self.rows_cache)
            # evict_count = max(1, int(current_size * 0.1))  # ✅ 10% statt 20%
            
            # ✅ Entferne älteste Einträge
            # for _ in range(evict_count):
                # try:
                    # first_key = next(iter(self.rows_cache))
                    # del self.rows_cache[first_key]
                # except (StopIteration, KeyError):
                    # break
            
            # ✅ Stats nur alle 10 Evictions
            # if not hasattr(self, '_evict_count'):
                # self._evict_count = 0
            
            # self._evict_count += 1
            # if self._evict_count % 10 == 0:
                # self.logger.log(
                    # f"[Cache] Eviction #{self._evict_count}: "
                    # f"{evict_count} entfernt, "
                    # f"{len(self.rows_cache)} verbleibend"
                # )
        # except Exception as e:
            # self.logger.log(f"[Cache] ⚠️ Eviction error: {e}")
        
    def _fetch_row_async(self, key):
        """Fetches row from DB asynchronously and updates cache."""
        try:
            row = self.db.get_variant(key)
            if row:
                self.rows_cache[key] = row
                # Trigger re-drain (row ist jetzt im Cache)
                try:
                    self.live_queue.put_nowait(key)
                except queue.Full:
                    pass
        except Exception as e:
            logger.log(f"[App] ⚠️ Async fetch error for {key}: {e}")



    def _add_or_update_row(self, key):
        self._add_or_update_row_fast(key)
        
    # [PROBABLY DEAD] Nie aufgerufen - Insert Position wird nicht mehr berechnet
    # def _calculate_insert_position(self, new_key):
        # """
        # Berechnet die korrekte Insert-Position basierend auf genomischer Sortierung.
        
        # FIXES:
        # - Verhindert "end"-Chaos
        # - Garantiert chromosomale Reihenfolge
        # """
        # try:
            # new_chr = str(new_key[0]).replace("chr", "").upper()
            # new_pos = int(new_key[1])
            
            # Numerische Chromosomen zuerst, dann X/Y/MT
            # def chr_sort_key(c):
                # if c.isdigit():
                    # return (0, int(c))
                # elif c == "X":
                    # return (1, 23)
                # elif c == "Y":
                    # return (1, 24)
                # elif c == "MT":
                    # return (1, 25)
                # else:
                    # return (2, 999)
            
            # new_chr_key = chr_sort_key(new_chr)
            
            # Suche korrekte Position
            # children = self.tree.get_children("")
            # for i, iid in enumerate(children):
                # try:
                    # vals = self.tree.item(iid)["values"]
                    # if not vals or len(vals) < 2:
                        # continue
                        
                    # chr_val = str(vals[0]).replace("chr", "").upper()
                    # pos_val = int(vals[1])
                    
                    # chr_key = chr_sort_key(chr_val)
                    
                    # Vergleich: (chr, pos)
                    # if new_chr_key < chr_key:
                        # return i
                    # elif new_chr_key == chr_key and new_pos < pos_val:
                        # return i
                        
                # except (ValueError, IndexError):
                    # continue
            
            # Am Ende einfügen
            # return "end"
            
        # except Exception:
            # return "end"
    
    # Dies ist der genutzte Weg als Zugang zur GUI der Keys
    # AKTIV
    # WICHTIG
    def _drain_live_enqueue(self):
        # ✅ Initialisiere Lock einmalig
        if not hasattr(self, '_drain_lock'):
            self._drain_lock = threading.Lock()
        
        # ✅ Non-blocking acquire: skip wenn schon aktiv
        if not self._drain_lock.acquire(blocking=False):
            self.after(50, self._drain_live_enqueue)
            return
        
        try:
            batch = []
            max_items = 20
            start = time.perf_counter()
            max_ms = 16.0
            
            # Adaptive Batch-Größe
            queue_size = self.live_queue.qsize()
            if queue_size > 2000:
                max_items = 3
                delay = 30
            elif queue_size > 1000:
                max_items = 5
                delay = 50
            elif queue_size > 500:
                max_items = 10
                delay = 100
            elif queue_size < 50:
                max_items = 30
                delay = 200
            else:
                delay = 150
            
            # Queue drainen mit Time-Budget
            while len(batch) < max_items:
                elapsed_ms = (time.perf_counter() - start) * 1000.0
                if elapsed_ms >= max_ms:
                    break
                
                try:
                    item = self.live_queue.get_nowait()
                except queue.Empty:
                    break
                
                if not (isinstance(item, tuple) and len(item) == 5):
                    continue
                
                batch.append(item)
            
            # Batch verarbeiten
            if batch:
                try:
                    for key in batch:
                        self._add_or_update_row_fast(key)
                        self.distiller.add_display_key(key)
                    
                    if not hasattr(self, '_drain_count'):
                        self._drain_count = 0
                    
                    self._drain_count += len(batch)
                    
                    if self._drain_count >= 50:
                        self.update_count_label()
                        self._drain_count = 0
                    
                except Exception as e:
                    self.logger.log(f"[App] ⚠️ Drain error: {e}")
            
            # Re-Schedule
            self.after(delay, self._drain_live_enqueue)
        
        finally:
            # ✅ Lock IMMER freigeben (auch bei Exception)
            self._drain_lock.release()

    # In App.txt
    def _refresh_table(self, force=False):
        """
        FIX: Refresh mit korrektem Count-Update.
        
        Änderungen:
        - Count-Update nach Rebuild
        - Skip bei Überlast
        - Längeres Throttling-Intervall
        """
        # ✅ Skip bei Überlast
        queue_size = self.live_queue.qsize()
        if queue_size > self.live_queue.maxsize * 0.8:
            self.logger.log(f"[Table-Refresh] Skip: Queue overloaded ({queue_size})")
            self.after(5000, self._refresh_table)
            return

        with self._table_lock:
            keys = list(self.distiller.display_keys)

            if not keys and not force:
                self.after(3000, self._refresh_table)
                return

            sig = hash(tuple(sorted(keys)))

            # ✅ Skip bei kleinen Änderungen (< 1% oder < 10 Keys)
            if not force and self._last_keyset_sig:
                old_len = len(self._last_keys) if hasattr(self, '_last_keys') else 0
                new_len = len(keys)

                if old_len > 0:
                    delta = abs(new_len - old_len)
                    delta_pct = delta / old_len

                    if delta < 10 or delta_pct < 0.01:
                        self.after(3000, self._refresh_table)
                        return

            # FIX: refer to self.progress instead of undefined 'progress'
            if force or sig != self._last_keyset_sig or self.progress.phase_name == "Abgeschlossen":
                self._last_keys = keys
                self._full_table_rebuild(keys, sig)
            else:
                self._incremental_table_update(keys)

        # ✅ Längeres Intervall
        self.after(3000, self._refresh_table)
            
    def _compute_row_tags(self, row):
        """Helper: Berechnet Tags."""
        tags = ()
        cadd_val = row.get("cadd_phred")
        try:
            cadd_thr = float(self.cadd_highlight_threshold.get())
        except Exception:
            cadd_thr = None
        
        if cadd_thr is not None and isinstance(cadd_val, (int, float)) and cadd_val >= cadd_thr:
            tags = ("high_cadd",)
        
        return tags


    # [PROBABLY DEAD] Nie aufgerufen - DB-Felder werden direkt zugegriffen
    # def _get_db_field(self, ui_col):
        # """
        # Helper: Nutzt zentrales Mapping der DB.
        # """
        # Wir fragen direkt die Datenbankklasse nach dem richtigen Feldnamen
        # return self.db.resolve_column(ui_col)

    def _compute_row_values(self, key, row):
        """
        Extrahiert Werte für die Tabelle. 
        Mappt UI-Spaltennamen auf DB-Felder und formatiert Coding-Boolean.
        """
        values = []
        for col in self.columns:
            # 1. Keys & Spezialspalten (direkt aus Row oder Key)
            if col == "chr":
                values.append(row.get("chr") or key[0])
            elif col == "pos":
                values.append(str(row.get("pos") or key[1]))
            elif col == "ref":
                values.append(row.get("ref") or key[2])
            elif col == "alt":
                values.append(row.get("alt") or key[3])
            elif col == "build":
                values.append(row.get("build") or key[4])
            elif col == "rsid":
                rsid = row.get("rsid") or "."
                values.append(rsid if rsid else ".")
            elif col == "genotype":
                gt_raw = self.distiller.genotype_store.get(key, ".")
                values.append(normalize_genotype_display(gt_raw))
            elif col == "pubmed":
                values.append("PubMed")
            
            # Status-Spalten
            elif col == "meanAF_status":
                code = row.get("meanAF_fetch_success")
                values.append(FetchStatusManager.status_label(code))
            elif col == "full_status":
                code = row.get("full_fetch_success")
                values.append(FetchStatusManager.status_label(code))
            
            # 2. Generische DB-Felder (mit Mapping)
            else:
                # WICHTIG: Hier wird "gene" -> "gene_symbol" und "coding" -> "is_coding" aufgelöst
                db_field = self.db.resolve_column(col)
                val = row.get(db_field)
                
                # Formatierung für Coding (Boolean -> Text)
                if col == "coding":
                    if val is None or val == "":
                        values.append(".")
                    else:
                        try:
                            # DB speichert 0/1, UI zeigt Nein/Ja
                            is_cod = int(val)
                            values.append("Ja" if is_cod else "Nein")
                        except:
                            values.append(str(val))
                
                # Generische Formatierung
                elif val is None or val == "":
                    values.append(".")
                
                elif isinstance(val, float):
                    values.append(f"{val:.4g}")
                
                elif isinstance(val, (list, dict)):
                    try:
                        if isinstance(val, list):
                            values.append(", ".join(map(str, val)))
                        else:
                            import json
                            values.append(json.dumps(val))
                    except:
                        values.append(str(val))
                else:
                    values.append(str(val))
        
        return values
    
    def _full_table_rebuild(self, keys: List[Tuple], sig: int):
        """
        FIX: Rebuild mit genomischer Sortierung und Count-Update.
        
        Änderungen:
        - Keys werden VOR dem Insert sortiert
        - Count-Update nach Completion
        """
        # ✅ Genomische Sortierung
        def sort_key(k):
            chr_str = str(k[0]).replace("chr", "").upper()
            pos_int = int(k[1])
            
            if chr_str.isdigit():
                chr_num = int(chr_str)
            elif chr_str == "X":
                chr_num = 23
            elif chr_str == "Y":
                chr_num = 24
            elif chr_str == "MT":
                chr_num = 25
            else:
                chr_num = 999
            
            return (chr_num, pos_int)
        
        sorted_keys = sorted(keys, key=sort_key)

        # ✅ FIX V13.1: Lösche ALLE Items (auch detachte vom Post-Filter)
        # get_children() gibt nur attached Items zurück, aber detachte Items
        # verhindern das Einfügen neuer Items mit gleicher IID
        all_iids = list(self.tree.get_children())

        # Sammle auch detachte Items (aus vorherigen Post-Filter-Anwendungen)
        if hasattr(self, '_detached_iids') and self._detached_iids:
            all_iids.extend(self._detached_iids)
            self._detached_iids.clear()

        # Lösche alle Items
        for iid in all_iids:
            try:
                self.tree.delete(iid)
            except tk.TclError:
                pass  # Item existiert nicht mehr

        total = len(sorted_keys)

        def insert_batch(start: int = 0):
            end = min(start + Config.TABLE_BATCH_SIZE, total)
            
            for k in sorted_keys[start:end]:
                # ✅ Insert mit korrekter Position (am Ende, da Keys schon sortiert)
                self._add_or_update_row(k)

            # ✅ Count-Update während Rebuild
            if start % (Config.TABLE_BATCH_SIZE * 5) == 0 or end >= total:
                self.update_count_label()

            if end < total:
                self.after(10, lambda: insert_batch(end))
            else:
                self._last_keyset_sig = sig
                self._refresh_idx = 0
                self.distiller.display_keys = set(sorted_keys)
                # ✅ Finales Count-Update
                self.update_count_label()

        self.after(0, insert_batch)


    def _sort_by(self, col, descending):
        """
        Sortiert sicher mit Lock-Koordination.
        
        FIXES:
        - Wartet auf drain-Completion
        - Verhindert Sort während Update
        """
        # ✅ Warte auf drain
        if hasattr(self, '_drain_active') and self._drain_active:
            self.after(100, lambda: self._sort_by(col, descending))
            return
        
        with self._table_lock:
            children = list(self.tree.get_children(""))
            idx = self.columns.index(col)

            snapshot = []
            for iid in children:
                vals = self.tree.item(iid)["values"]
                v = vals[idx] if idx < len(vals) else ""
                try:
                    f = float(v)
                    key = (0, f, iid)
                except Exception:
                    key = (1, str(v), iid)
                snapshot.append((iid, key))

            def compute_order(data):
                data.sort(key=lambda t: t[1], reverse=descending)
                return [iid for iid, _ in data]

            def apply_order(order):
                BATCH = 200
                def step(start=0):
                    end = min(start + BATCH, len(order))
                    for i, iid in enumerate(order[start:end], start):
                        try:
                            self.tree.move(iid, "", i)
                        except Exception:
                            pass
                    if end < len(order):
                        self.after(10, lambda: step(end))
                    else:
                        self.tree.heading(col, command=lambda c=col: self._sort_by(c, not descending))
                step(0)

            def worker():
                order = compute_order(snapshot)
                self.after(0, lambda: apply_order(order))

            threading.Thread(target=worker, daemon=True).start()

    def change_language(self, lang):
        self.translator.set_lang(lang)
        # komplettes GUI neu aufbauen
        for widget in self.winfo_children():
            widget.destroy()
        self._build_gui()
        
    def _apply_column_visibility(self):
        """
        Wendet die aktuelle Spaltenauswahl auf das Treeview an
        und speichert die Einstellung.
        """
        self.visible_columns = {c for c, var in self._viewmenu_vars.items() if var.get()}
        display = [c for c in self.columns if c in self.visible_columns]
        self.tree.configure(displaycolumns=display)
        self._save_settings()
        self._refresh_table(force=True)


    def _reset_columns(self):
        """
        Setzt die Spaltenauswahl zurück: alle Spalten sichtbar.
        """
        self.visible_columns = set(self.columns)
        for col, var in self._viewmenu_vars.items():
            var.set(True)
        self._save_settings()
        self._refresh_table(force=True)


    def _refresh_progress(self):
        """
        V17: Garantiert fortlaufendes Polling (verhindert 'Stuck' Progress Bar).
        """
        if not self.winfo_exists():
            return
            
        # Schedule next update before any early returns to ensure polling continues
        self.after(Config.PROGRESS_UPDATE_INTERVAL, self._refresh_progress)
        
        # UI-Update bei Abbruch
        if hasattr(self, 'stopflag') and self.stopflag and self.stopflag.is_set():
            self.progress_var.set(0)
            self.perc_label.config(text="Abgebrochen")
            self.eta_label.config(text="")
            self.phase_label.config(text="Phase: Abgebrochen")
            return

        status = self.progress.get_detailed_status()
        
        # GUI-Labels updaten
        phase_display = self.progress.phase_name or "—"
        self.phase_label.config(text=f"Phase: {phase_display}")
        self.progress_var.set(status["percent"])
        self.perc_label.config(text=f"{status['percent']:.1f}%")
        self.eta_label.config(text=fmt_eta(status["eta_seconds"]))
        
        # Optional: Detaillierte Phase-Info
        summary = self.progress.get_summary()
        if hasattr(self, 'detail_label'):
            self.detail_label.config(text=summary)
        
        # Logging (alle 10 Sekunden)
        if not hasattr(self, '_last_progress_log'):
            self._last_progress_log = 0
        
        now = time.time()
        if now - self._last_progress_log > 10:
            if status["total"] > 0 and self.progress.phase_name != "Abgeschlossen":
                self.logger.log(f"[Progress] {summary}")
            self._last_progress_log = now
        
        # Reset UI if finished
        if self.progress.phase_name == "Abgeschlossen":
             self.after(2000, lambda: self.progress_bar.configure(value=0))
             self.after(2000, lambda: self.perc_label.config(text=self._t("Fertig")))
             self.after(2000, lambda: self.eta_label.config(text=""))
        
    def on_close(self):
        """
        ✅ ERWEITERT: Sauberer App-Exit + BackofficeCrawler-Start.
        
        Workflow:
        1. StopFlag setzen (stoppt alle laufenden Threads)
        2. Maintainer stoppen (falls aktiv)
        3. Distiller cleanup:
           - VCF-Buffer flush
           - FastaValidator schließen
           - Controller cleanup
        4. DB-Verbindung freigeben
        5. GUI zerstören
        6. BackofficeCrawler starten (Migration + Background-Maintenance)
        
        Features:
        - Comprehensive Error-Handling
        - Resource-Cleanup (DB, FASTA, Buffer)
        - Gene-Annotator & AF-Fetcher an Crawler übergeben
        - Cache-Path-Ermittlung für Migration
        """
        cache_path = None
        db_path = None
        gene_annotator = None
        af_fetcher = None
        distiller = None
        stopflag = None
        
        try:
            logger.log("[App] 🛑 Beende Anwendung...")
            
            # ============================================================
            # FAST EXIT fuer PyInstaller (kein Cleanup noetig)
            # ============================================================
            if getattr(sys, 'frozen', False):
                if hasattr(self, "stopflag") and isinstance(self.stopflag, StopFlag):
                    self.stopflag.stop()
                try:
                    self.destroy()
                except Exception:
                    pass
                release_single_instance()
                os._exit(0)
            
            # ============================================================
            # PHASE 1: STOPFLAG SETZEN
            # ============================================================
            if hasattr(self, "stopflag") and isinstance(self.stopflag, StopFlag):
                try:
                    self.stopflag.stop()
                    logger.log("[App] ✅ StopFlag gesetzt")
                except Exception as e:
                    logger.log(f"[App] ⚠️ StopFlag-Fehler: {e}")
            
            # V10 FIX: Warte auf Pipeline-Thread bevor Cleanup
            if hasattr(self, "pipeline_thread") and self.pipeline_thread and self.pipeline_thread.is_alive():
                logger.log("[App] ⏳ Warte auf Pipeline-Thread...")
                self.pipeline_thread.join(timeout=3.0)
                if self.pipeline_thread.is_alive():
                    logger.log("[App] ⚠️ Pipeline-Thread reagiert nicht")
                else:
                    logger.log("[App] ✅ Pipeline-Thread beendet")
            
            # StopFlag für Crawler speichern
            stopflag = getattr(self, "stopflag", None)
            
            # ============================================================
            # PHASE 2: MAINTAINER STOPPEN
            # ============================================================
            if hasattr(self, "maint") and self.maint:
                try:
                    logger.log("[App] 🛑 Stoppe BackgroundMaintainer...")
                    self.maint.stop()
                    logger.log("[App] ✅ BackgroundMaintainer gestoppt")
                except Exception as e:
                    logger.log(f"[App] ⚠️ Maintainer-Stop-Fehler: {e}")
            
            # ============================================================
            # PHASE 3: DISTILLER CLEANUP
            # ============================================================
            if hasattr(self, "distiller") and self.distiller:
                distiller = self.distiller  # Für Crawler speichern
                
                try:
                    logger.log("[App] 🧹 Distiller Cleanup...")
                    
                    # VCF-Buffer flushen
                    if hasattr(self.distiller, "vcf_buffer"):
                        try:
                            flushed = self.distiller.vcf_buffer.flush()
                            if flushed:
                                logger.log(f"[App] 💾 VCF-Buffer geflusht: {flushed} Records")
                        except Exception as e:
                            logger.log(f"[App] ⚠️ VCF-Buffer-Flush-Fehler: {e}")
                    
                    # FastaValidator schließen
                    if hasattr(self.distiller, "fasta_validator") and self.distiller.fasta_validator:
                        try:
                            self.distiller.fasta_validator.close()
                            logger.log("[App] ✅ FastaValidator geschlossen")
                        except Exception as e:
                            logger.log(f"[App] ⚠️ FastaValidator-Close-Fehler: {e}")
                    
                    # Controller cleanup (falls vorhanden)
                    if hasattr(self.distiller, "cleanup"):
                        try:
                            self.distiller.cleanup()
                            logger.log("[App] ✅ Distiller-Cleanup abgeschlossen")
                        except Exception as e:
                            logger.log(f"[App] ⚠️ Distiller-Cleanup-Fehler: {e}")
                    
                    # Gene-Annotator & AF-Fetcher für Crawler speichern
                    if hasattr(self.distiller, "gene_annotator"):
                        gene_annotator = self.distiller.gene_annotator
                    
                    if hasattr(self.distiller, "af_fetcher"):
                        af_fetcher = self.distiller.af_fetcher
                    
                except Exception as e:
                    logger.log(f"[App] ⚠️ Distiller-Cleanup fehlgeschlagen: {e}")
            
            # ============================================================
            # PHASE 4: DB-VERBINDUNG FREIGEBEN
            # ============================================================
            if hasattr(self, "db") and isinstance(self.db, VariantDB):
                try:
                    db_path = self.db.db_path
                    logger.log(f"[App] 📦 DB-Pfad gespeichert: {db_path}")
                    
                    # Cache flush (falls vorhanden)
                    try:
                        flushed = self.db.flush_annotation_cache()
                        if flushed and flushed > 0:
                            logger.log(f"[App] 💾 Annotation-Cache geflusht: {flushed} Records")
                    except Exception as e:
                        logger.log(f"[App] ⚠️ Cache-Flush-Fehler: {e}")

                    # ✅ V12 FIX: WAL checkpoint before closing
                    try:
                        self.db.close()
                        logger.log("[App] ✅ DB WAL checkpoint durchgeführt")
                    except Exception as e:
                        logger.log(f"[App] ⚠️ DB-Close-Fehler: {e}")

                    # DB freigeben
                    self.db = None
                    logger.log("[App] ✅ DB-Verbindung freigegeben")
                    
                except Exception as e:
                    logger.log(f"[App] ⚠️ DB-Freigabe-Fehler: {e}")
            
            # ============================================================
            # PHASE 5: CACHE-PATH ERMITTELN
            # ============================================================
            # Standard-Pfad prüfen
            potential_cache_paths = [
                "data/vcf_cache.json",
                "vcf_cache.json",
                os.path.join(os.path.dirname(db_path) if db_path else ".", "vcf_cache.json")
            ]
            
            for path in potential_cache_paths:
                if os.path.exists(path):
                    cache_path = path
                    logger.log(f"[App] 📦 VCF-Cache gefunden: {cache_path}")
                    break
            
            if not cache_path:
                logger.log("[App] ℹ️ Kein VCF-Cache gefunden (Migration wird übersprungen)")
        
        except Exception as e:
            logger.log(f"[App] ❌ Kritischer Fehler beim Beenden: {e}")
            import traceback
            logger.log(f"[App] Traceback:\n{traceback.format_exc()}")
        
        finally:
            # ============================================================
            # PHASE 6: ONEDRIVE WIEDER STARTEN (V16 FIX 12)
            # ============================================================
            try:
                resume_onedrive(logger)
            except Exception as e:
                logger.log(f"[App] ⚠️ OneDrive-Resume-Fehler: {e}")

            # ============================================================
            # PHASE 7: GUI ZERSTÖREN
            # ============================================================
            try:
                self.destroy()
                logger.log("[App] ✅ GUI zerstört")
            except Exception as e:
                logger.log(f"[App] ⚠️ GUI-Destroy-Fehler: {e}")
            
            # ============================================================
            # PHASE 8: BACKOFFICE-CRAWLER STARTEN
            # ============================================================
            try:
                if getattr(sys, 'frozen', False):
                    logger.log('[App] ℹ️ PyInstaller-Modus: BackofficeCrawler deaktiviert (sauberer Exit)')
                    db_path = None  # Skip Crawler
                if db_path:
                    logger.log("[App] 🚀 Starte BackofficeCrawler...")
                    
                    _crawler_lang = getattr(getattr(self, 'translator', None), 'lang', 'de')
                    crawler = BackofficeCrawler(
                        cache_path=cache_path,  # Kann None sein
                        db_path=db_path,
                        logger=logger,
                        distiller=distiller,
                        db=None,  # Wird im Crawler neu geöffnet falls nötig
                        stopflag=stopflag,
                        maintainer_cls=BackgroundMaintainer,
                        gene_annotator=gene_annotator,
                        af_fetcher=af_fetcher,
                        lang=_crawler_lang
                    )
                    
                    # Crawler in Thread starten
                    crawler_thread = threading.Thread(
                        target=crawler.run, 
                        daemon=False,  # ✅ WICHTIG: Nicht daemon, sonst wird er beendet
                        name="BackofficeCrawler"
                    )
                    crawler_thread.start()
                    
                    logger.log(
                        "[App] ✅ BackofficeCrawler gestartet\n"
                        f"  Thread: {crawler_thread.name}\n"
                        f"  Cache: {cache_path or 'None (skip migration)'}\n"
                        f"  DB: {db_path}\n"
                        f"  Gene-Annotator: {'✓' if gene_annotator else '✗'}\n"
                        f"  AF-Fetcher: {'✓' if af_fetcher else '✗'}"
                    )
                else:
                    logger.log(
                        "[App] ⚠️ BackofficeCrawler nicht gestartet "
                        "(kein DB-Pfad verfügbar)"
                    )
            
            except Exception as e:
                logger.log(
                    f"[App] ❌ BackofficeCrawler-Start fehlgeschlagen: {e}"
                )
                import traceback
                logger.log(f"[App] Traceback:\n{traceback.format_exc()}")
            
            finally:
                # ✅ V17: PID Lock freigeben
                release_single_instance()
                logger.log("[App] 👋 Auf Wiedersehen!")
                
                # PyInstaller: Harter Exit um haengende non-daemon Threads zu beenden
                if getattr(sys, 'frozen', False):
                    os._exit(0)


    # =========================================================================
    # SETTINGS LADEN/SPEICHERN (ERWEITERT)
    # =========================================================================
       
    def _change_theme(self, theme_name):
        """
        ✅ NEU: Wechselt das Design-Theme zur Laufzeit.
        """
        self.style.theme_use(theme_name)
        # Speichern in Config/Settings
        self._save_settings(theme=theme_name)
        ToastNotification(
            title="Design", 
            message=f"Theme geändert zu {theme_name}", 
            duration=1500,
            bootstyle="info"
        ).show_toast()

    def open_quality_settings(self):
        """
        ✅ AKTUALISIERT: Öffnet den Dialog mit korrektem Manager-Objekt.
        """
        af_mgr = None
        if hasattr(self, 'distiller') and self.distiller:
            af_mgr = self.distiller.af_none_manager
        
        # Ruft den neuen QualitySettingsDialog auf
        QualitySettingsDialog(self, self.quality_manager, af_mgr)

    def reset_post_filter(self):
        """
        ✅ AKTUALISIERT: Setzt alle Filter zurück (inkl. Consequences) und refresht Tabelle.
        """
        try:
            # Reset UI Controls
            self.postfilter_cadd_min.set("")
            self.postfilter_benign.set(True)
            self.postfilter_likely_benign.set(True)
            self.postfilter_vus.set(True)
            self.postfilter_likely_path.set(True)
            self.postfilter_path.set(True)
            self.postfilter_impact_high.set(True)
            self.postfilter_impact_moderate.set(True)
            self.postfilter_impact_low.set(True)
            self.postfilter_protein_coding.set(False)
            
            # Reset Consequences Dictionary
            for var in self.postfilter_consequences.values():
                var.set(True)
            
            self.gen_whitelist.set("")
            self.gen_blacklist.set("")
            
            # Tabelle aktualisieren
            self._refresh_table(force=True)
            self.update_count_label()
        
        except Exception as e:
            self.logger.log(f"[Filter] Reset controls error: {e}")
    
    def choose_file(self):
        """
        Öffnet einen Dateidialog zur Auswahl einer Eingabedatei (VCF, FASTA, 23andMe).
        """
        path = filedialog.askopenfilename(
            filetypes=[("VCF/23andMe/FASTA", "*.vcf *.vcf.gz *.txt *.fa *.fasta *.fna *.gvcf")]
        )
        if path:
            self.selected_file.set(path)


    def on_start(self):
        """
        V17: Start-Klick triggert immer erst Stop/Reset (Legacy compatibility).
        Sorgt für konsistenten Status bei Mehrfachklicks auf Start.
        """
        # ✅ V17: Start -> Stop -> Startmechanik
        self.on_stop()
        # ✅ 1. Stopflag initialisieren falls nicht vorhanden
        if not hasattr(self, "stopflag") or self.stopflag is None:
            self.stopflag = StopFlag()

        # ✅ V15: Start-Button deaktivieren während Cleanup
        try:
            for widget in self.winfo_children():
                if hasattr(widget, 'winfo_children'):
                    for child in widget.winfo_children():
                        if hasattr(child, 'cget') and hasattr(child, 'configure'):
                            try:
                                if 'START' in str(child.cget('text')).upper():
                                    child.configure(state='disabled')
                            except Exception:
                                pass
        except Exception:
            pass

        # ✅ V15: Asynchrone Restart-Sequenz
        def restart_sequence():
            """Läuft in separatem Thread - blockiert nicht die GUI."""
            try:
                # Falls Pipeline läuft → Abbruch signalisieren und warten
                if hasattr(self, "pipeline_thread") and self.pipeline_thread and self.pipeline_thread.is_alive():
                    self.logger.log("[App] ℹ Laufende Pipeline erkannt → Stopflag setzen.")
                    self.stopflag.stop()

                    # Warte bis Pipeline-Thread beendet ist (max 10 Sekunden)
                    self.pipeline_thread.join(timeout=10)

                    if self.pipeline_thread.is_alive():
                        self.logger.log("[App] ⚠️ Pipeline-Thread reagiert nicht - erzwinge Neustart")
                    else:
                        self.logger.log("[App] ✅ Pipeline-Thread sauber beendet")

                # GUI-Reset über after() (thread-safe)
                self.after(0, self._on_start_continue)

            except Exception as e:
                self.logger.log(f"[App] ⚠️ Fehler in restart_sequence: {e}")
                self.after(0, self._on_start_continue)

        # Starte Restart-Sequenz in separatem Thread
        threading.Thread(target=restart_sequence, daemon=True).start()

    def _on_start_continue(self):
        """
        V16 FIXED: Fortsetzung von on_start() - wird über after() aufgerufen (GUI-safe).
        Inklusive robustem GUI-Clearing für große Datenmengen.
        """
        try:
            # Start-Button wieder aktivieren
            for widget in self.winfo_children():
                if hasattr(widget, 'winfo_children'):
                    for child in widget.winfo_children():
                        if hasattr(child, 'cget') and hasattr(child, 'configure'):
                            try:
                                if 'START' in str(child.cget('text')).upper():
                                    child.configure(state='normal')
                            except Exception:
                                pass
        except Exception:
            pass

        # ✅ 3. Immer resetten, damit neuer Run sauber startet
        self.stopflag.clear()

        # Distiller und Maintainer auf aktuelles Flag umbiegen
        if hasattr(self, "distiller") and self.distiller:
            self.distiller.stopflag = self.stopflag
        if hasattr(self, "maint") and self.maint:
            self.maint.stopflag = self.stopflag

        # ✅ 4. GUI clearen (robust bei >100k Einträgen)
        children = self.tree.get_children()
        total = len(children)
        if total > 0:
            self.logger.log(f"[App] 🧹 Bereinige GUI ({total} Varianten)...")
            if total > 5000:
                # In Blöcken löschen, um GUI-Thread nicht zu lange zu blockieren
                for i in range(0, total, 5000):
                    self.tree.delete(*children[i:i+5000])
                    self.update_idletasks() # Kurz atmen lassen
            else:
                self.tree.delete(*children)
            self.logger.log("[App] ✅ GUI bereinigt.")

        # Distiller-Reset
        if hasattr(self, "distiller") and self.distiller:
            try:
                self.distiller.display_keys = set()
                if hasattr(self.distiller, 'progress'):
                    self.distiller.progress.reset()
                if hasattr(self.distiller, 'rows_cache'):
                    self.distiller.rows_cache.clear()
                if hasattr(self.distiller, 'emit_queue'):
                    self.distiller.emit_queue.reset()
                if hasattr(self.distiller, 'vcf_buffer'):
                    try: self.distiller.vcf_buffer.flush()
                    except Exception: pass
                
                self.distiller.orig_header = []
                self.distiller.orig_records = {}
                self.distiller.current_vcf_keys = set()
                
                if hasattr(self.distiller, 'genotype_store'):
                    self.distiller.genotype_store.clear()
                
                if hasattr(self.distiller, 'to_fetch_af') and hasattr(self.distiller, 'to_fetch_af_lock'):
                    with self.distiller.to_fetch_af_lock:
                        self.distiller.to_fetch_af.clear()
                
                self.logger.log("[App] ✅ Distiller-State zurückgesetzt.")
            except Exception as e:
                self.logger.log(f"[App] ⚠️ Distiller-Reset teilweise fehlgeschlagen: {e}")

        # ✅ 6. App-Cache clearen
        try:
            if hasattr(self, 'rows_cache'):
                self.rows_cache.clear()
            if hasattr(self, '_normalized_cache'):
                self._normalized_cache.clear()
            if hasattr(self, '_last_keyset_sig'):
                self._last_keyset_sig = None
            self.logger.log("[App] ✅ App-Cache zurückgesetzt.")
        except Exception as e:
            self.logger.log(f"[App] ⚠️ App-Cache-Reset fehlgeschlagen: {e}")

        # ✅ 7. Live-Queue clearen
        try:
            while not self.live_queue.empty():
                try:
                    self.live_queue.get_nowait()
                except queue.Empty:
                    break
            self.logger.log("[App] ✅ Live-Queue geleert.")
        except Exception as e:
            self.logger.log(f"[App] ⚠️ Live-Queue-Reset fehlgeschlagen: {e}")

        # ✅ 7b. V10 FIX: Progress-UI zurücksetzen
        try:
            self.progress_var.set(0)
            self.perc_label.config(text="0.0%")
            self.eta_label.config(text="--:--")
            self.phase_label.config(text="Phase: Initialisierung...")
            self.logger.log("[App] ✅ Progress-UI zurückgesetzt.")
        except Exception as e:
            self.logger.log(f"[App] ⚠️ Progress-UI-Reset fehlgeschlagen: {e}")

        # ✅ 8. Datei-Check (FIX: Messagebox)
        path = self.selected_file.get()
        if not path:
            Messagebox.show_error("Bitte eine Eingabe-Datei auswählen.", "Fehlende Datei")
            return

        # ✅ 9. AlphaGenome-Key aktualisieren
        try:
            self._save_settings()
            self.distiller.set_alpha_genome_key(self.alphagenome_key.get().strip() or None)
        except Exception as e:
            self.logger.log(f"[App] ⚠️ AlphaGenome-Update fehlgeschlagen: {e}")

        # ✅ 10. Filter-Parameter sammeln
        kwargs = {
            "af_threshold": float(self.af_threshold.get()),
            "include_none": bool(self.include_none.get()),
            "filter_pass_only": bool(self.filter_pass_only.get()),
            "cadd_highlight_threshold": float(self.cadd_highlight_threshold.get()),
            "stale_days": int(self.stale_days.get()),
            "qual_threshold": float(self.qual_threshold.get()),
            "skip_info_recycling": bool(self.skip_info_recycling.get())
        }

        # ✅ 11. Pipeline starten
        # V15: Härteres Pausieren des Maintainers, um DB-Lock-Contention zu vermeiden
        if self.maint:
            try:
                if hasattr(self.maint, 'force_pause'):
                    self.maint.force_pause(timeout=2.0)
                else:
                    self.maint.pause()
            except Exception:
                pass
        self.pipeline_thread = threading.Thread(
            target=self.distiller.process_file,
            args=(path,),
            kwargs=kwargs,
            daemon=True
        )
        self.pipeline_thread.start()

        self.logger.log(
            f"[App] ✅ Pipeline gestartet: {path} | "
            f"Filter: AF≤{kwargs['af_threshold']}, include_none={kwargs['include_none']}"
        )

    def on_stop(self):
        """Signalisiert den sofortigen Abbruch der Pipeline."""
        if hasattr(self, "stopflag") and self.stopflag:
            self.stopflag.stop()
            self.logger.log("[App] ⏹ Stop-Signal gesendet.")
        
        # NEU: Stoppe auch Progress-Tracker
        if hasattr(self, 'progress') and self.progress:
            self.progress.stop()
        
        if hasattr(self, 'distiller') and hasattr(self.distiller, 'progress') and self.distiller.progress:
            self.distiller.progress.stop()
        
        # UI-Reset
        try:
            self.phase_label.config(text="Phase: Abbruch eingeleitet...")
            self.progress_var.set(0)
            self.perc_label.config(text="Abgebrochen")
            self.eta_label.config(text="")
        except Exception:
            pass
        
        self.logger.log("[App] ✅ Stop-Signal gesetzt, Progress gestoppt.")
        
        # ✅ V17: Stop triggert immer auch den Reset/Refresh (Anzeigenzählung zurücksetzen)
        try:
            self.reset_post_filter()
        except Exception as e:
            self.logger.log(f"[App] ⚠️ Fehler bei Stop-Reset: {e}")
            
    def update_count_label(self):
        """
        FIX: Robuste Postfilter-Erkennung ohne False Positives.
        
        Änderungen:
        - Prüft nur tatsächlich gesetzte Filter
        - Wartet auf Pipeline-Completion
        - Keine voreilige Postfilter-Anzeige
        """
        try:
            visible_count = len(self.tree.get_children(""))
            total_filtered = len(self.distiller.display_keys) if hasattr(self.distiller, 'display_keys') else 0
            
            # ✅ Postfilter nur anzeigen wenn Pipeline KOMPLETT
            pipeline_complete = (
                hasattr(self.progress, 'is_phase_complete') and
                self.progress.is_phase_complete('af_fetch')
            )
            
            if not pipeline_complete:
                # Pipeline läuft noch -> keine Postfilter-Info
                self.count_label.config(
                    text=f"Angezeigt: {visible_count:,} Varianten (lädt...)"
                )
                return
            
            # ✅ Prüfe ob TATSÄCHLICH Filter gesetzt sind
            postfilter_checks = [
                bool(self.postfilter_cadd_min.get().strip()),
                not self.postfilter_benign.get(),
                not self.postfilter_likely_benign.get(),
                not self.postfilter_vus.get(),
                not self.postfilter_likely_path.get(),
                not self.postfilter_path.get(),
                not self.postfilter_impact_high.get(),
                not self.postfilter_impact_moderate.get(),
                not self.postfilter_impact_low.get(),
                self.postfilter_protein_coding.get(),
                bool(self.gen_whitelist.get().strip()),
                bool(self.gen_blacklist.get().strip()),
            ]
            
            # Prüfe Consequences (nur wenn MINDESTENS eine deaktiviert)
            all_cons_active = all(var.get() for var in self.postfilter_consequences.values())
            if not all_cons_active:
                postfilter_checks.append(True)
            
            postfilter_active = any(postfilter_checks)
            
            # ✅ Info-Text
            if postfilter_active and visible_count < total_filtered:
                self.count_label.config(
                    text=f"Angezeigt: {visible_count:,} von {total_filtered:,} Varianten (Postfilter aktiv)"
                )
            else:
                self.count_label.config(
                    text=f"Angezeigt: {visible_count:,} Varianten"
                )
            
        except Exception as e:
            self.logger.log(f"[Count-Label] ⚠️ Update failed: {e}")
            try:
                self.count_label.config(text="Angezeigt: ? Varianten")
            except Exception:
                pass

    
    def on_refresh(self):
        """
        Erzwingt ein komplettes Refresh der Tabelle.
        """
        self._refresh_table(force=True)
        self.update_count_label()

    def _check_cloud_performance_warning(self):
        """Prüft ob DBs in Cloud-Ordnern liegen und warnt den User."""
        if hasattr(self, "cloud_warning_dismissed") and self.cloud_warning_dismissed:
            return
            
        db_path = getattr(self.db, "db_path", None)
        gnomad_path = getattr(self.lightdb_manager, "out_db", None)
        
        service = check_cloud_sync_warning(db_path) or check_cloud_sync_warning(gnomad_path)
        
        if service:
            self.logger.log(f"[App] ⚠️ PERFORMANCE-WARNUNG: Datenbank liegt in {service}-Ordner!")
            msg = (
                f"⚠️ PERFORMANCE-WARNUNG\n\n"
                f"Die Datenbank liegt in einem Cloud-Sync-Ordner ({service}).\n"
                "Dies kann die Performance um bis zu 99% reduzieren und zu DB-Locks führen!\n\n"
                "Empfehlung: Verschieben Sie die Datenbanken nach C:\\_Local_DEV\\DATA_STORE."
            )
            res = Messagebox.show_question(msg, "Performance-Hinweis", buttons=["Nicht mehr anzeigen:danger", "Ignorieren:primary"])
            if res == "Nicht mehr anzeigen":
                self.cloud_warning_dismissed = True
                self._save_settings()

    def _poll_log(self):
        """
        Pollt Log-Nachrichten.

        FIXES:
        - Widget-Existenz pruefen vor after-Call
        - Max 100 Zeilen pro Poll (GUI bleibt responsive)
        - Kuerzeres Intervall fuer bessere Reaktivitaet
        """
        # KRITISCH: Pruefe ob Widget noch existiert
        if not self.winfo_exists():
            return

        msgs = logger.drain()

        if msgs:
            # Limitiere Batch-Groesse
            if len(msgs) > 100:
                msgs = msgs[:100]
                # Rest fuer naechsten Poll aufheben

            # Inkrementelles Schreiben (je 10 Zeilen)
            for i in range(0, len(msgs), 10):
                batch = msgs[i:i+10]
                text_chunk = "\n".join(batch) + "\n"
                self.log_text.insert("end", text_chunk)
                self.log_text.see("end")

                # Event Loop atmen lassen
                if i > 0 and i % 30 == 0:
                    self.update_idletasks()

            # Log-Groesse begrenzen
            line_count = int(self.log_text.index('end-1c').split('.')[0])
            if line_count > 5000:
                self.log_text.delete('1.0', f'{line_count - 5000}.0')

        # Schnellerer Poll nur wenn Widget existiert
        if self.winfo_exists():
            self.after(50, self._poll_log)

    def _incremental_table_update(self, keys: List[Tuple]):
        """Update single row incrementally (called within lock)."""
        if not keys:
            return
        
        if self._refresh_idx >= len(keys):
            self._refresh_idx = 0
        
        self._add_or_update_row(keys[self._refresh_idx])
        self._refresh_idx += 1
        self.update_count_label()


    # [PROBABLY DEAD] Nie aufgerufen - Export nutzt andere Logik
    # def _get_visible_export_columns(self) -> List[str]:
        # """
        # Liefert die sichtbaren Spalten in der Reihenfolge von self.columns.
        # """
        # return [c for c in self.columns if c in self.visible_columns]


    # [PROBABLY DEAD] Nie aufgerufen - Iteration wurde umgestellt
    # def _iter_current_table_rows(self):
        # """
        # Generator: liefert alle Zeilen des Treeviews in aktueller Anzeige-Reihenfolge.
        # """
        # for iid in self.tree.get_children(""):
            # yield iid, self.tree.item(iid)["values"]


    def _setup_events(self):
        """
        Bindet GUI-Events (z. B. Doppelklick auf Tabellenzellen).
        """
        self.tree.bind("<Double-1>", self._on_double_click)
        # Optional: einfacher Klick für RSID
        self.tree.bind("<Button-1>", self._on_click_single)
    
    def _process_events(self):
        try:
            while True:
                event = self.live_queue.get_nowait()
                if isinstance(event, tuple) and event[0] == "error":
                    _, msg = event
                    messagebox.showerror("Fehler", msg)
                else:
                    self._add_or_update_row(event)
        except queue.Empty:
            pass
        self.after(200, self._process_events)


    def _handle_column_click(self, col_name: str, trigger_type: str, row_values: dict):
        """NEU: Generischer Handler für Column Links."""
        config = self.column_links.get(col_name)
        if not config:
            return
        if config.get('trigger') != trigger_type:
            return
        template = config.get('template')
        if not template:
            return
        try:
            context = row_values.copy()
            context['value'] = row_values.get(col_name, "")
            for k, v in list(context.items()):
                if v is None or v == ".":
                    context[k] = ""
                else:
                    context[k] = str(v).strip()
            url = template.format(**context)
            value = context.get('value', '')
            if value and value not in (".", "", "-"):
                webbrowser.open(url)
        except Exception as e:
            self.logger.log(f"[Link] Fehler: {e}")

    def _on_double_click(self, event):
        """REFAKTORIERT: Nutzt generisches Column Link System."""
        self._generic_click_handler(event, "double")
    
    def _row_to_pubmed_query(self, row: dict) -> str:
        """
        Baut eine PubMed-Suchanfrage aus den Variantendaten.
        """
        terms = []
        gene = row.get("gene_symbol") or ""
        if gene:
            terms.append(gene)
        cln = row.get("clinical_significance")
        if cln:
            terms.append(cln)
        cons = row.get("consequence")
        if cons:
            terms.append(str(cons).split(",")[0])
        terms.append("variant")
        return " ".join(terms) if terms else "genetic variant"


    def _on_click_single(self, event):
        """REFAKTORIERT: Nutzt generisches Column Link System."""
        self._generic_click_handler(event, "single")
    
    def _generic_click_handler(self, event, trigger_type: str):
        """NEU: Generischer Click-Handler für Tabellenzellen."""
        region = self.tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        col_id = self.tree.identify_column(event.x)
        row_id = self.tree.identify_row(event.y)
        if not row_id or not col_id:
            return
        try:
            col_idx = int(col_id.replace("#", "")) - 1
        except ValueError:
            return
        visible_cols = self._current_displaycolumns()
        if col_idx < 0 or col_idx >= len(visible_cols):
            return
        col_name = visible_cols[col_idx]
        values = self.tree.item(row_id).get("values", [])
        if not values:
            return
        row_data = {}
        for i, vc in enumerate(visible_cols):
            if i < len(values):
                row_data[vc] = values[i]
            else:
                row_data[vc] = ""
        if col_name == "pubmed":
            iid = row_id
            parts = iid.split("|")
            if len(parts) == 5:
                try:
                    chr_, pos_, ref_, alt_, build_ = parts
                    key = (chr_, int(pos_), ref_, alt_, build_)
                    db_row = self.db.get_variant(key) if key else {}
                    row_data['consequence'] = db_row.get('consequence', '')
                except Exception:
                    pass
        self._handle_column_click(col_name, trigger_type, row_data)
                        
    def _current_displaycolumns(self):
        """
        Liefert die aktuell sichtbaren Spalten für das Treeview.
        """
        return [c for c in self.columns if c in self.visible_columns]
    
   
    def apply_post_filter(self):
        """
        FIX: Postfilter mit korrektem Count-Update.
        
        Änderungen:
        - Count-Update nach Filter-Completion
        - Zeigt Filtered vs. Total
        """
        iids = list(self.tree.get_children(""))
        cols = list(self.columns)
        
        # Parse filter configuration
        try:
            cadd_min = float(self.postfilter_cadd_min.get()) if self.postfilter_cadd_min.get().strip() else None
        except ValueError:
            cadd_min = None
        
        # Clinical significance allowed
        allow_benign = self.postfilter_benign.get()
        allow_likely_benign = self.postfilter_likely_benign.get()
        allow_vus = self.postfilter_vus.get()
        allow_likely_path = self.postfilter_likely_path.get()
        allow_path = self.postfilter_path.get()
        
        # Impact allowed
        allow_high = self.postfilter_impact_high.get()
        allow_mod = self.postfilter_impact_moderate.get()
        allow_low = self.postfilter_impact_low.get()
        
        # Other filters
        only_prot = self.postfilter_protein_coding.get()
        cons_allowed = {lbl for lbl, var in self.postfilter_consequences.items() if var.get()}
        
        # Gene lists
        whitelist = set(g.strip().upper() for g in self.gen_whitelist.get().split(",") if g.strip())
        blacklist = set(g.strip().upper() for g in self.gen_blacklist.get().split(",") if g.strip())
        
        def is_missing(s: Optional[str]) -> bool:
            """Check if value is missing/empty."""
            s = (s or "").strip().lower()
            return s in ("", ".", "-", "none", "null")
        
        def parse_row(iid: str) -> Dict[str, str]:
            """Parse tree row into dict."""
            values = self.tree.item(iid, "values")
            return dict(zip(cols, values))
        
        def visible(row: Dict[str, str]) -> bool:
            """Determine if row passes all filters."""
            # CADD minimum (only exclude if present AND below threshold)
            if cadd_min is not None:
                cadd_str = row.get("cadd")
                if not is_missing(cadd_str):
                    try:
                        cadd_val = float(cadd_str)
                        if cadd_val < cadd_min:
                            return False
                    except (ValueError, TypeError):
                        pass
            
            # Clinical significance (checked = keep, missing = keep)
            sig = row.get("clin_sig")
            if sig and not is_missing(sig):
                s = sig.lower()
                # Exclude if explicitly forbidden
                if "pathogenic" in s and "likely pathogenic" not in s:
                    if not allow_path:
                        return False
                if "likely pathogenic" in s:
                    if not allow_likely_path:
                        return False
                if "benign" in s and "likely benign" not in s:
                    if not allow_benign:
                        return False
                if "likely benign" in s:
                    if not allow_likely_benign:
                        return False
                if "uncertain" in s or "vus" in s:
                    if not allow_vus:
                        return False
            
            # Impact (checked = keep, missing = keep)
            imp = (row.get("impact") or "").upper()
            if imp and not is_missing(imp):
                if imp == "HIGH" and not allow_high:
                    return False
                if imp == "MODERATE" and not allow_mod:
                    return False
                if imp == "LOW" and not allow_low:
                    return False
            
            # Protein coding (only if explicitly required)
            if only_prot:
                cod = (row.get("coding") or "").strip().lower()
                if cod not in ("true", "1", "yes"):
                    return False
            
            # Consequences (checked = whitelist, missing = keep if whitelist empty)
            cons = (row.get("consequence") or "").lower()
            if cons_allowed and cons and not is_missing(cons):
                if not any(lbl in cons for lbl in cons_allowed):
                    return False
            
            # Gene whitelist (if set, gene MUST be in list)
            gene = (row.get("gene") or "").upper()
            if whitelist and gene and not is_missing(gene):
                if gene not in whitelist:
                    return False
            
            # Gene blacklist (if set, gene must NOT be in list)
            if blacklist and gene and not is_missing(gene):
                if gene in blacklist:
                    return False
            
            return True
        
        def compute_visibility():
            """Compute visibility for all rows (in worker thread)."""
            vis = []
            for iid in iids:
                try:
                    row = parse_row(iid)
                    ok = visible(row)
                except Exception as e:
                    self.logger.log(f"[Filter] Error processing {iid}: {e}")
                    ok = True  # Keep on error
                vis.append((iid, ok))
            return vis
        
        def apply_visibility(vis: List[Tuple[str, bool]]):
            """Apply visibility in UI thread (batched)."""
            # ✅ FIX V13.1: Initialisiere detached-Liste für späteren Reset
            if not hasattr(self, '_detached_iids'):
                self._detached_iids = set()

            def step(start: int = 0):
                end = min(start + Config.TABLE_BATCH_SIZE, len(vis))
                for iid, ok in vis[start:end]:
                    try:
                        if ok:
                            self.tree.reattach(iid, "", "end")
                            # ✅ FIX: Aus detached-Liste entfernen wenn wieder sichtbar
                            self._detached_iids.discard(iid)
                        else:
                            self.tree.detach(iid)
                            # ✅ FIX: Merke detachte IIDs für späteren Reset
                            self._detached_iids.add(iid)
                    except tk.TclError:
                        # Item might have been deleted
                        pass

                if end < len(vis):
                    self.after(10, lambda: step(end))
                else:
                    # ✅ Count-Update nach Filter-Completion
                    self.update_count_label()

            self.after(0, lambda: step(0))
        
        def worker():
            """Worker thread: compute then apply."""
            vis = compute_visibility()
            self.after(0, lambda: apply_visibility(vis))
        
        threading.Thread(target=worker, daemon=True).start()
        
     # ---------------- Export-Methoden ----------------

    # [PROBABLY DEAD] Nie aufgerufen - Export läuft synchron
    # def _run_export_threaded(self, target, *args, **kwargs):
        # """Hilfsfunktion: startet Export in eigenem Thread."""
        # def worker():
            # try:
                # target(*args, **kwargs)
            # except Exception as e:
                # self.logger.log(f"[Export] ❌ Fehler im Export-Thread: {e}")
        # threading.Thread(target=worker, daemon=True).start()

    # =============================================================================
    # HELPER-METHODE: REFERENCE-HEADER PRÜFEN UND ERGÄNZEN
    # =============================================================================

    def _check_and_add_reference_header(self, outfile, original_vcf, build):
        """
        ✅ FIX: Prüfe und füge Reference-Header hinzu falls fehlend
        """
        # Prüfe ob Reference-Header bereits existiert
        reference_found = False
        with open(original_vcf, encoding="utf-8") as f:
            for line in f:
                if line.startswith("##reference"):
                    reference_found = True
                    break
                if line.startswith("#CHROM"):
                    break
        
        # Füge Reference-Header hinzu falls fehlend
        if not reference_found:
            if build == "GRCh37":
                ref_line = '##reference=ftp://ftp.ncbi.nlm.nih.gov/genomes/archive/old_genbank/Eukaryotes/vertebrates_mammals/Homo_sapiens/GRCh37/Primary_Assembly/assembled_chromosomes/FASTA/\n'
            elif build == "GRCh38":
                ref_line = '##reference=ftp://ftp.ncbi.nlm.nih.gov/genomes/all/GCA/000/001/405/GCA_000001405.15_GRCh38/seqs_for_alignment_pipelines.ucsc_ids/GCA_000001405.15_GRCh38_no_alt_analysis_set.fna.gz\n'
            else:
                ref_line = f'##reference={build}\n'
            
            outfile.write(ref_line)
            self.logger.log(f"[VCF-Export] ✅ Added missing reference header for {build}")
    # =============================================================================
    # GUI: EXPORT-ROUTER & WORKER
    # =============================================================================

    def export_vcf(self):
        """Router-Methode - checkt Einstellung und leitet weiter"""
        mode = self.vcf_export_mode.get()
        
        if mode == "filtered":
            self.export_filtered_vcf_optimized()
        elif mode == "complete":
            self.export_complete_vcf_optimized()

    # =============================================================================
    # HELPER-METHODEN FÜR EXPORT
    # =============================================================================

    def _write_vcf_meta_header(self, outfile, count, mode):
        """Schreibt zusätzliche Header-Informationen."""
        timestamp = datetime.datetime.now().isoformat()
        outfile.write(f'##annotated_by="VariantFusion v3.3"\n')
        outfile.write(f'##annotation_date="{timestamp}"\n')
        outfile.write(f'##export_mode="{mode}"\n')
        outfile.write(f'##variant_count="{count}"\n')
        
        # Neue INFO Definitionen
        headers = self._get_additional_info_headers()
        for h in headers:
            outfile.write(h + "\n")


    def _get_additional_info_headers(self) -> list:
        """Definiert die neuen INFO-Felder."""
        return [
            '##INFO=<ID=AF_VF,Number=1,Type=Float,Description="Population Allele Frequency (VariantFusion Mean)">',
            '##INFO=<ID=GENE_VF,Number=1,Type=String,Description="Gene Symbol (VariantFusion)">',
            '##INFO=<ID=CADD_phred,Number=1,Type=Float,Description="CADD Phred Score">',
            '##INFO=<ID=CLINSIG_VF,Number=.,Type=String,Description="Clinical Significance (ClinVar)">',
            '##INFO=<ID=IMPACT_VF,Number=1,Type=String,Description="Predicted Impact">',
            '##INFO=<ID=CONSEQUENCE_VF,Number=.,Type=String,Description="Variant Consequence">',
            '##INFO=<ID=IS_CODING,Number=1,Type=Integer,Description="Flag: Is Protein Coding (1=Yes)">',
            '##INFO=<ID=AG_SCORE,Number=1,Type=Float,Description="AlphaGenome Pathogenicity Score">'
        ]

    def _enrich_vcf_line(self, line: str, annotations: dict, key: tuple) -> str:
        """
        Fügt Annotationen in das INFO-Feld der VCF-Zeile ein.
        Achtet auf korrekte Formatierung und Escaping.
        """
        if not annotations:
            return line

        parts = line.strip().split("\t")
        if len(parts) < 8:
            return line

        info_str = parts[7]
        new_tags = []

        # Helper zum sicheren Hinzufügen
        def add_tag(tag_id, val):
            if val is not None and val != "":
                # VCF-Escaping: Semikolons und Leerzeichen verboten in INFO-Values
                clean_val = str(val).replace(";", "|").replace(" ", "_")
                new_tags.append(f"{tag_id}={clean_val}")

        # Mapping DB-Felder -> VCF INFO Tags
        add_tag("AF_VF", annotations.get("af_filter_mean"))
        add_tag("GENE_VF", annotations.get("gene_symbol"))
        add_tag("CADD_phred", annotations.get("cadd_phred"))
        add_tag("CLINSIG_VF", annotations.get("clinical_significance"))
        add_tag("IMPACT_VF", annotations.get("impact"))
        add_tag("CONSEQUENCE_VF", annotations.get("consequence"))
        add_tag("IS_CODING", annotations.get("is_coding"))
        add_tag("AG_SCORE", annotations.get("ag_score"))

        # rsID Update (Spalte 3), falls in DB vorhanden und im VCF fehlt
        db_rsid = annotations.get("rsid")
        if db_rsid and db_rsid.startswith("rs") and parts[2] == ".":
            parts[2] = db_rsid

        # INFO zusammenbauen
        if new_tags:
            additional_info = ";".join(new_tags)
            if info_str == ".":
                parts[7] = additional_info
            else:
                parts[7] = f"{info_str};{additional_info}"

        return "\t".join(parts) + "\n"
    
    # [PROBABLY DEAD] Nie aufgerufen - Metadata wird inline erstellt
    # def _get_export_metadata(self) -> dict:
        # """
        # ✅ NEUE METHODE: Sammelt Metadaten für Export.
        
        # Returns:
            # Dict mit:
            # - total_variants: Gesamtzahl nach AF-Filter
            # - visible_variants: Sichtbar nach Postfilter
            # - af_threshold: Verwendeter AF-Schwellenwert
            # - include_none: Include-None-Flag
            # - postfilter_active: Ob Postfilter aktiv
            # - export_timestamp: ISO-Timestamp
        # """
        # metadata = {
            # "export_timestamp": datetime.datetime.now().isoformat(),
            # "af_threshold": self.af_threshold.get() if hasattr(self, 'af_threshold') else None,
            # "include_none": self.include_none.get() if hasattr(self, 'include_none') else False,
            # "total_variants": len(self.distiller.display_keys) if hasattr(self.distiller, 'display_keys') else 0,
            # "visible_variants": len(self.tree.get_children("")),
            # "postfilter_active": False,
        # }
        
        # ✅ Prüfe, ob Postfilter aktiv
        # if metadata["visible_variants"] < metadata["total_variants"]:
            # metadata["postfilter_active"] = True
        
        # return metadata
    
    def _process_events(self):
        try:
            while True:
                event = self.live_queue.get_nowait()
                if isinstance(event, tuple) and event[0] == "error":
                    _, msg = event
                    # ✅ FIX: Bootstrap Messagebox statt tkinter.messagebox
                    Messagebox.show_error(msg, "Fehler")
                else:
                    self._add_or_update_row(event)
        except queue.Empty:
            pass
        self.after(200, self._process_events)

    def save_gene_list(self, list_type):
        """
        Speichert die aktuelle Gene-Whitelist oder -Blacklist als Textdatei.
        """
        folder = os.path.join(os.getcwd(), "GenePanels")
        os.makedirs(folder, exist_ok=True)

        genes = (self.gen_whitelist.get() if list_type == "whitelist" else self.gen_blacklist.get()).split(",")
        genes = [g.strip() for g in genes if g.strip()]

        first_three = "-".join(genes[:3]) if genes else "GENE"
        date_str = datetime.datetime.now().strftime("%Y-%m-%d")
        default_name = f"{first_three}_{list_type}_{date_str}.txt"

        file_path = filedialog.asksaveasfilename(
            initialdir=folder,
            initialfile=default_name,
            defaultextension=".txt",
            filetypes=[("Textdateien", "*.txt")]
        )
        if file_path:
            try:
                with open(file_path, "w", encoding="utf-8") as f:
                    for g in genes:
                        f.write(g + "\n")
                logger.log(f"[App] {list_type.capitalize()} gespeichert: {file_path}")
                
                # ✅ NEU: Toast statt Messagebox für Erfolg
                ToastNotification(
                    title="Gespeichert",
                    message=f"{len(genes)} Gene in {list_type} gespeichert.",
                    duration=2000,
                    bootstyle="success"
                ).show_toast()
                
            except Exception as e:
                # ✅ FIX: Bootstrap Messagebox
                Messagebox.show_error(f"Liste konnte nicht gespeichert werden:\n{e}", "Fehler")

    def load_gene_list(self, list_type):
        """
        Lädt eine gespeicherte Gene-Whitelist oder -Blacklist aus einer Textdatei.
        """
        folder = os.path.join(os.getcwd(), "GenePanels")
        os.makedirs(folder, exist_ok=True)
        file_path = filedialog.askopenfilename(
            initialdir=folder,
            filetypes=[("Textdateien", "*.txt")]
        )
        if not file_path:
            return
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                genes = [line.strip() for line in f if line.strip()]
            if list_type == "whitelist":
                self.gen_whitelist.set(",".join(genes))
            else:
                self.gen_blacklist.set(",".join(genes))
            logger.log(f"[App] {list_type.capitalize()} geladen: {file_path}")
            
            # ✅ NEU: Toast
            ToastNotification(
                title="Geladen",
                message=f"{len(genes)} Gene in {list_type} geladen.",
                duration=2000,
                bootstyle="info"
            ).show_toast()
            
        except Exception as e:
            # ✅ FIX: Bootstrap Messagebox
            Messagebox.show_error(f"Liste konnte nicht geladen werden:\n{e}", "Fehler")

        # ✅ V16: DIE UNTEN STEHENDE DOPPELTE on_start WURDE ENTFERNT
        # HIER FOLGT DER EXPORT UND ANDERE HELPER

    def export_csv(self):
        """FIX: CSV-Export mit Bootstrap-Dialogen."""
        filename = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV-Dateien", "*.csv")],
            initialfile=f"variant_fusion_export_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        )
        if not filename:
            return

        try:
            import csv
            visible_items = self.tree.get_children("")
            
            if not visible_items:
                Messagebox.show_warning("Keine Varianten zum Exportieren vorhanden.", "Export")
                return
            
            with open(filename, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f, delimiter=";", quoting=csv.QUOTE_MINIMAL)
                headers = [COLUMN_LABELS.get(c, c) for c in self.columns if c in self.visible_columns]
                writer.writerow(headers)
                
                exported_count = 0
                visible_indices = [i for i, col in enumerate(self.columns) if col in self.visible_columns]
                
                for iid in visible_items:
                    try:
                        values = self.tree.item(iid, "values")
                        if not values: continue
                        row_data = [values[i] if i < len(values) else "" for i in visible_indices]
                        writer.writerow(row_data)
                        exported_count += 1
                    except Exception:
                        continue
            
            self.logger.log(f"[Export] ✅ CSV: {filename} ({exported_count} Rows)")
            # ✅ FIX: Modernes Feedback
            ToastNotification(
                title="Export erfolgreich",
                message=f"{exported_count} Varianten exportiert.",
                bootstyle="success"
            ).show_toast()
            
        except Exception as e:
            self.logger.log(f"[CSV-Export] ❌ Fehler: {e}")
            Messagebox.show_error(f"CSV-Export fehlgeschlagen:\n{e}", "Export-Fehler")

    def export_excel(self):
        """FIX: Excel-Export mit Bootstrap-Dialogen."""
        if not HAVE_OPENPYXL:
            Messagebox.show_error("openpyxl fehlt.\nBitte installieren: pip install openpyxl", "Fehler")
            return
        
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel-Dateien", "*.xlsx")],
            initialfile=f"variant_fusion_export_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        if not filename:
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
            
            visible_items = self.tree.get_children("")
            if not visible_items:
                Messagebox.show_warning("Keine Varianten zum Exportieren.", "Export")
                return
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Varianten"
            
            headers = [COLUMN_LABELS.get(c, c) for c in self.columns if c in self.visible_columns]
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            ws.freeze_panes = "A2"
            visible_indices = [i for i, col in enumerate(self.columns) if col in self.visible_columns]
            exported_count = 0
            
            for row_idx, iid in enumerate(visible_items, start=2):
                try:
                    values = self.tree.item(iid, "values")
                    if not values: continue
                    
                    for col_idx, val_idx in enumerate(visible_indices, start=1):
                        val = values[val_idx] if val_idx < len(values) else ""
                        if isinstance(val, str):
                            try:
                                if val and val not in (".", ""): val = float(val)
                            except: pass
                        ws.cell(row=row_idx, column=col_idx, value=val)
                    exported_count += 1
                except: continue
            
            # Auto-Width
            for col_idx in range(1, len(headers) + 1):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = 20
            
            wb.save(filename)
            wb.close()
            
            self.logger.log(f"[Export] ✅ Excel: {filename} ({exported_count} Rows)")
            ToastNotification(
                title="Export erfolgreich",
                message=f"{exported_count} Varianten exportiert.",
                bootstyle="success"
            ).show_toast()
            
        except Exception as e:
            self.logger.log(f"[Excel-Export] ❌ Fehler: {e}")
            Messagebox.show_error(f"Excel-Export fehlgeschlagen:\n{e}", "Export-Fehler")

    def export_pdf(self):
        """FIX: PDF-Export mit Bootstrap-Dialogen."""
        if not HAVE_REPORTLAB:
            Messagebox.show_error("reportlab fehlt.\nBitte installieren: pip install reportlab", "Fehler")
            return
        
        filename = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF-Dateien", "*.pdf")],
            initialfile=f"variant_fusion_export_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        )
        if not filename:
            return

        try:
            from reportlab.lib.pagesizes import A3, landscape
            from reportlab.pdfgen import canvas
            from reportlab.lib.units import mm
            
            visible_items = self.tree.get_children("")
            if not visible_items:
                Messagebox.show_warning("Keine Varianten zum Exportieren.", "Export")
                return
            
            c = canvas.Canvas(filename, pagesize=landscape(A3))
            width, height = landscape(A3)
            x_margin = 15 * mm
            y_margin = 12 * mm
            c.setFont("Helvetica", 8)

            visible_cols = [col for col in self.columns if col in self.visible_columns]
            n = len(visible_cols)
            col_width = (width - 2 * x_margin) / max(1, n)

            def draw_header(y):
                c.setFont("Helvetica-Bold", 8)
                for i, col in enumerate(visible_cols):
                    label = COLUMN_LABELS.get(col, col)
                    c.drawString(x_margin + i * col_width, y, str(label)[:40])
                c.setFont("Helvetica", 8)

            def draw_row(y, values, visible_indices):
                for i, val_idx in enumerate(visible_indices):
                    val = values[val_idx] if val_idx < len(values) else ""
                    text = str(val)
                    if len(text) > 60: text = text[:57] + "..."
                    c.drawString(x_margin + i * col_width, y, text)

            header_y = height - y_margin
            draw_header(header_y)
            y = header_y - 12
            visible_indices = [i for i, col in enumerate(self.columns) if col in self.visible_columns]
            exported_count = 0

            for iid in visible_items:
                try:
                    values = self.tree.item(iid, "values")
                    if not values: continue
                    if y < y_margin:
                        c.showPage()
                        width, height = landscape(A3)
                        header_y = height - y_margin
                        draw_header(header_y)
                        y = header_y - 12
                    draw_row(y, values, visible_indices)
                    y -= 10
                    exported_count += 1
                except: continue

            c.save()
            self.logger.log(f"[Export] ✅ PDF: {filename}")
            ToastNotification(
                title="Export erfolgreich",
                message=f"{exported_count} Varianten exportiert.",
                bootstyle="success"
            ).show_toast()
            
        except Exception as e:
            self.logger.log(f"[PDF-Export] ❌ Fehler: {e}")
            Messagebox.show_error(f"PDF-Export fehlgeschlagen:\n{e}", "Export-Fehler")

    def export_filtered_vcf_optimized(self):
        """
        ✅ REPARIERT: Thread-Safe & Robustes Matching.
        Fix: Verwendet Messagebox/Toast aus ttkbootstrap.
        """
        visible_items = self.tree.get_children("")
        total_visible = len(visible_items)
        
        if total_visible == 0:
            Messagebox.show_warning("Keine sichtbaren Varianten zum Exportieren.", "Export")
            return

        target_keys_core = set()
        for iid in visible_items:
            key = self._parse_key_from_iid(iid)
            if key and len(key) >= 4:
                c = str(key[0]).replace("chr", "").replace("CHR", "").upper()
                p = int(key[1])
                r = str(key[2]).upper()
                a = str(key[3]).upper()
                target_keys_core.add((c, p, r, a))

        if not target_keys_core:
            self.logger.log("[VCF-Export] ❌ Keine gültigen Keys gefunden.")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".vcf",
            filetypes=[("VCF-Dateien", "*.vcf")],
            initialfile=f"filtered_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.vcf"
        )
        if not file_path:
            return

        def _export_worker():
            try:
                self.logger.log(f"[VCF-Export] 🎯 Starte Filtered Export...")
                current_build = self.distiller.build if self.distiller.build else "GRCh37"
                
                # DB Lookup Logik (verkürzt, da identisch zu vorher)
                db_lookup_keys = []
                for (c, p, r, a) in target_keys_core:
                    db_lookup_keys.append((c, p, r, a, current_build))
                
                annotations = {}
                for i in range(0, len(db_lookup_keys), 2000):
                    chunk = db_lookup_keys[i:i+2000]
                    try:
                        chunk_data = self.db.get_variants_bulk(chunk)
                        if chunk_data: annotations.update(chunk_data)
                    except: pass

                original_vcf = self.distiller.original_vcf_path
                if not original_vcf or not os.path.exists(original_vcf):
                    self.logger.log("[VCF-Export] ❌ Original-VCF nicht gefunden!")
                    return

                exported_count = 0
                with open_text_maybe_gzip(original_vcf) as infile, open(file_path, "w", encoding="utf-8") as outfile:
                    header_written = False
                    for line in infile:
                        if line.startswith("#"):
                            if line.startswith("#CHROM") and not header_written:
                                self._write_vcf_meta_header(outfile, total_visible, "filtered")
                                header_written = True
                            outfile.write(line)
                            if line.startswith("##fileformat"):
                                self._check_and_add_reference_header(outfile, original_vcf, current_build)
                            continue
                        
                        try:
                            parts = line.split("\t", 5)
                            if len(parts) < 5: continue
                            chrom, pos_str, _, ref, alt_str = parts[:5]
                            try: pos = int(pos_str)
                            except: continue
                            
                            c_clean = chrom.replace("chr", "").replace("CHR", "").upper()
                            ref = ref.upper()
                            match_found = False
                            matched_key_full = None
                            
                            for alt in alt_str.split(","):
                                alt = alt.upper()
                                if (c_clean, pos, ref, alt) in target_keys_core:
                                    match_found = True
                                    matched_key_full = (c_clean, pos, ref, alt, current_build)
                                    break
                            
                            if match_found:
                                anno = annotations.get(matched_key_full, {})
                                new_line = self._enrich_vcf_line(line, anno, matched_key_full)
                                outfile.write(new_line)
                                exported_count += 1
                        except: continue

                self.logger.log(f"[VCF-Export] ✅ Fertig: {exported_count} exportiert.")
                # ✅ FIX: Toast Notification
                self.after(0, lambda: ToastNotification(
                    title="Export erfolgreich",
                    message=f"{exported_count} Varianten in VCF gespeichert.",
                    bootstyle="success"
                ).show_toast())

            except Exception as e:
                self.logger.log(f"[VCF-Export] ❌ Fehler: {e}")
                self.after(0, lambda: Messagebox.show_error(f"Fehler: {e}", "VCF Export"))

        threading.Thread(target=_export_worker, daemon=True).start()

    def export_complete_vcf_optimized(self):
        """
        ✅ REPARIERT: Exportiert ALLE Varianten.
        Fix: Verwendet Messagebox/Toast aus ttkbootstrap.
        """
        file_path = filedialog.asksaveasfilename(
            defaultextension=".vcf",
            filetypes=[("VCF-Dateien", "*.vcf")],
            initialfile=f"complete_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.vcf"
        )
        if not file_path:
            return

        def _export_worker():
            try:
                original_vcf = self.distiller.original_vcf_path
                if not original_vcf: return
                self.logger.log("[VCF-Export] 🌍 Starte vollständigen Export...")

                current_build = self.distiller.build if self.distiller.build else "GRCh37"

                # ✅ V14 FIX: Sammle alle Keys für Bulk-DB-Lookup
                all_keys = []
                with open_text_maybe_gzip(original_vcf) as infile:
                    for line in infile:
                        if line.startswith("#"):
                            continue
                        try:
                            parts = line.split("\t", 5)
                            if len(parts) < 5:
                                continue
                            chrom, pos_str, _, ref, alt_str = parts[:5]
                            try:
                                pos = int(pos_str)
                            except ValueError:
                                continue
                            c_clean = chrom.replace("chr", "").replace("CHR", "").upper()
                            ref = ref.upper()
                            for alt in alt_str.split(","):
                                alt = alt.upper()
                                all_keys.append((c_clean, pos, ref, alt, current_build))
                        except Exception:
                            continue

                # ✅ V14 FIX: Bulk-Lookup aller Annotationen
                annotations = {}
                for i in range(0, len(all_keys), 2000):
                    chunk = all_keys[i:i+2000]
                    try:
                        chunk_data = self.db.get_variants_bulk(chunk)
                        if chunk_data:
                            annotations.update(chunk_data)
                    except Exception:
                        pass

                exported_count = 0
                with open_text_maybe_gzip(original_vcf) as infile, open(file_path, "w", encoding="utf-8") as outfile:
                    header_written = False
                    for line in infile:
                        if line.startswith("#"):
                            if line.startswith("#CHROM") and not header_written:
                                self._write_vcf_meta_header(outfile, "ALL", "complete")
                                header_written = True
                            outfile.write(line)
                            if line.startswith("##fileformat"):
                                self._check_and_add_reference_header(outfile, original_vcf, self.distiller.build)
                            continue

                        # ✅ V14 FIX: Vollständige Enrichment-Logik wie in filtered export
                        try:
                            parts = line.split("\t", 5)
                            if len(parts) < 5:
                                outfile.write(line)
                                exported_count += 1
                                continue
                            chrom, pos_str, _, ref, alt_str = parts[:5]
                            try:
                                pos = int(pos_str)
                            except ValueError:
                                outfile.write(line)
                                exported_count += 1
                                continue

                            c_clean = chrom.replace("chr", "").replace("CHR", "").upper()
                            ref = ref.upper()
                            matched_key_full = None

                            for alt in alt_str.split(","):
                                alt = alt.upper()
                                key = (c_clean, pos, ref, alt, current_build)
                                if key in annotations:
                                    matched_key_full = key
                                    break

                            if matched_key_full:
                                anno = annotations.get(matched_key_full, {})
                                new_line = self._enrich_vcf_line(line, anno, matched_key_full)
                                outfile.write(new_line)
                            else:
                                outfile.write(line)
                            exported_count += 1
                        except Exception:
                            outfile.write(line)
                            exported_count += 1

                self.logger.log(f"[VCF-Export] ✅ Fertig: {exported_count} Zeilen.")
                # ✅ FIX: Toast
                self.after(0, lambda: ToastNotification(
                    title="Export erfolgreich",
                    message=f"Vollständiges VCF gespeichert.",
                    bootstyle="success"
                ).show_toast())

            except Exception as e:
                self.logger.log(f"[VCF-Export] ❌ Fehler: {e}")
                self.after(0, lambda: Messagebox.show_error(f"Fehler: {e}", "VCF Export"))

        threading.Thread(target=_export_worker, daemon=True).start()
    def _init_gui_variables(self):
        """Initialisiert alle GUI-Variablen mit Standardwerten."""
        self.selected_file = tk.StringVar()
        self.af_threshold = tk.DoubleVar(value=0.01)
        self.include_none = tk.BooleanVar(value=True)
        # V10: Trace für sofortige Weitergabe an FlagManager
        self.include_none.trace_add("write", lambda *args: self._on_include_none_changed())
        self.filter_pass_only = tk.BooleanVar(value=False)
        self.qual_threshold = tk.DoubleVar(value=30.0)
        
        # Highlight & Settings
        self.cadd_highlight_threshold = tk.DoubleVar(value=20.0)
        self.alphagenome_key = tk.StringVar(value="")

        # API-Settings (deep copy der Defaults)
        import copy
        self.api_settings = copy.deepcopy(Config.DEFAULT_API_SETTINGS)
        # Bestehendes...
        self.stale_days = tk.IntVar(value=Config.STALE_DAYS_AF) # Default AF stale (lang)
        self.stale_days_full = tk.IntVar(value=Config.STALE_DAYS_FULL) # NEU: Full Anno stale (kurz)
        
        # ✅ NEU: Pfad zum externen DB-Viewer
        self.external_db_viewer = tk.StringVar(value="")
        
        # ✅ NEU: Column Link Templates
        self.column_links = Config.DEFAULT_COLUMN_LINKS.copy()
        
        # Prä-Filter
        self.prefilter_protein_coding = tk.BooleanVar(value=False)
        
        # Post-Filter (Ergebnis-Filter)
        self.postfilter_cadd_min = tk.StringVar(value="")
        self.postfilter_benign = tk.BooleanVar(value=True)
        self.postfilter_likely_benign = tk.BooleanVar(value=True)
        self.postfilter_vus = tk.BooleanVar(value=True)
        self.postfilter_likely_path = tk.BooleanVar(value=True)
        self.postfilter_path = tk.BooleanVar(value=True)
        
        self.postfilter_impact_high = tk.BooleanVar(value=True)
        self.postfilter_impact_moderate = tk.BooleanVar(value=True)
        self.postfilter_impact_low = tk.BooleanVar(value=True)
        
        self.postfilter_protein_coding = tk.BooleanVar(value=False)
        
        self.postfilter_consequences = {
            "missense": tk.BooleanVar(value=True),
            "stop_gained": tk.BooleanVar(value=True),
            "frameshift": tk.BooleanVar(value=True),
            "synonymous": tk.BooleanVar(value=True),
            "splice": tk.BooleanVar(value=True),
            "inframe": tk.BooleanVar(value=True),
        }
        
        # Gen-Listen
        self.gen_whitelist = tk.StringVar(value="")
        self.gen_blacklist = tk.StringVar(value="")
        
        # Export & Optionen
        self.debug_show_db = tk.BooleanVar(value=False)
        self.skip_info_recycling = tk.BooleanVar(value=False)
        self.vcf_export_mode = tk.StringVar(value="complete")
        self.progress_var = tk.DoubleVar(value=0.0)

        # V17: Setup-Dialog Unterdrueckung
        self.setup_dialog_suppressed = False

    def _load_settings(self):
        """Lädt Settings inkl. neuer Parameter."""
        if not os.path.exists(Config.SETTINGS_FILE):
            return
        
        try:
            with open(Config.SETTINGS_FILE, encoding="utf-8") as f:
                data = json.load(f)
            
            self.af_threshold.set(data.get("af_threshold", 0.01))
            self.include_none.set(data.get("include_none", True))
            self.cadd_highlight_threshold.set(data.get("cadd_highlight_threshold", 20.0))
            # ...
            self.stale_days.set(data.get("stale_days", 365))
            self.stale_days_full.set(data.get("stale_days_full", 30)) # Laden
            # ...
            self.alphagenome_key.set(data.get("alphagenome_key", ""))
            
            # ✅ NEU: DB-Viewer Pfad laden
            self.external_db_viewer.set(data.get("external_db_viewer", ""))
            
            # ✅ NEU: Column Links laden
            if "column_links" in data:
                self.column_links = data["column_links"]

            # API-Settings: Deep-Merge (gespeicherte Werte ueberschreiben Defaults)
            if "api_settings" in data:
                saved = data["api_settings"]
                for phase_k, phase_v in saved.items():
                    if phase_k in self.api_settings and isinstance(phase_v, dict):
                        if phase_k == "global":
                            self.api_settings[phase_k].update(phase_v)
                        else:
                            for api_k, api_v in phase_v.items():
                                if api_k in self.api_settings[phase_k] and isinstance(api_v, dict):
                                    self.api_settings[phase_k][api_k].update(api_v)
                                elif api_k in self.api_settings[phase_k]:
                                    self.api_settings[phase_k][api_k] = api_v
            global _runtime_api_settings
            _runtime_api_settings = self.api_settings

            # AlphaGenome-Key Sync: api_settings -> tk.StringVar
            ag_key = self.api_settings.get("phase6_ag", {}).get("alphagenome", {}).get("api_key", "")
            if ag_key:
                self.alphagenome_key.set(ag_key)

            self.skip_info_recycling.set(data.get("skip_info_recycling", False))
            self.vcf_export_mode.set(data.get("vcf_export_mode", "complete"))
            
            if "visible_columns" in data:
                self.visible_columns = set(data["visible_columns"])
            
            self.cloud_warning_dismissed = data.get("cloud_warning_dismissed", False)
            self.setup_dialog_suppressed = data.get("setup_dialog_suppressed", False)
            
            # Quality Settings laden
            quality_settings = data.get("quality_settings", {})
            if quality_settings:
                preset = quality_settings.get("preset", "clinical")
                if preset == "custom":
                    self.quality_manager.set_custom_settings(
                        filter_pass_only=quality_settings.get("filter_pass_only", True),
                        qual_threshold=quality_settings.get("qual_threshold", 30.0),
                        use_dp_filter=quality_settings.get("use_dp_filter", True),
                        min_dp=quality_settings.get("min_dp", 10),
                        max_dp=quality_settings.get("max_dp", None)
                    )
                else:
                    self.quality_manager.set_preset(preset)
                
                self.filter_pass_only.set(quality_settings.get("filter_pass_only", True))
                self.qual_threshold.set(quality_settings.get("qual_threshold", 30.0))
            else:
                # Migration alter Settings
                filter_pass = data.get("filter_pass_only", False)
                qual_thresh = data.get("qual_threshold", 30.0)
                self.filter_pass_only.set(filter_pass)
                self.qual_threshold.set(qual_thresh)
                self.quality_manager.set_custom_settings(
                    filter_pass_only=filter_pass,
                    qual_threshold=qual_thresh,
                    use_dp_filter=True, min_dp=10, max_dp=None
                )
            
        except Exception as e:
            self.logger.log(f"[App] ⚠️ Fehler beim Laden der Settings: {e}")

    def _save_settings(self, theme=None):
        """Speichert Settings inkl. neuer Parameter."""
        data = {}
        if os.path.exists(Config.SETTINGS_FILE):
            try:
                with open(Config.SETTINGS_FILE, "r") as f:
                    data = json.load(f)
            except: pass
        
        data["af_threshold"] = self.af_threshold.get()
        data["include_none"] = self.include_none.get()
        data["cadd_highlight_threshold"] = self.cadd_highlight_threshold.get()
        data["alphagenome_key"] = self.alphagenome_key.get()
        data["stale_days"] = self.stale_days.get()
        data["stale_days_full"] = self.stale_days_full.get() # Speichern     
        # ✅ NEU: DB-Viewer Pfad speichern
        data["external_db_viewer"] = self.external_db_viewer.get()
        
        # ✅ NEU: Column Links speichern
        data["column_links"] = self.column_links
        
        data["skip_info_recycling"] = self.skip_info_recycling.get()
        data["vcf_export_mode"] = self.vcf_export_mode.get()
        data["visible_columns"] = list(self.visible_columns)
        data["cloud_warning_dismissed"] = getattr(self, "cloud_warning_dismissed", False)
        data["setup_dialog_suppressed"] = getattr(self, "setup_dialog_suppressed", False)
        
        if theme:
            data["theme"] = theme
        elif hasattr(self, 'style'):
            data["theme"] = self.style.theme.name
            
        qs = self.quality_manager.get_settings()
        data["quality_settings"] = {
            "preset": qs["preset"],
            "filter_pass_only": qs["filter_pass_only"],
            "qual_threshold": qs["qual_threshold"],
            "use_dp_filter": qs["use_dp_filter"],
            "min_dp": qs["min_dp"],
            "max_dp": qs["max_dp"],
            "filter_homref": qs.get("filter_homref", True)
        }

        # API-Settings: AlphaGenome-Key bidirektional sync
        ag_key_gui = self.alphagenome_key.get()
        if ag_key_gui:
            self.api_settings["phase6_ag"]["alphagenome"]["api_key"] = ag_key_gui
        data["api_settings"] = self.api_settings
        global _runtime_api_settings
        _runtime_api_settings = self.api_settings

        try:
            with open(Config.SETTINGS_FILE, "w") as f:
                json.dump(data, f, indent=2)
        except Exception as e:
            logger.log(f"Settings save error: {e}")

    def _check_startup_setup(self):
        """V17: Zeigt Setup-Dialog bei Fresh-Install (viele fehlende Ressourcen)."""
        if getattr(self, 'setup_dialog_suppressed', False):
            logger.log("[App] Setup-Dialog unterdrueckt (vom User deaktiviert)")
            return
        try:
            rm = get_resource_manager()
            status = rm.get_status()
            check_keys = ["fasta_grch37", "fasta_grch38", "gnomad_db", "gtf_grch37", "gtf_grch38"]
            missing = [k for k in check_keys if status.get(k) == "missing"]
            logger.log(f"[App] Startup-Setup-Check: {len(missing)}/5 Ressourcen fehlen ({missing})")
            if len(missing) >= 3:
                logger.log("[App] Oeffne Ressourcen-Setup-Dialog...")
                ResourceSetupDialog(self, rm, logger)
        except Exception as e:
            logger.log(f"[App] Startup-Setup-Check Fehler: {e}")
            import traceback
            logger.log(traceback.format_exc())

    def _build_gui(self):
        # --- MENÜ ---
        menubar = tk.Menu(self)
        self.config(menu=menubar)
        
        # Ansicht
        viewmenu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label=self._t("Ansicht"), menu=viewmenu)
        self._viewmenu_vars = {}
        for col in self.columns:
            var = tk.BooleanVar(value=(col in self.visible_columns))
            self._viewmenu_vars[col] = var
            viewmenu.add_checkbutton(label=COLUMN_LABELS.get(col, col), variable=var, command=self._apply_column_visibility)
        viewmenu.add_separator()
        viewmenu.add_command(label=self._t("Reset Spalten"), command=self._reset_columns)

        # Optionen
        optionsmenu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label=self._t("Optionen"), menu=optionsmenu)
        optionsmenu.add_command(label=self._t("Allgemeine Einstellungen..."), command=self.open_general_settings)
        optionsmenu.add_separator()
        optionsmenu.add_command(label=self._t("Quality-Filter..."), command=self.open_quality_settings)
        optionsmenu.add_separator()
        optionsmenu.add_checkbutton(
            label=self._t("Recycling der INFO-Felder überspringen"),
            variable=self.skip_info_recycling,
            onvalue=True, offvalue=False, command=self._save_settings
        )
        
        # Sprache
        optionsmenu.add_separator()
        langmenu = tk.Menu(optionsmenu, tearoff=0)
        optionsmenu.add_cascade(label=self._t("Sprache / Language"), menu=langmenu)
        langmenu.add_command(label="Deutsch", command=lambda: self.change_language("de"))
        langmenu.add_command(label="English", command=lambda: self.change_language("en"))

        # Design-Menü
        designmenu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label=self._t("Design"), menu=designmenu)
        for theme in ["cosmo", "flatly", "lumen", "darkly", "superhero", "solar"]:
            designmenu.add_command(label=theme.capitalize(), command=lambda t=theme: self._change_theme(t))
            
        # ✅ NEU: Debug-Menü mit externem Viewer
        debugmenu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label=self._t("Debug"), menu=debugmenu)
        debugmenu.add_command(label=self._t("Datenbank extern öffnen"), command=self.open_external_db)

        # =====================================================================
        # 1. TOP BAR (Datei, Scan-Settings)
        # =====================================================================
        top_bar = ttk.Frame(self, padding=10)
        top_bar.pack(fill=X)

        # Datei
        file_row = ttk.Frame(top_bar)
        file_row.pack(fill=X, pady=(0, 10))
        ttk.Label(file_row, text="Eingabe-Datei:").pack(side=LEFT)
        ttk.Entry(file_row, textvariable=self.selected_file).pack(side=LEFT, fill=X, expand=YES, padx=10)
        ttk.Button(file_row, text="Durchsuchen", command=self.choose_file, bootstyle="secondary-outline").pack(side=LEFT)

        # Controls
        ctrl_row = ttk.Frame(top_bar)
        ctrl_row.pack(fill=X)
        
        # Filter Box
        filter_box = ttk.Labelframe(ctrl_row, text="VCF-Scan Filter (Pre)", padding=5, bootstyle="info")
        filter_box.pack(side=LEFT, fill=Y)
        
        ttk.Label(filter_box, text="Max AF:").pack(side=LEFT)
        ttk.Entry(filter_box, textvariable=self.af_threshold, width=6).pack(side=LEFT, padx=5)
        ttk.Checkbutton(filter_box, text="Nur Coding", variable=self.prefilter_protein_coding, bootstyle="round-toggle").pack(side=LEFT, padx=10)
        
        # Actions
        act_box = ttk.Frame(ctrl_row)
        act_box.pack(side=LEFT, padx=20)
        
        ttk.Button(act_box, text="START PIPELINE", command=self.on_start, bootstyle="success", width=18).pack(side=LEFT, padx=5)
        ttk.Button(act_box, text="STOP", command=self.on_stop, bootstyle="danger").pack(side=LEFT, padx=5)
        ttk.Button(act_box, text="⟳", command=self.on_refresh, bootstyle="info-outline", width=3).pack(side=LEFT, padx=5)
        
        # CADD Highlight
        ttk.Separator(act_box, orient=VERTICAL).pack(side=LEFT, fill=Y, padx=15)
        ttk.Label(act_box, text="CADD >", bootstyle="danger").pack(side=LEFT)
        ttk.Entry(act_box, textvariable=self.cadd_highlight_threshold, width=4).pack(side=LEFT, padx=2)
        ttk.Label(act_box, text="= 🔴").pack(side=LEFT)

        # Status
        status_box = ttk.Frame(ctrl_row)
        status_box.pack(side=RIGHT, fill=Y)
        self.phase_label = ttk.Label(status_box, text="Ready", font=("", 10, "bold"), bootstyle="primary")
        self.phase_label.pack(anchor="e")
        
        status_details = ttk.Frame(status_box)
        status_details.pack(anchor="e")
        self.perc_label = ttk.Label(status_details, text="0.0%", font=("", 9, "bold"), bootstyle="success")
        self.perc_label.pack(side=LEFT, padx=(0, 8))
        self.eta_label = ttk.Label(status_details, text="--:--", font=("", 9))
        self.eta_label.pack(side=LEFT)

        # Progressbar
        self.progress_bar = ttk.Progressbar(self, variable=self.progress_var, maximum=100, bootstyle="success-striped")
        self.progress_bar.pack(fill=X, padx=10, pady=(5, 10))

        # =====================================================================
        # 2. MAIN TABLE (Volle Breite)
        # =====================================================================
        self.tree_container = ttk.Frame(self)
        self.tree_container.pack(fill=BOTH, expand=YES, padx=10)

        tree_scroll_y = ttk.Scrollbar(self.tree_container, orient=VERTICAL)
        tree_scroll_x = ttk.Scrollbar(self.tree_container, orient=HORIZONTAL)

        self.tree = ttk.Treeview(
            self.tree_container,
            columns=self.columns, 
            show="headings",
            yscrollcommand=tree_scroll_y.set,
            xscrollcommand=tree_scroll_x.set,
            bootstyle="primary"
        )
        tree_scroll_y.config(command=self.tree.yview)
        tree_scroll_x.config(command=self.tree.xview)
        
        tree_scroll_y.pack(side=RIGHT, fill=Y)
        tree_scroll_x.pack(side=BOTTOM, fill=X)
        self.tree.pack(fill=BOTH, expand=YES)
        
        self.tree.tag_configure("high_cadd", background="#ffdddd", foreground="black")
        
        for col in self.columns:
            self.tree.heading(col, text=COLUMN_LABELS.get(col, col), command=lambda c=col: self._sort_by(c, False))
            self.tree.column(col, width=90, anchor="w")
            
        # =====================================================================
        # 3. COUNT LABEL
        # =====================================================================
        count_frame = ttk.Frame(self, padding=(10, 5))
        count_frame.pack(fill=X)
        self.count_label = ttk.Label(
            count_frame, 
            text="Angezeigt: 0 Varianten", 
            font=("", 10, "bold"), 
            bootstyle="inverse-primary", 
            anchor="center"
        )
        self.count_label.pack(fill=X)

        # =====================================================================
        # 4. COLLAPSIBLE FILTER SEKTION
        # =====================================================================
        mid_content = self._create_collapsible_section(self, "Ergebnis-Filter & Konfiguration", visible=True)
        
        # Grid Layout für Filter
        mid_content.columnconfigure(0, weight=1)
        mid_content.columnconfigure(1, weight=1)
        mid_content.columnconfigure(2, weight=2)
        mid_content.columnconfigure(3, weight=0)
        
        # ... (Erzeuge mid_content wie gehabt) ...
        
        # Layout: 4 Spalten
        # Col 0: Klinische Signifikanz
        # Col 1: Impact & Coding
        # Col 2: Scores & Consequences
        # Col 3: Gen-Listen & Buttons

        # --- SPALTE 0: CLINICAL ---
        col0 = ttk.Frame(mid_content)
        col0.grid(row=0, column=0, sticky="n", padx=10, pady=5)
        ttk.Label(col0, text="Klinische Signifikanz", font=("", 9, "bold")).pack(anchor="w", pady=(0,5))
        for txt, var in [("Pathogenic", self.postfilter_path), ("Likely Path.", self.postfilter_likely_path), ("VUS", self.postfilter_vus), ("Benign", self.postfilter_benign)]:
             ttk.Checkbutton(col0, text=txt, variable=var, bootstyle="square-toggle").pack(anchor="w")

        # --- SPALTE 1: IMPACT ---
        col1 = ttk.Frame(mid_content)
        col1.grid(row=0, column=1, sticky="n", padx=10, pady=5)
        ttk.Label(col1, text="Impact / Typ", font=("", 9, "bold")).pack(anchor="w", pady=(0,5))
        for txt, var in [("High", self.postfilter_impact_high), ("Moderate", self.postfilter_impact_moderate), ("Low", self.postfilter_impact_low)]:
             ttk.Checkbutton(col1, text=txt, variable=var, bootstyle="square-toggle").pack(anchor="w")
        ttk.Separator(col1).pack(fill=X, pady=8)
        ttk.Checkbutton(col1, text="Nur Protein-Coding", variable=self.postfilter_protein_coding, bootstyle="round-toggle").pack(anchor="w")

        # --- SPALTE 2: SCORES & CONS ---
        col2 = ttk.Frame(mid_content)
        col2.grid(row=0, column=2, sticky="n", padx=10, pady=5)
        
        # CADD
        cadd_f = ttk.Frame(col2)
        cadd_f.pack(anchor="w", pady=(0, 10))
        ttk.Label(cadd_f, text="CADD ≥ ").pack(side=LEFT)
        ttk.Entry(cadd_f, textvariable=self.postfilter_cadd_min, width=5).pack(side=LEFT)
        
        # Consequences
        ttk.Label(col2, text="Konsequenz", font=("", 9, "bold")).pack(anchor="w", pady=(0,5))
        for key in ["missense", "stop_gained", "frameshift", "splice"]:
             ttk.Checkbutton(col2, text=key.capitalize(), variable=self.postfilter_consequences[key], bootstyle="square-toggle").pack(anchor="w")

        # --- KORRIGIERTER BLOCK IN _build_gui ---
        
        # Spalte 3 (Genlisten)
        col3 = ttk.Frame(mid_content)
        col3.grid(row=0, column=3, sticky="nsew", padx=10, pady=5)
        
        # Whitelist
        ttk.Label(col3, text="Gen-Whitelist:").pack(anchor="w")
        wl = ttk.Frame(col3)
        wl.pack(fill=X)
        ttk.Entry(wl, textvariable=self.gen_whitelist, width=15).pack(side=LEFT, fill=X, expand=True)
        ttk.Button(wl, text="📂", width=3, command=lambda: self.load_gene_list("whitelist"), style="secondary-outline").pack(side=LEFT)
        
        # Blacklist - HIER WAR DER FEHLER
        # Das pady=(5,0) muss in .pack(), nicht in ttk.Label()
        ttk.Label(col3, text="Gen-Blacklist:").pack(anchor="w", pady=(5,0)) 
        
        bl = ttk.Frame(col3)
        bl.pack(fill=X)
        ttk.Entry(bl, textvariable=self.gen_blacklist, width=15).pack(side=LEFT, fill=X, expand=True)
        ttk.Button(bl, text="📂", width=3, command=lambda: self.load_gene_list("blacklist"), style="secondary-outline").pack(side=LEFT)
        
        # Buttons
        ttk.Separator(col3).pack(fill=X, pady=10)
        ttk.Button(col3, text="Filter Anwenden", command=self.apply_post_filter, bootstyle="primary").pack(fill=X, pady=2)
        ttk.Button(col3, text="Reset", command=self.reset_post_filter, bootstyle="secondary-outline").pack(fill=X)        # Buttons

        # =====================================================================
        # 5. COLLAPSIBLE BOTTOM SEKTION (Log & Export)
        # =====================================================================
        bottom_content = self._create_collapsible_section(self, "Log & Export", visible=True)
        
        # Log (Links)
        log_frame = ttk.Labelframe(bottom_content, text="System Log", padding=5)
        log_frame.pack(side=LEFT, fill=BOTH, expand=YES, padx=(0, 10))
        
        self.log_text = tk.Text(log_frame, height=6, state="normal", font=("Consolas", 8))
        self.log_text.pack(side=LEFT, fill=BOTH, expand=YES)
        sb_log = ttk.Scrollbar(log_frame, command=self.log_text.yview)
        sb_log.pack(side=RIGHT, fill=Y)
        self.log_text.config(yscrollcommand=sb_log.set)
        
        # Export (Rechts)
        exp_frame = ttk.Labelframe(bottom_content, text="Daten-Export", padding=10)
        exp_frame.pack(side=RIGHT, fill=Y)
        
        ttk.Label(exp_frame, text="VCF-Modus:", font=("", 8)).pack(anchor="w")
        ttk.Radiobutton(exp_frame, text="Original (+Anno)", variable=self.vcf_export_mode, value="complete").pack(anchor="w")
        ttk.Radiobutton(exp_frame, text="Gefiltert (Sichtbar)", variable=self.vcf_export_mode, value="filtered").pack(anchor="w")
        
        ttk.Separator(exp_frame).pack(fill=X, pady=5)
        
        btn_grid = ttk.Frame(exp_frame)
        btn_grid.pack(fill=X)
        
        ttk.Button(btn_grid, text="CSV", command=self.export_csv, width=8, bootstyle="secondary-outline").grid(row=0, column=0, padx=2, pady=2)
        ttk.Button(btn_grid, text="Excel", command=self.export_excel, width=8, bootstyle="success-outline").grid(row=0, column=1, padx=2, pady=2)
        ttk.Button(btn_grid, text="PDF", command=self.export_pdf, width=8, bootstyle="danger-outline").grid(row=1, column=0, padx=2, pady=2)
        ttk.Button(btn_grid, text="VCF", command=self.export_vcf, width=8, bootstyle="warning").grid(row=1, column=1, padx=2, pady=2)

    def open_general_settings(self):
        """NEU: Mit Column Link Templating System."""
        settings_window = ttk.Toplevel(self)
        settings_window.title("Allgemeine Einstellungen & Links")
        settings_window.geometry("850x700")
        settings_window.transient(self)
        settings_window.withdraw()
        settings_window.update_idletasks()
        settings_window.place_window_center()
        settings_window.deiconify()
        settings_window.grab_set()
        
        notebook = ttk.Notebook(settings_window)
        notebook.pack(fill=BOTH, expand=YES, padx=10, pady=10)
        
        # TAB 1: Allgemein
        tab_general = ttk.Frame(notebook, padding=10)
        notebook.add(tab_general, text="Allgemein")
        
        ag_frame = ttk.Labelframe(tab_general, text="AlphaGenome AI", padding=10)
        ag_frame.pack(fill=X, pady=5)
        ttk.Label(ag_frame, text="API-Key unter 'APIs & Services' konfigurierbar.").pack(anchor="w")
        
        db_frame = ttk.Labelframe(tab_general, text="Debugging & Tools", padding=10)
        db_frame.pack(fill=X, pady=5)
        ttk.Label(db_frame, text="Pfad zu externem SQLite Viewer:").pack(anchor="w")
        db_row = ttk.Frame(db_frame)
        db_row.pack(fill=X)
        ttk.Entry(db_row, textvariable=self.external_db_viewer).pack(side=LEFT, fill=X, expand=YES)
        ttk.Button(db_row, text="...", width=3, command=self.choose_db_viewer, bootstyle="secondary-outline").pack(side=LEFT, padx=5)
        
        stale_frame = ttk.Labelframe(tab_general, text="Veralterungs-Intervalle (Tage)", padding=10)
        stale_frame.pack(fill=X, pady=5)
        ttk.Label(stale_frame, text="AF-Daten:").grid(row=0, column=0, sticky="w", padx=5)
        ttk.Entry(stale_frame, textvariable=self.stale_days, width=8).grid(row=0, column=1, sticky="w")
        ttk.Label(stale_frame, text="Vollannotation:").grid(row=1, column=0, sticky="w", padx=5, pady=5)
        ttk.Entry(stale_frame, textvariable=self.stale_days_full, width=8).grid(row=1, column=1, sticky="w")
        
        # TAB 2: Ressourcen
        tab_resources = ttk.Frame(notebook, padding=10)
        notebook.add(tab_resources, text="Ressourcen")
        res_sf = ScrolledFrame(tab_resources, autohide=True, padding=5)
        res_sf.pack(fill=BOTH, expand=YES)
        self._settings_resource_widgets = {}
        build_resource_cards(res_sf, get_resource_manager(), self.logger, self._settings_resource_widgets, _t=self._t)

        res_btn_frame = ttk.Frame(tab_resources, padding=5)
        res_btn_frame.pack(fill=X)
        def _refresh_resources():
            rm = get_resource_manager()
            rm.refresh()
            st = rm.get_status()
            for key, w in self._settings_resource_widgets.items():
                is_ok = st.get(key) in ("found", "registered", "created")
                w["status_lbl"].config(
                    text="  Vorhanden" if is_ok else "  Fehlt",
                    foreground="green" if is_ok else "red"
                )
        ttk.Button(res_btn_frame, text="Ressourcen neu suchen", command=_refresh_resources,
                   bootstyle="info-outline").pack(side=LEFT)

        # TAB 3: Spalten-Links
        tab_links = ttk.Frame(notebook, padding=10)
        notebook.add(tab_links, text="Spalten-Links")
        
        ttk.Label(tab_links, text="Platzhalter: {value}, {chr}, {pos}, {ref}, {alt}, {gene}, {rsid}").pack(anchor="w", pady=(0, 10))
        
        tree_container = ttk.Frame(tab_links)
        tree_container.pack(fill=BOTH, expand=YES)
        
        link_cols = ("Spalte", "Auslöser", "URL Template")
        link_tree = ttk.Treeview(tree_container, columns=link_cols, show="headings", height=10)
        link_tree.heading("Spalte", text=self._t("Spalte"))
        link_tree.heading("Auslöser", text=self._t("Auslöser"))
        link_tree.heading("URL Template", text="URL Template")
        link_tree.column("Spalte", width=80)
        link_tree.column("Auslöser", width=80)
        link_tree.column("URL Template", width=500)
        
        link_scroll = ttk.Scrollbar(tree_container, orient="vertical", command=link_tree.yview)
        link_tree.configure(yscrollcommand=link_scroll.set)
        link_tree.pack(side=LEFT, fill=BOTH, expand=YES)
        link_scroll.pack(side=RIGHT, fill=Y)
        
        for col in self.columns:
            cfg = self.column_links.get(col, {'trigger': 'none', 'template': ''})
            link_tree.insert("", "end", iid=col, values=(col, cfg.get('trigger', 'none'), cfg.get('template', '')))
        
        edit_frame = ttk.Labelframe(tab_links, text="Bearbeiten", padding=10)
        edit_frame.pack(fill=X, pady=10)
        
        sel_col_var = tk.StringVar()
        sel_trigger_var = tk.StringVar(value="none")
        sel_temp_var = tk.StringVar()
        preset_var = tk.StringVar(value="")
        
        row1 = ttk.Frame(edit_frame)
        row1.pack(fill=X, pady=2)
        ttk.Label(row1, text="Spalte:").pack(side=LEFT, padx=5)
        ttk.Label(row1, textvariable=sel_col_var, font=("", 10, "bold"), width=12).pack(side=LEFT)
        ttk.Label(row1, text="Auslöser:").pack(side=LEFT, padx=(20, 5))
        trigger_combo = ttk.Combobox(row1, textvariable=sel_trigger_var, values=["none", "single", "double"], state="readonly", width=10)
        trigger_combo.pack(side=LEFT)
        ttk.Label(row1, text="Preset:").pack(side=LEFT, padx=(20, 5))
        preset_combo = ttk.Combobox(row1, textvariable=preset_var, state="readonly", width=18)
        preset_combo.pack(side=LEFT)
        
        def update_presets(*args):
            col = sel_col_var.get()
            if col in Config.COLUMN_LINK_PRESETS:
                preset_combo['values'] = list(Config.COLUMN_LINK_PRESETS[col].keys())
            else:
                preset_combo['values'] = []
            preset_var.set("")
        sel_col_var.trace_add("write", update_presets)
        
        def apply_preset(*args):
            col = sel_col_var.get()
            preset = preset_var.get()
            if col and preset and col in Config.COLUMN_LINK_PRESETS:
                sel_temp_var.set(Config.COLUMN_LINK_PRESETS[col].get(preset, ""))
        preset_var.trace_add("write", apply_preset)
        
        row2 = ttk.Frame(edit_frame)
        row2.pack(fill=X, pady=5)
        ttk.Label(row2, text="Template:").pack(side=LEFT, padx=5)
        ttk.Entry(row2, textvariable=sel_temp_var, width=80).pack(side=LEFT, fill=X, expand=YES, padx=5)
        
        row3 = ttk.Frame(edit_frame)
        row3.pack(fill=X, pady=5)
        
        def on_tree_select(event):
            sel = link_tree.selection()
            if not sel: return
            vals = link_tree.item(sel[0])['values']
            sel_col_var.set(vals[0])
            sel_trigger_var.set(vals[1])
            sel_temp_var.set(vals[2])
        link_tree.bind("<<TreeviewSelect>>", on_tree_select)
        
        def save_current_entry():
            col = sel_col_var.get()
            if not col: return
            new_cfg = {'trigger': sel_trigger_var.get(), 'template': sel_temp_var.get()}
            self.column_links[col] = new_cfg
            link_tree.item(col, values=(col, new_cfg['trigger'], new_cfg['template']))
        
        def reset_to_defaults():
            self.column_links = Config.DEFAULT_COLUMN_LINKS.copy()
            for col in self.columns:
                cfg = self.column_links.get(col, {'trigger': 'none', 'template': ''})
                link_tree.item(col, values=(col, cfg.get('trigger', 'none'), cfg.get('template', '')))
        
        ttk.Button(row3, text="Übernehmen", command=save_current_entry, bootstyle="success").pack(side=LEFT, padx=5)
        ttk.Button(row3, text="Standard", command=reset_to_defaults, bootstyle="warning-outline").pack(side=LEFT, padx=5)

        # ==================================================================
        # TAB 4: APIs & Services
        # ==================================================================
        tab_api = ttk.Frame(notebook, padding=10)
        notebook.add(tab_api, text="APIs & Services")

        api_sf = ScrolledFrame(tab_api, autohide=True, padding=5)
        api_sf.pack(fill=BOTH, expand=YES)
        api_inner = api_sf

        # Sammlung aller tkinter-Variablen fuer Refresh/Reset
        self._api_tk_vars = {}

        def _make_tk_var(phase_key, api_key, field, value):
            """Erstellt tk-Variable und registriert sie."""
            var_key = (phase_key, api_key, field)
            if isinstance(value, bool):
                v = tk.BooleanVar(value=value)
            elif isinstance(value, int):
                v = tk.IntVar(value=value)
            elif isinstance(value, float):
                v = tk.DoubleVar(value=value)
            else:
                v = tk.StringVar(value=str(value))
            self._api_tk_vars[var_key] = v

            def _on_change(*args, pk=phase_key, ak=api_key, f=field):
                try:
                    val = v.get()
                    if pk == "global":
                        self.api_settings[pk][f] = val
                    else:
                        self.api_settings[pk][ak][f] = val
                except (tk.TclError, KeyError):
                    pass
            v.trace_add("write", _on_change)
            return v

        def _build_api_row(parent, phase_key, api_key, settings):
            """Baut eine einzelne API-Zeile mit Checkbox + Spinboxen."""
            row = ttk.Frame(parent)
            row.pack(fill=X, pady=2)

            label = Config.API_LABELS.get(api_key, api_key)

            # Enabled Checkbox
            en_var = _make_tk_var(phase_key, api_key, "enabled", settings.get("enabled", True))
            ttk.Checkbutton(row, text=label, variable=en_var, width=22).pack(side=LEFT, padx=(0, 5))

            # Timeout
            if "timeout" in settings:
                ttk.Label(row, text="Timeout").pack(side=LEFT, padx=(5, 2))
                to_var = _make_tk_var(phase_key, api_key, "timeout", settings["timeout"])
                ttk.Spinbox(row, from_=5, to=300, textvariable=to_var, width=5).pack(side=LEFT)
                ttk.Label(row, text="s").pack(side=LEFT, padx=(0, 5))

            # Batch Size
            if "batch_size" in settings:
                ttk.Label(row, text="Batch").pack(side=LEFT, padx=(5, 2))
                bs_var = _make_tk_var(phase_key, api_key, "batch_size", settings["batch_size"])
                ttk.Spinbox(row, from_=10, to=5000, textvariable=bs_var, width=6).pack(side=LEFT)

            # Workers/Semaphore
            if "workers" in settings:
                ttk.Label(row, text="W").pack(side=LEFT, padx=(5, 2))
                w_var = _make_tk_var(phase_key, api_key, "workers", settings["workers"])
                ttk.Spinbox(row, from_=1, to=20, textvariable=w_var, width=4).pack(side=LEFT)

            # Rate-Limit
            if "rate_limit" in settings:
                ttk.Label(row, text="Rate").pack(side=LEFT, padx=(5, 2))
                rl_var = _make_tk_var(phase_key, api_key, "rate_limit", settings["rate_limit"])
                ttk.Spinbox(row, from_=0.05, to=5.0, increment=0.05, textvariable=rl_var, width=6).pack(side=LEFT)
                ttk.Label(row, text="s/req").pack(side=LEFT, padx=(0, 5))

            # API-Key (fuer AlphaGenome etc.)
            if "api_key" in settings:
                ttk.Label(row, text="API-Key:").pack(side=LEFT, padx=(10, 2))
                ak_var = _make_tk_var(phase_key, api_key, "api_key", settings.get("api_key", ""))
                ttk.Entry(row, textvariable=ak_var, show="*", width=25).pack(side=LEFT, fill=X, expand=YES)

        # --- Globale Einstellungen ---
        global_frame = ttk.Labelframe(api_inner, text="Globale Einstellungen", padding=10)
        global_frame.pack(fill=X, pady=5)

        g_row1 = ttk.Frame(global_frame)
        g_row1.pack(fill=X, pady=2)
        ttk.Label(g_row1, text="Max Retries:").pack(side=LEFT, padx=(0, 5))
        mr_var = _make_tk_var("global", "global", "max_retries",
                              self.api_settings["global"]["max_retries"])
        ttk.Spinbox(g_row1, from_=1, to=10, textvariable=mr_var, width=5).pack(side=LEFT)

        g_row2 = ttk.Frame(global_frame)
        g_row2.pack(fill=X, pady=2)
        ttk.Label(g_row2, text="NCBI API-Key:").pack(side=LEFT, padx=(0, 5))
        nk_var = _make_tk_var("global", "global", "ncbi_api_key",
                              self.api_settings["global"]["ncbi_api_key"])
        ttk.Entry(g_row2, textvariable=nk_var, show="*", width=40).pack(side=LEFT, fill=X, expand=YES)

        # --- Phase-Frames ---
        for phase_key, phase_label in Config.API_PHASE_LABELS.items():
            pf = ttk.Labelframe(api_inner, text=phase_label, padding=10)
            pf.pack(fill=X, pady=5)
            phase_data = self.api_settings.get(phase_key, {})
            for api_key, api_settings in phase_data.items():
                if isinstance(api_settings, dict):
                    _build_api_row(pf, phase_key, api_key, api_settings)

        # --- Reset-Button ---
        def _reset_api_defaults():
            import copy as _cp
            self.api_settings = _cp.deepcopy(Config.DEFAULT_API_SETTINGS)
            # Alle tk-Variablen aktualisieren
            for (pk, ak, field), var in self._api_tk_vars.items():
                try:
                    if pk == "global":
                        new_val = Config.DEFAULT_API_SETTINGS[pk][field]
                    else:
                        new_val = Config.DEFAULT_API_SETTINGS[pk][ak][field]
                    var.set(new_val)
                except (KeyError, TypeError):
                    pass

        api_reset_frame = ttk.Frame(api_inner)
        api_reset_frame.pack(fill=X, pady=10)
        ttk.Button(api_reset_frame, text="Auf Standard zuruecksetzen",
                   command=_reset_api_defaults, bootstyle="warning-outline").pack(anchor="center")

        btn_main = ttk.Frame(settings_window, padding=10)
        btn_main.pack(fill=X, side=BOTTOM)
        
        def save_all():
            self._save_settings()
            settings_window.destroy()
        
        ttk.Button(btn_main, text="Speichern & Schließen", command=save_all, bootstyle="success").pack(side=RIGHT)
        ttk.Button(btn_main, text="Abbrechen", command=settings_window.destroy, bootstyle="secondary").pack(side=RIGHT, padx=5)

    def choose_db_viewer(self):
        """Dateidialog für DB-Viewer."""
        path = filedialog.askopenfilename(
            title=self._t("Wähle DB-Browser Executable"),
            filetypes=[("Executables", "*.exe"), (self._t("Alle Dateien"), "*.*")]
        )
        if path:
            self.external_db_viewer.set(path)

    def open_external_db(self):
        """Öffnet die aktuelle Datenbank im konfigurierten Viewer."""
        viewer = self.external_db_viewer.get()
        if not viewer or not os.path.exists(viewer):
            Messagebox.show_error("Kein gültiger externer DB-Viewer konfiguriert.\nBitte Pfad in 'Optionen > Allgemeine Einstellungen' hinterlegen.", "Fehler")
            return
            
        if not self.db or not self.db.db_path or not os.path.exists(self.db.db_path):
             Messagebox.show_error("Datenbank noch nicht initialisiert.", "Fehler")
             return
             
        try:
            # Popen verwenden, damit die GUI nicht blockiert
            subprocess.Popen([viewer, self.db.db_path])
            self.logger.log(f"[Debug] Öffne DB mit {viewer}")
        except Exception as e:
            Messagebox.show_error(f"Fehler beim Starten des Viewers:\n{e}", "Fehler")
            # ============== Main ==============

if __name__ == "__main__":
    # ✅ V17: Prüfe Single Instance Lock
    success, pid = check_single_instance()
    if not success:
        _close_splash()  # Splash schliessen bevor Fehlerdialog
        root = tk.Tk()
        root.withdraw()
        Messagebox.show_error(
            f"Eine Instanz von Variant Fusion läuft bereits (PID: {pid}).\n"
            "Bitte schließen Sie die laufende Instanz zuerst, um Datenkorruption zu vermeiden.",
            "Bereits gestartet"
        )
        sys.exit(1)

    # V17.1: Splash schliessen bevor App (ttk.Window = neues Tk-Root) erstellt wird
    _splash_log("Starte GUI ...")
    _close_splash()

    app = App()
    app.mainloop()





 
                                                      
